# referrals/views.py
from django.shortcuts import render, redirect, get_object_or_404,HttpResponse
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import AffiliateLink, ReferralConversion, CommissionPayout
from .utils import make_code, calculate_commission_for
from django.views.decorators.http import require_POST
from decimal import Decimal
from django.db import models
from Admin_App.models import *
from django.views.decorators.csrf import csrf_exempt
import traceback
from django.template.loader import render_to_string
from CRM_Panel.models import *
from django.db.models.functions import Coalesce
from django.db.models import Count, Avg, Max, Min, Q,Sum
from datetime import datetime
from django.contrib import messages
import json
from django.http import JsonResponse
from django.urls import reverse
import hashlib
import re
from django.core.paginator import Paginator


def generate_property_fingerprint(property_no, building_name, locality, pincode=""):
    """
    Generates a normalized SHA-256 hash unique key based on physical location.
    Strips spaces and special characters for consistent matching.
    """
    def clean_text(val):
        if not val:
            return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()

    raw_string = f"{clean_text(property_no)}_{clean_text(building_name)}_{clean_text(locality)}_{clean_text(pincode)}"
    return hashlib.sha256(raw_string.encode('utf-8')).hexdigest()


def referral_redirect(request, code):
    """
    Route: /r/<code>/  - sets cookie and redirects to the affiliate's target URL.
    """
    aff = get_object_or_404(AffiliateLink, code=code, active=True)
    target = aff.get_target_url()
    response = redirect(target)
    response.set_cookie('referral', aff.code, max_age=60*60*24*30, samesite='Lax')
    return response

@login_required
@require_POST
def create_affiliate_link(request):
    owner = request.user
    target_type = request.POST.get('target_type', 'subscription')
    target_id = request.POST.get('target_id')
    code = make_code(prefix=f'RM{owner.id}')
    # ensure uniqueness
    while AffiliateLink.objects.filter(code=code).exists():
        code = make_code(prefix=f'RM{owner.id}')
    aff = AffiliateLink.objects.create(owner=owner, created_by=request.user, code=code, target_type=target_type, target_id=target_id)
    # return full absolute URL
    link = request.scheme + '://' + request.get_host() + aff.get_target_url()
    return JsonResponse({'ok': True, 'code': aff.code, 'link': link})

@login_required
@require_POST
def request_payout(request):
    code = request.POST.get('code')
    try:
        aff = AffiliateLink.objects.get(code=code, owner=request.user)
    except AffiliateLink.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)
    amount = Decimal(request.POST.get('amount') or 0)
    if amount <= 0:
        return JsonResponse({'ok': False, 'error': 'Invalid amount'}, status=400)
    payout = CommissionPayout.objects.create(affiliate=aff, total_amount=amount, status='requested')
    return JsonResponse({'ok': True, 'payout_id': str(payout.id)})

def is_staff(user):
    return user.is_staff

@login_required
@user_passes_test(is_staff)
@require_POST
def admin_create_affiliate_link(request):
    owner_id = request.POST.get('owner_id')
    from django.contrib.auth import get_user_model
    User = get_user_model()
    owner = get_object_or_404(User, pk=owner_id)
    code = make_code(prefix=f'RM{owner.id}')
    while AffiliateLink.objects.filter(code=code).exists():
        code = make_code(prefix=f'RM{owner.id}')
    aff = AffiliateLink.objects.create(owner=owner, created_by=request.user, code=code, target_type=request.POST.get('target_type','subscription'), target_id=request.POST.get('target_id'))
    link = request.scheme + '://' + request.get_host() + aff.get_target_url()
    return JsonResponse({'ok': True, 'code': aff.code, 'link': link})


def rm_dashboard(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    # If they are an Admin, we only care that they have an impersonate_id.
    # We don't care if their standard 'User_id' is missing.
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 4. Data Fetching
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/rm_dashboard.html", context)


############### Views start for user logout #####################

@csrf_exempt
def User_Logout(request):
    try:
        del request.session['User_id']
        del request.session['user_type']
        return JsonResponse({"status":"1",'msg': 'Logout Successfully '})
    except:
        print(traceback.format_exc())

######### Views end for user logout #############################


############ Views start for update rm profile #####################

def Update_Profile_Rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    # If they are an Admin, we only care that they have an impersonate_id.
    # We don't care if their standard 'User_id' is missing.
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 4. Data Fetching
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Profile/rm_profile.html", context)

############## Views end for update rm profile ########################


############### Views start for assign enquiries to Rm #######################

def Assign_Enquiry_Rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    # If they are an Admin, we only care that they have an impersonate_id.
    # We don't care if their standard 'User_id' is missing.
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 4. Data Fetching
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()


    enquiry_obj = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).order_by('-id')
    enquiry_obj_count = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    ############## Enquiries Stats By Source ##############################

    fb_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="facebook",assigned_to__id=user_obj.id).count()
    insta_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="instagram",assigned_to__id=user_obj.id).count()
    whatsapp_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="whatsapp",assigned_to__id=user_obj.id).count()
    google_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="google",assigned_to__id=user_obj.id).count()
    linkedin_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="linkedin",assigned_to__id=user_obj.id).count()
    twitter_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="twitter",assigned_to__id=user_obj.id).count()
    youtube_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="youtube",assigned_to__id=user_obj.id).count()
    referral_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="referral",assigned_to__id=user_obj.id).count()

    ########### Enquiry Stats by lead source ################################

    pending_obj_count = PropertyEnquiry.objects.filter(lead_status="Pending",assigned_to__id=user_obj.id).count()
    progress_obj_count = PropertyEnquiry.objects.filter(lead_status="In Progress",assigned_to__id=user_obj.id).count()
    hold_obj_count = PropertyEnquiry.objects.filter(lead_status="Hold",assigned_to__id=user_obj.id).count()
    closed_obj_count = PropertyEnquiry.objects.filter(lead_status="Closed",assigned_to__id=user_obj.id).count()
    cancelled_obj_count = PropertyEnquiry.objects.filter(lead_status="Cancelled",assigned_to__id=user_obj.id).count()

    rendered = render_to_string("rm_panel/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'property_enquiry_list':rendered,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Enquiry/assign_enquiry.html", context)

########### Views end for assig enquiries to RM ###########################


########### Views start for ajax for datewise filter for user #################

@csrf_exempt
def date_property_filter_user(request):
    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        lead_source = request.POST.get('lead_source')
        lead_status = request.POST.get('lead_status')
        role = request.POST.get('role')
        user_id = request.POST.get('user_id')

        if lead_source != "All" and lead_status != "All":
            # Case 1: Both filters applied
            enquiry_obj = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                utm_link__utm_source=lead_source,
                lead_status=lead_status,
                assigned_to__id=user_id
            ).order_by('-id')
            enquiry_obj_count= PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                utm_link__utm_source=lead_source,
                lead_status=lead_status,
                assigned_to__id=user_id
            ).count()
        
        elif lead_source != "All" and lead_status == "All":
            # Case 2: Only source filter, all statuses
            enquiry_obj = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                utm_link__utm_source=lead_source,
                assigned_to__id=user_id
            ).order_by('-id')
            enquiry_obj_count = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                utm_link__utm_source=lead_source,
                assigned_to__id=user_id
            ).count()
        
        elif lead_source == "All" and lead_status != "All":
            # Case 3: Only status filter, all sources
            enquiry_obj = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                lead_status=lead_status,
                assigned_to__id=user_id
            ).order_by('-id')
            enquiry_obj_count = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                lead_status=lead_status,
                assigned_to__id=user_id
            ).count()
        
        else:
            # Case 4: No filters (All sources, All statuses)
            enquiry_obj = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                assigned_to__id=user_id
            ).order_by('-id')
            enquiry_obj_count = PropertyEnquiry.objects.filter(
                enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,
                assigned_to__id=user_id
            ).count()


    ############## Enquiries Stats By Source ##############################

    fb_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="facebook",assigned_to__id=user_id).count()
    insta_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="instagram",assigned_to__id=user_id).count()
    whatsapp_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="whatsapp",assigned_to__id=user_id).count()
    google_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="google",assigned_to__id=user_id).count()
    linkedin_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="linkedin",assigned_to__id=user_id).count()
    twitter_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="twitter",assigned_to__id=user_id).count()
    youtube_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="youtube",assigned_to__id=user_id).count()
    referral_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,utm_link__utm_source="referral",assigned_to__id=user_id).count()

    ########### Enquiry Stats by lead source ################################

    pending_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,lead_status="Pending",assigned_to__id=user_id).count()
    progress_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,lead_status="In Progress",assigned_to__id=user_id).count()
    hold_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,lead_status="Hold",assigned_to__id=user_id).count()
    closed_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,lead_status="Closed",assigned_to__id=user_id).count()
    cancelled_obj_count = PropertyEnquiry.objects.filter(enquiry_date__gte=start_date,
                enquiry_date__lte=end_date,lead_status="Cancelled",assigned_to__id=user_id).count()

    if role == "Relationship Manager":
        rendered = render_to_string("rm_panel/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    elif role == "Agent":
        rendered = render_to_string("agent/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    elif role == "Agency/Builder":
        rendered = render_to_string("agency_panel/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    else:
        rendered = render_to_string("crm/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})

    return HttpResponse(rendered)

############ Views end for ajax for datewise filter for user #######################


############### Views start for update property enquiry ############################

def update_enquiry_rm(request,id):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    # If they are an Admin, we only care that they have an impersonate_id.
    # We don't care if their standard 'User_id' is missing.
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 4. Data Fetching
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    enquiry = PropertyEnquiry.objects.get(id=id)
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry':enquiry,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Enquiry/update_enquiry.html", context)

############ Views end for update property enquiry #################################


############ Views start for rental forms for RM ######################

def residential_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    # If they are an Admin, we only care that they have an impersonate_id.
    # We don't care if their standard 'User_id' is missing.
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    user_role = user_obj.user_role

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    

############### Views end foor rental forms for RM #####################


############## Views start for rental forms for RM #################### ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════

    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', '').strip()
    bhk_query = request.GET.get('bhk_type', '').strip()
    state_query = request.GET.get('state', '').strip()          # NEW
    furnish_query = request.GET.get('furnishing', '').strip()
    possession_query = request.GET.get('possession', '').strip()  # maps to availability_status

    prop_type_query = request.GET.get('property_type', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()     # listed_by_role
    uploaded_by_query = request.GET.get('uploaded_by', '').strip()  # NEW -> uploaded_by_role
    budget_query = request.GET.get('budget', '').strip()

    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query = request.GET.get('duplicate', '').strip()

    # ═══════════════════════════════════════
    # BASE QUERYSET & PERSISTENT SR.NO MAP
    # ═══════════════════════════════════════

    base_properties = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    absolute_ordered_ids = list(base_properties.values_list('id', flat=True))

    properties = base_properties

    # ═══════════════════════════════════════
    # GLOBAL SEARCH FILTER (Across ALL Fields + Sr.No)
    # ═══════════════════════════════════════

    if search_query:
        sr_no_query = Q()

        if search_query.isdigit():
            target_index = int(search_query) - 1
            if 0 <= target_index < len(absolute_ordered_ids):
                target_id = absolute_ordered_ids[target_index]
                sr_no_query = Q(id=target_id)

        properties = properties.filter(
            sr_no_query |
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(renting_option__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(bhk_type__icontains=search_query) |
            Q(locality_area__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(available_for__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(upload_file_name__icontains=search_query)
        )

    # ═══════════════════════════════════════
    # ADVANCED DROPDOWN FILTERS
    # ═══════════════════════════════════════

    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__iexact=city_query)

    if bhk_query and bhk_query != 'All BHK':
        properties = properties.filter(bhk_type__iexact=bhk_query)

    if state_query and state_query != 'All States':          # NEW
        properties = properties.filter(state__iexact=state_query)

    if furnish_query and furnish_query != 'All':
        properties = properties.filter(furnishing_status__iexact=furnish_query)

    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__iexact=possession_query)

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__iexact=prop_type_query)

    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':   # NEW
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_10k':
            properties = properties.filter(monthly_rent__lt=10000)
        elif budget_query == '10k_25k':
            properties = properties.filter(monthly_rent__gte=10000, monthly_rent__lte=25000)
        elif budget_query == '25k_50k':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=50000)
        elif budget_query == 'above_50k':
            properties = properties.filter(monthly_rent__gt=50000)

    # Date Filters
    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            properties = properties.filter(created_at__date__gte=from_date)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            properties = properties.filter(created_at__date__lte=to_date)

    # ═══════════════════════════════════════
    # PAGINATION & INJECT PERSISTENT SR.NO
    # ═══════════════════════════════════════
    filtered_count = properties.count()

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    id_to_srno = {pid: idx + 1 for idx, pid in enumerate(absolute_ordered_ids)}
    for prop in page_obj:
        prop.original_sr_no = id_to_srno[prop.id]

   

    # ═══════════════════════════════════════
    # EXPORT DATA (CSV / EXCEL)
    # ═══════════════════════════════════════
    if request.GET.get('download') in ['excel', 'csv']:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        from collections import OrderedDict
        import csv

        # ═══════════════════════════════════════
        # SINGLE SOURCE OF TRUTH — reuse the same field map the template
        # download and the importer already use, so LABELS stay in sync.
        # ═══════════════════════════════════════
        (field_map_sections, field_to_label_map, label_to_field_map,
         system_injected_map, helper_only_labels_map,
         decimal_fields_map, int_fields_map) = _residential_field_map()

        # ═══════════════════════════════════════
        # CANONICAL REQUIRED FIELDS — must be the literal same set used in
        # import_residential_excel's REQUIRED_FIELD_KEYS. Do NOT derive this
        # from _residential_field_map()'s per-field `required` flag — that
        # flag is used for template styling only and does not match the
        # actual mandatory inputs on the "Post Your Residential Rental" form
        # (e.g. it doesn't mark listed_by_name/email/contact as required,
        # but the Add form and the importer both do).
        # ═══════════════════════════════════════
        REQUIRED = {
            "property_type", "property_no", "bhk_type", "renting_option",
            "built_up_area", "bathrooms", "building_configuration", "total_floors",
            "furnishing_status", "available_for", "monthly_rent",
            "address", "city", "locality_area", "state", "pincode",
            "listed_by_name", "listed_by_contact", "listed_by_email", "listed_by_role",
        }

        # FIELD_LABELS starts from the field map (guarantees every shared,
        # re-importable column uses the exact label the importer expects).
        FIELD_LABELS = dict(field_to_label_map)

        # Export-only columns: appear on this report but are NOT part of the
        # import contract in _residential_field_map(). Never required, and
        # safely ignored (as unmatched headers) if this file is re-uploaded.
        FIELD_LABELS.update({

            "id": "Property ID",
            "property_title": "Property Title",
            "description": "Property Summary (Auto)",
            "rent_residential_desc": "Property Description (Auto)",
            "upload_file_name": "Uploaded File Name",
            "listing_status": "Listing Status",
            "approval_status": "Approval Status",
            "listed_elsewhere": "Listed Elsewhere?",
            "portal_name": "Portal Name",

            "property_unique_key": "Property Unique Key",
            "duplicate_count":  "Duplicate Count",
            "duplicate_group_id": "Duplicate Group ID",
            "is_duplicate": "Is Duplicate",
            "is_deleted": "Is Deleted",
            "deleted_at": "Deleted At",
            "deleted_by": "Deleted By",
            "created_at": "Created At",
        })



        sections = {


            "S.No": [
                "sr_no",
            ],
            "Listed By": [
                "listed_by_type", "listed_by_id", "listed_by_name", "listed_by_contact", "listed_by_email",
                "listed_by_role",
            ],
            "Basic Information": [
                "id", "property_title", "property_type", "property_no", "bhk_type", "renting_option",
                "built_up_area", "bathrooms", "balconies",
                "building_configuration", "total_floors", "facing_direction",
                "furnishing_status", "available_for"
            ],
            "Property Details": [
                "carpet_area", "city_zone", "ownership_type", "property_condition", "property_age",
                "wing_number", "building_name"
            ],
            "Availability Details": [
                "availability_status", "available_from", "lease_duration",
            ],
            "Pricing & Brokarage Details": [
                "monthly_rent","brokerage_percentage", "manual_brokerage", "advance_rent_month", "advance_rent_amount",
                "security_deposit_type", "security_deposit_amount",
                "maintenance_type", "monthy_maintenance_amount", "total_move_in_cost"
            ],
            "Property Location Details": [
                "address", "city", "locality_area", "property_landmark", "state",
                "pincode", "main_road_connectivity", "google_maps_link", "latitude", "longitude",
            ],
            "Amenities & Features": [
                "amenities",
            ],
            "Nearby Facilities": [
                "nearby_facilities",
            ],
            "Auto Property Summary & Description Generated Details": [
                "description",
                "rent_residential_desc",
            ],
            "Property Description(Added By User)": [
                "user_description",
            ],

            "Property Listed Elsewhere": [
                "listed_elsewhere", "portal_name",
                
            ],
            "Property Uploaded By(Auto Generated)": [
                "uploaded_by_name", "uploaded_by_email",
                "uploaded_by_contact", "uploaded_by_role", "upload_file_name",
            ],
            "Record Management": [
            "listing_status", "approval_status","property_unique_key",
            "duplicate_count", "duplicate_group_id" ,
            "is_duplicate" , 
            "is_deleted","deleted_at", "deleted_by", "created_at"
            ]
        }

        HINTS = {
            "sr_no":"Auto","id": "Property ID Auto-Generated",
            "property_title": "Auto_Generated Title", "property_type": "Apartment",
            "property_no": "e.g. B-402", "bhk_type": "e.g. 2bhk",
            "renting_option": "Full Property/Single Room/Shared Room", "built_up_area": "sq.ft",
            "bathrooms": "Number", "balconies": "Number", "building_configuration": "e.g. G+3",
            "total_floors": "Number", "facing_direction": "North/East", "furnishing_status": "Semi Furnished",
            "available_for": "Family/Bachelor", "carpet_area": "sq.ft", "city_zone": "North/South", "ownership_type": "Freehold",
            "property_condition": "Resale", "property_age": "1-3 Years",
            "wing_number": "e.g. A/B", "building_name": "Text", "availability_status": "Ready to Move",
            "available_from": "YYYY-MM-DD", "lease_duration": "11 Months", "monthly_rent": "₹","brokerage_percentage": "1%/Fixed Amount",
            "manual_brokerage": "e.g. 2.5%", "advance_rent_month": "0-11/fixed",
            "advance_rent_amount": "₹", "security_deposit_type": "0-11/fixed", "security_deposit_amount": "₹",
            "maintenance_type": "Included in Rent/Extra", "monthy_maintenance_amount": "₹",
            "total_move_in_cost": "₹", "address": "Full Address", "city": "Text", "locality_area": "Text",
            "property_landmark": "Optional", "state": "e.g. Maharashtra", "pincode": "6-digit",
            "main_road_connectivity": "Optional", "google_maps_link": "URL", "latitude": "21.1458",
            "longitude": "23.1458", "amenities": "Comma-sep",
            "nearby_facilities": "Comma-sep", "description": "Short Summary",
            "rent_residential_desc": "Long Rich Text", "user_description": "Added by user",  "listed_elsewhere": "Yes/No",
            "portal_name": "e.g. 99acres, MagicBricks",
            "listed_by_name": "Full Name", "listed_by_contact": "10 Digits", "listed_by_email": "email@example.com",
            "listed_by_role": "Owner/Agent/Admin", "listed_by_type": "Self/Other",
            "uploaded_by_name": "Admin Name", "uploaded_by_email": "Admin Email",
            "uploaded_by_contact": "Admin Contact", "uploaded_by_role": "Admin Role",
            "upload_file_name": "File Name", "listing_status": "Published/Draft", "approval_status": "Approved/Pending",
            "property_unique_key" : "Property unique id",
            "duplicate_count": "duplicate_count" ,
            "duplicate_group_id": "Duplicate Ggroup Id",
            "is_duplicate": "Yes/No", 
            "is_deleted": "Yes/No",
            "deleted_at": "YYYY-MM-DD (Auto)", "deleted_by": "Admin Name",
            "created_at": "YYYY-MM-DD"
        }

        BOOL_FIELDS = {"is_deleted", "is_duplicate"}

        ROLE_BROKERAGE_LABELS = {
            "admin": "EstateFlow Service Fee",
            "relationship manager": "Service Fee",
            "landlord": "Tenant Service Fee",
            "agent": "Brokerage",
            "agency/builder": "Service Fee",
            "builder": "Service Fee",
        }

        def brokerage_label_for(role):
            if not role:
                return "Brokerage"
            return ROLE_BROKERAGE_LABELS.get(str(role).strip().lower(), "Brokerage")

        def display_value(p, field):
            val = getattr(p, field, "")
            if field in ['available_from', 'created_at', 'deleted_at'] and val:
                try:
                    return val.strftime('%Y-%m-%d')
                except AttributeError:
                    return val
            if field in BOOL_FIELDS:
                return "Yes" if val else "No"
            # NEW: listings created directly on the site (not via bulk
            # import) never get an upload_file_name — label them clearly
            # instead of leaving the cell blank.
            if field == "upload_file_name":
                return val if val else "Web Listing UI Form"
            return val if val is not None else ""

        all_cols = []
        for sec, fields in sections.items():
            all_cols.extend([(sec, f) for f in fields])

        if request.GET.get('download') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rental Residential"

            HDR_BG, REQ_BG, OPT_BG = "667EEA", "FEF3C7", "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

            sec_spans = OrderedDict()
            for i, (sec, _) in enumerate(all_cols):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            role_col = None
            brokerage_col = None

            for ci, (sec, field) in enumerate(all_cols, 1):
                req = field in REQUIRED
                label = FIELD_LABELS.get(field, field)
                lc = ws.cell(row=2, column=ci, value=label + (" *" if req else ""))
                lc.font = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border = bdr

                hc = ws.cell(row=3, column=ci, value="")
                hc.font = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(16, len(label) // 2 + 6)

                if field == "listed_by_role":
                    role_col = ci
                if field == "brokerage_percentage":
                    brokerage_col = ci

            if brokerage_col:
                note = (
                    "The LABEL text shown above this field on the live form changes based on\n"
                    "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
                    "Admin -> EstateFlow Service Fee\n"
                    "Relationship Manager -> Service Fee\n"
                    "Landlord -> Tenant Service Fee\n"
                    "Agent -> Brokerage\n"
                    "Agency/Builder or Builder -> Service Fee\n"
                    "Any other role -> Brokerage (default)\n\n"
                    "See the 'Brokerage Label' column at the end of this sheet for the resolved value per row."
                )
                ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            for row_idx, p in enumerate(properties, start=4):
                for col_idx, (sec, field) in enumerate(all_cols, 1):
                    if field == "sr_no":
                        val = row_idx - 3   # row 4 -> Sr No 1
                    else:
                        val = display_value(p, field)

                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr

            if role_col:
                preview_col = len(all_cols) + 1
                pc = ws.cell(row=2, column=preview_col, value="Brokerage Label")
                pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
                pc.fill = PatternFill("solid", fgColor="FEF3C7")
                pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                pc.border = bdr
                ws.cell(row=3, column=preview_col, value="Auto-resolved from role")
                ws.cell(row=3, column=preview_col).font = Font(italic=True, color="64748B", name="Arial", size=8)
                ws.cell(row=3, column=preview_col).border = bdr
                ws.column_dimensions[get_column_letter(preview_col)].width = 22

                for row_idx, p in enumerate(properties, start=4):
                    label = brokerage_label_for(getattr(p, "listed_by_role", ""))
                    fc = ws.cell(row=row_idx, column=preview_col, value=label)
                    fc.alignment = Alignment(horizontal="center", vertical="center")
                    fc.border = bdr

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Properties_Data.xlsx"'
            wb.save(response)
            return response

        elif request.GET.get('download') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Rental_Residential_Properties_Data.csv"'
            writer = csv.writer(response)

            row1 = []
            current_sec = ""
            for sec, _ in all_cols:
                if sec != current_sec:
                    row1.append(f" {sec}")
                    current_sec = sec
                else:
                    row1.append("")
            row1.append("📋 Brokerage")
            writer.writerow(row1)

            header_row = [FIELD_LABELS.get(field, field) + (" *" if field in REQUIRED else "") for _, field in all_cols]
            header_row.append("Brokerage Label")
            writer.writerow(header_row)

            hint_row = ["" for _, field in all_cols]
            hint_row.append("")
            writer.writerow(hint_row)

            for p in properties:
                data_row = [display_value(p, field) for _, field in all_cols]
                data_row.append(brokerage_label_for(getattr(p, "listed_by_role", "")))
                writer.writerow(data_row)

            return response

    # ═══════════════════════════════════════
    # STATS & UNIQUE DROPDOWN DATA
    # ═══════════════════════════════════════

    all_props = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)

    total_count = all_props.count()

    unique_listing_status = all_props.exclude(listing_status__isnull=True).exclude(listing_status='').values_list('listing_status', flat=True).distinct()
    unique_approval_status = all_props.exclude(approval_status__isnull=True).exclude(approval_status='').values_list('approval_status', flat=True).distinct()
    unique_bhk = all_props.exclude(bhk_type__isnull=True).exclude(bhk_type='').values_list('bhk_type', flat=True).distinct()
    unique_cities = all_props.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_states = all_props.exclude(state__isnull=True).exclude(state='').values_list('state', flat=True).distinct()  # NEW
    unique_furnish = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values_list('furnishing_status', flat=True).distinct()
    unique_possession = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').values_list('availability_status', flat=True).distinct()
    unique_property_types = all_props.exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    unique_listed_roles = all_props.exclude(listed_by_role__isnull=True).exclude(listed_by_role='').values_list('listed_by_role', flat=True).distinct()   # NEW split
    unique_uploaded_roles = all_props.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()  # NEW split

    active_count = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    furnished_count = all_props.filter(furnishing_status__iexact='Furnished').count()
    available_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    city_count = all_props.exclude(city__isnull=True).exclude(city='').values('city').distinct().count()
    state_count = all_props.exclude(state__isnull=True).exclude(state='').values('state').distinct().count()  # NEW

    # Duplicate properties
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count = all_props.filter(is_duplicate=False).count()

# Listing status breakdown
    active_listing_count = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count = all_props.filter(listing_status__iexact='Rented').count()

# Approval status breakdown
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count = all_props.filter(approval_status__iexact='Rejected').count()

    # ═══════════════════════════════════════
    # RENT STATS
    # ═══════════════════════════════════════

    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )

    avg_rent = rent_stats['avg_rent']
    max_rent = rent_stats['max_rent']
    min_rent = rent_stats['min_rent']

    deposit_stats = all_props.exclude(security_deposit_amount__isnull=True).aggregate(avg_deposit=Avg('security_deposit_amount'))
    avg_deposit = deposit_stats['avg_deposit']

    area_stats = all_props.exclude(built_up_area__isnull=True).aggregate(avg_area=Avg('built_up_area'))
    avg_area = area_stats['avg_area']

    with_owner_count = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    with_images_count = all_props.filter(images__isnull=False).distinct().count()

    uploaded_files = all_props.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    # ═══════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════

    renting_option_qs = all_props.exclude(renting_option__isnull=True).exclude(renting_option='').values('renting_option').annotate(count=Count('id')).order_by('-count')
    renting_option_labels = json.dumps([item['renting_option'] for item in renting_option_qs])
    renting_option_data = json.dumps([item['count'] for item in renting_option_qs])

    rent_buckets = [
        ('Under ₹5k', 0, 5000), ('₹5k–10k', 5000, 10000), ('₹10k–20k', 10000, 20000),
        ('₹20k–30k', 20000, 30000), ('₹30k–50k', 30000, 50000), ('₹50k–1L', 50000, 100000),
        ('Above ₹1L', 100000, 999999999),
    ]

    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data = json.dumps([all_props.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values('furnishing_status').annotate(count=Count('id')).order_by('-count')
    furnishing_labels = json.dumps([item['furnishing_status'] for item in furnish_qs])
    furnishing_data = json.dumps([item['count'] for item in furnish_qs])

    prop_type_qs = all_props.exclude(property_type__isnull=True).exclude(property_type='').values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels = json.dumps([item['property_type'] for item in prop_type_qs])
    prop_type_data = json.dumps([item['count'] for item in prop_type_qs])

    # ═══════════════════════════════════════
    # KPI
    # ═══════════════════════════════════════

    occupied_count = all_props.filter(availability_status__iexact='Occupied').count()
    vacant_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100), 1) if total_count > 0 else 0
    vacancy_rate = round((vacant_count / total_count * 100), 1) if total_count > 0 else 0
    total_revenue = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0

    ready_to_move_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    short_lease_count = all_props.filter(lease_duration__icontains='6').count()
    long_lease_count = all_props.filter(lease_duration__icontains='12').count()
    new_property_count = all_props.filter(property_age__icontains='New').count()
    old_property_count = all_props.exclude(property_age__icontains='New').count()
    premium_properties_count = all_props.filter(monthly_rent__gte=50000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=15000).count()

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════

    context = {
        'user_obj': user_obj,
        'page_obj': page_obj,

        'search_query': search_query,
        'bhk_query': bhk_query,
        'city_query': city_query,
        'state_query': state_query,          # NEW
        'furnish_query': furnish_query,
        'possession_query': possession_query,
        'prop_type_query': prop_type_query,
        'listed_by_query': listed_by_query,
        'uploaded_by_query': uploaded_by_query,   # NEW
        'budget_query': budget_query,

        'unique_cities': unique_cities,
        'unique_states': unique_states,        # NEW
        'unique_furnish': unique_furnish,
        'unique_possession': unique_possession,
        'unique_property_types': unique_property_types,
        'unique_listed_roles': unique_listed_roles,     # NEW (split)
        'unique_uploaded_roles': unique_uploaded_roles, # NEW (split)

        'filtered_count': filtered_count,

        'total_count': total_count,
        'active_count': active_count,
        'furnished_count': furnished_count,
        'available_count': available_count,
        'city_count': city_count,
        'state_count': state_count,          # NEW

        'from_date': from_date_str,
        'to_date': to_date_str,

        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,

        'avg_deposit': avg_deposit,
        'avg_area': avg_area,

        'with_owner_count': with_owner_count,
        'with_images_count': with_images_count,

        'uploaded_files': uploaded_files,

        'renting_option_labels': renting_option_labels,   # renamed from bhk_labels
        'renting_option_data': renting_option_data,

        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,

        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,

        'prop_type_labels': prop_type_labels,
        'prop_type_data': prop_type_data,

        'occupied_count': occupied_count,
        'vacant_count': vacant_count,

        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,

        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,

        'ready_to_move_count': ready_to_move_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,

        'new_property_count': new_property_count,
        'old_property_count': old_property_count,

        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'bhk_query': bhk_query,
        'unique_bhk': unique_bhk,                              # ADD

        'listing_status_query': listing_status_query,          # ADD
        'approval_status_query': approval_status_query,        # ADD
        'duplicate_query': duplicate_query,                    # ADD

        'unique_listing_status': unique_listing_status,        # ADD
        'unique_approval_status': unique_approval_status,      # ADD

        'duplicate_properties_count': duplicate_properties_count,  # ADD
        'unique_properties_count': unique_properties_count,        # ADD

        'active_listing_count': active_listing_count,          # ADD
        'inactive_listing_count': inactive_listing_count,      # ADD
        'sold_listing_count': sold_listing_count,               # ADD
        'rented_listing_count': rented_listing_count,           # ADD
        'pending_approval_count': pending_approval_count,       # ADD
        'approved_count': approved_count,                       # ADD
        'rejected_count': rejected_count,                       # ADD
    }

    return render(
        request,
        'rm_panel/Reports/Rental/residential_list.html',
        context
    )

def residential_rm(request):
    # 1. Retrieve identity from browser session
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (admin_id == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = user_obj

    

    user_role = user_obj.user_role

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

            # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION (Who submitted the HTML form) ----------
            if admin_obj:
                uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
                uploader_email = getattr(admin_obj, 'email', '')
                uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
                uploader_role = "Admin"
                uploader_id = f"ADMIN_{admin_id}"
            elif user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "User"
                uploader_id = f"USER_{user_id}"
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION (Who owns/manages the listing) ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE (Checking LISTED BY, not UPLOADED BY)
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality_area') or request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_property_fingerprint(
                input_property_no, 
                input_building_name, 
                input_locality, 
                input_pincode
            )

            # 1. Direct Case-Insensitive Query for same unit in same locality/building
            direct_duplicates = RentalResidentialProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality_area__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            # Combine fingerprint match OR direct field match
            existing_duplicates = (
                RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                # Level 1: Hard Block if LISTED BY the exact same person (ID, Email, OR Phone match)
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        # INSTANT REJECTION: Same agent/owner cannot list the same property twice
                        messages.error(
                            request, 
                            f"Duplicate Blocked: This property (Unit {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        )
                        return redirect('residential')

                # Level 2: Different Agent/User listing the exact same physical unit -> Allow save & Flag
                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = RentalResidentialProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Residential",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                property_no=input_property_no,
                bhk_type=request.POST.get('bhk_type'),
                renting_option=request.POST.get('renting_option'),
                built_up_area=to_decimal(request.POST.get('built_up_area')),
                bathrooms=to_int(request.POST.get('bathrooms')),
                balconies=to_int(request.POST.get('balconies')),
                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int(request.POST.get('total_floors')),
                facing_direction=request.POST.get('facing_direction', request.POST.get('facing')),
                furnishing_status=request.POST.get('furnishing_status'),
                available_for=request.POST.get('available_for'),

                carpet_area=to_decimal(request.POST.get('carpet_area')),
                city_zone=request.POST.get('city_zone', request.POST.get('zone')),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition', request.POST.get('construction_status')),
                property_age=request.POST.get('property_age'),
                wing_number=request.POST.get('wing_number'),
                building_name=input_building_name,

                availability_status=request.POST.get('availability_status', request.POST.get('possession_status')),
                available_from=available_from,
                lease_duration=request.POST.get('lease_duration'),
                brokerage_percentage=request.POST.get('brokerage_percentage', request.POST.get('brokerage')),
                manual_brokerage=request.POST.get('manual_brokerage'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=to_int(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int(request.POST.get('security_deposit_amount', request.POST.get('security_deposit'))),
                maintenance_type=request.POST.get('maintenance_type'),
                monthy_maintenance_amount=to_int(request.POST.get('monthy_maintenance_amount', request.POST.get('maintenance_amount'))),
                total_move_in_cost=to_int(request.POST.get('total_move_in_cost')),

                address=request.POST.get('address'),
                city=request.POST.get('city'),
                locality_area=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=input_pincode,
                main_road_connectivity=request.POST.get('main_road_connectivity', request.POST.get('road_connectivity')),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                amenities=amenities,
                nearby_facilities=facilities,

                user_description=request.POST.get('user_description'),
                description=request.POST.get('description'),
                rent_residential_desc=request.POST.get('rent_residential_desc'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC ----------
            images = request.FILES.getlist('property_images[]')
            for index, img in enumerate(images[:10]):
                RentalResidentialImage.objects.create(
                    property=prop, 
                    image=img,
                    sequence_order=index
                )

            messages.success(request, "Property Added Successfully ")
            
            return redirect('residential_rm_list')

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            messages.error(request, f"Error while saving listing: {str(e)}")
            return redirect('residential_rm')
    
    context = {
        'user_obj': user_obj,
        'user_role': user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Rental/residential.html", context)

############# Views end for rental forms for RM #######################


############### Views start for rental edit forms for RM ###################

def residential_rm_edit(request, pk):
    prop = get_object_or_404(RentalResidentialProperty, id=pk)

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    if request.method == 'POST':
        try:
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # =====================================================
            # AUDIT LOGIC: Snapshot current state BEFORE modification
            # =====================================================
            tracked_fields_list = [
                'property_type', 'property_no', 'bhk_type', 'renting_option',
                'built_up_area', 'bathrooms', 'balconies', 'building_configuration',
                'total_floors', 'facing_direction', 'furnishing_status', 'available_for',
                'carpet_area', 'city_zone', 'ownership_type', 'property_condition',
                'property_age', 'wing_number', 'building_name',
                'availability_status', 'available_from', 'lease_duration',
                'brokerage_percentage', 'manual_brokerage', 'monthly_rent',
                'advance_rent_month', 'advance_rent_amount',
                'security_deposit_type', 'security_deposit_amount',
                'maintenance_type', 'monthy_maintenance_amount', 'total_move_in_cost',
                'address', 'city', 'locality_area', 'property_landmark', 'state', 'pincode',
                'main_road_connectivity', 'google_maps_link', 'latitude', 'longitude',
                'listed_by_id', 'listed_by_name', 'listed_by_email', 'listed_by_contact', 'listed_by_role',
                'listing_status', 'approval_status',
            ]
            old_state_snapshot = {field: str(getattr(prop, field, '')) for field in tracked_fields_list}

           
            # =====================================================
            prop.listed_by_type = request.POST.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to = request.POST.get('assigned_to', prop.assigned_to)
            # assigned_to_role is posted alongside assigned_to (mirrors the Add form)
            # and is available here if you want to store it on the model as well, e.g.:
            # prop.assigned_to_role = request.POST.get('assigned_to_role', getattr(prop, 'assigned_to_role', ''))
            prop.listed_by_id = (request.POST.get('listed_by_id') or prop.listed_by_id or '').strip()
            prop.listed_by_name = (request.POST.get('listed_by_name') or prop.listed_by_name or '').strip()
            prop.listed_by_email = (request.POST.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = (request.POST.get('listed_by_contact') or prop.listed_by_contact or '').strip()
            prop.listed_by_role = (request.POST.get('listed_by_role') or prop.listed_by_role or '').strip()

            # =====================================================
            # 2. BASIC INFORMATION
            # =====================================================
            prop.property_type = request.POST.get('property_type')
            prop.property_no = request.POST.get('property_no', prop.property_no)
            prop.bhk_type = request.POST.get('bhk_type')
            prop.renting_option = request.POST.get('renting_option')
            prop.built_up_area = to_decimal(request.POST.get('built_up_area'))
            prop.bathrooms = to_int(request.POST.get('bathrooms'))
            prop.balconies = to_int(request.POST.get('balconies'))
            prop.building_configuration = request.POST.get('building_configuration')
            prop.total_floors = to_int(request.POST.get('total_floors'))
            prop.facing_direction = request.POST.get('facing_direction')
            prop.furnishing_status = request.POST.get('furnishing_status')
            prop.available_for = request.POST.get('available_for')

            # =====================================================
            # 3. PROPERTY DETAILS
            # =====================================================
            prop.carpet_area = to_decimal(request.POST.get('carpet_area'))
            prop.city_zone = request.POST.get('city_zone')
            prop.ownership_type = request.POST.get('ownership_type')
            prop.property_condition = request.POST.get('property_condition')
            prop.property_age = request.POST.get('property_age')
            prop.wing_number = request.POST.get('wing_number')
            prop.building_name = request.POST.get('building_name')

            # =====================================================
            # 4. AVAILABILITY DETAILS
            # =====================================================
            prop.availability_status = request.POST.get('availability_status')

            available_from_raw = request.POST.get('available_from')
            if available_from_raw and available_from_raw.strip():
                try:
                    prop.available_from = datetime.strptime(available_from_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    prop.available_from = None
            else:
                prop.available_from = None

            prop.lease_duration = request.POST.get('lease_duration')
            prop.brokerage_percentage = request.POST.get('brokerage_percentage')
            prop.manual_brokerage = request.POST.get('manual_brokerage')

            # =====================================================
            # 5. PRICING DETAILS
            # =====================================================
            prop.monthly_rent = to_int(request.POST.get('monthly_rent'))
            prop.advance_rent_month = request.POST.get('advance_rent_month')
            prop.advance_rent_amount = to_int(request.POST.get('advance_rent_amount'))
            prop.security_deposit_type = request.POST.get('security_deposit_type')
            prop.security_deposit_amount = to_int(request.POST.get('security_deposit_amount'))
            prop.maintenance_type = request.POST.get('maintenance_type')
            prop.monthy_maintenance_amount = to_int(request.POST.get('monthy_maintenance_amount'))
            prop.total_move_in_cost = to_int(request.POST.get('total_move_in_cost'))

            # =====================================================
            # 6. LOCATION DETAILS
            # =====================================================
            prop.address = request.POST.get('address')
            prop.city = request.POST.get('city')
            prop.locality_area = request.POST.get('locality_area')
            prop.property_landmark = request.POST.get('property_landmark')
            prop.state = request.POST.get('state')
            prop.pincode = request.POST.get('pincode')
            prop.main_road_connectivity = request.POST.get('main_road_connectivity')
            prop.google_maps_link = request.POST.get('google_maps_link')
            prop.latitude = request.POST.get('latitude')
            prop.longitude = request.POST.get('longitude')

            # =====================================================
            # 7. AMENITIES & FACILITIES
            # (Unchanged — the modal-picker in the template still posts the exact
            # same field names amenities[] / nearby_facilities[], so no view
            # changes were required here.)
            # =====================================================
            prop.amenities = ",".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]'))

            # =====================================================
            # 8. DESCRIPTIONS
            # =====================================================
            prop.user_description = request.POST.get('user_description')

            # =====================================================
            # 9. LISTED ELSEWHERE
            # =====================================================
            prop.listed_elsewhere = request.POST.get('listed_elsewhere', 'No')
            prop.portal_name = request.POST.get('portal_name')

            # =====================================================
            # 10. LISTING STATUS & APPROVAL STATUS
            # =====================================================
            prop.listing_status = request.POST.get('listing_status', prop.listing_status)
            prop.approval_status = request.POST.get('approval_status', prop.approval_status)

            # description / rent_residential_desc are auto-regenerated in model.save()
            # via generate_auto_descriptions() — no need to set them from POST.

            prop.save()

            # =====================================================
            # 11. IMAGE UPLOAD & SEQUENCE LOGIC
            # =====================================================
            images = request.FILES.getlist('property_images[]')
            current_count = prop.images.count()
            for img in images:
                if current_count < 10:
                    RentalResidentialImage.objects.create(
                        property=prop, image=img, sequence_order=current_count
                    )
                    current_count += 1

            # =====================================================
            # AUDIT LOGIC: Diff dictionary
            # =====================================================
            modifications_diff = {}
            modified_fields_summary = []

            for field in tracked_fields_list:
                new_val = str(getattr(prop, field, ''))
                if old_state_snapshot[field] != new_val:
                    modifications_diff[field] = {
                        "old_value": old_state_snapshot[field],
                        "new_value": new_val
                    }
                    modified_fields_summary.append(field)

            # if modified_fields_summary:
            #     RentalActivityLog.objects.create(
            #         user_identity=admin_obj.email or admin_obj.username,
            #         user_role="Admin",
            #         action_type='UPDATE',
            #         property_id=prop.id,
            #         targeted_fields=", ".join(modified_fields_summary[:4]) + ("..." if len(modified_fields_summary) > 4 else ""),
            #         associated_file=prop.upload_file_name or "Web UI Form",
            #         action_payload=json.dumps(modifications_diff),
            #         ip_address=_get_client_ip(request),
            #         status='SUCCESS'
            #     )

            return JsonResponse({
                'status': 'success',
                'message': 'Property updated successfully!',
                'redirect_url': reverse('residential_rm_list')
            })

        except Exception as e:
            # RentalActivityLog.objects.create(
            #     user_identity=admin_obj.email if 'admin_obj' in locals() else "Unknown Session",
            #     user_role="Admin",
            #     action_type='UPDATE',
            #     property_id=pk,
            #     targeted_fields="System Errors",
            #     action_payload=f"Failed modification update runtime sequence execution target: {str(e)}",
            #     ip_address=_get_client_ip(request),
            #     status='FAILED'
            # )
            return JsonResponse({
                'status': 'error',
                'message': f"Failed to save data: {str(e)}"
            }, status=400)

    return render(request, 'rm_panel/Forms/Rental/residential_edit.html', {
        'property': prop,
        'user_obj': user_obj,
      
        
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all()
    })

########### Views end for rental edit forms for RM ###########################


############# Views start for commercial list for RM ################

def commercial_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Reports/Rental/commercial_list.html", context)

############ Views end for commercial list for RM #######################


########## Views start for commercial forms for RM #####################

def commercial_rm(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Rental/commercial.html", context)

########### Views end for commercial forms for RM ########################


############# Views start for pg forms for RM ########################

def pg_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

     # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Reports/Rental/pg_list.html", context)

############ Views end for pg forms for RM ############################


########## Views start for pg forms for RM #############################

def pg_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Rental/pg_coliving.html", context)

############# Views end for pg forms for RM ############################


############# Views start for resale property list ####################

def residential_resale_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    all_properties = (
        ResaleResidentialProperty.objects
        .filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
        .prefetch_related('images')
        .order_by('-created_at')
    )

    properties = all_properties

    ############# Filter Parameters Start ###################################

    # ── Read query params ────────────────────────────────────────────────────
    search_query    = request.GET.get('search', '').strip()
    prop_type       = request.GET.get('prop_type', '').strip()
    bhk_filter      = request.GET.get('bhk', '').strip()
    furnish         = request.GET.get('furnish', '').strip()
    occupancy       = request.GET.get('occupancy', '').strip()
    ownership       = request.GET.get('ownership', '').strip()
    negotiable      = request.GET.get('negotiable', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()
    listed_by_type  = request.GET.get('listed_by_type', '').strip()
    property_age    = request.GET.get('property_age', '').strip()
    facing_direction = request.GET.get('facing_direction', '').strip()
    society_type    = request.GET.get('society_type', '').strip()
    city_filter     = request.GET.get('city', '').strip()
    status_filter   = request.GET.get('listing_status', '').strip()

    # ── Apply filters ────────────────────────────────────────────────────────
    properties = all_properties

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query)           |
            Q(locality__icontains=search_query)       |
            Q(listed_by_name__icontains=search_query) |
            Q(bhk__icontains=search_query)            |
            Q(building_name__icontains=search_query)  |
            Q(property_no__icontains=search_query)    |
            Q(address__icontains=search_query)
        )

    if prop_type:
        properties = properties.filter(property_type=prop_type)

    if bhk_filter:
        properties = properties.filter(bhk=bhk_filter)

    if furnish:
        properties = properties.filter(furnishing_status=furnish)

    if occupancy:
        properties = properties.filter(occupancy_status=occupancy)

    if ownership:
        properties = properties.filter(ownership_status=ownership)

    if negotiable:
        properties = properties.filter(price_negotiable=negotiable)

    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)

    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    if listed_by_type:
        properties = properties.filter(listed_by_type=listed_by_type)

    if property_age:
        properties = properties.filter(property_age=property_age)

    if facing_direction:
        properties = properties.filter(facing_direction=facing_direction)

    if society_type:
        properties = properties.filter(society_type=society_type)

    if city_filter:
        properties = properties.filter(city__icontains=city_filter)

    if status_filter:
        properties = properties.filter(listing_status=status_filter)

    # ── Thumbnail + helper attributes for each filtered property ─────────────
    for prop in properties:
        prop.thumbnail = prop.images.first()

        prop.nearby_facilities_list = (
            [f.strip() for f in prop.nearby_facilities.split(',')]
            if prop.nearby_facilities else []
        )
        prop.amenities_list = (
            [a.strip() for a in prop.amenities.split(',')]
            if prop.amenities else []
        )
        prop.image_count = prop.images.count()
        prop.image_urls = [img.image.url for img in prop.images.all()]

    # ════════════════════════════════════════════════════════════════════════
    # KPI STATS (Calculated using correct DB layout keys)
    # ════════════════════════════════════════════════════════════════════════
    total_count = all_properties.count()

    # ── Row 1 — Inventory ────────────────────────────────────────────────────
    total_negotiable  = all_properties.filter(price_negotiable='yes').count()
    total_furnished   = all_properties.filter(furnishing_status='fully').count()
    total_freehold    = all_properties.filter(ownership_status='Self Owned').count()
    total_with_images = all_properties.filter(images__isnull=False).distinct().count()
    total_with_vacant = all_properties.filter(occupancy_status='Vacant').count()
    total_with_tenants = all_properties.filter(existing_tenants='yes').count()

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    negotiable_pct = pct(total_negotiable, total_count)
    furnished_pct  = pct(total_furnished, total_count)
    freehold_pct   = pct(total_freehold, total_count)
    images_pct     = pct(total_with_images, total_count)
    vacant_pct     = pct(total_with_vacant, total_count)

    # ── Row 2 — Pricing ──────────────────────────────────────────────────────
    price_agg = all_properties.aggregate(
        avg      = Avg('selling_price'),
        max_val  = Max('selling_price'),
        min_val  = Min('selling_price'),
        avg_sqft = Avg('price_per_sqft'),
        avg_area = Avg('builtup_area'),
        total_value = Coalesce(Sum('selling_price'), Decimal('0.00')),
    )
    avg_price      = price_agg['avg']
    max_price      = price_agg['max_val']
    min_price      = price_agg['min_val']
    avg_price_sqft = price_agg['avg_sqft']
    avg_builtup    = price_agg['avg_area']
    total_value    = price_agg['total_value']
    total_with_loan = all_properties.filter(property_loan='yes').count()

    # ── Row 3 — Legal & Status ───────────────────────────────────────────────
    no_dispute_count  = all_properties.filter(any_legal_dispute='no').count()
    dispute_count     = all_properties.filter(any_legal_dispute='yes').count()
    tax_pending_count = all_properties.filter(government_tax='yes').count()
    tenant_occupied   = all_properties.filter(existing_tenants='yes').count()
    premium_count     = all_properties.filter(selling_price__gte=10000000).count()  # >= 1 Cr
    budget_count      = all_properties.filter(selling_price__lt=3000000).count()    # < 30 L

    # ── Row 4 — Listing Quality ──────────────────────────────────────────────
    with_video_count = (
        all_properties
        .exclude(property_video__isnull=True)
        .exclude(property_video='')
        .count()
    )
    with_floor_plan = (
        all_properties
        .exclude(floor_plan__isnull=True)
        .exclude(floor_plan='')
        .count()
    )
    with_listed_by = (
        all_properties
        .exclude(listed_by_name__isnull=True)
        .exclude(listed_by_name='')
        .count()
    )
    with_uploaded_by = (
        all_properties
        .exclude(uploaded_by_name__isnull=True)
        .exclude(uploaded_by_name='')
        .count()
    )
    with_description = (
        all_properties
        .exclude(property_description__isnull=True)
        .exclude(property_description='')
        .count()
    )
    pending_approval = all_properties.filter(approval_status='Pending').count()
    approved_count   = all_properties.filter(approval_status='Approved').count()
    rejected_count   = all_properties.filter(approval_status='Rejected').count()

    # ── Charts ───────────────────────────────────────────────────────────────
    property_type_counts = dict(
        all_properties.values('property_type')
        .annotate(count=Count('id'))
        .values_list('property_type', 'count')
    )
    
    bhk_counts = dict(
        all_properties.values('bhk')
        .annotate(count=Count('id'))
        .values_list('bhk', 'count')
    )
    
    furnishing_counts = {
        'fully_furnished': all_properties.filter(furnishing_status='fully').count(),
        'semi_furnished': all_properties.filter(furnishing_status='semi').count(),
        'unfurnished': all_properties.filter(furnishing_status='unfurnished').count(),
    }
    
    occupancy_counts = dict(
        all_properties.values('occupancy_status')
        .annotate(count=Count('id'))
        .values_list('occupancy_status', 'count')
    )
    
    facing_counts = dict(
        all_properties.values('facing_direction')
        .annotate(count=Count('id'))
        .values_list('facing_direction', 'count')
    )
    
    society_type_counts = dict(
        all_properties.values('society_type')
        .annotate(count=Count('id'))
        .values_list('society_type', 'count')
    )
    
    ownership_counts = dict(
        all_properties.values('ownership_status')
        .annotate(count=Count('id'))
        .values_list('ownership_status', 'count')
    )
    
    listed_by_counts = dict(
        all_properties.values('listed_by_type')
        .annotate(count=Count('id'))
        .values_list('listed_by_type', 'count')
    )
    
    approval_status_counts = dict(
        all_properties.values('approval_status')
        .annotate(count=Count('id'))
        .values_list('approval_status', 'count')
    )

    # ── Monthly Property Listings (for trend chart) ──────────────────────────
    from django.db.models.functions import TruncMonth
    monthly_listings = (
        all_properties
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_labels = [m['month'].strftime('%b %Y') for m in monthly_listings] if monthly_listings else []
    monthly_counts = [m['count'] for m in monthly_listings] if monthly_listings else []

    # ── Unique values for Select2 searchable dropdowns ───────────────────────
    unique_prop_types = list(
        all_properties.values_list('property_type', flat=True)
        .distinct().order_by('property_type')
    )
    unique_bhk_values = list(
        all_properties.values_list('bhk', flat=True)
        .distinct().order_by('bhk')
    )
    unique_furnishings = list(
        all_properties.values_list('furnishing_status', flat=True)
        .distinct().order_by('furnishing_status')
    )
    unique_occupancies = list(
        all_properties.values_list('occupancy_status', flat=True)
        .distinct().order_by('occupancy_status')
    )
    unique_ownerships = list(
        all_properties.values_list('ownership_status', flat=True)
        .distinct().order_by('ownership_status')
    )
    unique_cities = list(
        all_properties.values_list('city', flat=True)
        .distinct().order_by('city')
    )
    unique_states = list(
        all_properties.values_list('state', flat=True)
        .distinct().order_by('state')
    )
    unique_facings = list(
        all_properties.values_list('facing_direction', flat=True)
        .distinct().order_by('facing_direction')
    )
    unique_society_types = list(
        all_properties.values_list('society_type', flat=True)
        .distinct().order_by('society_type')
    )

    # ── Uploaded Files (for tracking) ──────────────────────────────────────
    try:
        uploaded_files = (
            all_properties
            .exclude(upload_file_name__isnull=True)
            .exclude(upload_file_name='')
            .values_list('upload_file_name', flat=True)
            .distinct()
        )
    except Exception:
        uploaded_files = []
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'properties':properties,

        'filtered_count': properties.count(),
        'total_count': total_count,

        # Active search params
        'search_query': search_query,
        'prop_type_query': prop_type,
        'bhk_query': bhk_filter,
        'furnish_query': furnish,
        'occupancy_query': occupancy,
        'ownership_query': ownership,
        'negotiable_query': negotiable,
        'from_date': from_date,
        'to_date': to_date,
        'listed_by_type_query': listed_by_type,
        'property_age_query': property_age,
        'facing_direction_query': facing_direction,
        'society_type_query': society_type,
        'city_filter': city_filter,
        'status_filter': status_filter,

        # Row 1 — Inventory
        'total_negotiable': total_negotiable,
        'total_furnished': total_furnished,
        'total_freehold': total_freehold,
        'total_with_images': total_with_images,
        'total_with_vacant': total_with_vacant,
        'total_with_tenants': total_with_tenants,
        'negotiable_pct': negotiable_pct,
        'furnished_pct': furnished_pct,
        'freehold_pct': freehold_pct,
        'images_pct': images_pct,
        'vacant_pct': vacant_pct,

        # Row 2 — Pricing
        'avg_price': avg_price,
        'max_price': max_price,
        'min_price': min_price,
        'avg_price_sqft': avg_price_sqft,
        'avg_builtup': avg_builtup,
        'total_value': total_value,
        'total_with_loan': total_with_loan,

        # Row 3 — Legal
        'no_dispute_count': no_dispute_count,
        'dispute_count': dispute_count,
        'tax_pending_count': tax_pending_count,
        'tenant_occupied': tenant_occupied,
        'premium_count': premium_count,
        'budget_count': budget_count,
        'pending_approval': pending_approval,
        'approved_count': approved_count,
        'rejected_count': rejected_count,

        # Row 4 — Quality
        'with_video_count': with_video_count,
        'with_floor_plan': with_floor_plan,
        'with_listed_by': with_listed_by,
        'with_uploaded_by': with_uploaded_by,
        'with_description': with_description,

        # Charts
        'property_type_counts': property_type_counts,
        'bhk_counts': bhk_counts,
        'furnishing_counts': furnishing_counts,
        'occupancy_counts': occupancy_counts,
        'facing_counts': facing_counts,
        'society_type_counts': society_type_counts,
        'ownership_counts': ownership_counts,
        'listed_by_counts': listed_by_counts,
        'approval_status_counts': approval_status_counts,
        'monthly_labels': monthly_labels,
        'monthly_counts': monthly_counts,

        # Select2 unique options
        'unique_prop_types': unique_prop_types,
        'unique_bhk_values': unique_bhk_values,
        'unique_furnishings': unique_furnishings,
        'unique_occupancies': unique_occupancies,
        'unique_ownerships': unique_ownerships,
        'unique_cities': unique_cities,
        'unique_states': unique_states,
        'unique_facings': unique_facings,
        'unique_society_types': unique_society_types,

        'uploaded_files': uploaded_files,
    }
    
    return render(request, "rm_panel/Reports/Resale/residential_resale_list.html", context)


########### Views start for resale property form residential ################

def residential_resale_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/residential_resale.html", context)

########## Views end for resale property form residential ########################


########### Views start for update resale property form residential #################

def residential_resale_rm_update(request,id):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    property = ResaleResidentialProperty.objects.get(id=id)

    existing_images = property.images.all()

    facilities_obj = Facilities_Details.objects.all()
    ameneties_obj = Ameneties_Details.objects.all()

    amenities_list = []
    if property.amenities:
        amenities_list = [a.strip() for a in property.amenities.split(',') if a.strip()]
    
    # Parse facilities from stored comma-separated string
    nearby_facilities_list = []
    if property.nearby_facilities:
        nearby_facilities_list = [f.strip() for f in property.nearby_facilities.split(',') if f.strip()]

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj,
        'prop':property,
        'existing_images':existing_images,
        'facilities_obj':facilities_obj,
        'ameneties_obj':ameneties_obj,
        'amenities_list':amenities_list,
        'nearby_facilities_list':nearby_facilities_list,
    }
    
    return render(request, "rm_panel/Forms/Resale/residential_resale_edit.html", context)

############# Views end for update resale property form residential ###################


############ Views start for resale commercial property list ##################

def commercial_resale_rm_list(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Reports/Resale/commercial_list.html", context)

########### Views end for resale commercial property list ####################


########## Views start for resale property for commercial ####################

def commercial_resale_rm(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/commercial_resale.html", context)

######### Views end for resale property for commercial ########################


########## Views start  for plot resale property list ######################

def plot_resale_rm_list(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Reports/Resale/plot_list.html", context)

######### Views end for plot resale property list ###################


############# Views start for resale property from plot ##################

def plot_resale_rm(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Forms/Resale/plot_resale.html", context)

######### Views end for resale property from plot ########################


######### Views start for residential plot list from rm ###################

def plot_resale_res_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
    }
    
    return render(request, "rm_panel/Reports/Resale/Plots/res_plots_list.html", context)

########### Views end for residential plot list from rm ########################


############### Views start for residential plot list form ####################

def plot_resale_res_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/Plots/res_plots.html", context)

############# Views end for residential plot list form ###########################


########### Views start for commercial plot list for rm ########################

def plot_resale_comm_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
    }
    
    return render(request, "rm_panel/Reports/Resale/Plots/comm_plots_list.html", context)

######## Views end for commercial plot list for rm ###########################


############## Views start for commercial plot for rm ####################

def plot_resale_comm_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/Plots/comm_plots.html", context)

############ Views end for commercial plot for rm #######################


############ Views start for industrial plot list for rm ##################

def plot_resale_ind_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
    }
    
    return render(request, "rm_panel/Reports/Resale/Plots/ind_plots_list.html", context)

############ Views end for industrial plot list for rm #########################


############# Views start for industrial plot for rm ###########################

def plot_resale_ind_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/Plots/ind_plots.html", context)

############# Views end for industrial plot for rm #############################


########### Views start for industrial property list #########################

def industry_resale_rm_list(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
    }
    
    return render(request, "rm_panel/Reports/Resale/industrial_list.html", context)

############ Views end for industrial property list ############################


########### Views start for industrial resale property form #####################

def industry_resale_rm(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/industrial_resale.html", context)

############### Views end for industrial resale property form ######################


############## Views start for agricultural plot list for rm ####################

def plot_resale_agri_rm_list(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
    }
    
    return render(request, "rm_panel/Reports/Resale/Plots/agri_plots_list.html", context)

############## Views end for agricultural plot list for rm #####################


############ Views start for agricultural plot form #####################

def plot_resale_agri_rm(request):
    # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()


    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/Plots/agri_plots.html", context)

############# Views end for agricultural plot form ############################


############ Views start for agricultural resale property list #####################

def agriculture_resale_rm_list(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm
    }
    
    return render(request, "rm_panel/Reports/Resale/agricultural_list.html", context)

########## Views end for agricultural resale property list ##############################


############ Views start for agricultural resale property form ######################

def agriculture_resale_rm(request):
     # 1. Retrieve identity from browser session
    session_user_id = request.session.get('User_id')
    logged_in_role = request.session.get('user_type')

    #  2. UPDATED SECURITY CHECK
    
    is_valid_rm = (session_user_id and logged_in_role == "Relationship Manager")
    is_valid_admin = (logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_rm and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if logged_in_role == "Admin":
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = session_user_id 

    # 3. Data Fetching: Get the full user object for the template
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_rm = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_rm':enquiry_obj_rm,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "rm_panel/Forms/Resale/agricultural_resale.html", context)

############# Views end for agricultural resale property form ########################


@login_required
def user_affiliate_links(request):
    links = AffiliateLink.objects.filter(owner=request.user).order_by('-created_at')
    return render(request, 'rm_panel/user_links.html', {'links': links})

@login_required
@user_passes_test(is_staff)
def admin_affiliate_detail(request, code):
    aff = get_object_or_404(AffiliateLink, code=code)
    visits = aff.visits.all().order_by('-created_at')[:100]
    conversions = aff.conversions.all().order_by('-created_at')[:100]
    return render(request, 'rm_panel/admin_affiliate_detail.html', {'affiliate': aff, 'visits': visits, 'conversions': conversions})




# in your orders/views.py (example)
from RM_Dashboard.models import AffiliateLink, ReferralConversion
from RM_Dashboard.utils import calculate_commission_for
from decimal import Decimal
from django.db import models

def checkout_complete(request):
    # ... your order save logic here ...
    order = ...  # saved order instance
    ref_code = request.POST.get('referral_code') or request.session.get('referral_code') or request.COOKIES.get('referral')
    if ref_code:
        try:
            aff = AffiliateLink.objects.get(code=ref_code, active=True)
            amount = Decimal(getattr(order, 'total_amount', 0) or 0)
            comm = calculate_commission_for(aff, amount)
            ReferralConversion.objects.create(
                affiliate=aff,
                user=request.user if request.user.is_authenticated else None,
                order_id=str(order.pk),
                amount=amount,
                commission_amount=comm,
                status='pending'
            )
            AffiliateLink.objects.filter(pk=aff.pk).update(
                registration_count=models.F('registration_count')+1,
                conversion_count=models.F('conversion_count')+1,
                commission_total=models.F('commission_total')+comm
            )
        except AffiliateLink.DoesNotExist:
            pass
    # continue response




def affilate_page(request):
    return render(request,"rm_panel/affilate_page.html")
