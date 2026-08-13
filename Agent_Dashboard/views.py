from django.shortcuts import render,HttpResponse
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Agent_Dashboard .models import *
from django.contrib.auth.decorators import login_required
from Admin_App.models import *
from Landlord_Panel.views import calculate_profile_strength
from CRM_Panel.models import *
from django.template.loader import render_to_string
from django.db.models import Count, Avg, Max, Min, Q,Sum
from django.db.models.functions import Coalesce
from datetime import datetime
import json
from django.http import JsonResponse
from django.urls import reverse
import hashlib
import re
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from datetime import datetime, date
from django.db.models import Prefetch


from collections import defaultdict
from django.db.models.functions import TruncMonth

from django.shortcuts import render,HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Admin_App .models import *
from Main_App .models import *
from seo .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from openpyxl import load_workbook
from django.template.loader import render_to_string
import traceback
import json
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator  # ← ADD THIS
import csv
import csv
import json
from django.db.models import Count, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch
from django.utils.dateparse import parse_date

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect
from django.shortcuts import render,HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Admin_App .models import *
from Main_App .models import *
from seo .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from openpyxl import load_workbook
from django.template.loader import render_to_string
import traceback
import json
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator  # ← ADD THIS
import csv
import csv
import json
from django.db.models import Count, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect

# Create your views here.

########### Crime Officer Views#######

def generate_property_fingerprint(property_no, building_name, locality, pincode=""):
    """
    Generates a normalized SHA-256 hash unique key based on physical location.
    Strips spaces and special characters for consistent matching.
    """
    def clean_text(val):
        if not val:
            return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()

    raw_string = f"{clean_text(property_no)}_{clean_text(building_name)}_{clean_text(locality)}_{clean_text(pincode)}"
    return hashlib.sha256(raw_string.encode('utf-8')).hexdigest()


def Wallet_Recharge_agent(request):
    return render(request,"agent/Wallet_Recharge_agent.html") 


def agent_dashboard(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # Calculate profile strength based on the swapped user object
    completion_score = calculate_profile_strength(user_obj)

    enquiry_obj_agent = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        # Pass the object's actual role so the template renders the Agent UI correctly
        'user_role': user_obj.user_role,
        'profile_completion_percentage': completion_score,
        'enquiry_obj_agent':enquiry_obj_agent
    }
    
    return render(request, "agent/agent_dashboard.html", context) 


############## Views start for update agent profile #########################

def Update_Profile_Agent(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_agent = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_agent':enquiry_obj_agent
    }
    
    return render(request, "agent/Profile/agent_profile.html", context)

############# Views end for update agent profile ##############################


############ Views start for assign enquiries to agent ########################

def Assign_Enquiry_Agent(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    #Calculate profile strength based on the swapped user object

    completion_score = calculate_profile_strength(user_obj)

    enquiry_obj_agent = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()


    enquiry_obj = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).order_by('-id')
    enquiry_obj_count = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    ############## Enquiries Stats By Source ##############################

    fb_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="facebook",assigned_to__id=user_obj.id).count()
    insta_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="instagram",assigned_to__id=user_obj.id).count()
    whatsapp_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="whatsapp",assigned_to__id=user_obj.id).count()
    google_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="google",assigned_to__id=user_obj.id).count()
    linkedin_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="linkedin",assigned_to__id=user_obj.id).count()
    twitter_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="twitter",assigned_to__id=user_obj.id).count()
    youtube_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="youtube",assigned_to__id=user_obj.id).count()
    referral_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="referral",assigned_to__id=user_obj.id).count()

    ########### Enquiry Stats by lead source ################################

    pending_obj_count = PropertyEnquiry.objects.filter(lead_status="Pending",assigned_to__id=user_obj.id).count()
    progress_obj_count = PropertyEnquiry.objects.filter(lead_status="In Progress",assigned_to__id=user_obj.id).count()
    hold_obj_count = PropertyEnquiry.objects.filter(lead_status="Hold",assigned_to__id=user_obj.id).count()
    closed_obj_count = PropertyEnquiry.objects.filter(lead_status="Closed",assigned_to__id=user_obj.id).count()
    cancelled_obj_count = PropertyEnquiry.objects.filter(lead_status="Cancelled",assigned_to__id=user_obj.id).count()

    rendered = render_to_string("agent/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'property_enquiry_list':rendered,
        'enquiry_obj_agent':enquiry_obj_agent,
        'profile_completion_percentage': completion_score
    }
    
    return render(request, "agent/Enquiry/assign_enquiry.html", context)

############## Views end for assign enquiries to agent #########################


########### Views start for update property enquiry ##########################

def update_enquiry_agent(request,id):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry = PropertyEnquiry.objects.get(id=id)

    enquiry_obj_agent = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_agent':enquiry_obj_agent,
        'enquiry':enquiry
    }
    
    return render(request, "agent/Enquiry/update_enquiry.html", context)

############ Views end for update property enquiry #########################



def chat_agent(request):
    return render(request,"agent/chat_agent.html")



def Sponserproperty(request):
    return render(request,"agent/Sponserproperty.html")




def lead_purchase_create(request):
    if request.method == "POST":
        lead_id = request.POST.get("lead_id")
        wallet_deduction = request.POST.get("wallet_deduction")

        if not lead_id or not wallet_deduction:
            messages.error(request, "All fields are required.")
            return redirect("lead_purchase_create")

        LeadPurchase.objects.create(
            lead_id=lead_id,
            wallet_deduction=wallet_deduction,
        )
        messages.success(request, f"Lead {lead_id} purchased successfully!")
       # return redirect("lead_purchase_list")

    return render(request, "agent/lead_purchase.html")


def lead_purchase_list(request):
    purchases = LeadPurchase.objects.all().order_by("-created_at")
    return render(request, "lead_purchase_list.html", {"purchases": purchases})





def wallet_recharge_create(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_method = request.POST.get("payment_method")

        if not amount or not payment_method:
            messages.error(request, "All fields are required.")
            return redirect("wallet_recharge_create")

        WalletRecharge.objects.create(
            amount=amount,
            payment_method=payment_method,
        )
        messages.success(request, f"Wallet recharged with {amount} using {payment_method}!")
       # return redirect("wallet_recharge_list")

    return render(request, "agent/Wallet_Recharge_agent.html")


def wallet_recharge_list(request):
    recharges = WalletRecharge.objects.all().order_by("-created_at")
    return render(request, "agent/wallet_recharge_list.html", {"recharges": recharges})





def commission_report(request):
    lead_id = request.GET.get("lead_id")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    commissions = Commission.objects.all().order_by("-created_at")

    # Apply filters if provided
    if lead_id:
        commissions = commissions.filter(lead_id__icontains=lead_id)

    if from_date and to_date:
        commissions = commissions.filter(created_at__date__range=[from_date, to_date])
    elif from_date:  # only from date
        commissions = commissions.filter(created_at__date__gte=from_date)
    elif to_date:  # only to date
        commissions = commissions.filter(created_at__date__lte=to_date)

    context = {
        "commissions": commissions,
        "lead_id": lead_id or "",
        "from_date": from_date or "",
        "to_date": to_date or "",
    }
    return render(request, "agent/Commission_Report_Filter.html", context)




def subscription_overview(request):
    plans = SubscriptionPlan.objects.all().order_by('role', 'order')
    features = FeatureComparison.objects.all().order_by('order')
    # group plans by role as dict: {'Landlord': [...], ...}
    roles = {}
    for p in plans:
        roles.setdefault(p.role, []).append(p)
    return render(request, 'agent/overview.html', {
        'roles': roles,
        'features': features,
    })

def signup_submit(request):
    if request.method == 'POST':
        role = request.POST.get('role')
        plan_name = request.POST.get('selectedPlan')
        plan = SubscriptionPlan.objects.filter(name=plan_name, role=role).first()
        full_name = request.POST.get('owner_name') or request.POST.get('tenant_name') or request.POST.get('agency')
        # Collect other optional fields safely
        property_type = request.POST.get('property_type', '')
        amenities = request.POST.get('amenities', '')
        agency_name = request.POST.get('agency', '') if role == 'Agent' else ''
        service_areas = request.POST.get('service_areas', '') if role == 'Agent' else ''
        preferred_locations = request.POST.get('preferred_locations', '') if role == 'Tenant' else ''
        move_in_date = request.POST.get('move_in_date', None)
        if move_in_date == '':
            move_in_date = None

        signup = UserSignup.objects.create(
            role=role,
            selected_plan=plan,
            full_name=full_name,
            property_type=property_type,
            amenities=amenities,
            agency_name=agency_name,
            service_areas=service_areas,
            preferred_locations=preferred_locations,
            move_in_date=move_in_date
        )
        return redirect('agent:signup_success', pk=signup.pk)
    return redirect('agent:overview')

def signup_success(request, pk):
    signup = get_object_or_404(UserSignup, pk=pk)
    return render(request, 'agent/signup_success.html', {'signup': signup})

from Main_App .models import *


from django.http import HttpResponseForbidden

@login_required
def inquiry_list(request):
    user = request.user
    role = getattr(user, "role", "").lower()

    # Only agents can access this page
    if role != "agent":
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Get all properties posted by the current user (agent manages these)
    residential_props = ResidentialProperty.objects.filter(posted_by=user)
    commercial_props = CommercialProperty.objects.filter(posted_by=user)
    pg_props = PGProperty.objects.filter(posted_by=user)

    # Get inquiries for those properties
    residential_inquiries = ResidentialInquiry.objects.filter(residential_property__in=residential_props)
    commercial_inquiries = CommercialInquiry.objects.filter(commercial_property__in=commercial_props)
    pg_inquiries = PGInquiry.objects.filter(pg_property__in=pg_props)

    # Helper to prepare inquiry data
    def prepare_inquiry_data(inquiries, fk_field):
        data = []
        for inquiry in inquiries:
            prop = getattr(inquiry, fk_field)
            owner = prop.posted_by
            data.append({
                "inquiry_name": inquiry.name,
                "inquiry_email": inquiry.email,
                "inquiry_phone": inquiry.phone,
                "inquiry_message": inquiry.message,
                "property_title": getattr(prop, "property_title", "N/A"),
                "property_type": prop.__class__.__name__,
                "property_location": getattr(prop, "location", "N/A"),
                "property_price": getattr(prop, "price", "N/A"),
                "property_owner_name": getattr(owner, "full_name", owner.username),
                "property_owner_role": getattr(owner, "role", "N/A"),
                "lead_age": getattr(inquiry, "lead_age", "N/A"),
            })
        return data

    # Prepare inquiry data
    data = {
        "residential_inquiries": prepare_inquiry_data(residential_inquiries, "residential_property"),
        "commercial_inquiries": prepare_inquiry_data(commercial_inquiries, "commercial_property"),
        "pg_inquiries": prepare_inquiry_data(pg_inquiries, "pg_property"),
        "role": user.role,
        "user_full_name": getattr(user, "full_name", user.username),
    }

    template = "agent/inquiry_list.html"
    return render(request, template, data)







#################Views Start For Rental Residential Property###########################



def residential_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

   
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Rental/residential.html", context)





def residential_add_agent(request):
    # 1. Retrieve identity from browser session
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    

    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

 # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION (Who submitted the HTML form) ----------
            
            if user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "Agent"
                uploader_id = f"USER_{user_id}"
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION (Who owns/manages the listing) ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE (Checking LISTED BY, not UPLOADED BY)
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality_area') or request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_property_fingerprint(
                input_property_no, 
                input_building_name, 
                input_locality, 
                input_pincode
            )

            # 1. Direct Case-Insensitive Query for same unit in same locality/building
            direct_duplicates = RentalResidentialProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality_area__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            # Combine fingerprint match OR direct field match
            existing_duplicates = (
                RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                # Level 1: Hard Block if LISTED BY the exact same person (ID, Email, OR Phone match)
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        # INSTANT REJECTION: Same agent/owner cannot list the same property twice
                        messages.error(
                            request, 
                            f"Duplicate Blocked: This property (Unit {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        )
                        return redirect('residential_agent')

                # Level 2: Different Agent/User listing the exact same physical unit -> Allow save & Flag
                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = RentalResidentialProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Residential",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                property_no=input_property_no,
                bhk_type=request.POST.get('bhk_type'),
                renting_option=request.POST.get('renting_option'),
                built_up_area=to_decimal(request.POST.get('built_up_area')),
                bathrooms=to_int(request.POST.get('bathrooms')),
                balconies=to_int(request.POST.get('balconies')),
                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int(request.POST.get('total_floors')),
                facing_direction=request.POST.get('facing_direction', request.POST.get('facing')),
                furnishing_status=request.POST.get('furnishing_status'),
                available_for=request.POST.get('available_for'),

                carpet_area=to_decimal(request.POST.get('carpet_area')),
                city_zone=request.POST.get('city_zone', request.POST.get('zone')),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition', request.POST.get('construction_status')),
                property_age=request.POST.get('property_age'),
                wing_number=request.POST.get('wing_number'),
                building_name=input_building_name,

                availability_status=request.POST.get('availability_status', request.POST.get('possession_status')),
                available_from=available_from,
                lease_duration=request.POST.get('lease_duration'),
                brokerage_percentage=request.POST.get('brokerage_percentage', request.POST.get('brokerage')),
                manual_brokerage=request.POST.get('manual_brokerage'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=to_int(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int(request.POST.get('security_deposit_amount', request.POST.get('security_deposit'))),
                maintenance_type=request.POST.get('maintenance_type'),
                monthy_maintenance_amount=to_int(request.POST.get('monthy_maintenance_amount', request.POST.get('maintenance_amount'))),
                total_move_in_cost=to_int(request.POST.get('total_move_in_cost')),

                address=request.POST.get('address'),
                city=request.POST.get('city'),
                locality_area=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=input_pincode,
                main_road_connectivity=request.POST.get('main_road_connectivity', request.POST.get('road_connectivity')),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                amenities=amenities,
                nearby_facilities=facilities,

                user_description=request.POST.get('user_description'),
                description=request.POST.get('description'),
                rent_residential_desc=request.POST.get('rent_residential_desc'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None
            )

         
            # ---------- IMAGES MULTI-UPLOAD LOGIC (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'living':   'property_images_living[]',
                'bedroom':  'property_images_bedroom[]',
                'kitchen':  'property_images_kitchen[]',
                'bathroom': 'property_images_bathroom[]',
                'balcony':  'property_images_balcony[]',
                'others':   'property_images_others[]',
            }

            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 20:
                        break
                    RentalResidentialImage.objects.create(
                        property=prop,
                        image=img,
                        category=category,
                        sequence_order=cat_index,
                    )
                    saved_count += 1

          # ---------- PROPERTY VIDEO (UPLOAD, RM LINK, OR AUTO SLIDESHOW) ----------
            # ---------- PROPERTY VIDEO LOGIC (AUTO SLIDESHOW + UPLOAD / RM LINK) ----------
            video_option = request.POST.get('video_option') or request.POST.get('video_source') or 'auto'
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '').strip()

            # 1. ALWAYS auto-generate the slideshow row if >= 3 photos exist
            CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
            saved_images = list(RentalResidentialImage.objects.filter(property=prop))
            saved_images.sort(key=lambda img: (CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, img.sequence_order))
            image_paths = [img.image.path for img in saved_images if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)]

            if len(image_paths) >= 3:
                from Admin_App.utils import generate_property_slideshow

                output_relative_path = f"residential_rent/videos/auto_{prop.id}.mp4"
                try:
                    result_path = generate_property_slideshow(image_paths, output_relative_path)
                    if result_path:
                        RentalResidentialVideo.objects.update_or_create(
                            property=prop,
                            source='auto',
                            defaults={
                                'video': result_path,
                                'video_url': None
                                }
                        )
                except Exception as video_err:
                    import traceback
                    traceback.print_exc()

            # 2. Save Manual Upload Video as a separate row
            if video_option == 'upload' and uploaded_video:
                RentalResidentialVideo.objects.create(
                    property=prop,
                    video=uploaded_video,
                    source='uploaded'
                )

            # 3. Save RM Assisted Link Video as a separate row
            elif video_option == 'rm_assisted' and property_video_link:
                RentalResidentialVideo.objects.create(
                    property=prop,
                    video_url=property_video_link,
                    source='rm_assisted'
                )
            messages.success(request, "Property Added Successfully ")
            
            return redirect('residential_list_agent')

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            messages.error(request, f"Error while saving listing: {str(e)}")
            return redirect('residential_agent')

    return render(request, 'agent/Reports/Rental/rental_list.html', {
        'user_obj': user_obj,
        'user_role': user_role,
        
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    })






def _norm(val):
    return val.strip().lower() if val else ''
    

def rental_residential_view_agent(request, pk):
   

    # 1. Retrieve identity from browser session
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)



    prop = get_object_or_404(
        RentalResidentialProperty.objects.prefetch_related('images', 'faqs', 'video'),
        pk=pk
    )

    latest_properties = RentalResidentialProperty.objects.filter(
        is_deleted=False
    ).exclude(
        id=prop.id
    ).prefetch_related('faqs').order_by('-created_at')[:4]

    amenities_list = [x.strip() for x in prop.amenities.split(',')] if prop.amenities else []
    facilities_list = [x.strip() for x in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []

    # ---------- CATEGORY DISPLAY ORDER ----------
    CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
    CATEGORY_LABELS = {
        'exterior': 'Exterior / Building', 'living': 'Living / Dining', 'bedroom': 'Bedroom',
        'kitchen': 'Kitchen', 'bathroom': 'Bathroom', 'balcony': 'Balcony', 'others': 'Others',
    }

    # ---------- IMAGES ----------
    all_images = sorted(
        prop.images.all(),
        key=lambda img: (
            CATEGORY_ORDER.index(_norm(img.category)) if _norm(img.category) in CATEGORY_ORDER else 99,
            img.sequence_order
        )
    )

    # ---------- GROUPED BY CATEGORY ----------
    grouped_images = OrderedDict()
    for cat in CATEGORY_ORDER:
        cat_images = [img for img in all_images if _norm(img.category) == cat]
        if cat_images:
            grouped_images[cat] = {'label': CATEGORY_LABELS[cat], 'images': cat_images}

    # ---------- DEDICATED THUMBNAIL SLOTS ----------
    THUMB_PRIORITY = ['bathroom', 'living', 'bedroom', 'kitchen', 'balcony', 'others']

    def first_image_index(category):
        for idx, img in enumerate(all_images):
            if _norm(img.category) == category:
                return idx
        return None

    hero_category = _norm(all_images[0].category) if all_images else None
    used_categories = {hero_category} if hero_category else set()

    thumb1_idx = thumb2_idx = None
    for cat in THUMB_PRIORITY:
        if cat in used_categories:
            continue
        idx = first_image_index(cat)
        if idx is not None:
            if thumb1_idx is None:
                thumb1_idx = idx
            elif thumb2_idx is None:
                thumb2_idx = idx
                break
            used_categories.add(cat)

    thumb1_image = all_images[thumb1_idx] if thumb1_idx is not None else None
    thumb2_image = all_images[thumb2_idx] if thumb2_idx is not None else None

    # ---------- VIDEO HIERARCHY ENGINE ----------
    videos_queryset = prop.video.all() if hasattr(prop, 'video') else RentalResidentialVideo.objects.filter(property=prop)
    rm_video = videos_queryset.filter(source='rm_assisted', video_url__isnull=False).exclude(video_url='').first()
    manual_video = videos_queryset.filter(source='uploaded', video__isnull=False).first()
    auto_video = videos_queryset.filter(source='auto', video__isnull=False).first()

    selected_video = None
    video_display_mode = None

    if rm_video:
        selected_video = rm_video
        video_display_mode = 'rm_assisted'
    elif manual_video:
        selected_video = manual_video
        video_display_mode = 'manual'
    elif auto_video:
        selected_video = auto_video
        video_display_mode = 'auto'

    context = {
        'property': prop,
        'images': all_images,
        'grouped_images': grouped_images,
        'thumb1_image': thumb1_image,
        'thumb1_idx': thumb1_idx,
        'thumb2_image': thumb2_image,
        'thumb2_idx': thumb2_idx,
        'selected_video': selected_video,
        'video_display_mode': video_display_mode,
        'faqs': prop.faqs.all(),
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'latest_properties': latest_properties,
        

        'user_obj': user_obj,
       
        
    }
    return render(request, 'agent/Reports/Rental/rental_residential_detail.html', context)






def residential_agent_edit(request, pk):
    prop = get_object_or_404(RentalResidentialProperty, id=pk)

    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    if request.method == 'POST':
        try:
            p = request.POST

            def to_int(val):
                try: return int(str(val).replace(',', '').strip()) if val else None
                except: return None

            def to_decimal(val):
                try: return float(str(val).replace(',', '').strip()) if val else None
                except: return None

            # =====================================================
            # 1. AUDIT LOGIC: Snapshot state BEFORE modification
            # =====================================================
            tracked_fields_list = [
                'property_type', 'property_no', 'bhk_type', 'renting_option',
                'built_up_area', 'bathrooms', 'balconies', 'building_configuration',
                'total_floors', 'facing_direction', 'furnishing_status', 'available_for',
                'carpet_area', 'city_zone', 'ownership_type', 'property_condition',
                'property_age', 'wing_number', 'building_name',
                'availability_status', 'available_from', 'lease_duration',
                'brokerage_percentage', 'manual_brokerage', 'monthly_rent',
                'advance_rent_month', 'advance_rent_amount',
                'security_deposit_type', 'security_deposit_amount',
                'maintenance_type', 'monthy_maintenance_amount', 'total_move_in_cost',
                'address', 'city', 'locality_area', 'property_landmark', 'state', 'pincode',
                'main_road_connectivity', 'google_maps_link', 'latitude', 'longitude',
                'listed_by_id', 'listed_by_name', 'listed_by_email', 'listed_by_contact', 'listed_by_role',
                'listing_status', 'approval_status',
            ]
            old_state_snapshot = {field: str(getattr(prop, field, '')) for field in tracked_fields_list}

            # =====================================================
            # 2. LISTED BY & BASIC INFORMATION
            # =====================================================
            prop.listed_by_type = p.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to = p.get('assigned_to', prop.assigned_to)
            prop.listed_by_id = (p.get('listed_by_id') or prop.listed_by_id or '').strip()
            prop.listed_by_name = (p.get('listed_by_name') or prop.listed_by_name or '').strip()
            prop.listed_by_email = (p.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = (p.get('listed_by_contact') or prop.listed_by_contact or '').strip()
            prop.listed_by_role = (p.get('listed_by_role') or prop.listed_by_role or '').strip()

            prop.property_type = p.get('property_type')
            prop.property_no = (p.get('property_no') or prop.property_no or '').strip()
            prop.bhk_type = p.get('bhk_type')
            prop.renting_option = p.get('renting_option')
            prop.built_up_area = to_decimal(p.get('built_up_area'))
            prop.bathrooms = to_int(p.get('bathrooms'))
            prop.balconies = to_int(p.get('balconies'))
            prop.building_configuration = p.get('building_configuration')
            prop.total_floors = to_int(p.get('total_floors'))
            prop.facing_direction = p.get('facing_direction')
            prop.furnishing_status = p.get('furnishing_status')
            prop.available_for = p.get('available_for')

            # =====================================================
            # 3. PROPERTY DETAILS & LOCATION
            # =====================================================
            prop.carpet_area = to_decimal(p.get('carpet_area'))
            prop.city_zone = p.get('city_zone')
            prop.ownership_type = p.get('ownership_type')
            prop.property_condition = p.get('property_condition')
            prop.property_age = p.get('property_age')
            prop.wing_number = p.get('wing_number')
            prop.building_name = (p.get('building_name') or '').strip()

            prop.availability_status = p.get('availability_status')
            available_from_raw = p.get('available_from')
            if available_from_raw and available_from_raw.strip():
                try:
                    prop.available_from = datetime.strptime(available_from_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    prop.available_from = None
            else:
                prop.available_from = None

            prop.lease_duration = p.get('lease_duration')
            prop.brokerage_percentage = p.get('brokerage_percentage')
            prop.manual_brokerage = p.get('manual_brokerage')

            # =====================================================
            # 4. PRICING & LOCATION DETAILS
            # =====================================================
            prop.monthly_rent = to_int(p.get('monthly_rent'))
            prop.advance_rent_month = p.get('advance_rent_month')
            prop.advance_rent_amount = to_int(p.get('advance_rent_amount'))
            prop.security_deposit_type = p.get('security_deposit_type')
            prop.security_deposit_amount = to_int(p.get('security_deposit_amount'))
            prop.maintenance_type = p.get('maintenance_type')
            prop.monthy_maintenance_amount = to_int(p.get('monthy_maintenance_amount'))
            prop.total_move_in_cost = to_int(p.get('total_move_in_cost'))

            prop.address = p.get('address')
            prop.city = p.get('city')
            prop.locality_area = (p.get('locality_area') or p.get('locality') or '').strip()
            prop.property_landmark = p.get('property_landmark')
            prop.state = p.get('state')
            prop.pincode = (p.get('pincode') or '').strip()
            prop.main_road_connectivity = p.get('main_road_connectivity')
            prop.google_maps_link = p.get('google_maps_link')
            prop.latitude = p.get('latitude')
            prop.longitude = p.get('longitude')

            # =====================================================
            # 5. AMENITIES, DESCRIPTIONS & STATUS
            # =====================================================
            prop.amenities = ",".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]'))
            prop.user_description = p.get('user_description')
            prop.listed_elsewhere = p.get('listed_elsewhere', 'No')
            prop.portal_name = p.get('portal_name')
            prop.listing_status = p.get('listing_status', prop.listing_status)
            prop.approval_status = p.get('approval_status', prop.approval_status)

            prop.save()

            # =====================================================
            # 6. IMAGE DELETIONS
            # =====================================================
            delete_ids = request.POST.getlist('delete_image_ids[]')
            if delete_ids:
                RentalResidentialImage.objects.filter(id__in=delete_ids, property=prop).delete()

            # =====================================================
            # 7. NEW CATEGORIZED IMAGES UPLOAD LOGIC
            # =====================================================
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'living':   'property_images_living[]',
                'bedroom':  'property_images_bedroom[]',
                'kitchen':  'property_images_kitchen[]',
                'bathroom': 'property_images_bathroom[]',
                'balcony':  'property_images_balcony[]',
                'others':   'property_images_others[]',
            }
            existing_total = prop.images.count()
            new_images_added = 0

            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                existing_cat_count = prop.images.filter(category=category).count()
                for img in request.FILES.getlist(field_name):
                    if existing_total >= 20:
                        break
                    RentalResidentialImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=existing_cat_count
                    )
                    existing_cat_count += 1
                    existing_total += 1
                    new_images_added += 1

            images_changed = new_images_added > 0 or bool(delete_ids)

            # =====================================================
            # 8. VIDEO LOGIC: DIRECT ADMIN SELECTION
            # =====================================================
            # =====================================================
            # 8. VIDEO LOGIC: DIRECT ADMIN SELECTION
            # =====================================================
            video_option = p.get('video_option')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = p.get('property_video_link', '').strip()

            if video_option == 'upload' and uploaded_video:
                video_row, _ = RentalResidentialVideo.objects.get_or_create(
                    property=prop, 
                    source='uploaded'
                )
                video_row.video = uploaded_video
                video_row.video_url = None
                video_row.save()

            elif video_option == 'rm_assisted' and property_video_link:
                video_row, _ = RentalResidentialVideo.objects.get_or_create(
                    property=prop, 
                    source='rm_assisted'
                )
                video_row.video_url = property_video_link
                video_row.video = None
                video_row.save()

            # =====================================================
            # 9. AUTO SLIDESHOW: REFRESH ON PHOTO CHANGE OR REQUEST
            # =====================================================
            regenerate_requested = p.get('regenerate_slideshow') == 'on'

            if images_changed or regenerate_requested:
                CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
                saved_images = list(prop.images.all())
                saved_images.sort(key=lambda img: (
                    CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, 
                    img.sequence_order
                ))
                image_paths = [
                    img.image.path for img in saved_images
                    if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                ]

                if len(image_paths) >= 3:
               

                    output_relative_path = f"residential_rent/videos/auto_{prop.id}.mp4"
                    try:
                        result_path = generate_property_slideshow(image_paths, output_relative_path)
                        if result_path:
                            RentalResidentialVideo.objects.update_or_create(
                                property=prop,
                                source='auto',
                                defaults={
                                    'video': result_path,
                                    'video_url': None
                                    }
                            )
                    except Exception as video_err:
                        import traceback
                        traceback.print_exc()
            # =====================================================
            # 10. AUDIT LOGGING & DIFF DICTIONARY
            # =====================================================
            modifications_diff = {}
            modified_fields_summary = []

            for field in tracked_fields_list:
                new_val = str(getattr(prop, field, ''))
                if old_state_snapshot[field] != new_val:
                    modifications_diff[field] = {
                        "old_value": old_state_snapshot[field],
                        "new_value": new_val
                    }
                    modified_fields_summary.append(field)

            if modified_fields_summary:
                RentalActivityLog.objects.create(
                    user_identity=admin_obj.email or admin_obj.username,
                    user_role="Admin",
                    action_type='UPDATE',
                    property_id=prop.id,
                    targeted_fields=", ".join(modified_fields_summary[:4]) + ("..." if len(modified_fields_summary) > 4 else ""),
                    associated_file="Web UI Form",
                    action_payload=json.dumps(modifications_diff),
                    ip_address=_get_client_ip(request),
                    status='SUCCESS'
                )

            return JsonResponse({
                'status': 'success',
                'message': 'Residential Rental Property Updated Successfully',
                'redirect_url': reverse('residential_list_agent')
            })

        except Exception as e:
            print("ERROR IN RESIDENTIAL EDIT:", str(e))
            traceback.print_exc()

            return JsonResponse({'status': 'error', 'message': f"Failed to save data: {str(e)}"}, status=400)

    # ================= GET REQUEST CONTEXT =================
    residential_videos = RentalResidentialVideo.objects.filter(property=prop)
    videos_by_source = {v.source: v for v in residential_videos}

    images_by_category = {}
    for img in prop.images.all().order_by('category', 'sequence_order'):
        images_by_category.setdefault(img.category, []).append(img)

    context = {
        'property': prop,
        
 
        'user_obj':user_obj ,
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'uploaded_video': videos_by_source.get('uploaded'),
        'auto_video': videos_by_source.get('auto'),
        'rm_video': videos_by_source.get('rm_assisted'),
        'images_by_category': images_by_category,
        'selected_amenities': [a.strip() for a in (prop.amenities or '').split(',') if a.strip()],
        'selected_facilities': [f.strip() for f in (prop.nearby_facilities or '').split(',') if f.strip()],
    }
    return render(request, 'agent/Forms/Rental/residential_edit.html', context)




def residential_list_agent(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
   
    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════

    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', '').strip()
    bhk_query = request.GET.get('bhk_type', '').strip()
    state_query = request.GET.get('state', '').strip()          # NEW
    furnish_query = request.GET.get('furnishing', '').strip()
    possession_query = request.GET.get('possession', '').strip()  # maps to availability_status

    prop_type_query = request.GET.get('property_type', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()     # listed_by_role
    uploaded_by_query = request.GET.get('uploaded_by', '').strip()  # NEW -> uploaded_by_role
    budget_query = request.GET.get('budget', '').strip()

    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query = request.GET.get('duplicate', '').strip()

    # ═══════════════════════════════════════
    # BASE QUERYSET & PERSISTENT SR.NO MAP
    # ═══════════════════════════════════════

    base_properties = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    absolute_ordered_ids = list(base_properties.values_list('id', flat=True))

    properties = base_properties

    # ═══════════════════════════════════════
    # GLOBAL SEARCH FILTER (Across ALL Fields + Sr.No)
    # ═══════════════════════════════════════

    if search_query:
        sr_no_query = Q()

        if search_query.isdigit():
            target_index = int(search_query) - 1
            if 0 <= target_index < len(absolute_ordered_ids):
                target_id = absolute_ordered_ids[target_index]
                sr_no_query = Q(id=target_id)

        properties = properties.filter(
            sr_no_query |
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(renting_option__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(bhk_type__icontains=search_query) |
            Q(locality_area__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(available_for__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(upload_file_name__icontains=search_query)
        )

    # ═══════════════════════════════════════
    # ADVANCED DROPDOWN FILTERS
    # ═══════════════════════════════════════

    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__iexact=city_query)

    if bhk_query and bhk_query != 'All BHK':
        properties = properties.filter(bhk_type__iexact=bhk_query)

    if state_query and state_query != 'All States':          # NEW
        properties = properties.filter(state__iexact=state_query)

    if furnish_query and furnish_query != 'All':
        properties = properties.filter(furnishing_status__iexact=furnish_query)

    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__iexact=possession_query)

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__iexact=prop_type_query)

    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':   # NEW
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_10k':
            properties = properties.filter(monthly_rent__lt=10000)
        elif budget_query == '10k_25k':
            properties = properties.filter(monthly_rent__gte=10000, monthly_rent__lte=25000)
        elif budget_query == '25k_50k':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=50000)
        elif budget_query == 'above_50k':
            properties = properties.filter(monthly_rent__gt=50000)

    # Date Filters
    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            properties = properties.filter(created_at__date__gte=from_date)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            properties = properties.filter(created_at__date__lte=to_date)

    # ═══════════════════════════════════════
    # PAGINATION & INJECT PERSISTENT SR.NO
    # ═══════════════════════════════════════
    filtered_count = properties.count()

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    id_to_srno = {pid: idx + 1 for idx, pid in enumerate(absolute_ordered_ids)}
    for prop in page_obj:
        prop.original_sr_no = id_to_srno[prop.id]

   

    # ═══════════════════════════════════════
    # EXPORT DATA (CSV / EXCEL)
    # ═══════════════════════════════════════
    if request.GET.get('download') in ['excel', 'csv']:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        from collections import OrderedDict
        import csv

        # ═══════════════════════════════════════
        # SINGLE SOURCE OF TRUTH — reuse the same field map the template
        # download and the importer already use, so LABELS stay in sync.
        # ═══════════════════════════════════════
        (field_map_sections, field_to_label_map, label_to_field_map,
         system_injected_map, helper_only_labels_map,
         decimal_fields_map, int_fields_map) = _residential_field_map()

        # ═══════════════════════════════════════
        # CANONICAL REQUIRED FIELDS — must be the literal same set used in
        # import_residential_excel's REQUIRED_FIELD_KEYS. Do NOT derive this
        # from _residential_field_map()'s per-field `required` flag — that
        # flag is used for template styling only and does not match the
        # actual mandatory inputs on the "Post Your Residential Rental" form
        # (e.g. it doesn't mark listed_by_name/email/contact as required,
        # but the Add form and the importer both do).
        # ═══════════════════════════════════════
        REQUIRED = {
            "property_type", "property_no", "bhk_type", "renting_option",
            "built_up_area", "bathrooms", "building_configuration", "total_floors",
            "furnishing_status", "available_for", "monthly_rent",
            "address", "city", "locality_area", "state", "pincode",
            "listed_by_name", "listed_by_contact", "listed_by_email", "listed_by_role",
        }

        # FIELD_LABELS starts from the field map (guarantees every shared,
        # re-importable column uses the exact label the importer expects).
        FIELD_LABELS = dict(field_to_label_map)

        # Export-only columns: appear on this report but are NOT part of the
        # import contract in _residential_field_map(). Never required, and
        # safely ignored (as unmatched headers) if this file is re-uploaded.
        FIELD_LABELS.update({

            "id": "Property ID",
            "property_title": "Property Title",
            "description": "Property Summary (Auto)",
            "rent_residential_desc": "Property Description (Auto)",
            "upload_file_name": "Uploaded File Name",
            "listing_status": "Listing Status",
            "approval_status": "Approval Status",
            "listed_elsewhere": "Listed Elsewhere?",
            "portal_name": "Portal Name",

            "property_unique_key": "Property Unique Key",
            "duplicate_count":  "Duplicate Count",
            "duplicate_group_id": "Duplicate Group ID",
            "is_duplicate": "Is Duplicate",
            "is_deleted": "Is Deleted",
            "deleted_at": "Deleted At",
            "deleted_by": "Deleted By",
            "created_at": "Created At",
        })



        sections = {


            "S.No": [
                "sr_no",
            ],
            "Listed By": [
                "listed_by_type", "listed_by_id", "listed_by_name", "listed_by_contact", "listed_by_email",
                "listed_by_role",
            ],
            "Basic Information": [
                "id", "property_title", "property_type", "property_no", "bhk_type", "renting_option",
                "built_up_area", "bathrooms", "balconies",
                "building_configuration", "total_floors", "facing_direction",
                "furnishing_status", "available_for"
            ],
            "Property Details": [
                "carpet_area", "city_zone", "ownership_type", "property_condition", "property_age",
                "wing_number", "building_name"
            ],
            "Availability Details": [
                "availability_status", "available_from", "lease_duration",
            ],
            "Pricing & Brokarage Details": [
                "monthly_rent","brokerage_percentage", "manual_brokerage", "advance_rent_month", "advance_rent_amount",
                "security_deposit_type", "security_deposit_amount",
                "maintenance_type", "monthy_maintenance_amount", "total_move_in_cost"
            ],
            "Property Location Details": [
                "address", "city", "locality_area", "property_landmark", "state",
                "pincode", "main_road_connectivity", "google_maps_link", "latitude", "longitude",
            ],
            "Amenities & Features": [
                "amenities",
            ],
            "Nearby Facilities": [
                "nearby_facilities",
            ],
            "Auto Property Summary & Description Generated Details": [
                "description",
                "rent_residential_desc",
            ],
            "Property Description(Added By User)": [
                "user_description",
            ],

            "Property Listed Elsewhere": [
                "listed_elsewhere", "portal_name",
                
            ],
            "Property Uploaded By(Auto Generated)": [
                "uploaded_by_name", "uploaded_by_email",
                "uploaded_by_contact", "uploaded_by_role", "upload_file_name",
            ],
            "Record Management": [
            "listing_status", "approval_status","property_unique_key",
            "duplicate_count", "duplicate_group_id" ,
            "is_duplicate" , 
            "is_deleted","deleted_at", "deleted_by", "created_at"
            ]
        }

        HINTS = {
            "sr_no":"Auto","id": "Property ID Auto-Generated",
            "property_title": "Auto_Generated Title", "property_type": "Apartment",
            "property_no": "e.g. B-402", "bhk_type": "e.g. 2bhk",
            "renting_option": "Full Property/Single Room/Shared Room", "built_up_area": "sq.ft",
            "bathrooms": "Number", "balconies": "Number", "building_configuration": "e.g. G+3",
            "total_floors": "Number", "facing_direction": "North/East", "furnishing_status": "Semi Furnished",
            "available_for": "Family/Bachelor", "carpet_area": "sq.ft", "city_zone": "North/South", "ownership_type": "Freehold",
            "property_condition": "Resale", "property_age": "1-3 Years",
            "wing_number": "e.g. A/B", "building_name": "Text", "availability_status": "Ready to Move",
            "available_from": "YYYY-MM-DD", "lease_duration": "11 Months", "monthly_rent": "₹","brokerage_percentage": "1%/Fixed Amount",
            "manual_brokerage": "e.g. 2.5%", "advance_rent_month": "0-11/fixed",
            "advance_rent_amount": "₹", "security_deposit_type": "0-11/fixed", "security_deposit_amount": "₹",
            "maintenance_type": "Included in Rent/Extra", "monthy_maintenance_amount": "₹",
            "total_move_in_cost": "₹", "address": "Full Address", "city": "Text", "locality_area": "Text",
            "property_landmark": "Optional", "state": "e.g. Maharashtra", "pincode": "6-digit",
            "main_road_connectivity": "Optional", "google_maps_link": "URL", "latitude": "21.1458",
            "longitude": "23.1458", "amenities": "Comma-sep",
            "nearby_facilities": "Comma-sep", "description": "Short Summary",
            "rent_residential_desc": "Long Rich Text", "user_description": "Added by user",  "listed_elsewhere": "Yes/No",
            "portal_name": "e.g. 99acres, MagicBricks",
            "listed_by_name": "Full Name", "listed_by_contact": "10 Digits", "listed_by_email": "email@example.com",
            "listed_by_role": "Owner/Agent/Admin", "listed_by_type": "Self/Other",
            "uploaded_by_name": "Admin Name", "uploaded_by_email": "Admin Email",
            "uploaded_by_contact": "Admin Contact", "uploaded_by_role": "Admin Role",
            "upload_file_name": "File Name", "listing_status": "Published/Draft", "approval_status": "Approved/Pending",
            "property_unique_key" : "Property unique id",
            "duplicate_count": "duplicate_count" ,
            "duplicate_group_id": "Duplicate Ggroup Id",
            "is_duplicate": "Yes/No", 
            "is_deleted": "Yes/No",
            "deleted_at": "YYYY-MM-DD (Auto)", "deleted_by": "Admin Name",
            "created_at": "YYYY-MM-DD"
        }

        BOOL_FIELDS = {"is_deleted", "is_duplicate"}

        ROLE_BROKERAGE_LABELS = {
            "admin": "EstateFlow Service Fee",
            "relationship manager": "Service Fee",
            "landlord": "Tenant Service Fee",
            "agent": "Brokerage",
            "agency/builder": "Service Fee",
            "builder": "Service Fee",
        }

        def brokerage_label_for(role):
            if not role:
                return "Brokerage"
            return ROLE_BROKERAGE_LABELS.get(str(role).strip().lower(), "Brokerage")

        def display_value(p, field):
            val = getattr(p, field, "")
            if field in ['available_from', 'created_at', 'deleted_at'] and val:
                try:
                    return val.strftime('%Y-%m-%d')
                except AttributeError:
                    return val
            if field in BOOL_FIELDS:
                return "Yes" if val else "No"
            # NEW: listings created directly on the site (not via bulk
            # import) never get an upload_file_name — label them clearly
            # instead of leaving the cell blank.
            if field == "upload_file_name":
                return val if val else "Web Listing UI Form"
            return val if val is not None else ""

        all_cols = []
        for sec, fields in sections.items():
            all_cols.extend([(sec, f) for f in fields])

        if request.GET.get('download') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rental Residential"

            HDR_BG, REQ_BG, OPT_BG = "667EEA", "FEF3C7", "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

            sec_spans = OrderedDict()
            for i, (sec, _) in enumerate(all_cols):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            role_col = None
            brokerage_col = None

            for ci, (sec, field) in enumerate(all_cols, 1):
                req = field in REQUIRED
                label = FIELD_LABELS.get(field, field)
                lc = ws.cell(row=2, column=ci, value=label + (" *" if req else ""))
                lc.font = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border = bdr

                hc = ws.cell(row=3, column=ci, value="")
                hc.font = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(16, len(label) // 2 + 6)

                if field == "listed_by_role":
                    role_col = ci
                if field == "brokerage_percentage":
                    brokerage_col = ci

            if brokerage_col:
                note = (
                    "The LABEL text shown above this field on the live form changes based on\n"
                    "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
                    "Admin -> EstateFlow Service Fee\n"
                    "Relationship Manager -> Service Fee\n"
                    "Landlord -> Tenant Service Fee\n"
                    "Agent -> Brokerage\n"
                    "Agency/Builder or Builder -> Service Fee\n"
                    "Any other role -> Brokerage (default)\n\n"
                    "See the 'Brokerage Label' column at the end of this sheet for the resolved value per row."
                )
                ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            for row_idx, p in enumerate(properties, start=4):
                for col_idx, (sec, field) in enumerate(all_cols, 1):
                    if field == "sr_no":
                        val = row_idx - 3   # row 4 -> Sr No 1
                    else:
                        val = display_value(p, field)

                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr

            if role_col:
                preview_col = len(all_cols) + 1
                pc = ws.cell(row=2, column=preview_col, value="Brokerage Label")
                pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
                pc.fill = PatternFill("solid", fgColor="FEF3C7")
                pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                pc.border = bdr
                ws.cell(row=3, column=preview_col, value="Auto-resolved from role")
                ws.cell(row=3, column=preview_col).font = Font(italic=True, color="64748B", name="Arial", size=8)
                ws.cell(row=3, column=preview_col).border = bdr
                ws.column_dimensions[get_column_letter(preview_col)].width = 22

                for row_idx, p in enumerate(properties, start=4):
                    label = brokerage_label_for(getattr(p, "listed_by_role", ""))
                    fc = ws.cell(row=row_idx, column=preview_col, value=label)
                    fc.alignment = Alignment(horizontal="center", vertical="center")
                    fc.border = bdr

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Properties_Data.xlsx"'
            wb.save(response)
            return response

        elif request.GET.get('download') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Rental_Residential_Properties_Data.csv"'
            writer = csv.writer(response)

            row1 = []
            current_sec = ""
            for sec, _ in all_cols:
                if sec != current_sec:
                    row1.append(f" {sec}")
                    current_sec = sec
                else:
                    row1.append("")
            row1.append("📋 Brokerage")
            writer.writerow(row1)

            header_row = [FIELD_LABELS.get(field, field) + (" *" if field in REQUIRED else "") for _, field in all_cols]
            header_row.append("Brokerage Label")
            writer.writerow(header_row)

            hint_row = ["" for _, field in all_cols]
            hint_row.append("")
            writer.writerow(hint_row)

            for p in properties:
                data_row = [display_value(p, field) for _, field in all_cols]
                data_row.append(brokerage_label_for(getattr(p, "listed_by_role", "")))
                writer.writerow(data_row)

            return response

    # ═══════════════════════════════════════
    # STATS & UNIQUE DROPDOWN DATA
    # ═══════════════════════════════════════

    all_props = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)

    total_count = all_props.count()

    unique_listing_status = all_props.exclude(listing_status__isnull=True).exclude(listing_status='').values_list('listing_status', flat=True).distinct()
    unique_approval_status = all_props.exclude(approval_status__isnull=True).exclude(approval_status='').values_list('approval_status', flat=True).distinct()
    unique_bhk = all_props.exclude(bhk_type__isnull=True).exclude(bhk_type='').values_list('bhk_type', flat=True).distinct()
    unique_cities = all_props.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_states = all_props.exclude(state__isnull=True).exclude(state='').values_list('state', flat=True).distinct()  # NEW
    unique_furnish = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values_list('furnishing_status', flat=True).distinct()
    unique_possession = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').values_list('availability_status', flat=True).distinct()
    unique_property_types = all_props.exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    unique_listed_roles = all_props.exclude(listed_by_role__isnull=True).exclude(listed_by_role='').values_list('listed_by_role', flat=True).distinct()   # NEW split
    unique_uploaded_roles = all_props.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()  # NEW split

    active_count = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    furnished_count = all_props.filter(furnishing_status__iexact='Furnished').count()
    available_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    city_count = all_props.exclude(city__isnull=True).exclude(city='').values('city').distinct().count()
    state_count = all_props.exclude(state__isnull=True).exclude(state='').values('state').distinct().count()  # NEW

    # Duplicate properties
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count = all_props.filter(is_duplicate=False).count()

# Listing status breakdown
    active_listing_count = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count = all_props.filter(listing_status__iexact='Rented').count()

# Approval status breakdown
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count = all_props.filter(approval_status__iexact='Rejected').count()

    # ═══════════════════════════════════════
    # RENT STATS
    # ═══════════════════════════════════════

    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )

    avg_rent = rent_stats['avg_rent']
    max_rent = rent_stats['max_rent']
    min_rent = rent_stats['min_rent']

    deposit_stats = all_props.exclude(security_deposit_amount__isnull=True).aggregate(avg_deposit=Avg('security_deposit_amount'))
    avg_deposit = deposit_stats['avg_deposit']

    area_stats = all_props.exclude(built_up_area__isnull=True).aggregate(avg_area=Avg('built_up_area'))
    avg_area = area_stats['avg_area']

    with_owner_count = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    with_images_count = all_props.filter(images__isnull=False).distinct().count()

    uploaded_files = all_props.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    # ═══════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════

    renting_option_qs = all_props.exclude(renting_option__isnull=True).exclude(renting_option='').values('renting_option').annotate(count=Count('id')).order_by('-count')
    renting_option_labels = json.dumps([item['renting_option'] for item in renting_option_qs])
    renting_option_data = json.dumps([item['count'] for item in renting_option_qs])

    rent_buckets = [
        ('Under ₹5k', 0, 5000), ('₹5k–10k', 5000, 10000), ('₹10k–20k', 10000, 20000),
        ('₹20k–30k', 20000, 30000), ('₹30k–50k', 30000, 50000), ('₹50k–1L', 50000, 100000),
        ('Above ₹1L', 100000, 999999999),
    ]

    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data = json.dumps([all_props.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values('furnishing_status').annotate(count=Count('id')).order_by('-count')
    furnishing_labels = json.dumps([item['furnishing_status'] for item in furnish_qs])
    furnishing_data = json.dumps([item['count'] for item in furnish_qs])

    prop_type_qs = all_props.exclude(property_type__isnull=True).exclude(property_type='').values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels = json.dumps([item['property_type'] for item in prop_type_qs])
    prop_type_data = json.dumps([item['count'] for item in prop_type_qs])

    # ═══════════════════════════════════════
    # KPI
    # ═══════════════════════════════════════

    occupied_count = all_props.filter(availability_status__iexact='Occupied').count()
    vacant_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100), 1) if total_count > 0 else 0
    vacancy_rate = round((vacant_count / total_count * 100), 1) if total_count > 0 else 0
    total_revenue = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0

    ready_to_move_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    short_lease_count = all_props.filter(lease_duration__icontains='6').count()
    long_lease_count = all_props.filter(lease_duration__icontains='12').count()
    new_property_count = all_props.filter(property_age__icontains='New').count()
    old_property_count = all_props.exclude(property_age__icontains='New').count()
    premium_properties_count = all_props.filter(monthly_rent__gte=50000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=15000).count()

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════

    context = {
        'user_obj': user_obj,
        'page_obj': page_obj,

        'search_query': search_query,
        'bhk_query': bhk_query,
        'city_query': city_query,
        'state_query': state_query,          # NEW
        'furnish_query': furnish_query,
        'possession_query': possession_query,
        'prop_type_query': prop_type_query,
        'listed_by_query': listed_by_query,
        'uploaded_by_query': uploaded_by_query,   # NEW
        'budget_query': budget_query,

        'unique_cities': unique_cities,
        'unique_states': unique_states,        # NEW
        'unique_furnish': unique_furnish,
        'unique_possession': unique_possession,
        'unique_property_types': unique_property_types,
        'unique_listed_roles': unique_listed_roles,     # NEW (split)
        'unique_uploaded_roles': unique_uploaded_roles, # NEW (split)

        'filtered_count': filtered_count,

        'total_count': total_count,
        'active_count': active_count,
        'furnished_count': furnished_count,
        'available_count': available_count,
        'city_count': city_count,
        'state_count': state_count,          # NEW

        'from_date': from_date_str,
        'to_date': to_date_str,

        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,

        'avg_deposit': avg_deposit,
        'avg_area': avg_area,

        'with_owner_count': with_owner_count,
        'with_images_count': with_images_count,

        'uploaded_files': uploaded_files,

        'renting_option_labels': renting_option_labels,   # renamed from bhk_labels
        'renting_option_data': renting_option_data,

        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,

        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,

        'prop_type_labels': prop_type_labels,
        'prop_type_data': prop_type_data,

        'occupied_count': occupied_count,
        'vacant_count': vacant_count,

        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,

        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,

        'ready_to_move_count': ready_to_move_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,

        'new_property_count': new_property_count,
        'old_property_count': old_property_count,

        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'bhk_query': bhk_query,
        'unique_bhk': unique_bhk,                              # ADD

        'listing_status_query': listing_status_query,          # ADD
        'approval_status_query': approval_status_query,        # ADD
        'duplicate_query': duplicate_query,                    # ADD

        'unique_listing_status': unique_listing_status,        # ADD
        'unique_approval_status': unique_approval_status,      # ADD

        'duplicate_properties_count': duplicate_properties_count,  # ADD
        'unique_properties_count': unique_properties_count,        # ADD

        'active_listing_count': active_listing_count,          # ADD
        'inactive_listing_count': inactive_listing_count,      # ADD
        'sold_listing_count': sold_listing_count,               # ADD
        'rented_listing_count': rented_listing_count,           # ADD
        'pending_approval_count': pending_approval_count,       # ADD
        'approved_count': approved_count,                       # ADD
        'rejected_count': rejected_count,                       # ADD
    }

    return render(
        request,
        'agent/Reports/Rental/rental_list.html',
        context
    )




@require_POST
def rental_bulk_delete_agent(request):
    """Agent Advanced Bulk Delete (Soft Delete) — scoped to agent's own properties, supports Admin impersonation."""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved user
        # ======================================
        try:
            user_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {user_obj.user_id} | "
                f"Name: {user_obj.user_name} | "
                f"Email: {user_obj.user_email} | "
                f"Phone: {user_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {user_obj.user_id} | "
                f"Name: {user_obj.user_name} | "
                f"Email: {user_obj.user_email} | "
                f"Phone: {user_obj.user_phone} | "
                
                f"Role: {user_obj.user_role}"
            )
            user_role = user_obj.user_role or "Agent"

        # ======================================
        # 5. Parse Payload
        # ======================================
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Invalid request payload.'})

        delete_type = data.get('delete_type')

        # ======================================
        # 6. Base Queryset — SCOPED TO THIS AGENT ONLY
        # ======================================
        properties = RentalResidentialProperty.objects.filter(
            is_deleted=False,
            listed_by_id=user_obj.user_id
        )

        target_props = None
        criteria_label = ""

        if delete_type == 'delete_all':
            target_props = properties
            criteria_label = "All own properties"

        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            if not page_ids:
                return JsonResponse({'status': 'error', 'message': 'No properties selected on current page.'})
            target_props = properties.filter(id__in=page_ids)
            criteria_label = "Current page selection"

        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            if not from_date or not to_date:
                return JsonResponse({'status': 'error', 'message': 'Both From Date and To Date are required.'})
            target_props = properties.filter(available_from__range=[from_date, to_date])
            criteria_label = f"Date range {from_date} to {to_date}"

        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(available_from__gte=thirty_days_ago)
            criteria_label = "Latest month"

        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(available_from__lt=six_months_ago)
            criteria_label = "Older than 6 months"

        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '').strip()
            if not uploader:
                return JsonResponse({'status': 'error', 'message': 'Uploader detail is required.'})
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            criteria_label = f"Uploaded by '{uploader}'"

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '').strip()
            if not file_name:
                return JsonResponse({'status': 'error', 'message': 'Please select a file.'})
            target_props = properties.filter(upload_file_name=file_name)
            criteria_label = f"File '{file_name}'"

        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

        # ======================================
        # 7. Execute Soft Delete
        # ======================================
        count = target_props.count()

        if count == 0:
            return JsonResponse({
                'status': 'error',
                'message': f'No matching properties found for: {criteria_label}'
            })

        deleted_ids = list(target_props.values_list('id', flat=True))
        deleted_titles = list(target_props.values_list('property_title', flat=True))

        target_props.update(
            is_deleted=True,
            deleted_at=timezone.now(),
            deleted_by=user_identity
        )

        # ======================================
        # 8. Audit Log
        # ======================================
        RentalActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type='BULK_DELETE',
            property_id=",".join(str(i) for i in deleted_ids),
            targeted_fields='Entire Record(s) Purged',
            associated_file=criteria_label,
            action_payload=json.dumps({
                "delete_type": delete_type,
                "criteria": criteria_label,
                "deleted_count": count,
                "deleted_property_ids": deleted_ids,
                "deleted_property_titles": deleted_titles,
                "action": "bulk_soft_delete_to_recycle_bin",
            }),
            ip_address=_get_client_ip(request),
            status='SUCCESS'
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully moved {count} of your properties ({criteria_label}) to Recycle Bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



@require_POST
def rental_residential_delete_agent(request, pk):
    """Agent Soft Delete Property (supports Admin impersonation)"""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved user
        # ======================================
        try:
            user_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Agent not found.'
            })

        # Build identity string — note if this was an admin acting on the agent's behalf
        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {user_obj.user_id} | "
                f"Name: {user_obj.user_name} | "
                f"Email: {user_obj.user_email} | "
                f"Phone: {user_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {user_obj.user_id} | "
                f"Name: {user_obj.user_name} | "
                f"Email: {user_obj.user_email} | "
                f"Phone: {user_obj.user_phone} | "
                f"Role: {user_obj.user_role}"
            )
            user_role = user_obj.user_role or "Agent"

        # ======================================
        # 5. Fetch Property
        # ======================================
        prop = get_object_or_404(RentalResidentialProperty, id=pk)

        # ======================================
        # 6. Security Check — agent can delete only own property
        # ======================================
        if str(prop.listed_by_id) != str(user_obj.user_id):
            return JsonResponse({
                'status': 'error',
                'message': 'You can delete only your own properties.'
            })

        # ======================================
        # 7. Property Details (for audit log)
        # ======================================
        associated_origin_file = prop.upload_file_name if prop.upload_file_name else "Web UI Form"
        property_title_ref = prop.property_title if prop.property_title else "N/A"

        # ======================================
        # 8. Soft Delete
        # ======================================
        prop.is_deleted = True
        prop.deleted_at = timezone.now()
        prop.deleted_by = user_identity
        prop.save()

        # ======================================
        # 9. Audit Log
        # ======================================
        RentalActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type='DELETE',
            property_id=str(pk),
            targeted_fields='Entire Record Purged',
            associated_file=associated_origin_file,
            action_payload=json.dumps({
                "deleted_property_id": pk,
                "property_title": property_title_ref,
                "action": "soft_delete_to_recycle_bin",
            }),
            ip_address=_get_client_ip(request),
            status='SUCCESS'
        )

        return JsonResponse({
            'status': 'success',
            'message': 'Property deleted successfully!'
        })

    except RentalResidentialProperty.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Property not found.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


#################Views End For Rental Residential Property#####################################



########################Views Start For Rental residential listing dowload and import ####################################








def _residential_field_map():
    """Returns (sections, field_to_label, label_to_field, system_injected,
    helper_only_labels, decimal_fields, int_fields). Called fresh by both
    the download view and the import view — always local, never global."""
 
    sections = OrderedDict([
        ("Listed By", [
            ("listed_by_type",   "Listed By Type (Self/Other)", False),
            ("listed_by_id",     "Listed By Id", False),
            ("listed_by_name",   "Listed By Name", False),
            ("listed_by_email",  "Listed By Email", False),
            ("listed_by_contact","Listed By Contact", False),
            ("listed_by_role",   "Listed By Role", True),   # drives the brokerage label
        ]),
        ("Basic Information", [
            ("property_type",           "Property Type", True),
            ("property_no",             "Flat/House/Unit No", True),
            ("bhk_type",                "BHK Type", True),
            ("renting_option",          "Renting Option", True),
            ("built_up_area",           "Built-up Area (sq.ft)", True),
            ("bathrooms",                "Bathrooms", True),
            ("balconies",                "Balconies", False),
            ("building_configuration",  "Building Configuration", True),
            ("total_floors",             "Total Floors Constructed", True),
            ("facing_direction",        "Facing Direction", False),
            ("furnishing_status",       "Furnishing Status", True),
            ("available_for",           "Available For", True),
        ]),
        ("Property Details", [
            ("carpet_area",     "Carpet Area (sq.ft)", False),
            ("city_zone",       "City Zone", False),
            ("ownership_type",  "Ownership Type", False),
            ("property_condition", "Property Condition", False),
            ("property_age",    "Property Age (Years)", False),
            ("wing_number",     "Tower/Wing Number", False),
            ("building_name",   "Building/Society Name", False),
        ]),
        ("Availability Details", [
            ("availability_status",   "Availability Status", False),
            ("available_from",        "Available From", False),
            ("lease_duration",        "Lease Duration", False),
           
        ]),
        ("Pricing & Brokarage Details", [
            
            ("monthly_rent",              "Monthly Rent", True),
            ("brokerage_percentage",  "Brokerage / Service Fee", True),
            ("manual_brokerage",      "Fixed Brokerage Amount", False),
            ("advance_rent_month",        "Advance Rent Month", False),
            ("advance_rent_amount",       "Advance Rent Amount", False),
            ("security_deposit_type",     "Refundable Security Deposit", True),
            ("security_deposit_amount",   "Refundable Security Deposit Amount", False),
            ("maintenance_type",          "Maintenance Type", False),
            ("monthy_maintenance_amount", "Monthly Maintenance Amount", False),
            ("total_move_in_cost",        "Total Move In Cost", False),
        ]),
        ("Property Location Details", [
            ("address",                 "Address", True),
            ("city",                    "City", True),
            ("locality_area",           "Locality/Area", True),
            ("property_landmark",       "Property Landmark", False),
            ("state",                   "State", True),
            ("pincode",                 "Pincode", True),
            ("main_road_connectivity",  "Main Road Connectivity", False),
            # --- newly added: present on the form, were missing from the map ---
            ("google_maps_link",        "Google Maps Link", False),
            ("latitude",                 "Latitude", False),
            ("longitude",                "Longitude", False),
        ]),
        ("Amenities & Features", [
            ("amenities",          "Amenities (comma-separated)", False),
           
        ]),
        ("Nearby Facilities", [
           
            ("nearby_facilities",  "Nearby Facilities (comma-separated)", False),
        ]),
        ("Property Descriptions", [
            ("user_description", "Property Description", False),
        ]),
        ("Media & Listing Status", [
            ("listed_elsewhere", "Listed Elsewhere (Yes/No)", False),
            ("portal_name",      "Portal Name", False),
            ("created_at",          "Created At (Auto)", False),
        ]),
        
    ])
 
    field_to_label = {f: lbl for _, fields in sections.items() for f, lbl, _ in fields}
    label_to_field = {lbl.strip().lower(): f for _, fields in sections.items() for f, lbl, _ in fields}
 
    system_injected = {
        "created_at",
    }
    helper_only_labels = {"brokerage label preview (auto)"}
    decimal_fields = {"built_up_area", "carpet_area"}
    int_fields = {
        "bathrooms", "balconies", "total_floors", "monthly_rent",
        "advance_rent_amount", "security_deposit_amount",
        "monthy_maintenance_amount", "total_move_in_cost",
    }
 
    return sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields
 
 
def _sample_row_data():
    """One complete example value per column so every column in the
    downloaded template shows the expected format — nothing left blank."""
    return {
       
        "listed_by_id": "rm0943", "listed_by_name": "Vikas", "listed_by_email": "vikas@test.com",
        "listed_by_contact": "9876543210", "listed_by_role": "Relationship Manager",

        "property_type": "Apartment", "property_no": "B-402", "bhk_type": "2bhk" , "renting_option": "Full Property",
        "built_up_area": "1200", "bathrooms": "2", "balconies": "1",
        "building_configuration": "G+3", "total_floors": "4", "facing_direction": "East",
        "furnishing_status": "Semi Furnished", "available_for": "Family",
 
        "carpet_area": "950", "city_zone": "West Zone", "ownership_type": "Freehold",
        "property_condition": "Well Maintained", "property_age": "5", "wing_number": "A Wing",
        "building_name": "Green Valley Apartments",
 
        "availability_status": "Available Immediately", "available_from": "2026-07-15",
        "lease_duration": "11 Months",   # only used when brokerage_percentage = "Fixed Amount"
 
        "monthly_rent": "25000","brokerage_percentage": "1 Month Rent",
        "manual_brokerage": "", "advance_rent_month": "1", "advance_rent_amount": "",
        "security_deposit_type": "2", "security_deposit_amount": "",
        "maintenance_type": "Included in Rent", "monthy_maintenance_amount": "",
        "total_move_in_cost": "75000",
 
        "address": "Flat 402, Green Valley", "city": "Nagpur", "locality_area": "Dharampeth",
        "property_landmark": "Near Railway Station", "state": "Maharashtra", "pincode": "440010",
        "main_road_connectivity": "Within 500 Meters",
        "google_maps_link": "https://maps.google.com/?q=21.1458,79.0882",
        "latitude": "21.1458", "longitude": "79.0882",
 
        "amenities": "Wi-Fi, AC, Lift", "nearby_facilities": "Metro, Hospital",
 
        "user_description": "Spacious and well-lit apartment close to all amenities.",
 
        "listed_elsewhere": "No", "portal_name": "",
    }






def _normalize_label(raw):
    if raw is None:
        return ""
    text = str(raw).replace("\u00a0", " ")           
    text = text.replace(" *", "").strip()
    text = re.sub(r"\s+", " ", text)                  
    return text.lower()


def _find_header_row(ws, label_to_field, max_scan_rows=6):
    best_row, best_score = None, -1
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        score = sum(1 for v in row_vals if _normalize_label(v) in label_to_field)
        if score > best_score:
            best_row, best_score = r, score
    return best_row, best_score




def _get_client_ip(request):
    """Helper to safely fetch client IP address reference."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')






from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from django.db.models import Q









def _is_sample_data_row(obj_data, sample):
    """Detects a row that's still the unedited example data from the
    downloaded template (agent forgot to delete/replace it before
    uploading)."""
    match_count = 0
    total_checked = 0
    for field, sample_val in sample.items():
        sample_val = str(sample_val).strip()
        if not sample_val:
            continue
        total_checked += 1
        row_val = str(obj_data.get(field, "")).strip()
        if row_val == sample_val:
            match_count += 1
    if total_checked == 0:
        return False
    return (match_count / total_checked) >= 0.9


def _identity_conflicts_with_session(obj_data, session_identity):
    """Returns True only if the row EXPLICITLY names a different person
    than whoever is logged in. Blank fields are fine (they just mean
    "use my own info") — only a value that's typed in AND disagrees
    with the session counts as a conflict."""
    pairs = [
        ('listed_by_email', 'email'),
        ('listed_by_contact', 'contact'),
        ('listed_by_name', 'name'),
        ('listed_by_role', 'role'),
    ]
    for field, key in pairs:
        row_val = str(obj_data.get(field, '')).strip()
        if not row_val:
            continue
        session_val = str(session_identity.get(key, '')).strip()
        if row_val.lower() != session_val.lower():
            return True
    return False


# =====================================================================
# DOWNLOAD TEMPLATE (agent) — sample row locked/read-only
# =====================================================================

def download_residential_template_agent(request):
    """Download the upload template. The Listed By identity columns are
    optional and auto-detected: leave them blank (or fill in your own
    details exactly) to list under your own name. Typing someone else's
    details will cause that row to be skipped on upload. The sample row
    (row 4) is protected as read-only; data-entry rows (5+) stay fully
    editable."""

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()
    sample = _sample_row_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental Residential"

    HDR_BG, REQ_BG, OPT_BG, SAMP_BG = "667EEA", "FEF3C7", "F0FDF4", "ECFDF5"
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    col = 1
    role_col = None
    brokerage_col = None

    for section, fields in sections.items():
        start_col = col
        for field, label, required in fields:
            header_text = label + (" *" if required else "")
            c1 = ws.cell(row=2, column=col, value=header_text)
            c1.font = Font(bold=True, color="1E293B", name="Arial", size=9)
            c1.fill = PatternFill("solid", fgColor=REQ_BG if required else OPT_BG)
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = bdr

            sc = ws.cell(row=4, column=col, value=sample.get(field, ""))
            sc.font = Font(name="Arial", size=9, color="065F46")
            sc.fill = PatternFill("solid", fgColor=SAMP_BG)
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sc.border = bdr
            sc.protection = Protection(locked=True)  # sample row is reference-only

            ws.column_dimensions[get_column_letter(col)].width = max(18, len(label) // 2 + 6)

            if field == "listed_by_role":
                role_col = col
            if field == "brokerage_percentage":
                brokerage_col = col

            # NEW: identity is auto-detected, not enforced by typing "Self".
            if field in ("listed_by_id", "listed_by_name", "listed_by_email", "listed_by_contact"):
                ws.cell(row=2, column=col).comment = Comment(
                    "Leave BLANK to auto-list under your own logged-in profile.\n"
                    "If you do fill it in, it must exactly match your own name/\n"
                    "email/contact/role — a row naming a DIFFERENT person will be\n"
                    "skipped on upload and flagged as an alert.",
                    "System"
                )
            if field == "listed_by_type":
                ws.cell(row=2, column=col).comment = Comment(
                    "Optional / informational only. Whether a row is treated as\n"
                    "'Self' is auto-detected from the identity fields, not from\n"
                    "this column.",
                    "System"
                )

            col += 1

        end_col = col - 1
        hc = ws.cell(row=1, column=start_col, value=f"\U0001F4CB {section}")
        hc.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hc.fill = PatternFill("solid", fgColor=HDR_BG)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = PatternFill("solid", fgColor=HDR_BG)
            ws.cell(row=1, column=cc).border = bdr
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    # ---- live brokerage label preview column, appended at the end ----
    preview_col = col
    pc = ws.cell(row=2, column=preview_col, value="Brokerage Label Preview (auto)")
    pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
    pc.fill = PatternFill("solid", fgColor="FEF3C7")
    pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pc.border = bdr
    ws.column_dimensions[get_column_letter(preview_col)].width = 26

    role_letter = get_column_letter(role_col)
    formula = (
        f"=IFERROR(INDEX('Notes - Brokerage Label'!$B$4:$B$9,"
        f"MATCH(LOWER(TRIM({role_letter}4)),'Notes - Brokerage Label'!$C$4:$C$9,0)),\"Brokerage\")"
    )
    fcell = ws.cell(row=4, column=preview_col, value=formula)
    fcell.fill = PatternFill("solid", fgColor="FEF3C7")
    fcell.font = Font(bold=True, color="92400E", name="Arial", size=9)
    fcell.alignment = Alignment(horizontal="center", vertical="center")
    fcell.protection = Protection(locked=True)

    if brokerage_col:
        note = (
            "The LABEL text shown above this field on the live form changes based on\n"
            "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
            "Admin -> EstateFlow Service Fee\n"
            "Relationship Manager -> Service Fee\n"
            "Landlord -> Tenant Service Fee\n"
            "Agent -> Brokerage\n"
            "Agency/Builder or Builder -> Service Fee\n"
            "Any other role -> Brokerage (default)\n\n"
            "See 'Notes - Brokerage Label' sheet, and the live preview column at the end of this sheet."
        )
        ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    # ---- Notes sheet (lookup table the formula above reads from) ----
    notes = wb.create_sheet("Notes - Brokerage Label")
    notes.column_dimensions['A'].width = 26
    notes.column_dimensions['B'].width = 26
    notes.column_dimensions['C'].width = 4
    notes.sheet_view.showGridLines = False

    t = notes["A1"]
    notes.merge_cells("A1:B1")
    t.value = "Brokerage label — driven by Listed By Role"
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["Listed By Role", "Label shown on form"]
    for i, h in enumerate(hdrs, start=1):
        c = notes.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="1E293B", name="Arial")
        c.fill = PatternFill("solid", fgColor=OPT_BG)
        c.border = bdr

    role_rows = [
        ("Admin", "EstateFlow Service Fee"),
        ("Relationship Manager", "Service Fee"),
        ("Landlord", "Tenant Service Fee"),
        ("Agent", "Brokerage"),
        ("Agency/Builder", "Service Fee"),
        ("Builder", "Service Fee"),
    ]
    for r, (role, label) in enumerate(role_rows, start=4):
        notes.cell(row=r, column=1, value=role).border = bdr
        notes.cell(row=r, column=2, value=label).border = bdr
        notes.cell(row=r, column=3, value=f"=LOWER(TRIM(A{r}))")

    # ---- lock the sample row, unlock every real data-entry cell ----
    max_col = preview_col
    MAX_DATA_ROWS = 1000

    for c in range(1, max_col + 1):
        ws.cell(row=4, column=c).protection = Protection(locked=True)

    for r in range(5, 5 + MAX_DATA_ROWS):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Template.xlsx"'
    wb.save(response)
    return response


# =====================================================================
# IMPORT (agent) — auto-detect Self by identity match, skip+alert on mismatch
# =====================================================================

@csrf_exempt
@require_POST
def import_residential_excel_agent(request):
    excel_file = request.FILES.get("rental_file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx files allowed."}, status=400)

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()
    sample = _sample_row_data()

    # Identity fields are NOT in this list anymore — they're auto-derived
    # from the session / matched against it, never required as typed text.
    REQUIRED_FIELD_KEYS = [
        'property_type',
        'property_no',
        'bhk_type',
        'renting_option',
        'built_up_area',
        'bathrooms',
        'building_configuration',
        'total_floors',
        'furnishing_status',
        'available_for',
        'monthly_rent',
        'address',
        'city',
        'locality_area',
        'state',
        'pincode',
    ]

    def _field_label(field):
        return field_to_label.get(field) or field.replace('_', ' ').title()

    def _is_missing(val):
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    # ---- 1. Uploader Identity (system audit trail — who UPLOADED the file) ----
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    admin_obj = Admin_Login.objects.filter(id=admin_id).first() if admin_id else None
    user_obj = User_Details.objects.filter(id=user_id).first() if user_id else None

    uploader_name = uploader_email = uploader_contact = ""
    uploader_role = "Automated Engine"
    user_identity = "Automated Engine"

    if admin_obj:
        uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
        uploader_email = getattr(admin_obj, 'email', '')
        uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
        uploader_role = "Admin"
        user_identity = uploader_email or uploader_name
    elif user_obj:
        uploader_name = user_obj.user_name
        uploader_email = user_obj.user_email
        uploader_contact = user_obj.user_phone
        uploader_role = "User"
        user_identity = uploader_email or uploader_name

    # ---- 1b. Determine WHO IS ACTUALLY LOGGED IN — this is the identity
    # every row gets checked against. Covers three cases: a plain Agent
    # session, an Admin impersonating an Agent, and a plain Admin session
    # (Admin has full access and can upload under their own name too). ----
    session_user_id = request.session.get('User_id')
    session_admin_id = request.session.get('Admin_id')
    session_role = request.session.get('user_type')

    is_valid_agent = (session_user_id and session_role == "Agent")
    is_valid_admin_impersonating = (
        session_admin_id and session_role == "Admin" and 'impersonate_id' in request.session
    )

    agent_obj = None
    if is_valid_admin_impersonating:
        agent_obj = User_Details.objects.filter(id=request.session.get('impersonate_id')).first()
    elif is_valid_agent:
        agent_obj = User_Details.objects.filter(id=session_user_id).first()

    session_identity = None
    if agent_obj:
        session_identity = {
            'id': agent_obj.user_id,
            'name': agent_obj.user_name,
            'email': agent_obj.user_email,
            'contact': agent_obj.user_phone,
            'role': agent_obj.user_role,
        }
    elif admin_obj:
        session_identity = {
            'id': str(admin_obj.id),
            'name': uploader_name,
            'email': uploader_email,
            'contact': uploader_contact,
            'role': "Admin",
        }

    if not session_identity:
        return JsonResponse({
            "status": "error",
            "message": "No logged-in session found. This upload requires an active Agent or Admin session.",
        }, status=400)

    # ---- 2. Parse Excel ----
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Rental Residential"] if "Rental Residential" in wb.sheetnames else wb.active
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    header_row, matched_count = _find_header_row(ws, label_to_field)

    if matched_count == 0:
        return JsonResponse({
            "status": "error",
            "message": (
                "No recognizable column headers were found in this file. "
                "Please use the official template downloaded from "
                "'Download Template' — don't rename or re-order the header row."
            ),
        }, status=400)

    raw_headers = [cell.value for cell in ws[header_row]]
    field_headers = []
    unmatched_headers = []

    for h in raw_headers:
        norm = _normalize_label(h)
        if not norm:
            field_headers.append(None)
            continue
        if norm in helper_only_labels:
            field_headers.append(None)
            continue
        field = label_to_field.get(norm)
        field_headers.append(field)
        if field is None:
            unmatched_headers.append(str(h))

    data_start_row = header_row + 1

    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []      # missing required fields / sample row -> whole file rejected
    identity_mismatch_errors = []   # row names a different person -> that row is skipped, upload continues
    skipped_identity_mismatch = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        obj_data = {}
        for col_idx, field in enumerate(field_headers):
            if not field or field in system_injected:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip() != "":
                obj_data[field] = val

        if not obj_data:
            skipped_empty_after_mapping += 1
            continue

        # ---- reject the whole file if this is still the sample row ----
        if _is_sample_data_row(obj_data, sample):
            required_field_errors.append({
                "row": row_idx,
                "missing_fields": [
                    "This row still contains the example/sample data from the "
                    "downloaded template. Replace it with your actual property "
                    "details (or delete the row) before uploading."
                ],
            })
            continue

        # ---- Type Coercion (unchanged) ----
        if 'available_from' in obj_data:
            d_val = obj_data['available_from']
            if isinstance(d_val, str):
                c_str = d_val.strip().split(" ")[0]
                for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        obj_data['available_from'] = datetime.strptime(c_str, fmt).date()
                        break
                    except ValueError:
                        obj_data['available_from'] = None
            elif isinstance(d_val, datetime):
                obj_data['available_from'] = d_val.date()

        for f in int_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = int(float(str(obj_data[f]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    obj_data[f] = None

        for f in decimal_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = Decimal(str(obj_data[f]).replace(",", "").strip())
                except (InvalidOperation, ValueError):
                    obj_data[f] = None

        for f in ("advance_rent_month", "security_deposit_type"):
            if f in obj_data and obj_data[f] is not None and str(obj_data[f]).lower() != "fixed":
                try:
                    obj_data[f] = str(int(float(obj_data[f])))
                except (TypeError, ValueError):
                    obj_data[f] = str(obj_data[f]).strip()

        # ---- Required property fields (unrelated to identity) ----
        missing_fields = [
            _field_label(f) for f in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(f))
        ]
        if missing_fields:
            required_field_errors.append({
                "row": row_idx,
                "missing_fields": missing_fields,
            })
            continue

        # ---- NEW: identity auto-detection ----
        # Blank identity fields -> this row is automatically the logged-in
        # user's own listing. Any filled-in field that DISAGREES with the
        # logged-in identity means this row belongs to someone else -> skip
        # it and alert, rather than silently overwriting or rejecting the
        # whole file.
        if _identity_conflicts_with_session(obj_data, session_identity):
            l_role = str(obj_data.get('listed_by_role', '')).strip() or 'user'
            searched_info = filter(None, [
                str(obj_data.get('listed_by_name', '')).strip(),
                str(obj_data.get('listed_by_email', '')).strip(),
                str(obj_data.get('listed_by_contact', '')).strip(),
                str(obj_data.get('listed_by_role', '')).strip(),
            ])
            identity = " + ".join(searched_info) or "Unknown"
            identity_mismatch_errors.append({
                "row": row_idx,
                "errors": [
                    f"Listed By {l_role} '{identity}' does not match your logged-in "
                    f"profile. This upload only accepts your own listings — row skipped."
                ],
            })
            skipped_identity_mismatch += 1
            continue

        # Match (or blank) -> auto-fill from the session, ignore whatever
        # (if anything) was typed.
        obj_data['listed_by_id'] = session_identity['id']
        obj_data['listed_by_name'] = session_identity['name']
        obj_data['listed_by_email'] = session_identity['email']
        obj_data['listed_by_contact'] = session_identity['contact']
        obj_data['listed_by_role'] = session_identity['role']
        obj_data['listed_by_type'] = 'Self'

        parsed_rows.append({'row_idx': row_idx, 'data': obj_data})

    wb.close()

    # ---- Bail out if any row failed required-field / sample-row validation ----
    if required_field_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields "
                "or still contain sample data. Please fix these rows and re-upload. No records were saved."
            ),
            "row_errors": required_field_errors,
        }, status=400)

    # ---- Bail out if nothing usable was found ----
    if not parsed_rows and not identity_mismatch_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"0 usable data rows found. Detected header row {header_row}, "
                f"data expected from row {data_start_row} onward. "
                f"{skipped_empty_after_mapping} row(s) had values but none matched a known column."
            ),
            "unmatched_headers": unmatched_headers,
            "header_row_detected": header_row,
            "data_start_row_assumed": data_start_row,
        }, status=400)

    # ---- 3. Whole-file duplicate check (unchanged) ----
    file_name_exists = RentalResidentialProperty.objects.filter(
        upload_file_name=excel_file.name
    ).exists()

    # ---- 4. Write to DB (fingerprint-based duplicate engine) ----
    created, updated, skipped, errors = (
        0, 0, skipped_empty_after_mapping + skipped_identity_mismatch, []
    )
    duplicate_blocked_rows = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality_area', '')).strip()
        input_pincode = str(o_data.get('pincode', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_name = str(o_data.get('listed_by_name', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        fingerprint_key = generate_property_fingerprint(
            input_property_no,
            input_building_name,
            input_locality,
            input_pincode
        )

        direct_duplicates = RentalResidentialProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality_area__iexact=input_locality
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked - Unit '{input_property_no}' is already listed "
                    f"by/for {input_listed_by_name or 'this user'}. Row skipped; edit the existing listing instead."
                )
                skipped += 1
                continue

            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id
            )

        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["listing_type"] = o_data.get("listing_type") or "Rental"
        o_data["category"] = o_data.get("category") or "Residential"

        o_data["upload_file_name"] = excel_file.name
        o_data["uploaded_by_name"] = uploader_name
        o_data["uploaded_by_email"] = uploader_email
        o_data["uploaded_by_contact"] = uploader_contact
        o_data["uploaded_by_role"] = uploader_role

        try:
            RentalResidentialProperty.objects.create(**o_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    errors.extend(duplicate_blocked_rows)

    # NEW: fold identity-mismatch rows into the alert list
    for entry in identity_mismatch_errors:
        for msg in entry["errors"]:
            errors.append(f"Row {entry['row']}: {msg}")

    # ---- 5. Audit Log ----
    RentalActivityLog.objects.create(
        user_identity=user_identity,
        user_role=uploader_role,
        action_type='EXCEL_IMPORT',
        property_id="Multiple / Sheet Records",
        targeted_fields="bulk_action",
        associated_file=excel_file.name,
        action_payload=json.dumps({
            "filename": excel_file.name,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "duplicates_blocked": len(duplicate_blocked_rows),
            "identity_mismatches": len(identity_mismatch_errors),
            "errors_encountered": len(errors),
        }),
        ip_address=_get_client_ip(request),
        status='SUCCESS' if not errors else 'PARTIAL',
    )

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created, "updated": updated, "skipped": skipped,
        "duplicates_blocked": len(duplicate_blocked_rows),
        "identity_mismatches": len(identity_mismatch_errors),
        "error_count": len(errors), "errors": errors,
        "unmatched_headers": unmatched_headers,
        "header_row_detected": header_row,
        "data_start_row_used": data_start_row,
    })

########################Views ENDS For Rental residential listing dowload and import ####################################



#################Views Start For Rental Commericial Property####################################



def commercial_agent(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Rental/commercial.html", context)





def commercial_rental_add_agent(request):

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    

    # ---------- SAFE TYPE CONVERSIONS ----------
    def to_int(val, default=0):
        try: 
            return int(str(val).replace(',', '').strip())
        except: 
            return default

    def to_int_or_none(val):
        try: 
            return int(str(val).replace(',', '').strip())
        except: 
            return None

    def to_float_or_none(val):
        try: 
            return float(str(val).replace(',', '').strip())
        except: 
            return None

    if request.method == "POST":
        try:
            # ---------- UPLOADER IDENTIFICATION ----------
            user = User_Details.objects.get(id=user_id)
            uploader_name = user.user_name
            uploader_email = user.user_email
            uploader_phone = user.user_phone
            uploader_role = "Agent"
            uploader_id = f"USER_{user_id}"

            # ---------- LISTED BY IDENTIFICATION ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_phone).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

            # ==========================================================
            # DUPLICATE DETECTION ENGINE
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality') or request.POST.get('locality_area') or '').strip()
            input_city = (request.POST.get('city') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            # Generate unique fingerprint key
            fingerprint_key = f"{input_property_no}|{input_building_name}|{input_locality}|{input_city}|{input_pincode}".lower().replace(" ", "")

            # 1. Direct Case-Insensitive Query
            direct_duplicates = CommercialRentalProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            existing_duplicates = (
                CommercialRentalProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        return JsonResponse({
                            "status": "error", 
                            "message": f"Duplicate Blocked: This commercial unit ({input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing Property instead."
                        })

                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- AMENITIES & FACILITIES ----------
            amenities_list = ", ".join(request.POST.getlist('amenities[]'))
            facilities_list = ", ".join(request.POST.getlist('nearby_facilities[]'))

            # ---------- CREATE DATABASE OBJECT ----------
            prop = CommercialRentalProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Commercial",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                building_name=input_building_name,
                wing_number=request.POST.get('wing_number'),
                property_no=input_property_no,

                availability_status=request.POST.get('availability_status'),
                available_from=available_from,
                property_age=to_int_or_none(request.POST.get('property_age')),

                zone_type=request.POST.get('zone_type'),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition'),

                builtup_area=to_decimal_or_none(request.POST.get('builtup_area')),
                carpet_area=to_decimal_or_none(request.POST.get('carpet_area')),

                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                address=request.POST.get('address'),
                locality=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                city=request.POST.get('city'),
                state=request.POST.get('state'),
                location_hub=request.POST.get('location_hub'),
                pincode=input_pincode,
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                dg_ups_included=(request.POST.get('dg_ups_included') == 'on'),
                electricity_included=(request.POST.get('electricity_included') == 'on'),
                water_included=(request.POST.get('water_included') == 'on'),

                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int_or_none(request.POST.get('total_floors')),
                staircases=to_int_or_none(request.POST.get('staircases')),
                passenger_lifts=to_int(request.POST.get('passenger_lifts')),
                service_lifts=to_int(request.POST.get('service_lifts')),
                private_parking=to_int(request.POST.get('private_parking')),

                min_seats=to_int_or_none(request.POST.get('min_seats')),
                max_seats=to_int_or_none(request.POST.get('max_seats')),
                cabins=to_int_or_none(request.POST.get('cabins')),
                meeting_rooms=to_int_or_none(request.POST.get('meeting_rooms')),
                private_washroom=to_int(request.POST.get('private_washroom')),
                public_washroom=to_int(request.POST.get('public_washroom')),
                flooring_type=request.POST.get('flooring_type'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advanced_rent_type=request.POST.get('advanced_rent_type'),
                advanced_rent_amount=to_int_or_none(request.POST.get('advanced_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int_or_none(request.POST.get('security_deposit_amount')),
                maintenance_type=request.POST.get('maintenance_type'),
                maintenance_charges=to_int_or_none(request.POST.get('maintenance_charges')),
                total_move_in_cost=to_int_or_none(request.POST.get('total_move_in_cost')),
                negotiable=request.POST.get('negotiable', 'No'),
                lockin_period=to_int_or_none(request.POST.get('lockin_period')),
                rent_increase=to_float_or_none(request.POST.get('rent_increase')),

                amenities=amenities_list,
                nearby_facilities=facilities_list,
                user_description=request.POST.get('user_description'),
                property_summary=request.POST.get('property_summary'),
                property_description=request.POST.get('property_description'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_phone,
                uploaded_by_role=uploader_role,
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC (12 CATEGORIES) ----------
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'reception': 'property_images_reception[]',
                'workspace': 'property_images_workspace[]',
                'cabins': 'property_images_cabins[]',
                'meeting_room': 'property_images_meeting_room[]',
                'shop_showroom': 'property_images_shop_showroom[]',
                'warehouse': 'property_images_warehouse[]',
                'pantry': 'property_images_pantry[]',
                'washroom': 'property_images_washroom[]',
                'parking': 'property_images_parking[]',
                'amenities': 'property_images_amenities[]',
                'floor_plan': 'property_images_floor_plan[]',
            }
            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                for cat_index, img in enumerate(request.FILES.getlist(field_name)):
                    if saved_count >= 30:
                        break
                    CommercialRentalPropertyImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=cat_index
                    )
                    saved_count += 1

            # ---------- PROPERTY VIDEO (UPLOAD, AUTO, OR RM ASSISTED) ----------
            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '')

            # 1. ALWAYS auto-generate the slideshow (runs every time, unconditionally)
            saved_images = list(CommercialRentalPropertyImage.objects.filter(property=prop))
            image_paths = []
            for img_obj in saved_images:
                if img_obj.image and hasattr(img_obj.image, 'path') and os.path.exists(img_obj.image.path):
                    image_paths.append(img_obj.image.path)

            if len(image_paths) >= 3:
                try:
                    from Admin_App.utils import generate_property_slideshow
                    output_relative_path = f"commercial_rent/videos/auto_{prop.id}.mp4"
                    
                    result_path = generate_property_slideshow(image_paths, output_relative_path)
                    print("AUTO SLIDESHOW RESULT:", result_path)
                    
                    if result_path:
                        CommercialRentalVideo.objects.update_or_create(
                            property=prop,
                            source='auto',
                            defaults={
                                'video': result_path,
                                'video_status': 'Done',
                                'video_url': None
                            }
                        )
                except Exception as ve:
                    import traceback
                    print("COMMERCIAL VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()
            else:
                CommercialRentalVideo.objects.update_or_create(
                    property=prop,
                    source='auto',
                    defaults={
                        'video': None,
                        'video_status': 'Pending (Insufficient Images)',
                        'video_url': None
                    }
                )
            # 2. Manual Upload Logic (separate row, source='uploaded')
            if video_option == 'upload' and uploaded_video:
                if uploaded_video.size > 20 * 1024 * 1024:
                    CommercialRentalVideo.objects.create(
                        property=prop,
                        video=None,
                        source='uploaded',
                        video_status='Pending',
                        video_url=None
                    )
                else:
                    CommercialRentalVideo.objects.create(
                        property=prop,
                        video=uploaded_video,
                        source='uploaded',
                        video_status='Done',
                        video_url=None
                    )

            # 3. RM Assisted Link Logic (separate row, source='rm_assisted')
            elif video_option == 'rm_assisted':
                CommercialRentalVideo.objects.create(
                    property=prop,
                    video=None,
                    source='rm_assisted',
                    video_url=property_video_link,
                    video_status='Done' if property_video_link else 'Pending'
                )

            

            return JsonResponse({"status": "success", "message": "Commercial Rental Property Added Successfully"})

        except Exception as e:
            print("ERROR IN COMMERCIAL ADD:", str(e))
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})

    return render(request, 'agent/Forms/Rental/commercial.html'
   , {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        
        
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    })






def commercial_agent_edit(request, pk):
    prop = get_object_or_404(CommercialRentalProperty, id=pk)

    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    

    def to_int(val, default=0):
        try: return int(str(val).replace(',', '').strip())
        except: return default

    def to_int_or_none(val):
        try: return int(str(val).replace(',', '').strip())
        except: return None

    def to_float_or_none(val):
        try: return float(str(val).replace(',', '').strip())
        except: return None

    if request.method == "POST":
        try:
            p = request.POST

            date_val = p.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except Exception:
                    available_from = None
            else:
                available_from = None

            # ---------- LISTED BY ----------
            prop.listed_by_type    = p.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to       = p.get('assigned_to', prop.assigned_to)
            prop.listed_by_id      = p.get('listed_by_id') or prop.listed_by_id
            prop.listed_by_name    = p.get('listed_by_name') or prop.listed_by_name
            prop.listed_by_email   = (p.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = p.get('listed_by_contact') or prop.listed_by_contact
            prop.listed_by_role    = p.get('listed_by_role') or prop.listed_by_role

            # ---------- BASIC INFO ----------
            prop.property_title = p.get('property_title', prop.property_title)
            prop.property_type  = p.get('property_type')
            prop.building_name  = (p.get('building_name') or '').strip()
            prop.wing_number    = p.get('wing_number')
            prop.property_no    = (p.get('property_no') or '').strip()

            prop.availability_status = p.get('availability_status')
            prop.available_from      = available_from
            prop.property_age        = to_int_or_none(p.get('property_age'))

            prop.zone_type         = p.get('zone_type')
            prop.ownership_type    = p.get('ownership_type')
            prop.property_condition = p.get('property_condition')

            prop.builtup_area = to_decimal_or_none(p.get('builtup_area'))
            prop.carpet_area  = to_decimal_or_none(p.get('carpet_area'))

            # ---------- LOCATION ----------
            prop.address           = p.get('address')
            prop.locality          = (p.get('locality') or '').strip()
            prop.property_landmark = p.get('property_landmark')
            prop.city              = p.get('city')
            prop.state             = p.get('state')
            prop.location_hub      = p.get('location_hub')
            prop.pincode           = (p.get('pincode') or '').strip()
            prop.google_maps_link  = p.get('google_maps_link')
            prop.latitude          = p.get('latitude')
            prop.longitude         = p.get('longitude')

            # ---------- BUILDING SPECS ----------
            prop.dg_ups_included      = (p.get('dg_ups_included') == 'on')
            prop.electricity_included = (p.get('electricity_included') == 'on')
            prop.water_included       = (p.get('water_included') == 'on')

            prop.building_configuration = p.get('building_configuration')
            prop.total_floors    = to_int_or_none(p.get('total_floors'))
            prop.staircases      = to_int_or_none(p.get('staircases'))
            prop.passenger_lifts = to_int(p.get('passenger_lifts'))
            prop.service_lifts   = to_int(p.get('service_lifts'))
            prop.private_parking = to_int(p.get('private_parking'))

            prop.min_seats         = to_int_or_none(p.get('min_seats'))
            prop.max_seats         = to_int_or_none(p.get('max_seats'))
            prop.cabins            = to_int_or_none(p.get('cabins'))
            prop.meeting_rooms     = to_int_or_none(p.get('meeting_rooms'))
            prop.private_washroom  = to_int(p.get('private_washroom'))
            prop.public_washroom   = to_int(p.get('public_washroom'))
            prop.flooring_type     = p.get('flooring_type')

            # ---------- PRICING ----------
            prop.monthly_rent            = to_int(p.get('monthly_rent'))
            prop.brokerage_percentage    = p.get('brokerage_percentage')
            prop.manual_brokerage        = p.get('manual_brokerage')
            prop.advanced_rent_type      = p.get('advanced_rent_type')
            prop.advanced_rent_amount    = to_int_or_none(p.get('advanced_rent_amount'))
            prop.security_deposit_type   = p.get('security_deposit_type')
            prop.security_deposit_amount = to_int_or_none(p.get('security_deposit_amount'))
            prop.maintenance_type        = p.get('maintenance_type')
            prop.maintenance_charges     = to_int_or_none(p.get('maintenance_charges'))
            prop.total_move_in_cost      = to_int_or_none(p.get('total_move_in_cost'))
            prop.negotiable              = p.get('negotiable', prop.negotiable)
            prop.lockin_period           = to_int_or_none(p.get('lockin_period'))
            prop.rent_increase           = to_float_or_none(p.get('rent_increase'))

            # ---------- AMENITIES / DESCRIPTIONS ----------
            prop.amenities            = ", ".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities    = ", ".join(request.POST.getlist('nearby_facilities[]'))
            prop.user_description     = p.get('user_description')
            prop.property_summary     = p.get('property_summary') or prop.property_summary
            prop.property_description = p.get('property_description') or prop.property_description

            # ---------- LISTED ELSEWHERE ----------
            prop.listed_elsewhere = p.get('listed_elsewhere', 'No')
            prop.portal_name      = p.get('portal_name')
            prop.listing_status = p.get('listing_status')
            prop.approval_status = p.get('approval_status')

            prop.save()
            # ================= IMAGE DELETIONS =================
            delete_ids = request.POST.getlist('delete_image_ids[]')
            if delete_ids:
                CommercialRentalPropertyImage.objects.filter(id__in=delete_ids, property=prop).delete()

            # ================= NEW IMAGES BY CATEGORY =================
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'reception': 'property_images_reception[]',
                'workspace': 'property_images_workspace[]',
                'cabins': 'property_images_cabins[]',
                'meeting_room': 'property_images_meeting_room[]',
                'shop_showroom': 'property_images_shop_showroom[]',
                'warehouse': 'property_images_warehouse[]',
                'pantry': 'property_images_pantry[]',
                'washroom': 'property_images_washroom[]',
                'parking': 'property_images_parking[]',
                'amenities': 'property_images_amenities[]',
                'floor_plan': 'property_images_floor_plan[]',
            }
            existing_total = prop.images.count()
            new_images_added = 0

            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                existing_cat_count = prop.images.filter(category=category).count()
                for img in request.FILES.getlist(field_name):
                    if existing_total >= 30:
                        break
                    CommercialRentalPropertyImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=existing_cat_count
                    )
                    existing_cat_count += 1
                    existing_total += 1
                    new_images_added += 1

            images_changed = new_images_added > 0 or bool(delete_ids)

            # ================= VIDEO: SOURCE THE ADMIN ACTIVELY PICKED =================
            # This part stays exactly gated by video_option — only touches the source
            # the admin is currently interacting with (upload file OR RM link).
            video_option = p.get('video_option')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = p.get('property_video_link', '')

            if video_option == 'upload' and uploaded_video:
                video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='uploaded')
                if uploaded_video.size > 20 * 1024 * 1024:
                    video_row.video = None
                    video_row.video_status = 'Pending'
                else:
                    video_row.video = uploaded_video
                    video_row.video_status = 'Done'
                video_row.save()

            elif video_option == 'rm_assisted' and property_video_link:
                video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='rm_assisted')
                video_row.video_url = property_video_link
                video_row.video_status = 'Done'
                video_row.save()

            # ================= AUTO SLIDESHOW: ALWAYS KEEPS PACE WITH PHOTOS =================
            # Completely independent of video_option — the 'auto' row is a separate
            # database record from 'uploaded'/'rm_assisted', so it must refresh whenever
            # photos change, even if the admin is currently uploading a video manually.
            # This mirrors the add view, which regenerates the slideshow unconditionally.
            # ================= AUTO SLIDESHOW: ALWAYS KEEPS PACE WITH PHOTOS =================
            regenerate_requested = p.get('regenerate_slideshow') == 'on'

            if images_changed or regenerate_requested:
                saved_images = list(prop.images.all())
                image_paths = [
                    img.image.path for img in saved_images
                    if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                ]
                
                if len(image_paths) >= 3:
                    try:
                        from Admin_App.utils import generate_property_slideshow
                        import time

                        # Delete the old physical file before writing a new one, so we don't
                        # leave orphaned video files piling up on disk with every edit.
                        video_row = CommercialRentalVideo.objects.filter(property=prop, source='auto').first()
                        if video_row and video_row.video and video_row.video.name:
                            old_path = video_row.video.path
                            if os.path.exists(old_path):
                                try:
                                    os.remove(old_path)
                                except Exception as del_err:
                                    print("Could not delete old auto video:", del_err)

                        # Unique filename per regeneration to break browser caching
                        output_relative_path = f"commercial_rent/videos/auto_{prop.id}_{int(time.time())}.mp4"

                        result_path = generate_property_slideshow(image_paths, output_relative_path)
                        
                        if result_path:
                            CommercialRentalVideo.objects.update_or_create(
                                property=prop,
                                source='auto',
                                defaults={
                                    'video': result_path,
                                    'video_status': 'Done',
                                    'video_url': None
                                }
                            )
                    except Exception as ve:
                        import traceback
                        print("COMMERCIAL VIDEO REGEN FAILED:", str(ve))
                        traceback.print_exc()
                else:
                    CommercialRentalVideo.objects.update_or_create(
                        property=prop,
                        source='auto',
                        defaults={
                            'video': None,
                            'video_status': 'Pending (Insufficient Images)',
                            'video_url': None
                        }
                    )
            return JsonResponse({"status": "success", "message": "Commercial Rental Property Updated Successfully"})

        except Exception as e:
            print("ERROR IN COMMERCIAL EDIT:", str(e))
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})

    # ================= GET REQUEST =================
    videos_by_source = {v.source: v for v in prop.walkthrough_video.all()}
    images_by_category = {}
    for img in prop.images.all().order_by('category', 'sequence_order'):
        images_by_category.setdefault(img.category, []).append(img)

    context = {
        
        'prop': prop,
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
      
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
      
        'uploaded_video': videos_by_source.get('uploaded'),
        'auto_video': videos_by_source.get('auto'),
        'rm_video': videos_by_source.get('rm_assisted'),
        'images_by_category': images_by_category,
        'selected_amenities': [a.strip() for a in (prop.amenities or '').split(',') if a.strip()],
        'selected_facilities': [f.strip() for f in (prop.nearby_facilities or '').split(',') if f.strip()],
    }
    return render(request, 'agent/Forms/Rental/commercial_edit.html', context)


    
  



def commercial_view_agent(request, pk):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    prop = get_object_or_404(CommercialRentalProperty, pk=pk)

    amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []
    facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []

    # Video Hierarchy & Backend Saved Video Retrieval
    videos_queryset = prop.walkthrough_video.all()
    rm_video = videos_queryset.filter(source='rm_assisted', video_url__isnull=False).exclude(video_url='').first()
    manual_video = videos_queryset.filter(source='uploaded', video__isnull=False).first()
    auto_video = videos_queryset.filter(source='auto', video__isnull=False).first()

    selected_video = None
    video_display_mode = None

    if rm_video:
        selected_video = rm_video
        video_display_mode = 'rm_assisted'
    elif manual_video:
        selected_video = manual_video
        video_display_mode = 'manual'
    elif auto_video:
        selected_video = auto_video
        video_display_mode = 'auto'

    return render(request, 'agent/Reports/Rental/commercial_detail.html', {
        
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        
        'prop': prop,
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'selected_video': selected_video,
        'video_display_mode': video_display_mode,
    })






def commercial_list_agent(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    
    

    # ── 1. GET ALL SEARCH PARAMS ──
    search_query        = request.GET.get('search', '').strip()
    prop_type_query      = request.GET.get('property_type', '').strip()
    city_query           = request.GET.get('city', '').strip()
    zone_query           = request.GET.get('zone_type', '').strip()
    possession_query     = request.GET.get('possession', '').strip()
    listed_by_query      = request.GET.get('listed_by', '').strip()
    uploaded_by_query    = request.GET.get('uploaded_by', '').strip()
    budget_query         = request.GET.get('budget', '').strip()
    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query      = request.GET.get('duplicate', '').strip()
    from_date            = request.GET.get('from_date', '').strip()
    to_date              = request.GET.get('to_date', '').strip()

    # ── Base queryset ──
    try:
        properties = CommercialRentalProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    except Exception:
        properties = CommercialRentalProperty.objects.filter(listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    # ── 2. APPLY FILTERS ──
    # NOTE: search now checks across every meaningful text/identifier field,
    # not just title/city/locality/owner name.
    if search_query:
        properties = properties.filter(
            Q(id__icontains=search_query) |
            Q(property_unique_key__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(property_no__icontains=search_query) |
            Q(wing_number__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(zone_type__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_contact__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(listing_status__icontains=search_query) |
            Q(approval_status__icontains=search_query)
        )

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__icontains=prop_type_query)
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if zone_query and zone_query != 'All Zones':
        properties = properties.filter(zone_type__icontains=zone_query)
    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__icontains=possession_query)
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__icontains=listed_by_query)
    if uploaded_by_query and uploaded_by_query != 'All Roles':
        properties = properties.filter(uploaded_by_role__icontains=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)
    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if from_date:
        try:
            properties = properties.filter(created_at__date__gte=from_date)
        except Exception:
            pass
    if to_date:
        try:
            properties = properties.filter(created_at__date__lte=to_date)
        except Exception:
            pass

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_25k':
            properties = properties.filter(monthly_rent__lt=25000)
        elif budget_query == '25k_1L':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=100000)
        elif budget_query == '1L_5L':
            properties = properties.filter(monthly_rent__gte=100000, monthly_rent__lte=500000)
        elif budget_query == 'above_5L':
            properties = properties.filter(monthly_rent__gt=500000)

    properties = properties.prefetch_related(
        Prefetch(
            'walkthrough_video',
            queryset=CommercialRentalVideo.objects.all(),
            to_attr='all_videos'
        )
    )

    # ════════════════════════════════════════════════
    # ⛔ EXCEL / CSV DOWNLOAD BLOCK REMOVED
    # (the openpyxl / csv export code that used to live here has been
    #  deleted per request — this view no longer serves file downloads)
    # ════════════════════════════════════════════════

    # ── Pagination ──
    paginator      = Paginator(properties, 10)
    page_number    = request.GET.get('page', 1)
    page_obj       = paginator.get_page(page_number)
    filtered_count = properties.count()

    for prop in page_obj:
        videos_by_source = {v.source: v for v in prop.all_videos}
        prop.uploaded_video = videos_by_source.get('uploaded')
        prop.auto_video     = videos_by_source.get('auto')
        prop.rm_video       = videos_by_source.get('rm_assisted')

    # ════════════════════════════════════════════════
    # ALL-PROPS STATS (always on unfiltered dataset)
    # ════════════════════════════════════════════════
    try:
        all_props = CommercialRentalProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
    except Exception:
        all_props = CommercialRentalProperty.filter.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
         

    total_count = all_props.count()

    # ── Dropdown unique values ──
    unique_property_types = (all_props
        .exclude(property_type__isnull=True).exclude(property_type='')
        .values_list('property_type', flat=True).distinct().order_by('property_type'))
    unique_cities = (all_props
        .exclude(city__isnull=True).exclude(city='')
        .values_list('city', flat=True).distinct().order_by('city'))
    unique_zones = (all_props
        .exclude(zone_type__isnull=True).exclude(zone_type='')
        .values_list('zone_type', flat=True).distinct().order_by('zone_type'))
    unique_possession = (all_props
        .exclude(availability_status__isnull=True).exclude(availability_status='')
        .values_list('availability_status', flat=True).distinct())
    unique_roles = (all_props
        .exclude(listed_by_role__isnull=True).exclude(listed_by_role='')
        .values_list('listed_by_role', flat=True).distinct())
    unique_uploaded_roles = (all_props
        .exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='')
        .values_list('uploaded_by_role', flat=True).distinct())
    unique_listing_status = (all_props
        .exclude(listing_status__isnull=True).exclude(listing_status='')
        .values_list('listing_status', flat=True).distinct())
    unique_approval_status = (all_props
        .exclude(approval_status__isnull=True).exclude(approval_status='')
        .values_list('approval_status', flat=True).distinct())

    # ── Occupancy KPIs ──
    active_count   = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    occupied_count = all_props.filter(availability_status__iexact='currently_occupied').count()
    vacant_count   = all_props.filter(availability_status__iexact='available_immediately').count()
    occupancy_rate = round((occupied_count / total_count * 100)) if total_count > 0 else 0
    vacancy_rate   = round((vacant_count   / total_count * 100)) if total_count > 0 else 0

    # ── Revenue KPIs ──
    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )
    avg_rent = rent_stats['avg_rent'] or 0
    max_rent = rent_stats['max_rent'] or 0
    min_rent = rent_stats['min_rent'] or 0
    total_revenue          = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0
    avg_deposit = (all_props.exclude(security_deposit_amount__isnull=True)
                   .aggregate(avg=Avg('security_deposit_amount'))['avg'] or 0)

    try:
        avg_area = (all_props.exclude(builtup_area__isnull=True)
                    .aggregate(avg=Avg('builtup_area'))['avg'] or 0)
    except Exception:
        avg_area = 0

    # ── Business KPIs ──
    premium_properties_count    = all_props.filter(monthly_rent__gte=100000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=25000).count()
    short_lease_count           = all_props.filter(lockin_period=6).count()
    long_lease_count            = all_props.filter(lockin_period=12).count()
    with_owner_count            = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    city_count                  = (all_props.exclude(city__isnull=True).exclude(city='')
                                   .values('city').distinct().count())
    try:
        with_images_count = all_props.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    # ── Percentages ──
    verified_pct = round((with_owner_count     / total_count * 100)) if total_count > 0 else 0
    image_pct    = round((with_images_count    / total_count * 100)) if total_count > 0 else 0
    premium_pct  = round((premium_properties_count / total_count * 100)) if total_count > 0 else 0

    # ── NEW: Listing / Approval / Duplicate status KPIs ──
    active_listing_count   = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count     = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count   = all_props.filter(listing_status__iexact='Rented').count()
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count         = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count         = all_props.filter(approval_status__iexact='Rejected').count()
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count    = total_count - duplicate_properties_count

    # ── Uploaded file names (for bulk delete) ──
    try:
        uploaded_files = (all_props.exclude(upload_file_name__isnull=True)
                          .exclude(upload_file_name='')
                          .values_list('upload_file_name', flat=True).distinct())
    except Exception:
        uploaded_files = []

    # ════════════════════════════════════════════
    # CHART DATA  (JSON → template → JS)
    # ════════════════════════════════════════════

    prop_type_dist = list(
        all_props.exclude(property_type__isnull=True).exclude(property_type='')
        .values('property_type').annotate(cnt=Count('id')).order_by('-cnt')[:8]
    )
    prop_type_labels_json = json.dumps([x['property_type'] for x in prop_type_dist])
    prop_type_counts_json = json.dumps([x['cnt']           for x in prop_type_dist])

    rent_range_data = {
        'Under ₹25k':  all_props.filter(monthly_rent__lt=25000).count(),
        '₹25k–1L':     all_props.filter(monthly_rent__gte=25000,  monthly_rent__lt=100000).count(),
        '₹1L–5L':      all_props.filter(monthly_rent__gte=100000, monthly_rent__lt=500000).count(),
        'Above ₹5L':   all_props.filter(monthly_rent__gte=500000).count(),
    }
    rent_range_labels_json = json.dumps(list(rent_range_data.keys()))
    rent_range_counts_json = json.dumps(list(rent_range_data.values()))

    occupancy_json = json.dumps([occupied_count, vacant_count, max(0, total_count - occupied_count - vacant_count)])

    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_qs = (
        all_props.filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(revenue=Sum('monthly_rent'), cnt=Count('id'))
        .order_by('month')
    )
    monthly_labels_json  = json.dumps([x['month'].strftime('%b %Y') for x in monthly_qs])
    monthly_revenue_json = json.dumps([float(x['revenue'] or 0) for x in monthly_qs])

    total_tenants    = occupied_count
    collection_rate  = 0
    pending_payments = 0
    maintenance_req  = 0

    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        
        'page_obj': page_obj,

        'search_query':          search_query,
        'prop_type_query':       prop_type_query,
        'city_query':            city_query,
        'zone_query':            zone_query,
        'possession_query':      possession_query,
        'listed_by_query':       listed_by_query,
        'uploaded_by_query':     uploaded_by_query,
        'budget_query':          budget_query,
        'listing_status_query':  listing_status_query,
        'approval_status_query': approval_status_query,
        'duplicate_query':       duplicate_query,
        'from_date':             from_date,
        'to_date':               to_date,

        'unique_property_types':  unique_property_types,
        'unique_cities':          unique_cities,
        'unique_zones':           unique_zones,
        'unique_possession':      unique_possession,
        'unique_roles':           unique_roles,
        'unique_uploaded_roles':  unique_uploaded_roles,
        'unique_listing_status':  unique_listing_status,
        'unique_approval_status': unique_approval_status,
        'uploaded_files':         uploaded_files,

        'filtered_count':            filtered_count,
        'total_count':               total_count,
        'active_count':              active_count,
        'occupied_count':            occupied_count,
        'vacant_count':              vacant_count,
        'occupancy_rate':            occupancy_rate,
        'vacancy_rate':              vacancy_rate,
        'avg_rent':                  avg_rent,
        'max_rent':                  max_rent,
        'min_rent':                  min_rent,
        'total_revenue':             total_revenue,
        'total_security_deposit':    total_security_deposit,
        'avg_deposit':               avg_deposit,
        'avg_area':                  avg_area,
        'premium_properties_count':  premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'ready_to_move_count':       vacant_count,
        'short_lease_count':         short_lease_count,
        'long_lease_count':          long_lease_count,
        'with_owner_count':          with_owner_count,
        'with_images_count':         with_images_count,
        'city_count':                city_count,
        'verified_pct':              verified_pct,
        'image_pct':                 image_pct,
        'premium_pct':               premium_pct,

        # NEW status KPIs
        'active_listing_count':      active_listing_count,
        'inactive_listing_count':    inactive_listing_count,
        'sold_listing_count':        sold_listing_count,
        'rented_listing_count':      rented_listing_count,
        'pending_approval_count':    pending_approval_count,
        'approved_count':            approved_count,
        'rejected_count':            rejected_count,
        'duplicate_properties_count': duplicate_properties_count,
        'unique_properties_count':    unique_properties_count,

        'total_tenants':    total_tenants,
        'collection_rate':  collection_rate,
        'pending_payments': pending_payments,
        'maintenance_req':  maintenance_req,

        'prop_type_labels_json':  prop_type_labels_json,
        'prop_type_counts_json':  prop_type_counts_json,
        'rent_range_labels_json': rent_range_labels_json,
        'rent_range_counts_json': rent_range_counts_json,
        'occupancy_json':         occupancy_json,
        'monthly_labels_json':    monthly_labels_json,
        'monthly_revenue_json':   monthly_revenue_json,
    }
    return render(request, 'agent/Reports/Rental/commercial_list.html', context)






def to_decimal_or_none(value):
    if value in (None, "", "None"):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None




def export_commercial_rent_agent(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

  


    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    
    # ── 1. Re-apply the same search filters so the export matches the screen ──
    try:
        properties = CommercialRentalProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')
    except Exception:
        properties = CommercialRentalProperty.objects.filter(listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    search_query     = request.GET.get('search', '').strip()
    prop_type_query  = request.GET.get('property_type', '').strip()
    city_query       = request.GET.get('city', '').strip()
    zone_query       = request.GET.get('zone_type', '').strip()
    possession_query = request.GET.get('possession', '').strip()
    listed_by_query  = request.GET.get('listed_by', '').strip()
    budget_query     = request.GET.get('budget', '').strip()
    from_date        = request.GET.get('from_date', '').strip()
    to_date          = request.GET.get('to_date', '').strip()

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query)
        )
    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__icontains=prop_type_query)
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if zone_query and zone_query != 'All Zones':
        properties = properties.filter(zone_type__icontains=zone_query)
    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__icontains=possession_query)
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__icontains=listed_by_query)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)
    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_25k': properties = properties.filter(monthly_rent__lt=25000)
        elif budget_query == '25k_1L': properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=100000)
        elif budget_query == '1L_5L': properties = properties.filter(monthly_rent__gte=100000, monthly_rent__lte=500000)
        elif budget_query == 'above_5L': properties = properties.filter(monthly_rent__gt=500000)

    # ── 2. EXHAUSTIVE FIELD MAPPING (Every DB Field Included) ──
    # Format: (Section Name, Database Field Name, Is Required, Exact Header Name)
    EXPORT_COLS = [
        ("Sr.No", "sr_no", False, "Sr. No"),
        

        ("Listed By Details", "listed_by_type", False, "Listed By Type"),
        
        ("Listed By Details", "listed_by_id", False, "Listed By Id"),
        ("Listed By Details", "listed_by_name", False, "Listed By Name"),
        ("Listed By Details", "listed_by_email", False, "Listed By Email"),
        ("Listed By Details", "listed_by_contact", False, "Listed By Contact"),
        ("Listed By Details", "listed_by_role", True, "Listed By Role"),

        ("Basic Information", "id", False, "Property ID"),
        ("Basic Information", "property_title", False, "Property Title"),
        

        ("Basic Information", "property_type", True, "Property Type"),
        ("Basic Information", "building_name", True, "Building / Project Name"),
        ("Basic Information", "wing_number", False, "Tower/Wing Number"),
        ("Basic Information", "property_no", True, "Shop No/Godown/Unit No"),
        ("Basic Information", "availability_status", True, "Availability Status"),
        ("Basic Information", "available_from", False, "Available From"),
        ("Basic Information", "property_age", True, "Age of Property(In Years)"),
        ("Basic Information", "zone_type", False, "Zone Type"),
        ("Basic Information", "ownership_type", True, "Ownership Type"),
        ("Basic Information", "property_condition", False, "Property Condition"),

        ("Basic Information", "builtup_area", True, "Built-up Area (sq.ft)"),
        ("Basic Information", "carpet_area", False, "Carpet Area (sq.ft)"),
       

        ("Property Location Details", "address", True, "Property Address"),
        ("Property Location Details", "locality", True, "Area/Locality"),
        ("Property Location Details", "property_landmark", False, "Property Landmark"),
        ("Property Location Details", "city", True, "City"),
        ("Property Location Details", "state", True, "State"),
        ("Property Location Details", "location_hub", False, "Location Hub"),
        ("Property Location Details", "pincode", True, "Pincode"),
        ("Property Location Details", "google_maps_link", False, "Google Maps Link"),
        ("Property Location Details", "latitude", False, "Latitude"),
        ("Property Location Details", "longitude", False, "Longitude"),

        ("Property Specifications", "dg_ups_included", False, "DG & UPS Included?"),
        ("Property Specifications", "electricity_included", False, "Electricity Included?"),
        ("Property Specifications", "water_included", False, "Water Included?"),
        ("Property Specifications", "building_configuration", False, "Building Configuration"),
        ("Property Specifications", "total_floors", True, "Total Floors Constructed"),
        ("Property Specifications", "staircases", False, "Staircases"),
        ("Property Specifications", "passenger_lifts", False, "Passenger Lifts"),
        ("Property Specifications", "service_lifts", False, "Service Lifts"),
        ("Property Specifications", "private_parking", False, "Private Parking"),
        ("Property Specifications", "min_seats", False, "Minimum Seats Occupancy"),
        ("Property Specifications", "max_seats", False, "Maximum Seats Occupancy"),
        ("Property Specifications", "cabins", False, "Cabins"),
        ("Property Specifications", "meeting_rooms", False, "Meeting Rooms"),
        ("Property Specifications", "private_washroom", False, "Private Washrooms"),
        ("Property Specifications", "public_washroom", False, "Public Washrooms"),
        ("Property Specifications", "flooring_type", False, "Flooring Type"),

        ("Pricing Details", "monthly_rent", True, "Monthly Rent"),
        ("Pricing Details", "brokerage_percentage", True, "Brokerage"),
        ("Pricing Details", "manual_brokerage", False, "Enter Fixed Brokerage"),
        ("Pricing Details", "advanced_rent_type", True, "Advanced Rent Month"),
        ("Pricing Details", "advanced_rent_amount", False, "Advance Rent Amount"),
        ("Pricing Details", "security_deposit_type", True, "Refundable Security Deposit"),
        ("Pricing Details", "security_deposit_amount", False, "Refundable Security Deposit Amount"),
        ("Pricing Details", "maintenance_type", False, "Maintenance Type"),
        ("Pricing Details", "maintenance_charges", False, "Monthly Maintenance Amount"),
        ("Pricing Details", "total_move_in_cost", False, "Total Move In Cost"),
        ("Pricing Details", "negotiable", False, "Negotiable"),
        ("Pricing Details", "lockin_period", False, "Lock-in Period (months)"),
        ("Pricing Details", "rent_increase", False, "Rent Increase (%/year)"),

        ("Amenities & Facilities", "amenities", False, "Amenities"),
        ("Amenities & Facilities", "nearby_facilities", False, "Nearby Facilities"),
        
        
        ("Property Descriptions", "property_summary", False, "System Property Summary(Auto)"),
        ("Property Descriptions", "property_description", False, "System Property Description(Auto)"),
        ("Property Descriptions", "user_description", False, "Property Description(By User)"),

        ("Media & Listing Status", "listed_elsewhere", False, "Is Property Already Listed Elsewhere?"),
        ("Media & Listing Status", "portal_name", False, "Portal Name"),
        ("Media & Listing Status", "listing_status", False, "Listing Status"),
        ("Media & Listing Status", "approval_status", False, "Approval Status"),

        ("Data Uploadeded Via", "upload_file_name", False, "Upload File Name"),
        ("Property Uploaded By", "uploaded_by_name", False, "Uploaded By Name"),
        ("Property Uploaded By", "uploaded_by_email", False, "Uploaded By Email"),
        ("Property Uploaded By", "uploaded_by_contact", False, "Uploaded By Contact"),
        ("Property Uploaded By", "uploaded_by_role", False, "Uploaded By Role"),

        ("Database Audit", "created_at", False, "Created At"),
        
    ]

    export_format = request.GET.get('format', 'excel')
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Commercial_Rental_Full_Export_{timestamp}"

    # Helper for formatting values
    def format_val(val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(val, bool):
            return "Yes" if val else "No"
        elif val is None:
            return ""
        return str(val).strip()

    # ── 3. EXCEL EXPORT ──
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commercial Rental Database"

        HDR_BG  = "667EEA"
        REQ_BG  = "FEF3C7"
        OPT_BG  = "F0FDF4"
        thin = Side(style="thin", color="CBD5E1")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sec_spans = OrderedDict()
        for i, (sec, *_) in enumerate(EXPORT_COLS):
            sec_spans.setdefault(sec, []).append(i + 1)

        # Write Section Headers (Row 1)
        for sec, cols in sec_spans.items():
            c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
            
            for col_idx in cols[1:]:
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor=HDR_BG)
                ws.cell(row=1, column=col_idx).border = bdr
                
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

        # Write Field Headers (Row 2) - Exact string matches for importer!
        for ci, (_, _, req, header_name) in enumerate(EXPORT_COLS, 1):
            header_text = header_name + (" *" if req else "")
            lc = ws.cell(row=2, column=ci, value=header_text)
            lc.font = Font(bold=True, size=9)
            lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
            lc.border = bdr
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(header_text) + 5)

        # Write Data (Rows 3+)
        for row_idx, prop in enumerate(properties, start=3):
            for col_idx, (_, field, _, _) in enumerate(EXPORT_COLS, 1):
                
                # Custom logic for Serial Number and Upload File Name
                if field == "sr_no":
                    val = row_idx - 2
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center")

        # Layout adjustments
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # ── 4. CSV EXPORT ──
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        
        # Write exact Field Headers
        headers = [header_name + (" *" if req else "") for (_, _, req, header_name) in EXPORT_COLS]
        writer.writerow(headers)

        # Write Data Rows
        for row_idx, prop in enumerate(properties, start=1):
            row_data = []
            for (_, field, _, _) in EXPORT_COLS:
                if field == "sr_no":
                    val = row_idx
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                row_data.append(val)
            writer.writerow(row_data)

        return response




@require_POST
def commercial_delete_agent(request, pk):
    """Agent Soft Delete Commercial Property (supports Admin impersonation)"""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Agent not found.'
            })

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Fetch Property
        # ======================================
        prop = get_object_or_404(CommercialRentalProperty, pk=pk)

        # ======================================
        # 6. Security Check — agent can delete only own property
        # ======================================
        if str(prop.listed_by_id) != str(agent_obj.user_id):
            return JsonResponse({
                'status': 'error',
                'message': 'You can delete only your own properties.'
            })

        # ======================================
        # 7. Soft Delete
        # ======================================
        prop.is_deleted = True
        prop.deleted_at = timezone.now()
        prop.deleted_by = user_identity
        prop.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Moved to Recycle Bin successfully!'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })



@require_POST
def commercial_bulk_delete_agent(request):
    """Agent Advanced Bulk Delete (Soft Delete) — scoped to agent's own commercial properties, supports Admin impersonation."""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Parse Payload
        # ======================================
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Invalid request payload.'})

        delete_type = data.get('delete_type')

        # ======================================
        # 6. Base Queryset — SCOPED TO THIS AGENT ONLY
        # ======================================
        properties = CommercialRentalProperty.objects.filter(
            is_deleted=False,
            listed_by_id=agent_obj.user_id
        )

        target_props = None
        criteria_label = ""

        if delete_type == 'delete_all':
            target_props = properties
            criteria_label = "All own properties"

        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            if not page_ids:
                return JsonResponse({'status': 'error', 'message': 'No properties selected on current page.'})
            target_props = properties.filter(id__in=page_ids)
            criteria_label = "Current page selection"

        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            if not from_date or not to_date:
                return JsonResponse({'status': 'error', 'message': 'Both From Date and To Date are required.'})
            target_props = properties.filter(created_at__date__range=[from_date, to_date])
            criteria_label = f"Date range {from_date} to {to_date}"

        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)
            criteria_label = "Latest month"

        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)
            criteria_label = "Older than 6 months"

        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '').strip()
            if not uploader:
                return JsonResponse({'status': 'error', 'message': 'Uploader detail is required.'})
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            criteria_label = f"Uploaded by '{uploader}'"

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '').strip()
            if not file_name:
                return JsonResponse({'status': 'error', 'message': 'Please select a file.'})
            target_props = properties.filter(upload_file_name=file_name)
            criteria_label = f"File '{file_name}'"

        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

        # ======================================
        # 7. Execute Soft Delete
        # ======================================
        count = target_props.count()

        if count == 0:
            return JsonResponse({
                'status': 'error',
                'message': f'No matching properties found for: {criteria_label}'
            })

        target_props.update(
            is_deleted=True,
            deleted_at=timezone.now(),
            deleted_by=user_identity
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully moved {count} of your properties ({criteria_label}) to Recycle Bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



#################Views End For Rental Commericial  Property For Agent ###########################




############# Views start for Rental pg Module for Agent ###################################################




def pg_coliving_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Rental/pg_coliving.html", context)



def pg_agent_list(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    

    # ── READ FILTER STRINGS ──
    search_query          = request.GET.get('search', '').strip()
    pg_for_filter          = request.GET.get('pg_for', '').strip()
    city_filter            = request.GET.get('city', '').strip()
    furnish_filter         = request.GET.get('furnish', '').strip()
    meals_filter           = request.GET.get('meals', '').strip()
    sharing_filter         = request.GET.get('sharing', '').strip()
    listed_by_filter       = request.GET.get('listed_by', '').strip()
    uploaded_by_filter     = request.GET.get('uploaded_by', '').strip()
    listing_status_filter  = request.GET.get('listing_status', '').strip()
    approval_status_filter = request.GET.get('approval_status', '').strip()
    duplicate_filter       = request.GET.get('duplicate', '').strip()
    from_date              = request.GET.get('from_date', '').strip()
    to_date                = request.GET.get('to_date', '').strip()

    # ── FILTER SOFT DELETIONS ──
    all_props  = PGColivingProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
    properties = all_props.order_by('-created_at')

      


    # NOTE: search now checks every meaningful text/identifier field,
    # not just title/city/locality/owner name/contact/id.
    if search_query:
        properties = properties.filter(
            Q(id__icontains=search_query) |
            Q(property_unique_key__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(property_no__icontains=search_query) |
            Q(wing_number__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(pg_for__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(room_type__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_contact__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(listing_status__icontains=search_query) |
            Q(approval_status__icontains=search_query)
        )
    if pg_for_filter:
        properties = properties.filter(pg_for__iexact=pg_for_filter)
    if city_filter:
        properties = properties.filter(city__icontains=city_filter)
    if furnish_filter:
        properties = properties.filter(furnishing_status__iexact=furnish_filter)
    if meals_filter == 'Yes':
        properties = properties.filter(meals_available=True)
    elif meals_filter == 'No':
        properties = properties.filter(meals_available=False)
    if sharing_filter:
        properties = properties.filter(room_type__iexact=sharing_filter)

    if listed_by_filter and listed_by_filter != 'All Roles':
        properties = properties.filter(listed_by_role__icontains=listed_by_filter)
    if uploaded_by_filter and uploaded_by_filter != 'All Roles':
        properties = properties.filter(uploaded_by_role__icontains=uploaded_by_filter)
    if listing_status_filter and listing_status_filter != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_filter)
    if approval_status_filter and approval_status_filter != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_filter)

    if duplicate_filter == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_filter == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ── NEW: prefetch videos so uploaded_video / auto_video / rm_video
    # actually exist on each row instead of being undefined in the template.
    # PGColivingVideo.property has related_name="videos", so we prefetch
    # 'videos' (not 'video') and stash the whole list on each row as
    # `all_videos`, then split by `source` below. ──
    properties = properties.prefetch_related(
        Prefetch(
            'videos',
            queryset=PGColivingVideo.objects.all(),
            to_attr='all_videos'
        )
    )

    # ── PAGINATION SYSTEM ──
    paginator = Paginator(properties, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    # ── NEW: split each property's prefetched videos into the 3 source
    # buckets the template expects (mirrors commercial_list's pattern).
    # This MUST run on page_obj (after pagination), not on the full
    # `properties` queryset, so the prefetch is only evaluated for the
    # 10 rows actually shown on this page. ──
    for prop in page_obj:
        videos_by_source = {v.source: v for v in prop.all_videos}
        prop.uploaded_video = videos_by_source.get('uploaded')
        prop.auto_video     = videos_by_source.get('auto')
        prop.rm_video       = videos_by_source.get('rm_assisted')

    # ── SUMMARY CARD KPI AGGREGATIONS ──
    total_count    = properties.count()
    boys_count     = properties.filter(pg_for__iexact='Boys').count()
    girls_count    = properties.filter(pg_for__iexact='Girls').count()
    coliving_count = properties.filter(pg_for__iexact='Co-living').count()
    total_beds     = properties.aggregate(t=Sum('total_beds'))['t'] or 0
    city_count     = properties.values('city').distinct().count()

    boys_pct     = round((boys_count     / total_count * 100), 1) if total_count else 0
    girls_pct    = round((girls_count    / total_count * 100), 1) if total_count else 0
    coliving_pct = round((coliving_count / total_count * 100), 1) if total_count else 0

    rent_stats = properties.aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
        avg_dep=Avg('security_deposit_amount'),
        tot_rev=Sum('monthly_rent'),
        tot_dep=Sum('security_deposit_amount')
    )

    meals_available_count = properties.filter(meals_available=True).count()
    meals_pct     = round((meals_available_count / total_count * 100), 1) if total_count else 0
    furnished_count = properties.filter(furnishing_status__icontains='Fully').count()
    furnished_pct   = round((furnished_count / total_count * 100), 1) if total_count else 0
    single_room_count = properties.filter(room_type__iexact='single').count()
    shared_room_count = properties.filter(room_type__in=['double', 'triple', 'quad']).count()
    anytime_entry   = properties.filter(entry_24x7_allowed=True).count()
    visitors_allowed= properties.filter(visitors_allowed=True).count()
    premium_pg_count= properties.filter(monthly_rent__gte=10000).count()
    budget_pg_count = properties.filter(monthly_rent__lt=5000).count()
    with_owner_count= properties.exclude(Q(listed_by_name__isnull=True) | Q(listed_by_name='')).count()

    try:
        with_images_count = properties.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    # ── Listing / Approval / Duplicate status KPIs (unfiltered dataset) ──
    active_listing_count   = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count     = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count   = all_props.filter(listing_status__iexact='Rented').count()
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count         = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count         = all_props.filter(approval_status__iexact='Rejected').count()
    all_props_total        = all_props.count()
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count    = all_props_total - duplicate_properties_count

    uploaded_files = PGColivingProperty.objects.filter(
        is_deleted=False, upload_file_name__isnull=False
    ).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct().order_by('upload_file_name')

    # ── Dropdown unique values (roles / statuses) ──
    unique_roles = (all_props
        .exclude(listed_by_role__isnull=True).exclude(listed_by_role='')
        .values_list('listed_by_role', flat=True).distinct())
    unique_uploaded_roles = (all_props
        .exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='')
        .values_list('uploaded_by_role', flat=True).distinct())
    unique_listing_status = (all_props
        .exclude(listing_status__isnull=True).exclude(listing_status='')
        .values_list('listing_status', flat=True).distinct())
    unique_approval_status = (all_props
        .exclude(approval_status__isnull=True).exclude(approval_status='')
        .values_list('approval_status', flat=True).distinct())

    # ── CHARTS ──
    pg_for_qs     = properties.values('pg_for').annotate(c=Count('id')).order_by('-c')
    pg_for_labels = json.dumps([i['pg_for'] or 'Unspecified' for i in pg_for_qs])
    pg_for_data   = json.dumps([i['c'] for i in pg_for_qs])

    rent_buckets      = [('Under ₹3k', 0, 3000), ('₹3k–5k', 3000, 5000), ('₹5k–8k', 5000, 8000), ('₹8k–12k', 8000, 12000), ('Above ₹12k', 12000, 99999999)]
    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data   = json.dumps([properties.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs        = properties.values('furnishing_status').annotate(c=Count('id')).order_by('-c')
    furnishing_labels = json.dumps([i['furnishing_status'] or 'Unspecified' for i in furnish_qs])
    furnishing_data   = json.dumps([i['c'] for i in furnish_qs])

    city_qs     = properties.values('city').annotate(c=Count('id')).order_by('-c')[:5]
    city_labels = json.dumps([i['city'] or 'Unspecified' for i in city_qs])
    city_data   = json.dumps([i['c'] for i in city_qs])

    cities = all_props.values_list('city', flat=True).distinct().order_by('city')

    return render(request, 'agent/Reports/Rental/pg_list.html', {
        
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        
        'page_obj': page_obj,
        'search_query': search_query,
        'pg_for_filter': pg_for_filter,
        'city_filter': city_filter,
        'furnish_filter': furnish_filter,
        'meals_filter': meals_filter,
        'sharing_filter': sharing_filter,
        'listed_by_filter': listed_by_filter,
        'uploaded_by_filter': uploaded_by_filter,
        'listing_status_filter': listing_status_filter,
        'approval_status_filter': approval_status_filter,
        'duplicate_filter': duplicate_filter,
        'from_date': from_date,
        'to_date': to_date,
        'cities': cities,
        'unique_roles': unique_roles,
        'unique_uploaded_roles': unique_uploaded_roles,
        'unique_listing_status': unique_listing_status,
        'unique_approval_status': unique_approval_status,

        'total_count': total_count,
        'city_count': city_count,
        'boys_count': boys_count,
        'girls_count': girls_count,
        'coliving_count': coliving_count,
        'total_beds': total_beds,
        'boys_pct': boys_pct,
        'girls_pct': girls_pct,
        'coliving_pct': coliving_pct,
        'avg_rent': rent_stats['avg_rent'] or 0,
        'max_rent': rent_stats['max_rent'] or 0,
        'min_rent': rent_stats['min_rent'] or 0,
        'total_revenue': rent_stats['tot_rev'] or 0,
        'total_deposit': rent_stats['tot_dep'] or 0,
        'avg_deposit': rent_stats['avg_dep'] or 0,
        'meals_available_count': meals_available_count,
        'meals_pct': meals_pct,
        'furnished_count': furnished_count,
        'furnished_pct': furnished_pct,
        'single_room_count': single_room_count,
        'shared_room_count': shared_room_count,
        'non_veg_allowed': meals_available_count,
        'with_images_count': with_images_count,
        'anytime_entry': anytime_entry,
        'visitors_allowed': visitors_allowed,
        'premium_pg_count': premium_pg_count,
        'budget_pg_count': budget_pg_count,
        'with_owner_count': with_owner_count,
        'uploaded_files': uploaded_files,

        # status KPIs
        'active_listing_count': active_listing_count,
        'inactive_listing_count': inactive_listing_count,
        'sold_listing_count': sold_listing_count,
        'rented_listing_count': rented_listing_count,
        'pending_approval_count': pending_approval_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'duplicate_properties_count': duplicate_properties_count,
        'unique_properties_count': unique_properties_count,

        'pg_for_labels': pg_for_labels,
        'pg_for_data': pg_for_data,
        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,
        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,
        'city_labels': city_labels,
        'city_data': city_data,
    })



@csrf_exempt
def add_pg_agent(request):

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    try:
        
        
        user_obj = User_Details.objects.filter(id=user_id).first()
        uploader_name = getattr(user_obj, 'user_name', '') if user_obj else ''
        uploader_email = getattr(user_obj, 'user_email', '') if user_obj else ''
        uploader_contact = getattr(user_obj, 'user_phone', '') if user_obj else ''
        uploader_role = getattr(user_obj, 'user_role', 'User') if user_obj else 'Agent'
        uploader_id = f"USER_{user_id}"

        def get_list(name):
            return ",".join(request.POST.getlist(name))

        # ---------- CATEGORY IMAGES (validate count first, same as before) ----------
        IMAGE_CATEGORY_FIELDS = {
            'exterior': 'property_images_exterior[]',
            'single_room': 'property_images_single_room[]',
            'double_room': 'property_images_double_room[]',
            'multi_room': 'property_images_multi_room[]',
            'lounge': 'property_images_lounge[]',
            'kitchen': 'property_images_kitchen[]',
            'dining': 'property_images_dining[]',
            'washroom': 'property_images_washroom[]',
            'laundry': 'property_images_laundry[]',
            'parking': 'property_images_parking[]',
            'amenities': 'property_images_amenities[]',
            'floor_plan': 'property_images_floor_plan[]',
        }

        total_images = sum(len(request.FILES.getlist(f)) for f in IMAGE_CATEGORY_FIELDS.values())
        

        if total_images > 30:
           return JsonResponse({"status": "error", "message": "You can upload maximum 30 images"})

        # ---------- LISTED BY ----------
        input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
        input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
        input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
        input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
        input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

        # ---------- CORE LOCATION / IDENTITY FIELDS (used in fingerprint) ----------
        input_property_no = (request.POST.get('property_no') or '').strip()
        input_building_name = (request.POST.get('building_name') or '').strip()
        input_locality = (request.POST.get('locality') or '').strip()
        input_city = (request.POST.get('city') or '').strip()
        input_address = (request.POST.get('address') or '').strip()
        input_pincode = (request.POST.get('pincode') or '').strip()
        input_monthly_rent = _to_int_or_none(request.POST.get('room_rent'))

        # ---------- DUPLICATE DETECTION (mirrors model.save()'s key + commercial's approach) ----------
        fingerprint_key = build_pg_fingerprint(
            input_address, input_locality, input_city, input_property_no, input_pincode
        )

        direct_duplicates = PGColivingProperty.objects.filter(
            is_deleted=False,
            locality__iexact=input_locality,
        )
        if input_property_no:
            direct_duplicates = direct_duplicates.filter(property_no__iexact=input_property_no)
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            PGColivingProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    return JsonResponse({
                        "status": "error",
                        "message": (f"Duplicate Blocked: This PG (Unit {input_property_no}) is already "
                                    f"listed by/for {input_listed_by_name or 'this user'}. "
                                    f"Please edit the existing listing instead.")
                    })

            # Reuse an existing group id if the old rows already have one,
            # instead of overwriting it with this new record's fingerprint.
            existing_with_group = existing_duplicates.exclude(
                duplicate_group_id__isnull=True
            ).exclude(duplicate_group_id="").first()
            dup_group_id = existing_with_group.duplicate_group_id if existing_with_group else fingerprint_key

            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id,
            )

        # ---------- DATE PARSING ----------
        date_val = request.POST.get('available_from')
        available_from = None
        if date_val:
            try:
                available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
            except ValueError:
                available_from = None

        # ---------- MEALS / PILL-GROUP FIELDS (already comma-joined by the form's JS) ----------
        meals_available = _to_bool(request.POST.get('meals_available'))
        meal_offerings = (request.POST.get('meal_offerings') or '').strip()
        meal_speciality = (request.POST.get('meal_speciality') or '').strip()
        best_suited_for = (request.POST.get('best_suited_for') or '').strip()
        pg_for = (request.POST.get('pg_for') or '').strip()

        with transaction.atomic():

            pg = PGColivingProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="PG/Co-living",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                building_name=input_building_name,
                property_no=input_property_no,
                wing_number=request.POST.get('wing_number'),

                city=input_city,
                locality=input_locality,
                address=input_address,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=input_pincode,
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                pg_for=pg_for,
                furnishing_status=request.POST.get('furnishing_status'),
                best_suited_for=best_suited_for,

                meals_available=meals_available,
                meal_offerings=meal_offerings,
                meal_speciality=meal_speciality,

                notice_period=_to_int_or_none(request.POST.get('notice_period')),
                lockin_period=_to_int_or_none(request.POST.get('lockin_period')),
                minimum_stay=_to_int_or_none(request.POST.get('minimum_stay')) or 1,
                available_from=available_from,
                property_managed_by=request.POST.get('property_managed_by'),
                manager_stays=_to_bool(request.POST.get('manager_stays')),

                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                # ---- Step 2: room details, flattened directly onto the property ----
                room_type=request.POST.get('room_type'),
                total_beds=_to_int_or_none(request.POST.get('room_beds')),
                monthly_rent=input_monthly_rent,
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=_to_int_or_none(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=_to_int_or_none(request.POST.get('security_deposit_amount')),
                maintenance_type=request.POST.get('maintenance_type'),
                maintenance_amount=_to_int_or_none(request.POST.get('maintenance_amount')),
                # total_move_in_cost is recalculated in model.save() via calculate_move_in_cost()

                # ---- Step 3: regulations ----
                opposite_gender_visitors_allowed="opposite_gender_visitors_allowed" in request.POST,
                visitors_allowed="visitors_allowed" in request.POST,
                parents_guardians_allowed="parents_guardians_allowed" in request.POST,
                entry_24x7_allowed="entry_24x7_allowed" in request.POST,
                curfew_time=request.POST.get('curfew_time') or None,
                smoking_allowed="smoking_allowed" in request.POST,
                alcohol_consumption_allowed="alcohol_consumption_allowed" in request.POST,
                couples_allowed="couples_allowed" in request.POST,
                pets_allowed="pets_allowed" in request.POST,
                cooking_allowed="cooking_allowed" in request.POST,
                police_verification_required="police_verification_required" in request.POST,

                amenities=get_list("amenities[]"),
                nearby_facilities=get_list("nearby_facilities[]"),
                user_description=request.POST.get("user_description"),

                listed_elsewhere=request.POST.get("listed_elsewhere", "No"),
                portal_name=request.POST.get("portal_name"),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,

                listing_status="Pending",
                approval_status="Pending",
            )
            # NOTE: property_title, property_summary, property_description and the
            # 7 auto-FAQs are all generated inside PGColivingProperty.save() itself,
            # so nothing needs to be done for those here.

            # ---------- CATEGORY IMAGES ----------
            saved_count = 0
            global_sequence = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 30:
                        break
                    PGPropertyImage.objects.create(
                        property=pg, image=img, category=category, sequence_order=global_sequence
                    )
                    global_sequence += 1
                    saved_count += 1

            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '')

            # 1. Auto-generate the slideshow
            saved_images = list(PGPropertyImage.objects.filter(property=pg))
            image_paths = []
            for img_obj in saved_images:
                if img_obj.image and hasattr(img_obj.image, 'path') and os.path.exists(img_obj.image.path):
                    image_paths.append(img_obj.image.path)

            if len(image_paths) >= 3:
                try:
                    from Admin_App.utils import generate_property_slideshow
                    import time

                    # Define the output path securely with a timestamp
                    out_path = f"pg_coliving/videos/auto_{pg.id}_{int(time.time())}.mp4"

                    result_path = generate_property_slideshow(image_paths, out_path)
                    print("AUTO SLIDESHOW RESULT:", result_path)

                    if result_path:
                        PGColivingVideo.objects.update_or_create(
                            property=pg,
                            source='auto',
                            defaults={
                                'video': result_path,
                                'video_url': None
                            }
                        )
                except Exception as ve:
                    import traceback
                    print("PG VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()
            else:
                # Fallback if there are less than 3 images
                PGColivingVideo.objects.update_or_create(
                    property=pg,
                    source='auto',
                    defaults={
                        'video': None,
                        'video_url': None
                    }
                )
            # 2. Manual Upload Logic (separate row, source='uploaded')
            if video_option == 'upload' and uploaded_video:
                if uploaded_video.size > 20 * 1024 * 1024:
                    PGColivingVideo.objects.create(
                        property=pg,
                        video=None,
                        source='uploaded'
                    )
                else:
                    PGColivingVideo.objects.create(
                        property=pg,
                        video=uploaded_video,
                        source='uploaded'
                    )

            # 3. RM Assisted Link Logic (separate row, source='rm_assisted')
            elif video_option == 'rm_assisted':
                PGColivingVideo.objects.create(
                    property=pg,
                    video_url=property_video_link if property_video_link else None,
                    source='rm_assisted'
                )

            # ---------- ACTIVITY LOG ----------
            PGColivingActivityLog.objects.create(
                user_identity=uploader_name,
                user_role=uploader_role,
                action_type="CREATE",
                property_id=str(pg.id),
                action_payload=f"PG '{pg.property_title}' created by {uploader_name}",
                status="SUCCESS",
            )

        return JsonResponse({"status": "success", "message": "PG Added Successfully"})

    except Exception as e:
        try:
            print("ERROR IN PG ADD:", str(e).encode("ascii", "replace").decode())
            traceback.print_exc()
        except Exception:
            pass  # never let logging itself crash the response
        return JsonResponse(
            {"status": "error", "message": f"Server Error: {str(e)}"},
            status=500
        )




def _to_int(val, default=0):
    try:
        return int(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _to_int_or_none(val):
    try:
        return int(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _to_bool(val):
    return str(val).strip().lower() in ("true", "1", "yes", "on")

def build_pg_fingerprint(address, locality, city, property_no, pincode):
    """
    Mirrors PGColivingProperty.save()'s own key_source construction exactly,
    so the pre-create duplicate lookup matches what the model would generate.
    """
    key_source = f"{address}|{locality}|{city}|{property_no}|{pincode}"
    return key_source.strip().lower().replace(" ", "")







def pg_agent_edit(request, pk):
    property_id = pk
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    pg = get_object_or_404(PGColivingProperty, id=property_id)

    if request.method == "GET":
        # Context data required to render the edit form template safely
        amenities_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        

        # Group existing saved images by category so each tab/panel can show its own photos
        images_by_cat = defaultdict(list)
        for img in PGPropertyImage.objects.filter(property=pg).order_by('sequence_order', 'id'):
            images_by_cat[img.category].append(img)

        IMAGE_CATEGORY_DEFS_RAW = [
            ('exterior', 'Building Exterior', 'Front elevation, entrance, signage — up to 5 images'),
            ('single_room', 'Single Occupancy', 'Single sharing rooms — up to 4 images'),
            ('double_room', 'Double Sharing', 'Double sharing rooms — up to 5 images'),
            ('multi_room', 'Triple / Multi Sharing', 'Triple/quad sharing rooms — up to 5 images'),
            ('lounge', 'Common Lounge', 'Common area, TV lounge — up to 4 images'),
            ('kitchen', 'Shared Kitchen', 'Kitchen, cooking area — up to 3 images'),
            ('dining', 'Dining Area', 'Dining tables, seating — up to 3 images'),
            ('washroom', 'Bathroom / Washroom', 'Shared/attached washrooms — up to 4 images'),
            ('laundry', 'Laundry Area', 'Washing machine, drying area — up to 3 images'),
            ('parking', 'Parking', 'Two/four wheeler parking — up to 3 images'),
            ('amenities', 'Common Amenities', 'Gym, gaming zone, security — up to 5 images'),
            ('floor_plan', 'Floor Plan', 'Room/floor layout diagrams — up to 3 images'),
        ]

        image_category_defs = [
            {'cat': cat, 'label': label, 'subtext': subtext, 'existing': images_by_cat.get(cat, [])}
            for cat, label, subtext in IMAGE_CATEGORY_DEFS_RAW
        ]

        # Existing PG video, if any, so Step 4 can show "Current video source" block
        existing_video = PGColivingVideo.objects.filter(property=pg).order_by('-id').first()

        context = {
            'pg': pg,
            
            'ameneties_obj': amenities_obj,
            'facilities_obj': facilities_obj,
            
            'user_obj': user_obj,
            
            
           
            'image_category_defs': image_category_defs,
            'existing_video': existing_video,
            'selected_amenities': [a.strip() for a in (pg.amenities or "").split(",") if a.strip()],
            'selected_facilities': [f.strip() for f in (pg.nearby_facilities or "").split(",") if f.strip()],
        }
        return render(request, 'agent/Forms/Rental/pg_edit.html', context)

    # ---------- POST REQUEST PROCESSING (Update Logic) ----------
    try:
        def get_list(name):
            return ",".join(request.POST.getlist(name))

        def _to_int_or_none(val):
            try:
                return int(val) if val != '' and val is not None else None
            except (TypeError, ValueError):
                return None

        def _to_bool(val):
            return val in ["on", "true", "True", "1", True]

        IMAGE_CATEGORY_FIELDS = {
            'exterior': 'property_images_exterior[]', 'single_room': 'property_images_single_room[]',
            'double_room': 'property_images_double_room[]', 'multi_room': 'property_images_multi_room[]',
            'lounge': 'property_images_lounge[]', 'kitchen': 'property_images_kitchen[]',
            'dining': 'property_images_dining[]', 'washroom': 'property_images_washroom[]',
            'laundry': 'property_images_laundry[]', 'parking': 'property_images_parking[]',
            'amenities': 'property_images_amenities[]', 'floor_plan': 'property_images_floor_plan[]',
        }

        with transaction.atomic():
            pg.building_name = (request.POST.get('building_name') or '').strip()
            pg.property_no = (request.POST.get('property_no') or pg.property_no or '').strip()
            pg.city = (request.POST.get('city') or '').strip()
            pg.locality = (request.POST.get('locality') or '').strip()
            pg.address = (request.POST.get('address') or request.POST.get('property_address') or '').strip()
            pg.total_beds = _to_int_or_none(request.POST.get('room_beds') or request.POST.get('total_beds')) or pg.total_beds
            pg.pg_for = (request.POST.get('pg_for') or '').strip()
            pg.furnishing_status = request.POST.get('furnishing_status') or request.POST.get('furnishing_type') or pg.furnishing_status
            pg.best_suited_for = (request.POST.get('best_suited_for') or '').strip()

            pg.wing_number = (request.POST.get('wing_number') or '').strip()
            pg.property_landmark = (request.POST.get('property_landmark') or '').strip()
            pg.pincode = (request.POST.get('pincode') or '').strip()
            pg.state = (request.POST.get('state') or '').strip()
            pg.google_maps_link = (request.POST.get('google_maps_link') or '').strip()
            pg.latitude = (request.POST.get('latitude') or '').strip()
            pg.longitude = (request.POST.get('longitude') or '').strip()

            # Room type & pricing (Step 2)
            pg.room_type = (request.POST.get('room_type') or pg.room_type or '').strip()
            monthly_rent_val = _to_int_or_none(request.POST.get('room_rent') or request.POST.get('monthly_rent'))
            if monthly_rent_val is not None:
                pg.monthly_rent = monthly_rent_val
            pg.advance_rent_month = (request.POST.get('advance_rent_month') or '').strip()
            pg.advance_rent_amount = _to_int_or_none(request.POST.get('advance_rent_amount'))
            pg.security_deposit_type = (request.POST.get('security_deposit_type') or '').strip()
            pg.security_deposit_amount = _to_int_or_none(request.POST.get('security_deposit_amount'))
            pg.maintenance_type = (request.POST.get('maintenance_type') or '').strip()
            pg.maintenance_amount = _to_int_or_none(request.POST.get('maintenance_amount'))
            pg.total_move_in_cost = _to_int_or_none(request.POST.get('total_move_in_cost'))
            pg.brokerage_percentage = (request.POST.get('brokerage_percentage') or '').strip()
            pg.manual_brokerage = (request.POST.get('manual_brokerage') or '').strip()

            meals_available = _to_bool(request.POST.get('meals_available'))
            pg.meals_available = meals_available
            pg.meal_offerings = (request.POST.get('meal_offerings') or '').strip() if meals_available else None
            pg.meal_speciality = (request.POST.get('meal_speciality') or '').strip() if meals_available else None

            pg.notice_period = _to_int_or_none(request.POST.get('notice_period'))
            pg.lockin_period = _to_int_or_none(request.POST.get('lockin_period'))
            pg.minimum_stay = _to_int_or_none(request.POST.get('minimum_stay')) or 1

            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    pg.available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except ValueError:
                    pass

            pg.property_managed_by = request.POST.get('property_managed_by', pg.property_managed_by)
            pg.manager_stays = _to_bool(request.POST.get('manager_stays'))
            # Listed By
            listed_by_type = request.POST.get('listed_by_type', 'self')
            pg.listed_by_type = listed_by_type
            pg.listed_by_id = (request.POST.get('listed_by_id') or '').strip()
            pg.listed_by_name = (request.POST.get('listed_by_name') or '').strip()
            pg.listed_by_email = (request.POST.get('listed_by_email') or '').strip()
            pg.listed_by_contact = (request.POST.get('listed_by_contact') or '').strip()
            pg.listed_by_role = (request.POST.get('listed_by_role') or '').strip()
            if listed_by_type == 'other':
                pg.assigned_to = (request.POST.get('assigned_to') or '').strip()
            else:
                pg.assigned_to = ''

            # Regulations
            pg.opposite_gender_visitors_allowed = "opposite_gender_visitors_allowed" in request.POST or "opposite_sex_allowed" in request.POST
            pg.visitors_allowed = "visitors_allowed" in request.POST
            pg.parents_guardians_allowed = "parents_guardians_allowed" in request.POST or "guardian_allowed" in request.POST
            pg.entry_24x7_allowed = "entry_24x7_allowed" in request.POST or "any_time_allowed" in request.POST
            pg.curfew_time = request.POST.get('curfew_time') or pg.curfew_time
            pg.smoking_allowed = "smoking_allowed" in request.POST
            pg.alcohol_consumption_allowed = "alcohol_consumption_allowed" in request.POST or "drinking_allowed" in request.POST
            pg.couples_allowed = "couples_allowed" in request.POST
            pg.pets_allowed = "pets_allowed" in request.POST
            pg.cooking_allowed = "cooking_allowed" in request.POST
            pg.police_verification_required = "police_verification_required" in request.POST

            incoming_amenities = get_list("amenities[]")
            if incoming_amenities:
                pg.amenities = incoming_amenities
            incoming_facilities = get_list("facilities[]") or get_list("nearby_facilities[]")
            if incoming_facilities:
                pg.nearby_facilities = incoming_facilities

            pg.user_description = request.POST.get("user_description") or request.POST.get("property_description") or pg.user_description
            pg.owner_name = request.POST.get("owner_name", "")
            pg.contact_number = request.POST.get("contact_number", "")
            pg.email = request.POST.get("email", "")
            pg.alternate_contact = request.POST.get("alternate_contact", "")
            # Listed elsewhere / portal
            pg.listed_elsewhere = (request.POST.get('listed_elsewhere') or '').strip()
            pg.portal_name = (request.POST.get('portal_name') or '').strip()

            # Listing status & approval status (new fields added to Step 4)
            pg.listing_status = (request.POST.get('listing_status') or pg.listing_status or '').strip()
            pg.approval_status = (request.POST.get('approval_status') or pg.approval_status or '').strip()

            

            pg.save()

            # ---------- Handle removal of existing images ----------
            remove_ids = request.POST.getlist('remove_image_ids[]')
            if remove_ids:
                PGPropertyImage.objects.filter(property=pg, id__in=remove_ids).delete()

            # ---------- Category Images Handling (new uploads) ----------
            saved_count = PGPropertyImage.objects.filter(property=pg).count()
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                # Continue sequence numbering after existing images in this category
                existing_in_cat = PGPropertyImage.objects.filter(property=pg, category=category).count()
                for offset, img in enumerate(cat_images):
                    if saved_count >= 30:
                        break
                    PGPropertyImage.objects.create(
                        property=pg, image=img, category=category,
                        sequence_order=existing_in_cat + offset
                    )
                    saved_count += 1

            # ---------- Video handling ----------
            # ---------- Video handling ----------
            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video') or request.FILES.get('video')
            property_video_link = request.POST.get('property_video_link', '')
            delete_current_video = request.POST.get('delete_current_video')

            if delete_current_video:
                PGColivingVideo.objects.filter(property=pg).delete()

            saved_images = list(PGPropertyImage.objects.filter(property=pg))
            image_paths = [
                i.image.path for i in saved_images 
                if i.image and hasattr(i.image, 'path') and os.path.exists(i.image.path)
            ]

            # 1. Detect if images were added or removed in this request
            images_changed = False
            if request.POST.getlist('delete_image_ids[]'):
                images_changed = True
            for key in request.FILES.keys():
                if 'image' in key.lower():  # Catches property_images_bedroom, etc.
                    images_changed = True
                    break

            regenerate_slideshow = request.POST.get('regenerate_slideshow') == 'on'
            
            # 2. Strictly check if an ACTUAL video file exists, not just an empty database row
            has_existing_valid_video = PGColivingVideo.objects.filter(
                property=pg, source='auto'
            ).exclude(video='').exclude(video__isnull=True).exists()

            # 3. Regenerate if requested, missing, OR if images were changed
            should_generate = video_option == 'auto' and len(image_paths) >= 3 and (
                regenerate_slideshow or not has_existing_valid_video or images_changed
            )

            if should_generate:
                try:
                    from Admin_App.utils import generate_property_slideshow
                    import time
               

                    # Delete the old physical file before writing a new one
                    video_row = PGColivingVideo.objects.filter(property=pg, source='auto').first()
                    if video_row and video_row.video and video_row.video.name:
                        old_path = video_row.video.path
                        if os.path.exists(old_path):
                            try:
                                os.remove(old_path)
                            except Exception as del_err:
                                print("Could not delete old auto video:", del_err)

                    # Unique filename per regeneration to break browser caching
                    out_path = f"pg_coliving/videos/auto_{pg.id}_{int(time.time())}.mp4"
                    
                    result_path = generate_property_slideshow(image_paths, out_path)
                    print("AUTO SLIDESHOW RESULT:", result_path)
                    
                    if result_path:
                        PGColivingVideo.objects.update_or_create(
                            property=pg, 
                            source='auto', 
                            defaults={
                                'video': result_path, 
                                'video_url': None
                            }
                        )
                except Exception as ve:
                    
                    print("PG VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()
            elif video_option == 'auto' and len(image_paths) < 3:
                # Fallback: Clear it out if they requested 'auto' but removed too many images
                PGColivingVideo.objects.update_or_create(
                    property=pg,
                    source='auto',
                    defaults={
                        'video': None,
                        'video_url': None
                    }
                )

            if video_option == 'upload' and uploaded_video and uploaded_video.size <= 50 * 1024 * 1024:
                PGColivingVideo.objects.update_or_create(
                    property=pg, source='upload', defaults={'video': uploaded_video, 'video_url': None}
                )
            elif video_option == 'rm_assisted' and property_video_link:
                PGColivingVideo.objects.update_or_create(
                    property=pg, source='rm_assisted', defaults={'video_url': property_video_link, 'video': None}
                )

        return JsonResponse({
            "status": "success",
            "message": "PG Property Updated Successfully.",
            "redirect_url": reverse('pg_agent_list')
        })

    except Exception as e:
        print("ERROR IN PG EDIT:", str(e))
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})







def export_pg_coliving_agent(request):

    """Dedicated view for exporting PG/Coliving properties to CSV or Excel matching the template format and sequence with Sr.No."""
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    try:
        properties = PGColivingProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-created_at')
    except Exception:
        properties = PGColivingProperty.objects.filter(listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-created_at')

    # ── Apply Search Filters ──
    search_query = request.GET.get('search', '').strip()
    city_query   = request.GET.get('city', '').strip()
    pg_for_query = request.GET.get('pg_for', '').strip()
    from_date    = request.GET.get('from_date', '').strip()
    to_date      = request.GET.get('to_date', '').strip()

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(listed_by_name__icontains=search_query)
        )
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if pg_for_query and pg_for_query != 'All Types':
        properties = properties.filter(pg_for__icontains=pg_for_query)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ── EXACT MATCH TO TEMPLATE SEQUENCE, SECTIONS, & FIELD LABELS (+ SR.NO & AUDIT) ──
    EXPORT_COLS = [
        # 📋 Sr.No & Meta
        ("📋 Sr.No & Meta", "sr_no", False, "Sr.No"),

        # 📋 Listed By Details (6)
        ("📋 Listed By Details", "listed_by_type", False, "Listed By Type"),
        ("📋 Listed By Details", "listed_by_id", False, "Listed By Id"),
        ("📋 Listed By Details", "listed_by_name", False, "Listed By Name"),
        ("📋 Listed By Details", "listed_by_email", False, "Listed By Email"),
        ("📋 Listed By Details", "listed_by_contact", False, "Listed By Contact"),
        ("📋 Listed By Details", "listed_by_role", True, "Listed By Role *"),

        # 📋 Basic Information (15)
         
        ("📋 Basic Information", "id", False, "Property ID(System Generated)"),
        ("📋 Basic Information", "property_title", False, "Property Title(System Generated)"),
        ("📋 Basic Information", "building_name", False, "Building / Project / Society Name"),
        ("📋 Basic Information", "property_no", True, "Property No/Unit No *"),
        ("📋 Basic Information", "wing_number", False, "Tower/Wing Number"),
        ("📋 Basic Information", "city", True, "City *"),
        ("📋 Basic Information", "locality", True, "Locality *"),
        ("📋 Basic Information", "address", True, "Property Address *"),
        ("📋 Basic Information", "property_landmark", False, "Property Landmark"),
        ("📋 Basic Information", "pincode", True, "Pincode *"),
        ("📋 Basic Information", "state", True, "State *"),
        ("📋 Basic Information", "google_maps_link", False, "Google Maps Link"),
        ("📋 Basic Information", "latitude", False, "Latitude"),
        ("📋 Basic Information", "longitude", False, "Longitude"),
        ("📋 Basic Information", "pg_for", True, "PG for *"),
        ("📋 Basic Information", "furnishing_status", True, "Furnishing Status *"),
        ("📋 Basic Information", "best_suited_for", False, "Best Suited For"),

        
        ("📋 Basic Information", "meals_available", False, "Meals Available?"),
        ("📋 Basic Information", "meal_offerings", False, "Meal Offerings"),
        ("📋 Basic Information", "meal_speciality", False, "Meal Speciality"),
        ("📋 Basic Information", "notice_period", False, "Notice Period (Days)"),
        ("📋 Basic Information", "lockin_period", False, "Lock-in Period (Days)"),
        ("📋 Basic Information", "minimum_stay", True, "Minimum Stay (Months) *"),
        ("📋 Basic Information", "available_from", True, "Available From *"),
        ("📋 Basic Information", "property_managed_by", False, "Property Managed By"),
        ("📋 Basic Information", "manager_stays", False, "Property Manager Stays at Property?"),
   

        # 📋 Room Details & Pricing (10)
        ("📋 Room Details & Pricing", "room_type", True, "Room Type *"),
        ("📋 Room Details & Pricing", "total_beds", True, "Total Beds In Room *"),
        ("📋 Room Details & Pricing", "monthly_rent", True, "Rent Per Occupant (₹/Bed/Month/Person) *"),
        ("📋 Room Details & Pricing", "brokerage_percentage", True, "Brokerage *"),
        ("📋 Room Details & Pricing", "manual_brokerage", False, "Enter Fixed Brokerage"),
        ("📋 Room Details & Pricing", "advance_rent_month", False, "Advance Rent Month"),
        ("📋 Room Details & Pricing", "advance_rent_amount", False, "Advance Rent Amount"),
        ("📋 Room Details & Pricing", "security_deposit_type", True, "Refundable Room Security_Deposite (₹/Person) *"),
        ("📋 Room Details & Pricing", "security_deposit_amount", False, "Refundable Security Deposit Amount"),
        ("📋 Room Details & Pricing", "maintenance_type", False, "Maintenance Type"),
        ("📋 Room Details & Pricing", "maintenance_amount", False, "Monthly Maintenance Amount"),
        ("📋 Room Details & Pricing", "total_move_in_cost", False, "Estimated Move-In Cost (Per Occupant)"),

        # 📋 PG Regulations (11)
        ("📋 PG Regulations", "opposite_gender_visitors_allowed", False, "Opposite Gender Visitors Allowed?"),
        ("📋 PG Regulations", "visitors_allowed", False, "Visitors Allowed?"),
        ("📋 PG Regulations", "parents_guardians_allowed", False, "Parents/Guardians Allowed?"),
        ("📋 PG Regulations", "entry_24x7_allowed", False, "24x7 Entry Allowed?"),
        ("📋 PG Regulations", "curfew_time", False, "Curfew Time"),
        ("📋 PG Regulations", "smoking_allowed", False, "Smoking Allowed?"),
        ("📋 PG Regulations", "alcohol_consumption_allowed", False, "Alcohol Consumption Allowed?"),
        ("📋 PG Regulations", "couples_allowed", False, "Couples Allowed?"),
        ("📋 PG Regulations", "pets_allowed", False, "Pets Allowed?"),
        ("📋 PG Regulations", "cooking_allowed", False, "Cooking Allowed?"),
        ("📋 PG Regulations", "police_verification_required", False, "Police Verification Required?"),

        # 📋 Amenities & Facilities (2)
        ("📋 Amenities", "amenities", False, "Amenities"),
        ("📋 Nearby Facilities", "nearby_facilities", False, "Nearby Facilities"),

        # 📋 Description (1)
        ("📋 Description", "property_summary", False, "Property Summary(Auto)"),
        ("📋 Description", "property_description", False, "Property Description(Auto)"),
        ("📋 Description", "user_description", False, "Property Description(Added By User)"),

        # 📋 Media & Listing Status (3)
        ("📋 Media & Listing Status", "listed_elsewhere", False, "Is Property Already Listed Elsewhere?"),
        ("📋 Media & Listing Status", "portal_name", False, "Portal Name"),
        

       
        ("⚙️ System Meta", "uploaded_by_name", False, "Uploaded By (Name)"),
        ("⚙️ System Meta", "uploaded_by_email", False, "Uploaded By (Email)"),
        ("⚙️ System Meta", "uploaded_by_contact", False, "Uploaded By (Contact)"),
        ("⚙️ System Meta", "uploaded_by_role", False, "Uploaded By (Role)"),
        
        ("⚙️ System Meta", "created_at", False, "Created At"),
       
        ("⚙️ System Meta", "listing_status", False, "Listing Status"),
        ("⚙️ System Meta", "approval_status", False, "Approval Status"),
    ]

    export_format = request.GET.get('format', 'excel')

    # ── EXCEL EXPORT ──
    if export_format == 'excel':
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PG Coliving"

            HDR_BG, META_BG = "1F4E79", "475569"
            REQ_BG, OPT_BG = "FFD7D7", "F8FAFC"
            thin = Side(style="thin", color="BBBBBB")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

            contiguous_sections = []
            for i, (sec, *_) in enumerate(EXPORT_COLS):
                col_idx = i + 1
                if not contiguous_sections or contiguous_sections[-1]['sec'] != sec:
                    contiguous_sections.append({'sec': sec, 'start': col_idx, 'end': col_idx})
                else:
                    contiguous_sections.append({'sec': sec, 'start': contiguous_sections[-1]['start'], 'end': col_idx})
                    contiguous_sections.pop(-2)

            for item in contiguous_sections:
                sec = item['sec']
                start_col = item['start']
                end_col = item['end']

                fill_color = META_BG if "System Meta" in sec else HDR_BG

                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.border = bdr

                top_cell = ws.cell(row=1, column=start_col)
                top_cell.value = sec
                top_cell.font = Font(bold=True, color="FFFFFF", size=10)
                top_cell.alignment = Alignment(horizontal="center", vertical="center")

                if start_col < end_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            for ci, (sec, field, req, hint) in enumerate(EXPORT_COLS, 1):
                lc = ws.cell(row=2, column=ci, value=hint) 
                lc.font, lc.border = Font(bold=True, size=9), bdr
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(ci)].width = max(18, len(hint) * 0.9)

            for row_idx, prop in enumerate(properties, start=3):
                for col_idx, (_, field_name, _, _) in enumerate(EXPORT_COLS, start=1):
                    if field_name == "sr_no":
                        val = row_idx - 2
                    elif field_name == "upload_file_name":
                        raw_file = getattr(prop, field_name, "")
                        val = raw_file if raw_file else "Web Listing UI Form"
                    else:
                        val = getattr(prop, field_name, "") if hasattr(prop, field_name) else ""
                    
                    try:
                        if val is True: val = "True"
                        elif val is False: val = "False"
                        elif hasattr(val, 'strftime'): 
                            val = val.strftime('%Y-%m-%d %H:%M:%S') if "at" in field_name else val.strftime('%Y-%m-%d')
                        elif isinstance(val, list): val = ", ".join(map(str, val))
                        elif hasattr(val, 'url'): val = val.url if val else ""
                    except ValueError: 
                        val = ""
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                    cell.alignment, cell.border = Alignment(vertical="center", wrap_text=True), bdr

            ws.row_dimensions[1].height, ws.row_dimensions[2].height = 28, 36
            ws.freeze_panes = "A3"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="PG_Coliving_Template_Export.xlsx"'
            return response
            
        except Exception as e:
            return HttpResponse(f"<pre>ERROR GENERATING EXCEL:\n{str(e)}\n{traceback.format_exc()}</pre>", status=500)

    # ── CSV EXPORT ──
    elif export_format == 'csv':
        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="PG_Coliving_Template_Export.csv"'
            writer = csv.writer(response)
            
            writer.writerow([hint for _, _, _, hint in EXPORT_COLS])
            
            for row_idx, prop in enumerate(properties, start=1):
                row_data = []
                for _, field_name, _, _ in EXPORT_COLS:
                    if field_name == "sr_no":
                        val = row_idx
                    elif field_name == "upload_file_name":
                        raw_file = getattr(prop, field_name, "")
                        val = raw_file if raw_file else "Web Listing UI Form"
                    else:
                        val = getattr(prop, field_name, "") if hasattr(prop, field_name) else ""
                    
                    try:
                        if val is True: val = "True"
                        elif val is False: val = "False"
                        elif hasattr(val, 'strftime'): 
                            val = val.strftime('%Y-%m-%d %H:%M:%S') if "at" in field_name else val.strftime('%Y-%m-%d')
                        elif isinstance(val, list): val = ", ".join(map(str, val))
                        elif hasattr(val, 'url'): val = val.url if val else ""
                    except ValueError: 
                        val = ""
                        
                    row_data.append(str(val) if val is not None else "")
                writer.writerow(row_data)
                
            return response
            
        except Exception as e:
            return HttpResponse(f"<pre>ERROR GENERATING CSV:\n{str(e)}\n{traceback.format_exc()}</pre>", status=500)





def pg_coliving_view_agent(request, pk):
    
    pg = get_object_or_404(PGColivingProperty, pk=pk)
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # 2. Parse Comma-Separated Strings into Lists for HTML "Chips"
    def split_to_list(db_string):
        return [x.strip() for x in db_string.split(',') if x.strip()] if db_string else []

    # --- CATEGORY-WISE IMAGE GROUPING ---
    category_labels = dict(PGPropertyImage.CATEGORY_CHOICES)
    images_queryset = pg.images.all()
    
    grouped_images = {}
    for img in images_queryset:
        cat_key = img.category
        if cat_key not in grouped_images:
            grouped_images[cat_key] = {
                'label': category_labels.get(cat_key, cat_key.title()),
                'images': []
            }
        grouped_images[cat_key]['images'].append(img)

    # --- VIDEO HIERARCHY ENGINE ---
    videos_queryset = pg.videos.all()
    rm_video = videos_queryset.filter(source='rm_assisted', video_url__isnull=False).exclude(video_url='').first()
    manual_video = videos_queryset.filter(source='uploaded', video__isnull=False).first()
    auto_video = videos_queryset.filter(source='auto', video__isnull=False).first()

    selected_video = None
    video_display_mode = None

    if rm_video:
        selected_video = rm_video
        video_display_mode = 'rm_assisted'
    elif manual_video:
        selected_video = manual_video
        video_display_mode = 'manual'
    elif auto_video:
        selected_video = auto_video
        video_display_mode = 'auto'

    context = {
        
        'pg': pg,
        'user_obj': user_obj,
      
        
        'starting_rent': pg.monthly_rent or 0,
        'sharing_type_list': split_to_list(pg.room_type), # Mapped to flattened field
        'best_suited_list': split_to_list(pg.best_suited_for),
        'amenities_list': split_to_list(pg.amenities),
        'facilities_list': split_to_list(pg.nearby_facilities),
        'meal_offerings_list': split_to_list(pg.meal_offerings),
        'meal_speciality_list': split_to_list(pg.meal_speciality),
        'grouped_images': grouped_images,
        'selected_video': selected_video,
        'video_display_mode': video_display_mode,
    }

    return render(request, "agent/Reports/Rental/pg_coliving_view.html", context)




@require_POST
def pg_coliving_delete_agent(request, pk):
    """Agent Soft Delete PG/Coliving Property (supports Admin impersonation)"""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Agent not found.'
            })

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )

        # ======================================
        # 5. Fetch Property
        # ======================================
        pg = get_object_or_404(PGColivingProperty, id=pk)

        # ======================================
        # 6. Security Check — agent can delete only own property
        # ======================================
        if str(pg.listed_by_id) != str(agent_obj.user_id):
            return JsonResponse({
                'status': 'error',
                'message': 'You can delete only your own properties.'
            })

        # ======================================
        # 7. Soft Delete
        # ======================================
        pg.is_deleted = True
        pg.deleted_at = timezone.now()
        pg.deleted_by = user_identity
        pg.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Moved to Recycle Bin successfully!'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })



@require_POST
def pg_bulk_delete_agent(request):
    """Agent Advanced Bulk Delete (Soft Delete) — scoped to agent's own PG/Coliving properties, supports Admin impersonation."""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'}, status=404)

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone} | "
                f"Role: {agent_obj.user_role}"
            )

        # ======================================
        # 5. Parse Payload
        # ======================================
        if not request.body:
            return JsonResponse({'status': 'error', 'message': 'Empty request body payload.'}, status=400)

        data = json.loads(request.body)
        delete_type = data.get('delete_type')

        # ======================================
        # 6. Base Queryset — SCOPED TO THIS AGENT ONLY
        # ======================================
        properties = PGColivingProperty.objects.filter(
            is_deleted=False,
            listed_by_id=agent_obj.user_id
        )
        target_props = properties.none()

        if delete_type == 'delete_all':
            target_props = properties

        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(pg_property_id__in=page_ids)

        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            target_props = properties.filter(created_at__date__range=[from_date, to_date])

        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)

        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)

        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '')
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_contact__icontains=uploader) |
                Q(owner_name__icontains=uploader)
            )

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            target_props = properties.filter(upload_file_name=file_name)
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'}, status=400)

        count = target_props.count()
        if count == 0:
            return JsonResponse({'status': 'success', 'message': 'No matching active records found to process.'})

        target_props.update(
            is_deleted=True,
            deleted_at=timezone.now(),
            deleted_by=user_identity
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully moved {count} PG properties to the Recycle Bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)





############# Views END for Rental pg Module for Agent ###################################################











#################Views Start For Resale Residential Listing Property###########################



def residential_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    



    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale/residential_resale.html", context)


############## Views start for residential resale form #####################

def residential_resale_agent_view(request,id):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    property = ResaleResidentialProperty.objects.get(id=id)

    # Convert comma-separated string datasets into lists for badge generation loop arrays
    facilities_list = [f.strip() for f in property.nearby_facilities.split(',')] if property.nearby_facilities else []
    amenities_list = [a.strip() for a in property.amenities.split(',')] if property.amenities else []
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'prop':property,
        "images": property.images.all(),
        "facilities_list": facilities_list,
        "amenities_list": amenities_list,
    }
    
    return render(request, "agent/Forms/Resale/residential_resale_view.html", context)

############ Views end for residential resale form ############################


############### Views start for residential resale edit form ###################

def residential_resale_agent_edit(request,id):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    property = ResaleResidentialProperty.objects.get(id=id)

    existing_images = property.images.all()

    facilities_obj = Facilities_Details.objects.all()
    ameneties_obj = Ameneties_Details.objects.all()

    amenities_list = []
    if property.amenities:
        amenities_list = [a.strip() for a in property.amenities.split(',') if a.strip()]
    
    # Parse facilities from stored comma-separated string
    nearby_facilities_list = []
    if property.nearby_facilities:
        nearby_facilities_list = [f.strip() for f in property.nearby_facilities.split(',')]
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        "amenities_list": amenities_list,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj,
        'prop':property,
        'existing_images':existing_images,
        'facilities_obj':facilities_obj,
        'ameneties_obj':ameneties_obj,
        'amenities_list':amenities_list,
        'nearby_facilities_list':nearby_facilities_list, 
    }
    
    return render(request, "agent/Forms/Resale/residential_resale_edit.html", context)

############# Views end for residential resale edit form #####################


def residential_resale_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    all_properties = (
        ResaleResidentialProperty.objects
        .filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
        .prefetch_related('images')
        .order_by('-created_at')
    )

    properties = all_properties

    ############# Filter Parameters Start ###################################

    # ── Read query params ────────────────────────────────────────────────────
    search_query    = request.GET.get('search', '').strip()
    prop_type       = request.GET.get('prop_type', '').strip()
    bhk_filter      = request.GET.get('bhk', '').strip()
    furnish         = request.GET.get('furnish', '').strip()
    occupancy       = request.GET.get('occupancy', '').strip()
    ownership       = request.GET.get('ownership', '').strip()
    negotiable      = request.GET.get('negotiable', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()
    listed_by_type  = request.GET.get('listed_by_type', '').strip()
    property_age    = request.GET.get('property_age', '').strip()
    facing_direction = request.GET.get('facing_direction', '').strip()
    society_type    = request.GET.get('society_type', '').strip()
    city_filter     = request.GET.get('city', '').strip()
    status_filter   = request.GET.get('listing_status', '').strip()

    # ── Apply filters ────────────────────────────────────────────────────────
    properties = all_properties

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query)           |
            Q(locality__icontains=search_query)       |
            Q(listed_by_name__icontains=search_query) |
            Q(bhk__icontains=search_query)            |
            Q(building_name__icontains=search_query)  |
            Q(property_no__icontains=search_query)    |
            Q(address__icontains=search_query)
        )

    if prop_type:
        properties = properties.filter(property_type=prop_type)

    if bhk_filter:
        properties = properties.filter(bhk=bhk_filter)

    if furnish:
        properties = properties.filter(furnishing_status=furnish)

    if occupancy:
        properties = properties.filter(occupancy_status=occupancy)

    if ownership:
        properties = properties.filter(ownership_status=ownership)

    if negotiable:
        properties = properties.filter(price_negotiable=negotiable)

    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)

    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    if listed_by_type:
        properties = properties.filter(listed_by_type=listed_by_type)

    if property_age:
        properties = properties.filter(property_age=property_age)

    if facing_direction:
        properties = properties.filter(facing_direction=facing_direction)

    if society_type:
        properties = properties.filter(society_type=society_type)

    if city_filter:
        properties = properties.filter(city__icontains=city_filter)

    if status_filter:
        properties = properties.filter(listing_status=status_filter)

    # ── Thumbnail + helper attributes for each filtered property ─────────────
    for prop in properties:
        prop.thumbnail = prop.images.first()

        prop.nearby_facilities_list = (
            [f.strip() for f in prop.nearby_facilities.split(',')]
            if prop.nearby_facilities else []
        )
        prop.amenities_list = (
            [a.strip() for a in prop.amenities.split(',')]
            if prop.amenities else []
        )
        prop.image_count = prop.images.count()
        prop.image_urls = [img.image.url for img in prop.images.all()]

    # ════════════════════════════════════════════════════════════════════════
    # KPI STATS (Calculated using correct DB layout keys)
    # ════════════════════════════════════════════════════════════════════════
    total_count = all_properties.count()

    # ── Row 1 — Inventory ────────────────────────────────────────────────────
    total_negotiable  = all_properties.filter(price_negotiable='yes').count()
    total_furnished   = all_properties.filter(furnishing_status='fully').count()
    total_freehold    = all_properties.filter(ownership_status='Self Owned').count()
    total_with_images = all_properties.filter(images__isnull=False).distinct().count()
    total_with_vacant = all_properties.filter(occupancy_status='Vacant').count()
    total_with_tenants = all_properties.filter(existing_tenants='yes').count()

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    negotiable_pct = pct(total_negotiable, total_count)
    furnished_pct  = pct(total_furnished, total_count)
    freehold_pct   = pct(total_freehold, total_count)
    images_pct     = pct(total_with_images, total_count)
    vacant_pct     = pct(total_with_vacant, total_count)

    # ── Row 2 — Pricing ──────────────────────────────────────────────────────
    price_agg = all_properties.aggregate(
        avg      = Avg('selling_price'),
        max_val  = Max('selling_price'),
        min_val  = Min('selling_price'),
        avg_sqft = Avg('price_per_sqft'),
        avg_area = Avg('builtup_area'),
        total_value = Coalesce(Sum('selling_price'), Decimal('0.00')),
    )
    avg_price      = price_agg['avg']
    max_price      = price_agg['max_val']
    min_price      = price_agg['min_val']
    avg_price_sqft = price_agg['avg_sqft']
    avg_builtup    = price_agg['avg_area']
    total_value    = price_agg['total_value']
    total_with_loan = all_properties.filter(property_loan='yes').count()

    # ── Row 3 — Legal & Status ───────────────────────────────────────────────
    no_dispute_count  = all_properties.filter(any_legal_dispute='no').count()
    dispute_count     = all_properties.filter(any_legal_dispute='yes').count()
    tax_pending_count = all_properties.filter(government_tax='yes').count()
    tenant_occupied   = all_properties.filter(existing_tenants='yes').count()
    premium_count     = all_properties.filter(selling_price__gte=10000000).count()  # >= 1 Cr
    budget_count      = all_properties.filter(selling_price__lt=3000000).count()    # < 30 L

    # ── Row 4 — Listing Quality ──────────────────────────────────────────────
    with_video_count = (
        all_properties
        .exclude(property_video__isnull=True)
        .exclude(property_video='')
        .count()
    )
    with_floor_plan = (
        all_properties
        .exclude(floor_plan__isnull=True)
        .exclude(floor_plan='')
        .count()
    )
    with_listed_by = (
        all_properties
        .exclude(listed_by_name__isnull=True)
        .exclude(listed_by_name='')
        .count()
    )
    with_uploaded_by = (
        all_properties
        .exclude(uploaded_by_name__isnull=True)
        .exclude(uploaded_by_name='')
        .count()
    )
    with_description = (
        all_properties
        .exclude(property_description__isnull=True)
        .exclude(property_description='')
        .count()
    )
    pending_approval = all_properties.filter(approval_status='Pending').count()
    approved_count   = all_properties.filter(approval_status='Approved').count()
    rejected_count   = all_properties.filter(approval_status='Rejected').count()

    # ── Charts ───────────────────────────────────────────────────────────────
    property_type_counts = dict(
        all_properties.values('property_type')
        .annotate(count=Count('id'))
        .values_list('property_type', 'count')
    )
    
    bhk_counts = dict(
        all_properties.values('bhk')
        .annotate(count=Count('id'))
        .values_list('bhk', 'count')
    )
    
    furnishing_counts = {
        'fully_furnished': all_properties.filter(furnishing_status='fully').count(),
        'semi_furnished': all_properties.filter(furnishing_status='semi').count(),
        'unfurnished': all_properties.filter(furnishing_status='unfurnished').count(),
    }
    
    occupancy_counts = dict(
        all_properties.values('occupancy_status')
        .annotate(count=Count('id'))
        .values_list('occupancy_status', 'count')
    )
    
    facing_counts = dict(
        all_properties.values('facing_direction')
        .annotate(count=Count('id'))
        .values_list('facing_direction', 'count')
    )
    
    society_type_counts = dict(
        all_properties.values('society_type')
        .annotate(count=Count('id'))
        .values_list('society_type', 'count')
    )
    
    ownership_counts = dict(
        all_properties.values('ownership_status')
        .annotate(count=Count('id'))
        .values_list('ownership_status', 'count')
    )
    
    listed_by_counts = dict(
        all_properties.values('listed_by_type')
        .annotate(count=Count('id'))
        .values_list('listed_by_type', 'count')
    )
    
    approval_status_counts = dict(
        all_properties.values('approval_status')
        .annotate(count=Count('id'))
        .values_list('approval_status', 'count')
    )

    # ── Monthly Property Listings (for trend chart) ──────────────────────────
    from django.db.models.functions import TruncMonth
    monthly_listings = (
        all_properties
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_labels = [m['month'].strftime('%b %Y') for m in monthly_listings] if monthly_listings else []
    monthly_counts = [m['count'] for m in monthly_listings] if monthly_listings else []

    # ── Unique values for Select2 searchable dropdowns ───────────────────────
    unique_prop_types = list(
        all_properties.values_list('property_type', flat=True)
        .distinct().order_by('property_type')
    )
    unique_bhk_values = list(
        all_properties.values_list('bhk', flat=True)
        .distinct().order_by('bhk')
    )
    unique_furnishings = list(
        all_properties.values_list('furnishing_status', flat=True)
        .distinct().order_by('furnishing_status')
    )
    unique_occupancies = list(
        all_properties.values_list('occupancy_status', flat=True)
        .distinct().order_by('occupancy_status')
    )
    unique_ownerships = list(
        all_properties.values_list('ownership_status', flat=True)
        .distinct().order_by('ownership_status')
    )
    unique_cities = list(
        all_properties.values_list('city', flat=True)
        .distinct().order_by('city')
    )
    unique_states = list(
        all_properties.values_list('state', flat=True)
        .distinct().order_by('state')
    )
    unique_facings = list(
        all_properties.values_list('facing_direction', flat=True)
        .distinct().order_by('facing_direction')
    )
    unique_society_types = list(
        all_properties.values_list('society_type', flat=True)
        .distinct().order_by('society_type')
    )

    # ── Uploaded Files (for tracking) ──────────────────────────────────────
    try:
        uploaded_files = (
            all_properties
            .exclude(upload_file_name__isnull=True)
            .exclude(upload_file_name='')
            .values_list('upload_file_name', flat=True)
            .distinct()
        )
    except Exception:
        uploaded_files = []

    
    context = {
        'user_obj': user_obj,
        'properties':properties,

        'filtered_count': properties.count(),
        'total_count': total_count,

        # Active search params
        'search_query': search_query,
        'prop_type_query': prop_type,
        'bhk_query': bhk_filter,
        'furnish_query': furnish,
        'occupancy_query': occupancy,
        'ownership_query': ownership,
        'negotiable_query': negotiable,
        'from_date': from_date,
        'to_date': to_date,
        'listed_by_type_query': listed_by_type,
        'property_age_query': property_age,
        'facing_direction_query': facing_direction,
        'society_type_query': society_type,
        'city_filter': city_filter,
        'status_filter': status_filter,

        # Row 1 — Inventory
        'total_negotiable': total_negotiable,
        'total_furnished': total_furnished,
        'total_freehold': total_freehold,
        'total_with_images': total_with_images,
        'total_with_vacant': total_with_vacant,
        'total_with_tenants': total_with_tenants,
        'negotiable_pct': negotiable_pct,
        'furnished_pct': furnished_pct,
        'freehold_pct': freehold_pct,
        'images_pct': images_pct,
        'vacant_pct': vacant_pct,

        # Row 2 — Pricing
        'avg_price': avg_price,
        'max_price': max_price,
        'min_price': min_price,
        'avg_price_sqft': avg_price_sqft,
        'avg_builtup': avg_builtup,
        'total_value': total_value,
        'total_with_loan': total_with_loan,

        # Row 3 — Legal
        'no_dispute_count': no_dispute_count,
        'dispute_count': dispute_count,
        'tax_pending_count': tax_pending_count,
        'tenant_occupied': tenant_occupied,
        'premium_count': premium_count,
        'budget_count': budget_count,
        'pending_approval': pending_approval,
        'approved_count': approved_count,
        'rejected_count': rejected_count,

        # Row 4 — Quality
        'with_video_count': with_video_count,
        'with_floor_plan': with_floor_plan,
        'with_listed_by': with_listed_by,
        'with_uploaded_by': with_uploaded_by,
        'with_description': with_description,

        # Charts
        'property_type_counts': property_type_counts,
        'bhk_counts': bhk_counts,
        'furnishing_counts': furnishing_counts,
        'occupancy_counts': occupancy_counts,
        'facing_counts': facing_counts,
        'society_type_counts': society_type_counts,
        'ownership_counts': ownership_counts,
        'listed_by_counts': listed_by_counts,
        'approval_status_counts': approval_status_counts,
        'monthly_labels': monthly_labels,
        'monthly_counts': monthly_counts,

        # Select2 unique options
        'unique_prop_types': unique_prop_types,
        'unique_bhk_values': unique_bhk_values,
        'unique_furnishings': unique_furnishings,
        'unique_occupancies': unique_occupancies,
        'unique_ownerships': unique_ownerships,
        'unique_cities': unique_cities,
        'unique_states': unique_states,
        'unique_facings': unique_facings,
        'unique_society_types': unique_society_types,

        'uploaded_files': uploaded_files,
        
    }

    return render(request, 'agent/Reports/Resale/residential_resale_list.html', context)






#################Views Start For Resale Commercial Listing Property###########################

def commercial_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale/commercial_resale.html", context)


def commercial_resale_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agent/Reports/Resale/commercial_list.html', context)




#################Views Start For Resale Industrial Listing Property###########################

def industrial_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale/industrial_resale.html", context)


def industrial_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agent/Reports/Resale/industrial_list.html', context)


#################Views Start For Resale Agricultural Listing Property###########################


def agricultural_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale/agricultural_resale.html", context)


def agricultural_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agent/Reports/Resale/agricultural_list.html', context)




#################Views Start For Resale Plot Residential Listing Property###########################


def residential_plot_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale_Plot/residential_plot_resale.html", context)


def residential_plot_resale_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }

    return render(request, 'agent/Reports/Resale_Plot/residential_plot_resale_list.html', context)




#################Views Start For Resale Plot Commericial Listing Property###########################


def commercial_plot_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    return render(request, "agent/Forms/Resale_Plot/commercial_plot_resale.html", context)


def commercial_plot_resale_list_agent(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }

    return render(request, 'agent/Reports/Resale_Plot/commercial_plot_resale_list.html', context)




############# Views Start for Resale Plot Industrial Module for Agent ######################################



def industrial_plot_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale_Plot/industrial_plot_resale.html", context)
    

def generate_industrial_plot_fingerprint(property_no, industrial_estate_name, locality, pincode):
    """Fingerprint used ONLY by the bulk-import duplicate engine (mirrors
    rental's generate_property_fingerprint). Distinct from the model's own
    save()-time property_unique_key, which keys on address/locality/city/
    plot_area/property_no."""
    key_source = f"{property_no}|{industrial_estate_name}|{locality}|{pincode}"
    return key_source.strip().lower().replace(" ", "")



def industrial_plot_resale_add_agent(request):
    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            def to_date(val):
                if not val:
                    return None
                try:
                    return datetime.strptime(val, "%Y-%m-%d").date()
                except:
                    return None

            # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION (Who submitted the HTML form) ----------
           
            if user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "Landlord"
                
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION (Who owns/manages the listing) ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE (Checking LISTED BY, not UPLOADED BY)
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_industrial_estate_name = (request.POST.get('industrial_estate_name') or '').strip()
            input_locality = (request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_industrial_plot_fingerprint(
                input_property_no,
                input_industrial_estate_name,
                input_locality,
                input_pincode
            )

            # 1. Direct Case-Insensitive Query for same plot in same locality/estate
            direct_duplicates = IndustrialPlotResaleProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            )
            if input_industrial_estate_name:
                direct_duplicates = direct_duplicates.filter(industrial_estate_name__iexact=input_industrial_estate_name)

            # Combine fingerprint match OR direct field match
            existing_duplicates = (
                IndustrialPlotResaleProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                # Level 1: Hard Block if LISTED BY the exact same person (ID, Email, OR Phone match)
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and
                               existing_prop.listed_by_id.strip() == input_listed_by_id)

                    same_email = (existing_prop.listed_by_email and input_listed_by_email and
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)

                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                    if same_id or same_email or same_contact:
                        return JsonResponse({
                            'status': 'error',
                            'message': f"Duplicate Blocked: This plot (No {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        })

                # Level 2: Different Agent/User listing the exact same physical plot -> Allow save & Flag
                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True,
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = IndustrialPlotResaleProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Resale",
                category="Plot",
                sub_category=request.POST.get('sub_category', 'Industrial'),
              

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_no=input_property_no,

                plot_area=to_decimal(request.POST.get('plot_area')),
                property_type=request.POST.get('property_type'),
                land_use=request.POST.get('land_use'),
                industrial_zone_type=request.POST.get('industrial_zone_type'),
                industrial_estate_name=input_industrial_estate_name,
                na_status=request.POST.get('na_status'),
                layout_approval_status=request.POST.get('layout_approval_status'),
                industrial_fsi=request.POST.get('industrial_fsi'),

                plot_frontage=to_decimal(request.POST.get('plot_frontage')),
                plot_depth=to_decimal(request.POST.get('plot_depth')),
                plot_shape=request.POST.get('plot_shape'),
                plot_road_facing=request.POST.get('plot_road_facing'),
                road_width=request.POST.get('road_width'),
                corner_plot=request.POST.get('corner_plot', 'no'),

                power_supply=request.POST.get('power_supply'),
                power_load_kva=to_int(request.POST.get('power_load_kva')),
                industrial_water_supply=request.POST.get('industrial_water_supply'),
                effluent_treatment=request.POST.get('effluent_treatment'),
                industry_type_permissible=request.POST.get('industry_type_permissible'),
                plot_fencing=request.POST.get('plot_fencing'),
                loading_dock=request.POST.get('loading_dock'),
                current_possession_status=request.POST.get('current_possession_status'),

                selling_price=to_int(request.POST.get('selling_price')),
                price_per_sqft=to_int(request.POST.get('price_per_sqft')),
                price_negotiable=request.POST.get('price_negotiable', 'no'),
                additional_charges=request.POST.get('additional_charges'),
                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                ownership_type=request.POST.get('ownership_type'),
                ownership_document_type=request.POST.get('ownership_document_type'),
                other_document_type=request.POST.get('other_document_type'),
                midc_allotment=request.POST.get('midc_allotment'),
                midc_transfer_noc=request.POST.get('midc_transfer_noc'),
                environmental_clearance=request.POST.get('environmental_clearance'),
                rera_status=request.POST.get('rera_status'),
                title_clearance=request.POST.get('title_clearance'),
                property_encumbrance_status=request.POST.get('property_encumbrance_status'),

                property_tax_status=request.POST.get('property_tax_status'),
                outstanding_tax_amount=to_int(request.POST.get('outstanding_tax_amount')),
                pending_since=to_date(request.POST.get('pending_since')),
                property_loan_status=request.POST.get('property_loan_status'),
                financing_bank=request.POST.get('financing_bank'),
                outstanding_loan_amount=to_int(request.POST.get('outstanding_loan_amount')),
                sanctioning_authority=request.POST.get('sanctioning_authority'),

                amenities=amenities,
                nearby_facilities=nearby_facilities,
                user_description=request.POST.get('user_description'),
                property_summary=request.POST.get('property_summary'),
                property_description=request.POST.get('property_description'),

                state=request.POST.get('state'),
                city=request.POST.get('city'),
                locality=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                pincode=input_pincode,
                address=request.POST.get('address'),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                encumbrance_cert=request.FILES.get('encumbrance_cert'),
                layout_plan=request.FILES.get('layout_plan'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None,
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'full_plot':         'plot_images_full_plot[]',
                'main_entrance':     'plot_images_main_entrance[]',
                'boundary_fencing':  'plot_images_boundary_fencing[]',
                'road_facing':       'plot_images_road_facing[]',
                'approach_road':     'plot_images_approach_road[]',
                'truck_access':      'plot_images_truck_access[]',
                'industrial_estate': 'plot_images_industrial_estate[]',
                'electricity_infra': 'plot_images_electricity_infra[]',
                'water_infra':       'plot_images_water_infra[]',
                'aerial_drone':      'plot_images_aerial_drone[]',
                'layout_site_plan':  'plot_images_layout_site_plan[]',
            }

            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 25:
                        break
                    IndustrialPlotResaleImage.objects.create(
                        property=prop,
                        image=img,
                        category=category,
                        sequence_order=cat_index,
                    )
                    saved_count += 1

            # ---------- PLOT VIDEO (UPLOAD, RM LINK, OR AUTO SLIDESHOW) ----------
            video_option = request.POST.get('video_option') or request.POST.get('video_source') or 'auto'
            uploaded_video = request.FILES.get('property_video') or request.FILES.get('social_video')
            property_video_link = request.POST.get('property_video_link', '').strip()

            # 1. ALWAYS auto-generate the slideshow row if >= 3 photos exist
            CATEGORY_ORDER = list(IMAGE_CATEGORY_FIELDS.keys())
            saved_images = list(IndustrialPlotResaleImage.objects.filter(property=prop))
            saved_images.sort(key=lambda img: (CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, img.sequence_order))
            image_paths = [img.image.path for img in saved_images if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)]

            if len(image_paths) >= 3:
                output_relative_path = f"industrial_plot/videos/auto_{prop.id}.mp4"
                try:
                    result_path = generate_property_slideshow(image_paths, output_relative_path)
                    if result_path:
                        IndustrialPlotResaleVideo.objects.update_or_create(
                            property=prop,
                            source='auto',
                            defaults={
                                'video': result_path,
                                'video_url': None
                            }
                        )
                except Exception:
                    import traceback
                    traceback.print_exc()

            # 2. Save Manual Upload Video as a separate row
            if video_option == 'upload' and uploaded_video:
                IndustrialPlotResaleVideo.objects.create(
                    property=prop,
                    video=uploaded_video,
                    source='uploaded'
                )

            # 3. Save RM Assisted Link Video as a separate row
            elif video_option == 'rm_assisted' and property_video_link:
                IndustrialPlotResaleVideo.objects.create(
                    property=prop,
                    video_url=property_video_link,
                    source='rm_assisted'
                )

            return JsonResponse({
                'status': 'success',
                'message': "Industrial Plot Listing Added Successfully"
            })

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            return JsonResponse({
                'status': 'error',
                'message': f"Error while saving listing: {str(e)}"
            })

    return render(request, 'agent/Reports/Resale_Plot/industrial_plot_resale_list.html', {
        
        'user_obj': user_obj,
       
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'image_category_choices': IndustrialPlotResaleImage.CATEGORY_CHOICES,
    })



def industrial_plot_resale_edit_agent(request, pk):
    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    prop = get_object_or_404(IndustrialPlotResaleProperty, id=pk, is_deleted=False)

    if request.method == 'POST':
        try:
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            def to_date(val):
                if not val:
                    return None
                try:
                    return datetime.strptime(val, "%Y-%m-%d").date()
                except:
                    return None

            amenities = ",".join(request.POST.getlist('amenities[]'))
            nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            input_property_no = (request.POST.get('property_no') or '').strip()
            input_industrial_estate_name = (request.POST.get('industrial_estate_name') or '').strip()
            input_locality = (request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            input_listed_by_id = (request.POST.get('listed_by_id') or '').strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or '').strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or '').strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or '').strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or '').strip()

            # ---------- DUPLICATE CHECK (excluding this property itself) ----------
            fingerprint_key = generate_industrial_plot_fingerprint(
                input_property_no, input_industrial_estate_name, input_locality, input_pincode
            )

            direct_duplicates = IndustrialPlotResaleProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            ).exclude(id=prop.id)
            if input_industrial_estate_name:
                direct_duplicates = direct_duplicates.filter(industrial_estate_name__iexact=input_industrial_estate_name)

            existing_duplicates = (
                IndustrialPlotResaleProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False).exclude(id=prop.id)
                | direct_duplicates
            ).distinct()

            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                if same_id or same_email or same_contact:
                    return JsonResponse({
                        'status': 'error',
                        'message': f"Duplicate Blocked: This plot (No {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}."
                    })

            # ---------- UPDATE FIELDS ----------
            prop.property_unique_key = fingerprint_key

            prop.listed_by_type = request.POST.get('listed_by_type')
            prop.assigned_to = request.POST.get('assigned_to')
            prop.listed_by_id = input_listed_by_id
            prop.listed_by_name = input_listed_by_name
            prop.listed_by_email = input_listed_by_email
            prop.listed_by_contact = input_listed_by_contact
            prop.listed_by_role = input_listed_by_role

            prop.property_title = request.POST.get('property_title')
            prop.property_no = input_property_no

            prop.plot_area = to_decimal(request.POST.get('plot_area'))
            prop.property_type = request.POST.get('property_type')
            prop.land_use = request.POST.get('land_use')
            prop.industrial_zone_type = request.POST.get('industrial_zone_type')
            prop.industrial_estate_name = input_industrial_estate_name
            prop.na_status = request.POST.get('na_status')
            prop.layout_approval_status = request.POST.get('layout_approval_status')
            prop.industrial_fsi = request.POST.get('industrial_fsi')

            prop.plot_frontage = to_decimal(request.POST.get('plot_frontage'))
            prop.plot_depth = to_decimal(request.POST.get('plot_depth'))
            prop.plot_shape = request.POST.get('plot_shape')
            prop.plot_road_facing = request.POST.get('plot_road_facing')
            prop.road_width = request.POST.get('road_width')
            prop.corner_plot = request.POST.get('corner_plot', 'no')

            prop.power_supply = request.POST.get('power_supply')
            prop.power_load_kva = to_int(request.POST.get('power_load_kva'))
            prop.industrial_water_supply = request.POST.get('industrial_water_supply')
            prop.effluent_treatment = request.POST.get('effluent_treatment')
            prop.industry_type_permissible = request.POST.get('industry_type_permissible')
            prop.plot_fencing = request.POST.get('plot_fencing')
            prop.loading_dock = request.POST.get('loading_dock')
            prop.current_possession_status = request.POST.get('current_possession_status')

            prop.selling_price = to_int(request.POST.get('selling_price'))
            prop.price_per_sqft = to_int(request.POST.get('price_per_sqft'))
            prop.price_negotiable = request.POST.get('price_negotiable', 'no')
            prop.additional_charges = request.POST.get('additional_charges')
            prop.brokerage_percentage = request.POST.get('brokerage_percentage')
            prop.manual_brokerage = request.POST.get('manual_brokerage')

            prop.ownership_type = request.POST.get('ownership_type')
            prop.ownership_document_type = request.POST.get('ownership_document_type')
            prop.other_document_type = request.POST.get('other_document_type')
            prop.midc_allotment = request.POST.get('midc_allotment')
            prop.midc_transfer_noc = request.POST.get('midc_transfer_noc')
            prop.environmental_clearance = request.POST.get('environmental_clearance')
            prop.rera_status = request.POST.get('rera_status')
            prop.title_clearance = request.POST.get('title_clearance')
            prop.property_encumbrance_status = request.POST.get('property_encumbrance_status')

            prop.property_tax_status = request.POST.get('property_tax_status')
            prop.outstanding_tax_amount = to_int(request.POST.get('outstanding_tax_amount'))
            prop.pending_since = to_date(request.POST.get('pending_since'))
            prop.property_loan_status = request.POST.get('property_loan_status')
            prop.financing_bank = request.POST.get('financing_bank')
            prop.outstanding_loan_amount = to_int(request.POST.get('outstanding_loan_amount'))
            prop.sanctioning_authority = request.POST.get('sanctioning_authority')

            prop.amenities = amenities
            prop.nearby_facilities = nearby_facilities
            prop.user_description = request.POST.get('user_description')

            prop.state = request.POST.get('state')
            prop.city = request.POST.get('city')
            prop.locality = input_locality
            prop.property_landmark = request.POST.get('property_landmark')
            prop.pincode = input_pincode
            prop.address = request.POST.get('plot_address')
            prop.google_maps_link = request.POST.get('google_maps_link')
            prop.latitude = request.POST.get('plot_latitude')
            prop.longitude = request.POST.get('plot_longitude')

            if request.FILES.get('encumbrance_cert'):
                prop.encumbrance_cert = request.FILES.get('encumbrance_cert')
            if request.FILES.get('layout_plan'):
                prop.layout_plan = request.FILES.get('layout_plan')

            prop.listed_elsewhere = request.POST.get('listed_elsewhere', 'No')
            prop.portal_name = request.POST.get('portal_name')
            # ... other fields ...
            prop.listed_elsewhere = request.POST.get('listed_elsewhere', 'No')
            prop.portal_name = request.POST.get('portal_name')
            
            # --- ADD THESE TWO LINES ---
            prop.listing_status = request.POST.get('listing_status')
            prop.approval_status = request.POST.get('approval_status')

           

            prop.save()

            # ---------- REMOVE IMAGES MARKED FOR DELETION ----------
            remove_image_ids = request.POST.getlist('delete_image_ids[]') or request.POST.getlist('remove_images[]')
            if remove_image_ids:
                IndustrialPlotResaleImage.objects.filter(id__in=remove_image_ids, property=prop).delete()

            # ---------- ADD NEWLY UPLOADED IMAGES (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'full_plot':         'plot_images_full_plot[]',
                'main_entrance':     'plot_images_main_entrance[]',
                'boundary_fencing':  'plot_images_boundary_fencing[]',
                'road_facing':       'plot_images_road_facing[]',
                'approach_road':     'plot_images_approach_road[]',
                'truck_access':      'plot_images_truck_access[]',
                'industrial_estate': 'plot_images_industrial_estate[]',
                'electricity_infra': 'plot_images_electricity_infra[]',
                'water_infra':       'plot_images_water_infra[]',
                'aerial_drone':      'plot_images_aerial_drone[]',
                'layout_site_plan':  'plot_images_layout_site_plan[]',
            }

            existing_count = IndustrialPlotResaleImage.objects.filter(property=prop).count()
            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)

                new_image_objs = []
                if cat_images:
                    start_seq = IndustrialPlotResaleImage.objects.filter(property=prop, category=category).count()
                    for idx, img in enumerate(cat_images):
                        if existing_count + saved_count >= 25:
                            break
                        new_obj = IndustrialPlotResaleImage.objects.create(
                            property=prop,
                            image=img,
                            category=category,
                            sequence_order=start_seq + idx,
                        )
                        new_image_objs.append(new_obj)
                        saved_count += 1

                # Apply the drag-and-drop order (existing + new images mixed),
                # e.g. image_order_full_plot[] = ["existing:12", "new:0", "existing:14", ...]
                order_tokens = request.POST.getlist(f'image_order_{category}[]')
                for position, token in enumerate(order_tokens):
                    if token.startswith('existing:'):
                        img_id = token.split(':', 1)[1]
                        IndustrialPlotResaleImage.objects.filter(
                            id=img_id, property=prop, category=category
                        ).update(sequence_order=position)
                    elif token.startswith('new:'):
                        new_idx = int(token.split(':', 1)[1])
                        if 0 <= new_idx < len(new_image_objs):
                            new_image_objs[new_idx].sequence_order = position
                            new_image_objs[new_idx].save(update_fields=['sequence_order'])

            # ---------- VIDEO ----------
            # Handle Video Deletion
            if request.POST.get('delete_current_video') == '1':
                prop.video.filter(source='uploaded').delete()

            video_option = request.POST.get('video_option') or 'auto'
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '').strip()
            regenerate_slideshow = request.POST.get('regenerate_slideshow') == 'on'

            if uploaded_video:
                # A real video file was attached in this request — always save it,
                # regardless of what the (sometimes stale) radio value says.
                IndustrialPlotResaleVideo.objects.update_or_create(
                    property=prop, source='uploaded',
                    defaults={'video': uploaded_video, 'video_url': None}
                )
            elif video_option == 'rm_assisted' and property_video_link:
                IndustrialPlotResaleVideo.objects.update_or_create(
                    property=prop, source='rm_assisted',
                    defaults={'video_url': property_video_link, 'video': None}
                )
            elif video_option == 'auto':
                existing_auto = prop.video.filter(source='auto').first()
                if regenerate_slideshow or not existing_auto:
                    CATEGORY_ORDER = list(IMAGE_CATEGORY_FIELDS.keys())
                    all_images = list(IndustrialPlotResaleImage.objects.filter(property=prop))
                    all_images.sort(key=lambda img: (
                        CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99,
                        img.sequence_order
                    ))
                    image_paths = [
                        img.image.path for img in all_images
                        if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                    ]
                    if len(image_paths) >= 3:
                        output_relative_path = f"industrial_plot/videos/auto_{prop.id}.mp4"
                        try:
                            result_path = generate_property_slideshow(image_paths, output_relative_path)
                            if result_path:
                                IndustrialPlotResaleVideo.objects.update_or_create(
                                    property=prop, source='auto',
                                    defaults={'video': result_path, 'video_url': None}
                                )
                        except Exception:
                            import traceback
                            traceback.print_exc()
            return JsonResponse({
    'status': 'success',
    'message': "Industrial Plot Listing Updated Successfully",
    "redirect_url": reverse('industrial_plot_resale_list_agent')
})
        except Exception as e:
            print("ERROR DETECTED:", str(e))
            return JsonResponse({
                'status': 'error',
                'message': f"Error while updating listing: {str(e)}"
            })

    # ---------- GET: render prefilled form ----------
    existing_images = IndustrialPlotResaleImage.objects.filter(property=prop).order_by('category', 'sequence_order')
    images_by_category = {}
    for img in existing_images:
        images_by_category.setdefault(img.category, []).append(img)
        
    existing_image_total = existing_images.count()

    # Query existing videos to pass to template
    uploaded_video = prop.video.filter(source='uploaded').first()
    auto_video = prop.video.filter(source='auto').first()
    rm_video = prop.video.filter(source='rm_assisted').first()

    return render(request, 'agent/Reports/Resale_Plot/industrial_plot_resale_edit.html', {
       
        'user_obj': user_obj,
        
        'property': prop,
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'images_by_category': images_by_category,
        'existing_image_total': existing_image_total,
        'uploaded_video': uploaded_video,
        'auto_video': auto_video,
        'rm_video': rm_video,
        'selected_amenities': (prop.amenities or '').split(',') if prop.amenities else [],
        'selected_facilities': (prop.nearby_facilities or '').split(',') if prop.nearby_facilities else [],
    })



def industrial_plot_resale_view_agent(request, pk):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    # UPDATED: We use prefetch_related on the base queryset so Django fetches 
    # all images, videos, and FAQs in one go before returning the 404 check.
    queryset = IndustrialPlotResaleProperty.objects.prefetch_related('images', 'video', 'faqs')
    plot = get_object_or_404(queryset, id=pk, is_deleted=False)
 
    if not plot.faqs.exists():
        plot.generate_auto_faqs()
 
    amenities_list = [a.strip() for a in plot.amenities.split(',')] if plot.amenities else []
    facilities_list = [f.strip() for f in plot.nearby_facilities.split(',')] if plot.nearby_facilities else []
 
    # Price/sq.m — precomputed here rather than doing float math in the template
    price_per_sqm = None
    if plot.selling_price and plot.plot_area and plot.plot_area > 0:
        try:
            price_per_sqm = round(float(plot.selling_price) / float(plot.plot_area))
        except (TypeError, ZeroDivisionError):
            price_per_sqm = None

    # Extracted explicitly so you can use {% for vid in videos %} in your template
    # Extract videos and get the first one for the preview
    videos = plot.video.all()
    selected_video = videos.first() if videos.exists() else None

    # Get all images and sort them
    images = plot.images.all().order_by('category', 'sequence_order')
    
    # Group images by category for the filtering tabs
    grouped_images = {}
    for img in images:
        cat = img.category
        if cat not in grouped_images:
            # We use get_category_display() if it exists to show the readable label
            grouped_images[cat] = {
                'label': img.get_category_display() if hasattr(img, 'get_category_display') else cat.replace('_', ' ').title(),
                'images': []
            }
        grouped_images[cat]['images'].append(img)
 
    return render(request, 'agent/Reports/Resale_Plot/industrial_plot_resale_view.html', {
        'user_obj': user_obj,
        
        'plot': plot,
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'price_per_sqm': price_per_sqm,
        'images': images,
        'grouped_images': grouped_images,
        'selected_video': selected_video,
    })


def industrial_plot_resale_list_agent(request):

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    
    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    # UPDATED: Added prefetch_related to load images and videos in a single query
    qs = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).prefetch_related('images', 'video').order_by("-created_at")

    # ---- Search & Filters ----
    search_query = request.GET.get("search", "").strip()
    city_query = request.GET.get("city", "").strip()
    locality_query = request.GET.get("locality", "").strip()
    property_type_query = request.GET.get("property_type", "").strip()
    road_facing_query = request.GET.get("road_facing", "").strip()
    corner_plot_query = request.GET.get("corner_plot", "").strip()
    loan_query = request.GET.get("loan", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()
    
    # NEW FILTERS
    listed_by_query = request.GET.get("listed_by", "").strip()
    uploaded_by_query = request.GET.get("uploaded_by", "").strip()
    listing_status_query = request.GET.get("listing_status", "").strip()
    approval_status_query = request.GET.get("approval_status", "").strip()
    duplicate_query = request.GET.get("duplicate", "").strip()

    # Apply Filters
    if listed_by_query and listed_by_query != 'All Roles':
        qs = qs.filter(listed_by_role__iexact=listed_by_query)
    
    if uploaded_by_query and uploaded_by_query != 'All Roles':
        qs = qs.filter(uploaded_by_role__iexact=uploaded_by_query)
        
    if listing_status_query and listing_status_query != 'All Status':
        qs = qs.filter(listing_status__iexact=listing_status_query)
        
    if approval_status_query and approval_status_query != 'All Approval':
        qs = qs.filter(approval_status__iexact=approval_status_query)
        
    if duplicate_query == 'duplicates_only':
        qs = qs.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        qs = qs.filter(is_duplicate=False)

    if search_query:
        qs = qs.filter(
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(property_no__icontains=search_query)
        )

    if city_query:
        qs = qs.filter(city__icontains=city_query)

    if locality_query:
        qs = qs.filter(locality__icontains=locality_query)

    if property_type_query:
        qs = qs.filter(property_type=property_type_query)

    if road_facing_query:
        qs = qs.filter(plot_road_facing__icontains=road_facing_query)

    if corner_plot_query in ("yes", "no"):
        qs = qs.filter(corner_plot__iexact=corner_plot_query)

    if loan_query == "yes":
        qs = qs.filter(property_loan_status__iexact="Loan Running")
    elif loan_query == "no":
        qs = qs.exclude(property_loan_status__iexact="Loan Running")

    if min_price:
        try:
            qs = qs.filter(selling_price__gte=int(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            qs = qs.filter(selling_price__lte=int(max_price))
        except ValueError:
            pass

    if from_date:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(from_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    if to_date:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(to_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    total_properties = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).count()
    filtered_count = qs.count()

    # ---- KPI Cards (Inventory & Plot Features) ----
    active_listings = qs.filter(selling_price__isnull=False).count()
    midc_count = qs.filter(property_type="midc_industrial").count()
    warehouse_count = qs.filter(property_type="warehouse_logistics").count()
    sez_count = qs.filter(property_type="sez_plot").count()

    corner_plot_count = qs.filter(corner_plot__iexact="yes").count()
    fenced_plot_count = qs.exclude(
        Q(plot_fencing__isnull=True) | Q(plot_fencing="") | Q(plot_fencing="none")
    ).count()
    finance_ready_count = qs.filter(property_loan_status__iexact="No Active Loan").count()

    # ---- KPI Cards (Listing, Approval & Duplicates) ----
    active_listing_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Active").count()
    inactive_listing_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Inactive").count()
    sold_listing_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Sold").count()
    rented_listing_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Rented").count()

    pending_approval_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False, approval_status__iexact="Pending").count()
    approved_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, approval_status__iexact="Approved").count()
    rejected_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, approval_status__iexact="Rejected").count()

    duplicate_properties_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, is_duplicate=True).count()
    unique_properties_count = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, is_duplicate=False).count()


    # ---- Pagination ----
    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    # give each row a stable serial number across pages
    start_index = (page_obj.number - 1) * paginator.per_page
    for i, obj in enumerate(page_obj.object_list, start=1):
        obj.original_sr_no = start_index + i

    # ---- Dropdown option sources ----
    unique_cities = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(city__isnull=True).exclude(city="").values_list("city", flat=True).distinct().order_by("city")
    unique_property_types = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(property_type__isnull=True).exclude(property_type="").values_list("property_type", flat=True).distinct().order_by("property_type")
    uploaded_files = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(upload_file_name__isnull=True).exclude(upload_file_name="").values_list("upload_file_name", flat=True).distinct()
    
    unique_listed_roles = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(listed_by_role__isnull=True).exclude(listed_by_role="").values_list("listed_by_role", flat=True).distinct().order_by("listed_by_role")
    unique_uploaded_roles = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role="").values_list("uploaded_by_role", flat=True).distinct().order_by("uploaded_by_role")
    unique_listing_status = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(listing_status__isnull=True).exclude(listing_status="").values_list("listing_status", flat=True).distinct().order_by("listing_status")
    unique_approval_status = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(approval_status__isnull=True).exclude(approval_status="").values_list("approval_status", flat=True).distinct().order_by("approval_status")
    
    # ---- Ensure Context dictionary has ALL variables ----
    context = {
        "properties": page_obj,
        "page_obj": page_obj,
        "total_properties": total_properties,
        "filtered_count": filtered_count,
        "active_listings": active_listings,
        "midc_count": midc_count,
        "warehouse_count": warehouse_count,
        "sez_count": sez_count,
        "corner_plot_count": corner_plot_count,
        "fenced_plot_count": fenced_plot_count,
        "finance_ready_count": finance_ready_count,
        
        # New Status KPI Counts
        "active_listing_count": active_listing_count,
        "inactive_listing_count": inactive_listing_count,
        "sold_listing_count": sold_listing_count,
        "rented_listing_count": rented_listing_count,
        "pending_approval_count": pending_approval_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "duplicate_properties_count": duplicate_properties_count,
        "unique_properties_count": unique_properties_count,

        # Dropdown Options
        "unique_cities": unique_cities,
        "unique_property_types": unique_property_types,
        "uploaded_files": uploaded_files,
        "unique_listed_roles": unique_listed_roles,
        "unique_uploaded_roles": unique_uploaded_roles,
        "unique_listing_status": unique_listing_status,
        "unique_approval_status": unique_approval_status,
        
        # Selected Queries
        "search_query": search_query,
        "city_query": city_query,
        "locality_query": locality_query,
        "property_type_query": property_type_query,
        "road_facing_query": road_facing_query,
        "corner_plot_query": corner_plot_query,
        "loan_query": loan_query,
        "min_price": min_price,
        "max_price": max_price,
        "listed_by_query": listed_by_query,
        "uploaded_by_query": uploaded_by_query,
        "listing_status_query": listing_status_query,
        "approval_status_query": approval_status_query,
        "duplicate_query": duplicate_query,
        
        # Base Objects
        
        'ameneties_obj': ameneties_obj,
        'facilities_obj': facilities_obj,
        'user_obj': user_obj,
       
    }
    return render(request, "agent/Reports/Resale_Plot/industrial_plot_resale_list.html", context)





@require_POST
def industrial_plot_resale_delete_agent(request, pk):
    """Agent Soft Delete Industrial Plot Resale Property (supports Admin impersonation)"""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Fetch Property
        # ======================================
        plot = get_object_or_404(IndustrialPlotResaleProperty, id=pk)

        # ======================================
        # 6. Security Check — agent can delete only own property
        # ======================================
        if str(plot.listed_by_id) != str(agent_obj.user_id):
            return JsonResponse({
                'status': 'error',
                'message': 'You can delete only your own properties.'
            })

        # ======================================
        # 7. Soft Delete
        # ======================================
        plot.is_deleted = True
        plot.deleted_at = timezone.now()
        plot.deleted_by = user_identity
        plot.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])

        IndustrialPlotResaleActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type="DELETE",
            property_id=plot.id,
            action_payload=json.dumps({"reason": "Manual delete via Agent list page"}),
            status="SUCCESS",
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Property {plot.id} moved to recycle bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



@require_POST
def industrial_plot_resale_bulk_delete_agent(request):
    """Agent Advanced Bulk Delete (Soft Delete) — scoped to agent's own industrial plots, supports Admin impersonation."""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Parse Payload
        # ======================================
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (ValueError, UnicodeDecodeError):
            return JsonResponse({"status": "error", "message": "Invalid payload."}, status=400)

        delete_type = payload.get("delete_type")

        # ======================================
        # 6. Base Queryset — SCOPED TO THIS AGENT ONLY
        # ======================================
        qs = IndustrialPlotResaleProperty.objects.filter(
            is_deleted=False,
            listed_by_id=agent_obj.user_id
        )

        if delete_type == "current_page":
            ids = payload.get("page_ids", [])
            qs = qs.filter(id__in=ids)

        elif delete_type == "date_range":
            from_date = payload.get("from_date")
            to_date = payload.get("to_date")
            if not from_date or not to_date:
                return JsonResponse({"status": "error", "message": "Both dates are required."}, status=400)
            qs = qs.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)

        elif delete_type == "latest_month":
            today = timezone.now().date()
            first_of_month = today.replace(day=1)
            qs = qs.filter(created_at__date__gte=first_of_month)

        elif delete_type == "old_data":
            cutoff = timezone.now() - timedelta(days=180)
            qs = qs.filter(created_at__lt=cutoff)

        elif delete_type == "by_uploader":
            text = payload.get("uploader_text", "").strip()
            if not text:
                return JsonResponse({"status": "error", "message": "Uploader detail is required."}, status=400)
            qs = qs.filter(
                Q(uploaded_by_name__icontains=text) |
                Q(uploaded_by_email__icontains=text) |
                Q(uploaded_by_role__icontains=text)
            )

        elif delete_type == "by_file":
            file_name = payload.get("file_name", "").strip()
            if not file_name:
                return JsonResponse({"status": "error", "message": "File name is required."}, status=400)
            qs = qs.filter(upload_file_name=file_name)

        elif delete_type == "delete_all":
            pass  # qs already scoped to this agent's non-deleted rows

        else:
            return JsonResponse({"status": "error", "message": "Unknown delete criteria."}, status=400)

        count = qs.count()
        if count == 0:
            return JsonResponse({"status": "error", "message": "No matching records found to delete."}, status=400)

        qs.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=user_identity)

        IndustrialPlotResaleActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type="DELETE",
            property_id="Multiple / Bulk Action",
            action_payload=json.dumps({"delete_type": delete_type, "records_deleted": count}),
            status="SUCCESS",
        )

        return JsonResponse({"status": "success", "message": f"{count} record(s) deleted successfully."})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})





def export_industrial_plot_resale_agent(request):

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    # ── 1. Re-apply the same search filters so the export matches the screen ──
    try:
        properties = IndustrialPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')
    except Exception:
        properties = IndustrialPlotResaleProperty.objects.all(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    search_query        = request.GET.get("search", "").strip()
    city_query          = request.GET.get("city", "").strip()
    locality_query      = request.GET.get("locality", "").strip()
    property_type_query = request.GET.get("property_type", "").strip()
    road_facing_query   = request.GET.get("road_facing", "").strip()
    corner_plot_query   = request.GET.get("corner_plot", "").strip()
    loan_query          = request.GET.get("loan", "").strip()
    min_price           = request.GET.get("min_price", "").strip()
    max_price           = request.GET.get("max_price", "").strip()
    from_date           = request.GET.get("from_date", "").strip()
    to_date             = request.GET.get("to_date", "").strip()
    listed_by_query     = request.GET.get("listed_by", "").strip()
    uploaded_by_query   = request.GET.get("uploaded_by", "").strip()
    listing_status_query= request.GET.get("listing_status", "").strip()
    approval_status_query= request.GET.get("approval_status", "").strip()
    duplicate_query     = request.GET.get("duplicate", "").strip()

    # Apply Filters
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)
    
    if uploaded_by_query and uploaded_by_query != 'All Roles':
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)
        
    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)
        
    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)
        
    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if search_query:
        properties = properties.filter(
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(property_no__icontains=search_query)
        )

    if city_query:
        properties = properties.filter(city__icontains=city_query)

    if locality_query:
        properties = properties.filter(locality__icontains=locality_query)

    if property_type_query:
        properties = properties.filter(property_type=property_type_query)

    if road_facing_query:
        properties = properties.filter(plot_road_facing__icontains=road_facing_query)

    if corner_plot_query in ("yes", "no"):
        properties = properties.filter(corner_plot__iexact=corner_plot_query)

    if loan_query == "yes":
        properties = properties.filter(property_loan_status__iexact="Loan Running")
    elif loan_query == "no":
        properties = properties.exclude(property_loan_status__iexact="Loan Running")

    if min_price:
        try:
            properties = properties.filter(selling_price__gte=int(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            properties = properties.filter(selling_price__lte=int(max_price))
        except ValueError:
            pass

    if from_date:
        try:
            properties = properties.filter(created_at__date__gte=datetime.strptime(from_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    if to_date:
        try:
            properties = properties.filter(created_at__date__lte=datetime.strptime(to_date, "%Y-%m-%d").date())
        except ValueError:
            pass


  
    EXPORT_COLS = [
        ("Sr.No", "sr_no", False, "Sr. No"),
        ("Listed By", "listed_by_type", False, "Listed By Type (Self/Other)"),
        ("Listed By", "listed_by_id", False, "Listed By Id"),
        ("Listed By", "listed_by_name", False, "Listed By Name"),
        ("Listed By", "listed_by_email", False, "Listed By Email"),
        ("Listed By", "listed_by_contact", False, "Listed By Contact"),
        ("Listed By", "listed_by_role", True, "Listed By Role"),

        ("System Generated Property Id(Auto)", "id", False, "Property Id"),
        ("Basic Industrial Plot Information", "property_title", False, "Property Title"),
        ("Basic Industrial Plot Information", "property_no", True, "Plot / Survey Number"),
        ("Basic Industrial Plot Information", "plot_area", True, "Plot Area (sq.m)"),
        ("Basic Industrial Plot Information", "property_type", True, "Industrial Plot Type"),

        ("Industrial Zone, Estate & Authority Details", "land_use", True, "Industrial Zone / Land Use"),
        ("Industrial Zone, Estate & Authority Details", "industrial_zone_type", True, "Industrial Zone Type"),
        ("Industrial Zone, Estate & Authority Details", "industrial_estate_name", False, "Industrial Estate / Park Name"),
        ("Industrial Zone, Estate & Authority Details", "na_status", True, "NA Status / Industrial Conversion"),
        ("Industrial Zone, Estate & Authority Details", "layout_approval_status", False, "Industrial Authority / Estate Approval"),
        ("Industrial Zone, Estate & Authority Details", "industrial_fsi", False, "Permissible FSI / Coverage"),

        ("Industrial Plot Specifications & Infrastructure", "plot_frontage", False, "Plot Frontage / Width (m)"),
        ("Industrial Plot Specifications & Infrastructure", "plot_depth", False, "Depth / Length (m)"),
        ("Industrial Plot Specifications & Infrastructure", "plot_shape", False, "Plot Shape"),
        ("Industrial Plot Specifications & Infrastructure", "plot_road_facing", True, "Road Access & Entry"),
        ("Industrial Plot Specifications & Infrastructure", "road_width", False, "Road Width (Front of Plot)"),
        ("Industrial Plot Specifications & Infrastructure", "corner_plot", False, "Corner Plot (yes/no)"),
        ("Industrial Plot Specifications & Infrastructure", "power_supply", False, "Power Supply Available"),
        ("Industrial Plot Specifications & Infrastructure", "power_load_kva", False, "Power Load Sanctioned (KVA)"),
        ("Industrial Plot Specifications & Infrastructure", "industrial_water_supply", False, "Water Supply"),
        ("Industrial Plot Specifications & Infrastructure", "effluent_treatment", False, "Drain / Effluent Facility"),
        ("Industrial Plot Specifications & Infrastructure", "industry_type_permissible", False, "Permissible Industry Type"),
        ("Industrial Plot Specifications & Infrastructure", "plot_fencing", False, "Compound Wall / Security Fencing"),
        ("Industrial Plot Specifications & Infrastructure", "loading_dock", False, "Loading / Unloading Facility"),
        ("Industrial Plot Specifications & Infrastructure", "current_possession_status", False, "Current Plot / Shed Status"),

        ("Pricing Details", "selling_price", True, "Selling Price"),
        ("Pricing Details", "price_per_sqft", False, "Price per Sq.m(Auto-calculated)"),
        ("Pricing Details", "price_negotiable", False, "Is the Price Negotiable (Yes/No)"),
        ("Pricing Details", "additional_charges", False, "Additional Industrial Charges"),
        ("Pricing Details", "brokerage_percentage", True, "Brokerage / Service Fee"),
        ("Pricing Details", "manual_brokerage", False, "Fixed Brokerage Amount"),

        ("Legal, Title & Industrial Authority Details", "ownership_type", False, "Ownership / Tenure Type"),
        ("Legal, Title & Industrial Authority Details", "ownership_document_type", False, "Primary Title / Ownership Document"),
        ("Legal, Title & Industrial Authority Details", "other_document_type", False, "Specify Other Ownership Document Type"),
        ("Legal, Title & Industrial Authority Details", "midc_allotment", False, "MIDC / Authority Allotment Letter"),
        ("Legal, Title & Industrial Authority Details", "midc_transfer_noc", False, "MIDC Transfer / NOC Status"),
        ("Legal, Title & Industrial Authority Details", "environmental_clearance", False, "Environmental Clearance (EC)"),
        ("Legal, Title & Industrial Authority Details", "rera_status", False, "RERA / Industrial Authority Status"),
        ("Legal, Title & Industrial Authority Details", "title_clearance", False, "Title Clarity / Marketability"),
        ("Legal, Title & Industrial Authority Details", "property_encumbrance_status", False, "Encumbrance / Charge Status"),
        ("Legal, Title & Industrial Authority Details", "property_tax_status", False, "Industrial Property Tax / Assessment Status"),
        ("Legal, Title & Industrial Authority Details", "outstanding_tax_amount", False, "Outstanding Tax / Dues Amount"),
        ("Legal, Title & Industrial Authority Details", "pending_since", False, "Dues Pending Since"),
        ("Legal, Title & Industrial Authority Details", "property_loan_status", False, "Existing Loan / Mortgage on Property"),
        ("Legal, Title & Industrial Authority Details", "financing_bank", False, "Lender Bank / NBFC Name"),
        ("Legal, Title & Industrial Authority Details", "outstanding_loan_amount", False, "Outstanding Loan Amount"),
        ("Legal, Title & Industrial Authority Details", "sanctioning_authority", False, "Industrial Authority / Estate Developer"),

        ("Amenities & Facilities", "amenities", False, "Amenities (comma-separated)"),
        ("Nearby Facilities", "nearby_facilities", False, "Nearby Facilities (comma-separated)"),
        
        ("Property Descriptions(Added By User)", "user_description", False, "Property Description"),
        ("Property Summary(Auto)", "property_summary", False, "Property Summary"),
        ("Property Description(Auto)", "property_description", False, "Property Description"),


        ("Property Location Details", "address", True, "Complete Industrial Plot Address"),
        ("Property Location Details", "locality", True, "MIDC Phase / Industrial Area Name"),
        ("Property Location Details", "property_landmark", False, "Nearest Highway / Railway / Port Reference"),
        ("Property Location Details", "city", True, "City / District"),
        ("Property Location Details", "state", True, "State"),
        ("Property Location Details", "pincode", False, "PIN Code"),
        ("Property Location Details", "google_maps_link", False, "Google Maps Link"),
        ("Property Location Details", "latitude", False, "Latitude"),
        ("Property Location Details", "longitude", False, "Longitude"),

        ("Property Listed Elsewhere", "listed_elsewhere", False, "Listed Elsewhere (Yes/No)"),
        ("Property Listed Elsewhere", "portal_name", False, "Portal Name"),


      
        ("Media & Listing Status", "listing_status", False, "Listing Status"),
        ("Media & Listing Status", "approval_status", False, "Approval Status"),

        ("Data Uploadeded Via", "upload_file_name", False, "Upload File Name"),
        ("Property Uploaded By", "uploaded_by_name", False, "Uploaded By Name"),
        ("Property Uploaded By", "uploaded_by_email", False, "Uploaded By Email"),
        ("Property Uploaded By", "uploaded_by_contact", False, "Uploaded By Contact"),
        ("Property Uploaded By", "uploaded_by_role", False, "Uploaded By Role"),

        ("Database Audit", "created_at", False, "Created At"),
        ("Database Audit", "updated_at", False, "Updated At"),
        ("Database Audit", "is_deleted", False, "Is Deleted"),
        ("Database Audit", "deleted_at", False, "Deleted At"),
        ("Database Audit", "deleted_by", False, "Deleted By"),
        ("Database Audit", "is_duplicate", False, "Is Duplicate"),
        ("Database Audit", "duplicate_count", False, "Duplicate Count"),
        ("Database Audit", "duplicate_group_id", False, "Duplicate Group ID"),
        ("Database Audit", "property_unique_key", False, "Property Unique Key"),

        ("Brokerage Label", "get_brokerage_label", False, "Brokerage Label Preview (auto)"),
    ]

    export_format = request.GET.get('format', 'excel')
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Industrial_Plot_Resale_Full_Export_{timestamp}"

    # Helper for formatting values
    def format_val(val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(val, bool):
            return "Yes" if val else "No"
        elif val is None:
            return ""
        return str(val).strip()

    # ── 3. EXCEL EXPORT ──
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Industrial Plot Resale DB"

        HDR_BG  = "667EEA"
        REQ_BG  = "FEF3C7"
        OPT_BG  = "F0FDF4"
        thin = Side(style="thin", color="CBD5E1")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sec_spans = OrderedDict()
        for i, (sec, *_) in enumerate(EXPORT_COLS):
            sec_spans.setdefault(sec, []).append(i + 1)

        # Write Section Headers (Row 1)
        for sec, cols in sec_spans.items():
            c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
            
            for col_idx in cols[1:]:
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor=HDR_BG)
                ws.cell(row=1, column=col_idx).border = bdr
                
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

        # Write Field Headers (Row 2)
        for ci, (_, _, req, header_name) in enumerate(EXPORT_COLS, 1):
            lc = ws.cell(row=2, column=ci, value=header_name)
            lc.font = Font(bold=True, size=9)
            lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
            lc.border = bdr
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(header_name) + 5)

        # Write Data (Rows 3+)
        for row_idx, prop in enumerate(properties, start=3):
            for col_idx, (_, field, _, _) in enumerate(EXPORT_COLS, 1):
                
                # <-- ADDED SR_NO LOGIC HERE
                if field == "sr_no":
                    val = row_idx - 2
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                elif field == "get_brokerage_label":
                    val = format_val(getattr(prop, field)()) if callable(getattr(prop, field, None)) else ""
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center")

        # Layout adjustments
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # ── 4. CSV EXPORT ──
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        
        # Write exact Field Headers
        headers = [header_name for (_, _, _, header_name) in EXPORT_COLS]
        writer.writerow(headers)

        # Write Data Rows
        for row_idx, prop in enumerate(properties, start=1):
            row_data = []
            for (_, field, _, _) in EXPORT_COLS:
                
                # <-- ADDED SR_NO LOGIC HERE
                if field == "sr_no":
                    val = row_idx
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                elif field == "get_brokerage_label":
                    val = format_val(getattr(prop, field)()) if callable(getattr(prop, field, None)) else ""
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                row_data.append(val)
            writer.writerow(row_data)

        return response


############# Views End for Resale Plot Industrial Module for Agent ######################################




############# Views Start for Resale Plot Agriculture Module for Agent ######################################



def agricultural_plot_resale_agent(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agent/Forms/Resale_Plot/agricultural_plot_resale.html", context)

    

def generate_agricultural_plot_fingerprint(property_no, locality, city, pincode):
    """Fingerprint used ONLY by the bulk-import duplicate engine (mirrors
    industrial's generate_industrial_plot_fingerprint). Distinct from the
    model's own save()-time property_unique_key, which keys on
    address/locality/city/plot_area/property_no."""
    key_source = f"{property_no}|{locality}|{city}|{pincode}"
    return key_source.strip().lower().replace(" ", "")



def agricultural_plot_resale_add_agent(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            def to_date(val):
                if not val:
                    return None
                try:
                    return datetime.strptime(val, "%Y-%m-%d").date()
                except:
                    return None

            # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION ----------
            
            if user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "Agent"
                
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_locality = (request.POST.get('locality') or '').strip()
            input_city = (request.POST.get('city') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_agricultural_plot_fingerprint(
                input_property_no,
                input_locality,
                input_city,
                input_pincode
            )

            direct_duplicates = AgriculturalPlotResaleProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            )
            if input_city:
                direct_duplicates = direct_duplicates.filter(city__iexact=input_city)

            existing_duplicates = (
                AgriculturalPlotResaleProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                    if same_id or same_email or same_contact:
                        return JsonResponse({
                            'status': 'error',
                            'message': f"Duplicate Blocked: This land parcel ({input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        })

                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True,
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = AgriculturalPlotResaleProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Resale",
                category="Plot",
                sub_category=request.POST.get('sub_category', 'Agricultural'),

                

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_no=input_property_no,

                plot_area=to_decimal(request.POST.get('plot_area')),
                agr_area_unit=request.POST.get('agr_area_unit', 'acre'),
                property_type=request.POST.get('property_type'),
                land_use=request.POST.get('land_use'),
                na_status=request.POST.get('na_status'),
                layout_approval_status=request.POST.get('layout_approval_status'),

                plot_frontage=to_decimal(request.POST.get('plot_frontage')),
                plot_depth=to_decimal(request.POST.get('plot_depth')),
                plot_shape=request.POST.get('plot_shape'),
                plot_road_facing=request.POST.get('plot_road_facing'),
                road_width=request.POST.get('road_width'),
                corner_plot=request.POST.get('corner_plot', 'no'),

                soil_type=request.POST.get('soil_type'),
                current_crop=request.POST.get('current_crop'),
                irrigation_source=request.POST.get('irrigation_source'),
                agr_electricity=request.POST.get('agr_electricity'),
                highway_distance=request.POST.get('highway_distance'),
                land_topography=request.POST.get('land_topography'),
                govt_scheme=request.POST.get('govt_scheme'),
                plot_fencing=request.POST.get('plot_fencing'),
                current_possession_status=request.POST.get('current_possession_status'),

                selling_price=to_int(request.POST.get('selling_price')),
                price_negotiable=request.POST.get('price_negotiable', 'no'),
                additional_charges=request.POST.get('additional_charges'),
                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                ownership_type=request.POST.get('ownership_type'),
                ownership_document_type=request.POST.get('ownership_document_type'),
                other_document_type=request.POST.get('other_document_type'),
                rera_status=request.POST.get('rera_status'),
                title_clearance=request.POST.get('title_clearance'),
                property_encumbrance_status=request.POST.get('property_encumbrance_status'),

                satbara_available=request.POST.get('satbara_available'),
                khate_utara=request.POST.get('khate_utara'),
                section63_clearance=request.POST.get('section63_clearance'),

                property_tax_status=request.POST.get('property_tax_status'),
                outstanding_tax_amount=to_int(request.POST.get('outstanding_tax_amount')),
                pending_since=to_date(request.POST.get('pending_since')),
                property_loan_status=request.POST.get('property_loan_status'),
                financing_bank=request.POST.get('financing_bank'),
                outstanding_loan_amount=to_int(request.POST.get('outstanding_loan_amount')),
                sanctioning_authority=request.POST.get('sanctioning_authority'),

                amenities=amenities,
                nearby_facilities=nearby_facilities,
                user_description=request.POST.get('user_description'),

                state=request.POST.get('state'),
                city=input_city,
                locality=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                pincode=input_pincode,
                address=request.POST.get('address'),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                encumbrance_cert=request.FILES.get('encumbrance_cert'),
                layout_plan=request.FILES.get('layout_plan'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None,
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'front_view':         'plot_images_front_view[]',
                'full_plot':          'plot_images_full_plot[]',
                'farm_gate':          'plot_images_farm_gate[]',
                'boundary_fencing':   'plot_images_boundary_fencing[]',
                'road_facing':        'plot_images_road_facing[]',
                'cultivated_area':    'plot_images_cultivated_area[]',
                'irrigation_source':  'plot_images_irrigation_source[]',
                'borewell_well':      'plot_images_borewell_well[]',
                'electricity_infra':  'plot_images_electricity_infra[]',
                'farmhouse_shed':     'plot_images_farmhouse_shed[]',
                'aerial_drone':       'plot_images_aerial_drone[]',
                'layout_site_plan':   'plot_images_layout_site_plan[]',
            }

            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 25:
                        break
                    AgriculturalPlotResaleImage.objects.create(
                        property=prop,
                        image=img,
                        category=category,
                        sequence_order=cat_index,
                    )
                    saved_count += 1

            # ---------- PLOT VIDEO (UPLOAD, RM LINK, OR AUTO SLIDESHOW) ----------
            video_option = request.POST.get('video_option') or request.POST.get('video_source') or 'auto'
            uploaded_video = request.FILES.get('property_video') or request.FILES.get('social_video')
            property_video_link = request.POST.get('property_video_link', '').strip()

            # 1. ALWAYS auto-generate the slideshow row if >= 3 photos exist
            CATEGORY_ORDER = list(IMAGE_CATEGORY_FIELDS.keys())
            saved_images = list(AgriculturalPlotResaleImage.objects.filter(property=prop))
            saved_images.sort(key=lambda img: (CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, img.sequence_order))
            image_paths = [img.image.path for img in saved_images if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)]

            if len(image_paths) >= 3:
                output_relative_path = f"agricultural_plot/videos/auto_{prop.id}.mp4"
                try:
                    result_path = generate_property_slideshow(image_paths, output_relative_path)
                    if result_path:
                        AgriculturalPlotResaleVideo.objects.update_or_create(
                            property=prop,
                            source='auto',
                            defaults={
                                'video': result_path,
                                'video_url': None
                            }
                        )
                except Exception:
                    import traceback
                    traceback.print_exc()

            # 2. Save Manual Upload Video as a separate row
            if video_option == 'upload' and uploaded_video:
                AgriculturalPlotResaleVideo.objects.create(
                    property=prop,
                    video=uploaded_video,
                    source='uploaded'
                )

            # 3. Save RM Assisted Link Video as a separate row
            elif video_option == 'rm_assisted' and property_video_link:
                AgriculturalPlotResaleVideo.objects.create(
                    property=prop,
                    video_url=property_video_link,
                    source='rm_assisted'
                )

            return JsonResponse({
                'status': 'success',
                'message': "Agricultural Plot Listing Added Successfully"
            })

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            return JsonResponse({
                'status': 'error',
                'message': f"Error while saving listing: {str(e)}"
            })

    return render(request, 'agent/Reports/Resale_Plot/agricultural_plot_resale_list.html', {
        
        'user_obj': user_obj,
        
      
    
    
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'image_category_choices': AgriculturalPlotResaleImage.CATEGORY_CHOICES,
    })



def agricultural_plot_resale_edit_agent(request, pk):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    prop = get_object_or_404(AgriculturalPlotResaleProperty, id=pk, is_deleted=False)

    if request.method == 'POST':
        try:
            def to_int(val):
                try: return int(val) if val else None
                except: return None

            def to_decimal(val):
                try: return float(val) if val else None
                except: return None

            def to_date(val):
                if not val: return None
                try: return datetime.strptime(val, "%Y-%m-%d").date()
                except: return None

            amenities = ",".join(request.POST.getlist('amenities[]'))
            nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            input_property_no = (request.POST.get('property_no') or '').strip()
            input_locality = (request.POST.get('locality') or '').strip()
            input_city = (request.POST.get('city') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            input_listed_by_id = (request.POST.get('listed_by_id') or '').strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or '').strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or '').strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or '').strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or '').strip()

            # ---------- DUPLICATE CHECK (excluding this property) ----------
            fingerprint_key = generate_agricultural_plot_fingerprint(
                input_property_no, input_locality, input_city, input_pincode
            )

            direct_duplicates = AgriculturalPlotResaleProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            ).exclude(id=prop.id)
            if input_city:
                direct_duplicates = direct_duplicates.filter(city__iexact=input_city)

            existing_duplicates = (
                AgriculturalPlotResaleProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False).exclude(id=prop.id)
                | direct_duplicates
            ).distinct()

            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                if same_id or same_email or same_contact:
                    return JsonResponse({
                        'status': 'error',
                        'message': f"Duplicate Blocked: This land ({input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}."
                    })

            # ---------- UPDATE FIELDS ----------
            prop.property_unique_key = fingerprint_key

            prop.listed_by_type = request.POST.get('listed_by_type')
            prop.assigned_to = request.POST.get('assigned_to')
            prop.listed_by_id = input_listed_by_id
            prop.listed_by_name = input_listed_by_name
            prop.listed_by_email = input_listed_by_email
            prop.listed_by_contact = input_listed_by_contact
            prop.listed_by_role = input_listed_by_role

            prop.property_no = input_property_no

            prop.plot_area = to_decimal(request.POST.get('plot_area'))
            prop.agr_area_unit = request.POST.get('agr_area_unit', 'acre')
            prop.property_type = request.POST.get('property_type')
            prop.land_use = request.POST.get('land_use')
            prop.na_status = request.POST.get('na_status')
            prop.layout_approval_status = request.POST.get('layout_approval_status')

            prop.plot_frontage = to_decimal(request.POST.get('plot_frontage'))
            prop.plot_depth = to_decimal(request.POST.get('plot_depth'))
            prop.plot_shape = request.POST.get('plot_shape')
            prop.plot_road_facing = request.POST.get('plot_road_facing')
            prop.road_width = request.POST.get('road_width')
            prop.corner_plot = request.POST.get('corner_plot', 'no')

            prop.soil_type = request.POST.get('soil_type')
            prop.current_crop = request.POST.get('current_crop')
            prop.irrigation_source = request.POST.get('irrigation_source')
            prop.agr_electricity = request.POST.get('agr_electricity')
            prop.highway_distance = request.POST.get('highway_distance')
            prop.land_topography = request.POST.get('land_topography')
            prop.govt_scheme = request.POST.get('govt_scheme')
            prop.plot_fencing = request.POST.get('plot_fencing')
            prop.current_possession_status = request.POST.get('current_possession_status')

            prop.selling_price = to_int(request.POST.get('selling_price'))
            prop.price_negotiable = request.POST.get('price_negotiable', 'no')
            prop.additional_charges = request.POST.get('additional_charges')
            prop.brokerage_percentage = request.POST.get('brokerage_percentage')
            prop.manual_brokerage = request.POST.get('manual_brokerage')

            prop.ownership_type = request.POST.get('ownership_type')
            prop.ownership_document_type = request.POST.get('ownership_document_type')
            prop.other_document_type = request.POST.get('other_document_type')
            prop.rera_status = request.POST.get('rera_status')
            prop.title_clearance = request.POST.get('title_clearance')
            prop.property_encumbrance_status = request.POST.get('property_encumbrance_status')

            prop.satbara_available = request.POST.get('satbara_available')
            prop.khate_utara = request.POST.get('khate_utara')
            prop.section63_clearance = request.POST.get('section63_clearance')

            prop.property_tax_status = request.POST.get('property_tax_status')
            prop.outstanding_tax_amount = to_int(request.POST.get('outstanding_tax_amount'))
            prop.pending_since = to_date(request.POST.get('pending_since'))
            prop.property_loan_status = request.POST.get('property_loan_status')
            prop.financing_bank = request.POST.get('financing_bank')
            prop.outstanding_loan_amount = to_int(request.POST.get('outstanding_loan_amount'))
            prop.sanctioning_authority = request.POST.get('sanctioning_authority')

            prop.amenities = amenities
            prop.nearby_facilities = nearby_facilities
            prop.user_description = request.POST.get('user_description')

            prop.state = request.POST.get('state')
            prop.city = input_city
            prop.locality = input_locality
            prop.property_landmark = request.POST.get('property_landmark')
            prop.pincode = input_pincode
            prop.address = request.POST.get('address')
            prop.google_maps_link = request.POST.get('google_maps_link')
            prop.latitude = request.POST.get('latitude')
            prop.longitude = request.POST.get('longitude')

            if request.FILES.get('encumbrance_cert'):
                prop.encumbrance_cert = request.FILES.get('encumbrance_cert')
            if request.FILES.get('layout_plan'):
                prop.layout_plan = request.FILES.get('layout_plan')

            prop.listed_elsewhere = request.POST.get('listed_elsewhere', 'No')
            prop.portal_name = request.POST.get('portal_name')

            prop.listing_status = request.POST.get('listing_status')
            prop.approval_status = request.POST.get('approval_status')

            prop.save()

            # ---------- REMOVE IMAGES MARKED FOR DELETION ----------
            remove_image_ids = request.POST.getlist('delete_image_ids[]') or request.POST.getlist('remove_images[]')
            if remove_image_ids:
                AgriculturalPlotResaleImage.objects.filter(id__in=remove_image_ids, property=prop).delete()

            # ---------- ADD NEWLY UPLOADED IMAGES (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'front_view':         'plot_images_front_view[]',
                'full_plot':          'plot_images_full_plot[]',
                'farm_gate':          'plot_images_farm_gate[]',
                'boundary_fencing':   'plot_images_boundary_fencing[]',
                'road_facing':        'plot_images_road_facing[]',
                'cultivated_area':    'plot_images_cultivated_area[]',
                'irrigation_source':  'plot_images_irrigation_source[]',
                'borewell_well':      'plot_images_borewell_well[]',
                'electricity_infra':  'plot_images_electricity_infra[]',
                'farmhouse_shed':     'plot_images_farmhouse_shed[]',
                'aerial_drone':       'plot_images_aerial_drone[]',
                'layout_site_plan':   'plot_images_layout_site_plan[]',
            }

            existing_count = AgriculturalPlotResaleImage.objects.filter(property=prop).count()
            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                new_image_objs = []
                if cat_images:
                    start_seq = AgriculturalPlotResaleImage.objects.filter(property=prop, category=category).count()
                    for idx, img in enumerate(cat_images):
                        if existing_count + saved_count >= 25:
                            break
                        new_obj = AgriculturalPlotResaleImage.objects.create(
                            property=prop, image=img, category=category, sequence_order=start_seq + idx,
                        )
                        new_image_objs.append(new_obj)
                        saved_count += 1

                order_tokens = request.POST.getlist(f'image_order_{category}[]')
                for position, token in enumerate(order_tokens):
                    if token.startswith('existing:'):
                        img_id = token.split(':', 1)[1]
                        AgriculturalPlotResaleImage.objects.filter(
                            id=img_id, property=prop, category=category
                        ).update(sequence_order=position)
                    elif token.startswith('new:'):
                        new_idx = int(token.split(':', 1)[1])
                        if 0 <= new_idx < len(new_image_objs):
                            new_image_objs[new_idx].sequence_order = position
                            new_image_objs[new_idx].save(update_fields=['sequence_order'])

            # ---------- VIDEO ----------
            if request.POST.get('delete_current_video') == '1':
                prop.video.filter(source='uploaded').delete()

            video_option = request.POST.get('video_option') or 'auto'
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '').strip()
            regenerate_slideshow = request.POST.get('regenerate_slideshow') == 'on'

            if uploaded_video:
                AgriculturalPlotResaleVideo.objects.update_or_create(
                    property=prop, source='uploaded',
                    defaults={'video': uploaded_video, 'video_url': None}
                )
            elif video_option == 'rm_assisted' and property_video_link:
                AgriculturalPlotResaleVideo.objects.update_or_create(
                    property=prop, source='rm_assisted',
                    defaults={'video_url': property_video_link, 'video': None}
                )
            elif video_option == 'auto':
                existing_auto = prop.video.filter(source='auto').first()
                if regenerate_slideshow or not existing_auto:
                    CATEGORY_ORDER = list(IMAGE_CATEGORY_FIELDS.keys())
                    all_images = list(AgriculturalPlotResaleImage.objects.filter(property=prop))
                    all_images.sort(key=lambda img: (
                        CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99,
                        img.sequence_order
                    ))
                    image_paths = [
                        img.image.path for img in all_images
                        if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                    ]
                    if len(image_paths) >= 3:
                        output_relative_path = f"agricultural_plot/videos/auto_{prop.id}.mp4"
                        try:
                            result_path = generate_property_slideshow(image_paths, output_relative_path)
                            if result_path:
                                AgriculturalPlotResaleVideo.objects.update_or_create(
                                    property=prop, source='auto',
                                    defaults={'video': result_path, 'video_url': None}
                                )
                        except Exception:
                            import traceback
                            traceback.print_exc()

            return JsonResponse({
                'status': 'success',
                'message': "Agricultural Plot Listing Updated Successfully",
                'redirect_url': reverse('agricultural_plot_resale_list_agent')
            })

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            return JsonResponse({'status': 'error', 'message': f"Error while updating listing: {str(e)}"})

    # ---------- GET: render prefilled form ----------
    existing_images = AgriculturalPlotResaleImage.objects.filter(property=prop).order_by('category', 'sequence_order')
    images_by_category = {}
    for img in existing_images:
        images_by_category.setdefault(img.category, []).append(img)

    existing_image_total = existing_images.count()

    uploaded_video = prop.video.filter(source='uploaded').first()
    auto_video = prop.video.filter(source='auto').first()
    rm_video = prop.video.filter(source='rm_assisted').first()

    return render(request, 'agent/Reports/Resale_Plot/agricultural_plot_resale_edit.html', {
        
        
        'user_obj': user_obj,
        
        
    
      
        'property': prop,
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'images_by_category': images_by_category,
        'existing_image_total': existing_image_total,
        'uploaded_video': uploaded_video,
        'auto_video': auto_video,
        'rm_video': rm_video,
        'selected_amenities': (prop.amenities or '').split(',') if prop.amenities else [],
        'selected_facilities': (prop.nearby_facilities or '').split(',') if prop.nearby_facilities else [],
    })



def export_agricultural_plot_resale_agent(request):
    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    # ── 1. Re-apply the same search filters so the export matches the screen ──
    try:
        properties = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')
    except Exception:
        properties = AgriculturalPlotResaleProperty.objects.all(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    search_query        = request.GET.get("search", "").strip()
    city_query          = request.GET.get("city", "").strip()
    locality_query      = request.GET.get("locality", "").strip()
    property_type_query = request.GET.get("property_type", "").strip()
    road_facing_query   = request.GET.get("road_facing", "").strip()
    corner_plot_query   = request.GET.get("corner_plot", "").strip()
    loan_query          = request.GET.get("loan", "").strip()
    min_price           = request.GET.get("min_price", "").strip()
    max_price           = request.GET.get("max_price", "").strip()
    from_date           = request.GET.get("from_date", "").strip()
    to_date              = request.GET.get("to_date", "").strip()
    listed_by_query     = request.GET.get("listed_by", "").strip()
    uploaded_by_query   = request.GET.get("uploaded_by", "").strip()
    listing_status_query= request.GET.get("listing_status", "").strip()
    approval_status_query= request.GET.get("approval_status", "").strip()
    duplicate_query     = request.GET.get("duplicate", "").strip()

    # Apply Filters
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if search_query:
        properties = properties.filter(
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(property_no__icontains=search_query)
        )

    if city_query:
        properties = properties.filter(city__icontains=city_query)

    if locality_query:
        properties = properties.filter(locality__icontains=locality_query)

    if property_type_query:
        properties = properties.filter(property_type=property_type_query)

    if road_facing_query:
        properties = properties.filter(plot_road_facing__icontains=road_facing_query)

    if corner_plot_query in ("yes", "no"):
        properties = properties.filter(corner_plot__iexact=corner_plot_query)

    if loan_query == "yes":
        properties = properties.filter(property_loan_status__iexact="Loan Running")
    elif loan_query == "no":
        properties = properties.exclude(property_loan_status__iexact="Loan Running")

    if min_price:
        try:
            properties = properties.filter(selling_price__gte=int(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            properties = properties.filter(selling_price__lte=int(max_price))
        except ValueError:
            pass

    if from_date:
        try:
            properties = properties.filter(created_at__date__gte=datetime.strptime(from_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    if to_date:
        try:
            properties = properties.filter(created_at__date__lte=datetime.strptime(to_date, "%Y-%m-%d").date())
        except ValueError:
            pass


    EXPORT_COLS = [
        ("Sr.No", "sr_no", False, "Sr. No"),
        ("Listed By", "listed_by_type", False, "Listed By Type (Self/Other)"),
        ("Listed By", "listed_by_id", False, "Listed By Id"),
        ("Listed By", "listed_by_name", False, "Listed By Name"),
        ("Listed By", "listed_by_email", False, "Listed By Email"),
        ("Listed By", "listed_by_contact", False, "Listed By Contact"),
        ("Listed By", "listed_by_role", True, "Listed By Role"),

        ("System Generated Property Id(Auto)", "id", False, "Property Id"),
        ("Basic Agricultural Land Information", "property_title", False, "Property Title"),
        ("Basic Agricultural Land Information", "property_no", True, "Gat / Gut / Khasra / Survey Number"),
        ("Basic Agricultural Land Information", "plot_area", True, "Land Area (acre)"),
       
        ("Basic Agricultural Land Information", "property_type", True, "Agricultural Land Type"),

        ("Revenue Classification & Legal Status", "land_use", True, "Revenue / Land Classification"),
        ("Revenue Classification & Legal Status", "na_status", True, "NA Conversion Status"),
        ("Revenue Classification & Legal Status", "layout_approval_status", False, "Revenue / Panchayat Sanction"),

      

        ("Land Specifications & Agricultural Details", "plot_frontage", False, "Frontage / Width (ft)"),
        ("Land Specifications & Agricultural Details", "plot_depth", False, "Depth / Length (ft)"),
        ("Land Specifications & Agricultural Details", "plot_shape", False, "Plot Shape"),
        ("Land Specifications & Agricultural Details", "plot_road_facing", True, "Road / Track Access to Land"),
        ("Land Specifications & Agricultural Details", "road_width", False, "Access Road / Track Width)"),
        ("Land Specifications & Agricultural Details", "corner_plot", False, "Corner Plot (Two Road Access)?"),
        ("Land Specifications & Agricultural Details", "agr_area_unit", False, "Land Measurement Unit"),

        ("Land Specifications & Agricultural Details", "soil_type", False, "Soil Type"),
        ("Land Specifications & Agricultural Details", "current_crop", False, "Current Crop / Plantation"),
        ("Land Specifications & Agricultural Details", "irrigation_source", False, "Irrigation / Water Source"),
        ("Land Specifications & Agricultural Details", "agr_electricity", False, "Electricity on Land"),
        ("Land Specifications & Agricultural Details", "highway_distance", False, "Distance from Nearest Highway / Road"),
        ("Land Specifications & Agricultural Details", "land_topography", False, "Land Level / Topography"),
        ("Land Specifications & Agricultural Details", "govt_scheme", False, "Is Land Under Any Government Scheme?"),
        ("Land Specifications & Agricultural Details", "plot_fencing", False, "Land Demarcation / Boundary"),
        ("Land Specifications & Agricultural Details", "current_possession_status", False, "Current Land Use / Occupancy"),

        ("Pricing Details", "selling_price", True, "Selling Price"),
        ("Pricing Details", "price_per_unit", False, "Rate per Unit (Auto-calculated)"),
        ("Pricing Details", "price_negotiable", False, "Is the Price Negotiable?"),
        ("Pricing Details", "additional_charges", False, "Transaction / Transfer Charges"),
        ("Pricing Details", "brokerage_percentage", True, "Brokerage / Service Fee"),
        ("Pricing Details", "manual_brokerage", False, "Fixed Brokerage Amount"),

        ("Legal, Title & Approval Details", "ownership_type", False, "Ownership / Tenure Type"),
        ("Legal, Title & Approval Details", "ownership_document_type", False, "Primary Title / Ownership Document"),
        ("Legal, Title & Approval Details", "other_document_type", False, "Specify Other Ownership Document Type"),
        ("Legal, Title & Approval Details", "rera_status", False, "RERA / Revenue Department Status"),
        ("Legal, Title & Approval Details", "title_clearance", False, "Title Clarity / Marketability"),
        ("Legal, Title & Approval Details", "property_encumbrance_status", False, "Encumbrance / Charge Status"),
        ("Legal, Title & Approval Details", "satbara_available", False, "7/12 Utara (Satbara) Available?"),
        ("Legal, Title & Approval Details", "khate_utara", False, "8A Khate Utara (Land Record) Available?"),
        ("Legal, Title & Approval Details", "section63_clearance", False, "Section 63 / 63-A Clearance (Non-Agriculturist Purchase)?"),
      
        
        
        ("Legal, Title & Approval Details", "property_tax_status", False, "Land Revenue / Khajana Status"),
        ("Legal, Title & Approval Details", "outstanding_tax_amount", False, "Outstanding Tax Amount (₹)"),
        ("Legal, Title & Approval Details", "pending_since", False, "Tax Dues Pending Since"),
        ("Legal, Title & Approval Details", "property_loan_status", False, "Existing Loan / Mortgage on Property"),
        ("Legal, Title & Approval Details", "financing_bank", False, "Lender Bank / NBFC Name"),
        ("Legal, Title & Approval Details", "outstanding_loan_amount", False, "Existing Loan Outstanding Loan Amount (₹)"),
        ("Legal, Title & Approval Details", "sanctioning_authority", False, "Revenue Authority / Talathi / Tehsil Office"),

        ("Amenities & Facilities", "amenities", False, "Amenities (comma-separated)"),
        ("Nearby Facilities", "nearby_facilities", False, "Nearby Facilities (comma-separated)"),

        ("Property Descriptions(Added By User)", "user_description", False, "Property Description"),
        ("Property Summary(Auto)", "property_summary", False, "Property Summary"),
        ("Property Description(Auto)", "property_description", False, "Property Description"),

        ("Location Details", "address", True, "Complete Land Location / Revenue Address"),
        ("Location Details", "locality", True, "Village / Mouza / Taluka Name"),
        ("Location Details", "property_landmark", False, "Nearest Village / Highway / City Landmark"),
        ("Location Details", "city", True, "City / District"),
        ("Location Details", "state", True, "State"),
        ("Location Details", "pincode", False, "PIN Code"),
        ("Location Details", "google_maps_link", False, "Google Maps Link"),
        ("Location Details", "latitude", False, "Latitude"),
        ("Location Details", "longitude", False, "Longitude"),

        ("Property Listed Elsewhere", "listed_elsewhere", False, "Is Property Already Listed Elsewhere?"),
        ("Property Listed Elsewhere", "portal_name", False, "Portal Name"),

        ("Media & Listing Status", "listing_status", False, "Listing Status"),
        ("Media & Listing Status", "approval_status", False, "Approval Status"),

        ("Data Uploadeded Via", "upload_file_name", False, "Upload File Name"),
        ("Property Uploaded By", "uploaded_by_name", False, "Uploaded By Name"),
        ("Property Uploaded By", "uploaded_by_email", False, "Uploaded By Email"),
        ("Property Uploaded By", "uploaded_by_contact", False, "Uploaded By Contact"),
        ("Property Uploaded By", "uploaded_by_role", False, "Uploaded By Role"),

        ("Database Audit", "created_at", False, "Created At"),
        ("Database Audit", "updated_at", False, "Updated At"),
        ("Database Audit", "is_deleted", False, "Is Deleted"),
        ("Database Audit", "deleted_at", False, "Deleted At"),
        ("Database Audit", "deleted_by", False, "Deleted By"),
        ("Database Audit", "is_duplicate", False, "Is Duplicate"),
        ("Database Audit", "duplicate_count", False, "Duplicate Count"),
        ("Database Audit", "duplicate_group_id", False, "Duplicate Group ID"),
        ("Database Audit", "property_unique_key", False, "Property Unique Key"),

        ("Brokerage Label", "get_brokerage_label", False, "Brokerage Label Preview (auto)"),
    ]

    export_format = request.GET.get('format', 'excel')
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Agricultural_Plot_Resale_Full_Export_{timestamp}"

    # Helper for formatting values
    def format_val(val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(val, bool):
            return "Yes" if val else "No"
        elif val is None:
            return ""
        return str(val).strip()

    # ── 3. EXCEL EXPORT ──
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agricultural Plot Resale DB"

        HDR_BG  = "667EEA"
        REQ_BG  = "FEF3C7"
        OPT_BG  = "F0FDF4"
        thin = Side(style="thin", color="CBD5E1")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sec_spans = OrderedDict()
        for i, (sec, *_) in enumerate(EXPORT_COLS):
            sec_spans.setdefault(sec, []).append(i + 1)

        # Write Section Headers (Row 1)
        for sec, cols in sec_spans.items():
            c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr

            for col_idx in cols[1:]:
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor=HDR_BG)
                ws.cell(row=1, column=col_idx).border = bdr

            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

        # Write Field Headers (Row 2)
        for ci, (_, _, req, header_name) in enumerate(EXPORT_COLS, 1):
            lc = ws.cell(row=2, column=ci, value=header_name)
            lc.font = Font(bold=True, size=9)
            lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
            lc.border = bdr
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(header_name) + 5)

        # Write Data (Rows 3+)
        for row_idx, prop in enumerate(properties, start=3):
            for col_idx, (_, field, _, _) in enumerate(EXPORT_COLS, 1):

                if field == "sr_no":
                    val = row_idx - 2
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                elif field == "get_brokerage_label":
                    val = format_val(getattr(prop, field)()) if callable(getattr(prop, field, None)) else ""
                else:
                    val = format_val(getattr(prop, field, ""))

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center")

        # Layout adjustments
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # ── 4. CSV EXPORT ──
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)

        # Write exact Field Headers
        headers = [header_name for (_, _, _, header_name) in EXPORT_COLS]
        writer.writerow(headers)

        # Write Data Rows
        for row_idx, prop in enumerate(properties, start=1):
            row_data = []
            for (_, field, _, _) in EXPORT_COLS:

                if field == "sr_no":
                    val = row_idx
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                elif field == "get_brokerage_label":
                    val = format_val(getattr(prop, field)()) if callable(getattr(prop, field, None)) else ""
                else:
                    val = format_val(getattr(prop, field, ""))

                row_data.append(val)
            writer.writerow(row_data)

        return response





def agricultural_plot_resale_view_agent(request, pk):
    
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    # UPDATED: prefetch_related on the base queryset so Django fetches
    # all images, videos, and FAQs in one go before returning the 404 check.
    queryset = AgriculturalPlotResaleProperty.objects.prefetch_related('images', 'video', 'faqs')
    plot = get_object_or_404(queryset, id=pk, is_deleted=False)

    if not plot.faqs.exists():
        plot.generate_auto_faqs()

    amenities_list = [a.strip() for a in plot.amenities.split(',')] if plot.amenities else []
    facilities_list = [f.strip() for f in plot.nearby_facilities.split(',')] if plot.nearby_facilities else []

    # Rate per unit (e.g. Rate per Acre / Rate per Hectare) — precomputed here
    # rather than doing float math in the template. plot.price_per_unit is also
    # auto-calculated in model.save(), this just guards against stale/null values.
    price_per_unit = plot.price_per_unit
    if price_per_unit is None and plot.selling_price and plot.plot_area and plot.plot_area > 0:
        try:
            price_per_unit = round(float(plot.selling_price) / float(plot.plot_area))
        except (TypeError, ZeroDivisionError):
            price_per_unit = None

    rate_label = plot.get_rate_label()  # e.g. "Rate per Acre"

    # Extract videos and get the first one for the preview
    videos = plot.video.all()
    selected_video = videos.first() if videos.exists() else None

    # Get all images and sort them
    images = plot.images.all().order_by('category', 'sequence_order')

    # Group images by category for the filtering tabs
    grouped_images = {}
    for img in images:
        cat = img.category
        if cat not in grouped_images:
            grouped_images[cat] = {
                'label': img.get_category_display() if hasattr(img, 'get_category_display') else cat.replace('_', ' ').title(),
                'images': []
            }
        grouped_images[cat]['images'].append(img)

    return render(request, 'agent/Reports/Resale_Plot/agricultural_plot_resale_view.html', {
        'user_obj': user_obj,
        
        
        'plot': plot,
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'price_per_unit': price_per_unit,
        'rate_label': rate_label,
        'images': images,
        'grouped_images': grouped_images,
        'selected_video': selected_video,
    })







def agricultural_plot_resale_list_agent(request):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agent")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    qs = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).prefetch_related('images', 'video').order_by("-created_at")

    # ---- Search & Filters ----
    search_query = request.GET.get("search", "").strip()
    city_query = request.GET.get("city", "").strip()
    locality_query = request.GET.get("locality", "").strip()
    property_type_query = request.GET.get("property_type", "").strip()
    road_facing_query = request.GET.get("road_facing", "").strip()
    corner_plot_query = request.GET.get("corner_plot", "").strip()
    loan_query = request.GET.get("loan", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    listed_by_query = request.GET.get("listed_by", "").strip()
    uploaded_by_query = request.GET.get("uploaded_by", "").strip()
    listing_status_query = request.GET.get("listing_status", "").strip()
    approval_status_query = request.GET.get("approval_status", "").strip()
    duplicate_query = request.GET.get("duplicate", "").strip()

    if listed_by_query and listed_by_query != 'All Roles':
        qs = qs.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':
        qs = qs.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        qs = qs.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        qs = qs.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        qs = qs.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        qs = qs.filter(is_duplicate=False)

    if search_query:
        qs = qs.filter(
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(property_no__icontains=search_query)
        )

    if city_query:
        qs = qs.filter(city__icontains=city_query)

    if locality_query:
        qs = qs.filter(locality__icontains=locality_query)

    if property_type_query:
        qs = qs.filter(property_type=property_type_query)

    if road_facing_query:
        qs = qs.filter(plot_road_facing__icontains=road_facing_query)

    if corner_plot_query in ("yes", "no"):
        qs = qs.filter(corner_plot__iexact=corner_plot_query)

    if loan_query == "yes":
        qs = qs.filter(property_loan_status__iexact="Loan Running")
    elif loan_query == "no":
        qs = qs.exclude(property_loan_status__iexact="Loan Running")

    if min_price:
        try:
            qs = qs.filter(selling_price__gte=int(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            qs = qs.filter(selling_price__lte=int(max_price))
        except ValueError:
            pass

    if from_date:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(from_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    if to_date:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(to_date, "%Y-%m-%d").date())
        except ValueError:
            pass

    total_properties = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).count()
   
    filtered_count = qs.count()
    # ---- KPI Cards (Inventory & Land Features) ----
    active_listings = qs.filter(selling_price__isnull=False).count()
    irrigated_count = qs.filter(property_type="irrigated_land").count()
    orchard_count = qs.filter(property_type="orchard_plantation").count()
    na_converted_count = qs.filter(na_status="NA Converted").count()

    corner_plot_count = qs.filter(corner_plot__iexact="yes").count()
    fenced_plot_count = qs.exclude(
        Q(plot_fencing__isnull=True) | Q(plot_fencing="") | Q(plot_fencing="none")
    ).count()
    finance_ready_count = qs.filter(property_loan_status__iexact="No Active Loan").count()
    satbara_available_count = qs.filter(satbara_available__iexact="Yes").count()

    # ---- KPI Cards (Listing, Approval & Duplicates) ----
    active_listing_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Active").count()
    inactive_listing_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Inactive").count()
    sold_listing_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Sold").count()
    rented_listing_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, listing_status__iexact="Rented").count()

    pending_approval_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, approval_status__iexact="Pending").count()
    approved_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, approval_status__iexact="Approved").count()
    rejected_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, approval_status__iexact="Rejected").count()

    duplicate_properties_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, is_duplicate=True).count()
    unique_properties_count = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role, is_duplicate=False).count()

    # ---- Pagination ----
    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    start_index = (page_obj.number - 1) * paginator.per_page
    for i, obj in enumerate(page_obj.object_list, start=1):
        obj.original_sr_no = start_index + i

    # ---- Dropdown option sources ----
    unique_cities = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(city__isnull=True).exclude(city="").values_list("city", flat=True).distinct().order_by("city")
    unique_property_types = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(property_type__isnull=True).exclude(property_type="").values_list("property_type", flat=True).distinct().order_by("property_type")
    uploaded_files = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(upload_file_name__isnull=True).exclude(upload_file_name="").values_list("upload_file_name", flat=True).distinct()

    unique_listed_roles = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(listed_by_role__isnull=True).exclude(listed_by_role="").values_list("listed_by_role", flat=True).distinct().order_by("listed_by_role")
    unique_uploaded_roles = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role="").values_list("uploaded_by_role", flat=True).distinct().order_by("uploaded_by_role")
    unique_listing_status = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(listing_status__isnull=True).exclude(listing_status="").values_list("listing_status", flat=True).distinct().order_by("listing_status")
    unique_approval_status = AgriculturalPlotResaleProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).exclude(approval_status__isnull=True).exclude(approval_status="").values_list("approval_status", flat=True).distinct().order_by("approval_status")

    context = {
        "properties": page_obj,
        "page_obj": page_obj,
        "total_properties": total_properties,
        "filtered_count": filtered_count,
        "active_listings": active_listings,
        "irrigated_count": irrigated_count,
        "orchard_count": orchard_count,
        "na_converted_count": na_converted_count,
        "corner_plot_count": corner_plot_count,
        "fenced_plot_count": fenced_plot_count,
        "finance_ready_count": finance_ready_count,
        "satbara_available_count": satbara_available_count,

        "active_listing_count": active_listing_count,
        "inactive_listing_count": inactive_listing_count,
        "sold_listing_count": sold_listing_count,
        "rented_listing_count": rented_listing_count,
        "pending_approval_count": pending_approval_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "duplicate_properties_count": duplicate_properties_count,
        "unique_properties_count": unique_properties_count,

        "unique_cities": unique_cities,
        "unique_property_types": unique_property_types,
        "uploaded_files": uploaded_files,
        "unique_listed_roles": unique_listed_roles,
        "unique_uploaded_roles": unique_uploaded_roles,
        "unique_listing_status": unique_listing_status,
        "unique_approval_status": unique_approval_status,

        "search_query": search_query,
        "city_query": city_query,
        "locality_query": locality_query,
        "property_type_query": property_type_query,
        "road_facing_query": road_facing_query,
        "corner_plot_query": corner_plot_query,
        "loan_query": loan_query,
        "min_price": min_price,
        "max_price": max_price,
        "listed_by_query": listed_by_query,
        "uploaded_by_query": uploaded_by_query,
        "listing_status_query": listing_status_query,
        "approval_status_query": approval_status_query,
        "duplicate_query": duplicate_query,

        
        'ameneties_obj': ameneties_obj,
        'facilities_obj': facilities_obj,
        'user_obj': user_obj,
      
        
      
        'image_category_choices': AgriculturalPlotResaleImage.CATEGORY_CHOICES,
    }
    return render(request, "agent/Reports/Resale_Plot/agricultural_plot_resale_list.html", context)




@require_POST
def agricultural_plot_resale_bulk_delete_agent(request):
    """Agent Advanced Bulk Delete (Soft Delete) — scoped to agent's own agricultural plots, supports Admin impersonation."""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Parse Payload
        # ======================================
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (ValueError, UnicodeDecodeError):
            return JsonResponse({"status": "error", "message": "Invalid payload."}, status=400)

        delete_type = payload.get("delete_type")

        # ======================================
        # 6. Base Queryset — SCOPED TO THIS AGENT ONLY
        # ======================================
        qs = AgriculturalPlotResaleProperty.objects.filter(
            is_deleted=False,
            listed_by_id=agent_obj.user_id
        )

        if delete_type == "current_page":
            ids = payload.get("page_ids", [])
            qs = qs.filter(id__in=ids)

        elif delete_type == "date_range":
            from_date = payload.get("from_date")
            to_date = payload.get("to_date")
            if not from_date or not to_date:
                return JsonResponse({"status": "error", "message": "Both dates are required."}, status=400)
            qs = qs.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)

        elif delete_type == "latest_month":
            today = timezone.now().date()
            first_of_month = today.replace(day=1)
            qs = qs.filter(created_at__date__gte=first_of_month)

        elif delete_type == "old_data":
            cutoff = timezone.now() - timedelta(days=180)
            qs = qs.filter(created_at__lt=cutoff)

        elif delete_type == "by_uploader":
            text = payload.get("uploader_text", "").strip()
            if not text:
                return JsonResponse({"status": "error", "message": "Uploader detail is required."}, status=400)
            qs = qs.filter(
                Q(uploaded_by_name__icontains=text) |
                Q(uploaded_by_email__icontains=text) |
                Q(uploaded_by_role__icontains=text)
            )

        elif delete_type == "by_file":
            file_name = payload.get("file_name", "").strip()
            if not file_name:
                return JsonResponse({"status": "error", "message": "File name is required."}, status=400)
            qs = qs.filter(upload_file_name=file_name)

        elif delete_type == "delete_all":
            pass  # qs already scoped to this agent's non-deleted rows

        else:
            return JsonResponse({"status": "error", "message": "Unknown delete criteria."}, status=400)

        count = qs.count()
        if count == 0:
            return JsonResponse({"status": "error", "message": "No matching records found to delete."}, status=400)

        qs.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=user_identity)

        AgriculturalPlotResaleActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type="DELETE",
            property_id="Multiple / Bulk Action",
            action_payload=json.dumps({"delete_type": delete_type, "records_deleted": count}),
            status="SUCCESS",
        )

        return JsonResponse({"status": "success", "message": f"{count} record(s) deleted successfully."})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



@require_POST
def agricultural_plot_resale_delete_agent(request, pk):
    """Agent Soft Delete Agricultural Plot Resale Property (supports Admin impersonation)"""

    try:
        # ======================================
        # 1. Retrieve identity from browser session
        # ======================================
        user_id = request.session.get('User_id')
        admin_id = request.session.get('Admin_id')
        logged_in_role = request.session.get('user_type')

        # ======================================
        # 2. VIP Access Control
        # ======================================
        is_valid_agent = (user_id and logged_in_role == "Agent")
        is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

        if not is_valid_agent and not is_valid_admin:
            return JsonResponse({
                'status': 'error',
                'message': 'Unauthorized access.'
            }, status=401)

        # ======================================
        # 3. The ID Swap
        # ======================================
        if is_valid_admin:
            dashboard_user_id = request.session.get('impersonate_id')
        else:
            dashboard_user_id = user_id

        # ======================================
        # 4. Fetch the resolved agent
        # ======================================
        try:
            agent_obj = User_Details.objects.get(id=dashboard_user_id)
        except User_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found.'})

        if is_valid_admin:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            admin_label = admin_obj.username if hasattr(admin_obj, 'username') else f"Admin #{admin_id}"
            user_identity = (
                f"[Impersonated] Acting Admin: {admin_label} | "
                f"On behalf of User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = "Admin (Impersonating Agent)"
        else:
            user_identity = (
                f"User ID: {agent_obj.user_id} | "
                f"Name: {agent_obj.user_name} | "
                f"Email: {agent_obj.user_email} | "
                f"Phone: {agent_obj.user_phone}"
            )
            user_role = agent_obj.user_role or "Agent"

        # ======================================
        # 5. Fetch Property
        # ======================================
        plot = get_object_or_404(AgriculturalPlotResaleProperty, id=pk)

        # ======================================
        # 6. Security Check — agent can delete only own property
        # ======================================
        if str(plot.listed_by_id) != str(agent_obj.user_id):
            return JsonResponse({
                'status': 'error',
                'message': 'You can delete only your own properties.'
            })

        # ======================================
        # 7. Soft Delete
        # ======================================
        plot.is_deleted = True
        plot.deleted_at = timezone.now()
        plot.deleted_by = user_identity
        plot.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])

        AgriculturalPlotResaleActivityLog.objects.create(
            user_identity=user_identity,
            user_role=user_role,
            action_type="DELETE",
            property_id=plot.id,
            action_payload=json.dumps({"reason": "Manual delete via Agent list page"}),
            status="SUCCESS",
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Property {plot.id} moved to recycle bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

    



############# Views End for Resale Plot Agriculture Module for Agent ######################################