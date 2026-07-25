from django.shortcuts import render,HttpResponse,redirect,get_object_or_404
from Admin_App.models import *
from Landlord_Panel.views import calculate_profile_strength
from django.template.loader import render_to_string
from CRM_Panel.models import *
from django.db.models import Count, Avg, Max, Min, Q,Sum
from django.db.models.functions import Coalesce
from datetime import datetime
from django.contrib import messages
import json
from django.http import JsonResponse
from django.urls import reverse
import hashlib
import re
from django.core.paginator import Paginator


from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from django.db.models import Q

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch
from django.utils.dateparse import parse_date

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect
from django.shortcuts import render,HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Admin_App .models import *
from Main_App .models import *
from seo .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from openpyxl import load_workbook
from django.template.loader import render_to_string
import traceback
import json
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator  # ← ADD THIS
import csv
import csv
import json
from django.db.models import Count, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect



# Create your views here.

def generate_property_fingerprint(property_no, building_name, locality, pincode=""):
    """
    Generates a normalized SHA-256 hash unique key based on physical location.
    Strips spaces and special characters for consistent matching.
    """
    def clean_text(val):
        if not val:
            return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()

    raw_string = f"{clean_text(property_no)}_{clean_text(building_name)}_{clean_text(locality)}_{clean_text(pincode)}"
    return hashlib.sha256(raw_string.encode('utf-8')).hexdigest()

############## Views start for agency dashboard ######################

def Agency_Dashboard(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # Calculate profile strength based on the swapped user object
    completion_score = calculate_profile_strength(user_obj)

    enquiry_obj_agency = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    context = {
        'user_obj': user_obj,
        # Pass the object's actual role so the template renders the Agency UI correctly
        'user_role': user_obj.user_role,
        'profile_completion_percentage': completion_score,
        'enquiry_obj_agency':enquiry_obj_agency
    }
    
    return render(request, "agency_panel/agency_dashboard.html", context)

############# Views end for agency dashboard ###########################


############# Views start for update agency profile #####################

def Update_Profile_Agency(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_agency = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_agency':enquiry_obj_agency
    }
    
    return render(request, "agency_panel/Profile/agency_profile.html", context)

############# Views end for update agency profile ###########################


############## Views start for assign enquiries to agency/builder #################

def Assign_Enquiry_Agency(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry_obj_agency = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    # Calculate profile strength based on the swapped user object
    completion_score = calculate_profile_strength(user_obj)

    enquiry_obj = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).order_by('-id')
    enquiry_obj_count = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()

    ############## Enquiries Stats By Source ##############################

    fb_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="facebook",assigned_to__id=user_obj.id).count()
    insta_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="instagram",assigned_to__id=user_obj.id).count()
    whatsapp_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="whatsapp",assigned_to__id=user_obj.id).count()
    google_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="google",assigned_to__id=user_obj.id).count()
    linkedin_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="linkedin",assigned_to__id=user_obj.id).count()
    twitter_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="twitter",assigned_to__id=user_obj.id).count()
    youtube_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="youtube",assigned_to__id=user_obj.id).count()
    referral_obj_count = PropertyEnquiry.objects.filter(utm_link__utm_source="referral",assigned_to__id=user_obj.id).count()

    ########### Enquiry Stats by lead source ################################

    pending_obj_count = PropertyEnquiry.objects.filter(lead_status="Pending",assigned_to__id=user_obj.id).count()
    progress_obj_count = PropertyEnquiry.objects.filter(lead_status="In Progress",assigned_to__id=user_obj.id).count()
    hold_obj_count = PropertyEnquiry.objects.filter(lead_status="Hold",assigned_to__id=user_obj.id).count()
    closed_obj_count = PropertyEnquiry.objects.filter(lead_status="Closed",assigned_to__id=user_obj.id).count()
    cancelled_obj_count = PropertyEnquiry.objects.filter(lead_status="Cancelled",assigned_to__id=user_obj.id).count()

    rendered = render_to_string("agency_panel/render_to_string/R_Enquiry/r_t_s_enquiry.html",{'enquiry_obj':enquiry_obj,'enquiry_obj_count':enquiry_obj_count,'fb_obj_count':fb_obj_count,'insta_obj_count':insta_obj_count,'whatsapp_obj_count':whatsapp_obj_count,'google_obj_count':google_obj_count,'linkedin_obj_count':linkedin_obj_count,'twitter_obj_count':twitter_obj_count,'youtube_obj_count':youtube_obj_count,'referral_obj_count':referral_obj_count,'pending_obj_count':pending_obj_count,'progress_obj_count':progress_obj_count,'hold_obj_count':hold_obj_count,'closed_obj_count':closed_obj_count,'cancelled_obj_count':cancelled_obj_count})
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'profile_completion_percentage': completion_score,
        'property_enquiry_list':rendered,
        'enquiry_obj_agency':enquiry_obj_agency
    }
    
    return render(request, "agency_panel/Enquiry/assign_enquiry.html", context)

############# Views end for assign enquiries to agency/builder ####################


############ Views start for update property enquiry #########################

def update_enquiry_agency(request,id):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    enquiry = PropertyEnquiry.objects.get(id=id)

    enquiry_obj_agency = PropertyEnquiry.objects.filter(assigned_to__id=user_obj.id).count()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'enquiry_obj_agency':enquiry_obj_agency,
        'enquiry':enquiry
    }
    
    return render(request, "agency_panel/Enquiry/update_enquiry.html", context)

############# Views end for update property enquiry #################################



#################Views Start For Rental Residential Property###########################



def residential_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()


    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

            # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION (Who submitted the HTML form) ----------
            if admin_obj:
                uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
                uploader_email = getattr(admin_obj, 'email', '')
                uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
                uploader_role = "Admin"
                uploader_id = f"ADMIN_{admin_id}"
            elif user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "User"
                uploader_id = f"USER_{user_id}"
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION (Who owns/manages the listing) ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE (Checking LISTED BY, not UPLOADED BY)
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality_area') or request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_property_fingerprint(
                input_property_no, 
                input_building_name, 
                input_locality, 
                input_pincode
            )

            # 1. Direct Case-Insensitive Query for same unit in same locality/building
            direct_duplicates = RentalResidentialProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality_area__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            # Combine fingerprint match OR direct field match
            existing_duplicates = (
                RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                # Level 1: Hard Block if LISTED BY the exact same person (ID, Email, OR Phone match)
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        # INSTANT REJECTION: Same agent/owner cannot list the same property twice
                        messages.error(
                            request, 
                            f"Duplicate Blocked: This property (Unit {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        )
                        return redirect('residential')

                # Level 2: Different Agent/User listing the exact same physical unit -> Allow save & Flag
                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = RentalResidentialProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Residential",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                property_no=input_property_no,
                bhk_type=request.POST.get('bhk_type'),
                renting_option=request.POST.get('renting_option'),
                built_up_area=to_decimal(request.POST.get('built_up_area')),
                bathrooms=to_int(request.POST.get('bathrooms')),
                balconies=to_int(request.POST.get('balconies')),
                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int(request.POST.get('total_floors')),
                facing_direction=request.POST.get('facing_direction', request.POST.get('facing')),
                furnishing_status=request.POST.get('furnishing_status'),
                available_for=request.POST.get('available_for'),

                carpet_area=to_decimal(request.POST.get('carpet_area')),
                city_zone=request.POST.get('city_zone', request.POST.get('zone')),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition', request.POST.get('construction_status')),
                property_age=request.POST.get('property_age'),
                wing_number=request.POST.get('wing_number'),
                building_name=input_building_name,

                availability_status=request.POST.get('availability_status', request.POST.get('possession_status')),
                available_from=available_from,
                lease_duration=request.POST.get('lease_duration'),
                brokerage_percentage=request.POST.get('brokerage_percentage', request.POST.get('brokerage')),
                manual_brokerage=request.POST.get('manual_brokerage'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=to_int(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int(request.POST.get('security_deposit_amount', request.POST.get('security_deposit'))),
                maintenance_type=request.POST.get('maintenance_type'),
                monthy_maintenance_amount=to_int(request.POST.get('monthy_maintenance_amount', request.POST.get('maintenance_amount'))),
                total_move_in_cost=to_int(request.POST.get('total_move_in_cost')),

                address=request.POST.get('address'),
                city=request.POST.get('city'),
                locality_area=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=input_pincode,
                main_road_connectivity=request.POST.get('main_road_connectivity', request.POST.get('road_connectivity')),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                amenities=amenities,
                nearby_facilities=facilities,

                user_description=request.POST.get('user_description'),
                description=request.POST.get('description'),
                rent_residential_desc=request.POST.get('rent_residential_desc'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC ----------
            images = request.FILES.getlist('property_images[]')
            for index, img in enumerate(images[:10]):
                RentalResidentialImage.objects.create(
                    property=prop, 
                    image=img,
                    sequence_order=index
                )

            messages.success(request, "Property Added Successfully ")
            
            return redirect('rental_list_agency')

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            messages.error(request, f"Error while saving listing: {str(e)}")
            return redirect('residential_agency')

    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Rental/residential.html", context)


def residential_agency_edit(request, pk):
    prop = get_object_or_404(RentalResidentialProperty, id=pk)

    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    if request.method == 'POST':
        try:
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # =====================================================
            # AUDIT LOGIC: Snapshot current state BEFORE modification
            # =====================================================
            tracked_fields_list = [
                'property_type', 'property_no', 'bhk_type', 'renting_option',
                'built_up_area', 'bathrooms', 'balconies', 'building_configuration',
                'total_floors', 'facing_direction', 'furnishing_status', 'available_for',
                'carpet_area', 'city_zone', 'ownership_type', 'property_condition',
                'property_age', 'wing_number', 'building_name',
                'availability_status', 'available_from', 'lease_duration',
                'brokerage_percentage', 'manual_brokerage', 'monthly_rent',
                'advance_rent_month', 'advance_rent_amount',
                'security_deposit_type', 'security_deposit_amount',
                'maintenance_type', 'monthy_maintenance_amount', 'total_move_in_cost',
                'address', 'city', 'locality_area', 'property_landmark', 'state', 'pincode',
                'main_road_connectivity', 'google_maps_link', 'latitude', 'longitude',
                'listed_by_id', 'listed_by_name', 'listed_by_email', 'listed_by_contact', 'listed_by_role',
                'listing_status', 'approval_status',
            ]
            old_state_snapshot = {field: str(getattr(prop, field, '')) for field in tracked_fields_list}

           
            # =====================================================
            prop.listed_by_type = request.POST.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to = request.POST.get('assigned_to', prop.assigned_to)
            # assigned_to_role is posted alongside assigned_to (mirrors the Add form)
            # and is available here if you want to store it on the model as well, e.g.:
            # prop.assigned_to_role = request.POST.get('assigned_to_role', getattr(prop, 'assigned_to_role', ''))
            prop.listed_by_id = (request.POST.get('listed_by_id') or prop.listed_by_id or '').strip()
            prop.listed_by_name = (request.POST.get('listed_by_name') or prop.listed_by_name or '').strip()
            prop.listed_by_email = (request.POST.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = (request.POST.get('listed_by_contact') or prop.listed_by_contact or '').strip()
            prop.listed_by_role = (request.POST.get('listed_by_role') or prop.listed_by_role or '').strip()

            # =====================================================
            # 2. BASIC INFORMATION
            # =====================================================
            prop.property_type = request.POST.get('property_type')
            prop.property_no = request.POST.get('property_no', prop.property_no)
            prop.bhk_type = request.POST.get('bhk_type')
            prop.renting_option = request.POST.get('renting_option')
            prop.built_up_area = to_decimal(request.POST.get('built_up_area'))
            prop.bathrooms = to_int(request.POST.get('bathrooms'))
            prop.balconies = to_int(request.POST.get('balconies'))
            prop.building_configuration = request.POST.get('building_configuration')
            prop.total_floors = to_int(request.POST.get('total_floors'))
            prop.facing_direction = request.POST.get('facing_direction')
            prop.furnishing_status = request.POST.get('furnishing_status')
            prop.available_for = request.POST.get('available_for')

            # =====================================================
            # 3. PROPERTY DETAILS
            # =====================================================
            prop.carpet_area = to_decimal(request.POST.get('carpet_area'))
            prop.city_zone = request.POST.get('city_zone')
            prop.ownership_type = request.POST.get('ownership_type')
            prop.property_condition = request.POST.get('property_condition')
            prop.property_age = request.POST.get('property_age')
            prop.wing_number = request.POST.get('wing_number')
            prop.building_name = request.POST.get('building_name')

            # =====================================================
            # 4. AVAILABILITY DETAILS
            # =====================================================
            prop.availability_status = request.POST.get('availability_status')

            available_from_raw = request.POST.get('available_from')
            if available_from_raw and available_from_raw.strip():
                try:
                    prop.available_from = datetime.strptime(available_from_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    prop.available_from = None
            else:
                prop.available_from = None

            prop.lease_duration = request.POST.get('lease_duration')
            prop.brokerage_percentage = request.POST.get('brokerage_percentage')
            prop.manual_brokerage = request.POST.get('manual_brokerage')

            # =====================================================
            # 5. PRICING DETAILS
            # =====================================================
            prop.monthly_rent = to_int(request.POST.get('monthly_rent'))
            prop.advance_rent_month = request.POST.get('advance_rent_month')
            prop.advance_rent_amount = to_int(request.POST.get('advance_rent_amount'))
            prop.security_deposit_type = request.POST.get('security_deposit_type')
            prop.security_deposit_amount = to_int(request.POST.get('security_deposit_amount'))
            prop.maintenance_type = request.POST.get('maintenance_type')
            prop.monthy_maintenance_amount = to_int(request.POST.get('monthy_maintenance_amount'))
            prop.total_move_in_cost = to_int(request.POST.get('total_move_in_cost'))

            # =====================================================
            # 6. LOCATION DETAILS
            # =====================================================
            prop.address = request.POST.get('address')
            prop.city = request.POST.get('city')
            prop.locality_area = request.POST.get('locality_area')
            prop.property_landmark = request.POST.get('property_landmark')
            prop.state = request.POST.get('state')
            prop.pincode = request.POST.get('pincode')
            prop.main_road_connectivity = request.POST.get('main_road_connectivity')
            prop.google_maps_link = request.POST.get('google_maps_link')
            prop.latitude = request.POST.get('latitude')
            prop.longitude = request.POST.get('longitude')

            # =====================================================
            # 7. AMENITIES & FACILITIES
            # (Unchanged — the modal-picker in the template still posts the exact
            # same field names amenities[] / nearby_facilities[], so no view
            # changes were required here.)
            # =====================================================
            prop.amenities = ",".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]'))

            # =====================================================
            # 8. DESCRIPTIONS
            # =====================================================
            prop.user_description = request.POST.get('user_description')

            # =====================================================
            # 9. LISTED ELSEWHERE
            # =====================================================
            prop.listed_elsewhere = request.POST.get('listed_elsewhere', 'No')
            prop.portal_name = request.POST.get('portal_name')

            # =====================================================
            # 10. LISTING STATUS & APPROVAL STATUS
            # =====================================================
            prop.listing_status = request.POST.get('listing_status', prop.listing_status)
            prop.approval_status = request.POST.get('approval_status', prop.approval_status)

            # description / rent_residential_desc are auto-regenerated in model.save()
            # via generate_auto_descriptions() — no need to set them from POST.

            prop.save()

            # =====================================================
            # 11. IMAGE UPLOAD & SEQUENCE LOGIC
            # =====================================================
            images = request.FILES.getlist('property_images[]')
            current_count = prop.images.count()
            for img in images:
                if current_count < 10:
                    RentalResidentialImage.objects.create(
                        property=prop, image=img, sequence_order=current_count
                    )
                    current_count += 1

            # =====================================================
            # AUDIT LOGIC: Diff dictionary
            # =====================================================
            modifications_diff = {}
            modified_fields_summary = []

            for field in tracked_fields_list:
                new_val = str(getattr(prop, field, ''))
                if old_state_snapshot[field] != new_val:
                    modifications_diff[field] = {
                        "old_value": old_state_snapshot[field],
                        "new_value": new_val
                    }
                    modified_fields_summary.append(field)

            # if modified_fields_summary:
            #     RentalActivityLog.objects.create(
            #         user_identity=admin_obj.email or admin_obj.username,
            #         user_role="Admin",
            #         action_type='UPDATE',
            #         property_id=prop.id,
            #         targeted_fields=", ".join(modified_fields_summary[:4]) + ("..." if len(modified_fields_summary) > 4 else ""),
            #         associated_file=prop.upload_file_name or "Web UI Form",
            #         action_payload=json.dumps(modifications_diff),
            #         ip_address=_get_client_ip(request),
            #         status='SUCCESS'
            #     )

            return JsonResponse({
                'status': 'success',
                'message': 'Property updated successfully!',
                'redirect_url': reverse('rental_list_agency')
            })

        except Exception as e:
            RentalActivityLog.objects.create(
                user_identity=admin_obj.email if 'admin_obj' in locals() else "Unknown Session",
                user_role="Admin",
                action_type='UPDATE',
                property_id=pk,
                targeted_fields="System Errors",
                action_payload=f"Failed modification update runtime sequence execution target: {str(e)}",
                ip_address=_get_client_ip(request),
                status='FAILED'
            )
            return JsonResponse({
                'status': 'error',
                'message': f"Failed to save data: {str(e)}"
            }, status=400)

    return render(request, 'agency_panel/Forms/Rental/residential_edit.html', {
        'property': prop,
        'user_obj': user_obj,
      
        
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all()
    })


def residential_agency_view(request, pk):
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agent = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agent, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agent and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agent's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agent is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)


    # Prefetch core and related assets for the current listing
    prop = get_object_or_404(
        RentalResidentialProperty.objects.prefetch_related('images', 'faqs'), 
        pk=pk
    )
    
    # Context cross-linking: Pull the latest uploaded properties along with their dynamic FAQs
    latest_properties = RentalResidentialProperty.objects.filter(
        is_deleted=False
    ).exclude(
        id=prop.id
    ).prefetch_related('faqs').order_by('-created_at')[:4]

    # Convert comma-separated string arrays smoothly
    amenities_list = [x.strip() for x in prop.amenities.split(',')] if prop.amenities else []
    
    # FIX APPLIED HERE: Changed prop.facilities to prop.nearby_facilities to match the updated model
    facilities_list = [x.strip() for x in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []

    context = {
        'property': prop,
        'images': prop.images.all(),
        'faqs': prop.faqs.all(), # Dynamic property FAQs
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'latest_properties': latest_properties,
        'user_obj': user_obj # Direct cross-linking hook
    }
    return render(request,
        'agency_panel/Forms/Rental/residential_view.html',
        context)




def rental_list_agency(request):
       # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════

    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', '').strip()
    bhk_query = request.GET.get('bhk_type', '').strip()
    state_query = request.GET.get('state', '').strip()          # NEW
    furnish_query = request.GET.get('furnishing', '').strip()
    possession_query = request.GET.get('possession', '').strip()  # maps to availability_status

    prop_type_query = request.GET.get('property_type', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()     # listed_by_role
    uploaded_by_query = request.GET.get('uploaded_by', '').strip()  # NEW -> uploaded_by_role
    budget_query = request.GET.get('budget', '').strip()

    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query = request.GET.get('duplicate', '').strip()

    # ═══════════════════════════════════════
    # BASE QUERYSET & PERSISTENT SR.NO MAP
    # ═══════════════════════════════════════

    base_properties = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role).order_by('-id')

    absolute_ordered_ids = list(base_properties.values_list('id', flat=True))

    properties = base_properties

    # ═══════════════════════════════════════
    # GLOBAL SEARCH FILTER (Across ALL Fields + Sr.No)
    # ═══════════════════════════════════════

    if search_query:
        sr_no_query = Q()

        if search_query.isdigit():
            target_index = int(search_query) - 1
            if 0 <= target_index < len(absolute_ordered_ids):
                target_id = absolute_ordered_ids[target_index]
                sr_no_query = Q(id=target_id)

        properties = properties.filter(
            sr_no_query |
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(renting_option__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(bhk_type__icontains=search_query) |
            Q(locality_area__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(available_for__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(upload_file_name__icontains=search_query)
        )

    # ═══════════════════════════════════════
    # ADVANCED DROPDOWN FILTERS
    # ═══════════════════════════════════════

    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__iexact=city_query)

    if bhk_query and bhk_query != 'All BHK':
        properties = properties.filter(bhk_type__iexact=bhk_query)

    if state_query and state_query != 'All States':          # NEW
        properties = properties.filter(state__iexact=state_query)

    if furnish_query and furnish_query != 'All':
        properties = properties.filter(furnishing_status__iexact=furnish_query)

    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__iexact=possession_query)

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__iexact=prop_type_query)

    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':   # NEW
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_10k':
            properties = properties.filter(monthly_rent__lt=10000)
        elif budget_query == '10k_25k':
            properties = properties.filter(monthly_rent__gte=10000, monthly_rent__lte=25000)
        elif budget_query == '25k_50k':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=50000)
        elif budget_query == 'above_50k':
            properties = properties.filter(monthly_rent__gt=50000)

    # Date Filters
    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            properties = properties.filter(created_at__date__gte=from_date)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            properties = properties.filter(created_at__date__lte=to_date)

    # ═══════════════════════════════════════
    # PAGINATION & INJECT PERSISTENT SR.NO
    # ═══════════════════════════════════════
    filtered_count = properties.count()

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    id_to_srno = {pid: idx + 1 for idx, pid in enumerate(absolute_ordered_ids)}
    for prop in page_obj:
        prop.original_sr_no = id_to_srno[prop.id]

   

    # ═══════════════════════════════════════
    # EXPORT DATA (CSV / EXCEL)
    # ═══════════════════════════════════════
    if request.GET.get('download') in ['excel', 'csv']:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        from collections import OrderedDict
        import csv

        # ═══════════════════════════════════════
        # SINGLE SOURCE OF TRUTH — reuse the same field map the template
        # download and the importer already use, so LABELS stay in sync.
        # ═══════════════════════════════════════
        (field_map_sections, field_to_label_map, label_to_field_map,
         system_injected_map, helper_only_labels_map,
         decimal_fields_map, int_fields_map) = _residential_field_map()

        # ═══════════════════════════════════════
        # CANONICAL REQUIRED FIELDS — must be the literal same set used in
        # import_residential_excel's REQUIRED_FIELD_KEYS. Do NOT derive this
        # from _residential_field_map()'s per-field `required` flag — that
        # flag is used for template styling only and does not match the
        # actual mandatory inputs on the "Post Your Residential Rental" form
        # (e.g. it doesn't mark listed_by_name/email/contact as required,
        # but the Add form and the importer both do).
        # ═══════════════════════════════════════
        REQUIRED = {
            "property_type", "property_no", "bhk_type", "renting_option",
            "built_up_area", "bathrooms", "building_configuration", "total_floors",
            "furnishing_status", "available_for", "monthly_rent",
            "address", "city", "locality_area", "state", "pincode",
            "listed_by_name", "listed_by_contact", "listed_by_email", "listed_by_role",
        }

        # FIELD_LABELS starts from the field map (guarantees every shared,
        # re-importable column uses the exact label the importer expects).
        FIELD_LABELS = dict(field_to_label_map)

        # Export-only columns: appear on this report but are NOT part of the
        # import contract in _residential_field_map(). Never required, and
        # safely ignored (as unmatched headers) if this file is re-uploaded.
        FIELD_LABELS.update({

            "id": "Property ID",
            "property_title": "Property Title",
            "description": "Property Summary (Auto)",
            "rent_residential_desc": "Property Description (Auto)",
            "upload_file_name": "Uploaded File Name",
            "listing_status": "Listing Status",
            "approval_status": "Approval Status",
            "listed_elsewhere": "Listed Elsewhere?",
            "portal_name": "Portal Name",

            
            "created_at": "Created At",
        })



        sections = {


            "S.No": [
                "sr_no",
            ],
            "Listed By": [
                "listed_by_type", "listed_by_id", "listed_by_name", "listed_by_contact", "listed_by_email",
                "listed_by_role",
            ],
            "Basic Information": [
                "id", "property_title", "property_type", "property_no", "bhk_type", "renting_option",
                "built_up_area", "bathrooms", "balconies",
                "building_configuration", "total_floors", "facing_direction",
                "furnishing_status", "available_for"
            ],
            "Property Details": [
                "carpet_area", "city_zone", "ownership_type", "property_condition", "property_age",
                "wing_number", "building_name"
            ],
            "Availability Details": [
                "availability_status", "available_from", "lease_duration",
            ],
            "Pricing & Brokarage Details": [
                "monthly_rent","brokerage_percentage", "manual_brokerage", "advance_rent_month", "advance_rent_amount",
                "security_deposit_type", "security_deposit_amount",
                "maintenance_type", "monthy_maintenance_amount", "total_move_in_cost"
            ],
            "Property Location Details": [
                "address", "city", "locality_area", "property_landmark", "state",
                "pincode", "main_road_connectivity", "google_maps_link", "latitude", "longitude",
            ],
            "Amenities & Features": [
                "amenities",
            ],
            "Nearby Facilities": [
                "nearby_facilities",
            ],
            "Auto Property Summary & Description Generated Details": [
                "description",
                "rent_residential_desc",
            ],
            "Property Description(Added By User)": [
                "user_description",
            ],

            "Property Listed Elsewhere": [
                "listed_elsewhere", "portal_name",
                
            ],
           
            "Record Management": [
            "listing_status", "approval_status",
            "created_at"
            ]
        }

        HINTS = {
            "sr_no":"Auto","id": "Property ID Auto-Generated",
            "property_title": "Auto_Generated Title", "property_type": "Apartment",
            "property_no": "e.g. B-402", "bhk_type": "e.g. 2bhk",
            "renting_option": "Full Property/Single Room/Shared Room", "built_up_area": "sq.ft",
            "bathrooms": "Number", "balconies": "Number", "building_configuration": "e.g. G+3",
            "total_floors": "Number", "facing_direction": "North/East", "furnishing_status": "Semi Furnished",
            "available_for": "Family/Bachelor", "carpet_area": "sq.ft", "city_zone": "North/South", "ownership_type": "Freehold",
            "property_condition": "Resale", "property_age": "1-3 Years",
            "wing_number": "e.g. A/B", "building_name": "Text", "availability_status": "Ready to Move",
            "available_from": "YYYY-MM-DD", "lease_duration": "11 Months", "monthly_rent": "₹","brokerage_percentage": "1%/Fixed Amount",
            "manual_brokerage": "e.g. 2.5%", "advance_rent_month": "0-11/fixed",
            "advance_rent_amount": "₹", "security_deposit_type": "0-11/fixed", "security_deposit_amount": "₹",
            "maintenance_type": "Included in Rent/Extra", "monthy_maintenance_amount": "₹",
            "total_move_in_cost": "₹", "address": "Full Address", "city": "Text", "locality_area": "Text",
            "property_landmark": "Optional", "state": "e.g. Maharashtra", "pincode": "6-digit",
            "main_road_connectivity": "Optional", "google_maps_link": "URL", "latitude": "21.1458",
            "longitude": "23.1458", "amenities": "Comma-sep",
            "nearby_facilities": "Comma-sep", "description": "Short Summary",
            "rent_residential_desc": "Long Rich Text", "user_description": "Added by user",  "listed_elsewhere": "Yes/No",
            "portal_name": "e.g. 99acres, MagicBricks",
            "listed_by_name": "Full Name", "listed_by_contact": "10 Digits", "listed_by_email": "email@example.com",
            "listed_by_role": "Owner/Agent/Admin", "listed_by_type": "Self/Other",
            
            "upload_file_name": "File Name", "listing_status": "Published/Draft", "approval_status": "Approved/Pending",
           
            "created_at": "YYYY-MM-DD"
        }

        BOOL_FIELDS = {"is_deleted", "is_duplicate"}

        ROLE_BROKERAGE_LABELS = {
            "admin": "EstateFlow Service Fee",
            "relationship manager": "Service Fee",
            "landlord": "Tenant Service Fee",
            "agent": "Brokerage",
            "agency/builder": "Service Fee",
            "builder": "Service Fee",
        }

        def brokerage_label_for(role):
            if not role:
                return "Brokerage"
            return ROLE_BROKERAGE_LABELS.get(str(role).strip().lower(), "Brokerage")

        def display_value(p, field):
            val = getattr(p, field, "")
            if field in ['available_from', 'created_at', 'deleted_at'] and val:
                try:
                    return val.strftime('%Y-%m-%d')
                except AttributeError:
                    return val
            if field in BOOL_FIELDS:
                return "Yes" if val else "No"
            # NEW: listings created directly on the site (not via bulk
            # import) never get an upload_file_name — label them clearly
            # instead of leaving the cell blank.
            if field == "upload_file_name":
                return val if val else "Web Listing UI Form"
            return val if val is not None else ""

        all_cols = []
        for sec, fields in sections.items():
            all_cols.extend([(sec, f) for f in fields])

        if request.GET.get('download') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rental Residential"

            HDR_BG, REQ_BG, OPT_BG = "667EEA", "FEF3C7", "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

            sec_spans = OrderedDict()
            for i, (sec, _) in enumerate(all_cols):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            role_col = None
            brokerage_col = None

            for ci, (sec, field) in enumerate(all_cols, 1):
                req = field in REQUIRED
                label = FIELD_LABELS.get(field, field)
                lc = ws.cell(row=2, column=ci, value=label + (" *" if req else ""))
                lc.font = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border = bdr

                hc = ws.cell(row=3, column=ci, value="")
                hc.font = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(16, len(label) // 2 + 6)

                if field == "listed_by_role":
                    role_col = ci
                if field == "brokerage_percentage":
                    brokerage_col = ci

            if brokerage_col:
                note = (
                    "The LABEL text shown above this field on the live form changes based on\n"
                    "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
                    "Admin -> EstateFlow Service Fee\n"
                    "Relationship Manager -> Service Fee\n"
                    "Landlord -> Tenant Service Fee\n"
                    "Agent -> Brokerage\n"
                    "Agency/Builder or Builder -> Service Fee\n"
                    "Any other role -> Brokerage (default)\n\n"
                    "See the 'Brokerage Label' column at the end of this sheet for the resolved value per row."
                )
                ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            for row_idx, p in enumerate(properties, start=4):
                for col_idx, (sec, field) in enumerate(all_cols, 1):
                    if field == "sr_no":
                        val = row_idx - 3   # row 4 -> Sr No 1
                    else:
                        val = display_value(p, field)

                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr

            if role_col:
                preview_col = len(all_cols) + 1
                pc = ws.cell(row=2, column=preview_col, value="Brokerage Label")
                pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
                pc.fill = PatternFill("solid", fgColor="FEF3C7")
                pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                pc.border = bdr
                ws.cell(row=3, column=preview_col, value="Auto-resolved from role")
                ws.cell(row=3, column=preview_col).font = Font(italic=True, color="64748B", name="Arial", size=8)
                ws.cell(row=3, column=preview_col).border = bdr
                ws.column_dimensions[get_column_letter(preview_col)].width = 22

                for row_idx, p in enumerate(properties, start=4):
                    label = brokerage_label_for(getattr(p, "listed_by_role", ""))
                    fc = ws.cell(row=row_idx, column=preview_col, value=label)
                    fc.alignment = Alignment(horizontal="center", vertical="center")
                    fc.border = bdr

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Properties_Data.xlsx"'
            wb.save(response)
            return response

        elif request.GET.get('download') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Rental_Residential_Properties_Data.csv"'
            writer = csv.writer(response)

            row1 = []
            current_sec = ""
            for sec, _ in all_cols:
                if sec != current_sec:
                    row1.append(f" {sec}")
                    current_sec = sec
                else:
                    row1.append("")
            row1.append("📋 Brokerage")
            writer.writerow(row1)

            header_row = [FIELD_LABELS.get(field, field) + (" *" if field in REQUIRED else "") for _, field in all_cols]
            header_row.append("Brokerage Label")
            writer.writerow(header_row)

            hint_row = ["" for _, field in all_cols]
            hint_row.append("")
            writer.writerow(hint_row)

            for p in properties:
                data_row = [display_value(p, field) for _, field in all_cols]
                data_row.append(brokerage_label_for(getattr(p, "listed_by_role", "")))
                writer.writerow(data_row)

            return response

    # ═══════════════════════════════════════
    # STATS & UNIQUE DROPDOWN DATA
    # ═══════════════════════════════════════

    all_props = RentalResidentialProperty.objects.filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)

    total_count = all_props.count()

    unique_listing_status = all_props.exclude(listing_status__isnull=True).exclude(listing_status='').values_list('listing_status', flat=True).distinct()
    unique_approval_status = all_props.exclude(approval_status__isnull=True).exclude(approval_status='').values_list('approval_status', flat=True).distinct()
    unique_bhk = all_props.exclude(bhk_type__isnull=True).exclude(bhk_type='').values_list('bhk_type', flat=True).distinct()
    unique_cities = all_props.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_states = all_props.exclude(state__isnull=True).exclude(state='').values_list('state', flat=True).distinct()  # NEW
    unique_furnish = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values_list('furnishing_status', flat=True).distinct()
    unique_possession = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').values_list('availability_status', flat=True).distinct()
    unique_property_types = all_props.exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    unique_listed_roles = all_props.exclude(listed_by_role__isnull=True).exclude(listed_by_role='').values_list('listed_by_role', flat=True).distinct()   # NEW split
    unique_uploaded_roles = all_props.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()  # NEW split

    active_count = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    furnished_count = all_props.filter(furnishing_status__iexact='Furnished').count()
    available_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    city_count = all_props.exclude(city__isnull=True).exclude(city='').values('city').distinct().count()
    state_count = all_props.exclude(state__isnull=True).exclude(state='').values('state').distinct().count()  # NEW

    # Duplicate properties
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count = all_props.filter(is_duplicate=False).count()

# Listing status breakdown
    active_listing_count = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count = all_props.filter(listing_status__iexact='Rented').count()

# Approval status breakdown
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count = all_props.filter(approval_status__iexact='Rejected').count()

    # ═══════════════════════════════════════
    # RENT STATS
    # ═══════════════════════════════════════

    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )

    avg_rent = rent_stats['avg_rent']
    max_rent = rent_stats['max_rent']
    min_rent = rent_stats['min_rent']

    deposit_stats = all_props.exclude(security_deposit_amount__isnull=True).aggregate(avg_deposit=Avg('security_deposit_amount'))
    avg_deposit = deposit_stats['avg_deposit']

    area_stats = all_props.exclude(built_up_area__isnull=True).aggregate(avg_area=Avg('built_up_area'))
    avg_area = area_stats['avg_area']

    with_owner_count = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    with_images_count = all_props.filter(images__isnull=False).distinct().count()

    uploaded_files = all_props.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    # ═══════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════

    renting_option_qs = all_props.exclude(renting_option__isnull=True).exclude(renting_option='').values('renting_option').annotate(count=Count('id')).order_by('-count')
    renting_option_labels = json.dumps([item['renting_option'] for item in renting_option_qs])
    renting_option_data = json.dumps([item['count'] for item in renting_option_qs])

    rent_buckets = [
        ('Under ₹5k', 0, 5000), ('₹5k–10k', 5000, 10000), ('₹10k–20k', 10000, 20000),
        ('₹20k–30k', 20000, 30000), ('₹30k–50k', 30000, 50000), ('₹50k–1L', 50000, 100000),
        ('Above ₹1L', 100000, 999999999),
    ]

    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data = json.dumps([all_props.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values('furnishing_status').annotate(count=Count('id')).order_by('-count')
    furnishing_labels = json.dumps([item['furnishing_status'] for item in furnish_qs])
    furnishing_data = json.dumps([item['count'] for item in furnish_qs])

    prop_type_qs = all_props.exclude(property_type__isnull=True).exclude(property_type='').values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels = json.dumps([item['property_type'] for item in prop_type_qs])
    prop_type_data = json.dumps([item['count'] for item in prop_type_qs])

    # ═══════════════════════════════════════
    # KPI
    # ═══════════════════════════════════════

    occupied_count = all_props.filter(availability_status__iexact='Occupied').count()
    vacant_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100), 1) if total_count > 0 else 0
    vacancy_rate = round((vacant_count / total_count * 100), 1) if total_count > 0 else 0
    total_revenue = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0

    ready_to_move_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    short_lease_count = all_props.filter(lease_duration__icontains='6').count()
    long_lease_count = all_props.filter(lease_duration__icontains='12').count()
    new_property_count = all_props.filter(property_age__icontains='New').count()
    old_property_count = all_props.exclude(property_age__icontains='New').count()
    premium_properties_count = all_props.filter(monthly_rent__gte=50000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=15000).count()

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════

    context = {
        'user_obj': user_obj,
        'page_obj': page_obj,

        'search_query': search_query,
        'bhk_query': bhk_query,
        'city_query': city_query,
        'state_query': state_query,          # NEW
        'furnish_query': furnish_query,
        'possession_query': possession_query,
        'prop_type_query': prop_type_query,
        'listed_by_query': listed_by_query,
        'uploaded_by_query': uploaded_by_query,   # NEW
        'budget_query': budget_query,

        'unique_cities': unique_cities,
        'unique_states': unique_states,        # NEW
        'unique_furnish': unique_furnish,
        'unique_possession': unique_possession,
        'unique_property_types': unique_property_types,
        'unique_listed_roles': unique_listed_roles,     # NEW (split)
        'unique_uploaded_roles': unique_uploaded_roles, # NEW (split)

        'filtered_count': filtered_count,

        'total_count': total_count,
        'active_count': active_count,
        'furnished_count': furnished_count,
        'available_count': available_count,
        'city_count': city_count,
        'state_count': state_count,          # NEW

        'from_date': from_date_str,
        'to_date': to_date_str,

        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,

        'avg_deposit': avg_deposit,
        'avg_area': avg_area,

        'with_owner_count': with_owner_count,
        'with_images_count': with_images_count,

        'uploaded_files': uploaded_files,

        'renting_option_labels': renting_option_labels,   # renamed from bhk_labels
        'renting_option_data': renting_option_data,

        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,

        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,

        'prop_type_labels': prop_type_labels,
        'prop_type_data': prop_type_data,

        'occupied_count': occupied_count,
        'vacant_count': vacant_count,

        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,

        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,

        'ready_to_move_count': ready_to_move_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,

        'new_property_count': new_property_count,
        'old_property_count': old_property_count,

        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'bhk_query': bhk_query,
        'unique_bhk': unique_bhk,                              # ADD

        'listing_status_query': listing_status_query,          # ADD
        'approval_status_query': approval_status_query,        # ADD
        'duplicate_query': duplicate_query,                    # ADD

        'unique_listing_status': unique_listing_status,        # ADD
        'unique_approval_status': unique_approval_status,      # ADD

        'duplicate_properties_count': duplicate_properties_count,  # ADD
        'unique_properties_count': unique_properties_count,        # ADD

        'active_listing_count': active_listing_count,          # ADD
        'inactive_listing_count': inactive_listing_count,      # ADD
        'sold_listing_count': sold_listing_count,               # ADD
        'rented_listing_count': rented_listing_count,           # ADD
        'pending_approval_count': pending_approval_count,       # ADD
        'approved_count': approved_count,                       # ADD
        'rejected_count': rejected_count,                       # ADD
    }

    return render(
        request,
        'agency_panel/Reports/Rental/rental_list.html',
        context
    )




#################Views End For Rental Residential Property###########################


#################Views Start For Rental Commericial Property###########################




def commercial_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Rental/commercial.html", context)


def commercial_list_agency(request):
       # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(
        request,
        'agency_panel/Reports/Rental/commercial_list.html',
        context
    )




#################Views End For Rental Commericial  Property###########################


#################Views Start For Rental Pg_coliving Property###########################



def pg_coliving_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Rental/pg_coliving.html", context)


def pg_list_agency(request):
       # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(
        request,
        'agency_panel/Reports/Rental/pg_list.html',
        context
    )





#################Views Start For Resale Residential Listing Property###########################



def residential_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale/residential_resale.html", context)


############## Views start for resale residential view form #####################

def residential_resale_agency_view(request,id):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    property = ResaleResidentialProperty.objects.get(id=id)

    # Convert comma-separated string datasets into lists for badge generation loop arrays
    facilities_list = [f.strip() for f in property.nearby_facilities.split(',')] if property.nearby_facilities else []
    amenities_list = [a.strip() for a in property.amenities.split(',')] if property.amenities else []
 
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        "images": property.images.all(),
        "facilities_list": facilities_list,
        "amenities_list": amenities_list,
    }
    
    return render(request, "agency_panel/Forms/Resale/residential_resale_view.html", context)

########### Views end for resale residential view form ######################


def residential_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    all_properties = (
        ResaleResidentialProperty.objects
        .filter(is_deleted=False,listed_by_id=user_obj.user_id,listed_by_name=user_obj.user_name,listed_by_email=user_obj.user_email,listed_by_contact=user_obj.user_phone,listed_by_role=user_obj.user_role)
        .prefetch_related('images')
        .order_by('-created_at')
    )

    properties = all_properties

    ############# Filter Parameters Start ###################################

    # ── Read query params ────────────────────────────────────────────────────
    search_query    = request.GET.get('search', '').strip()
    prop_type       = request.GET.get('prop_type', '').strip()
    bhk_filter      = request.GET.get('bhk', '').strip()
    furnish         = request.GET.get('furnish', '').strip()
    occupancy       = request.GET.get('occupancy', '').strip()
    ownership       = request.GET.get('ownership', '').strip()
    negotiable      = request.GET.get('negotiable', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()
    listed_by_type  = request.GET.get('listed_by_type', '').strip()
    property_age    = request.GET.get('property_age', '').strip()
    facing_direction = request.GET.get('facing_direction', '').strip()
    society_type    = request.GET.get('society_type', '').strip()
    city_filter     = request.GET.get('city', '').strip()
    status_filter   = request.GET.get('listing_status', '').strip()

    # ── Apply filters ────────────────────────────────────────────────────────
    properties = all_properties

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query)           |
            Q(locality__icontains=search_query)       |
            Q(listed_by_name__icontains=search_query) |
            Q(bhk__icontains=search_query)            |
            Q(building_name__icontains=search_query)  |
            Q(property_no__icontains=search_query)    |
            Q(address__icontains=search_query)
        )

    if prop_type:
        properties = properties.filter(property_type=prop_type)

    if bhk_filter:
        properties = properties.filter(bhk=bhk_filter)

    if furnish:
        properties = properties.filter(furnishing_status=furnish)

    if occupancy:
        properties = properties.filter(occupancy_status=occupancy)

    if ownership:
        properties = properties.filter(ownership_status=ownership)

    if negotiable:
        properties = properties.filter(price_negotiable=negotiable)

    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)

    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    if listed_by_type:
        properties = properties.filter(listed_by_type=listed_by_type)

    if property_age:
        properties = properties.filter(property_age=property_age)

    if facing_direction:
        properties = properties.filter(facing_direction=facing_direction)

    if society_type:
        properties = properties.filter(society_type=society_type)

    if city_filter:
        properties = properties.filter(city__icontains=city_filter)

    if status_filter:
        properties = properties.filter(listing_status=status_filter)

    # ── Thumbnail + helper attributes for each filtered property ─────────────
    for prop in properties:
        prop.thumbnail = prop.images.first()

        prop.nearby_facilities_list = (
            [f.strip() for f in prop.nearby_facilities.split(',')]
            if prop.nearby_facilities else []
        )
        prop.amenities_list = (
            [a.strip() for a in prop.amenities.split(',')]
            if prop.amenities else []
        )
        prop.image_count = prop.images.count()
        prop.image_urls = [img.image.url for img in prop.images.all()]

    # ════════════════════════════════════════════════════════════════════════
    # KPI STATS (Calculated using correct DB layout keys)
    # ════════════════════════════════════════════════════════════════════════
    total_count = all_properties.count()

    # ── Row 1 — Inventory ────────────────────────────────────────────────────
    total_negotiable  = all_properties.filter(price_negotiable='yes').count()
    total_furnished   = all_properties.filter(furnishing_status='fully').count()
    total_freehold    = all_properties.filter(ownership_status='Self Owned').count()
    total_with_images = all_properties.filter(images__isnull=False).distinct().count()
    total_with_vacant = all_properties.filter(occupancy_status='Vacant').count()
    total_with_tenants = all_properties.filter(existing_tenants='yes').count()

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    negotiable_pct = pct(total_negotiable, total_count)
    furnished_pct  = pct(total_furnished, total_count)
    freehold_pct   = pct(total_freehold, total_count)
    images_pct     = pct(total_with_images, total_count)
    vacant_pct     = pct(total_with_vacant, total_count)

    # ── Row 2 — Pricing ──────────────────────────────────────────────────────
    price_agg = all_properties.aggregate(
        avg      = Avg('selling_price'),
        max_val  = Max('selling_price'),
        min_val  = Min('selling_price'),
        avg_sqft = Avg('price_per_sqft'),
        avg_area = Avg('builtup_area'),
        total_value = Coalesce(Sum('selling_price'), Decimal('0.00')),
    )
    avg_price      = price_agg['avg']
    max_price      = price_agg['max_val']
    min_price      = price_agg['min_val']
    avg_price_sqft = price_agg['avg_sqft']
    avg_builtup    = price_agg['avg_area']
    total_value    = price_agg['total_value']
    total_with_loan = all_properties.filter(property_loan='yes').count()

    # ── Row 3 — Legal & Status ───────────────────────────────────────────────
    no_dispute_count  = all_properties.filter(any_legal_dispute='no').count()
    dispute_count     = all_properties.filter(any_legal_dispute='yes').count()
    tax_pending_count = all_properties.filter(government_tax='yes').count()
    tenant_occupied   = all_properties.filter(existing_tenants='yes').count()
    premium_count     = all_properties.filter(selling_price__gte=10000000).count()  # >= 1 Cr
    budget_count      = all_properties.filter(selling_price__lt=3000000).count()    # < 30 L

    # ── Row 4 — Listing Quality ──────────────────────────────────────────────
    with_video_count = (
        all_properties
        .exclude(property_video__isnull=True)
        .exclude(property_video='')
        .count()
    )
    with_floor_plan = (
        all_properties
        .exclude(floor_plan__isnull=True)
        .exclude(floor_plan='')
        .count()
    )
    with_listed_by = (
        all_properties
        .exclude(listed_by_name__isnull=True)
        .exclude(listed_by_name='')
        .count()
    )
    with_uploaded_by = (
        all_properties
        .exclude(uploaded_by_name__isnull=True)
        .exclude(uploaded_by_name='')
        .count()
    )
    with_description = (
        all_properties
        .exclude(property_description__isnull=True)
        .exclude(property_description='')
        .count()
    )
    pending_approval = all_properties.filter(approval_status='Pending').count()
    approved_count   = all_properties.filter(approval_status='Approved').count()
    rejected_count   = all_properties.filter(approval_status='Rejected').count()

    # ── Charts ───────────────────────────────────────────────────────────────
    property_type_counts = dict(
        all_properties.values('property_type')
        .annotate(count=Count('id'))
        .values_list('property_type', 'count')
    )
    
    bhk_counts = dict(
        all_properties.values('bhk')
        .annotate(count=Count('id'))
        .values_list('bhk', 'count')
    )
    
    furnishing_counts = {
        'fully_furnished': all_properties.filter(furnishing_status='fully').count(),
        'semi_furnished': all_properties.filter(furnishing_status='semi').count(),
        'unfurnished': all_properties.filter(furnishing_status='unfurnished').count(),
    }
    
    occupancy_counts = dict(
        all_properties.values('occupancy_status')
        .annotate(count=Count('id'))
        .values_list('occupancy_status', 'count')
    )
    
    facing_counts = dict(
        all_properties.values('facing_direction')
        .annotate(count=Count('id'))
        .values_list('facing_direction', 'count')
    )
    
    society_type_counts = dict(
        all_properties.values('society_type')
        .annotate(count=Count('id'))
        .values_list('society_type', 'count')
    )
    
    ownership_counts = dict(
        all_properties.values('ownership_status')
        .annotate(count=Count('id'))
        .values_list('ownership_status', 'count')
    )
    
    listed_by_counts = dict(
        all_properties.values('listed_by_type')
        .annotate(count=Count('id'))
        .values_list('listed_by_type', 'count')
    )
    
    approval_status_counts = dict(
        all_properties.values('approval_status')
        .annotate(count=Count('id'))
        .values_list('approval_status', 'count')
    )

    # ── Monthly Property Listings (for trend chart) ──────────────────────────
    from django.db.models.functions import TruncMonth
    monthly_listings = (
        all_properties
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_labels = [m['month'].strftime('%b %Y') for m in monthly_listings] if monthly_listings else []
    monthly_counts = [m['count'] for m in monthly_listings] if monthly_listings else []

    # ── Unique values for Select2 searchable dropdowns ───────────────────────
    unique_prop_types = list(
        all_properties.values_list('property_type', flat=True)
        .distinct().order_by('property_type')
    )
    unique_bhk_values = list(
        all_properties.values_list('bhk', flat=True)
        .distinct().order_by('bhk')
    )
    unique_furnishings = list(
        all_properties.values_list('furnishing_status', flat=True)
        .distinct().order_by('furnishing_status')
    )
    unique_occupancies = list(
        all_properties.values_list('occupancy_status', flat=True)
        .distinct().order_by('occupancy_status')
    )
    unique_ownerships = list(
        all_properties.values_list('ownership_status', flat=True)
        .distinct().order_by('ownership_status')
    )
    unique_cities = list(
        all_properties.values_list('city', flat=True)
        .distinct().order_by('city')
    )
    unique_states = list(
        all_properties.values_list('state', flat=True)
        .distinct().order_by('state')
    )
    unique_facings = list(
        all_properties.values_list('facing_direction', flat=True)
        .distinct().order_by('facing_direction')
    )
    unique_society_types = list(
        all_properties.values_list('society_type', flat=True)
        .distinct().order_by('society_type')
    )

    # ── Uploaded Files (for tracking) ──────────────────────────────────────
    try:
        uploaded_files = (
            all_properties
            .exclude(upload_file_name__isnull=True)
            .exclude(upload_file_name='')
            .values_list('upload_file_name', flat=True)
            .distinct()
        )
    except Exception:
        uploaded_files = []

    context = {
        'user_obj': user_obj,
        'properties':properties,

        'filtered_count': properties.count(),
        'total_count': total_count,

        # Active search params
        'search_query': search_query,
        'prop_type_query': prop_type,
        'bhk_query': bhk_filter,
        'furnish_query': furnish,
        'occupancy_query': occupancy,
        'ownership_query': ownership,
        'negotiable_query': negotiable,
        'from_date': from_date,
        'to_date': to_date,
        'listed_by_type_query': listed_by_type,
        'property_age_query': property_age,
        'facing_direction_query': facing_direction,
        'society_type_query': society_type,
        'city_filter': city_filter,
        'status_filter': status_filter,

        # Row 1 — Inventory
        'total_negotiable': total_negotiable,
        'total_furnished': total_furnished,
        'total_freehold': total_freehold,
        'total_with_images': total_with_images,
        'total_with_vacant': total_with_vacant,
        'total_with_tenants': total_with_tenants,
        'negotiable_pct': negotiable_pct,
        'furnished_pct': furnished_pct,
        'freehold_pct': freehold_pct,
        'images_pct': images_pct,
        'vacant_pct': vacant_pct,

        # Row 2 — Pricing
        'avg_price': avg_price,
        'max_price': max_price,
        'min_price': min_price,
        'avg_price_sqft': avg_price_sqft,
        'avg_builtup': avg_builtup,
        'total_value': total_value,
        'total_with_loan': total_with_loan,

        # Row 3 — Legal
        'no_dispute_count': no_dispute_count,
        'dispute_count': dispute_count,
        'tax_pending_count': tax_pending_count,
        'tenant_occupied': tenant_occupied,
        'premium_count': premium_count,
        'budget_count': budget_count,
        'pending_approval': pending_approval,
        'approved_count': approved_count,
        'rejected_count': rejected_count,

        # Row 4 — Quality
        'with_video_count': with_video_count,
        'with_floor_plan': with_floor_plan,
        'with_listed_by': with_listed_by,
        'with_uploaded_by': with_uploaded_by,
        'with_description': with_description,

        # Charts
        'property_type_counts': property_type_counts,
        'bhk_counts': bhk_counts,
        'furnishing_counts': furnishing_counts,
        'occupancy_counts': occupancy_counts,
        'facing_counts': facing_counts,
        'society_type_counts': society_type_counts,
        'ownership_counts': ownership_counts,
        'listed_by_counts': listed_by_counts,
        'approval_status_counts': approval_status_counts,
        'monthly_labels': monthly_labels,
        'monthly_counts': monthly_counts,

        # Select2 unique options
        'unique_prop_types': unique_prop_types,
        'unique_bhk_values': unique_bhk_values,
        'unique_furnishings': unique_furnishings,
        'unique_occupancies': unique_occupancies,
        'unique_ownerships': unique_ownerships,
        'unique_cities': unique_cities,
        'unique_states': unique_states,
        'unique_facings': unique_facings,
        'unique_society_types': unique_society_types,

        'uploaded_files': uploaded_files,
        
    }

    return render(request, 'agency_panel/Reports/Resale/residential_resale_list.html', context)






#################Views Start For Resale Commercial Listing Property###########################

def commercial_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale/commercial_resale.html", context)


def commercial_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale/commercial_list.html', context)




#################Views Start For Resale Industrial Listing Property###########################

def industrial_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale/industrial_resale.html", context)


def industrial_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale/industrial_list.html', context)


#################Views Start For Resale Agricultural Listing Property###########################


def agricultural_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale/agricultural_resale.html", context)


def agricultural_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale/agricultural_list.html', context)






#################Views Start For Resale Plot Residential Listing Property###########################


def residential_plot_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale_Plot/residential_plot_resale.html", context)


def residential_plot_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale_Plot/residential_plot_resale_list.html', context)




#################Views Start For Resale Plot Commericial Listing Property###########################


def commercial_plot_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale_Plot/commercial_plot_resale.html", context)


def commercial_plot_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale_Plot/commercial_plot_resale_list.html', context)



#################Views Start For Resale Plot Industrial Listing Property###########################



def industrial_plot_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale_Plot/industrial_plot_resale.html", context)


def industrial_plot_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale_Plot/industrial_plot_resale_list.html', context)





#################Views Start For Resale Plot Agricultural Listing Property###########################



def agricultural_plot_resale_agency(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "agency_panel/Forms/Resale_Plot/agricultural_plot_resale.html", context)


def agricultural_plot_resale_list_agency(request):
   

      # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2.  VIP Access Control
    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Agency/Builder, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_agency and not is_valid_admin:
        return redirect('login') 

    # 3.  The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Agency's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Agency is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════
  

    # ═══════════════════════════════════════
   

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════
    context = {
        'user_obj': user_obj,
        
    }

    return render(request, 'agency_panel/Reports/Resale_Plot/agricultural_plot_resale_list.html', context)





#################Views Start For Residential Rental Agency/Builder  Listing Download & Upload excel file ###########################




def _residential_field_map():
    """Returns (sections, field_to_label, label_to_field, system_injected,
    helper_only_labels, decimal_fields, int_fields). Called fresh by both
    the download view and the import view — always local, never global."""
 
    sections = OrderedDict([
        ("Listed By", [
            ("listed_by_type",   "Listed By Type (Self/Other)", False),
            ("listed_by_id",     "Listed By Id", False),
            ("listed_by_name",   "Listed By Name", False),
            ("listed_by_email",  "Listed By Email", False),
            ("listed_by_contact","Listed By Contact", False),
            ("listed_by_role",   "Listed By Role", True),   # drives the brokerage label
        ]),
        ("Basic Information", [
            ("property_type",           "Property Type", True),
            ("property_no",             "Flat/House/Unit No", True),
            ("bhk_type",                "BHK Type", True),
            ("renting_option",          "Renting Option", True),
            ("built_up_area",           "Built-up Area (sq.ft)", True),
            ("bathrooms",                "Bathrooms", True),
            ("balconies",                "Balconies", False),
            ("building_configuration",  "Building Configuration", True),
            ("total_floors",             "Total Floors Constructed", True),
            ("facing_direction",        "Facing Direction", False),
            ("furnishing_status",       "Furnishing Status", True),
            ("available_for",           "Available For", True),
        ]),
        ("Property Details", [
            ("carpet_area",     "Carpet Area (sq.ft)", False),
            ("city_zone",       "City Zone", False),
            ("ownership_type",  "Ownership Type", False),
            ("property_condition", "Property Condition", False),
            ("property_age",    "Property Age (Years)", False),
            ("wing_number",     "Tower/Wing Number", False),
            ("building_name",   "Building/Society Name", False),
        ]),
        ("Availability Details", [
            ("availability_status",   "Availability Status", False),
            ("available_from",        "Available From", False),
            ("lease_duration",        "Lease Duration", False),
           
        ]),
        ("Pricing & Brokarage Details", [
            
            ("monthly_rent",              "Monthly Rent", True),
            ("brokerage_percentage",  "Brokerage / Service Fee", True),
            ("manual_brokerage",      "Fixed Brokerage Amount", False),
            ("advance_rent_month",        "Advance Rent Month", False),
            ("advance_rent_amount",       "Advance Rent Amount", False),
            ("security_deposit_type",     "Refundable Security Deposit", True),
            ("security_deposit_amount",   "Refundable Security Deposit Amount", False),
            ("maintenance_type",          "Maintenance Type", False),
            ("monthy_maintenance_amount", "Monthly Maintenance Amount", False),
            ("total_move_in_cost",        "Total Move In Cost", False),
        ]),
        ("Property Location Details", [
            ("address",                 "Address", True),
            ("city",                    "City", True),
            ("locality_area",           "Locality/Area", True),
            ("property_landmark",       "Property Landmark", False),
            ("state",                   "State", True),
            ("pincode",                 "Pincode", True),
            ("main_road_connectivity",  "Main Road Connectivity", False),
            # --- newly added: present on the form, were missing from the map ---
            ("google_maps_link",        "Google Maps Link", False),
            ("latitude",                 "Latitude", False),
            ("longitude",                "Longitude", False),
        ]),
        ("Amenities & Features", [
            ("amenities",          "Amenities (comma-separated)", False),
           
        ]),
        ("Nearby Facilities", [
           
            ("nearby_facilities",  "Nearby Facilities (comma-separated)", False),
        ]),
        ("Property Descriptions", [
            ("user_description", "Property Description", False),
        ]),
        ("Media & Listing Status", [
            ("listed_elsewhere", "Listed Elsewhere (Yes/No)", False),
            ("portal_name",      "Portal Name", False),
            ("created_at",          "Created At (Auto)", False),
        ]),
        
    ])
 
    field_to_label = {f: lbl for _, fields in sections.items() for f, lbl, _ in fields}
    label_to_field = {lbl.strip().lower(): f for _, fields in sections.items() for f, lbl, _ in fields}
 
    system_injected = {
        "created_at",
    }
    helper_only_labels = {"brokerage label preview (auto)"}
    decimal_fields = {"built_up_area", "carpet_area"}
    int_fields = {
        "bathrooms", "balconies", "total_floors", "monthly_rent",
        "advance_rent_amount", "security_deposit_amount",
        "monthy_maintenance_amount", "total_move_in_cost",
    }
 
    return sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields
 
 
def _sample_row_data():
    """One complete example value per column so every column in the
    downloaded template shows the expected format — nothing left blank."""
    return {
       
        "listed_by_id": "rm0943", "listed_by_name": "Vikas", "listed_by_email": "vikas@test.com",
        "listed_by_contact": "9876543210", "listed_by_role": "Relationship Manager",

        "property_type": "Apartment", "property_no": "B-402", "bhk_type": "2bhk" , "renting_option": "Full Property",
        "built_up_area": "1200", "bathrooms": "2", "balconies": "1",
        "building_configuration": "G+3", "total_floors": "4", "facing_direction": "East",
        "furnishing_status": "Semi Furnished", "available_for": "Family",
 
        "carpet_area": "950", "city_zone": "West Zone", "ownership_type": "Freehold",
        "property_condition": "Well Maintained", "property_age": "5", "wing_number": "A Wing",
        "building_name": "Green Valley Apartments",
 
        "availability_status": "Available Immediately", "available_from": "2026-07-15",
        "lease_duration": "11 Months",   # only used when brokerage_percentage = "Fixed Amount"
 
        "monthly_rent": "25000","brokerage_percentage": "1 Month Rent",
        "manual_brokerage": "", "advance_rent_month": "1", "advance_rent_amount": "",
        "security_deposit_type": "2", "security_deposit_amount": "",
        "maintenance_type": "Included in Rent", "monthy_maintenance_amount": "",
        "total_move_in_cost": "75000",
 
        "address": "Flat 402, Green Valley", "city": "Nagpur", "locality_area": "Dharampeth",
        "property_landmark": "Near Railway Station", "state": "Maharashtra", "pincode": "440010",
        "main_road_connectivity": "Within 500 Meters",
        "google_maps_link": "https://maps.google.com/?q=21.1458,79.0882",
        "latitude": "21.1458", "longitude": "79.0882",
 
        "amenities": "Wi-Fi, AC, Lift", "nearby_facilities": "Metro, Hospital",
 
        "user_description": "Spacious and well-lit apartment close to all amenities.",
 
        "listed_elsewhere": "No", "portal_name": "",
    }






def _normalize_label(raw):
    if raw is None:
        return ""
    text = str(raw).replace("\u00a0", " ")           
    text = text.replace(" *", "").strip()
    text = re.sub(r"\s+", " ", text)                  
    return text.lower()


def _find_header_row(ws, label_to_field, max_scan_rows=6):
    best_row, best_score = None, -1
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        score = sum(1 for v in row_vals if _normalize_label(v) in label_to_field)
        if score > best_score:
            best_row, best_score = r, score
    return best_row, best_score




def _get_client_ip(request):
    """Helper to safely fetch client IP address reference."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')














def _is_sample_data_row(obj_data, sample):
    """Detects a row that's still the unedited example data from the
    downloaded template (agent forgot to delete/replace it before
    uploading)."""
    match_count = 0
    total_checked = 0
    for field, sample_val in sample.items():
        sample_val = str(sample_val).strip()
        if not sample_val:
            continue
        total_checked += 1
        row_val = str(obj_data.get(field, "")).strip()
        if row_val == sample_val:
            match_count += 1
    if total_checked == 0:
        return False
    return (match_count / total_checked) >= 0.9


def _identity_conflicts_with_session(obj_data, session_identity):
    """Returns True only if the row EXPLICITLY names a different person
    than whoever is logged in. Blank fields are fine (they just mean
    "use my own info") — only a value that's typed in AND disagrees
    with the session counts as a conflict."""
    pairs = [
        ('listed_by_email', 'email'),
        ('listed_by_contact', 'contact'),
        ('listed_by_name', 'name'),
        ('listed_by_role', 'role'),
    ]
    for field, key in pairs:
        row_val = str(obj_data.get(field, '')).strip()
        if not row_val:
            continue
        session_val = str(session_identity.get(key, '')).strip()
        if row_val.lower() != session_val.lower():
            return True
    return False







# =====================================================================
# DOWNLOAD TEMPLATE (AGENCY/BUILDER)
# =====================================================================

def download_residential_template_agency(request):
    """Download the upload template for Agency/Builder users. Listed By
    identity columns (name/email/contact/role) are MANDATORY and must
    exactly match your own registered Agency profile — a row that's
    blank or names someone else is skipped on upload. The sample row
    (row 4) is protected as read-only; data-entry rows (5+) stay fully
    editable."""

    # ---- SECURITY CHECK (agency pattern) ----
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id')
    logged_in_role = request.session.get('user_type')

    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_agency and not is_valid_admin:
        return redirect('login')

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()
    sample = _sample_row_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental Residential"

    HDR_BG, REQ_BG, OPT_BG, SAMP_BG = "667EEA", "FEF3C7", "F0FDF4", "ECFDF5"
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    col = 1
    role_col = None
    brokerage_col = None

    for section, fields in sections.items():
        start_col = col
        for field, label, required in fields:
            header_text = label + (" *" if required else "")
            c1 = ws.cell(row=2, column=col, value=header_text)
            c1.font = Font(bold=True, color="1E293B", name="Arial", size=9)
            c1.fill = PatternFill("solid", fgColor=REQ_BG if required else OPT_BG)
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = bdr

            sc = ws.cell(row=4, column=col, value=sample.get(field, ""))
            sc.font = Font(name="Arial", size=9, color="065F46")
            sc.fill = PatternFill("solid", fgColor=SAMP_BG)
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sc.border = bdr
            sc.protection = Protection(locked=True)

            ws.column_dimensions[get_column_letter(col)].width = max(18, len(label) // 2 + 6)

            if field == "listed_by_role":
                role_col = col
            if field == "brokerage_percentage":
                brokerage_col = col

            if field in ("listed_by_name", "listed_by_email", "listed_by_contact", "listed_by_role"):
                ws.cell(row=2, column=col).comment = Comment(
                    "MANDATORY. Must exactly match your own registered Agency/Builder\n"
                    "profile (name / email / contact / role). Leaving this blank, or entering\n"
                    "a different person's details, will cause the row to be SKIPPED\n"
                    "on upload and flagged as an alert.",
                    "System"
                )
            if field == "listed_by_id":
                ws.cell(row=2, column=col).comment = Comment(
                    "Optional — your Listed By ID is confirmed from your session once\n"
                    "the name/email/contact/role above match your profile.",
                    "System"
                )
            if field == "listed_by_type":
                ws.cell(row=2, column=col).comment = Comment(
                    "Informational only. Every accepted row is treated as your own\n"
                    "(Self) listing once the identity fields are verified.",
                    "System"
                )

            col += 1

        end_col = col - 1
        hc = ws.cell(row=1, column=start_col, value=f"\U0001F4CB {section}")
        hc.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hc.fill = PatternFill("solid", fgColor=HDR_BG)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = PatternFill("solid", fgColor=HDR_BG)
            ws.cell(row=1, column=cc).border = bdr
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    preview_col = col
    pc = ws.cell(row=2, column=preview_col, value="Brokerage Label Preview (auto)")
    pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
    pc.fill = PatternFill("solid", fgColor="FEF3C7")
    pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pc.border = bdr
    ws.column_dimensions[get_column_letter(preview_col)].width = 26

    role_letter = get_column_letter(role_col)
    formula = (
        f"=IFERROR(INDEX('Notes - Brokerage Label'!$B$4:$B$9,"
        f"MATCH(LOWER(TRIM({role_letter}4)),'Notes - Brokerage Label'!$C$4:$C$9,0)),\"Brokerage\")"
    )
    fcell = ws.cell(row=4, column=preview_col, value=formula)
    fcell.fill = PatternFill("solid", fgColor="FEF3C7")
    fcell.font = Font(bold=True, color="92400E", name="Arial", size=9)
    fcell.alignment = Alignment(horizontal="center", vertical="center")
    fcell.protection = Protection(locked=True)

    if brokerage_col:
        note = (
            "The LABEL text shown above this field on the live form changes based on\n"
            "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
            "Admin -> EstateFlow Service Fee\n"
            "Relationship Manager -> Service Fee\n"
            "Landlord -> Tenant Service Fee\n"
            "Agent -> Brokerage\n"
            "Agency/Builder or Builder -> Service Fee\n"
            "Any other role -> Brokerage (default)\n\n"
            "See 'Notes - Brokerage Label' sheet, and the live preview column at the end of this sheet."
        )
        ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    notes = wb.create_sheet("Notes - Brokerage Label")
    notes.column_dimensions['A'].width = 26
    notes.column_dimensions['B'].width = 26
    notes.column_dimensions['C'].width = 4
    notes.sheet_view.showGridLines = False

    t = notes["A1"]
    notes.merge_cells("A1:B1")
    t.value = "Brokerage label — driven by Listed By Role"
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["Listed By Role", "Label shown on form"]
    for i, h in enumerate(hdrs, start=1):
        c = notes.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="1E293B", name="Arial")
        c.fill = PatternFill("solid", fgColor=OPT_BG)
        c.border = bdr

    role_rows = [
        ("Admin", "EstateFlow Service Fee"),
        ("Relationship Manager", "Service Fee"),
        ("Landlord", "Tenant Service Fee"),
        ("Agent", "Brokerage"),
        ("Agency/Builder", "Service Fee"),
        ("Builder", "Service Fee"),
    ]
    for r, (role, label) in enumerate(role_rows, start=4):
        notes.cell(row=r, column=1, value=role).border = bdr
        notes.cell(row=r, column=2, value=label).border = bdr
        notes.cell(row=r, column=3, value=f"=LOWER(TRIM(A{r}))")

    # ---- lock sample row, unlock data-entry rows, ACTUALLY enforce protection ----
    max_col = preview_col
    MAX_DATA_ROWS = 1000

    for c in range(1, max_col + 1):
        ws.cell(row=4, column=c).protection = Protection(locked=True)

    for r in range(5, 5 + MAX_DATA_ROWS):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.password = "estateflow2026"     # FIX: actually enforce protection
    ws.protection.formatCells = True               # FIX: was False (inverted bug)
    ws.protection.formatColumns = True              # FIX
    ws.protection.formatRows = True                 # FIX
    ws.protection.insertRows = True                  # FIX
    ws.protection.deleteRows = True                   # FIX
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Template.xlsx"'
    wb.save(response)
    return response


# =====================================================================
# IMPORT (AGENCY/BUILDER) — explicit identity MATCH required, else skip + alert
# =====================================================================

@csrf_exempt
@require_POST
def import_residential_excel_agency(request):
    excel_file = request.FILES.get("rental_file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx files allowed."}, status=400)

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()
    sample = _sample_row_data()

    REQUIRED_FIELD_KEYS = [
        'property_type', 'property_no', 'bhk_type', 'renting_option',
        'built_up_area', 'bathrooms', 'building_configuration', 'total_floors',
        'furnishing_status', 'available_for', 'monthly_rent',
        'address', 'city', 'locality_area', 'state', 'pincode',
    ]

    def _field_label(field):
        return field_to_label.get(field) or field.replace('_', ' ').title()

    def _is_missing(val):
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    # ---- 1. SECURITY CHECK (agency pattern, exactly as provided) ----
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id')
    logged_in_role = request.session.get('user_type')

    is_valid_agency = (user_id and logged_in_role == "Agency/Builder")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    if not is_valid_agency and not is_valid_admin:
        return JsonResponse({"status": "error", "message": "Not authorized. Please log in again."}, status=403)

    # ---- 2. The ID Swap ----
    if is_valid_admin:
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        dashboard_user_id = user_id

    # ---- 3. Data Fetching: the Agency whose listings this upload belongs to ----
    try:
        agency_obj = User_Details.objects.get(id=dashboard_user_id)
    except User_Details.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Could not find the Agency/Builder profile for this session.",
        }, status=400)

    admin_obj = Admin_Login.objects.filter(id=admin_id).first() if admin_id else None

    if is_valid_admin and admin_obj:
        uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
        uploader_email = getattr(admin_obj, 'email', '')
        uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
        uploader_role = "Admin"
        user_identity = uploader_email or uploader_name
    else:
        uploader_name = agency_obj.user_name
        uploader_email = agency_obj.user_email
        uploader_contact = agency_obj.user_phone
        uploader_role = "User"
        user_identity = uploader_email or uploader_name

    # Identity every row's Listed By fields must MATCH — always the Agency
    # (agency_obj), never the impersonating admin.
    session_identity = {
        'id': agency_obj.user_id,
        'name': agency_obj.user_name,
        'email': agency_obj.user_email,
        'contact': agency_obj.user_phone,
        'role': agency_obj.user_role,
    }

    # ---- 4. Parse Excel ----
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Rental Residential"] if "Rental Residential" in wb.sheetnames else wb.active
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    header_row, matched_count = _find_header_row(ws, label_to_field)

    if matched_count == 0:
        return JsonResponse({
            "status": "error",
            "message": (
                "No recognizable column headers were found in this file. "
                "Please use the official template downloaded from "
                "'Download Template' — don't rename or re-order the header row."
            ),
        }, status=400)

    raw_headers = [cell.value for cell in ws[header_row]]
    field_headers = []
    unmatched_headers = []

    for h in raw_headers:
        norm = _normalize_label(h)
        if not norm:
            field_headers.append(None)
            continue
        if norm in helper_only_labels:
            field_headers.append(None)
            continue
        field = label_to_field.get(norm)
        field_headers.append(field)
        if field is None:
            unmatched_headers.append(str(h))

    data_start_row = header_row + 1

    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []
    identity_mismatch_errors = []
    skipped_identity_mismatch = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        obj_data = {}
        for col_idx, field in enumerate(field_headers):
            if not field or field in system_injected:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip() != "":
                obj_data[field] = val

        if not obj_data:
            skipped_empty_after_mapping += 1
            continue

        if _is_sample_data_row(obj_data, sample):
            required_field_errors.append({
                "row": row_idx,
                "missing_fields": [
                    "This row still contains the example/sample data from the "
                    "downloaded template. Replace it with your actual property "
                    "details (or delete the row) before uploading."
                ],
            })
            continue

        if 'available_from' in obj_data:
            d_val = obj_data['available_from']
            if isinstance(d_val, str):
                c_str = d_val.strip().split(" ")[0]
                for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        obj_data['available_from'] = datetime.strptime(c_str, fmt).date()
                        break
                    except ValueError:
                        obj_data['available_from'] = None
            elif isinstance(d_val, datetime):
                obj_data['available_from'] = d_val.date()

        for f in int_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = int(float(str(obj_data[f]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    obj_data[f] = None

        for f in decimal_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = Decimal(str(obj_data[f]).replace(",", "").strip())
                except (InvalidOperation, ValueError):
                    obj_data[f] = None

        for f in ("advance_rent_month", "security_deposit_type"):
            if f in obj_data and obj_data[f] is not None and str(obj_data[f]).lower() != "fixed":
                try:
                    obj_data[f] = str(int(float(obj_data[f])))
                except (TypeError, ValueError):
                    obj_data[f] = str(obj_data[f]).strip()

        missing_fields = [
            _field_label(f) for f in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(f))
        ]
        if missing_fields:
            required_field_errors.append({
                "row": row_idx,
                "missing_fields": missing_fields,
            })
            continue

        if _identity_conflicts_with_session(obj_data, session_identity):
            typed_name = str(obj_data.get('listed_by_name', '')).strip()
            typed_email = str(obj_data.get('listed_by_email', '')).strip()
            typed_contact = str(obj_data.get('listed_by_contact', '')).strip()
            typed_role = str(obj_data.get('listed_by_role', '')).strip()

            if not any([typed_name, typed_email, typed_contact, typed_role]):
                reason = "Listed By details (name/email/contact/role) are blank."
            else:
                identity = " + ".join(filter(None, [typed_name, typed_email, typed_contact, typed_role])) or "Unknown"
                reason = f"Listed By '{identity}' does not match your logged-in Agency/Builder profile."

            identity_mismatch_errors.append({
                "row": row_idx,
                "errors": [f"{reason} Row skipped — this upload only accepts your own listings."],
            })
            skipped_identity_mismatch += 1
            continue

        obj_data['listed_by_id'] = session_identity['id']
        obj_data['listed_by_name'] = session_identity['name']
        obj_data['listed_by_email'] = session_identity['email']
        obj_data['listed_by_contact'] = session_identity['contact']
        obj_data['listed_by_role'] = session_identity['role']
        obj_data['listed_by_type'] = 'Self'

        parsed_rows.append({'row_idx': row_idx, 'data': obj_data})

    wb.close()

    if required_field_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields "
                "or still contain sample data. Please fix these rows and re-upload. No records were saved."
            ),
            "row_errors": required_field_errors,
        }, status=400)

    if not parsed_rows and not identity_mismatch_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"0 usable data rows found. Detected header row {header_row}, "
                f"data expected from row {data_start_row} onward. "
                f"{skipped_empty_after_mapping} row(s) had values but none matched a known column."
            ),
            "unmatched_headers": unmatched_headers,
            "header_row_detected": header_row,
            "data_start_row_assumed": data_start_row,
        }, status=400)

    file_name_exists = RentalResidentialProperty.objects.filter(
        upload_file_name=excel_file.name
    ).exists()

    created, updated, skipped, errors = (
        0, 0, skipped_empty_after_mapping + skipped_identity_mismatch, []
    )
    duplicate_blocked_rows = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality_area', '')).strip()
        input_pincode = str(o_data.get('pincode', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_name = str(o_data.get('listed_by_name', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        fingerprint_key = generate_property_fingerprint(
            input_property_no, input_building_name, input_locality, input_pincode
        )

        direct_duplicates = RentalResidentialProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality_area__iexact=input_locality
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked - Unit '{input_property_no}' is already listed "
                    f"by/for {input_listed_by_name or 'this user'}. Row skipped; edit the existing listing instead."
                )
                skipped += 1
                continue

            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id
            )

        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["listing_type"] = o_data.get("listing_type") or "Rental"
        o_data["category"] = o_data.get("category") or "Residential"

        o_data["upload_file_name"] = excel_file.name
        o_data["uploaded_by_name"] = uploader_name
        o_data["uploaded_by_email"] = uploader_email
        o_data["uploaded_by_contact"] = uploader_contact
        o_data["uploaded_by_role"] = uploader_role

        try:
            RentalResidentialProperty.objects.create(**o_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    errors.extend(duplicate_blocked_rows)

    for entry in identity_mismatch_errors:
        for msg in entry["errors"]:
            errors.append(f"Row {entry['row']}: {msg}")

    RentalActivityLog.objects.create(
        user_identity=user_identity,
        user_role=uploader_role,
        action_type='EXCEL_IMPORT',
        property_id="Multiple / Sheet Records",
        targeted_fields="bulk_action",
        associated_file=excel_file.name,
        action_payload=json.dumps({
            "filename": excel_file.name,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "duplicates_blocked": len(duplicate_blocked_rows),
            "identity_mismatches": len(identity_mismatch_errors),
            "errors_encountered": len(errors),
        }),
        ip_address=_get_client_ip(request),
        status='SUCCESS' if not errors else 'PARTIAL',
    )

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created, "updated": updated, "skipped": skipped,
        "duplicates_blocked": len(duplicate_blocked_rows),
        "identity_mismatches": len(identity_mismatch_errors),
        "error_count": len(errors), "errors": errors,
        "unmatched_headers": unmatched_headers,
        "header_row_detected": header_row,
        "data_start_row_used": data_start_row,
    })


def _identity_conflicts_with_session(obj_data, session_identity):
    """Returns True only if the row EXPLICITLY names a different person
    than whoever is logged in. Blank fields are fine (they just mean
    "use my own info") — only a value that's typed in AND disagrees
    with the session counts as a conflict."""
    pairs = [
        ('listed_by_email', 'email'),
        ('listed_by_contact', 'contact'),
        ('listed_by_name', 'name'),
        ('listed_by_role', 'role'),
    ]
    for field, key in pairs:
        row_val = str(obj_data.get(field, '')).strip()
        if not row_val:
            continue
        session_val = str(session_identity.get(key, '')).strip()
        if row_val.lower() != session_val.lower():
            return True
    return False












#################Views End For Residential Rental Agency/Builder  Listing Download & Upload excel file ###########################