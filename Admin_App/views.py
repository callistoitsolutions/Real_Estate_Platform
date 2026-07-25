from django.shortcuts import render,HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Admin_App .models import *
from Main_App .models import *
from seo .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from openpyxl import load_workbook
from django.template.loader import render_to_string
import traceback
import json
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator  # ← ADD THIS
import csv
import csv
from django.db.models import Prefetch
import json
from django.db.models import Count, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

########### Crime Officer Views#######

def _float(val):
    try:
        return float(val) if val not in (None, '') else None
    except:
        return None
def profile_update(request):
    return render(request,"admin_user/profile_update.html")

def chat(request):
    return render(request,"admin_user/chat.html")

def admin_approval_form(request):
    return render(request,"admin_user/admin_approval_form.html")

def referral_closing(request):
    return render(request,"admin_user/referral_closing.html")

def rm_register(request):
    return render(request,"admin_user/rm_register.html")


def Property_Review(request):
    return render(request,"admin_user/Property_Review.html")  

def Lead_Status_Update(request):
    return render(request,"admin_user/Lead_Status_Update.html")  

def Lead_Assignment(request):
    return render(request,"admin_user/Lead_Assignment.html")

def Wallet_Top_up(request):
    return render(request,"admin_user/Wallet_Top_up.html")

def GST_Invoice(request):
    return render(request,"admin_user/GST_Invoice.html")

def Subscription_Purchase(request):
    return render(request,"admin_user/Subscription_Purchase.html")

def other(request):
    return render(request,"admin_user/other.html")

def inquiry(request):
    return render(request,"admin_user/inquiry.html")

def commision_hold_table(request):
    return render(request,"admin_user/commision_hold_table.html")


def Commission_Hold_Release(request):
    return render(request,"admin_user/Commission_Hold_Release.html")

def dynamic_page_report(request):
    return render(request,"admin_user/dynamic_page_report.html")


def dynamic_page_edit(request):
   
    if request.method == 'POST':
        title = request.POST.get('title')
        seo_meta = request.POST.get('seo_meta')
        body = request.POST.get('body')
        image = request.FILES.get('image')
        DynamicPage.objects.create(
            title=title,
            seo_meta=seo_meta,
            body=body,
            image=image
        )
      #  return redirect('dynamicpage_list')
    #return render(request, 'dynamicpage_form.html')

    return render(request,"admin_user/dynamic_page_edit.html")


#def blog_list(request):
   # return render(request,"admin_user/blog_list.html")





def comission_structure_setup(request):
    if request.method == "POST":
        role = request.POST.get("role")
        rate_type = request.POST.get("rateType")
        commission_value = request.POST.get("commissionValue")
        deduction = request.POST.get("deduction")
        from_date = request.POST.get("fromDate")
        to_date = request.POST.get("toDate")
        release_option = request.POST.get("releaseOption")
        custom_release_date = request.POST.get("customReleaseDate")

        CommissionStructure.objects.create(
            role=role,
            rate_type=rate_type,
            commission_value=commission_value,
            deduction=deduction or None,
            from_date=from_date,
            to_date=to_date,
            release_option=release_option,
            custom_release_date=custom_release_date or None,
        )

        return redirect("comission_structure_setup")  # reload page after save

    commission_list = CommissionStructure.objects.all().order_by("-created_at")
    return render(request, "admin_user/comission_structure_setup.html", {"commission_list": commission_list})



def seo_meta_tag(request):
    if request.method == "POST":
        page_name = request.POST.get("page_name")
        meta_title = request.POST.get("meta_title")
        canonical_url = request.POST.get("canonical_url")
        meta_description = request.POST.get("meta_description")
        keywords = request.POST.get("keywords")

        # Save in DB
        SeoMetaTag.objects.create(
            page_name=page_name,
            meta_title=meta_title,
            canonical_url=canonical_url,
            meta_description=meta_description,
            keywords=keywords,
        )
        messages.success(request, "SEO Meta Tag added successfully!")
        #return redirect("seo_meta_tag_list")

    # Display saved data
    seo_tags = SeoMetaTag.objects.all()
    return render(request, "admin_user/seo_meta_tag.html", {"seo_tags": seo_tags})


def seo_meta_tag_list(request):
   # seo_tags = SEOMetaTag.objects.all().order_by("-created_at")
    return render(request, "admin_user/seo_meta_tag_list.html")


def admin_page(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,"admin_user/admin_page.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')
   

def index2(request):
    return render(request,"admin_user/index2.html")


########## Views start for impersonation url for super admin ######################

@csrf_exempt
def Impersonate(request):
    if request.method == "POST":
        target_id = request.POST.get('target_user_id')

        
        
        if target_id:
            # 1. Save ID to session securely
            request.session['impersonate_id'] = target_id
            
            target_user = User_Details.objects.get(id=target_id)
            
                
            # 2. Determine the correct URL based on their role
            if target_user.user_role == 'Relationship Manager':
                url = reverse('rm_dashboard')
            elif target_user.user_role == 'Landlord':
                url = reverse('landlord_dashboard')
            elif target_user.user_role == 'Tenant':
                url = reverse('Tenant_App:tenant_Dashboard')
            elif target_user.user_role == 'Buyer':
                url = reverse('Buyer_Dashboard')
            elif target_user.user_role == 'Agent':
                url = reverse('agent_dashboard')
            elif target_user.user_role == 'Agency/Builder':
                url = reverse('Agency_Dashboard')
            elif target_user.user_role == 'Vendor':
                url = reverse('Vendors:vendors_Dashboard')            
            
            # 3. Send the URL back to the JavaScript
            return JsonResponse({'status': 'success', 'redirect_url': url})
            
    return JsonResponse({'status': 'error', 'msg': 'Unauthorized request'})

############ Views end for impersonation url for super admin ##########################


############ Views start for live statistical tracking #####################

def get_live_traffic(request):
    if request.session.get('user_type') != 'Admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    five_minutes_ago = now() - timedelta(minutes=5)
    
    # Clean up old users
    ActiveVisitor.objects.filter(last_seen__lt=five_minutes_ago).delete()

    # Count active users from SQL
    return JsonResponse({
        'desktop': ActiveVisitor.objects.filter(device_type='desktop').count(),
        'mobile': ActiveVisitor.objects.filter(device_type='mobile').count(),
        'tablet': ActiveVisitor.objects.filter(device_type='tablet').count()
    })

############# Production Version ##############################

# from django.core.cache import cache

# def get_live_traffic(request):
#     if request.session.get('user_type') != 'Admin':
#         return JsonResponse({'error': 'Unauthorized'}, status=403)

#     # Ask Redis for all active user keys
#     active_keys = cache.keys("active_user_*")
    
#     desktop_count, mobile_count, tablet_count = 0, 0, 0

#     # Tally them up from RAM
#     for key in active_keys:
#         device = cache.get(key)
#         if device == 'desktop': desktop_count += 1
#         elif device == 'mobile': mobile_count += 1
#         elif device == 'tablet': tablet_count += 1

#     return JsonResponse({
#         'desktop': desktop_count,
#         'mobile': mobile_count,
#         'tablet': tablet_count
#     })

############ Views end for live statistical tracking #############################


############# Views start for global search ########################

def global_search(request):
    query = request.GET.get('q', '')
    results_list = []

    if len(query) >= 2:
        
        property_tables = [
            {'model': RentalResidentialProperty, 'label': 'Residential (Rent)', 'url_name': 'residential_detail'},
            {'model': CommercialRentalProperty, 'label': 'Commercial (Rent)', 'url_name': 'commercial_detail'},
            # {'model': PGColivingProperty, 'label': 'PG / Co-living', 'url_name': 'pg_detail'},
            {'model': ResaleResidentialProperty, 'label': 'Residential (Resale)', 'url_name': 'residential_resale_detail'},
            {'model': ResaleResidentialProperty, 'label': 'Commercial (Resale)', 'url_name': 'commercial_resale_detail'},
            {'model': PlotSaleProperty, 'label': 'Plot / Land', 'url_name': 'plot_detail'},
            {'model': IndustrialResaleProperty, 'label': 'Industrial', 'url_name': 'industrial_detail'},
            {'model': AgriculturalResaleProperty, 'label': 'Agricultural', 'url_name': 'agricultural_detail'},
        ]

        # 🟢 1. SEARCH PROPERTIES (By Title, Location, Price, City, Status, etc.)
        for table in property_tables:
            ModelClass = table['model']
            
            matches = ModelClass.objects.filter(            # Search by City       # Search by Rent/Price amount
                Q(uploaded_by_name=query) |          # Search by Status (e.g., "Active")
                Q(uploaded_by_email=query) |           # Search by Status (e.g., "Active")
                Q(uploaded_by_contact=query)|            # Search by Status (e.g., "Active")
                Q(uploaded_by_role=query)            # Search by Status (e.g., "Active")
                # Add as many Q() | as you want here!
            )[:3] 
            
            for match in matches:
                results_list.append({
                    'title': f"{match.title} - {match.location}",
                    'type': table['label'],
                    'url': reverse(table['url_name'], args=[match.id]) 
                })
        
        # 🟢 2. SEARCH USERS (By Name, Email, Phone, Role, etc.)
        users = User_Details.objects.filter(
            Q(user_name__icontains=query) | 
            Q(user_email__icontains=query) |
            Q(user_phone__icontains=query) |       # Search Last Name  # Search Phone Number
            Q(user_state__icontains=query) |        # Search by Role (e.g., "Tenant")
            Q(user_city__icontains=query)  |       # Search by Role (e.g., "Tenant")
            Q(user_role__icontains=query)         # Search by Role (e.g., "Tenant")
        )[:5]
        
       # 🟢 Create a map that connects the exact database role to its URLs.py name
        role_url_map = {
            'Tenant': 'Update_Tenant',     # Replace 'tenant_detail' with actual url name
            'Landlord': 'Update_Landlord', # Replace 'landlord_detail' with actual url name
            'Buyer': 'Update_Buyer',
            'Agent': 'Update_Agent',
            'Agency': 'Update_Agency',
            'Vendor': 'Update_Vendor',
            'Relationship Manager': 'Update_RM',
        }

        for user in users:
            # 🟢 Look up the correct URL name based on the user's role
            url_name = role_url_map.get(user.user_role)
            
            # If the role exists in our map, generate the real link. 
            # If not, fall back to '#' so the server doesn't crash.
            if url_name:
                final_url = reverse(url_name, args=[user.id])
            else:
                final_url = '#'

            results_list.append({
                'title': f"{user.user_name} ({user.user_email})",
                'type': user.user_role, 
                'url': final_url 
            })

        results_list = results_list[:10]

    return JsonResponse({'results': results_list})

########## Views end for global search ########################


############## Views start for notifications ########################

def get_todays_notifications(request):
    today = datetime.today()
    master_feed = []

    # 🟢 1. Create the Map linking Roles to their specific URLs
    role_url_map = {
        'Relationship Manager': 'Update_RM',             
        'Landlord': 'Update_Landlord', 
        'Tenant': 'Update_Tenant',     
        'Buyer': 'Update_Buyer',
        'Agent': 'Update_Agent',
        'Agency/Builder': 'Update_Agency',
        'Vendor': 'Update_Vendor',
    }

    # ==========================================
    # 2. FETCH NEW USERS
    # ==========================================
    recent_users = User_Details.objects.filter(user_register_date=today).order_by('-id')[:10]
    
    for user in recent_users:
        
        #  3. Look up the correct URL name based on the user's role
        url_name = role_url_map.get(user.user_role)
        user_url = '#' # Default fallback

        if url_name:
            try:    
                user_url = reverse(url_name, args=[user.id])
            except NoReverseMatch:
                pass # If URL isn't built yet, it stays '#' safely

        master_feed.append({
            'category': 'user', 
            'title': f"New {user.user_role} Registered",
            'desc': user.user_email, # Or user_name, whichever you prefer
            'timestamp': user.user_register_date, 
            'time': user.user_register_time, 
            'url': user_url #  Plugs in the dynamic URL!
        })


    # ==========================================
    # 3. FETCH NEW CONTACTS ENQUIRIES 
    # ==========================================
    recent_contacts = Contact_Enquiry.objects.filter(contact_enquiry_date=today).order_by('-contact_enquiry_date')[:10]
    for con in recent_contacts:

        contact_url = reverse("View_Contact_Enquiry", args=[con.id])
        
        master_feed.append({
            'category': 'contact', 
            'title': "New Contact Enquiry",
            'desc': f"{con.contact_name} and contact {con.contact_phone}",
            'timestamp': con.contact_enquiry_date,
            'time': con.contact_enquiry_time,
            'url': contact_url 
        })


    # ==========================================
    # 3. FETCH NEW SUBSCRIPTIONS (Optional)
    # ==========================================
    # recent_subs = Subscriptions.objects.filter(purchased_at__date=today).order_by('-purchased_at')[:10]
    # for sub in recent_subs:
    #     master_feed.append({
    #         'category': 'sub', 
    #         'title': "Plan Purchased",
    #         'desc': f"{sub.plan_name} by User ID {sub.user_id}",
    #         'timestamp': sub.purchased_at,
    #         'time': timezone.localtime(sub.purchased_at).strftime("%I:%M %p"),
    #         'url': '#' 
    #     })


    # ==========================================
    # 4. SORT AND FINALIZE
    # ==========================================
    master_feed.sort(key=lambda x: x['timestamp'], reverse=True)
    final_feed = master_feed[:10]

    return JsonResponse({
        'notifications': final_feed
    })

############# Views endd for notifications ###############################


def index3(request):
    return render(request,"admin_user/index3.html")


def data(request):
    return render(request,"admin_user/data.html")

def commercial_table(request):
    return render(request,"admin_user/commercial_table.html")

def pg_co_table(request):
    return render(request,"admin_user/pg_co_table.html")

    

def residential(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/residential.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')




@csrf_exempt
def get_user_data(request):
    assigned_to = request.POST.get('assigned_to')
    
    if not assigned_to:
        return JsonResponse({'error': 'No user selected'}, status=400)
    
    # Check if it's a self selection (just an ID without role)
    if '-' not in assigned_to:
        try:
            user_obj = User_Details.objects.get(id=assigned_to)
            data = {
                'user_id': user_obj.user_id,
                'name': getattr(user_obj, 'user_name', ''), # Fallback if naming differs
                'email': getattr(user_obj, 'user_email', ''),
                'contact': getattr(user_obj, 'user_phone', ''),
                'role': getattr(user_obj, 'user_role', '')
            }
            return JsonResponse(data)
        except User_Details.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
    
    # For other users (format: user_id-role)
    split = assigned_to.split("-")
    user_id = split[0]
    user_role = split[1]

    try:
        user_obj = User_Details.objects.get(id=user_id, user_role=user_role)
        data = {
            'user_id': user_obj.user_id,
            'name': getattr(user_obj, 'user_name', ''),
            'email': getattr(user_obj, 'user_email', ''),
            'contact': getattr(user_obj, 'user_phone', ''),
            'role': getattr(user_obj, 'user_role', '')
        }
        return JsonResponse(data)
    except User_Details.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)


def commercial(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/commercial.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


def pg_coliving(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/pg_coliving.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


############### Views start for contact enquiries list #####################

def Contact_Enquiries_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        contacts_en_obj = Contact_Enquiry.objects.all().order_by('-contact_enquiry_date')
        contacts_en_obj_count = Contact_Enquiry.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Contact/r_t_s_enquiry.html",{'contacts_en_obj':contacts_en_obj,'contacts_en_obj_count':contacts_en_obj_count})

        context = {'admin_obj':admin_obj,'contact_enquiries_list':rendered}

        return render(request,"admin_user/Contact_Enquiry/contact_enquiry.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for contact enquiries list ###########################


############ Views start for delete contact enquiry ######################

@csrf_exempt
def Delete_Contact_Enquiry(request):
    try:
        try:
            enquiry_id = request.POST.get('enquiry_id')
            Contact_Enquiry.objects.filter(id=enquiry_id).delete()
            return JsonResponse({'status':'1', 'msg':'Contact enquiry details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})



############## Views end for delete contact enquiry ######################


############ Views start for view contact enquiries ####################

def  View_Contact_Enquiry(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        contact = Contact_Enquiry.objects.get(id=id)

        context = {'admin_obj':admin_obj,'contact':contact}

        return render(request,"admin_user/Contact_Enquiry/view_enquiry.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for view contact enquiries ######################


############# Views start for upload contact enquiries data via excel #################

@csrf_exempt
def Contacts_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('contact_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({"status": "0", "msg": "Please upload .xlsx or .xls file only"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            skipped_count = 0
            
            # Get headers from row 2 (for debugging)
            headers = []
            for cell in sheet[2]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)
            
            print("Headers found:", headers)
            
            # Find column indices based on header names (flexible)
            col_indices = {
                'name': None,
                'email': None,
                'phone': None,
                'city': None,
                'enquiry_for': None,
                'property_type': None,
                'budget_range': None,
                'message': None,
                'contact_mode': None,
                'contact_time': None,
                'enquiry_date': None,
                'enquiry_time': None,
            }
            
            for idx, header in enumerate(headers):
                if not header:
                    continue
                header_lower = header.lower().strip()
                if header_lower == 'name':
                    col_indices['name'] = idx
                elif header_lower == 'email':
                    col_indices['email'] = idx
                elif header_lower == 'phone number':
                    col_indices['phone'] = idx
                elif header_lower == 'city':
                    col_indices['city'] = idx
                elif header_lower == 'enquiry for':
                    col_indices['enquiry_for'] = idx
                elif header_lower == 'property type':
                    col_indices['property_type'] = idx
                elif header_lower == 'budget range':
                    col_indices['budget_range'] = idx
                elif header_lower == 'message':
                    col_indices['message'] = idx
                elif header_lower == 'contact mode':
                    col_indices['contact_mode'] = idx
                elif header_lower == 'contact time':
                    col_indices['contact_time'] = idx
                elif header_lower == 'enquiry date':
                    col_indices['enquiry_date'] = idx
                elif header_lower == 'enquiry time':
                    col_indices['enquiry_time'] = idx
            
            print("Column indices:", col_indices)
            
            # Start from row 3 (data starts after headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip completely empty rows
                    if not any(row):
                        continue
                    
                    print(f"Row {row_idx} raw data: {row[:15] if len(row) > 14 else row}")
                    
                    # Get values using column indices (or fixed indices if not found)
                    name = row[col_indices['name']] if col_indices['name'] is not None and len(row) > col_indices['name'] else (row[2] if len(row) > 2 else None)
                    email = row[col_indices['email']] if col_indices['email'] is not None and len(row) > col_indices['email'] else (row[3] if len(row) > 3 else None)
                    phone = row[col_indices['phone']] if col_indices['phone'] is not None and len(row) > col_indices['phone'] else (row[4] if len(row) > 4 else None)
                    city = row[col_indices['city']] if col_indices['city'] is not None and len(row) > col_indices['city'] else (row[5] if len(row) > 5 else None)
                    enquiry_for = row[col_indices['enquiry_for']] if col_indices['enquiry_for'] is not None and len(row) > col_indices['enquiry_for'] else (row[6] if len(row) > 6 else None)
                    property_type = row[col_indices['property_type']] if col_indices['property_type'] is not None and len(row) > col_indices['property_type'] else (row[7] if len(row) > 7 else None)
                    budget_range = row[col_indices['budget_range']] if col_indices['budget_range'] is not None and len(row) > col_indices['budget_range'] else (row[8] if len(row) > 8 else None)
                    message = row[col_indices['message']] if col_indices['message'] is not None and len(row) > col_indices['message'] else (row[9] if len(row) > 9 else None)
                    contact_mode = row[col_indices['contact_mode']] if col_indices['contact_mode'] is not None and len(row) > col_indices['contact_mode'] else (row[10] if len(row) > 10 else None)
                    contact_time = row[col_indices['contact_time']] if col_indices['contact_time'] is not None and len(row) > col_indices['contact_time'] else (row[11] if len(row) > 11 else None)
                    enquiry_date_value = row[col_indices['enquiry_date']] if col_indices['enquiry_date'] is not None and len(row) > col_indices['enquiry_date'] else (row[12] if len(row) > 12 else None)
                    enquiry_time_value = row[col_indices['enquiry_time']] if col_indices['enquiry_time'] is not None and len(row) > col_indices['enquiry_time'] else (row[13] if len(row) > 13 else None)
                    
                    # Skip if no name
                    if not name:
                        skipped_count += 1
                        print(f"Row {row_idx}: No name found, skipping")
                        continue
                    
                    # Convert name to string and strip
                    name = str(name).strip()
                    
                    # Skip if name is a header value
                    if name.lower() in ['name', 'sr. no.', 'sr no', 'actions', 'none']:
                        skipped_count += 1
                        continue
                    
                    # Skip if name is a number (Sr. No.)
                    if name.isdigit():
                        print(f"Row {row_idx}: Name '{name}' is a number (likely Sr. No.), skipping")
                        skipped_count += 1
                        continue
                    
                    # Handle phone number
                    if phone:
                        if isinstance(phone, (int, float)):
                            phone = str(int(phone))
                        else:
                            phone = str(phone).replace('-', '').replace(' ', '').strip()
                    else:
                        print(f"Row {row_idx}: No phone number, skipping")
                        error_count += 1
                        continue
                    
                    # Split budget range
                    contact_start_budget = None
                    contact_end_budget = None
                    
                    if budget_range and budget_range not in ['-', '---', None, 'None']:
                        budget_str = str(budget_range).strip()
                        if '-' in budget_str:
                            parts = budget_str.split('-')
                            if len(parts) == 2:
                                contact_start_budget = parts[0].strip() if parts[0] else None
                                contact_end_budget = parts[1].strip() if parts[1] else None
                        else:
                            contact_start_budget = budget_str
                    
                    # Handle empty values
                    email = None if not email or str(email) in ['---', 'None', ''] else str(email).strip()
                    city = None if not city or str(city) in ['---', 'None', ''] else str(city).strip()
                    enquiry_for = None if not enquiry_for or str(enquiry_for) in ['---', 'None', ''] else str(enquiry_for).strip()
                    property_type = None if not property_type or str(property_type) in ['---', 'None', ''] else str(property_type).strip()
                    message = None if not message or str(message) in ['---', 'None', ''] else str(message).strip()
                    contact_mode = None if not contact_mode or str(contact_mode) in ['---', 'None', ''] else str(contact_mode).strip()
                    contact_time = None if not contact_time or str(contact_time) in ['---', 'None', ''] else str(contact_time).strip()
                    
                    # Handle Enquiry Date
                    enquiry_date = None
                    if enquiry_date_value and str(enquiry_date_value).strip() not in ['', '---', '-', 'None']:
                        try:
                            if isinstance(enquiry_date_value, (date, datetime)):
                                enquiry_date = enquiry_date_value.date() if isinstance(enquiry_date_value, datetime) else enquiry_date_value
                            else:
                                date_str = str(enquiry_date_value).strip()
                                if ',' in date_str:
                                    enquiry_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    enquiry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    enquiry_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                        except Exception as e:
                            print(f"Row {row_idx}: Date parse error: {e}")
                    
                    # Handle Enquiry Time
                    enquiry_time = None
                    if enquiry_time_value and str(enquiry_time_value).strip() not in ['', '---', '-', 'None']:
                        try:
                            time_str = str(enquiry_time_value).strip()
                            # Handle formats like "12:14 p.m." or "12:14 PM"
                            time_str = time_str.replace('p.m.', 'PM').replace('a.m.', 'AM').replace('.', '')
                            if 'PM' in time_str or 'AM' in time_str:
                                enquiry_time = datetime.strptime(time_str, '%I:%M %p').time()
                            else:
                                enquiry_time = datetime.strptime(time_str, '%H:%M').time()
                        except Exception as e:
                            print(f"Row {row_idx}: Time parse error: {e}")
                    
                    print(f"Row {row_idx}: Importing - Name: {name}, Phone: {phone}, Date: {enquiry_date}, Time: {enquiry_time}")
                    
                    # Create record with all fields
                    Contact_Enquiry.objects.create(
                        contact_name=name,
                        contact_phone=phone,
                        contact_email=email,
                        contact_city=city,
                        contact_en_title=enquiry_for,
                        contact_en_type=property_type,
                        contact_start_budget=contact_start_budget,
                        contact_end_budget=contact_end_budget,
                        contact_message=message,
                        contact_mode=contact_mode,
                        contact_time=contact_time,
                        contact_enquiry_date=enquiry_date,
                        contact_enquiry_time=enquiry_time
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error in row {row_idx}: {e}")
                    continue
            
            # Prepare response message
            msg = f"Successfully imported {success_count} contact enquiries."
            if error_count > 0:
                msg += f" Failed: {error_count}"
            if skipped_count > 0:
                msg += f" Skipped: {skipped_count}"
            
            return JsonResponse({
                "status": "1",
                "msg": msg,
                "success_count": success_count,
                "error_count": error_count,
                "skipped_count": skipped_count
            })
            
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({
                "status": "0",
                "msg": f"Error: {str(e)}"
            })
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})
  

############### Views end for upload contact enquiries data via excel ###################
   

############## Views start for ameneties list ##########################

def Ameneties_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all().order_by('-amenties_date')
        ameneties_obj_count = Ameneties_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Ameneties/r_t_s_ameneties.html",{'ameneties_obj':ameneties_obj,'ameneties_obj_count':ameneties_obj_count})

        context = {'admin_obj':admin_obj,'ameneties_list':rendered}
        return render(request,"admin_user/Ameneties/ameneties_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

########### Views end for ameneties list ################################


########### Views start for ajax for add/update ameneties #####################

@csrf_exempt
def Ameneties_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['amenties_date'] = datetime.today()
        data['amenties_time'] = datetime.now()
        Ameneties_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Ameneties Details added successfully"})

    # UPDATE MODE
    else:
        try:
            ameneties = Ameneties_Details.objects.get(id=data['id'])
        except Ameneties_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Ameneties Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(ameneties, key, value)

        ameneties.save()
        return JsonResponse({"status":"1", "msg" : f"Ameneties Details updated successfully"})

############ Views end for ajax for add/update ameneties #########################


############# Views start for upload ameneties data via excel ##################

@csrf_exempt
def Ameneties_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('ameneties_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            
            # Data starts from row 3
            for row in sheet.iter_rows(min_row=3, values_only=True):
                try:
                    if not any(row):
                        continue
                    
                    amenties_icon = row[2] if len(row) > 2 else None
                    amenties_name = row[3] if len(row) > 3 else None
                    added_date_value = row[4] if len(row) > 4 else None
                    
                    if not amenties_icon or not amenties_name:
                        error_count += 1
                        continue
                    
                    # Clean data
                    amenties_icon = str(amenties_icon).strip()
                    amenties_name = str(amenties_name).strip()
                    
                    # Handle date: from Excel or today's date
                    amenties_date = datetime.today()
                    if added_date_value and str(added_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(added_date_value, (date, datetime)):
                                amenties_date = added_date_value.date() if isinstance(added_date_value, datetime) else added_date_value
                            else:
                                date_str = str(added_date_value).strip()
                                if ',' in date_str:
                                    amenties_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    amenties_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except:
                            pass
                    
                    # Update or create
                    Ameneties_Details.objects.update_or_create(
                        amenties_name=amenties_name,
                        defaults={
                            "amenties_icon": amenties_icon,
                            "amenties_date": amenties_date,
                            "amenties_time": datetime.now().time()
                        }
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} amenities. Failed: {error_count}"
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

############## Views end for upload ameneties date via excel #######################


########### Views start for delete ameneties data #########################

@csrf_exempt
def Delete_Ameneties(request):
    try:
        try:
            ameneties_id = request.POST.get('ameneties_id')
            Ameneties_Details.objects.filter(id=ameneties_id).delete()
            return JsonResponse({'status':'1', 'msg':'Ameneties details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    

############ Views end for delete ameneties data ############################


########## Views start for update ameneties data ####################

def Update_Ameneties(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties = Ameneties_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'ameneties':ameneties}
        return render(request,'admin_user/Ameneties/update_ameneties.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

########### Views end for update ameneties data #####################


############# Views start for nearby facilities list #####################

def Facilities_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        facilities_obj = Facilities_Details.objects.all().order_by('-facilities_date')
        facilities_obj_count = Facilities_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Facilities/r_t_s_facilities.html",{'facilities_obj':facilities_obj,'facilities_obj_count':facilities_obj_count})

        context = {'admin_obj':admin_obj,'facilities_list':rendered}
        return render(request,"admin_user/Nearby_Facility/facilities_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for nearby facilities list ##########################


############# Views start for ajax for add/update nearby facilities ##############

@csrf_exempt
def Facilities_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['facilities_date'] = datetime.today()
        data['facilities_time'] = datetime.now()
        Facilities_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Nearby Facilities Details added successfully"})

    # UPDATE MODE
    else:
        try:
            facilities = Facilities_Details.objects.get(id=data['id'])
        except Facilities_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Facilities Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(facilities, key, value)

        facilities.save()
        return JsonResponse({"status":"1", "msg" : f"Nearby Facilities Details updated successfully"})

############# Views end for ajax for add/update nearby facilities #################


########### Views start for upload facilities data via excel ######################

@csrf_exempt
def Facilities_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('facilities_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            
            # Data starts from row 3
            for row in sheet.iter_rows(min_row=3, values_only=True):
                try:
                    if not any(row):
                        continue
                    
                    facilities_icon = row[2] if len(row) > 2 else None
                    facilities_name = row[3] if len(row) > 3 else None
                    added_date_value = row[4] if len(row) > 4 else None
                    
                    if not facilities_icon or not facilities_name:
                        error_count += 1
                        continue
                    
                    # Clean data
                    facilities_icon = str(facilities_icon).strip()
                    facilities_name = str(facilities_name).strip()
                    
                    # Handle date: from Excel or today's date
                    facilities_date = datetime.today()
                    if added_date_value and str(added_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(added_date_value, (date, datetime)):
                                facilities_date = added_date_value.date() if isinstance(added_date_value, datetime) else added_date_value
                            else:
                                date_str = str(added_date_value).strip()
                                if ',' in date_str:
                                    facilities_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    facilities_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except:
                            pass
                    
                    # Update or create
                    Facilities_Details.objects.update_or_create(
                        facilities_name=facilities_name,
                        defaults={
                            "facilities_icon": facilities_icon,
                            "facilities_date": facilities_date,
                            "facilities_time": datetime.now().time()
                        }
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} nearby facilities. Failed: {error_count}"
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

########### Views end for upload facilities data via excel ########################


############# Views start for delete facilities data ######################

@csrf_exempt
def Delete_Facilities(request):
    try:
        try:
            facilities_id = request.POST.get('facilities_id')
            Facilities_Details.objects.filter(id=facilities_id).delete()
            return JsonResponse({'status':'1', 'msg':'Facilities details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############# Views end for delete facilities data ###########################


############### Views start for update facilities data ########################

def Update_Facilities(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        facilities = Facilities_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'facilities':facilities}
        return render(request,'admin_user/Nearby_Facility/update_facilities.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for update facilities data #########################


########## Views start for vendor services list ########################

def Services_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        services_obj = Service_Type_Details.objects.all().order_by('-service_upload_date')
        services_obj_count = Service_Type_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Services/r_t_s_services.html",{'services_obj':services_obj,'services_obj_count':services_obj_count})

        context = {'admin_obj':admin_obj,'services_list':rendered}

        return render(request,"admin_user/Service_Type/service_type_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for vendor services list ###########################


########## Views start for ajax for add/update service types ###################

@csrf_exempt
def Services_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['service_upload_date'] = datetime.today()
        data['service_upload_time'] = datetime.now()
        Service_Type_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Service Type Details added successfully"})

    # UPDATE MODE
    else:
        try:
            services = Service_Type_Details.objects.get(id=data['id'])
        except Service_Type_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Service Type Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(services, key, value)

        services.save()
        return JsonResponse({"status":"1", "msg" : f"Service Type Details updated successfully"})

########## Views end for ajax for add/update service types ########################


############ Views start for upload service type details via excel ###################

@csrf_exempt
def Services_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('services_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            
            # Data starts from row 3
            for row in sheet.iter_rows(min_row=3, values_only=True):
                try:
                    if not any(row):
                        continue
                    
                    service_id = row[2] if len(row) > 2 else None
                    service_name = row[3] if len(row) > 3 else None
                    added_date_value = row[4] if len(row) > 4 else None
                    
                    if not service_id or not service_name:
                        error_count += 1
                        continue
                    
                    # Clean data
                    service_id = str(service_id).strip()
                    service_name = str(service_name).strip()
                    
                    # Handle date: from Excel or today's date
                    service_upload_date = datetime.today()
                    if added_date_value and str(added_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(added_date_value, (date, datetime)):
                                service_upload_date = added_date_value.date() if isinstance(added_date_value, datetime) else added_date_value
                            else:
                                date_str = str(added_date_value).strip()
                                if ',' in date_str:
                                    service_upload_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    service_upload_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except:
                            pass
                    
                    # Update or create
                    Service_Type_Details.objects.update_or_create(
                        service_id=service_id,
                        defaults={
                            "service_name": service_name,
                            "service_upload_date": service_upload_date,
                            "service_upload_time": datetime.now().time()
                        }
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} vendor services. Failed: {error_count}"
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

############ Views end for upload service type details via excel ######################


########### Views start for delete vendor service details ##########################

@csrf_exempt
def Delete_Services(request):
    try:
        try:
            services_id = request.POST.get('services_id')
            Service_Type_Details.objects.filter(id=services_id).delete()
            return JsonResponse({'status':'1', 'msg':'Services type details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############## Views end for delete vendor service details #########################


########## Views start for update service details ########################

def Update_Services(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        service = Service_Type_Details.objects.get(id=id)
        context = {'service':service,'admin_obj':admin_obj}

        return render(request,"admin_user/Service_Type/update_service_type.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for update service details ######################


############## Views start for normal faqs list ######################

def Faqs_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        faqs_obj = NormalFAQ.objects.all().order_by('-faq_date')
        faqs_obj_count = NormalFAQ.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_FAQ/r_t_s_faq.html",{'faqs_obj':faqs_obj,'faqs_obj_count':faqs_obj_count})

        context = {'admin_obj':admin_obj,'faqs_list':rendered}

        return render(request,"admin_user/FAQ/display_faq.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

######### Views end for normal faqs list ###############################


########### Views start for add normal faq #############################

def Add_FAQ(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        context = {'admin_obj':admin_obj}

        return render(request,"admin_user/FAQ/add_faq.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for add normal faq #########################


############## Views start for ajax for normal faq ####################

@csrf_exempt
def Faq_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['faq_date'] = datetime.today()
        data['faq_time'] = datetime.now()
        NormalFAQ.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"FAQ Details added successfully"})

    # UPDATE MODE
    else:
        try:
            faqs = NormalFAQ.objects.get(id=data['id'])
        except NormalFAQ.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'FAQ Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(faqs, key, value)

        faqs.save()
        return JsonResponse({"status":"1", "msg" : f"FAQ Details updated successfully"})

############ Views end for ajax for normal faq ##########################


############# Views start for upload faq data via excel #######################

@csrf_exempt
def Faq_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('faq_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            
            # Data starts from row 3
            for row in sheet.iter_rows(min_row=3, values_only=True):
                try:
                    if not any(row):
                        continue
                    
                    faq_question = row[2] if len(row) > 2 else None
                    faq_answer = row[3] if len(row) > 3 else None
                    added_date_value = row[4] if len(row) > 4 else None
                    
                    if not faq_question or not faq_answer:
                        error_count += 1
                        continue
                    
                    # Clean data
                    faq_question = str(faq_question).strip()
                    faq_answer = str(faq_answer).strip()
                    
                    # Handle date: from Excel or today's date
                    faq_date = datetime.today()
                    if added_date_value and str(added_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(added_date_value, (date, datetime)):
                                faq_date = added_date_value.date() if isinstance(added_date_value, datetime) else added_date_value
                            else:
                                date_str = str(added_date_value).strip()
                                if ',' in date_str:
                                    faq_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    faq_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except:
                            pass
                    
                    # Update or create
                    NormalFAQ.objects.update_or_create(
                        faq_question=faq_question,
                        faq_answer=faq_answer,
                        faq_date=faq_date,
                        faq_time=datetime.now().time()
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} FAQS. Failed: {error_count}"
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

############# Views end for upload faq data via excel #########################


############## Views start for delete faqs #########################

@csrf_exempt
def Delete_Faqs(request):
    try:
        try:
            faq_id = request.POST.get('faq_id')
            NormalFAQ.objects.filter(id=faq_id).delete()
            return JsonResponse({'status':'1', 'msg':'FAQ details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############# Views end for delete faqs ###############################


############## Views start for update faqs #######################

def Update_Faqs(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        faq = NormalFAQ.objects.get(id=id)

        context = {'admin_obj':admin_obj,'faq':faq}

        return render(request,"admin_user/FAQ/update_faq.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for update faqs ############################


############ Views start for subscription packages list ###################

def Subscriptions_Packages_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        packages_obj = Package_Details.objects.all().order_by('-package_upload_date')
        packages_obj_count = Package_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Subscription/r_t_s_packages.html",{'packages_obj':packages_obj,'packages_obj_count':packages_obj_count})

        context = {'admin_obj':admin_obj,'packages_list':rendered}

        return render(request,"admin_user/Subscription/packages_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

########## Views end for subscription packages list ########################


########### Views start for ajax for add/edit packages #####################

@csrf_exempt
def Packages_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['package_upload_date'] = datetime.today()
        data['package_upload_time'] = datetime.now()
        Package_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Package Details added successfully"})

    # UPDATE MODE
    else:
        try:
            packages = Package_Details.objects.get(id=data['id'])
        except Package_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Packages Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(packages, key, value)

        packages.save()
        return JsonResponse({"status":"1", "msg" : f"Packages Details updated successfully"})

############ Views end for ajax for add/edit packages ##########################

########### Views start for delete packages ########################

@csrf_exempt
def Delete_Packages(request):
    try:
        try:
            package_id = request.POST.get('package_id')
            Package_Details.objects.filter(id=package_id).delete()
            return JsonResponse({'status':'1', 'msg':'Packages details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    

########### Views end for delete packages ########################


############# Views start for update packages #######################

def Update_Packages(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        package = Package_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'package':package}

        return render(request,"admin_user/Subscription/update_packages.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end foor update packages ##########################

############ Views start for subscription plan types list ########################

def Subscriptions_Plans_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        plans_obj = Plan_Details.objects.all().order_by('-plan_upload_date')
        plans_obj_count = Plan_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Subscription/r_t_s_plans.html",{'plans_obj':plans_obj,'plans_obj_count':plans_obj_count})

        context = {'admin_obj':admin_obj,'plans_list':rendered}

        return render(request,"admin_user/Subscription/plans_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for subscription plan types list #####################


########## Views start for ajax for add/edit plans  #######################

@csrf_exempt
def Plans_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['plan_upload_date'] = datetime.today()
        data['plan_upload_time'] = datetime.now()
        Plan_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Plan Details added successfully"})

    # UPDATE MODE
    else:
        try:
            plans = Plan_Details.objects.get(id=data['id'])
        except Plan_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Plans Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(plans, key, value)

        plans.save()
        return JsonResponse({"status":"1", "msg" : f"Plans Details updated successfully"})

############ Views end for ajax for add/edit plans #######################


############ Views start for delete plans ########################

@csrf_exempt
def Delete_Plans(request):
    try:
        try:
            plan_id = request.POST.get('plan_id')
            Plan_Details.objects.filter(id=plan_id).delete()
            return JsonResponse({'status':'1', 'msg':'Plan details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})


########### Views end for delete plans ################################


############## Views start for update plans ########################

def Update_Plans(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        plan = Plan_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'plan':plan}

        return render(request,"admin_user/Subscription/update_plans.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

########## Views end for update plans #########################


############## Views start for subscriptions list ##########################

def Subscriptions_List(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        subscriptions_obj = Subscription_Details.objects.all().order_by('-id')
        subscriptions_obj_count = Subscription_Details.objects.all().count()

        rendered = render_to_string("admin_user/render_to_string/R_Subscription/r_t_s_subsciption.html",{'subscriptions_obj':subscriptions_obj,'subscriptions_obj_count':subscriptions_obj_count})

        context = {'admin_obj':admin_obj,'subscriptions_list':rendered}

        return render(request,"admin_user/Subscription/subscriptions_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for subscriptions list ##########################


############### Views start for add subscriptions ########################

def Add_Subscriptions(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        context = {'admin_obj':admin_obj}
        return render(request,"admin_user/Subscription/add_subscription.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for add subscriptions ###########################


############ Views start for ajax for add/update subscriptions ################

@csrf_exempt
def Subscriptions_Ajax(request):
    data = request.POST.dict()

    if data.get('id') == "":
        data.pop("id", None)        
        data['plan_upload_date'] = datetime.today()
        data['plan_upload_time'] = datetime.now()
        Subscription_Details.objects.create(**data)
        return JsonResponse({"status":"1", "msg" : f"Subscription Details added successfully"})

    # UPDATE MODE
    else:
        try:
            subscriptions = Subscription_Details.objects.get(id=data['id'])
        except Subscription_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Subscription Details not found'})


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(subscriptions, key, value)

        subscriptions.save()
        return JsonResponse({"status":"1", "msg" : f"Subscriptions Details updated successfully"})

########### Views end for ajax for add/update subscriptions ######################


############## Views start for delete subscriptions #####################

@csrf_exempt
def Delete_Subscriptions(request):
    try:
        try:
            subscription_id = request.POST.get('subscription_id')
            Subscription_Details.objects.filter(id=subscription_id).delete()
            return JsonResponse({'status':'1', 'msg':'Subscription type details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

########### Views end for delete subscriptions ########################


############## Views start for update subscriptions #########################

def Update_Subscriptions(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        subscription = Subscription_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'subscription':subscription}
        return render(request,"admin_user/Subscription/update_subscription.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for update subscriptions ##########################


########### Views start for upload subscription details via excel ###############

@csrf_exempt
def Subscriptions_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('subscriptions_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({"status": "0", "msg": "Please upload .xlsx or .xls file only"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            skipped_count = 0
            
            # Data starts from row 3 (row 1=title, row 2=headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    if not any(row):
                        continue
                    
                    # Column mapping
                    package_name = row[2] if len(row) > 2 else None
                    plan_type = row[3] if len(row) > 3 else None
                    plan_duration = row[4] if len(row) > 4 else None
                    plan_for = row[5] if len(row) > 5 else None
                    plan_base_price = row[6] if len(row) > 6 else None
                    plan_offer_price = row[7] if len(row) > 7 else None
                    plan_discount = row[8] if len(row) > 8 else None
                    plan_max_listings = row[9] if len(row) > 9 else None
                    plan_offer_start_date = row[10] if len(row) > 10 else None
                    plan_offer_end_date = row[11] if len(row) > 11 else None
                    plan_desc = row[12] if len(row) > 12 else None
                    added_date_value = row[13] if len(row) > 13 else None
                    
                    if not package_name or not plan_type:
                        skipped_count += 1
                        print(f"Row {row_idx}: Missing package_name or plan_type, skipping")
                        continue
                    
                    # Clean data - handle None values
                    package_name = str(package_name).strip()
                    plan_type = str(plan_type).strip() if plan_type else ''
                    plan_duration = str(plan_duration).strip() if plan_duration else ''
                    plan_for = str(plan_for).strip() if plan_for else ''
                    
                    # Handle numeric fields (convert to int/float)
                    try:
                        plan_base_price = float(plan_base_price) if plan_base_price else 0
                    except:
                        plan_base_price = 0
                    
                    try:
                        plan_offer_price = float(plan_offer_price) if plan_offer_price else 0
                    except:
                        plan_offer_price = 0
                    
                    try:
                        plan_discount = float(plan_discount) if plan_discount else 0
                    except:
                        plan_discount = 0
                    
                    try:
                        plan_max_listings = int(plan_max_listings) if plan_max_listings else 0
                    except:
                        plan_max_listings = 0
                    
                    # Handle description
                    plan_desc = str(plan_desc).strip() if plan_desc else ''
                    
                    # Handle dates
                    def parse_date(date_value):
                        if not date_value or str(date_value).strip() in ['', '---', '-']:
                            return None
                        try:
                            if isinstance(date_value, (date, datetime)):
                                return date_value.date() if isinstance(date_value, datetime) else date_value
                            else:
                                date_str = str(date_value).strip()
                                if ',' in date_str:
                                    return datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    return datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    return datetime.strptime(date_str, '%d/%m/%Y').date()
                        except:
                            return None
                        return None
                    
                    plan_offer_start = parse_date(plan_offer_start_date)
                    plan_offer_end = parse_date(plan_offer_end_date)
                    
                    # Handle upload date (default to today)
                    plan_upload_date = datetime.now().date()
                    if added_date_value and str(added_date_value).strip() not in ['', '---', '-']:
                        parsed = parse_date(added_date_value)
                        if parsed:
                            plan_upload_date = parsed
                    
                    print(f"Row {row_idx}: Importing - {package_name} ({plan_type})")
                    
                    # Update or create
                    obj, created = Subscription_Details.objects.update_or_create(
                        package_name=package_name,
                        defaults={
                            "plan_type": plan_type,
                            "plan_duration": plan_duration,
                            "plan_for": plan_for,
                            "plan_base_price": plan_base_price,
                            "plan_offer_price": plan_offer_price,
                            "plan_discount": plan_discount,
                            "plan_max_listings": plan_max_listings,
                            "plan_offer_start_date": plan_offer_start,
                            "plan_offer_end_date": plan_offer_end,
                            "plan_desc": plan_desc,
                            "plan_upload_date": plan_upload_date,
                            "plan_upload_time": datetime.now().time()
                        }
                    )
                    
                    success_count += 1
                    print(f"Row {row_idx}: {'Created' if created else 'Updated'} - {package_name}")
                    
                except Exception as e:
                    error_count += 1
                    print(f"Row {row_idx} error: {e}")
            
            msg = f"Successfully imported {success_count} Subscriptions."
            if error_count > 0:
                msg += f" Failed: {error_count}"
            if skipped_count > 0:
                msg += f" Skipped: {skipped_count}"
            
            return JsonResponse({
                "status": "1",
                "msg": msg,
                "success_count": success_count,
                "error_count": error_count,
                "skipped_count": skipped_count
            })
            
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

########### Views end for upload subscriptions data via excel ######################


############  Views start for rental property list ########################



############ Views end for rental property list ###########################


########### Views start for commercial property list ###################







# ─────────────────────────────────────────────
#  Helper converters
# ─────────────────────────────────────────────

def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def _int(val):
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None

def _float(val):
    try:
        return float(str(val).strip())
    except (TypeError, ValueError):
        return None

def _bool(val):
    if val is None:
        return False
    return str(val).strip().lower() in ('true', '1', 'yes')

def _date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except ValueError:
        try:
            return datetime.strptime(str(val).strip(), "%d-%m-%Y").date()
        except ValueError:
            return None


# ─────────────────────────────────────────────
#  Column map — matches Excel template exactly
# ─────────────────────────────────────────────



# ─────────────────────────────────────────────
#  Commercial List View
# ─────────────────────────────────────────────









def export_commercial_rent(request):
    """Dedicated view for exporting commercial properties to CSV or Excel.
       Includes EVERY database field for full backup and seamless re-upload."""
    
    # ── 1. Re-apply the same search filters so the export matches the screen ──
    try:
        properties = CommercialRentalProperty.objects.filter(is_deleted=False).order_by('-id')
    except Exception:
        properties = CommercialRentalProperty.objects.all().order_by('-id')

    search_query     = request.GET.get('search', '').strip()
    prop_type_query  = request.GET.get('property_type', '').strip()
    city_query       = request.GET.get('city', '').strip()
    zone_query       = request.GET.get('zone_type', '').strip()
    possession_query = request.GET.get('possession', '').strip()
    listed_by_query  = request.GET.get('listed_by', '').strip()
    budget_query     = request.GET.get('budget', '').strip()
    from_date        = request.GET.get('from_date', '').strip()
    to_date          = request.GET.get('to_date', '').strip()

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query)
        )
    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__icontains=prop_type_query)
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if zone_query and zone_query != 'All Zones':
        properties = properties.filter(zone_type__icontains=zone_query)
    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__icontains=possession_query)
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__icontains=listed_by_query)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)
    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_25k': properties = properties.filter(monthly_rent__lt=25000)
        elif budget_query == '25k_1L': properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=100000)
        elif budget_query == '1L_5L': properties = properties.filter(monthly_rent__gte=100000, monthly_rent__lte=500000)
        elif budget_query == 'above_5L': properties = properties.filter(monthly_rent__gt=500000)

    # ── 2. EXHAUSTIVE FIELD MAPPING (Every DB Field Included) ──
    # Format: (Section Name, Database Field Name, Is Required, Exact Header Name)
    EXPORT_COLS = [
        ("Sr.No", "sr_no", False, "Sr. No"),
        

        ("Listed By Details", "listed_by_type", False, "Listed By Type"),
        
        ("Listed By Details", "listed_by_id", False, "Listed By Id"),
        ("Listed By Details", "listed_by_name", False, "Listed By Name"),
        ("Listed By Details", "listed_by_email", False, "Listed By Email"),
        ("Listed By Details", "listed_by_contact", False, "Listed By Contact"),
        ("Listed By Details", "listed_by_role", True, "Listed By Role"),

        ("Basic Information", "id", False, "Property ID"),
        ("Basic Information", "property_title", False, "Property Title"),
        

        ("Basic Information", "property_type", True, "Property Type"),
        ("Basic Information", "building_name", True, "Building / Project Name"),
        ("Basic Information", "wing_number", False, "Tower/Wing Number"),
        ("Basic Information", "property_no", True, "Shop No/Godown/Unit No"),
        ("Basic Information", "availability_status", True, "Availability Status"),
        ("Basic Information", "available_from", False, "Available From"),
        ("Basic Information", "property_age", True, "Age of Property(In Years)"),
        ("Basic Information", "zone_type", False, "Zone Type"),
        ("Basic Information", "ownership_type", True, "Ownership Type"),
        ("Basic Information", "property_condition", False, "Property Condition"),

        ("Basic Information", "builtup_area", True, "Built-up Area (sq.ft)"),
        ("Basic Information", "carpet_area", False, "Carpet Area (sq.ft)"),
       

        ("Property Location Details", "address", True, "Property Address"),
        ("Property Location Details", "locality", True, "Area/Locality"),
        ("Property Location Details", "property_landmark", False, "Property Landmark"),
        ("Property Location Details", "city", True, "City"),
        ("Property Location Details", "state", True, "State"),
        ("Property Location Details", "location_hub", False, "Location Hub"),
        ("Property Location Details", "pincode", True, "Pincode"),
        ("Property Location Details", "google_maps_link", False, "Google Maps Link"),
        ("Property Location Details", "latitude", False, "Latitude"),
        ("Property Location Details", "longitude", False, "Longitude"),

        ("Property Specifications", "dg_ups_included", False, "DG & UPS Included?"),
        ("Property Specifications", "electricity_included", False, "Electricity Included?"),
        ("Property Specifications", "water_included", False, "Water Included?"),
        ("Property Specifications", "building_configuration", False, "Building Configuration"),
        ("Property Specifications", "total_floors", True, "Total Floors Constructed"),
        ("Property Specifications", "staircases", False, "Staircases"),
        ("Property Specifications", "passenger_lifts", False, "Passenger Lifts"),
        ("Property Specifications", "service_lifts", False, "Service Lifts"),
        ("Property Specifications", "private_parking", False, "Private Parking"),
        ("Property Specifications", "min_seats", False, "Minimum Seats Occupancy"),
        ("Property Specifications", "max_seats", False, "Maximum Seats Occupancy"),
        ("Property Specifications", "cabins", False, "Cabins"),
        ("Property Specifications", "meeting_rooms", False, "Meeting Rooms"),
        ("Property Specifications", "private_washroom", False, "Private Washrooms"),
        ("Property Specifications", "public_washroom", False, "Public Washrooms"),
        ("Property Specifications", "flooring_type", False, "Flooring Type"),

        ("Pricing Details", "monthly_rent", True, "Monthly Rent"),
        ("Pricing Details", "brokerage_percentage", True, "Brokerage"),
        ("Pricing Details", "manual_brokerage", False, "Enter Fixed Brokerage"),
        ("Pricing Details", "advanced_rent_type", True, "Advanced Rent Month"),
        ("Pricing Details", "advanced_rent_amount", False, "Advance Rent Amount"),
        ("Pricing Details", "security_deposit_type", True, "Refundable Security Deposit"),
        ("Pricing Details", "security_deposit_amount", False, "Refundable Security Deposit Amount"),
        ("Pricing Details", "maintenance_type", False, "Maintenance Type"),
        ("Pricing Details", "maintenance_charges", False, "Monthly Maintenance Amount"),
        ("Pricing Details", "total_move_in_cost", False, "Total Move In Cost"),
        ("Pricing Details", "negotiable", False, "Negotiable"),
        ("Pricing Details", "lockin_period", False, "Lock-in Period (months)"),
        ("Pricing Details", "rent_increase", False, "Rent Increase (%/year)"),

        ("Amenities & Facilities", "amenities", False, "Amenities"),
        ("Amenities & Facilities", "nearby_facilities", False, "Nearby Facilities"),
        
        
        ("Property Descriptions", "property_summary", False, "System Property Summary(Auto)"),
        ("Property Descriptions", "property_description", False, "System Property Description(Auto)"),
        ("Property Descriptions", "user_description", False, "Property Description(By User)"),

        ("Media & Listing Status", "listed_elsewhere", False, "Is Property Already Listed Elsewhere?"),
        ("Media & Listing Status", "portal_name", False, "Portal Name"),
        ("Media & Listing Status", "listing_status", False, "Listing Status"),
        ("Media & Listing Status", "approval_status", False, "Approval Status"),

        ("Data Uploadeded Via", "upload_file_name", False, "Upload File Name"),
        ("Property Uploaded By", "uploaded_by_name", False, "Uploaded By Name"),
        ("Property Uploaded By", "uploaded_by_email", False, "Uploaded By Email"),
        ("Property Uploaded By", "uploaded_by_contact", False, "Uploaded By Contact"),
        ("Property Uploaded By", "uploaded_by_role", False, "Uploaded By Role"),

        ("Database Audit", "created_at", False, "Created At"),
        ("Database Audit", "updated_at", False, "Updated At"),
        ("Database Audit", "is_deleted", False, "Is Deleted"),
        ("Database Audit", "deleted_at", False, "Deleted At"),
        ("Database Audit", "deleted_by", False, "Deleted By"),
        ("Database Audit", "is_duplicate", False, "Is Duplicate"),
        ("Database Audit", "duplicate_count", False, "Duplicate Count"),
        ("Database Audit", "duplicate_group_id", False, "Duplicate Group ID"),
        ("Database Audit", "property_unique_key", False, "Property Unique Key"),
    ]

    export_format = request.GET.get('format', 'excel')
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Commercial_Rental_Full_Export_{timestamp}"

    # Helper for formatting values
    def format_val(val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(val, bool):
            return "Yes" if val else "No"
        elif val is None:
            return ""
        return str(val).strip()

    # ── 3. EXCEL EXPORT ──
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commercial Rental Database"

        HDR_BG  = "667EEA"
        REQ_BG  = "FEF3C7"
        OPT_BG  = "F0FDF4"
        thin = Side(style="thin", color="CBD5E1")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sec_spans = OrderedDict()
        for i, (sec, *_) in enumerate(EXPORT_COLS):
            sec_spans.setdefault(sec, []).append(i + 1)

        # Write Section Headers (Row 1)
        for sec, cols in sec_spans.items():
            c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
            
            for col_idx in cols[1:]:
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor=HDR_BG)
                ws.cell(row=1, column=col_idx).border = bdr
                
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

        # Write Field Headers (Row 2) - Exact string matches for importer!
        for ci, (_, _, req, header_name) in enumerate(EXPORT_COLS, 1):
            header_text = header_name + (" *" if req else "")
            lc = ws.cell(row=2, column=ci, value=header_text)
            lc.font = Font(bold=True, size=9)
            lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
            lc.border = bdr
            lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(header_text) + 5)

        # Write Data (Rows 3+)
        for row_idx, prop in enumerate(properties, start=3):
            for col_idx, (_, field, _, _) in enumerate(EXPORT_COLS, 1):
                
                # Custom logic for Serial Number and Upload File Name
                if field == "sr_no":
                    val = row_idx - 2
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center")

        # Layout adjustments
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # ── 4. CSV EXPORT ──
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        
        # Write exact Field Headers
        headers = [header_name + (" *" if req else "") for (_, _, req, header_name) in EXPORT_COLS]
        writer.writerow(headers)

        # Write Data Rows
        for row_idx, prop in enumerate(properties, start=1):
            row_data = []
            for (_, field, _, _) in EXPORT_COLS:
                if field == "sr_no":
                    val = row_idx
                elif field == "upload_file_name":
                    raw_val = getattr(prop, field, "")
                    val = "Web UI Listing Form" if not raw_val or raw_val.strip() == "" else raw_val
                else:
                    val = format_val(getattr(prop, field, ""))
                    
                row_data.append(val)
            writer.writerow(row_data)

        return response



from django.db.models.functions import TruncMonth

def commercial_list(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return render(request, 'home_page/Adminlogin.html')

    print(">>> COMMERCIAL LIST VIEW IS RUNNING! <<<")

    # ── 1. GET ALL SEARCH PARAMS ──
    search_query    = request.GET.get('search', '').strip()
    prop_type_query = request.GET.get('property_type', '').strip()
    city_query      = request.GET.get('city', '').strip()
    zone_query      = request.GET.get('zone_type', '').strip()
    possession_query= request.GET.get('possession', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()
    budget_query    = request.GET.get('budget', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()

    # ── Base queryset ──
    try:
        properties = CommercialRentalProperty.objects.filter(is_deleted=False).order_by('-id')
    except Exception:
        properties = CommercialRentalProperty.objects.all().order_by('-id')

    # ── 2. APPLY FILTERS ──
    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(listed_by_name__icontains=search_query)
        )

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__icontains=prop_type_query)
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if zone_query and zone_query != 'All Zones':
        properties = properties.filter(zone_type__icontains=zone_query)
    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__icontains=possession_query)
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__icontains=listed_by_query)

    if from_date:
        try:
            properties = properties.filter(created_at__date__gte=from_date)
        except Exception:
            pass
    if to_date:
        try:
            properties = properties.filter(created_at__date__lte=to_date)
        except Exception:
            pass

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_25k':
            properties = properties.filter(monthly_rent__lt=25000)
        elif budget_query == '25k_1L':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=100000)
        elif budget_query == '1L_5L':
            properties = properties.filter(monthly_rent__gte=100000, monthly_rent__lte=500000)
        elif budget_query == 'above_5L':
            properties = properties.filter(monthly_rent__gt=500000)

    properties = properties.prefetch_related(
    Prefetch(
        'walkthrough_video',
        queryset=CommercialRentalVideo.objects.all(),
        to_attr='all_videos'
    )
)

    # ── Exhaustive Field Definition for Exports ──
    EXPORT_COLS = [
        ("System Data", "id", False, "Database ID"),
        ("System Data", "property_unique_key", False, "Duplicate Checker Key"),
        ("System Data", "is_duplicate", False, "True/False"),
        ("System Data", "property_title", False, "Auto Generated Title"),
        
        ("Basic Info", "property_type", True, "office-space / shop / warehouse / industrial / land"),
        ("Basic Info", "property_condition", True, "bare-shell / warm-shell / fitted / furnished"),
        ("Basic Info", "city", True, "City name"),
        ("Basic Info", "locality", True, "Area/locality"),
        ("Basic Info", "address", True, "Complete address"),
        ("Basic Info", "building_name", True, "Building/project name"),
        ("Basic Info", "wing_number", False, "Wing/Tower"),
        ("Basic Info", "property_no", True, "Internal Unit No"),
        ("Basic Info", "availability_status", True, "Available Status"),
        ("Basic Info", "available_from", False, "YYYY-MM-DD"),
        ("Basic Info", "property_age", True, "Age in Years"),
        ("Basic Info", "zone_type", False, "industrial / commercial / mixed_use"),
        ("Basic Info", "location_hub", False, "it_park / business_district"),
        ("Basic Info", "ownership_type", True, "freehold / leasehold"),
        
        ("Area & Pricing", "builtup_area", True, "Number in sq.ft"),
        ("Area & Pricing", "carpet_area", False, "Number in sq.ft"),
        ("Area & Pricing", "monthly_rent", True, "Monthly rent in ₹"),
        ("Area & Pricing", "advanced_rent_type", False, "Advance Type"),
        ("Area & Pricing", "advanced_rent_amount", False, "Advance Amount"),
        ("Area & Pricing", "security_deposit_type", False, "Deposit Type"),
        ("Area & Pricing", "security_deposit_amount", False, "Deposit Amount"),
        ("Area & Pricing", "maintenance_type", False, "Included/Extra"),
        ("Area & Pricing", "maintenance_charges", False, "Monthly maintenance in ₹"),
        ("Area & Pricing", "total_move_in_cost", False, "Total ₹"),
        ("Area & Pricing", "negotiable", False, "Yes / No"),
        ("Area & Pricing", "brokerage_percentage", False, "Brokerage Terms"),
        ("Area & Pricing", "manual_brokerage", False, "Manual Amount"),
        ("Area & Pricing", "dg_ups_included", False, "true / false"),
        ("Area & Pricing", "electricity_included", False, "true / false"),
        ("Area & Pricing", "water_included", False, "true / false"),
        ("Area & Pricing", "lockin_period", False, "Lock-in months"),
        ("Area & Pricing", "rent_increase", False, "% per year"),
        
        ("Building", "building_configuration", False, "e.g., G+3"),
        ("Building", "total_floors", False, "Total floors"),
        ("Building", "staircases", False, "Number of staircases"),
        ("Building", "passenger_lifts", False, "Number (0 if none)"),
        ("Building", "service_lifts", False, "Number (0 if none)"),
        ("Building", "private_parking", False, "Private parking spots"),
        ("Building", "min_seats", False, "Min seating"),
        ("Building", "max_seats", False, "Max seating"),
        ("Building", "cabins", False, "Number of cabins"),
        ("Building", "meeting_rooms", False, "Number of meeting rooms"),
        ("Building", "private_washroom", False, "Number (0 if none)"),
        ("Building", "public_washroom", False, "Number (0 if none)"),
        ("Building", "flooring_type", False, "marble / vitrified / granite / wooden / ceramic"),
        
        ("Amenities", "amenities", True, "Comma-separated"),
        ("Amenities", "nearby_facilities", True, "Comma-separated"),
        ("Amenities", "property_summary", False, "Short Summary"),
        ("Amenities", "property_description", False, "Auto Generated Detailed description"),
        ("Amenities", "user_description", False, "User Added Detailed description"),
        
        ("Media & Contact", "listed_by_type", True, "Self/Other"),
        ("Media & Contact", "listed_by_name", True, "Full name"),
        ("Media & Contact", "listed_by_contact", True, "Contact Number"),
        ("Media & Contact", "listed_by_email", True, "Email"),
        ("Media & Contact", "listed_by_role", True, "Role"),
        
        ("Uploader Tracking", "uploaded_by_name", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_email", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_contact", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_role", False, "Auto-filled"),
        
        ("System Tracking", "created_at", False, "Datetime created"),
        ("System Tracking", "is_deleted", False, "True/False"),
        ("System Tracking", "upload_file_name", False, "Source Form"),
    ]

    # ════════════════════════════════════════════════
    # 🛑 ROBUST EXPORT LOGIC 🛑
    # ════════════════════════════════════════════════

    # ── Excel Download ──
    if request.GET.get('download') == 'excel':
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Commercial Rental"

            HDR_BG  = "667EEA"
            REQ_BG  = "FEF3C7"
            OPT_BG  = "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Row 1: Section Banners
            sec_spans = OrderedDict()
            for i, (sec, *_) in enumerate(EXPORT_COLS):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill      = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border    = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            # Rows 2 & 3: Labels and Hints
            for ci, (sec, field, req, hint) in enumerate(EXPORT_COLS, 1):
                lc = ws.cell(row=2, column=ci, value=field + (" *" if req else ""))
                lc.font      = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill      = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border    = bdr

                hc = ws.cell(row=3, column=ci, value=hint)
                hc.font      = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill      = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border    = bdr

                ws.column_dimensions[get_column_letter(ci)].width = max(18, len(field) + 4)

            # Rows 4+: Database Data Injection
            for row_idx, prop in enumerate(properties, start=4):
                for col_idx, (_, field_name, _, _) in enumerate(EXPORT_COLS, start=1):
                    val = getattr(prop, field_name, "")
                    
                    if val is True: val = "Yes"
                    elif val is False: val = "No"
                    elif hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d %H:%M')
                    elif isinstance(val, list): val = ", ".join(map(str, val))
                    elif hasattr(val, 'url'): val = val.url if val else ""
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = bdr

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            response = HttpResponse(
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="Commercial_Rental_Listing_Export.xlsx"'
            return response

        except Exception as e:
            error_msg = f"ERROR GENERATING EXCEL:\n\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            return HttpResponse(f"<pre>{error_msg}</pre>", status=500)

    # ── CSV Download ──
    if request.GET.get('download') == 'csv':
        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="commercial_rental_properties.csv"'
            writer = csv.writer(response)
            
            writer.writerow([field for _, field, _, _ in EXPORT_COLS])
            
            for prop in properties:
                row_data = []
                for _, field_name, _, _ in EXPORT_COLS:
                    val = getattr(prop, field_name, "")
                    if val is True: val = "Yes"
                    elif val is False: val = "No"
                    elif hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d %H:%M')
                    elif isinstance(val, list): val = ", ".join(map(str, val))
                    elif hasattr(val, 'url'): val = val.url if val else ""
                    row_data.append(val if val is not None else "")
                writer.writerow(row_data)
            
            return response
            
        except Exception as e:
            error_msg = f"ERROR GENERATING CSV:\n\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            return HttpResponse(f"<pre>{error_msg}</pre>", status=500)

    # ── Pagination ──
    paginator    = Paginator(properties, 10)
    page_number  = request.GET.get('page', 1)
    page_obj     = paginator.get_page(page_number)
    filtered_count = properties.count()

    for prop in page_obj:
        videos_by_source = {v.source: v for v in prop.all_videos}
        prop.uploaded_video = videos_by_source.get('uploaded')
        prop.auto_video     = videos_by_source.get('auto')
        prop.rm_video       = videos_by_source.get('rm_assisted')

    # ════════════════════════════════════════════════
    # ALL-PROPS STATS (always on unfiltered dataset)
    # ════════════════════════════════════════════════
    try:
        all_props = CommercialRentalProperty.objects.filter(is_deleted=False)
    except Exception:
        all_props = CommercialRentalProperty.objects.all()

    total_count = all_props.count()

    # ── Dropdown unique values ──
    unique_property_types = (all_props
        .exclude(property_type__isnull=True).exclude(property_type='')
        .values_list('property_type', flat=True).distinct().order_by('property_type'))
    unique_cities = (all_props
        .exclude(city__isnull=True).exclude(city='')
        .values_list('city', flat=True).distinct().order_by('city'))
    unique_zones = (all_props
        .exclude(zone_type__isnull=True).exclude(zone_type='')
        .values_list('zone_type', flat=True).distinct().order_by('zone_type'))
    unique_possession = (all_props
        .exclude(availability_status__isnull=True).exclude(availability_status='')
        .values_list('availability_status', flat=True).distinct())
    unique_roles = (all_props
        .exclude(listed_by_role__isnull=True).exclude(listed_by_role='')
        .values_list('listed_by_role', flat=True).distinct())

    # ── Occupancy KPIs ──
    active_count   = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    occupied_count = all_props.filter(availability_status__iexact='currently_occupied').count()
    vacant_count   = all_props.filter(availability_status__iexact='available_immediately').count()
    occupancy_rate = round((occupied_count / total_count * 100)) if total_count > 0 else 0
    vacancy_rate   = round((vacant_count   / total_count * 100)) if total_count > 0 else 0

    # ── Revenue KPIs ──
    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )
    avg_rent = rent_stats['avg_rent'] or 0
    max_rent = rent_stats['max_rent'] or 0
    min_rent = rent_stats['min_rent'] or 0
    total_revenue          = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0
    avg_deposit = (all_props.exclude(security_deposit_amount__isnull=True)
                   .aggregate(avg=Avg('security_deposit_amount'))['avg'] or 0)

    try:
        avg_area = (all_props.exclude(builtup_area__isnull=True)
                    .aggregate(avg=Avg('builtup_area'))['avg'] or 0)
    except Exception:
        avg_area = 0

    # ── Business KPIs ──
    premium_properties_count    = all_props.filter(monthly_rent__gte=100000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=25000).count()
    short_lease_count           = all_props.filter(lockin_period=6).count()
    long_lease_count            = all_props.filter(lockin_period=12).count()
    with_owner_count            = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    city_count                  = (all_props.exclude(city__isnull=True).exclude(city='')
                                   .values('city').distinct().count())
    try:
        with_images_count = all_props.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    # ── Percentages ──
    verified_pct = round((with_owner_count     / total_count * 100)) if total_count > 0 else 0
    image_pct    = round((with_images_count    / total_count * 100)) if total_count > 0 else 0
    premium_pct  = round((premium_properties_count / total_count * 100)) if total_count > 0 else 0

    # ── Uploaded file names (for bulk delete) ──
    try:
        uploaded_files = (all_props.exclude(upload_file_name__isnull=True)
                          .exclude(upload_file_name='')
                          .values_list('upload_file_name', flat=True).distinct())
    except Exception:
        uploaded_files = []

    # ════════════════════════════════════════════
    # CHART DATA  (JSON → template → JS)
    # ════════════════════════════════════════════

    prop_type_dist = list(
        all_props.exclude(property_type__isnull=True).exclude(property_type='')
        .values('property_type').annotate(cnt=Count('id')).order_by('-cnt')[:8]
    )
    prop_type_labels_json = json.dumps([x['property_type'] for x in prop_type_dist])
    prop_type_counts_json = json.dumps([x['cnt']           for x in prop_type_dist])

    rent_range_data = {
        'Under ₹25k':  all_props.filter(monthly_rent__lt=25000).count(),
        '₹25k–1L':     all_props.filter(monthly_rent__gte=25000,  monthly_rent__lt=100000).count(),
        '₹1L–5L':      all_props.filter(monthly_rent__gte=100000, monthly_rent__lt=500000).count(),
        'Above ₹5L':   all_props.filter(monthly_rent__gte=500000).count(),
    }
    rent_range_labels_json = json.dumps(list(rent_range_data.keys()))
    rent_range_counts_json = json.dumps(list(rent_range_data.values()))

    occupancy_json = json.dumps([occupied_count, vacant_count, max(0, total_count - occupied_count - vacant_count)])

    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_qs = (
        all_props.filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(revenue=Sum('monthly_rent'), cnt=Count('id'))
        .order_by('month')
    )
    monthly_labels_json  = json.dumps([x['month'].strftime('%b %Y') for x in monthly_qs])
    monthly_revenue_json = json.dumps([float(x['revenue'] or 0) for x in monthly_qs])

    total_tenants    = occupied_count
    collection_rate  = 0
    pending_payments = 0
    maintenance_req  = 0

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,

        'search_query':     search_query,
        'prop_type_query':  prop_type_query,
        'city_query':       city_query,
        'zone_query':       zone_query,
        'possession_query': possession_query,
        'listed_by_query':  listed_by_query,
        'budget_query':     budget_query,
        'from_date':        from_date,
        'to_date':          to_date,

        'unique_property_types': unique_property_types,
        'unique_cities':         unique_cities,
        'unique_zones':          unique_zones,
        'unique_possession':     unique_possession,
        'unique_roles':          unique_roles,
        'uploaded_files':        uploaded_files,

        'filtered_count':            filtered_count,
        'total_count':               total_count,
        'active_count':              active_count,
        'occupied_count':            occupied_count,
        'vacant_count':              vacant_count,
        'occupancy_rate':            occupancy_rate,
        'vacancy_rate':              vacancy_rate,
        'avg_rent':                  avg_rent,
        'max_rent':                  max_rent,
        'min_rent':                  min_rent,
        'total_revenue':             total_revenue,
        'total_security_deposit':    total_security_deposit,
        'avg_deposit':               avg_deposit,
        'avg_area':                  avg_area,
        'premium_properties_count':  premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'ready_to_move_count':       vacant_count,
        'short_lease_count':         short_lease_count,
        'long_lease_count':          long_lease_count,
        'with_owner_count':          with_owner_count,
        'with_images_count':         with_images_count,
        'city_count':                city_count,
        'verified_pct':              verified_pct,
        'image_pct':                 image_pct,
        'premium_pct':               premium_pct,

        'total_tenants':    total_tenants,
        'collection_rate':  collection_rate,
        'pending_payments': pending_payments,
        'maintenance_req':  maintenance_req,

        'prop_type_labels_json':  prop_type_labels_json,
        'prop_type_counts_json':  prop_type_counts_json,
        'rent_range_labels_json': rent_range_labels_json,
        'rent_range_counts_json': rent_range_counts_json,
        'occupancy_json':         occupancy_json,
        'monthly_labels_json':    monthly_labels_json,
        'monthly_revenue_json':   monthly_revenue_json,
    }
    return render(request, 'admin_user/Reports/Rental/commercial_list.html', context)





def commercial_reports(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return render(request, 'home_page/Adminlogin.html')


    print(">>> YES! THE VIEW IS RUNNING! <<<")
    print(f">>> THE DOWNLOAD PARAMETER IS: {request.GET.get('download')} <<<")

    # ── 1. GET ALL SEARCH PARAMS ──
    search_query    = request.GET.get('search', '').strip()
    prop_type_query = request.GET.get('property_type', '').strip()
    city_query      = request.GET.get('city', '').strip()
    zone_query      = request.GET.get('zone_type', '').strip()
    possession_query= request.GET.get('possession', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()
    budget_query    = request.GET.get('budget', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()

    # ── Base queryset ──
    try:
        properties = CommercialRentalProperty.objects.filter(is_deleted=False).order_by('-id')
    except Exception:
        properties = CommercialRentalProperty.objects.all().order_by('-id')

    # ── 2. APPLY FILTERS ──
    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(area_locality__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__icontains=prop_type_query)
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if zone_query and zone_query != 'All Zones':
        properties = properties.filter(zone_type__icontains=zone_query)
    if possession_query and possession_query != 'All Status':
        properties = properties.filter(possession_status__icontains=possession_query)
    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(uploaded_by_role__icontains=listed_by_query)

    if from_date:
        try:
            properties = properties.filter(created_at__date__gte=from_date)
        except Exception:
            pass
    if to_date:
        try:
            properties = properties.filter(created_at__date__lte=to_date)
        except Exception:
            pass

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_25k':
            properties = properties.filter(expected_rent__lt=25000)
        elif budget_query == '25k_1L':
            properties = properties.filter(expected_rent__gte=25000, expected_rent__lte=100000)
        elif budget_query == '1L_5L':
            properties = properties.filter(expected_rent__gte=100000, expected_rent__lte=500000)
        elif budget_query == 'above_5L':
            properties = properties.filter(expected_rent__gt=500000)

    # ── Exhaustive Field Definition for Exports ──
    EXPORT_COLS = [
        ("System Data", "id", False, "Database ID"),
        ("System Data", "commercial_rental_id", False, "System Generated ID"),
        ("System Data", "property_title", False, "Auto Generated Title"),
        
        ("Basic Info", "property_type", True, "office-space / shop / warehouse / industrial / land"),
        ("Basic Info", "property_condition", True, "bare-shell / warm-shell / fitted / furnished"),
        ("Basic Info", "city", True, "City name"),
        ("Basic Info", "area_locality", True, "Area/locality"),
        ("Basic Info", "property_address", True, "Complete address"),
        ("Basic Info", "building_name", True, "Building/project name"),
        ("Basic Info", "possession_status", True, "ready-to-move / under-construction"),
        ("Basic Info", "available_from", False, "YYYY-MM-DD"),
        ("Basic Info", "age_of_property", True, "0-1 / 1-3 / 3-5 / 5-10 / 10+"),
        ("Basic Info", "zone_type", False, "industrial / commercial / residential / special-economic"),
        ("Basic Info", "location_hub", False, "it-park / business-district / mall / standalone"),
        ("Basic Info", "ownership_type", True, "freehold / leasehold / co-operative"),
        ("Basic Info", "construction_status", False, "new / resale"),
        
        ("Area & Pricing", "builtup_area", True, "Number in sq.ft"),
        ("Area & Pricing", "carpet_area", False, "Number in sq.ft"),
        ("Area & Pricing", "expected_rent", True, "Monthly rent in ₹"),
        ("Area & Pricing", "security_deposit", False, "Deposit in ₹"),
        ("Area & Pricing", "maintenance_charges", False, "Monthly maintenance in ₹"),
        ("Area & Pricing", "negotiable", False, "Yes / No"),
        ("Area & Pricing", "brokerage", False, "Yes / No"),
        ("Area & Pricing", "brokerage_percentage", False, "1% / 1.5% / 2% / Negotiable / Manual"),
        ("Area & Pricing", "manual_brokerage", False, "e.g. 2.5%"),
        ("Area & Pricing", "dg_ups_included", False, "true / false"),
        ("Area & Pricing", "electricity_included", False, "true / false"),
        ("Area & Pricing", "water_included", False, "true / false"),
        ("Area & Pricing", "lockin_period", False, "Lock-in months"),
        ("Area & Pricing", "rent_increase", False, "% per year"),
        
        ("Building", "total_floors", False, "Total floors"),
        ("Building", "your_floor", False, "Floor of property"),
        ("Building", "staircases", False, "Number of staircases"),
        ("Building", "passenger_lifts", False, "Number (0 if none)"),
        ("Building", "service_lifts", False, "Number (0 if none)"),
        ("Building", "private_parking", False, "Private parking spots"),
        ("Building", "min_seats", False, "Min seating"),
        ("Building", "max_seats", False, "Max seating"),
        ("Building", "cabins", False, "Number of cabins"),
        ("Building", "meeting_rooms", False, "Number of meeting rooms"),
        ("Building", "private_washroom", False, "Number (0 if none)"),
        ("Building", "public_washroom", False, "Number (0 if none)"),
        ("Building", "flooring_type", False, "marble / vitrified / granite / wooden / ceramic"),
        
        ("Amenities", "amenities", True, "Comma-separated"),
        ("Amenities", "nearby_facilities", True, "Comma-separated"),
        ("Amenities", "property_summary", False, "Short Summary"),
        ("Amenities", "property_description", False, "Auto Generated Detailed description"),
        ("Amenities", "property_description", False, "User Added Detailed description"),
        
        ("Media & Contact", "video", False, "Video file path"),
        ("Media & Contact", "owner_name", True, "Full name"),
        ("Media & Contact", "contact_number", True, "+91 XXXXXXXXXX"),
        ("Media & Contact", "email", True, "email@example.com"),
        ("Media & Contact", "alternate_contact", False, "+91 XXXXXXXXXX"),
        
        ("Uploader Tracking", "uploaded_by_name", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_email", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_contact", False, "Auto-filled"),
        ("Uploader Tracking", "uploaded_by_role", False, "Auto-filled"),
        
        ("System Tracking", "created_at", False, "Datetime created"),
        ("System Tracking", "is_deleted", False, "True/False"),
        ("System Tracking", "deleted_at", False, "Datetime deleted"),
        ("System Tracking", "upload_file_name", False, "Bulk upload source"),
        ("System Tracking", "upload_file_hash", False, "File hash"),
        ("System Tracking", "deleted_by", False, "Deleted by user"),
    ]

    # ════════════════════════════════════════════════
    # 🛑 ROBUST EXPORT LOGIC 🛑
    # ════════════════════════════════════════════════

    # ── Excel Download ──
    if request.GET.get('download') == 'excel':
        try:
            print("--- STARTING EXCEL GENERATION ---")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Commercial Rental"

            HDR_BG  = "667EEA"
            REQ_BG  = "FEF3C7"
            OPT_BG  = "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Row 1: Section Banners
            sec_spans = OrderedDict()
            for i, (sec, *_) in enumerate(EXPORT_COLS):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill      = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border    = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            # Rows 2 & 3: Labels and Hints
            for ci, (sec, field, req, hint) in enumerate(EXPORT_COLS, 1):
                lc = ws.cell(row=2, column=ci, value=field + (" *" if req else ""))
                lc.font      = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill      = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border    = bdr

                hc = ws.cell(row=3, column=ci, value=hint)
                hc.font      = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill      = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border    = bdr

                ws.column_dimensions[get_column_letter(ci)].width = max(18, len(field) + 4)

            print(f"--- FETCHING DATA FOR EXCEL. ROWS FOUND: {properties.count()} ---")
            # Rows 4+: Database Data Injection
            for row_idx, prop in enumerate(properties, start=4):
                for col_idx, (_, field_name, _, _) in enumerate(EXPORT_COLS, start=1):
                    val = getattr(prop, field_name, "")
                    
                    if val is True: val = "Yes"
                    elif val is False: val = "No"
                    elif hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d %H:%M')
                    elif isinstance(val, list): val = ", ".join(map(str, val))
                    elif hasattr(val, 'url'): val = val.url if val else ""
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = bdr

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            print("--- EXCEL FILE BUILT, SENDING RESPONSE ---")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            response = HttpResponse(
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="Commercial_Rental_Listing_Export.xlsx"'
            return response

        except Exception as e:
            # THIS IS CRITICAL: If it fails, it will output the exact error to your browser screen!
            error_msg = f"ERROR GENERATING EXCEL:\n\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            return HttpResponse(f"<pre>{error_msg}</pre>", status=500)

    # ── CSV Download ──
    if request.GET.get('download') == 'csv':
        try:
            print("--- STARTING CSV GENERATION ---")
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="commercial_rental_properties.csv"'
            writer = csv.writer(response)
            
            writer.writerow([field for _, field, _, _ in EXPORT_COLS])
            
            print(f"--- FETCHING DATA FOR CSV. ROWS FOUND: {properties.count()} ---")
            for prop in properties:
                row_data = []
                for _, field_name, _, _ in EXPORT_COLS:
                    val = getattr(prop, field_name, "")
                    if val is True: val = "Yes"
                    elif val is False: val = "No"
                    elif hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d %H:%M')
                    elif isinstance(val, list): val = ", ".join(map(str, val))
                    elif hasattr(val, 'url'): val = val.url if val else ""
                    row_data.append(val if val is not None else "")
                writer.writerow(row_data)
            
            print("--- CSV FILE BUILT, SENDING RESPONSE ---")
            return response
            
        except Exception as e:
            error_msg = f"ERROR GENERATING CSV:\n\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            return HttpResponse(f"<pre>{error_msg}</pre>", status=500)


    # ════════════════════════════════════════════════
    # END EXPORT LOGIC
    # ════════════════════════════════════════════════

    # ── Pagination ──
    paginator    = Paginator(properties, 10)
    page_number  = request.GET.get('page', 1)
    page_obj     = paginator.get_page(page_number)
    filtered_count = properties.count()

    # ════════════════════════════════════════════════
    # ALL-PROPS STATS (always on unfiltered dataset)
    # ════════════════════════════════════════════════
    try:
        all_props = CommercialRentalProperty.objects.filter(is_deleted=False)
    except Exception:
        all_props = CommercialRentalProperty.objects.all()

    total_count = all_props.count()

    # ── Dropdown unique values ──
    unique_property_types = (all_props
        .exclude(property_type__isnull=True).exclude(property_type='')
        .values_list('property_type', flat=True).distinct().order_by('property_type'))
    unique_cities = (all_props
        .exclude(city__isnull=True).exclude(city='')
        .values_list('city', flat=True).distinct().order_by('city'))
    unique_zones = (all_props
        .exclude(zone_type__isnull=True).exclude(zone_type='')
        .values_list('zone_type', flat=True).distinct().order_by('zone_type'))
    unique_possession = (all_props
        .exclude(possession_status__isnull=True).exclude(possession_status='')
        .values_list('possession_status', flat=True).distinct())
    unique_roles = (all_props
        .exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='')
        .values_list('uploaded_by_role', flat=True).distinct())

    # ── Occupancy KPIs ──
    active_count   = all_props.exclude(possession_status__isnull=True).exclude(possession_status='').count()
    occupied_count = all_props.filter(possession_status__iexact='Occupied').count()
    vacant_count   = all_props.filter(possession_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100)) if total_count > 0 else 0
    vacancy_rate   = round((vacant_count   / total_count * 100)) if total_count > 0 else 0

    # ── Revenue KPIs ──
    rent_stats = all_props.exclude(expected_rent__isnull=True).aggregate(
        avg_rent=Avg('expected_rent'),
        max_rent=Max('expected_rent'),
        min_rent=Min('expected_rent'),
    )
    avg_rent = rent_stats['avg_rent'] or 0
    max_rent = rent_stats['max_rent'] or 0
    min_rent = rent_stats['min_rent'] or 0
    total_revenue          = all_props.aggregate(total=Sum('expected_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit'))['total'] or 0
    avg_deposit = (all_props.exclude(security_deposit__isnull=True)
                   .aggregate(avg=Avg('security_deposit'))['avg'] or 0)

    try:
        avg_area = (all_props.exclude(builtup_area__isnull=True)
                    .aggregate(avg=Avg('builtup_area'))['avg'] or 0)
    except Exception:
        avg_area = 0

    # ── Business KPIs ──
    premium_properties_count    = all_props.filter(expected_rent__gte=100000).count()
    affordable_properties_count = all_props.filter(expected_rent__lt=25000).count()
    short_lease_count           = all_props.filter(lockin_period__icontains='6').count()
    long_lease_count            = all_props.filter(lockin_period__icontains='12').count()
    with_owner_count            = all_props.exclude(owner_name__isnull=True).exclude(owner_name='').count()
    city_count                  = (all_props.exclude(city__isnull=True).exclude(city='')
                                   .values('city').distinct().count())
    try:
        with_images_count = all_props.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    # ── Percentages ──
    verified_pct = round((with_owner_count     / total_count * 100)) if total_count > 0 else 0
    image_pct    = round((with_images_count    / total_count * 100)) if total_count > 0 else 0
    premium_pct  = round((premium_properties_count / total_count * 100)) if total_count > 0 else 0

    # ── Uploaded file names (for bulk delete) ──
    try:
        uploaded_files = (all_props.exclude(upload_file_name__isnull=True)
                          .exclude(upload_file_name='')
                          .values_list('upload_file_name', flat=True).distinct())
    except Exception:
        uploaded_files = []

    # ════════════════════════════════════════════
    # CHART DATA  (JSON → template → JS)
    # ════════════════════════════════════════════

    prop_type_dist = list(
        all_props.exclude(property_type__isnull=True).exclude(property_type='')
        .values('property_type').annotate(cnt=Count('id')).order_by('-cnt')[:8]
    )
    prop_type_labels_json = json.dumps([x['property_type'] for x in prop_type_dist])
    prop_type_counts_json = json.dumps([x['cnt']           for x in prop_type_dist])

    rent_range_data = {
        'Under ₹25k':  all_props.filter(expected_rent__lt=25000).count(),
        '₹25k–1L':     all_props.filter(expected_rent__gte=25000,  expected_rent__lt=100000).count(),
        '₹1L–5L':      all_props.filter(expected_rent__gte=100000, expected_rent__lt=500000).count(),
        'Above ₹5L':   all_props.filter(expected_rent__gte=500000).count(),
    }
    rent_range_labels_json = json.dumps(list(rent_range_data.keys()))
    rent_range_counts_json = json.dumps(list(rent_range_data.values()))

    occupancy_json = json.dumps([occupied_count, vacant_count, max(0, total_count - occupied_count - vacant_count)])

    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_qs = (
        all_props.filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(revenue=Sum('expected_rent'), cnt=Count('id'))
        .order_by('month')
    )
    monthly_labels_json  = json.dumps([x['month'].strftime('%b %Y') for x in monthly_qs])
    monthly_revenue_json = json.dumps([float(x['revenue'] or 0) for x in monthly_qs])

    total_tenants    = occupied_count
    collection_rate  = 0
    pending_payments = 0
    maintenance_req  = 0

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,

        'search_query':     search_query,
        'prop_type_query':  prop_type_query,
        'city_query':       city_query,
        'zone_query':       zone_query,
        'possession_query': possession_query,
        'listed_by_query':  listed_by_query,
        'budget_query':     budget_query,
        'from_date':        from_date,
        'to_date':          to_date,

        'unique_property_types': unique_property_types,
        'unique_cities':         unique_cities,
        'unique_zones':          unique_zones,
        'unique_possession':     unique_possession,
        'unique_roles':          unique_roles,
        'uploaded_files':        uploaded_files,

        'filtered_count':            filtered_count,
        'total_count':               total_count,
        'active_count':              active_count,
        'occupied_count':            occupied_count,
        'vacant_count':              vacant_count,
        'occupancy_rate':            occupancy_rate,
        'vacancy_rate':              vacancy_rate,
        'avg_rent':                  avg_rent,
        'max_rent':                  max_rent,
        'min_rent':                  min_rent,
        'total_revenue':             total_revenue,
        'total_security_deposit':    total_security_deposit,
        'avg_deposit':               avg_deposit,
        'avg_area':                  avg_area,
        'premium_properties_count':  premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'ready_to_move_count':       vacant_count,
        'short_lease_count':         short_lease_count,
        'long_lease_count':          long_lease_count,
        'with_owner_count':          with_owner_count,
        'with_images_count':         with_images_count,
        'city_count':                city_count,
        'verified_pct':              verified_pct,
        'image_pct':                 image_pct,
        'premium_pct':               premium_pct,

        'total_tenants':    total_tenants,
        'collection_rate':  collection_rate,
        'pending_payments': pending_payments,
        'maintenance_req':  maintenance_req,

        'prop_type_labels_json':  prop_type_labels_json,
        'prop_type_counts_json':  prop_type_counts_json,
        'rent_range_labels_json': rent_range_labels_json,
        'rent_range_counts_json': rent_range_counts_json,
        'occupancy_json':         occupancy_json,
        'monthly_labels_json':    monthly_labels_json,
        'monthly_revenue_json':   monthly_revenue_json,
    }
    return render(request, 'admin_user/Reports/Rental/commercial_reports.html', context)
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────

def _decimal(v):
    if v is None or str(v).strip() == '': return None
    try: return float(str(v))
    except: return None


def _email(v):
    return _str(v)




COMMERCIAL_RENTAL_COLUMN_MAP = [
    ("property_type",         "property_type",         _str),
    ("property_condition",    "property_condition",    _str),
    ("city",                  "city",                  _str),
    ("area_locality",         "area_locality",         _str),
    ("property_address",      "property_address",      _str),
    ("building_name",         "building_name",         _str),
    ("possession_status",     "possession_status",     _str),
    ("available_from",        "available_from",        _date),
    ("age_of_property",       "age_of_property",       _str),
    ("zone_type",             "zone_type",             _str),
    ("location_hub",          "location_hub",          _str),
    ("ownership_type",        "ownership_type",        _str),
    ("construction_status",   "construction_status",   _str),
    ("builtup_area",          "builtup_area",          _int),
    ("carpet_area",           "carpet_area",           _int),
    ("expected_rent",         "expected_rent",         _int),
    ("security_deposit",      "security_deposit",      _int),
    ("maintenance_charges",   "maintenance_charges",   _int),
    ("negotiable",            "negotiable",            _bool),
    ("brokerage",             "brokerage",             _str),
    ("brokerage_percentage",  "brokerage_percentage",  _str),
    ("manual_brokerage",      "manual_brokerage",      _str),
    ("dg_ups_included",       "dg_ups_included",       _bool),
    ("electricity_included",  "electricity_included",  _bool),
    ("water_included",        "water_included",        _bool),
    ("lockin_period",         "lockin_period",         _int),
    ("rent_increase",         "rent_increase",         _decimal),
    ("total_floors",          "total_floors",          _int),
    ("your_floor",            "your_floor",            _int),
    ("staircases",            "staircases",            _int),
    ("passenger_lifts",       "passenger_lifts",       _int),
    ("service_lifts",         "service_lifts",         _int),
    ("private_parking",       "private_parking",       _int),
    ("min_seats",             "min_seats",             _int),
    ("max_seats",             "max_seats",             _int),
    ("cabins",                "cabins",                _int),
    ("meeting_rooms",         "meeting_rooms",         _int),
    ("private_washroom",      "private_washroom",      _int),
    ("public_washroom",       "public_washroom",       _int),
    ("flooring_type",         "flooring_type",         _str),
    ("amenities",             "amenities",             _str),
    ("nearby_facilities",     "nearby_facilities",     _str),
    ("property_summary",      "property_summary",      _str),
    ("owner_name",            "owner_name",            _str),
    ("contact_number",        "contact_number",        _str),
    ("email",                 "email",                 _email),
    ("alternate_contact",     "alternate_contact",     _str),
    ("uploaded_by_name",      "uploaded_by_name",      _str),
    ("uploaded_by_email",     "uploaded_by_email",     _str),
    ("uploaded_by_contact",   "uploaded_by_contact",   _str),
    ("uploaded_by_role",      "uploaded_by_role",      _str),
]

COMMERCIAL_RENTAL_REQUIRED_DEFAULTS = {
    'property_type': 'office-space', 'city': '', 'area_locality': '',
    'property_address': '', 'building_name': '', 'possession_status': 'ready-to-move',
    'age_of_property': '0-1', 'property_condition': 'bare-shell',
    'ownership_type': 'freehold', 'builtup_area': 0, 'expected_rent': 0,
    'owner_name': '', 'contact_number': '', 'email': '',
}


@require_POST
def import_commercial_excel(request):
    excel_file = request.FILES.get("commercial_file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx files accepted."}, status=400)

    try:
        wb, ws, headers = _load_new_excel(excel_file)
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    created, errors = 0, []
    file_name = excel_file.name

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        obj_data = {**COMMERCIAL_RENTAL_REQUIRED_DEFAULTS, "upload_file_name": file_name}
        row_error = None

        for excel_col, model_field, converter in COMMERCIAL_RENTAL_COLUMN_MAP:
            raw = _get_cell(row, headers.get(excel_col))
            try:
                val = converter(raw)
                if val is not None:
                    obj_data[model_field] = val
            except Exception as e:
                row_error = f"Row {row_idx} '{excel_col}': {e}"
                break

        if row_error:
            errors.append(row_error)
            continue

        # amenities / nearby_facilities are JSONField — convert comma-sep string → list
        for fld in ('amenities', 'nearby_facilities'):
            raw_str = obj_data.get(fld)
            if isinstance(raw_str, str) and raw_str:
                obj_data[fld] = [x.strip() for x in raw_str.split(',') if x.strip()]
            else:
                obj_data[fld] = []

        try:
            CommercialRentalProperty.objects.create(**obj_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} DB: {e}")

    wb.close()
    return JsonResponse({
        "status": "success",
        "message": f"{created} record(s) imported. {len(errors)} error(s).",
        "created": created, "error_count": len(errors), "errors": errors[:20],
    })


def download_commercial_template(request):
    """New-style Commercial Rental template: Row1=banners, Row2=labels, Row3=hints, Row4=sample."""
   

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commercial Rental"

    HDR_BG  = "667EEA"   # section banner – indigo
    REQ_BG  = "FEF3C7"   # required field – amber
    OPT_BG  = "F0FDF4"   # optional field – green-tint
    SAMP_BG = "ECFDF5"   # sample row – mint

    thin = Side(style="thin", color="CBD5E1")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # (section, excel_col, required, hint, sample)
    COLS = [
        # ── Basic Info ────────────────────────────────────────────────────────
        ("Basic Info", "property_type",       True,  "office-space / shop / warehouse / industrial / land",          "office-space"),
        ("Basic Info", "property_condition",  True,  "bare-shell / warm-shell / fitted / furnished",                 "bare-shell"),
        ("Basic Info", "city",                True,  "City name",                                                     "Mumbai"),
        ("Basic Info", "area_locality",       True,  "Area/locality",                                                 "BKC"),
        ("Basic Info", "property_address",    True,  "Complete address",                                              "Tower A, BKC, Mumbai"),
        ("Basic Info", "building_name",       True,  "Building/project name",                                         "Platina Tower"),
        ("Basic Info", "possession_status",   True,  "ready-to-move / under-construction",                           "ready-to-move"),
        ("Basic Info", "available_from",      False, "YYYY-MM-DD",                                                    "2025-08-01"),
        ("Basic Info", "age_of_property",     True,  "0-1 / 1-3 / 3-5 / 5-10 / 10+",                               "1-3"),
        ("Basic Info", "zone_type",           False, "industrial / commercial / residential / special-economic",      "commercial"),
        ("Basic Info", "location_hub",        False, "it-park / business-district / mall / standalone",              "business-district"),
        ("Basic Info", "ownership_type",      True,  "freehold / leasehold / co-operative",                          "freehold"),
        ("Basic Info", "construction_status", False, "new / resale",                                                  "resale"),
        # ── Area & Pricing ────────────────────────────────────────────────────
        ("Area & Pricing", "builtup_area",         True,  "Number in sq.ft",                                         "2000"),
        ("Area & Pricing", "carpet_area",          False, "Number in sq.ft",                                         "1700"),
        ("Area & Pricing", "expected_rent",        True,  "Monthly rent in ₹",                                      "85000"),
        ("Area & Pricing", "security_deposit",     False, "Deposit in ₹",                                           "500000"),
        ("Area & Pricing", "maintenance_charges",  False, "Monthly maintenance in ₹",                               "5000"),
        ("Area & Pricing", "negotiable",           False, "Yes / No",                                                "Yes"),
        ("Area & Pricing", "brokerage",            False, "Yes / No",                                                "No"),
        ("Area & Pricing", "brokerage_percentage", False, "1% / 1.5% / 2% / Negotiable / Manual",                   ""),
        ("Area & Pricing", "manual_brokerage",     False, "e.g. 2.5% (if Manual)",                                  ""),
        ("Area & Pricing", "dg_ups_included",      False, "true / false",                                           "false"),
        ("Area & Pricing", "electricity_included", False, "true / false",                                           "false"),
        ("Area & Pricing", "water_included",       False, "true / false",                                           "false"),
        ("Area & Pricing", "lockin_period",        False, "Lock-in months",                                         "6"),
        ("Area & Pricing", "rent_increase",        False, "% per year e.g. 5",                                      "5"),
        # ── Building ──────────────────────────────────────────────────────────
        ("Building", "total_floors",     False, "Total floors in building",        "10"),
        ("Building", "your_floor",       False, "Floor of this property",          "4"),
        ("Building", "staircases",       False, "Number of staircases",            "2"),
        ("Building", "passenger_lifts",  False, "Number (use 0 if none)",          "2"),
        ("Building", "service_lifts",    False, "Number (use 0 if none)",          "1"),
        ("Building", "private_parking",  False, "Number of private parking spots", "2"),
        ("Building", "min_seats",        False, "Minimum seating capacity",        "20"),
        ("Building", "max_seats",        False, "Maximum seating capacity",        "50"),
        ("Building", "cabins",           False, "Number of cabins",                "5"),
        ("Building", "meeting_rooms",    False, "Number of meeting rooms",         "2"),
        ("Building", "private_washroom", False, "Number (use 0)",                  "1"),
        ("Building", "public_washroom",  False, "Number (use 0)",                  "2"),
        ("Building", "flooring_type",    False, "marble / vitrified / granite / wooden / ceramic", "vitrified"),
        # ── Amenities ─────────────────────────────────────────────────────────
        ("Amenities/Property Description", "amenities",          True,  "Comma-sep e.g. Wi-Fi,AC,CCTV,Generator",  "Wi-Fi,AC,CCTV"),
        ("Amenities/Property Description", "nearby_facilities",  True,  "Comma-sep e.g. Metro,Bank,Parking",        "Metro,Bank"),
        ("Amenities/Property Description", "user_description",   False, "Short plain-text description added By user", "My Property Near ATM...."),
        # ── Contact ───────────────────────────────────────────────────────────
        ("Contact", "owner_name",        True,  "Full name",           "Rahul Mehta"),
        ("Contact", "contact_number",    True,  "+91 XXXXXXXXXX",      "9876543210"),
        ("Contact", "email",             True,  "email@example.com",   "rahul@email.com"),
        ("Contact", "alternate_contact", False, "+91 XXXXXXXXXX",      ""),
        ("Contact", "uploaded_by_name",  False, "Auto-filled",         ""),
        ("Contact", "uploaded_by_email", False, "Auto-filled",         ""),
        ("Contact", "uploaded_by_contact",False,"Auto-filled",         ""),
        ("Contact", "uploaded_by_role",  False, "Auto-filled",         ""),
    ]

    # ── Row 1: section banners ────────────────────────────────────────────────
    sec_spans = OrderedDict()
    for i, (sec, *_) in enumerate(COLS):
        sec_spans.setdefault(sec, []).append(i + 1)

    for sec, cols in sec_spans.items():
        c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr
        if len(cols) > 1:
            ws.merge_cells(start_row=1, start_column=cols[0],
                           end_row=1,   end_column=cols[-1])

    # ── Rows 2 / 3 / 4 ───────────────────────────────────────────────────────
    for ci, (sec, field, req, hint, sample) in enumerate(COLS, 1):
        # Row 2 – label
        lc = ws.cell(row=2, column=ci, value=field + (" *" if req else ""))
        lc.font      = Font(bold=True, color="1E293B", name="Arial", size=9)
        lc.fill      = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lc.border    = bdr

        # Row 3 – hint
        hc = ws.cell(row=3, column=ci, value=hint)
        hc.font      = Font(italic=True, color="64748B", name="Arial", size=8)
        hc.fill      = PatternFill("solid", fgColor="FFFFFF")
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border    = bdr

        # Row 4 – sample
        sc = ws.cell(row=4, column=ci, value=sample)
        sc.font      = Font(name="Arial", size=9, color="065F46")
        sc.fill      = PatternFill("solid", fgColor=SAMP_BG)
        sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sc.border    = bdr

        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(field) + 4)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[4].height = 26
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Commercial_Rental_Listing_Template.xlsx"'
    return response




############### Views end for commercial property list ########################


######### Views start for pg co living rental list ########################



   





# ─────────────────────────────────────────────────────────────
# LIST VIEW
# ─────────────────────────────────────────────────────────────





# ─────────────────────────────────────────────────────────────
# DELETE VIEW  (POST only — called via JS fetch)
# ─────────────────────────────────────────────────────────────

########### Views end for pg co living rental list ########################


def residential_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale/residential_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')
   

def commercial_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()

        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale/commercial_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


def plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale/plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')






def industrial_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()

        user_obj = User_Details.objects.all()

        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale/industrial_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')
   

def agricultural_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()

        user_obj = User_Details.objects.all()


        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale/agricultural_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')




############# Views start for resale residential property list ###################













############ Views end for resale residential list #######################



########### Views start for display rm list ##########################

@csrf_exempt
def rm_list(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        rm_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Relationship Manager").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Relationship Manager").exists():
            rm_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Relationship Manager").count()
            rendered = render_to_string("admin_user/render_to_string/R_RM/r_t_s_rm.html",{'rm_obj':rm_obj,'rm_obj_count':rm_obj_count,'Role':'Relationship Manager'})
            return HttpResponse(rendered)
        else:
            return HttpResponse("error")

            
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        rm_obj = User_Details.objects.filter(user_role="Relationship Manager").order_by('-user_register_date')
        rm_obj_count = User_Details.objects.filter(user_role="Relationship Manager").count()

        rendered = render_to_string("admin_user/render_to_string/R_RM/r_t_s_rm.html",{'rm_obj':rm_obj,'rm_obj_count':rm_obj_count,'Role':'Relationship Manager'})

        context = {'admin_obj':admin_obj,'rm_list':rendered}
        return render(request,'admin_user/RM/rm_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for display rm list ###########################


############ Views start for add rm ############################

def Add_RM(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/RM/add_rm.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

################ Views end for add rm ###########################


########### Views start for data upload functtionality via excel ##############

@csrf_exempt
def Rm_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('rm_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Relationship Manager"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name (CORRECT)
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email (CORRECT)
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    user_state = row[9] if len(row) > 9 else None  # Column 9: State
                    user_city = row[10] if len(row) > 10 else None  # Column 10: City
                    user_address = row[11] if len(row) > 11 else None  # Column 11: Address
                    register_date_value = row[12] if len(row) > 12 else None  # Column 12: Register Date
                    register_time_value = row[13] if len(row) > 13 else None  # Column 13: Register Time
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    register_time = datetime.now().time()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass
                    
                    # If time provided in Excel, try to parse it
                    if register_time_value and str(register_time_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_time_value, time):
                                register_time = register_time_value
                            else:
                                time_str = str(register_time_value).strip()
                                if ':' in time_str:
                                    time_str = time_str.replace('a.m.', '').replace('p.m.', '').replace('AM', '').replace('PM', '').strip()
                                    try:
                                        register_time = datetime.strptime(time_str, '%I:%M').time()
                                    except:
                                        try:
                                            register_time = datetime.strptime(time_str, '%H:%M').time()
                                        except:
                                            pass
                        except:
                            pass
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = register_time
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=register_time
                        )
                        
                        # Generate USER_ID: EF-{ID}-{YY}
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Relationship Managers. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

########## Views end for data upload functionality via excel #######################


########### Views start for delete rm details ######################

@csrf_exempt
def Delete_RM(request):
    try:
        try:
            rm_id = request.POST.get('rm_id')
            User_Details.objects.filter(id=rm_id).delete()
            return JsonResponse({'status':'1', 'msg':'RM details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############ Views end for delete rm details #########################


########### Views start for update rm details #########################

def Update_RM(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        rm = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'rm':rm}
        return render(request,'admin_user/RM/update_rm.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

######### Views end for update rm details #############################


########## Views start for ajax for add/update rm functionality #################4

@csrf_exempt
def User_Ajax(request):
    data = request.POST.dict()

    # --- HANDLE NEW FIELDS ---
    # 1. Handle Checkboxes: .dict() fails on lists, so we use .getlist
    user_operational_scope = data.get('user_operational_scope')
    if user_operational_scope == 'all':
        data['selected_regions'] = "All Over India"
    else:
        # We sent it as a JSON string from AJAX
        regions_raw = request.POST.get('selected_regions')
        try:
            regions_list = json.loads(regions_raw)
            data['selected_regions'] = ", ".join(regions_list)
        except (json.JSONDecodeError, TypeError):
            data['selected_regions'] = ""

    if data.get('id') == "":
        data.pop("id", None)
        data['user_profile'] = request.FILES.get('user_profile')        
        data['user_register_date'] = datetime.today()
        data['user_register_time'] = datetime.now()
        if User_Details.objects.filter(user_phone=data['user_phone']).exists():
            return JsonResponse({"status":"0", "msg" : f"User with this phone number already exists"})
        elif User_Details.objects.filter(user_email=data['user_email']).exists():
            return JsonResponse({"status":"0", "msg" : f"User with this email address already exists"})
        else:
            User_Details.objects.create(**data)
            return JsonResponse({"status":"1", "msg" : f"User Details added successfully"})

    # UPDATE MODE
    else:
        try:
            rm = User_Details.objects.get(id=data['id'])
        except User_Details.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'User Details not found'})
       
        data['user_profile'] = request.FILES.get('user_profile')

        if request.FILES.get('user_profile'):
            data['user_profile'] = request.FILES.get('user_profile')
        else:
            data.pop('user_profile', None)


        # Update withdraw fields (unchanged)
        for key, value in data.items():
            setattr(rm, key, value)

        rm.save()
        return JsonResponse({"status":"1", "msg" : f"User Details updated successfully"})

########### Views end for ajax for add/update rm functionality ###################


############ Views start for ajax for delete bulk users #####################

@csrf_exempt
def Users_Bulk_Delete(request):
    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        
        #  BUG FIX 1: This was incorrectly pulling 'delete_type' before
        role = data.get('role') 
        
        # Base query: Get all users of the selected role
        users = User_Details.objects.filter(user_role=role)

        print(f"--- Bulk Delete Request: Type={delete_type}, Role={role} ---")
        
        if delete_type == 'delete_all':
            count = users.count()
            users.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted ALL {count} users ({role}).'})

            
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            # Assuming your date field is 'user_register_date'. Adjust if necessary.
            target_users = users.filter(user_register_date__range=[from_date, to_date])
            count = target_users.count()
            target_users.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} users ({role}) in date range.'})
            
            
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            #  BUG FIX 2 & 3: Used the base 'users' queryset so it respects the role filter
            target_users = users.filter(user_register_date__gte=thirty_days_ago)
            count = target_users.count()
            target_users.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} users ({role}) from the last 30 days.'})
            
            
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            #  BUG FIX 2 & 3: Used the base 'users' queryset so it respects the role filter
            target_users = users.filter(user_register_date__lt=six_months_ago)
            count = target_users.count()
            target_users.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} older users ({role}).'})
            

        elif delete_type == 'current_page':
            # This handles the front-end 'current_page' logic if you pass page_ids
            page_ids = data.get('page_ids', [])
            target_users = users.filter(id__in=page_ids)
            count = target_users.count()
            target_users.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} users ({role}) from the current page.'})

            
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'})

############# Views end for ajax for delete bulk users #######################


########### Views start for display landlords list ###################

@csrf_exempt
def Landlord_List(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        landlord_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Landlord").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Landlord").exists():
            landlord_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Landlord").count()
            rendered = render_to_string("admin_user/render_to_string/R_Landlord/r_t_s_landlord.html",{'landlord_obj':landlord_obj,'landlord_obj_count':landlord_obj_count,'Role':'Landlord'})

            return HttpResponse(rendered)
        else:
            return HttpResponse("error")


    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        landlord_obj = User_Details.objects.filter(user_role="Landlord").order_by('-user_register_date')
        landlord_obj_count = User_Details.objects.filter(user_role="Landlord").count()

        rendered = render_to_string("admin_user/render_to_string/R_Landlord/r_t_s_landlord.html",{'landlord_obj':landlord_obj,'landlord_obj_count':landlord_obj_count,'Role':'Landlord'})

        context = {'admin_obj':admin_obj,'landlords_list':rendered}

        return render(request,'admin_user/Landlord/landlord_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for display landlords list ######################


############ Views start for add landlords #####################

def Add_Landlord(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/Landlord/add_landlord.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

########### Views end for add landlords ########################


########### Views start for upload landlord data functionality via excel ###############

@csrf_exempt
def Landlord_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('landlord_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Landlord"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name (CORRECT)
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email (CORRECT)
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    user_state = row[9] if len(row) > 9 else None  # Column 9: State
                    user_city = row[10] if len(row) > 10 else None  # Column 10: City
                    user_address = row[11] if len(row) > 11 else None  # Column 11: Address
                    register_date_value = row[12] if len(row) > 12 else None  # Column 12: Register Date
                    register_time_value = row[13] if len(row) > 13 else None  # Column 13: Register Time
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    register_time = datetime.now().time()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass
                    
                    # If time provided in Excel, try to parse it
                    if register_time_value and str(register_time_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_time_value, time):
                                register_time = register_time_value
                            else:
                                time_str = str(register_time_value).strip()
                                if ':' in time_str:
                                    time_str = time_str.replace('a.m.', '').replace('p.m.', '').replace('AM', '').replace('PM', '').strip()
                                    try:
                                        register_time = datetime.strptime(time_str, '%I:%M').time()
                                    except:
                                        try:
                                            register_time = datetime.strptime(time_str, '%H:%M').time()
                                        except:
                                            pass
                        except:
                            pass
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = register_time
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=register_time
                        )
                        
                        # Generate USER_ID: EF-{ID}-{YY}
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Landlords. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

############ Views end for upload landlord data functionality via excel #######


############ Views start for delete landlord details ########################

@csrf_exempt
def Delete_Landlord(request):
    try:
        try:
            landlord_id = request.POST.get('landlord_id')
            User_Details.objects.filter(id=landlord_id).delete()
            return JsonResponse({'status':'1', 'msg':'Landlord details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

########## Views end for delete landlord details ##############################


############### Views start for update landlord details #####################

def Update_Landlord(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        landlord = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'landlord':landlord}
        return render(request,'admin_user/Landlord/update_landlord.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for update landlord details ###########################


######### Views start for display tenants list #####################

@csrf_exempt
def Tenant_List(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        tenant_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Tenant").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Tenant").exists():
            tenant_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Tenant").count()

            rendered = render_to_string("admin_user/render_to_string/R_Tenant/r_t_s_tenant.html",{'tenant_obj':tenant_obj,'tenant_obj_count':tenant_obj_count,'Role':'Tenant'})
            
            return HttpResponse(rendered)
        else:
            return HttpResponse("error")


    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        tenant_obj = User_Details.objects.filter(user_role="Tenant").order_by('-user_register_date')
        tenant_obj_count = User_Details.objects.filter(user_role="Tenant").count()

        rendered = render_to_string("admin_user/render_to_string/R_Tenant/r_t_s_tenant.html",{'tenant_obj':tenant_obj,'tenant_obj_count':tenant_obj_count,'Role':'Tenant'})


        context = {'admin_obj':admin_obj,'tenants_list':rendered}
       
        return render(request,'admin_user/Tenant/tenant_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for display tenants list ########################


############ Views start for add tenants ######################

def Add_Tenant(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/Tenant/add_tenant.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

######## Views end for add tenants ##########################


########## Views start for upload tenant data functionality via excel ##############

@csrf_exempt
def Tenant_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('tenant_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Tenant"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name (CORRECT)
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email (CORRECT)
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    user_state = row[9] if len(row) > 9 else None  # Column 9: State
                    user_city = row[10] if len(row) > 10 else None  # Column 10: City
                    user_address = row[11] if len(row) > 11 else None  # Column 11: Address
                    register_date_value = row[12] if len(row) > 12 else None  # Column 12: Register Date
                    register_time_value = row[13] if len(row) > 13 else None  # Column 13: Register Time
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    register_time = datetime.now().time()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass
                    
                    # If time provided in Excel, try to parse it
                    if register_time_value and str(register_time_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_time_value, time):
                                register_time = register_time_value
                            else:
                                time_str = str(register_time_value).strip()
                                if ':' in time_str:
                                    time_str = time_str.replace('a.m.', '').replace('p.m.', '').replace('AM', '').replace('PM', '').strip()
                                    try:
                                        register_time = datetime.strptime(time_str, '%I:%M').time()
                                    except:
                                        try:
                                            register_time = datetime.strptime(time_str, '%H:%M').time()
                                        except:
                                            pass
                        except:
                            pass
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = register_time
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=register_time
                        )
                        
                        # Generate USER_ID: EF-{ID}-{YY}
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Tenants. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

######### Views end for upload tenant data functionality via excel ####################


########### Views start for delete tenant details #######################

@csrf_exempt
def Delete_Tenant(request):
    try:
        try:
            tenant_id = request.POST.get('tenant_id')
            User_Details.objects.filter(id=tenant_id).delete()
            return JsonResponse({'status':'1', 'msg':'Tenant details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()

########### Views end for delete tenant details ############################


############ Views start for update tenant details ###################

def Update_Tenant(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        tenant = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'tenant':tenant}
        return render(request,'admin_user/Tenant/update_tenant.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')


############### Views start for display buyers list ####################

@csrf_exempt
def Buyer_List(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        buyer_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Buyer").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Buyer").exists():
            buyer_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Buyer").count()

            rendered = render_to_string("admin_user/render_to_string/R_Buyer/r_t_s_buyer.html",{'buyer_obj':buyer_obj,'buyer_obj_count':buyer_obj_count,'Role':'Buyer'})

            return HttpResponse(rendered)
        else:
            return HttpResponse("error")


    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        buyer_obj = User_Details.objects.filter(user_role="Buyer").order_by('-user_register_date')
        buyer_obj_count = User_Details.objects.filter(user_role="Buyer").count()

        rendered = render_to_string("admin_user/render_to_string/R_Buyer/r_t_s_buyer.html",{'buyer_obj':buyer_obj,'buyer_obj_count':buyer_obj_count,'Role':'Buyer'})

        context = {'admin_obj':admin_obj,'buyer_list':rendered}

        return render(request,'admin_user/Buyer/buyer_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for display buyers list #########################


############# Views start for add buyers ########################

def Add_Buyer(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/Buyer/add_buyer.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for add buyers ###############################


############# Views start for buyer data functionality via excel ####################

@csrf_exempt
def Buyer_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('buyer_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Buyer"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name (CORRECT)
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email (CORRECT)
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    user_state = row[9] if len(row) > 9 else None  # Column 9: State
                    user_city = row[10] if len(row) > 10 else None  # Column 10: City
                    user_address = row[11] if len(row) > 11 else None  # Column 11: Address
                    register_date_value = row[12] if len(row) > 12 else None  # Column 12: Register Date
                    register_time_value = row[13] if len(row) > 13 else None  # Column 13: Register Time
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    register_time = datetime.now().time()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass
                    
                    # If time provided in Excel, try to parse it
                    if register_time_value and str(register_time_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_time_value, time):
                                register_time = register_time_value
                            else:
                                time_str = str(register_time_value).strip()
                                if ':' in time_str:
                                    time_str = time_str.replace('a.m.', '').replace('p.m.', '').replace('AM', '').replace('PM', '').strip()
                                    try:
                                        register_time = datetime.strptime(time_str, '%I:%M').time()
                                    except:
                                        try:
                                            register_time = datetime.strptime(time_str, '%H:%M').time()
                                        except:
                                            pass
                        except:
                            pass
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = register_time
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=register_time
                        )
                        
                        # Generate USER_ID: EF-{ID}-{YY}
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Buyers. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})


######### Views end for buyer data functionality via excel ###########################


############ Views start for delete buyer details #######################

@csrf_exempt
def Delete_Buyer(request):
    try:
        try:
            buyer_id = request.POST.get('buyer_id')
            User_Details.objects.filter(id=buyer_id).delete()
            return JsonResponse({'status':'1', 'msg':'Buyer details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()

########## Views end for delete buyer details ###########################


########### Views start for update buyer details ###########################

def Update_Buyer(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        buyer = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'buyer':buyer}
       
        return render(request,'admin_user/Buyer/update_buyer.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for update buyer details ###############################


######### Views start for display agents list ##################

@csrf_exempt
def Agent_List(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        agent_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agent").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agent").exists():
            agent_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agent").count()

            rendered = render_to_string("admin_user/render_to_string/R_Agent/r_t_s_agent.html",{'agent_obj':agent_obj,'agent_obj_count':agent_obj_count,'Role':'Agent'})

            return HttpResponse(rendered)
        else:
            return HttpResponse("error")
        
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        agent_obj = User_Details.objects.filter(user_role="Agent").order_by('-user_register_date')
        agent_obj_count = User_Details.objects.filter(user_role="Agent").count()

        rendered = render_to_string("admin_user/render_to_string/R_Agent/r_t_s_agent.html",{'agent_obj':agent_obj,'agent_obj_count':agent_obj_count,'Role':'Agent'})


        context = {'admin_obj':admin_obj,'agent_list':rendered}
        return render(request,'admin_user/Agent/agent_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for display agents list #################


############ Views start for add agents #################

def Add_Agent(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/Agent/add_agent.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for add agents #######################


########## Views start for upload agent data functionality via excel ###############

@csrf_exempt
def Agent_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('agent_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Agent"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # ================================================================
                    # COLUMN MAPPING (Based on your Agents Report structure)
                    # ================================================================
                    # Column 0: Actions (empty)
                    # Column 1: Sr. No.
                    # Column 2: Profile
                    # Column 3: User Id
                    # Column 4: Role (contains "Agent")
                    # Column 5: Name (contains "Anita Chacko")
                    # Column 6: Email Address (contains "anita.chacko@example.com")
                    # Column 7: Phone Number
                    # Column 8: Password
                    # Column 9: Agency Name
                    # Column 10: License Number
                    # Column 11: State
                    # Column 12: City
                    # Column 13: Address
                    # Column 14: Register Date
                    # ================================================================
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    agency_name = row[9] if len(row) > 9 else None  # Column 9: Agency Name
                    license_number = row[10] if len(row) > 10 else None  # Column 10: License Number
                    user_state = row[11] if len(row) > 11 else None  # Column 11: State
                    user_city = row[12] if len(row) > 12 else None  # Column 12: City
                    user_address = row[13] if len(row) > 13 else None  # Column 13: Address
                    register_date_value = row[14] if len(row) > 14 else None  # Column 14: Register Date
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    agency_name = str(agency_name).strip() if agency_name and str(agency_name) != '---' else None
                    license_number = str(license_number).strip() if license_number and str(license_number) != '---' else None
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass  # Keep today's date if parsing fails
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_agency_name = agency_name
                        existing_user.user_license_number = license_number
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = datetime.now().time()
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_agency_name=agency_name,
                            user_license_number=license_number,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=datetime.now().time()
                        )
                        
                        # --- GENERATE USER_ID WITH FORMAT: EF-{ID}-{YY} ---
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]  # Get last 2 digits of year (e.g., 26 for 2026)
                        
                        # Format: EF-{user.id}-{YY}
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        
                        # Update the user with the generated user_id
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Agents. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]  # Limit to first 10 errors for response
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})

########### Views end for upload agent data functionaity via excel #####################


############# Views start for delete agent ##############################

@csrf_exempt
def Delete_Agent(request):
    try:
        try:
            agent_id = request.POST.get('agent_id')
            User_Details.objects.filter(id=agent_id).delete()
            return JsonResponse({'status':'1', 'msg':'Agent details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()

########## Views ennd for delete agent ###################################


########### Views start for update agent details ################

def Update_Agent(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        agent = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'agent':agent}

        return render(request,'admin_user/Agent/update_agent.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

########## Views end for update agent details #####################


########## Views start for display agency list #########################

@csrf_exempt
def Agency_List(request):

    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        agency_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agency/Builder").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agency/Builder").exists():
            
            agency_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Agency/Builder").count()

            rendered = render_to_string("admin_user/render_to_string/R_Agency/r_t_s_agency.html",{'agency_obj':agency_obj,'agency_obj_count':agency_obj_count,'Role':'Agency'})

            return HttpResponse(rendered)
        else:
            return HttpResponse("error")


    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        agency_obj = User_Details.objects.filter(user_role="Agency/Builder").order_by('-user_register_date')
        agency_obj_count = User_Details.objects.filter(user_role="Agency/Builder").count()

        rendered = render_to_string("admin_user/render_to_string/R_Agency/r_t_s_agency.html",{'agency_obj':agency_obj,'agency_obj_count':agency_obj_count,'Role':'Agency'})

        context = {'admin_obj':admin_obj,'agency_list':rendered}
        return render(request,'admin_user/Agency/agency_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for display agency list ############################


############### Views start for add agency ########################

def Add_Agency(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        context = {'admin_obj':admin_obj}
        return render(request,'admin_user/Agency/add_agency.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############ Views end for add agency ##########################


########## Views start for upload agency data functionality via excel #################

@csrf_exempt
def Agency_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('agency_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Please select an Excel file"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            error_details = []
            
            FIXED_ROLE = "Agency/Builder"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    # Skip empty rows
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # ================================================================
                    # COLUMN MAPPING (Based on your Agents Report structure)
                    # ================================================================
                    # Column 0: Actions (empty)
                    # Column 1: Sr. No.
                    # Column 2: Profile
                    # Column 3: User Id
                    # Column 4: Role (contains "Agent")
                    # Column 5: Name (contains "Anita Chacko")
                    # Column 6: Email Address (contains "anita.chacko@example.com")
                    # Column 7: Phone Number
                    # Column 8: Password
                    # Column 9: Agency Name
                    # Column 10: License Number
                    # Column 11: State
                    # Column 12: City
                    # Column 13: Address
                    # Column 14: Register Date
                    # ================================================================
                    
                    # Extract values with correct column indices
                    user_role = row[4] if len(row) > 4 else None  # Column 4: Role
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    agency_name = row[9] if len(row) > 9 else None  # Column 9: Agency Name
                    license_number = row[10] if len(row) > 10 else None  # Column 10: License Number
                    user_state = row[11] if len(row) > 11 else None  # Column 11: State
                    user_city = row[12] if len(row) > 12 else None  # Column 12: City
                    user_address = row[13] if len(row) > 13 else None  # Column 13: Address
                    register_date_value = row[14] if len(row) > 14 else None  # Column 14: Register Date
                    
                    # Clean and validate user_name (from Name column - index 5)
                    if user_name and str(user_name).strip() not in ['', '---', '-']:
                        user_name = str(user_name).strip()
                    else:
                        user_name = None
                    
                    # Clean and validate user_email (from Email column - index 6)
                    if user_email and str(user_email).strip() not in ['', '---', '-']:
                        user_email = str(user_email).strip()
                    else:
                        user_email = None
                    
                    if not user_name or not user_phone:
                        error_count += 1
                        error_details.append(f"Row {row_idx}: Missing Name or Phone (Name: {user_name}, Phone: {user_phone})")
                        continue
                    
                    # Clean phone
                    user_phone = str(int(user_phone)) if isinstance(user_phone, (int, float)) else str(user_phone).replace('-', '').strip()
                    
                    # Clean other fields
                    user_name = str(user_name).strip()
                    user_password = str(int(user_password)) if isinstance(user_password, (int, float)) else str(user_password).split('.')[0].strip() if user_password else 'default123'
                    agency_name = str(agency_name).strip() if agency_name and str(agency_name) != '---' else None
                    license_number = str(license_number).strip() if license_number and str(license_number) != '---' else None
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass  # Keep today's date if parsing fails
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_agency_name = agency_name
                        existing_user.user_license_number = license_number
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = datetime.now().time()
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_agency_name=agency_name,
                            user_license_number=license_number,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=datetime.now().time()
                        )
                        
                        # --- GENERATE USER_ID WITH FORMAT: EF-{ID}-{YY} ---
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]  # Get last 2 digits of year (e.g., 26 for 2026)
                        
                        # Format: EF-{user.id}-{YY}
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        
                        # Update the user with the generated user_id
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error at row {row_idx}: {str(e)}")
            
            return JsonResponse({
                "status": "1",
                "msg": f"Successfully imported {success_count} Agencies/Builders. Failed: {error_count}",
                "success_count": success_count,
                "error_count": error_count,
                "error_details": error_details[:10]  # Limit to first 10 errors for response
            })
            
        except Exception as e:
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid request method"})


############# Views end for upload agency data functionality via excel ##################


############## Views start for delete agency ###########################

@csrf_exempt
def Delete_Agency(request):
    try:
        try:
            agency_id = request.POST.get('agency_id')
            User_Details.objects.filter(id=agency_id).delete()
            return JsonResponse({"status":"1", "msg" : "Agency Details Deleted Successfully..."})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()

############ Views end for delete agency ################################


########### Views start for update agency ###########################

def Update_Agency(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        agency = User_Details.objects.get(id=id)
       
        context = {'admin_obj':admin_obj,'agency':agency}

        return render(request,'admin_user/Agency/update_agency.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

######### Views end for update agency ###############################


########## Views start for display vendors list ##################

@csrf_exempt
def Vendor_List(request):


    if request.method=="POST":
        start_date= request.POST.get('start_date')
        end_date= request.POST.get('end_date')
        vendor_obj = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Vendor").order_by("-id")
        if User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Vendor").exists():
            vendor_obj_count = User_Details.objects.filter(user_register_date__gte=start_date,user_register_date__lte=end_date,user_role="Vendor").count()

            rendered = render_to_string("admin_user/render_to_string/R_Vendor/r_t_s_vendor.html",{'vendor_obj':vendor_obj,'vendor_obj_count':vendor_obj_count,'Role':'Vendor'})

            return HttpResponse(rendered)
        else:
            return HttpResponse("error")


    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        vendor_obj = User_Details.objects.filter(user_role="Vendor").order_by('-user_register_date')
        vendor_obj_count = User_Details.objects.filter(user_role="Vendor").count()

        rendered = render_to_string("admin_user/render_to_string/R_Vendor/r_t_s_vendor.html",{'vendor_obj':vendor_obj,'vendor_obj_count':vendor_obj_count,'Role':'Vendor'})

        context = {'admin_obj':admin_obj,'vendors_list':rendered}

        return render(request,'admin_user/Vendor/vendor_list.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

######### Views end for display vendors list ######################


########### Views start for add vendor #####################

def Add_Vendor(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        services_obj = Service_Type_Details.objects.all()

        context = {'admin_obj':admin_obj,'services_obj':services_obj}
        return render(request,'admin_user/Vendor/add_vendor.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############### Views end for add vendor #######################


############# Views start for upload vendor data functionality via excel ###############

@csrf_exempt
def Vendor_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('vendor_file')
            
            if not excel_file:
                return JsonResponse({"status": "0", "msg": "Excel file not found"})
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({"status": "0", "msg": "Please upload .xlsx or .xls file only"})
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            success_count = 0
            error_count = 0
            skipped_count = 0
            error_details = []
            
            FIXED_ROLE = "Vendor"
            
            # Data starts from row 3
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    if not any(row) or len(row) < 18:
                        skipped_count += 1
                        continue
                   
                    
                    user_service_type = row[4] if len(row) > 4 else None  # Column 4: Service Type
                    user_name = row[5] if len(row) > 5 else None  # Column 5: Name
                    user_email = row[6] if len(row) > 6 else None  # Column 6: Email
                    user_phone = row[7] if len(row) > 7 else None  # Column 7: Phone
                    user_password = row[8] if len(row) > 8 else None  # Column 8: Password
                    user_state = row[9] if len(row) > 9 else None  # Column 9: State
                    user_city = row[10] if len(row) > 10 else None  # Column 10: City
                    user_address = row[11] if len(row) > 11 else None  # Column 11: Address
                    user_company_name = row[12] if len(row) > 12 else None  # Column 12: Company Name
                    user_pan_number = row[13] if len(row) > 13 else None  # Column 13: PAN
                    user_gstin_number = row[14] if len(row) > 14 else None  # Column 14: GSTIN
                    operational_scope = row[15] if len(row) > 15 else None  # Column 15: Operational Scope
                    selected_regions = row[16] if len(row) > 16 else None  # Column 16: Selected Regions
                    register_date_value = row[17] if len(row) > 17 else None  # Column 17: Register Date
                    register_time_value = row[18] if len(row) > 18 else None  # Column 18: Register Time
                    
                    # Skip if no name or phone
                    if not user_name or not user_phone:
                        skipped_count += 1
                        error_details.append(f"Row {row_idx}: Missing name or phone")
                        continue
                    
                    # Skip if name is "View Profile" (Profile column value)
                    user_name_str = str(user_name).strip()
                    if user_name_str == 'View Profile' or user_name_str == 'None' or user_name_str.isdigit():
                        error_details.append(f"Row {row_idx}: Invalid name '{user_name_str}'")
                        skipped_count += 1
                        continue
                    
                    # Clean phone
                    if isinstance(user_phone, (int, float)):
                        user_phone = str(int(user_phone))
                    else:
                        user_phone = str(user_phone).replace('-', '').replace(' ', '').strip()
                    
                    # Clean password
                    if user_password:
                        if isinstance(user_password, (int, float)):
                            user_password = str(int(user_password))
                        else:
                            user_password = str(user_password).split('.')[0].strip()
                    else:
                        user_password = 'default123'
                    
                    # Clean email
                    user_email = str(user_email).strip() if user_email and str(user_email) != '---' else None
                    
                    # Clean address fields
                    user_name = str(user_name).strip()
                    user_state = str(user_state).strip() if user_state and str(user_state) != '---' else None
                    user_city = str(user_city).strip() if user_city and str(user_city) != '---' else None
                    user_address = str(user_address).strip() if user_address and str(user_address) != '---' else None
                    
                    # Clean vendor fields
                    user_service_type = str(user_service_type).strip() if user_service_type and str(user_service_type) != '---' else None
                    user_company_name = str(user_company_name).strip() if user_company_name and str(user_company_name) != '---' else None
                    user_pan_number = str(user_pan_number).strip().upper() if user_pan_number and str(user_pan_number) != '---' else None
                    user_gstin_number = str(user_gstin_number).strip().upper() if user_gstin_number and str(user_gstin_number) != '---' else None
                    operational_scope = str(operational_scope).strip() if operational_scope and str(operational_scope) != '---' else None
                    
                    # Clean selected_regions
                    if selected_regions and str(selected_regions) != '---':
                        selected_regions = str(selected_regions).strip()
                    else:
                        selected_regions = None
                    
                    # Register date: Today's date as default
                    register_date = datetime.now().date()
                    register_time = datetime.now().time()
                    
                    # If date provided in Excel, try to parse it
                    if register_date_value and str(register_date_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_date_value, (date, datetime)):
                                register_date = register_date_value.date() if isinstance(register_date_value, datetime) else register_date_value
                            else:
                                date_str = str(register_date_value).strip()
                                if ',' in date_str:
                                    register_date = datetime.strptime(date_str, '%B %d, %Y').date()
                                elif '-' in date_str:
                                    register_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                elif '/' in date_str:
                                    register_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass
                    
                    # If time provided in Excel, try to parse it
                    if register_time_value and str(register_time_value).strip() not in ['', '---', '-']:
                        try:
                            if isinstance(register_time_value, time):
                                register_time = register_time_value
                            else:
                                time_str = str(register_time_value).strip()
                                if ':' in time_str:
                                    time_str = time_str.replace('a.m.', '').replace('p.m.', '').replace('AM', '').replace('PM', '').strip()
                                    try:
                                        register_time = datetime.strptime(time_str, '%I:%M').time()
                                    except:
                                        try:
                                            register_time = datetime.strptime(time_str, '%H:%M').time()
                                        except:
                                            pass
                        except:
                            pass
                    
                    # Check if user exists by phone number and role
                    existing_user = User_Details.objects.filter(user_phone=user_phone, user_role=FIXED_ROLE).first()
                    
                    if existing_user:
                        # --- UPDATE MODE: DO NOT CHANGE user_id ---
                        existing_user.user_name = user_name
                        existing_user.user_email = user_email
                        existing_user.user_role = FIXED_ROLE
                        existing_user.user_service_type = user_service_type
                        existing_user.user_company_name = user_company_name
                        existing_user.user_pan_number = user_pan_number
                        existing_user.user_gstin_number = user_gstin_number
                        existing_user.user_operational_scope = operational_scope
                        existing_user.selected_regions = selected_regions
                        existing_user.user_state = user_state
                        existing_user.user_city = user_city
                        existing_user.user_address = user_address
                        existing_user.user_password = user_password
                        existing_user.user_register_date = register_date
                        existing_user.user_register_time = register_time
                        existing_user.save()
                        
                        success_count += 1
                    else:
                        # --- CREATE MODE: GENERATE USER_ID ---
                        new_user = User_Details.objects.create(
                            user_name=user_name,
                            user_email=user_email,
                            user_phone=user_phone,
                            user_role=FIXED_ROLE,
                            user_service_type=user_service_type,
                            user_company_name=user_company_name,
                            user_pan_number=user_pan_number,
                            user_gstin_number=user_gstin_number,
                            user_operational_scope=operational_scope,
                            selected_regions=selected_regions,
                            user_state=user_state,
                            user_city=user_city,
                            user_address=user_address,
                            user_password=user_password,
                            user_register_date=register_date,
                            user_register_time=register_time
                        )
                        
                        # --- GENERATE USER_ID WITH FORMAT: EF-{ID}-{YY} ---
                        current_year = datetime.now().year
                        year_suffix = str(current_year)[-2:]  # Get last 2 digits of year (e.g., 26 for 2026)
                        
                        # Format: EF-{user.id}-{YY}
                        user_id = f"EF-{new_user.id}-{year_suffix}"
                        
                        # Update the user with the generated user_id
                        new_user.user_id = user_id
                        new_user.save()
                        
                        success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Row {row_idx} error: {e}")
            
            msg = f"Successfully imported {success_count} Vendors."
            if error_count > 0:
                msg += f" Failed: {error_count}"
            if skipped_count > 0:
                msg += f" Skipped: {skipped_count}"
            
            return JsonResponse({
                "status": "1",
                "msg": msg,
                "success_count": success_count,
                "error_count": error_count,
                "skipped_count": skipped_count,
                "error_details": error_details[:10]
            })
            
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({"status": "0", "msg": f"Error: {str(e)}"})
    
    return JsonResponse({"status": "0", "msg": "Invalid Request"})


############ Views end for upload vendor data functionality via excel #################


########## Views start for delete vendor ##########################

@csrf_exempt
def Delete_Vendor(request):
    try:
        try:
            vendor_id = request.POST.get('vendor_id')
            User_Details.objects.filter(id=vendor_id).delete()
            return JsonResponse({'status':'1', 'msg':'Vendor details deleted successfully...'})
        except:
            traceback.print_exc()
            return JsonResponse({"status":"0", "msg" : "Something went wrong..."})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############# Views end for delete vendor ###########################


############## Views start for update vendor #######################

def Update_Vendor(request,id):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        services_obj = Service_Type_Details.objects.all()
        vendor = User_Details.objects.get(id=id)

        context = {'admin_obj':admin_obj,'services_obj':services_obj,'vendor':vendor}

        return render(request,'admin_user/Vendor/update_vendor.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

######### Views end for update vendor ##########################


############ Views start for update profile page ########################

def Update_Profile_Admin(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)

        context = {'admin_obj':admin_obj}

        return render(request,'admin_user/Profile/profile_admin.html',context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for update profile page ###########################


############ Views start for ajax for update profile #######################

@csrf_exempt
def Admin_Profile_Ajax(request):
    data=request.POST.dict()
    try:
        Admin_Login.objects.get(id=data['id'])
        Admin_Login.objects.filter(id=data['id']).update(**data)
        return JsonResponse({"status":"1", "msg" : f"Profile updated successfully"})
    except:
        traceback.print_exc()
        return JsonResponse({"status":"0", "msg" : "Something went wrong..."})

############## Views end for ajax for update profile ##########################



def broadcast_email(request):
    return render(request,"admin_user/broadcast_email.html")



def send_message(request):
   
   
    if request.method == 'POST':
        WhatsAppMessage.objects.create(
            phone_number=request.POST.get('phone_number'),
            template=request.POST.get('template'),
            message=request.POST.get('message')
        )
       # return redirect('whatsapp_message:message_list')

   # return render(request, 'whatsapp_message/message_form.html')
    return render(request,"admin_user/send_message.html")



def commision_release_cycle(request):
    return render(request,"admin_user/commision_release_cycle.html")





from django.core.files.images import get_image_dimensions

def hero_section(request):
    if request.method == "POST":
        title = request.POST.get("title")
        subtitle = request.POST.get("subtitle")
        title_font_size = request.POST.get("title_font_size")
        subtitle_font_size = request.POST.get("subtitle_font_size")
        text_color = request.POST.get("text_color")
        overlay_color = request.POST.get("overlay_color")
        is_active = True if request.POST.get("is_active") == "on" else False

        background_image = request.FILES.get("background_image")

        # Validate image
        if background_image:
            # File size (max 2MB)
            if background_image.size > 2 * 1024 * 1024:
                messages.error(request, "Image size must be under 2MB")
                return redirect("hero_section_form")

            # Resolution (min 1200x600)
            width, height = get_image_dimensions(background_image)
            if width < 1200 or height < 600:
                messages.error(request, "Image resolution must be at least 1200x600 pixels")
                return redirect("hero_section_form")

        # Save to DB
        HeroSection.objects.create(
            title=title,
            subtitle=subtitle,
            title_font_size=title_font_size,
            subtitle_font_size=subtitle_font_size,
            text_color=text_color,
            overlay_color=overlay_color,
            background_image=background_image,
            is_active=is_active
        )

        messages.success(request, "Hero section saved successfully!")
       # return redirect("home")

    return render(request, "admin_user/hero_section.html")



def hero_section_list(request):
    heros = HeroSection.objects.all().order_by("-id")
    return render(request, "admin_user/hero_section_list.html", {"heros": heros})


def hero_section_edit(request, pk):
    hero = get_object_or_404(HeroSection, pk=pk)

    if request.method == "POST":
        hero.title = request.POST.get("title")
        hero.subtitle = request.POST.get("subtitle")
        hero.title_font_size = request.POST.get("title_font_size")
        hero.subtitle_font_size = request.POST.get("subtitle_font_size")
        hero.text_color = request.POST.get("text_color")
        hero.overlay_color = request.POST.get("overlay_color")
        hero.is_active = True if request.POST.get("is_active") == "on" else False

        if "background_image" in request.FILES:
            hero.background_image = request.FILES["background_image"]

        hero.save()
        messages.success(request, "Hero section updated successfully!")
        return redirect("hero_section_list")

    return render(request, "admin_user/hero_section_edit.html", {"hero": hero})


def hero_section_delete(request, pk):
    hero = get_object_or_404(HeroSection, pk=pk)
    hero.delete()
    messages.success(request, "Hero section deleted successfully!")
    return redirect("hero_section_list")


def hero_section_toggle(request, pk):
    hero = get_object_or_404(HeroSection, pk=pk)
    hero.is_active = not hero.is_active
    hero.save()
    messages.success(request, f"{hero.title} status updated successfully.")
    return redirect("hero_section_list")



from django.utils.text import slugify





# services/views.py



# ADD view
def add_about(request):
    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        intro_badge = request.POST.get("intro_badge")
        intro_heading = request.POST.get("intro_heading")
        intro_content = request.POST.get("intro_content")
        founder_name = request.POST.get("founder_name")
        founder_role = request.POST.get("founder_role")
        founder_quote = request.POST.get("founder_quote")

        founder_image = request.FILES.get("founder_image")
        main_image = request.FILES.get("main_image")
        overlay_image = request.FILES.get("overlay_image")

        AboutPage.objects.create(
            title=title,
            description=description,
            intro_badge=intro_badge,
            intro_heading=intro_heading,
            intro_content=intro_content,
            founder_name=founder_name,
            founder_role=founder_role,
            founder_quote=founder_quote,
            founder_image=founder_image,
            main_image=main_image,
            overlay_image=overlay_image,
            years_of_excellence=request.POST.get("years_of_excellence", 1),
        )
      #  return redirect("about")  # show frontend about page

    return render(request, "admin_user/add_about.html")


# EDIT view
def edit_about(request, pk):
    about = get_object_or_404(AboutPage, pk=pk)

    if request.method == "POST":
        about.title = request.POST.get("title")
        about.description = request.POST.get("description")
        about.intro_badge = request.POST.get("intro_badge")
        about.intro_heading = request.POST.get("intro_heading")
        about.intro_content = request.POST.get("intro_content")
        about.founder_name = request.POST.get("founder_name")
        about.founder_role = request.POST.get("founder_role")
        about.founder_quote = request.POST.get("founder_quote")
        about.years_of_excellence = request.POST.get("years_of_excellence", 1)

        if request.FILES.get("founder_image"):
            about.founder_image = request.FILES.get("founder_image")
        if request.FILES.get("main_image"):
            about.main_image = request.FILES.get("main_image")
        if request.FILES.get("overlay_image"):
            about.overlay_image = request.FILES.get("overlay_image")

        about.save()
        return redirect("about")

    return render(request, "edit_about.html", {"about": about})





def achievements_page(request):
    achievements = Achievement.objects.all()

    if request.method == "POST":
        pk = request.POST.get("pk")
        if pk:
            achievement = get_object_or_404(Achievement, pk=pk)
            achievement.icon_class = request.POST.get("icon_class")
            achievement.number = request.POST.get("number")
            achievement.suffix = request.POST.get("suffix")
            achievement.label = request.POST.get("label")
            achievement.order = request.POST.get("order", 0)
            achievement.save()
        else:
            Achievement.objects.create(
                icon_class=request.POST.get("icon_class"),
                number=request.POST.get("number"),
                suffix=request.POST.get("suffix"),
                label=request.POST.get("label"),
                order=request.POST.get("order", 0),
            )
        #return redirect("achievements_page")

    return render(request, "admin_user/achievements_page.html", {"achievements": achievements})





def faq_list_admin(request):
    faqs = FAQ.objects.all().order_by('-created_at')
    return render(request, 'faq_list_admin.html', {'faqs': faqs})

()
def faq_add(request):
    if request.method == 'POST':
        question = request.POST['question']
        answer = request.POST['answer']
        FAQ.objects.create(question=question, answer=answer)
       # return redirect('faq_list_admin')
    return render(request, 'admin_user/faq_add.html')


def faq_edit(request, faq_id):
    faq = get_object_or_404(FAQ, id=faq_id)
    if request.method == 'POST':
        faq.question = request.POST['question']
        faq.answer = request.POST['answer']
        faq.save()
        return redirect('faq_list_admin')
    return render(request, 'faq_edit.html', {'faq': faq})


def faq_delete(request, faq_id):
    faq = get_object_or_404(FAQ, id=faq_id)
    if request.method == 'POST':
        faq.delete()
        return redirect('faq_list_admin')
    return render(request, 'faq_delete.html', {'faq': faq})

def faq_list_public(request):
    faqs = FAQ.objects.all().order_by('-created_at')
    return render(request, 'faq_list_public.html', {'faqs': faqs})





def timeline_page(request):
    timeline_items = TimelineItem.objects.all

    if request.method == "POST":
        pk = request.POST.get("pk")
        if pk:
            timeline = get_object_or_404(TimelineItem, pk=pk)
            timeline.year = request.POST.get("year")
            timeline.title = request.POST.get("title")
            timeline.description = request.POST.get("description")
            timeline.order = request.POST.get("order", 0)
            timeline.save()
        else:
            TimelineItem.objects.create(
                year=request.POST.get("year"),
                title=request.POST.get("title"),
                description=request.POST.get("description"),
                order=request.POST.get("order", 0),
            )
        #return redirect("timeline_page")

    return render(request, "admin_user/timeline_page.html", {"timeline_items": timeline_items})




def add_ad(request):
    if request.method == "POST":
        title = request.POST.get("title")
        category = request.POST.get("category")
        image = request.FILES.get("image")
        short_description = request.POST.get("short_description")
        detail_content = request.POST.get("detail_content")
        badge_text = request.POST.get("badge_text")
        badge_icon = request.POST.get("badge_icon")
        special_offer_title = request.POST.get("special_offer_title")
        special_offer_description = request.POST.get("special_offer_description")
        text_size_heading = request.POST.get("text_size_heading")
        text_size_paragraph = request.POST.get("text_size_paragraph")
        slug = request.POST.get("slug")

        ad = Ad(
            title=title,
            category=category,
            image=image,
            short_description=short_description,
            detail_content=detail_content,
            badge_text=badge_text,
            badge_icon=badge_icon,
            special_offer_title=special_offer_title,
            special_offer_description=special_offer_description,
            text_size_heading=text_size_heading,
            text_size_paragraph=text_size_paragraph,
            slug=slug
        )
        ad.save()
     #   return redirect("ad_list")

    return render(request, "admin_user/add_ad.html")









# seo/views.py




def toggle_seo_status(request, pk):
    seo_page = get_object_or_404(LocationSEO, pk=pk)
    seo_page.is_active = not seo_page.is_active
    seo_page.save()
    return redirect("seo_list")


def delete_seo_page(request, pk):
    seo_page = get_object_or_404(LocationSEO, pk=pk)
    seo_page.delete()
    return redirect("seo_list")




def edit_seo_page(request, pk):
    seo_page = get_object_or_404(LocationSEO, pk=pk)
   
    if request.method == "POST":
        seo_page.meta_title = request.POST.get("meta_title")
        seo_page.meta_description = request.POST.get("meta_description")
        seo_page.primary_keyword = request.POST.get("primary_keyword")
        seo_page.secondary_keywords = request.POST.get("secondary_keywords")
        seo_page.intro_html = request.POST.get("intro_html")
        seo_page.noindex = "noindex" in request.POST
        seo_page.is_active = "is_active" in request.POST
        seo_page.save()
        return redirect("seo_list")

    return render(request, "admin_user/seo_edit.html", {"page": seo_page})


def services_list1(request):
    services = LocationSEO.objects.filter(page_type="service", is_active=True)
    return render(request, "admin_user/services_list1.html", {"services": services})




def plans_list(request):
    plans = Plan.objects.all().order_by('-id')
    return render(request, "admin_user/plans_list.html", {'plans': plans})

def plan_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        plan_type = request.POST.get('plan_type')
        desc = request.POST.get('description')
        base_price = request.POST.get('base_price') or 0
        roles_visible = request.POST.get('roles_visible') or ''
        is_active = bool(request.POST.get('is_active'))
        plan = Plan.objects.create(
            name=name, plan_type=plan_type, description=desc,
            base_price=Decimal(base_price), roles_visible=roles_visible, is_active=is_active
        )
        return redirect('plans_list')
    return render(request, 'admin_user/plan_add.html')

def plan_edit(request, pid):
    plan = get_object_or_404(Plan, id=pid)
    if request.method == 'POST':
        plan.name = request.POST.get('name')
        plan.plan_type = request.POST.get('plan_type')
        plan.description = request.POST.get('description')
        plan.base_price = Decimal(request.POST.get('base_price') or 0)
        plan.roles_visible = request.POST.get('roles_visible') or ''
        plan.is_active = bool(request.POST.get('is_active'))
        plan.save()
        return redirect('plans_list')
    return render(request, 'admin_user/plan_edit.html', {'plan': plan})




# --- Add-On Create View ---
def addon_create(request):
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        price = request.POST['price']
        applicableroles = request.POST.get('applicableroles', '')
        isactive = 'isactive' in request.POST

        Addon.objects.create(
            name=name,
            description=description,
            price=price,
            applicableroles=applicableroles,
            isactive=isactive,
        )
        messages.success(request, "Add-On created successfully!")
        return redirect('addon_list')

    return render(request, 'admin_user/addon_create.html')


# --- Add-On List View ---








 

##################################RESIDENTIAL RENTAL LISTING VEIW SECTION START##############################




def rental_residential_logs_view(request):
    # Base query initialization
    logs = RentalActivityLog.objects.all()

    # Capture URL GET search query parameters
    user_query = request.GET.get('user_query', '').strip()
    file_query = request.GET.get('file_query', '').strip()
    location_query = request.GET.get('location_query', '').strip()
    property_type = request.GET.get('property_type', '').strip()
    bhk_type = request.GET.get('bhk_type', '').strip()
    min_budget = request.GET.get('min_budget', '').strip()
    max_budget = request.GET.get('max_budget', '').strip()
    field_target = request.GET.get('field_target', '').strip()
    action_type = request.GET.get('action_type', '').strip()
    month_filter = request.GET.get('month_filter', '').strip()
    
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    start_time_str = request.GET.get('start_time', '').strip()
    end_time_str = request.GET.get('end_time', '').strip()

    # 1. Build Property Specifications Pool Filter First
    property_pool = RentalResidentialProperty.objects.filter(is_deleted=False)
    has_property_filter = False

    if property_type:
        property_pool = property_pool.filter(property_type__icontains=property_type)
        has_property_filter = True
    if bhk_type:
        property_pool = property_pool.filter(bhk_type__icontains=bhk_type)
        has_property_filter = True
    if location_query:
        property_pool = property_pool.filter(
            Q(city__icontains=location_query) | 
            Q(locality__icontains=location_query) | 
            Q(address__icontains=location_query)
        )
        has_property_filter = True
    if min_budget and min_budget.isdigit():
        property_pool = property_pool.filter(monthly_rent__gte=int(min_budget))
        has_property_filter = True
    if max_budget and max_budget.isdigit():
        property_pool = property_pool.filter(monthly_rent__lte=int(max_budget))
        has_property_filter = True

    matched_property_ids = list(property_pool.values_list('rental_residential_id', flat=True))

    # 2. Apply Filters Across the Audit Logs Queryset
    log_conditions = Q()

    if user_query:
        log_conditions &= (Q(user_identity__icontains=user_query) | Q(user_role__icontains=user_query))
    if file_query:
        clean_file = file_query.replace('.xlsx', '').replace('.xls', '').strip()
        log_conditions &= Q(associated_file__icontains=clean_file)
    if action_type:
        log_conditions &= Q(action_type__iexact=action_type)
    if month_filter and month_filter.isdigit():
        log_conditions &= Q(timestamp__month=int(month_filter))
    if field_target:
        if field_target == 'city_locality':
            log_conditions &= (Q(targeted_fields__icontains='city') | Q(targeted_fields__icontains='locality'))
        elif field_target == 'owner_contact':
            log_conditions &= (Q(targeted_fields__icontains='owner') | Q(targeted_fields__icontains='contact'))
        else:
            log_conditions &= Q(targeted_fields__icontains=field_target)

    # Date-wise & Time-wise Precision Datetime Combiner Lookups
    if start_date_str:
        try:
            start_date_parsed = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            if start_time_str:
                start_time_parsed = datetime.strptime(start_time_str, "%H:%M").time()
                log_conditions &= Q(timestamp__gte=datetime.combine(start_date_parsed, start_time_parsed))
            else:
                log_conditions &= Q(timestamp__date__gte=start_date_parsed)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date_parsed = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            if end_time_str:
                end_time_parsed = datetime.strptime(end_time_str, "%H:%M").time()
                log_conditions &= Q(timestamp__lte=datetime.combine(end_date_parsed, end_time_parsed))
            else:
                log_conditions &= Q(timestamp__date__lte=end_date_parsed)
        except ValueError:
            pass

    # Merge search targets dynamically
    if has_property_filter:
        payload_q = Q(property_id__in=matched_property_ids)
        if bhk_type: payload_q |= Q(action_payload__icontains=bhk_type)
        if property_type: payload_q |= Q(action_payload__icontains=property_type)
        if location_query: payload_q |= Q(action_payload__icontains=location_query)
        log_conditions &= payload_q

    filtered_logs = logs.filter(log_conditions)

    # 3. Compile the Comprehensive Final Properties Presentation Array
    logs_property_ids = filtered_logs.exclude(property_id="Multiple / Sheet Records").values_list('property_id', flat=True).distinct()
    
    final_properties_queryset = RentalResidentialProperty.objects.filter(
        Q(rental_residential_id__in=logs_property_ids) | Q(rental_residential_id__in=matched_property_ids),
        is_deleted=False
    ).prefetch_related('images').distinct()

    # Calculate real-time totals for the Property Counter Badge
    properties_filtered_count = final_properties_queryset.count()
    properties_total_count = RentalResidentialProperty.objects.filter(is_deleted=False).count()

    # 4. Compute Metrics for the Top KPI Cards Block
    total_logs_count = filtered_logs.count()
    update_logs_count = filtered_logs.filter(action_type__iexact='UPDATE').count()
    delete_logs_count = filtered_logs.filter(action_type__iexact='DELETE').count()
    import_logs_count = filtered_logs.filter(action_type__iexact='EXCEL_IMPORT').count()

    # 5. LIVE EXPORT LOGIC FOR CSV AND EXCEL TRANSFERS
    download_format = request.GET.get('download', '').strip()
    if download_format in ['csv', 'excel']:
        filename_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if download_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="filtered_activity_logs_{filename_stamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Sr.No', 'Timestamp', 'Property ID', 'Operator Identity', 'User System Role', 'Action', 'Changed Fields', 'Associated Excel File', 'Client IP Address', 'Execution Status'])
            for idx, item in enumerate(filtered_logs, 1):
                writer.writerow([idx, item.timestamp.strftime("%Y-%m-%d %H:%M"), item.property_id, item.user_identity, item.user_role, item.action_type, item.targeted_fields, item.associated_file, item.ip_address, item.status])
            return response
            
        elif download_format == 'excel':
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="filtered_properties_matrix_{filename_stamp}.xls"'
            writer = csv.writer(response, delimiter='\t')
            writer.writerow(['Property ID', 'Property Title', 'Property Type', 'BHK Config Type', 'Monthly Rent', 'Security Deposit', 'City', 'Locality', 'Uploaded By', 'Source File Name'])
            for item in final_properties_queryset:
                writer.writerow([item.rental_residential_id, item.property_title, item.property_type, item.bhk_type, item.monthly_rent, item.security_deposit, item.city, item.locality, item.uploaded_by_name, item.upload_file_name])
            return response

    # 6. Pagination System Engine Layout Slices
    log_paginator = Paginator(filtered_logs, 15)
    log_records_list = log_paginator.get_page(request.GET.get('log_page', 1))

    prop_paginator = Paginator(final_properties_queryset, 10)
    page_obj = prop_paginator.get_page(request.GET.get('prop_page', 1))

    context = {
        'log_records_list': log_records_list,
        'page_obj': page_obj,
        'properties_filtered_count': properties_filtered_count,
        'properties_total_count': properties_total_count,
        'total_logs_count': total_logs_count,
        'update_logs_count': update_logs_count,
        'delete_logs_count': delete_logs_count,
        'import_logs_count': import_logs_count,
    }
    return render(request, "admin_user/Reports/Rental/rental_residential_activity_logs.html", context)





def _get_client_ip(request):
    """Helper to safely fetch client IP address reference."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')



def _get_deleter_name(request):
    """Helper function to get the name of the person deleting the property."""
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if admin_id:
        try:
            admin = Admin_Login.objects.get(id=admin_id)
            return f"{admin.name} (Admin)"
        except Admin_Login.DoesNotExist:
            return "Unknown Admin"
            
    elif user_id:
        try:
            user = User_Details.objects.get(id=user_id)
            return f"{user.user_name} (User)"
        except User_Details.DoesNotExist:
            return "Unknown User"
            
    return "System"







import hashlib
import re


def generate_property_fingerprint(property_no, building_name, locality, pincode=""):
    """
    Generates a normalized SHA-256 hash unique key based on physical location.
    Strips spaces and special characters for consistent matching.
    """
    def clean_text(val):
        if not val:
            return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(val)).upper()

    raw_string = f"{clean_text(property_no)}_{clean_text(building_name)}_{clean_text(locality)}_{clean_text(pincode)}"
    return hashlib.sha256(raw_string.encode('utf-8')).hexdigest()


def rental_residential_add(request):
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    admin_obj = None
    user_obj = None

    if admin_id:
        admin_obj = Admin_Login.objects.filter(id=admin_id).first()

    if user_id:
        user_obj = User_Details.objects.filter(id=user_id).first()

    if not admin_obj and not user_obj:
        return render(request, 'home_page/Adminlogin.html')

    if request.method == 'POST':
        try:
            # ---------- SAFE TYPE CONVERSIONS ----------
            def to_int(val):
                try:
                    return int(val) if val else None
                except:
                    return None

            def to_decimal(val):
                try:
                    return float(val) if val else None
                except:
                    return None

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

            # ---------- AMENITIES & FACILITIES ----------
            amenities = ",".join(request.POST.getlist('amenities[]'))
            facilities = ",".join(request.POST.getlist('nearby_facilities[]')) or ",".join(request.POST.getlist('facilities[]'))

            # ---------- UPLOADER IDENTIFICATION (Who submitted the HTML form) ----------
            if admin_obj:
                uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
                uploader_email = getattr(admin_obj, 'email', '')
                uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
                uploader_role = "Admin"
                uploader_id = f"ADMIN_{admin_id}"
            elif user_obj:
                uploader_name = user_obj.user_name
                uploader_email = user_obj.user_email
                uploader_contact = user_obj.user_phone
                uploader_role = "User"
                uploader_id = f"USER_{user_id}"
            else:
                uploader_name, uploader_email, uploader_contact, uploader_role, uploader_id = "", "", "", "", ""

            # ---------- LISTED BY IDENTIFICATION (Who owns/manages the listing) ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ==========================================================
            # DUPLICATE DETECTION ENGINE (Checking LISTED BY, not UPLOADED BY)
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality_area') or request.POST.get('locality') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            fingerprint_key = generate_property_fingerprint(
                input_property_no, 
                input_building_name, 
                input_locality, 
                input_pincode
            )

            # 1. Direct Case-Insensitive Query for same unit in same locality/building
            direct_duplicates = RentalResidentialProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality_area__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            # Combine fingerprint match OR direct field match
            existing_duplicates = (
                RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                # Level 1: Hard Block if LISTED BY the exact same person (ID, Email, OR Phone match)
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        # INSTANT REJECTION: Same agent/owner cannot list the same property twice
                        messages.error(
                            request, 
                            f"Duplicate Blocked: This property (Unit {input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing instead."
                        )
                        return redirect('residential')

                # Level 2: Different Agent/User listing the exact same physical unit -> Allow save & Flag
                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- CREATE DATABASE OBJECT ----------
            prop = RentalResidentialProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Residential",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                property_no=input_property_no,
                bhk_type=request.POST.get('bhk_type'),
                renting_option=request.POST.get('renting_option'),
                built_up_area=to_decimal(request.POST.get('built_up_area')),
                bathrooms=to_int(request.POST.get('bathrooms')),
                balconies=to_int(request.POST.get('balconies')),
                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int(request.POST.get('total_floors')),
                facing_direction=request.POST.get('facing_direction', request.POST.get('facing')),
                furnishing_status=request.POST.get('furnishing_status'),
                available_for=request.POST.get('available_for'),

                carpet_area=to_decimal(request.POST.get('carpet_area')),
                city_zone=request.POST.get('city_zone', request.POST.get('zone')),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition', request.POST.get('construction_status')),
                property_age=request.POST.get('property_age'),
                wing_number=request.POST.get('wing_number'),
                building_name=input_building_name,

                availability_status=request.POST.get('availability_status', request.POST.get('possession_status')),
                available_from=available_from,
                lease_duration=request.POST.get('lease_duration'),
                brokerage_percentage=request.POST.get('brokerage_percentage', request.POST.get('brokerage')),
                manual_brokerage=request.POST.get('manual_brokerage'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=to_int(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int(request.POST.get('security_deposit_amount', request.POST.get('security_deposit'))),
                maintenance_type=request.POST.get('maintenance_type'),
                monthy_maintenance_amount=to_int(request.POST.get('monthy_maintenance_amount', request.POST.get('maintenance_amount'))),
                total_move_in_cost=to_int(request.POST.get('total_move_in_cost')),

                address=request.POST.get('address'),
                city=request.POST.get('city'),
                locality_area=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=input_pincode,
                main_road_connectivity=request.POST.get('main_road_connectivity', request.POST.get('road_connectivity')),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                amenities=amenities,
                nearby_facilities=facilities,

                user_description=request.POST.get('user_description'),
                description=request.POST.get('description'),
                rent_residential_desc=request.POST.get('rent_residential_desc'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=None
            )

         
            # ---------- IMAGES MULTI-UPLOAD LOGIC (CATEGORY WISE) ----------
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'living':   'property_images_living[]',
                'bedroom':  'property_images_bedroom[]',
                'kitchen':  'property_images_kitchen[]',
                'bathroom': 'property_images_bathroom[]',
                'balcony':  'property_images_balcony[]',
                'others':   'property_images_others[]',
            }

            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 20:
                        break
                    RentalResidentialImage.objects.create(
                        property=prop,
                        image=img,
                        category=category,
                        sequence_order=cat_index,
                    )
                    saved_count += 1

          # ---------- PROPERTY VIDEO (UPLOAD, RM LINK, OR AUTO SLIDESHOW) ----------
            # ---------- PROPERTY VIDEO LOGIC (AUTO SLIDESHOW + UPLOAD / RM LINK) ----------
            video_option = request.POST.get('video_option') or request.POST.get('video_source') or 'auto'
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '').strip()

            # 1. ALWAYS auto-generate the slideshow row if >= 3 photos exist
            CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
            saved_images = list(RentalResidentialImage.objects.filter(property=prop))
            saved_images.sort(key=lambda img: (CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, img.sequence_order))
            image_paths = [img.image.path for img in saved_images if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)]

            if len(image_paths) >= 3:
                try:
                    from .utils.video_generator import generate_property_slideshow
                    import time
                    out_path = f"residential_rent/videos/auto_{prop.id}_{int(time.time())}.mp4"
                    result = generate_property_slideshow(image_paths, out_path)
                    if result:
                        RentalResidentialVideo.objects.create(
                            property=prop,
                            video=out_path,
                            source='auto'
                        )
                except Exception as ve:
                    print("RESIDENTIAL VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()

            # 2. Save Manual Upload Video as a separate row
            if video_option == 'upload' and uploaded_video:
                RentalResidentialVideo.objects.create(
                    property=prop,
                    video=uploaded_video,
                    source='uploaded'
                )

            # 3. Save RM Assisted Link Video as a separate row
            elif video_option == 'rm_assisted' and property_video_link:
                RentalResidentialVideo.objects.create(
                    property=prop,
                    video_url=property_video_link,
                    source='rm_assisted'
                )
            messages.success(request, "Property Added Successfully ")
            
            return redirect('residential_list')

        except Exception as e:
            print("ERROR DETECTED:", str(e))
            messages.error(request, f"Error while saving listing: {str(e)}")
            return redirect('residential')

    return render(request, 'admin_user/Reports/Rental/rental_list.html', {
        'admin_obj': admin_obj,
        'user_obj': user_obj,
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all()
    })




from django.utils.dateparse import parse_date





def rental_reports(request):

    session_id = request.session.get('Admin_id')

    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════

    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', '').strip()
    state_query = request.GET.get('state', '').strip()          # NEW
    furnish_query = request.GET.get('furnishing', '').strip()
    possession_query = request.GET.get('possession', '').strip()  # maps to availability_status

    prop_type_query = request.GET.get('property_type', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()     # listed_by_role
    uploaded_by_query = request.GET.get('uploaded_by', '').strip()  # NEW -> uploaded_by_role
    budget_query = request.GET.get('budget', '').strip()

    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query = request.GET.get('duplicate', '').strip()

    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    # ═══════════════════════════════════════
    # BASE QUERYSET & PERSISTENT SR.NO MAP
    # ═══════════════════════════════════════

    base_properties = RentalResidentialProperty.objects.filter(is_deleted=False).order_by('-id')
    absolute_ordered_ids = list(base_properties.values_list('id', flat=True))

    properties = base_properties

    # ═══════════════════════════════════════
    # GLOBAL SEARCH FILTER (Across ALL Fields + Sr.No)
    # ═══════════════════════════════════════

    if search_query:
        sr_no_query = Q()

        if search_query.isdigit():
            target_index = int(search_query) - 1
            if 0 <= target_index < len(absolute_ordered_ids):
                target_id = absolute_ordered_ids[target_index]
                sr_no_query = Q(id=target_id)

        properties = properties.filter(
            sr_no_query |
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(renting_option__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality_area__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(available_for__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(upload_file_name__icontains=search_query)
        )

    # ═══════════════════════════════════════
    # ADVANCED DROPDOWN FILTERS
    # ═══════════════════════════════════════

    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__iexact=city_query)

    if state_query and state_query != 'All States':          # NEW
        properties = properties.filter(state__iexact=state_query)

    if furnish_query and furnish_query != 'All':
        properties = properties.filter(furnishing_status__iexact=furnish_query)

    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__iexact=possession_query)

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__iexact=prop_type_query)

    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':   # NEW
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_10k':
            properties = properties.filter(monthly_rent__lt=10000)
        elif budget_query == '10k_25k':
            properties = properties.filter(monthly_rent__gte=10000, monthly_rent__lte=25000)
        elif budget_query == '25k_50k':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=50000)
        elif budget_query == 'above_50k':
            properties = properties.filter(monthly_rent__gt=50000)

    # Date Filters
    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            properties = properties.filter(created_at__date__gte=from_date)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            properties = properties.filter(created_at__date__lte=to_date)

    # ═══════════════════════════════════════
    # PAGINATION & INJECT PERSISTENT SR.NO
    # ═══════════════════════════════════════
    filtered_count = properties.count()

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for prop in page_obj:
        prop.original_sr_no = absolute_ordered_ids.index(prop.id) + 1

    # ═══════════════════════════════════════
    # EXPORT DATA (CSV / EXCEL)
    # ═══════════════════════════════════════
    if request.GET.get('download') in ['excel', 'csv']:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import OrderedDict
        import csv

        sections = {
            "Basic Info": [
                "id", "property_title", "property_type", "renting_option",
                "built_up_area", "bathrooms", "balconies",
                "building_configuration", "total_floors", "facing_direction",
                "furnishing_status", "available_for"
            ],
            "Property Details": [
                "city_zone", "ownership_type", "property_condition", "property_age",
                "carpet_area", "wing_number", "building_name"
            ],
            "Availability": [
                "availability_status", "available_from", "lease_duration",
                "brokerage_percentage", "manual_brokerage"
            ],
            "Pricing": [
                "monthly_rent", "advance_rent_month", "advance_rent_amount",
                "security_deposit_type", "security_deposit_amount",
                "maintenance_type", "monthy_maintenance_amount", "total_move_in_cost"
            ],
            "Location": [
                "address", "city", "locality_area", "property_landmark", "state",
                "pincode", "main_road_connectivity", "google_maps_link"
            ],
            "Description & Features": [
                "amenities", "nearby_facilities", "description",
                "rent_residential_desc", "user_description"
            ],
            "Listed By Info": [
                "listed_by_name", "listed_by_contact", "listed_by_email",
                "listed_by_role", "listed_by_type"
            ],
            "System Data": [
                "uploaded_by_name", "uploaded_by_email",
                "uploaded_by_contact", "uploaded_by_role", "upload_file_name", "created_at"
            ],
        }

        HINTS = {
            "id": "Property ID Auto-Generated",
            "property_title": "Auto_Generated Title", "property_type": "Apartment",
            "renting_option": "Full Property/Single Room/Shared Room", "built_up_area": "sq.ft",
            "bathrooms": "Number", "balconies": "Number", "building_configuration": "e.g. G+3",
            "total_floors": "Number", "facing_direction": "North/East", "furnishing_status": "Semi Furnished",
            "available_for": "Family/Bachelor", "city_zone": "North/South", "ownership_type": "Freehold",
            "property_condition": "Resale", "property_age": "1-3 Years", "carpet_area": "sq.ft",
            "wing_number": "e.g. A/B", "building_name": "Text", "availability_status": "Ready to Move",
            "available_from": "YYYY-MM-DD", "lease_duration": "11 Months", "brokerage_percentage": "1%/Fixed Amount",
            "manual_brokerage": "e.g. 2.5%", "monthly_rent": "₹", "advance_rent_month": "0-11/fixed",
            "advance_rent_amount": "₹", "security_deposit_type": "0-11/fixed", "security_deposit_amount": "₹",
            "maintenance_type": "Included in Rent/Extra", "monthy_maintenance_amount": "₹",
            "total_move_in_cost": "₹", "address": "Full Address", "city": "Text", "locality_area": "Text",
            "property_landmark": "Optional", "state": "e.g. Maharashtra", "pincode": "6-digit",
            "main_road_connectivity": "Optional", "google_maps_link": "URL", "amenities": "Comma-sep",
            "nearby_facilities": "Comma-sep", "description": "Short Summary",
            "rent_residential_desc": "Long Rich Text", "user_description": "Added by user",
            "listed_by_name": "Full Name", "listed_by_contact": "10 Digits", "listed_by_email": "email@example.com",
            "listed_by_role": "Owner/Agent/Admin", "listed_by_type": "Self/Other",
            "uploaded_by_name": "Admin Name", "uploaded_by_email": "Admin Email",
            "uploaded_by_contact": "Admin Contact", "uploaded_by_role": "Admin Role",
            "upload_file_name": "File Name", "created_at": "YYYY-MM-DD"
        }

        REQUIRED = {
            "property_type", "renting_option", "built_up_area", "bathrooms",
            "furnishing_status", "available_for", "monthly_rent",
            "address", "city", "locality_area", "state", "pincode",
            "listed_by_name", "listed_by_contact", "listed_by_email"
        }

        all_cols = []
        for sec, fields in sections.items():
            all_cols.extend([(sec, f) for f in fields])

        if request.GET.get('download') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rental Residential"

            HDR_BG, REQ_BG, OPT_BG = "667EEA", "FEF3C7", "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

            sec_spans = OrderedDict()
            for i, (sec, _) in enumerate(all_cols):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            for ci, (sec, field) in enumerate(all_cols, 1):
                req = field in REQUIRED
                lc = ws.cell(row=2, column=ci, value=field + (" *" if req else ""))
                lc.font = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border = bdr

                hc = ws.cell(row=3, column=ci, value=HINTS.get(field, ""))
                hc.font = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(16, len(field) + 4)

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            for row_idx, p in enumerate(properties, start=4):
                for col_idx, (sec, field) in enumerate(all_cols, 1):
                    val = getattr(p, field, "")
                    if field in ['available_from', 'created_at'] and val:
                        val = val.strftime('%Y-%m-%d')

                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Rental_Properties_Data.xlsx"'
            wb.save(response)
            return response

        elif request.GET.get('download') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Rental_Properties_Data.csv"'
            writer = csv.writer(response)

            row1 = []
            current_sec = ""
            for sec, _ in all_cols:
                if sec != current_sec:
                    row1.append(f"📋 {sec}")
                    current_sec = sec
                else:
                    row1.append("")
            writer.writerow(row1)

            writer.writerow([field + (" *" if field in REQUIRED else "") for _, field in all_cols])
            writer.writerow([HINTS.get(field, "") for _, field in all_cols])

            for p in properties:
                data_row = []
                for _, field in all_cols:
                    val = getattr(p, field, "")
                    if field in ['available_from', 'created_at'] and val:
                        val = val.strftime('%Y-%m-%d')
                    data_row.append(val)
                writer.writerow(data_row)

            return response

    # ═══════════════════════════════════════
    # STATS & UNIQUE DROPDOWN DATA
    # ═══════════════════════════════════════

    all_props = RentalResidentialProperty.objects.filter(is_deleted=False)
    total_count = all_props.count()

    unique_cities = all_props.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_states = all_props.exclude(state__isnull=True).exclude(state='').values_list('state', flat=True).distinct()  # NEW
    unique_furnish = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values_list('furnishing_status', flat=True).distinct()
    unique_possession = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').values_list('availability_status', flat=True).distinct()
    unique_property_types = all_props.exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    unique_listed_roles = all_props.exclude(listed_by_role__isnull=True).exclude(listed_by_role='').values_list('listed_by_role', flat=True).distinct()   # NEW split
    unique_uploaded_roles = all_props.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()  # NEW split

    active_count = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    furnished_count = all_props.filter(furnishing_status__iexact='Furnished').count()
    available_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    city_count = all_props.exclude(city__isnull=True).exclude(city='').values('city').distinct().count()
    state_count = all_props.exclude(state__isnull=True).exclude(state='').values('state').distinct().count()  # NEW

    # ═══════════════════════════════════════
    # RENT STATS
    # ═══════════════════════════════════════

    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )

    avg_rent = rent_stats['avg_rent']
    max_rent = rent_stats['max_rent']
    min_rent = rent_stats['min_rent']

    deposit_stats = all_props.exclude(security_deposit_amount__isnull=True).aggregate(avg_deposit=Avg('security_deposit_amount'))
    avg_deposit = deposit_stats['avg_deposit']

    area_stats = all_props.exclude(built_up_area__isnull=True).aggregate(avg_area=Avg('built_up_area'))
    avg_area = area_stats['avg_area']

    with_owner_count = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    with_images_count = all_props.filter(images__isnull=False).distinct().count()

    uploaded_files = all_props.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    # ═══════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════

    renting_option_qs = all_props.exclude(renting_option__isnull=True).exclude(renting_option='').values('renting_option').annotate(count=Count('id')).order_by('-count')
    renting_option_labels = json.dumps([item['renting_option'] for item in renting_option_qs])
    renting_option_data = json.dumps([item['count'] for item in renting_option_qs])

    rent_buckets = [
        ('Under ₹5k', 0, 5000), ('₹5k–10k', 5000, 10000), ('₹10k–20k', 10000, 20000),
        ('₹20k–30k', 20000, 30000), ('₹30k–50k', 30000, 50000), ('₹50k–1L', 50000, 100000),
        ('Above ₹1L', 100000, 999999999),
    ]

    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data = json.dumps([all_props.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values('furnishing_status').annotate(count=Count('id')).order_by('-count')
    furnishing_labels = json.dumps([item['furnishing_status'] for item in furnish_qs])
    furnishing_data = json.dumps([item['count'] for item in furnish_qs])

    prop_type_qs = all_props.exclude(property_type__isnull=True).exclude(property_type='').values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels = json.dumps([item['property_type'] for item in prop_type_qs])
    prop_type_data = json.dumps([item['count'] for item in prop_type_qs])

    # ═══════════════════════════════════════
    # KPI
    # ═══════════════════════════════════════

    occupied_count = all_props.filter(availability_status__iexact='Occupied').count()
    vacant_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100), 1) if total_count > 0 else 0
    vacancy_rate = round((vacant_count / total_count * 100), 1) if total_count > 0 else 0
    total_revenue = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0

    ready_to_move_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    short_lease_count = all_props.filter(lease_duration__icontains='6').count()
    long_lease_count = all_props.filter(lease_duration__icontains='12').count()
    new_property_count = all_props.filter(property_age__icontains='New').count()
    old_property_count = all_props.exclude(property_age__icontains='New').count()
    premium_properties_count = all_props.filter(monthly_rent__gte=50000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=15000).count()

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,

        'search_query': search_query,
        'city_query': city_query,
        'state_query': state_query,          # NEW
        'furnish_query': furnish_query,
        'possession_query': possession_query,
        'prop_type_query': prop_type_query,
        'listed_by_query': listed_by_query,
        'uploaded_by_query': uploaded_by_query,   # NEW
        'budget_query': budget_query,

        'unique_cities': unique_cities,
        'unique_states': unique_states,        # NEW
        'unique_furnish': unique_furnish,
        'unique_possession': unique_possession,
        'unique_property_types': unique_property_types,
        'unique_listed_roles': unique_listed_roles,     # NEW (split)
        'unique_uploaded_roles': unique_uploaded_roles, # NEW (split)

        'filtered_count': filtered_count,

        'total_count': total_count,
        'active_count': active_count,
        'furnished_count': furnished_count,
        'available_count': available_count,
        'city_count': city_count,
        'state_count': state_count,          # NEW

        'from_date': from_date_str,
        'to_date': to_date_str,

        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,

        'avg_deposit': avg_deposit,
        'avg_area': avg_area,

        'with_owner_count': with_owner_count,
        'with_images_count': with_images_count,

        'uploaded_files': uploaded_files,

        'renting_option_labels': renting_option_labels,   # renamed from bhk_labels
        'renting_option_data': renting_option_data,

        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,

        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,

        'prop_type_labels': prop_type_labels,
        'prop_type_data': prop_type_data,

        'occupied_count': occupied_count,
        'vacant_count': vacant_count,

        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,

        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,

        'ready_to_move_count': ready_to_move_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,

        'new_property_count': new_property_count,
        'old_property_count': old_property_count,

        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
    }

    return render(
        request,
        'admin_user/Reports/Rental/rental_reports.html',
        context
    )






def rental_list(request):

    session_id = request.session.get('Admin_id')

    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ═══════════════════════════════════════
    # GET SEARCH & FILTERS
    # ═══════════════════════════════════════

    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', '').strip()
    bhk_query = request.GET.get('bhk_type', '').strip()
    state_query = request.GET.get('state', '').strip()          # NEW
    furnish_query = request.GET.get('furnishing', '').strip()
    possession_query = request.GET.get('possession', '').strip()  # maps to availability_status

    prop_type_query = request.GET.get('property_type', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()     # listed_by_role
    uploaded_by_query = request.GET.get('uploaded_by', '').strip()  # NEW -> uploaded_by_role
    budget_query = request.GET.get('budget', '').strip()

    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    listing_status_query = request.GET.get('listing_status', '').strip()
    approval_status_query = request.GET.get('approval_status', '').strip()
    duplicate_query = request.GET.get('duplicate', '').strip()

    # ═══════════════════════════════════════
    # BASE QUERYSET & PERSISTENT SR.NO MAP
    # ═══════════════════════════════════════

    base_properties = RentalResidentialProperty.objects.filter(is_deleted=False).order_by('-id')
    absolute_ordered_ids = list(base_properties.values_list('id', flat=True))

    properties = base_properties

    # ═══════════════════════════════════════
    # GLOBAL SEARCH FILTER (Across ALL Fields + Sr.No)
    # ═══════════════════════════════════════

    if search_query:
        sr_no_query = Q()

        if search_query.isdigit():
            target_index = int(search_query) - 1
            if 0 <= target_index < len(absolute_ordered_ids):
                target_id = absolute_ordered_ids[target_index]
                sr_no_query = Q(id=target_id)

        properties = properties.filter(
            sr_no_query |
            Q(id__icontains=search_query) |
            Q(property_title__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(renting_option__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(bhk_type__icontains=search_query) |
            Q(locality_area__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(pincode__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) |
            Q(listed_by_contact__icontains=search_query) |
            Q(listed_by_email__icontains=search_query) |
            Q(listed_by_role__icontains=search_query) |
            Q(availability_status__icontains=search_query) |
            Q(furnishing_status__icontains=search_query) |
            Q(available_for__icontains=search_query) |
            Q(uploaded_by_name__icontains=search_query) |
            Q(uploaded_by_email__icontains=search_query) |
            Q(uploaded_by_role__icontains=search_query) |
            Q(upload_file_name__icontains=search_query)
        )

    # ═══════════════════════════════════════
    # ADVANCED DROPDOWN FILTERS
    # ═══════════════════════════════════════

    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__iexact=city_query)

    if bhk_query and bhk_query != 'All BHK':
        properties = properties.filter(bhk_type__iexact=bhk_query)

    if state_query and state_query != 'All States':          # NEW
        properties = properties.filter(state__iexact=state_query)

    if furnish_query and furnish_query != 'All':
        properties = properties.filter(furnishing_status__iexact=furnish_query)

    if possession_query and possession_query != 'All Status':
        properties = properties.filter(availability_status__iexact=possession_query)

    if prop_type_query and prop_type_query != 'All Types':
        properties = properties.filter(property_type__iexact=prop_type_query)

    if listed_by_query and listed_by_query != 'All Roles':
        properties = properties.filter(listed_by_role__iexact=listed_by_query)

    if uploaded_by_query and uploaded_by_query != 'All Roles':   # NEW
        properties = properties.filter(uploaded_by_role__iexact=uploaded_by_query)

    if listing_status_query and listing_status_query != 'All Status':
        properties = properties.filter(listing_status__iexact=listing_status_query)

    if approval_status_query and approval_status_query != 'All Approval':
        properties = properties.filter(approval_status__iexact=approval_status_query)

    if duplicate_query == 'duplicates_only':
        properties = properties.filter(is_duplicate=True)
    elif duplicate_query == 'unique_only':
        properties = properties.filter(is_duplicate=False)

    if budget_query and budget_query != 'All Budgets':
        if budget_query == 'under_10k':
            properties = properties.filter(monthly_rent__lt=10000)
        elif budget_query == '10k_25k':
            properties = properties.filter(monthly_rent__gte=10000, monthly_rent__lte=25000)
        elif budget_query == '25k_50k':
            properties = properties.filter(monthly_rent__gte=25000, monthly_rent__lte=50000)
        elif budget_query == 'above_50k':
            properties = properties.filter(monthly_rent__gt=50000)

    # Date Filters
    if from_date_str:
        from_date = parse_date(from_date_str)
        if from_date:
            properties = properties.filter(created_at__date__gte=from_date)

    if to_date_str:
        to_date = parse_date(to_date_str)
        if to_date:
            properties = properties.filter(created_at__date__lte=to_date)

    # ═══════════════════════════════════════
    # PAGINATION & INJECT PERSISTENT SR.NO
    # ═══════════════════════════════════════
    filtered_count = properties.count()

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    id_to_srno = {pid: idx + 1 for idx, pid in enumerate(absolute_ordered_ids)}
    for prop in page_obj:
        prop.original_sr_no = id_to_srno[prop.id]

   

    # ═══════════════════════════════════════
    # EXPORT DATA (CSV / EXCEL)
    # ═══════════════════════════════════════
    if request.GET.get('download') in ['excel', 'csv']:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        from collections import OrderedDict
        import csv

        # ═══════════════════════════════════════
        # SINGLE SOURCE OF TRUTH — reuse the same field map the template
        # download and the importer already use, so LABELS stay in sync.
        # ═══════════════════════════════════════
        (field_map_sections, field_to_label_map, label_to_field_map,
         system_injected_map, helper_only_labels_map,
         decimal_fields_map, int_fields_map) = _residential_field_map()

        # ═══════════════════════════════════════
        # CANONICAL REQUIRED FIELDS — must be the literal same set used in
        # import_residential_excel's REQUIRED_FIELD_KEYS. Do NOT derive this
        # from _residential_field_map()'s per-field `required` flag — that
        # flag is used for template styling only and does not match the
        # actual mandatory inputs on the "Post Your Residential Rental" form
        # (e.g. it doesn't mark listed_by_name/email/contact as required,
        # but the Add form and the importer both do).
        # ═══════════════════════════════════════
        REQUIRED = {
            "property_type", "property_no", "bhk_type", "renting_option",
            "built_up_area", "bathrooms", "building_configuration", "total_floors",
            "furnishing_status", "available_for", "monthly_rent",
            "address", "city", "locality_area", "state", "pincode",
            "listed_by_name", "listed_by_contact", "listed_by_email", "listed_by_role",
        }

        # FIELD_LABELS starts from the field map (guarantees every shared,
        # re-importable column uses the exact label the importer expects).
        FIELD_LABELS = dict(field_to_label_map)

        # Export-only columns: appear on this report but are NOT part of the
        # import contract in _residential_field_map(). Never required, and
        # safely ignored (as unmatched headers) if this file is re-uploaded.
        FIELD_LABELS.update({

            "id": "Property ID",
            "property_title": "Property Title",
            "description": "Property Summary (Auto)",
            "rent_residential_desc": "Property Description (Auto)",
            "upload_file_name": "Uploaded File Name",
            "listing_status": "Listing Status",
            "approval_status": "Approval Status",
            "listed_elsewhere": "Listed Elsewhere?",
            "portal_name": "Portal Name",

            "property_unique_key": "Property Unique Key",
            "duplicate_count":  "Duplicate Count",
            "duplicate_group_id": "Duplicate Group ID",
            "is_duplicate": "Is Duplicate",
            "is_deleted": "Is Deleted",
            "deleted_at": "Deleted At",
            "deleted_by": "Deleted By",
            "created_at": "Created At",
        })



        sections = {


            "S.No": [
                "sr_no",
            ],
            "Listed By": [
                "listed_by_type", "listed_by_id", "listed_by_name", "listed_by_contact", "listed_by_email",
                "listed_by_role",
            ],
            "Basic Information": [
                "id", "property_title", "property_type", "property_no", "bhk_type", "renting_option",
                "built_up_area", "bathrooms", "balconies",
                "building_configuration", "total_floors", "facing_direction",
                "furnishing_status", "available_for"
            ],
            "Property Details": [
                "carpet_area", "city_zone", "ownership_type", "property_condition", "property_age",
                "wing_number", "building_name"
            ],
            "Availability Details": [
                "availability_status", "available_from", "lease_duration",
            ],
            "Pricing & Brokarage Details": [
                "monthly_rent","brokerage_percentage", "manual_brokerage", "advance_rent_month", "advance_rent_amount",
                "security_deposit_type", "security_deposit_amount",
                "maintenance_type", "monthy_maintenance_amount", "total_move_in_cost"
            ],
            "Property Location Details": [
                "address", "city", "locality_area", "property_landmark", "state",
                "pincode", "main_road_connectivity", "google_maps_link", "latitude", "longitude",
            ],
            "Amenities & Features": [
                "amenities",
            ],
            "Nearby Facilities": [
                "nearby_facilities",
            ],
            "Auto Property Summary & Description Generated Details": [
                "description",
                "rent_residential_desc",
            ],
            "Property Description(Added By User)": [
                "user_description",
            ],

            "Property Listed Elsewhere": [
                "listed_elsewhere", "portal_name",
                
            ],
            "Property Uploaded By(Auto Generated)": [
                "uploaded_by_name", "uploaded_by_email",
                "uploaded_by_contact", "uploaded_by_role", "upload_file_name",
            ],
            "Record Management": [
            "listing_status", "approval_status","property_unique_key",
            "duplicate_count", "duplicate_group_id" ,
            "is_duplicate" , 
            "is_deleted","deleted_at", "deleted_by", "created_at"
            ]
        }

        HINTS = {
            "sr_no":"Auto","id": "Property ID Auto-Generated",
            "property_title": "Auto_Generated Title", "property_type": "Apartment",
            "property_no": "e.g. B-402", "bhk_type": "e.g. 2bhk",
            "renting_option": "Full Property/Single Room/Shared Room", "built_up_area": "sq.ft",
            "bathrooms": "Number", "balconies": "Number", "building_configuration": "e.g. G+3",
            "total_floors": "Number", "facing_direction": "North/East", "furnishing_status": "Semi Furnished",
            "available_for": "Family/Bachelor", "carpet_area": "sq.ft", "city_zone": "North/South", "ownership_type": "Freehold",
            "property_condition": "Resale", "property_age": "1-3 Years",
            "wing_number": "e.g. A/B", "building_name": "Text", "availability_status": "Ready to Move",
            "available_from": "YYYY-MM-DD", "lease_duration": "11 Months", "monthly_rent": "₹","brokerage_percentage": "1%/Fixed Amount",
            "manual_brokerage": "e.g. 2.5%", "advance_rent_month": "0-11/fixed",
            "advance_rent_amount": "₹", "security_deposit_type": "0-11/fixed", "security_deposit_amount": "₹",
            "maintenance_type": "Included in Rent/Extra", "monthy_maintenance_amount": "₹",
            "total_move_in_cost": "₹", "address": "Full Address", "city": "Text", "locality_area": "Text",
            "property_landmark": "Optional", "state": "e.g. Maharashtra", "pincode": "6-digit",
            "main_road_connectivity": "Optional", "google_maps_link": "URL", "latitude": "21.1458",
            "longitude": "23.1458", "amenities": "Comma-sep",
            "nearby_facilities": "Comma-sep", "description": "Short Summary",
            "rent_residential_desc": "Long Rich Text", "user_description": "Added by user",  "listed_elsewhere": "Yes/No",
            "portal_name": "e.g. 99acres, MagicBricks",
            "listed_by_name": "Full Name", "listed_by_contact": "10 Digits", "listed_by_email": "email@example.com",
            "listed_by_role": "Owner/Agent/Admin", "listed_by_type": "Self/Other",
            "uploaded_by_name": "Admin Name", "uploaded_by_email": "Admin Email",
            "uploaded_by_contact": "Admin Contact", "uploaded_by_role": "Admin Role",
            "upload_file_name": "File Name", "listing_status": "Published/Draft", "approval_status": "Approved/Pending",
            "property_unique_key" : "Property unique id",
            "duplicate_count": "duplicate_count" ,
            "duplicate_group_id": "Duplicate Ggroup Id",
            "is_duplicate": "Yes/No", 
            "is_deleted": "Yes/No",
            "deleted_at": "YYYY-MM-DD (Auto)", "deleted_by": "Admin Name",
            "created_at": "YYYY-MM-DD"
        }

        BOOL_FIELDS = {"is_deleted", "is_duplicate"}

        ROLE_BROKERAGE_LABELS = {
            "admin": "EstateFlow Service Fee",
            "relationship manager": "Service Fee",
            "landlord": "Tenant Service Fee",
            "agent": "Brokerage",
            "agency/builder": "Service Fee",
            "builder": "Service Fee",
        }

        def brokerage_label_for(role):
            if not role:
                return "Brokerage"
            return ROLE_BROKERAGE_LABELS.get(str(role).strip().lower(), "Brokerage")

        def display_value(p, field):
            val = getattr(p, field, "")
            if field in ['available_from', 'created_at', 'deleted_at'] and val:
                try:
                    return val.strftime('%Y-%m-%d')
                except AttributeError:
                    return val
            if field in BOOL_FIELDS:
                return "Yes" if val else "No"
            # NEW: listings created directly on the site (not via bulk
            # import) never get an upload_file_name — label them clearly
            # instead of leaving the cell blank.
            if field == "upload_file_name":
                return val if val else "Web Listing UI Form"
            return val if val is not None else ""

        all_cols = []
        for sec, fields in sections.items():
            all_cols.extend([(sec, f) for f in fields])

        if request.GET.get('download') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rental Residential"

            HDR_BG, REQ_BG, OPT_BG = "667EEA", "FEF3C7", "F0FDF4"
            thin = Side(style="thin", color="CBD5E1")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

            sec_spans = OrderedDict()
            for i, (sec, _) in enumerate(all_cols):
                sec_spans.setdefault(sec, []).append(i + 1)

            for sec, cols in sec_spans.items():
                c = ws.cell(row=1, column=cols[0], value=f"📋 {sec}")
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                c.fill = PatternFill("solid", fgColor=HDR_BG)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bdr
                if len(cols) > 1:
                    ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])

            role_col = None
            brokerage_col = None

            for ci, (sec, field) in enumerate(all_cols, 1):
                req = field in REQUIRED
                label = FIELD_LABELS.get(field, field)
                lc = ws.cell(row=2, column=ci, value=label + (" *" if req else ""))
                lc.font = Font(bold=True, color="1E293B", name="Arial", size=9)
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lc.border = bdr

                hc = ws.cell(row=3, column=ci, value="")
                hc.font = Font(italic=True, color="64748B", name="Arial", size=8)
                hc.fill = PatternFill("solid", fgColor="FFFFFF")
                hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                hc.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(16, len(label) // 2 + 6)

                if field == "listed_by_role":
                    role_col = ci
                if field == "brokerage_percentage":
                    brokerage_col = ci

            if brokerage_col:
                note = (
                    "The LABEL text shown above this field on the live form changes based on\n"
                    "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
                    "Admin -> EstateFlow Service Fee\n"
                    "Relationship Manager -> Service Fee\n"
                    "Landlord -> Tenant Service Fee\n"
                    "Agent -> Brokerage\n"
                    "Agency/Builder or Builder -> Service Fee\n"
                    "Any other role -> Brokerage (default)\n\n"
                    "See the 'Brokerage Label' column at the end of this sheet for the resolved value per row."
                )
                ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 36
            ws.row_dimensions[3].height = 42
            ws.freeze_panes = "A4"

            for row_idx, p in enumerate(properties, start=4):
                for col_idx, (sec, field) in enumerate(all_cols, 1):
                    if field == "sr_no":
                        val = row_idx - 3   # row 4 -> Sr No 1
                    else:
                        val = display_value(p, field)

                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr

            if role_col:
                preview_col = len(all_cols) + 1
                pc = ws.cell(row=2, column=preview_col, value="Brokerage Label")
                pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
                pc.fill = PatternFill("solid", fgColor="FEF3C7")
                pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                pc.border = bdr
                ws.cell(row=3, column=preview_col, value="Auto-resolved from role")
                ws.cell(row=3, column=preview_col).font = Font(italic=True, color="64748B", name="Arial", size=8)
                ws.cell(row=3, column=preview_col).border = bdr
                ws.column_dimensions[get_column_letter(preview_col)].width = 22

                for row_idx, p in enumerate(properties, start=4):
                    label = brokerage_label_for(getattr(p, "listed_by_role", ""))
                    fc = ws.cell(row=row_idx, column=preview_col, value=label)
                    fc.alignment = Alignment(horizontal="center", vertical="center")
                    fc.border = bdr

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Properties_Data.xlsx"'
            wb.save(response)
            return response

        elif request.GET.get('download') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Rental_Residential_Properties_Data.csv"'
            writer = csv.writer(response)

            row1 = []
            current_sec = ""
            for sec, _ in all_cols:
                if sec != current_sec:
                    row1.append(f"📋 {sec}")
                    current_sec = sec
                else:
                    row1.append("")
            row1.append("📋 Brokerage")
            writer.writerow(row1)

            header_row = [FIELD_LABELS.get(field, field) + (" *" if field in REQUIRED else "") for _, field in all_cols]
            header_row.append("Brokerage Label")
            writer.writerow(header_row)

            hint_row = ["" for _, field in all_cols]
            hint_row.append("")
            writer.writerow(hint_row)

            for p in properties:
                data_row = [display_value(p, field) for _, field in all_cols]
                data_row.append(brokerage_label_for(getattr(p, "listed_by_role", "")))
                writer.writerow(data_row)

            return response

    # ═══════════════════════════════════════
    # STATS & UNIQUE DROPDOWN DATA
    # ═══════════════════════════════════════

    all_props = RentalResidentialProperty.objects.filter(is_deleted=False)
    total_count = all_props.count()

    unique_listing_status = all_props.exclude(listing_status__isnull=True).exclude(listing_status='').values_list('listing_status', flat=True).distinct()
    unique_approval_status = all_props.exclude(approval_status__isnull=True).exclude(approval_status='').values_list('approval_status', flat=True).distinct()
    unique_bhk = all_props.exclude(bhk_type__isnull=True).exclude(bhk_type='').values_list('bhk_type', flat=True).distinct()
    unique_cities = all_props.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_states = all_props.exclude(state__isnull=True).exclude(state='').values_list('state', flat=True).distinct()  # NEW
    unique_furnish = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values_list('furnishing_status', flat=True).distinct()
    unique_possession = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').values_list('availability_status', flat=True).distinct()
    unique_property_types = all_props.exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    unique_listed_roles = all_props.exclude(listed_by_role__isnull=True).exclude(listed_by_role='').values_list('listed_by_role', flat=True).distinct()   # NEW split
    unique_uploaded_roles = all_props.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()  # NEW split

    active_count = all_props.exclude(availability_status__isnull=True).exclude(availability_status='').count()
    furnished_count = all_props.filter(furnishing_status__iexact='Furnished').count()
    available_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    city_count = all_props.exclude(city__isnull=True).exclude(city='').values('city').distinct().count()
    state_count = all_props.exclude(state__isnull=True).exclude(state='').values('state').distinct().count()  # NEW

    # Duplicate properties
    duplicate_properties_count = all_props.filter(is_duplicate=True).count()
    unique_properties_count = all_props.filter(is_duplicate=False).count()

# Listing status breakdown
    active_listing_count = all_props.filter(listing_status__iexact='Active').count()
    inactive_listing_count = all_props.filter(listing_status__iexact='Inactive').count()
    sold_listing_count = all_props.filter(listing_status__iexact='Sold').count()
    rented_listing_count = all_props.filter(listing_status__iexact='Rented').count()

# Approval status breakdown
    pending_approval_count = all_props.filter(approval_status__iexact='Pending').count()
    approved_count = all_props.filter(approval_status__iexact='Approved').count()
    rejected_count = all_props.filter(approval_status__iexact='Rejected').count()

    # ═══════════════════════════════════════
    # RENT STATS
    # ═══════════════════════════════════════

    rent_stats = all_props.exclude(monthly_rent__isnull=True).aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
    )

    avg_rent = rent_stats['avg_rent']
    max_rent = rent_stats['max_rent']
    min_rent = rent_stats['min_rent']

    deposit_stats = all_props.exclude(security_deposit_amount__isnull=True).aggregate(avg_deposit=Avg('security_deposit_amount'))
    avg_deposit = deposit_stats['avg_deposit']

    area_stats = all_props.exclude(built_up_area__isnull=True).aggregate(avg_area=Avg('built_up_area'))
    avg_area = area_stats['avg_area']

    with_owner_count = all_props.exclude(listed_by_name__isnull=True).exclude(listed_by_name='').count()
    with_images_count = all_props.filter(images__isnull=False).distinct().count()

    uploaded_files = all_props.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    # ═══════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════

    renting_option_qs = all_props.exclude(renting_option__isnull=True).exclude(renting_option='').values('renting_option').annotate(count=Count('id')).order_by('-count')
    renting_option_labels = json.dumps([item['renting_option'] for item in renting_option_qs])
    renting_option_data = json.dumps([item['count'] for item in renting_option_qs])

    rent_buckets = [
        ('Under ₹5k', 0, 5000), ('₹5k–10k', 5000, 10000), ('₹10k–20k', 10000, 20000),
        ('₹20k–30k', 20000, 30000), ('₹30k–50k', 30000, 50000), ('₹50k–1L', 50000, 100000),
        ('Above ₹1L', 100000, 999999999),
    ]

    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data = json.dumps([all_props.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs = all_props.exclude(furnishing_status__isnull=True).exclude(furnishing_status='').values('furnishing_status').annotate(count=Count('id')).order_by('-count')
    furnishing_labels = json.dumps([item['furnishing_status'] for item in furnish_qs])
    furnishing_data = json.dumps([item['count'] for item in furnish_qs])

    prop_type_qs = all_props.exclude(property_type__isnull=True).exclude(property_type='').values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels = json.dumps([item['property_type'] for item in prop_type_qs])
    prop_type_data = json.dumps([item['count'] for item in prop_type_qs])

    # ═══════════════════════════════════════
    # KPI
    # ═══════════════════════════════════════

    occupied_count = all_props.filter(availability_status__iexact='Occupied').count()
    vacant_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    occupancy_rate = round((occupied_count / total_count * 100), 1) if total_count > 0 else 0
    vacancy_rate = round((vacant_count / total_count * 100), 1) if total_count > 0 else 0
    total_revenue = all_props.aggregate(total=Sum('monthly_rent'))['total'] or 0
    total_security_deposit = all_props.aggregate(total=Sum('security_deposit_amount'))['total'] or 0

    ready_to_move_count = all_props.filter(availability_status__iexact='Ready to Move').count()
    short_lease_count = all_props.filter(lease_duration__icontains='6').count()
    long_lease_count = all_props.filter(lease_duration__icontains='12').count()
    new_property_count = all_props.filter(property_age__icontains='New').count()
    old_property_count = all_props.exclude(property_age__icontains='New').count()
    premium_properties_count = all_props.filter(monthly_rent__gte=50000).count()
    affordable_properties_count = all_props.filter(monthly_rent__lt=15000).count()

    # ═══════════════════════════════════════
    # CONTEXT MAP
    # ═══════════════════════════════════════

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,

        'search_query': search_query,
        'bhk_query': bhk_query,
        'city_query': city_query,
        'state_query': state_query,          # NEW
        'furnish_query': furnish_query,
        'possession_query': possession_query,
        'prop_type_query': prop_type_query,
        'listed_by_query': listed_by_query,
        'uploaded_by_query': uploaded_by_query,   # NEW
        'budget_query': budget_query,

        'unique_cities': unique_cities,
        'unique_states': unique_states,        # NEW
        'unique_furnish': unique_furnish,
        'unique_possession': unique_possession,
        'unique_property_types': unique_property_types,
        'unique_listed_roles': unique_listed_roles,     # NEW (split)
        'unique_uploaded_roles': unique_uploaded_roles, # NEW (split)

        'filtered_count': filtered_count,

        'total_count': total_count,
        'active_count': active_count,
        'furnished_count': furnished_count,
        'available_count': available_count,
        'city_count': city_count,
        'state_count': state_count,          # NEW

        'from_date': from_date_str,
        'to_date': to_date_str,

        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,

        'avg_deposit': avg_deposit,
        'avg_area': avg_area,

        'with_owner_count': with_owner_count,
        'with_images_count': with_images_count,

        'uploaded_files': uploaded_files,

        'renting_option_labels': renting_option_labels,   # renamed from bhk_labels
        'renting_option_data': renting_option_data,

        'rent_range_labels': rent_range_labels,
        'rent_range_data': rent_range_data,

        'furnishing_labels': furnishing_labels,
        'furnishing_data': furnishing_data,

        'prop_type_labels': prop_type_labels,
        'prop_type_data': prop_type_data,

        'occupied_count': occupied_count,
        'vacant_count': vacant_count,

        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,

        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,

        'ready_to_move_count': ready_to_move_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,

        'new_property_count': new_property_count,
        'old_property_count': old_property_count,

        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'bhk_query': bhk_query,
        'unique_bhk': unique_bhk,                              # ADD

        'listing_status_query': listing_status_query,          # ADD
        'approval_status_query': approval_status_query,        # ADD
        'duplicate_query': duplicate_query,                    # ADD

        'unique_listing_status': unique_listing_status,        # ADD
        'unique_approval_status': unique_approval_status,      # ADD

        'duplicate_properties_count': duplicate_properties_count,  # ADD
        'unique_properties_count': unique_properties_count,        # ADD

        'active_listing_count': active_listing_count,          # ADD
        'inactive_listing_count': inactive_listing_count,      # ADD
        'sold_listing_count': sold_listing_count,               # ADD
        'rented_listing_count': rented_listing_count,           # ADD
        'pending_approval_count': pending_approval_count,       # ADD
        'approved_count': approved_count,                       # ADD
        'rejected_count': rejected_count,                       # ADD
    }

    return render(
        request,
        'admin_user/Reports/Rental/rental_list.html',
        context
    )






def rental_bulk_delete(request):
    """Handles Advanced Bulk Deletions for Rental Properties (Soft Delete)."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
        
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')
    
    if not admin_id and not user_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'})

    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        
        # 👈 1. Get the deleter's name
        deleter_name = _get_deleter_name(request)
        
        # IMPORTANT: Only target properties that are NOT currently in the Recycle Bin
        properties = RentalResidentialProperty.objects.filter(is_deleted=False)
        
        if delete_type == 'delete_all':
            count = properties.count()
            # 👈 2. Add deleted_by=deleter_name to the update query
            properties.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved ALL {count} properties to Recycle Bin.'})
            
        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(id__in=page_ids)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from current page to Recycle Bin.'})
            
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            target_props = properties.filter(available_from__range=[from_date, to_date])
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties in date range to Recycle Bin.'})
            
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(available_from__gte=thirty_days_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from the last 30 days to Recycle Bin.'})
            
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(available_from__lt=six_months_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} older properties to Recycle Bin.'})
            
        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '')
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) | 
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties uploaded by {uploader} to Recycle Bin.'})

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            target_props = properties.filter(upload_file_name=file_name) 
            count = target_props.count()
            if count == 0:
                return JsonResponse({'status': 'error', 'message': f'No properties found for file: {file_name}'})
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from {file_name} to Recycle Bin.'})
            
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})





@require_POST
def rental_residential_delete(request, pk):
    """Soft delete a single property."""
    try:
        session_id = request.session.get('Admin_id')
        user_identity = "Unknown Admin"
        if session_id:
            try:
                admin_obj = Admin_Login.objects.get(id=session_id)
                user_identity = admin_obj.email or admin_obj.username
            except Admin_Login.DoesNotExist:
                pass

        prop = RentalResidentialProperty.objects.get(id=pk)

        # Snapshot file source metadata before deletion mapping executions
        associated_origin_file = prop.upload_file_name or "Web UI Form"
        property_title_ref = prop.property_title or "—"

        prop.is_deleted = True
        prop.deleted_at = timezone.now()
        if hasattr(prop, 'deleted_by'):
            prop.deleted_by = user_identity

        prop.save()

        # =====================================================
        # AUDIT LOGIC: Operational Deletion Trace Serialization
        # =====================================================
        RentalActivityLog.objects.create(
            user_identity=user_identity,
            user_role="Admin",
            action_type='DELETE',
            property_id=pk,
            targeted_fields="Entire Record Purged",
            associated_file=associated_origin_file,
            action_payload=json.dumps({
                "deleted_property_id": pk,
                "property_title": property_title_ref,
                "action": "soft_delete_to_recycle_bin"
            }),
            ip_address=_get_client_ip(request),
            status='SUCCESS'
        )

        return JsonResponse({
            'status': 'success',
            'message': 'Moved to Recycle Bin successfully!'
        })

    except RentalResidentialProperty.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Property not found.'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })





def calculate_retention(queryset):
    now = timezone.now()
    items = list(queryset)

    for item in items:
        if item.deleted_at:
            deleted_at = item.deleted_at
            # Safeguard against offset-naive vs offset-aware datetime comparison errors
            if timezone.is_naive(deleted_at):
                deleted_at = timezone.make_aware(deleted_at)
            expiry_date = deleted_at + timedelta(days=30)
            item.days_left = max((expiry_date - now).days, 0)
        else:
            item.days_left = 30
    return items

from django.db import transaction

# ── UNIFIED GLOBAL RECYCLE BIN VIEW ──────────────────────────────────────────
def global_recycle_bin(request):
    """Unified Recycle Bin displaying deleted items from all property modules."""
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    rental_deleted = calculate_retention(RentalResidentialProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    commercial_deleted = calculate_retention(CommercialRentalProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    pg_deleted = calculate_retention(PGColivingProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    resale_deleted = calculate_retention(ResaleResidentialProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    commercial_resale_deleted = calculate_retention(CommercialResaleProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    plot_sale_deleted = calculate_retention(PlotSaleProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    industrial_resale_deleted = calculate_retention(IndustrialResaleProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))
    agricultural_resale_deleted = calculate_retention(AgriculturalResaleProperty.objects.filter(is_deleted=True).order_by('-deleted_at'))

    context = {
        'rental_deleted': rental_deleted, 'rental_count': len(rental_deleted),
        'commercial_deleted': commercial_deleted, 'commercial_count': len(commercial_deleted),
        'pg_deleted': pg_deleted, 'pg_count': len(pg_deleted),
        'resale_deleted': resale_deleted, 'resale_count': len(resale_deleted),
        'commercial_resale_deleted': commercial_resale_deleted, 'commercial_resale_count': len(commercial_resale_deleted),
        'plot_sale_deleted': plot_sale_deleted, 'plot_sale_count': len(plot_sale_deleted),
        'industrial_resale_deleted': industrial_resale_deleted, 'industrial_resale_count': len(industrial_resale_deleted),
        'agricultural_resale_deleted': agricultural_resale_deleted, 'agricultural_resale_count': len(agricultural_resale_deleted),
        'total_deleted_all': (
            len(rental_deleted) + len(commercial_deleted) + len(pg_deleted) +
            len(resale_deleted) + len(commercial_resale_deleted) + len(plot_sale_deleted) +
            len(industrial_resale_deleted) + len(agricultural_resale_deleted)
        )
    }
    return render(request, 'admin_user/Reports/Rental/global_recycle_bin.html', context)

def system_audit_logs(request):
    """A completely separate view for tracking Deletion and Restore Audit Logs."""
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)
    deletion_logs = []
    restore_logs = []

    models_map = {
        'Rental Residential': RentalResidentialProperty,
        'Commercial Rental': CommercialRentalProperty,
        'PG / Co-living': PGColivingProperty,
        'Resale Residential': ResaleResidentialProperty,
        'Commercial Resale': CommercialResaleProperty,
        'Plot/Land': PlotSaleProperty,
        'Industrial Resale': IndustrialResaleProperty,
        'Agricultural Resale': AgriculturalResaleProperty,
    }

    def get_display_id(obj):
        # Look for custom model unique IDs before falling back to PK
        id_fields = [
            'rental_residential_id', 'commercial_rental_id', 'pg_property_id', 
            'property_id', 'commercial_id', 'plot_property_id', 
            'industrial_id', 'agri_property_id'
        ]
        for field in id_fields:
            if hasattr(obj, field) and getattr(obj, field):
                return getattr(obj, field)
        return obj.pk

    def get_display_title(obj):
        if hasattr(obj, 'property_title') and obj.property_title:
            return obj.property_title
        elif hasattr(obj, 'title') and obj.title:
            return obj.title
        elif hasattr(obj, 'plot_title') and obj.plot_title:
            return obj.plot_title
        return 'N/A'

    for module_name, model in models_map.items():
        # 1. Fetch Deletion Logs
        try:
            deleted_items = model.objects.filter(is_deleted=True).exclude(deleted_at__isnull=True)
            for p in deleted_items:
                deletion_logs.append({
                    'module': module_name,
                    'id': get_display_id(p),
                    'title': get_display_title(p),
                    'by': getattr(p, 'deleted_by', 'System Admin') or 'System Admin',
                    'date': p.deleted_at
                })
        except Exception:
            pass

        # 2. Fetch Restore Logs (if restored_at is later implemented in models)
        try:
            restored_items = model.objects.filter(restored_at__isnull=False)
            for p in restored_items:
                restore_logs.append({
                    'module': module_name,
                    'id': get_display_id(p),
                    'title': get_display_title(p),
                    'by': getattr(p, 'restored_by', 'System Admin') or 'System Admin',
                    'date': p.restored_at
                })
        except Exception:
            pass

    # Sort both lists by date (Newest logs first)
    deletion_logs.sort(key=lambda x: x['date'] if x['date'] else timezone.now(), reverse=True)
    restore_logs.sort(key=lambda x: x['date'] if x['date'] else timezone.now(), reverse=True)

    context = {
        'admin_obj': admin_obj,
        'deletion_logs': deletion_logs,
        'restore_logs': restore_logs,
        'deletion_count': len(deletion_logs),
        'restore_count': len(restore_logs),
        'total_logs': len(deletion_logs) + len(restore_logs)
    }

    return render(request, 'admin_user/Reports/Rental/audit_logs.html', context)





@csrf_exempt
@require_POST
def bulk_restore_route(request, property_type):
    """
    Unified bulk-restore engine. Reverses soft-deletion status across 
    all 8 real estate modules using efficient database batch updates.
    """
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

    try:
        # Complete dictionary mapping for all active property models
        model_map = {
            'rental-residential': RentalResidentialProperty,
            'commercial': CommercialRentalProperty,
            'pg-coliving': PGColivingProperty,
            'resale-residential': ResaleResidentialProperty,
            'commercial-resale': CommercialResaleProperty,
            'plot-sale': PlotSaleProperty,
            'industrial-resale': IndustrialResaleProperty,
            'agricultural-resale': AgriculturalResaleProperty,
        }

        # Guard against invalid module requests
        if property_type not in model_map:
            return JsonResponse({
                'status': 'error', 
                'message': f'Invalid property module selector: {property_type}'
            }, status=400)

        TargetModel = model_map[property_type]
        
        # Grab only the records currently marked as soft-deleted
        deleted_queryset = TargetModel.objects.filter(is_deleted=True)
        count = deleted_queryset.count()

        if count == 0:
            readable_name = property_type.replace("-", " ").title()
            return JsonResponse({
                'status': 'success', 
                'message': f'The bin for {readable_name} has no items to restore.'
            })

        # Perform a high-speed database batch update
        deleted_queryset.update(
            is_deleted=False,
            deleted_at=None,
            deleted_by=None
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully restored {count} properties back to active inventory status.'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': f'Database operational failure during restoration: {str(e)}'
        }, status=500)


@csrf_exempt
@require_POST
def bulk_hard_delete_properties(request, property_type):

    """
    Unified hard-delete engine. Permanently deletes soft-deleted records 
    across all 8 real estate modules and purges attached media files from disk.
    """
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

    try:
        # 1. Map endpoint slugs to their respective Models
        model_map = {
            'rental-residential': RentalResidentialProperty,
            'commercial': CommercialRentalProperty,
            'pg-coliving': PGColivingProperty,
            'resale-residential': ResaleResidentialProperty,
            'commercial-resale': CommercialResaleProperty,
            'plot-sale': PlotSaleProperty,
            'industrial-resale': IndustrialResaleProperty,   
            'agricultural-resale': AgriculturalResaleProperty,
        }

        if property_type not in model_map:
            return JsonResponse({'status': 'error', 'message': f'Invalid property module context: {property_type}'}, status=400)

        TargetModel = model_map[property_type]
        
        # 2. Grab all items waiting in the recycle bin for this specific model
        deleted_queryset = TargetModel.objects.filter(is_deleted=True)
        count = deleted_queryset.count()

        if count == 0:
            return JsonResponse({
                'status': 'success', 
                'message': f'The bin for {property_type.replace("-", " ").title()} is already completely empty.'
            })

        # 3. Media Cleanup Loop (Prevents unlinked/orphaned images/videos from filling up the disk)
        for property_obj in deleted_queryset:
            # Drop related images if the model utilizes a related images manager
            if hasattr(property_obj, 'images'):
                try:
                    for img in property_obj.images.all():
                        if img.image:
                            img.image.delete(save=False)
                except Exception:
                    pass  # Fail gracefully if relationship structure differs

            # Drop standalone video files if present
            if hasattr(property_obj, 'video') and property_obj.video:
                try:
                    property_obj.video.delete(save=False)
                except Exception:
                    pass

            # Drop standalone image files if present (e.g., plot layouts, agri maps)
            if hasattr(property_obj, 'image') and property_obj.image:
                try:
                    property_obj.image.delete(save=False)
                except Exception:
                    pass

        # 4. Wipe rows from the database permanently
        deleted_queryset.delete()

        return JsonResponse({
            'status': 'success',
            'message': f'Permanently purged {count} records and all associated assets from the database.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Database operational failure: {str(e)}'}, status=500)




@require_POST
def rental_restore(request, pk):
    """Restore property from recycle bin."""

    try:
        prop = RentalResidentialProperty.objects.get(
            id=pk,
            is_deleted=True
        )

        prop.is_deleted = False
        prop.deleted_at = None

        if hasattr(prop, 'deleted_by'):
            prop.deleted_by = None

        prop.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Property successfully restored!'
        })

    except RentalResidentialProperty.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Property not found.'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })


@require_POST
def rental_hard_delete(request, pk):
    print("PK RECEIVED:", pk)

    try:
        prop = RentalResidentialProperty.objects.get(
            id=pk,
            is_deleted=True
        )

        print("FOUND PROPERTY:", prop.id)

        prop.delete()

        return JsonResponse({
            'status': 'success',
            'message': 'Property permanently deleted.'
        })

    except RentalResidentialProperty.DoesNotExist:
        print("PROPERTY NOT FOUND:", pk)

        return JsonResponse({
            'status': 'error',
            'message': 'Property not found.'
        })
# ─────────────────────────────────────────────
#  Helper converters
# ─────────────────────────────────────────────

def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _int(val):
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None


def _decimal(val):
    try:
        cleaned = str(val).replace(",", "").replace("₹", "").strip()
        return Decimal(cleaned)
    except Exception:
        return None


def _bigint(val):
    try:
        cleaned = str(val).replace(",", "").replace("₹", "").strip()
        return int(float(cleaned))
    except (TypeError, ValueError):
        return None


def _date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────
#  Column → model field mapping
# ─────────────────────────────────────────────

def _email(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or '@' not in s:
        return None
    return s



def _load_new_excel(excel_file):
    """
    Returns (ws, headers_dict) where headers_dict maps clean_col_name → col_index (1-based).
    Caller iterates ws.iter_rows(min_row=4, values_only=True) for data.
    """
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active
 
    headers = {}
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    for col_idx, cell_val in enumerate(header_row, 1):
        if cell_val:
            key = str(cell_val).strip().rstrip(' *').strip().lower()
            headers[key] = col_idx
    return wb, ws, headers
 
 
def _get_cell(row, col_idx):
    """Safe cell fetch from a values_only row tuple."""
    if col_idx and col_idx - 1 < len(row):
        return row[col_idx - 1]
    return None
 
 
# =============================================================================
# 1. RENTAL RESIDENTIAL
# =============================================================================
 
RENTAL_RES_COLUMN_MAP = [
    # ==========================================
    # BASIC INFO
    # ==========================================
    
   
    ("property_type",        "property_type",        _str),
    ("bhk_type",             "bhk_type",             _str),
    ("renting_option",       "renting_option",       _str),
    ("built_up_area",        "built_up_area",        _decimal),
    ("bathrooms",            "bathrooms",            _int),
    ("balconies",            "balconies",            _int),
    ("floor_number",         "floor_number",         _str),
    ("total_floors",         "total_floors",         _int),
    ("facing",               "facing",               _str),
    ("furnishing_status",    "furnishing_status",    _str),
    ("available_for",        "available_for",        _str),

    # ==========================================
    # PROPERTY DETAILS
    # ==========================================
    ("zone",                 "zone",                 _str),
    ("ownership_type",       "ownership_type",       _str),
    ("construction_status",  "construction_status",  _str),
    ("property_age",         "property_age",         _str),
    ("carpet_area",          "carpet_area",          _decimal),
    ("plot_area",            "plot_area",            _decimal),
    ("building_name",        "building_name",        _str),

    # ==========================================
    # AVAILABILITY DETAILS
    # ==========================================
    ("possession_status",    "possession_status",    _str),
    ("available_from",       "available_from",       _date),
    ("lease_duration",       "lease_duration",       _str),
    ("brokerage",            "brokerage",            _str),
    ("brokerage_percentage", "brokerage_percentage", _str),
    ("manual_brokerage",     "manual_brokerage",     _str),

    # ==========================================
    # PRICING DETAILS
    # ==========================================
    ("monthly_rent",         "monthly_rent",         _bigint),
    ("security_deposit",     "security_deposit",     _bigint),
    ("maintenance_type",     "maintenance_type",     _str),
    ("maintenance_amount",   "maintenance_amount",   _bigint),

    # ==========================================
    # LOCATION DETAILS
    # ==========================================
    ("address",              "address",              _str),
    ("city",                 "city",                 _str),
    ("locality",             "locality",             _str),
    ("state",                "state",                _str),
    ("pincode",              "pincode",              _str),
    ("road_connectivity",    "road_connectivity",    _str),

    # ==========================================
    # AMENITIES & DESCRIPTION
    # ==========================================
    ("amenities",            "amenities",            _str),
    ("facilities",           "facilities",           _str),
    ("description",          "description",          _str),
    ("rent_residential_desc","rent_residential_desc",_str),

    # ==========================================
    # OWNER DETAILS
    # ==========================================
    ("owner_name",           "owner_name",           _str),
    ("contact_number",       "contact_number",       _str),
    ("email",                "email",                _email),
    ("alternate_contact",    "alternate_contact",    _str),

    # ==========================================
    # SYSTEM / TRACKING (Optional for Excel Upload)
    # ==========================================
    ("uploaded_by_name",     "uploaded_by_name",     _str),
    ("uploaded_by_email",    "uploaded_by_email",    _str),
    ("uploaded_by_contact",  "uploaded_by_contact",  _str),
    ("uploaded_by_role",     "uploaded_by_role",     _str),
    ("created_at",           "created_at",           _date), 
]













from collections import OrderedDict









# =====================================================================
# NOTE ON THE BUG YOU HIT:
# "'list' object has no attribute 'items'" happens when something else
# in your views.py ALSO defines a module-level name called SECTIONS (or
# HINTS / SAMPLE / etc) — very likely another Excel template generator
# for Commercial/Plot/PG listings. Python module-level names are shared
# across the whole file, so the last one loaded silently overwrites this
# one. Fix: everything below is now built INSIDE a function
# (_residential_field_map), not at module level, so it can never collide
# with anything else in your file, no matter how many other generators
# reuse the same names.
# =====================================================================



def _residential_field_map():
    """Returns (sections, field_to_label, label_to_field, system_injected,
    helper_only_labels, decimal_fields, int_fields). Called fresh by both
    the download view and the import view — always local, never global."""
 
    sections = OrderedDict([
        ("Listed By", [
            ("listed_by_type",   "Listed By Type (Self/Other)", False),
            ("listed_by_id",     "Listed By Id", False),
            ("listed_by_name",   "Listed By Name", False),
            ("listed_by_email",  "Listed By Email", False),
            ("listed_by_contact","Listed By Contact", False),
            ("listed_by_role",   "Listed By Role", True),   # drives the brokerage label
        ]),
        ("Basic Information", [
            ("property_type",           "Property Type", True),
            ("property_no",             "Flat/House/Unit No", True),
            ("bhk_type",                "BHK Type", True),
            ("renting_option",          "Renting Option", True),
            ("built_up_area",           "Built-up Area (sq.ft)", True),
            ("bathrooms",                "Bathrooms", True),
            ("balconies",                "Balconies", False),
            ("building_configuration",  "Building Configuration", True),
            ("total_floors",             "Total Floors Constructed", True),
            ("facing_direction",        "Facing Direction", False),
            ("furnishing_status",       "Furnishing Status", True),
            ("available_for",           "Available For", True),
        ]),
        ("Property Details", [
            ("carpet_area",     "Carpet Area (sq.ft)", False),
            ("city_zone",       "City Zone", False),
            ("ownership_type",  "Ownership Type", False),
            ("property_condition", "Property Condition", False),
            ("property_age",    "Property Age (Years)", False),
            ("wing_number",     "Tower/Wing Number", False),
            ("building_name",   "Building/Society Name", False),
        ]),
        ("Availability Details", [
            ("availability_status",   "Availability Status", False),
            ("available_from",        "Available From", False),
            ("lease_duration",        "Lease Duration", False),
           
        ]),
        ("Pricing & Brokarage Details", [
            
            ("monthly_rent",              "Monthly Rent", True),
            ("brokerage_percentage",  "Brokerage / Service Fee", True),
            ("manual_brokerage",      "Fixed Brokerage Amount", False),
            ("advance_rent_month",        "Advance Rent Month", False),
            ("advance_rent_amount",       "Advance Rent Amount", False),
            ("security_deposit_type",     "Refundable Security Deposit", True),
            ("security_deposit_amount",   "Refundable Security Deposit Amount", False),
            ("maintenance_type",          "Maintenance Type", False),
            ("monthy_maintenance_amount", "Monthly Maintenance Amount", False),
            ("total_move_in_cost",        "Total Move In Cost", False),
        ]),
        ("Property Location Details", [
            ("address",                 "Address", True),
            ("city",                    "City", True),
            ("locality_area",           "Locality/Area", True),
            ("property_landmark",       "Property Landmark", False),
            ("state",                   "State", True),
            ("pincode",                 "Pincode", True),
            ("main_road_connectivity",  "Main Road Connectivity", False),
            # --- newly added: present on the form, were missing from the map ---
            ("google_maps_link",        "Google Maps Link", False),
            ("latitude",                 "Latitude", False),
            ("longitude",                "Longitude", False),
        ]),
        ("Amenities & Features", [
            ("amenities",          "Amenities (comma-separated)", False),
           
        ]),
        ("Nearby Facilities", [
           
            ("nearby_facilities",  "Nearby Facilities (comma-separated)", False),
        ]),
        ("Property Descriptions", [
            ("user_description", "Property Description", False),
        ]),
        ("Media & Listing Status", [
            ("listed_elsewhere", "Listed Elsewhere (Yes/No)", False),
            ("portal_name",      "Portal Name", False),
        ]),
        ("Property Uploaded By(Auto Generated)", [
            ("uploaded_by_name",    "Uploaded By Name (Auto)", False),
            ("uploaded_by_email",   "Uploaded By Email (Auto)", False),
            ("uploaded_by_contact", "Uploaded By Contact (Auto)", False),
            ("uploaded_by_role",    "Uploaded By Role (Auto)", False),
            ("created_at",          "Created At (Auto)", False),
        ]),
    ])
 
    field_to_label = {f: lbl for _, fields in sections.items() for f, lbl, _ in fields}
    label_to_field = {lbl.strip().lower(): f for _, fields in sections.items() for f, lbl, _ in fields}
 
    system_injected = {
        "uploaded_by_name", "uploaded_by_email", "uploaded_by_contact",
        "uploaded_by_role", "created_at",
    }
    helper_only_labels = {"brokerage label preview (auto)"}
    decimal_fields = {"built_up_area", "carpet_area"}
    int_fields = {
        "bathrooms", "balconies", "total_floors", "monthly_rent",
        "advance_rent_amount", "security_deposit_amount",
        "monthy_maintenance_amount", "total_move_in_cost",
    }
 
    return sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields
 
 
def _sample_row_data():
    """One complete example value per column so every column in the
    downloaded template shows the expected format — nothing left blank."""
    return {
        "listed_by_type": "self", "listed_by_role": "Relationship Manager",
        "listed_by_id": "rm0943", "listed_by_name": "Vikas", "listed_by_email": "vikas@test.com",
        "listed_by_contact": "9876543210",
 
        "property_type": "Apartment", "property_no": "B-402", "bhk_type": "2bhk" , "renting_option": "Full Property",
        "built_up_area": "1200", "bathrooms": "2", "balconies": "1",
        "building_configuration": "G+3", "total_floors": "4", "facing_direction": "East",
        "furnishing_status": "Semi Furnished", "available_for": "Family",
 
        "carpet_area": "950", "city_zone": "West Zone", "ownership_type": "Freehold",
        "property_condition": "Well Maintained", "property_age": "5", "wing_number": "A Wing",
        "building_name": "Green Valley Apartments",
 
        "availability_status": "Available Immediately", "available_from": "2026-07-15",
        "lease_duration": "11 Months",   # only used when brokerage_percentage = "Fixed Amount"
 
        "monthly_rent": "25000","brokerage_percentage": "1 Month Rent",
        "manual_brokerage": "", "advance_rent_month": "1", "advance_rent_amount": "",
        "security_deposit_type": "2", "security_deposit_amount": "",
        "maintenance_type": "Included in Rent", "monthy_maintenance_amount": "",
        "total_move_in_cost": "75000",
 
        "address": "Flat 402, Green Valley", "city": "Nagpur", "locality_area": "Dharampeth",
        "property_landmark": "Near Railway Station", "state": "Maharashtra", "pincode": "440010",
        "main_road_connectivity": "Within 500 Meters",
        "google_maps_link": "https://maps.google.com/?q=21.1458,79.0882",
        "latitude": "21.1458", "longitude": "79.0882",
 
        "amenities": "Wi-Fi, AC, Lift", "nearby_facilities": "Metro, Hospital",
 
        "user_description": "Spacious and well-lit apartment close to all amenities.",
 
        "listed_elsewhere": "No", "portal_name": "",
    }

# =====================================================================
# DOWNLOAD TEMPLATE
# =====================================================================

def download_residential_template(request):
    """Download the upload template — column headers are the same
    human-readable labels used on the actual form, not raw field names.
    Includes a live 'brokerage label preview' formula so staff can see
    the label change instantly when they type a different role.

    Row 4 (sample data) is LOCKED — visible for reference only, cannot
    be edited or deleted. Rows 5+ are unlocked for actual data entry.
    """

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()
    sample = _sample_row_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental Residential"

    HDR_BG, REQ_BG, OPT_BG, SAMP_BG = "667EEA", "FEF3C7", "F0FDF4", "ECFDF5"
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    col = 1
    role_col = None
    brokerage_col = None

    for section, fields in sections.items():
        start_col = col
        for field, label, required in fields:
            header_text = label + (" *" if required else "")
            c1 = ws.cell(row=2, column=col, value=header_text)
            c1.font = Font(bold=True, color="1E293B", name="Arial", size=9)
            c1.fill = PatternFill("solid", fgColor=REQ_BG if required else OPT_BG)
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = bdr

            sc = ws.cell(row=4, column=col, value=sample.get(field, ""))
            sc.font = Font(name="Arial", size=9, color="065F46")
            sc.fill = PatternFill("solid", fgColor=SAMP_BG)
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sc.border = bdr
            sc.protection = Protection(locked=True)  # <-- NEW: sample stays read-only

            ws.column_dimensions[get_column_letter(col)].width = max(18, len(label) // 2 + 6)

            if field == "listed_by_role":
                role_col = col
            if field == "brokerage_percentage":
                brokerage_col = col
            col += 1

        end_col = col - 1
        hc = ws.cell(row=1, column=start_col, value=f"\U0001F4CB {section}")
        hc.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hc.fill = PatternFill("solid", fgColor=HDR_BG)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = PatternFill("solid", fgColor=HDR_BG)
            ws.cell(row=1, column=cc).border = bdr
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    # ---- NEW: unlock data-entry rows (5+) so users can fill them in ----
    total_cols = col - 1
    unlocked = Protection(locked=False)
    MAX_DATA_ROWS = 500  # adjust if you expect more than 500 rows of data
    for r in range(5, 5 + MAX_DATA_ROWS):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).protection = unlocked

    # ---- live brokerage label preview column, appended at the end ----
    preview_col = col
    pc = ws.cell(row=2, column=preview_col, value="Brokerage Label Preview (auto)")
    pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
    pc.fill = PatternFill("solid", fgColor="FEF3C7")
    pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pc.border = bdr
    ws.column_dimensions[get_column_letter(preview_col)].width = 26

    role_letter = get_column_letter(role_col)
    formula = (
        f"=IFERROR(INDEX('Notes - Brokerage Label'!$B$4:$B$9,"
        f"MATCH(LOWER(TRIM({role_letter}4)),'Notes - Brokerage Label'!$C$4:$C$9,0)),\"Brokerage\")"
    )
    fcell = ws.cell(row=4, column=preview_col, value=formula)
    fcell.fill = PatternFill("solid", fgColor="FEF3C7")
    fcell.font = Font(bold=True, color="92400E", name="Arial", size=9)
    fcell.alignment = Alignment(horizontal="center", vertical="center")
    fcell.protection = Protection(locked=True)  # <-- NEW: preview formula stays read-only too

    if brokerage_col:
        note = (
            "The LABEL text shown above this field on the live form changes based on\n"
            "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
            "Admin -> EstateFlow Service Fee\n"
            "Relationship Manager -> Service Fee\n"
            "Landlord -> Tenant Service Fee\n"
            "Agent -> Brokerage\n"
            "Agency/Builder or Builder -> Service Fee\n"
            "Any other role -> Brokerage (default)\n\n"
            "See 'Notes - Brokerage Label' sheet, and the live preview column at the end of this sheet."
        )
        ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    # ---- Notes sheet (lookup table the formula above reads from) ----
    notes = wb.create_sheet("Notes - Brokerage Label")
    notes.column_dimensions['A'].width = 26
    notes.column_dimensions['B'].width = 26
    notes.column_dimensions['C'].width = 4
    notes.sheet_view.showGridLines = False

    t = notes["A1"]
    notes.merge_cells("A1:B1")
    t.value = "Brokerage label — driven by Listed By Role"
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["Listed By Role", "Label shown on form"]
    for i, h in enumerate(hdrs, start=1):
        c = notes.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="1E293B", name="Arial")
        c.fill = PatternFill("solid", fgColor=OPT_BG)
        c.border = bdr

    role_rows = [
        ("Admin", "EstateFlow Service Fee"),
        ("Relationship Manager", "Service Fee"),
        ("Landlord", "Tenant Service Fee"),
        ("Agent", "Brokerage"),
        ("Agency/Builder", "Service Fee"),
        ("Builder", "Service Fee"),
    ]
    for r, (role, label) in enumerate(role_rows, start=4):
        notes.cell(row=r, column=1, value=role).border = bdr
        notes.cell(row=r, column=2, value=label).border = bdr
        notes.cell(row=r, column=3, value=f"=LOWER(TRIM(A{r}))")

    # ---- NEW: lock the whole sheet, keep only rows 5+ editable ----
    ws.protection.sheet = True
    ws.protection.formatColumns = True   # allow resizing columns
    ws.protection.formatRows = True      # allow resizing rows
    ws.protection.insertRows = False     # block inserting rows (protects layout)
    ws.protection.deleteRows = False     # block deleting rows (protects the sample row)
    ws.protection.autoFilter = False
    ws.protection.sort = False

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Rental_Residential_Template.xlsx"'
    wb.save(response)
    return response

# =====================================================================
# IMPORT
# =====================================================================




# from .helpers import _residential_field_map, _get_client_ip  # <-- adjust import to wherever you kept it




def _normalize_label(raw):
    if raw is None:
        return ""
    text = str(raw).replace("\u00a0", " ")           
    text = text.replace(" *", "").strip()
    text = re.sub(r"\s+", " ", text)                  
    return text.lower()


def _find_header_row(ws, label_to_field, max_scan_rows=6):
    best_row, best_score = None, -1
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        score = sum(1 for v in row_vals if _normalize_label(v) in label_to_field)
        if score > best_score:
            best_row, best_score = r, score
    return best_row, best_score



@csrf_exempt
@require_POST
def import_residential_excel(request):
    excel_file = request.FILES.get("rental_file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx files allowed."}, status=400)

    sections, field_to_label, label_to_field, system_injected, helper_only_labels, decimal_fields, int_fields = _residential_field_map()

    # -------------------------------------------------------------------
    # REQUIRED FIELDS - must mirror the `required` inputs in the Add form.
    # (Uses field_to_label from _residential_field_map() so the error
    # messages show the same column names as your Excel template.)
    # -------------------------------------------------------------------
    REQUIRED_FIELD_KEYS = [
    'property_type',
    'property_no',
    'bhk_type',
    'renting_option',
    'built_up_area',
    'bathrooms',
    'building_configuration',
    'total_floors',
    'furnishing_status',
    'available_for',
    'monthly_rent',
    
    'address',
    'city',
    'locality_area',
    'state',
    'pincode',
   
    'listed_by_id',
    'listed_by_name',
    'listed_by_email',
    'listed_by_contact',
    'listed_by_role',
]

    def _field_label(field):
        return field_to_label.get(field) or field.replace('_', ' ').title()

    def _is_missing(val):
        """Treat None / empty-string as missing. Does NOT treat 0 / '0' as missing,
        so numeric fields like bathrooms=0 aren't incorrectly flagged."""
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    # ---- 1. Uploader Identity ----
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    admin_obj = Admin_Login.objects.filter(id=admin_id).first() if admin_id else None
    user_obj = User_Details.objects.filter(id=user_id).first() if user_id else None

    uploader_name = uploader_email = uploader_contact = ""
    uploader_role = "Automated Engine"
    user_identity = "Automated Engine"

    if admin_obj:
        uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
        uploader_email = getattr(admin_obj, 'email', '')
        uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
        uploader_role = "Admin"
        user_identity = uploader_email or uploader_name
    elif user_obj:
        uploader_name = user_obj.user_name
        uploader_email = user_obj.user_email
        uploader_contact = user_obj.user_phone
        uploader_role = "User"
        user_identity = uploader_email or uploader_name

    # ---- 2. Parse Excel ----
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Rental Residential"] if "Rental Residential" in wb.sheetnames else wb.active
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    header_row, matched_count = _find_header_row(ws, label_to_field)

    if matched_count == 0:
        return JsonResponse({
            "status": "error",
            "message": (
                "No recognizable column headers were found in this file. "
                "Please use the official template downloaded from "
                "'Download Template' — don't rename or re-order the header row."
            ),
        }, status=400)

    raw_headers = [cell.value for cell in ws[header_row]]
    field_headers = []
    unmatched_headers = []

    for h in raw_headers:
        norm = _normalize_label(h)
        if not norm:
            field_headers.append(None)
            continue
        if norm in helper_only_labels:
            field_headers.append(None)
            continue
        field = label_to_field.get(norm)
        field_headers.append(field)
        if field is None:
            unmatched_headers.append(str(h))

    data_start_row = header_row + 1

    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []       
    listed_by_mismatch_errors = []   
    skipped_listed_by_mismatch = 0   

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        obj_data = {}
        for col_idx, field in enumerate(field_headers):
            if not field or field in system_injected:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip() != "":
                obj_data[field] = val

        if not obj_data:
            skipped_empty_after_mapping += 1
            continue

        # ---- Type Coercion (unchanged) ----
        if 'available_from' in obj_data:
            d_val = obj_data['available_from']
            if isinstance(d_val, str):
                c_str = d_val.strip().split(" ")[0]
                for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        obj_data['available_from'] = datetime.strptime(c_str, fmt).date()
                        break
                    except ValueError:
                        obj_data['available_from'] = None
            elif isinstance(d_val, datetime):
                obj_data['available_from'] = d_val.date()

        for f in int_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = int(float(str(obj_data[f]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    obj_data[f] = None

        for f in decimal_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = Decimal(str(obj_data[f]).replace(",", "").strip())
                except (InvalidOperation, ValueError):
                    obj_data[f] = None

        for f in ("advance_rent_month", "security_deposit_type"):
            if f in obj_data and obj_data[f] is not None and str(obj_data[f]).lower() != "fixed":
                try:
                    obj_data[f] = str(int(float(obj_data[f])))
                except (TypeError, ValueError):
                    obj_data[f] = str(obj_data[f]).strip()

        # ---- REQUIRED-FIELD VALIDATION (new) ----
        row_errors = []

        missing_fields = [
            _field_label(f) for f in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(f))
        ]
        if missing_fields:
            required_field_errors.append({
                "row": row_idx,
                "missing_fields": missing_fields,
            })
            # Don't add to parsed_rows yet — we may reject the whole file below.
            continue

        ############## Condition start to check admin and user validations ###########

        l_role = str(obj_data.get('listed_by_role', '')).strip().title()
        l_email = str(obj_data.get('listed_by_email', '')).strip().lower()
        l_contact = str(obj_data.get('listed_by_contact', '')).strip()
        l_name = str(obj_data.get('listed_by_name', '')).strip()
        l_id = str(obj_data.get('listed_by_id', '')).strip()

        assigned_to = ""

        # Only perform the DB check if at least one identifying field is provided
        if l_email or l_contact or l_name or l_id:
            is_registered = False

            if l_role.lower() == 'admin':         
                admin_query = Q()
                if l_email:
                    admin_query &= Q(email=l_email)
                if l_contact:
                    admin_query &= Q(phone=l_contact)
                if l_name:
                    admin_query &= Q(name__iexact=l_name)
                if l_id and l_id.isdigit():
                    admin_query &= Q(id=l_id)
                

                if admin_query:
                    
                    is_registered = Admin_Login.objects.filter(admin_query).exists()

            else:
                user_query = Q()
                if l_email:
                    user_query &= Q(user_email=l_email)
                if l_contact:
                    user_query &= Q(user_phone=l_contact)
                if l_name:
                    user_query &= Q(user_name__iexact=l_name)
                if l_id:
                    user_query &= Q(user_id=l_id)

                if user_query:
                    if l_role:
                        is_registered = User_Details.objects.filter(user_query, user_role__iexact=l_role).exists()
                        matched_user = User_Details.objects.filter(user_query, user_role__iexact=l_role).first()
                    else:
                        is_registered = User_Details.objects.filter(user_query).exists()
                        matched_user = User_Details.objects.filter(user_query).first()

                    # If a user was found, they are registered and we can extract their data
                    if matched_user:
                        is_registered = True
                        
                        # Store it in your exact format: user.id-user.role
                        # Note: use matched_user.id (or matched_user.user_id depending on your primary key)
                        assigned_to = f"{matched_user.id}-{matched_user.user_role}"
                        
                        obj_data['assigned_to'] = assigned_to
                    else:
                        is_registered = False

            if not is_registered:
                searched_info = filter(None, [l_id, l_name, l_email, l_contact, l_role])
                identity = " + ".join(searched_info) or "Unknown"
                row_errors.append(
                    f"Listed By {l_role or 'user'} '{identity}' is not present in our records. "
                    f"Please register this {l_role or 'user'} first, then re-upload this row."
                )

        ########### Condition end to check admin and user validations ###################

        # ---- NEW: reject this row if the Listed By identity failed validation ----
        if row_errors:
            listed_by_mismatch_errors.append({
                "row": row_idx,
                "errors": row_errors,
            })
            skipped_listed_by_mismatch += 1
            continue

        parsed_rows.append({'row_idx': row_idx, 'data': obj_data})

    wb.close()

    
    if required_field_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields. "
                "Please fill in every required column (as marked * on the Add Listing form) for "
                "all rows and re-upload the file. No records were saved."
            ),
            "row_errors": required_field_errors,
        }, status=400)

    # ---- Bail out if nothing usable was found ----
    if not parsed_rows and not listed_by_mismatch_errors:
        return JsonResponse({
            "status": "error",
            "message": (
                f"0 usable data rows found. Detected header row {header_row}, "
                f"data expected from row {data_start_row} onward. "
                f"{skipped_empty_after_mapping} row(s) had values but none matched a known column."
            ),
            "unmatched_headers": unmatched_headers,
            "header_row_detected": header_row,
            "data_start_row_assumed": data_start_row,
        }, status=400)

    
    file_name_exists = RentalResidentialProperty.objects.filter(
        upload_file_name=excel_file.name
    ).exists()

    # ---- 4. Write to DB (fingerprint-based duplicate engine, ported from rental_residential_add) ----
    created, updated, skipped, errors = 0, 0, skipped_empty_after_mapping + skipped_listed_by_mismatch, []
    duplicate_blocked_rows = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality_area', '')).strip()
        input_pincode = str(o_data.get('pincode', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_name = str(o_data.get('listed_by_name', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        fingerprint_key = generate_property_fingerprint(
            input_property_no,
            input_building_name,
            input_locality,
            input_pincode
        )

        # Direct case-insensitive match on the same unit in the same locality/building
        direct_duplicates = RentalResidentialProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality_area__iexact=input_locality
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            RentalResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            hard_blocked = False
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and
                        existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and
                            existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            # <-- OUTSIDE the for-loop now, still inside `for item in parsed_rows:`
            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked - Unit '{input_property_no}' is already listed "
                    f"by/for {input_listed_by_name or 'this user'}. Row skipped; edit the existing listing instead."
                )
                skipped += 1
                continue   # now genuinely skips this row of `for item in parsed_rows:`

            # Level 2: Different agent/user listing the exact same physical unit -> allow save & flag
            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id
            )

        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["listing_type"] = o_data.get("listing_type") or "Rental"
        o_data["category"] = o_data.get("category") or "Residential"

        o_data["upload_file_name"] = excel_file.name
        o_data["uploaded_by_name"] = uploader_name
        o_data["uploaded_by_email"] = uploader_email
        o_data["uploaded_by_contact"] = uploader_contact
        o_data["uploaded_by_role"] = uploader_role

        try:
            RentalResidentialProperty.objects.create(**o_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    # Fold hard-blocked rows into the errors list so they're visible in the response
    errors.extend(duplicate_blocked_rows)

    # NEW: fold Listed By mismatches into the errors list so they're visible too
    for entry in listed_by_mismatch_errors:
        for msg in entry["errors"]:
            errors.append(f"Row {entry['row']}: {msg}")

    # ---- 5. Audit Log ----
    RentalActivityLog.objects.create(
        user_identity=user_identity,
        user_role=uploader_role,
        action_type='EXCEL_IMPORT',
        property_id="Multiple / Sheet Records",
        targeted_fields="bulk_action",
        associated_file=excel_file.name,
        action_payload=json.dumps({
            "filename": excel_file.name,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "duplicates_blocked": len(duplicate_blocked_rows),
            "listed_by_mismatches": len(listed_by_mismatch_errors),
            "errors_encountered": len(errors),
        }),
        ip_address=_get_client_ip(request),
        status='SUCCESS' if not errors else 'PARTIAL',
    )

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created, "updated": updated, "skipped": skipped,
        "duplicates_blocked": len(duplicate_blocked_rows),
        "listed_by_mismatches": len(listed_by_mismatch_errors),
        "error_count": len(errors), "errors": errors,
        "unmatched_headers": unmatched_headers,
        "header_row_detected": header_row,
        "data_start_row_used": data_start_row,
    })




from collections import OrderedDict



def _norm(value):
    """Normalize any category string for safe comparison: lowercase + stripped."""
    return (value or '').strip().lower()


def rental_residential_view(request, pk):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    prop = get_object_or_404(
        RentalResidentialProperty.objects.prefetch_related('images', 'faqs'),
        pk=pk
    )

    latest_properties = RentalResidentialProperty.objects.filter(
        is_deleted=False
    ).exclude(
        id=prop.id
    ).prefetch_related('faqs').order_by('-created_at')[:4]

    amenities_list = [x.strip() for x in prop.amenities.split(',')] if prop.amenities else []
    facilities_list = [x.strip() for x in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []

    # ---------- CATEGORY DISPLAY ORDER ----------
    CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
    CATEGORY_LABELS = {
        'exterior': 'Exterior / Building', 'living': 'Living / Dining', 'bedroom': 'Bedroom',
        'kitchen': 'Kitchen', 'bathroom': 'Bathroom', 'balcony': 'Balcony', 'others': 'Others',
    }

    # ---------- IMAGES (sorted using NORMALIZED category so case/whitespace in the DB can't break ordering) ----------
    all_images = sorted(
        prop.images.all(),
        key=lambda img: (
            CATEGORY_ORDER.index(_norm(img.category)) if _norm(img.category) in CATEGORY_ORDER else 99,
            img.sequence_order
        )
    )

    # ---------- GROUPED BY CATEGORY (normalized matching) ----------
    grouped_images = OrderedDict()
    for cat in CATEGORY_ORDER:
        cat_images = [img for img in all_images if _norm(img.category) == cat]
        if cat_images:
            grouped_images[cat] = {'label': CATEGORY_LABELS[cat], 'images': cat_images}

    # ---------- DEDICATED THUMBNAIL SLOTS (normalized matching — THIS is what was silently failing before) ----------
    THUMB_PRIORITY = ['bathroom', 'living', 'bedroom', 'kitchen', 'balcony', 'others']

    def first_image_index(category):
        for idx, img in enumerate(all_images):
            if _norm(img.category) == category:
                return idx
        return None

    hero_category = _norm(all_images[0].category) if all_images else None
    used_categories = {hero_category} if hero_category else set()

    thumb1_idx = thumb2_idx = None
    for cat in THUMB_PRIORITY:
        if cat in used_categories:
            continue
        idx = first_image_index(cat)
        if idx is not None:
            if thumb1_idx is None:
                thumb1_idx = idx
            elif thumb2_idx is None:
                thumb2_idx = idx
                break
            used_categories.add(cat)

    thumb1_image = all_images[thumb1_idx] if thumb1_idx is not None else None
    thumb2_image = all_images[thumb2_idx] if thumb2_idx is not None else None

    # ---------- VIDEO ----------
    video_obj = RentalResidentialVideo.objects.filter(property=prop).first()

    context = {
        'property': prop,
        'images': all_images,
        'grouped_images': grouped_images,
        'thumb1_image': thumb1_image,
        'thumb1_idx': thumb1_idx,
        'thumb2_image': thumb2_image,
        'thumb2_idx': thumb2_idx,
        'video_obj': video_obj,
        'faqs': prop.faqs.all(),
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'latest_properties': latest_properties,
        'admin_obj': admin_obj,
        # TEMP DEBUG — remove once confirmed. Shows raw category values as stored in DB.
        'debug_categories': [(img.id, repr(img.category)) for img in all_images],
    }
    return render(request, 'admin_user/Reports/Rental/rental_residential_detail.html', context)


def _get_client_ip(request):
    """Helper to safely fetch client IP address reference."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')








def rental_residential_edit(request, pk):
    prop = get_object_or_404(RentalResidentialProperty, id=pk)

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    if request.method == 'POST':
        try:
            p = request.POST

            def to_int(val):
                try: return int(str(val).replace(',', '').strip()) if val else None
                except: return None

            def to_decimal(val):
                try: return float(str(val).replace(',', '').strip()) if val else None
                except: return None

            # =====================================================
            # 1. AUDIT LOGIC: Snapshot state BEFORE modification
            # =====================================================
            tracked_fields_list = [
                'property_type', 'property_no', 'bhk_type', 'renting_option',
                'built_up_area', 'bathrooms', 'balconies', 'building_configuration',
                'total_floors', 'facing_direction', 'furnishing_status', 'available_for',
                'carpet_area', 'city_zone', 'ownership_type', 'property_condition',
                'property_age', 'wing_number', 'building_name',
                'availability_status', 'available_from', 'lease_duration',
                'brokerage_percentage', 'manual_brokerage', 'monthly_rent',
                'advance_rent_month', 'advance_rent_amount',
                'security_deposit_type', 'security_deposit_amount',
                'maintenance_type', 'monthy_maintenance_amount', 'total_move_in_cost',
                'address', 'city', 'locality_area', 'property_landmark', 'state', 'pincode',
                'main_road_connectivity', 'google_maps_link', 'latitude', 'longitude',
                'listed_by_id', 'listed_by_name', 'listed_by_email', 'listed_by_contact', 'listed_by_role',
                'listing_status', 'approval_status',
            ]
            old_state_snapshot = {field: str(getattr(prop, field, '')) for field in tracked_fields_list}

            # =====================================================
            # 2. LISTED BY & BASIC INFORMATION
            # =====================================================
            prop.listed_by_type = p.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to = p.get('assigned_to', prop.assigned_to)
            prop.listed_by_id = (p.get('listed_by_id') or prop.listed_by_id or '').strip()
            prop.listed_by_name = (p.get('listed_by_name') or prop.listed_by_name or '').strip()
            prop.listed_by_email = (p.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = (p.get('listed_by_contact') or prop.listed_by_contact or '').strip()
            prop.listed_by_role = (p.get('listed_by_role') or prop.listed_by_role or '').strip()

            prop.property_type = p.get('property_type')
            prop.property_no = (p.get('property_no') or prop.property_no or '').strip()
            prop.bhk_type = p.get('bhk_type')
            prop.renting_option = p.get('renting_option')
            prop.built_up_area = to_decimal(p.get('built_up_area'))
            prop.bathrooms = to_int(p.get('bathrooms'))
            prop.balconies = to_int(p.get('balconies'))
            prop.building_configuration = p.get('building_configuration')
            prop.total_floors = to_int(p.get('total_floors'))
            prop.facing_direction = p.get('facing_direction')
            prop.furnishing_status = p.get('furnishing_status')
            prop.available_for = p.get('available_for')

            # =====================================================
            # 3. PROPERTY DETAILS & LOCATION
            # =====================================================
            prop.carpet_area = to_decimal(p.get('carpet_area'))
            prop.city_zone = p.get('city_zone')
            prop.ownership_type = p.get('ownership_type')
            prop.property_condition = p.get('property_condition')
            prop.property_age = p.get('property_age')
            prop.wing_number = p.get('wing_number')
            prop.building_name = (p.get('building_name') or '').strip()

            prop.availability_status = p.get('availability_status')
            available_from_raw = p.get('available_from')
            if available_from_raw and available_from_raw.strip():
                try:
                    prop.available_from = datetime.strptime(available_from_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    prop.available_from = None
            else:
                prop.available_from = None

            prop.lease_duration = p.get('lease_duration')
            prop.brokerage_percentage = p.get('brokerage_percentage')
            prop.manual_brokerage = p.get('manual_brokerage')

            # =====================================================
            # 4. PRICING & LOCATION DETAILS
            # =====================================================
            prop.monthly_rent = to_int(p.get('monthly_rent'))
            prop.advance_rent_month = p.get('advance_rent_month')
            prop.advance_rent_amount = to_int(p.get('advance_rent_amount'))
            prop.security_deposit_type = p.get('security_deposit_type')
            prop.security_deposit_amount = to_int(p.get('security_deposit_amount'))
            prop.maintenance_type = p.get('maintenance_type')
            prop.monthy_maintenance_amount = to_int(p.get('monthy_maintenance_amount'))
            prop.total_move_in_cost = to_int(p.get('total_move_in_cost'))

            prop.address = p.get('address')
            prop.city = p.get('city')
            prop.locality_area = (p.get('locality_area') or p.get('locality') or '').strip()
            prop.property_landmark = p.get('property_landmark')
            prop.state = p.get('state')
            prop.pincode = (p.get('pincode') or '').strip()
            prop.main_road_connectivity = p.get('main_road_connectivity')
            prop.google_maps_link = p.get('google_maps_link')
            prop.latitude = p.get('latitude')
            prop.longitude = p.get('longitude')

            # =====================================================
            # 5. AMENITIES, DESCRIPTIONS & STATUS
            # =====================================================
            prop.amenities = ",".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities = ",".join(request.POST.getlist('nearby_facilities[]'))
            prop.user_description = p.get('user_description')
            prop.listed_elsewhere = p.get('listed_elsewhere', 'No')
            prop.portal_name = p.get('portal_name')
            prop.listing_status = p.get('listing_status', prop.listing_status)
            prop.approval_status = p.get('approval_status', prop.approval_status)

            prop.save()

            # =====================================================
            # 6. IMAGE DELETIONS
            # =====================================================
            delete_ids = request.POST.getlist('delete_image_ids[]')
            if delete_ids:
                RentalResidentialImage.objects.filter(id__in=delete_ids, property=prop).delete()

            # =====================================================
            # 7. NEW CATEGORIZED IMAGES UPLOAD LOGIC
            # =====================================================
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'living':   'property_images_living[]',
                'bedroom':  'property_images_bedroom[]',
                'kitchen':  'property_images_kitchen[]',
                'bathroom': 'property_images_bathroom[]',
                'balcony':  'property_images_balcony[]',
                'others':   'property_images_others[]',
            }
            existing_total = prop.images.count()
            new_images_added = 0

            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                existing_cat_count = prop.images.filter(category=category).count()
                for img in request.FILES.getlist(field_name):
                    if existing_total >= 20:
                        break
                    RentalResidentialImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=existing_cat_count
                    )
                    existing_cat_count += 1
                    existing_total += 1
                    new_images_added += 1

            images_changed = new_images_added > 0 or bool(delete_ids)

            # =====================================================
            # 8. VIDEO LOGIC: DIRECT ADMIN SELECTION
            # =====================================================
            # =====================================================
            # 8. VIDEO LOGIC: DIRECT ADMIN SELECTION
            # =====================================================
            video_option = p.get('video_option')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = p.get('property_video_link', '').strip()

            if video_option == 'upload' and uploaded_video:
                video_row, _ = RentalResidentialVideo.objects.get_or_create(
                    property=prop, 
                    source='uploaded'
                )
                video_row.video = uploaded_video
                video_row.video_url = None
                video_row.save()

            elif video_option == 'rm_assisted' and property_video_link:
                video_row, _ = RentalResidentialVideo.objects.get_or_create(
                    property=prop, 
                    source='rm_assisted'
                )
                video_row.video_url = property_video_link
                video_row.video = None
                video_row.save()

            # =====================================================
            # 9. AUTO SLIDESHOW: REFRESH ON PHOTO CHANGE OR REQUEST
            # =====================================================
            regenerate_requested = p.get('regenerate_slideshow') == 'on'

            if images_changed or regenerate_requested:
                CATEGORY_ORDER = ['exterior', 'living', 'bedroom', 'kitchen', 'bathroom', 'balcony', 'others']
                saved_images = list(prop.images.all())
                saved_images.sort(key=lambda img: (
                    CATEGORY_ORDER.index(img.category) if img.category in CATEGORY_ORDER else 99, 
                    img.sequence_order
                ))
                image_paths = [
                    img.image.path for img in saved_images
                    if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                ]

                if len(image_paths) >= 3:
                    try:
                        from .utils.video_generator import generate_property_slideshow

                        video_row, _ = RentalResidentialVideo.objects.get_or_create(
                            property=prop, 
                            source='auto'
                        )

                        # Clean up old file to save disk space
                        if video_row.video and video_row.video.name:
                            old_path = video_row.video.path
                            if os.path.exists(old_path):
                                try: os.remove(old_path)
                                except Exception as del_err: print("Could not delete old auto video:", del_err)

                        # Timestamped file path to prevent browser caching stale videos
                        out_path = f"residential_rent/videos/auto_{prop.id}_{int(time.time())}.mp4"

                        result = generate_property_slideshow(image_paths, out_path)
                        if result:
                            video_row.video = out_path
                            video_row.save()
                    except Exception as ve:
                        print("RESIDENTIAL VIDEO REGEN FAILED:", str(ve))
                        traceback.print_exc()
            # =====================================================
            # 10. AUDIT LOGGING & DIFF DICTIONARY
            # =====================================================
            modifications_diff = {}
            modified_fields_summary = []

            for field in tracked_fields_list:
                new_val = str(getattr(prop, field, ''))
                if old_state_snapshot[field] != new_val:
                    modifications_diff[field] = {
                        "old_value": old_state_snapshot[field],
                        "new_value": new_val
                    }
                    modified_fields_summary.append(field)

            if modified_fields_summary:
                RentalActivityLog.objects.create(
                    user_identity=admin_obj.email or admin_obj.username,
                    user_role="Admin",
                    action_type='UPDATE',
                    property_id=prop.id,
                    targeted_fields=", ".join(modified_fields_summary[:4]) + ("..." if len(modified_fields_summary) > 4 else ""),
                    associated_file="Web UI Form",
                    action_payload=json.dumps(modifications_diff),
                    ip_address=_get_client_ip(request),
                    status='SUCCESS'
                )

            return JsonResponse({
                'status': 'success',
                'message': 'Residential Rental Property Updated Successfully',
                'redirect_url': reverse('residential_list')
            })

        except Exception as e:
            print("ERROR IN RESIDENTIAL EDIT:", str(e))
            traceback.print_exc()

            return JsonResponse({'status': 'error', 'message': f"Failed to save data: {str(e)}"}, status=400)

    # ================= GET REQUEST CONTEXT =================
    residential_videos = RentalResidentialVideo.objects.filter(property=prop)
    videos_by_source = {v.source: v for v in residential_videos}

    images_by_category = {}
    for img in prop.images.all().order_by('category', 'sequence_order'):
        images_by_category.setdefault(img.category, []).append(img)

    context = {
        'property': prop,
        'admin_obj': admin_obj,
        'user_obj': User_Details.objects.all(),
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'uploaded_video': videos_by_source.get('uploaded'),
        'auto_video': videos_by_source.get('auto'),
        'rm_video': videos_by_source.get('rm_assisted'),
        'images_by_category': images_by_category,
        'selected_amenities': [a.strip() for a in (prop.amenities or '').split(',') if a.strip()],
        'selected_facilities': [f.strip() for f in (prop.nearby_facilities or '').split(',') if f.strip()],
    }
    return render(request, 'admin_user/rental_residential_edit.html', context)


##################################RESIDENTIAL RENTAL LISTING VEIW SECTION END##############################

 ######################START VIEW SECTION OF RENTAL COMMERCIAL VIEW SECTION####################################










def commercial_rental_add(request):
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    # ---------- SAFE TYPE CONVERSIONS ----------
    def to_int(val, default=0):
        try: 
            return int(str(val).replace(',', '').strip())
        except: 
            return default

    def to_int_or_none(val):
        try: 
            return int(str(val).replace(',', '').strip())
        except: 
            return None

    def to_float_or_none(val):
        try: 
            return float(str(val).replace(',', '').strip())
        except: 
            return None

    if request.method == "POST":
        try:
            # ---------- UPLOADER IDENTIFICATION ----------
            if admin_id:
                admin = Admin_Login.objects.get(id=admin_id)
                uploader_name = getattr(admin, 'name', '') or getattr(admin, 'username', '')
                uploader_email = getattr(admin, 'email', '')
                uploader_phone = getattr(admin, 'phone', '') or getattr(admin, 'mobile', '')
                uploader_role = "Admin"
                uploader_id = f"ADMIN_{admin_id}"
            else:
                user = User_Details.objects.get(id=user_id)
                uploader_name = user.user_name
                uploader_email = user.user_email
                uploader_phone = user.user_phone
                uploader_role = "User"
                uploader_id = f"USER_{user_id}"

            # ---------- LISTED BY IDENTIFICATION ----------
            input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
            input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
            input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
            input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_phone).strip()
            input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

            # ---------- DATE PARSING ----------
            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except:
                    available_from = None
            else:
                available_from = None

            # ==========================================================
            # DUPLICATE DETECTION ENGINE
            # ==========================================================
            input_property_no = (request.POST.get('property_no') or '').strip()
            input_building_name = (request.POST.get('building_name') or '').strip()
            input_locality = (request.POST.get('locality') or request.POST.get('locality_area') or '').strip()
            input_city = (request.POST.get('city') or '').strip()
            input_pincode = (request.POST.get('pincode') or '').strip()

            # Generate unique fingerprint key
            fingerprint_key = f"{input_property_no}|{input_building_name}|{input_locality}|{input_city}|{input_pincode}".lower().replace(" ", "")

            # 1. Direct Case-Insensitive Query
            direct_duplicates = CommercialRentalProperty.objects.filter(
                is_deleted=False,
                property_no__iexact=input_property_no,
                locality__iexact=input_locality
            )
            if input_building_name:
                direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

            existing_duplicates = (
                CommercialRentalProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False) | direct_duplicates
            ).distinct()

            is_dup_flag = False
            dup_group_id = fingerprint_key
            total_dup_count = 1

            if existing_duplicates.exists():
                for existing_prop in existing_duplicates:
                    same_id = (existing_prop.listed_by_id and input_listed_by_id and 
                               existing_prop.listed_by_id.strip() == input_listed_by_id)
                    
                    same_email = (existing_prop.listed_by_email and input_listed_by_email and 
                                  existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                    
                    same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and 
                                    existing_prop.listed_by_contact.strip() == input_listed_by_contact)
                    
                    if same_id or same_email or same_contact:
                        return JsonResponse({
                            "status": "error", 
                            "message": f"Duplicate Blocked: This commercial unit ({input_property_no}) is already listed by/for {input_listed_by_name or 'this user'}. Please edit the existing listing Property instead."
                        })

                is_dup_flag = True
                total_dup_count = existing_duplicates.count() + 1
                existing_duplicates.update(
                    is_duplicate=True, 
                    duplicate_count=total_dup_count,
                    duplicate_group_id=dup_group_id
                )

            # ---------- AMENITIES & FACILITIES ----------
            amenities_list = ", ".join(request.POST.getlist('amenities[]'))
            facilities_list = ", ".join(request.POST.getlist('nearby_facilities[]'))

            # ---------- CREATE DATABASE OBJECT ----------
            prop = CommercialRentalProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="Commercial",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                property_title=request.POST.get('property_title'),
                property_type=request.POST.get('property_type'),
                building_name=input_building_name,
                wing_number=request.POST.get('wing_number'),
                property_no=input_property_no,

                availability_status=request.POST.get('availability_status'),
                available_from=available_from,
                property_age=to_int_or_none(request.POST.get('property_age')),

                zone_type=request.POST.get('zone_type'),
                ownership_type=request.POST.get('ownership_type'),
                property_condition=request.POST.get('property_condition'),

                builtup_area=to_int(request.POST.get('builtup_area')),
                carpet_area=to_int_or_none(request.POST.get('carpet_area')),

                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                address=request.POST.get('address'),
                locality=input_locality,
                property_landmark=request.POST.get('property_landmark'),
                city=request.POST.get('city'),
                state=request.POST.get('state'),
                location_hub=request.POST.get('location_hub'),
                pincode=input_pincode,
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                dg_ups_included=(request.POST.get('dg_ups_included') == 'on'),
                electricity_included=(request.POST.get('electricity_included') == 'on'),
                water_included=(request.POST.get('water_included') == 'on'),

                building_configuration=request.POST.get('building_configuration'),
                total_floors=to_int_or_none(request.POST.get('total_floors')),
                staircases=to_int_or_none(request.POST.get('staircases')),
                passenger_lifts=to_int(request.POST.get('passenger_lifts')),
                service_lifts=to_int(request.POST.get('service_lifts')),
                private_parking=to_int(request.POST.get('private_parking')),

                min_seats=to_int_or_none(request.POST.get('min_seats')),
                max_seats=to_int_or_none(request.POST.get('max_seats')),
                cabins=to_int_or_none(request.POST.get('cabins')),
                meeting_rooms=to_int_or_none(request.POST.get('meeting_rooms')),
                private_washroom=to_int(request.POST.get('private_washroom')),
                public_washroom=to_int(request.POST.get('public_washroom')),
                flooring_type=request.POST.get('flooring_type'),

                monthly_rent=to_int(request.POST.get('monthly_rent')),
                advanced_rent_type=request.POST.get('advanced_rent_type'),
                advanced_rent_amount=to_int_or_none(request.POST.get('advanced_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=to_int_or_none(request.POST.get('security_deposit_amount')),
                maintenance_type=request.POST.get('maintenance_type'),
                maintenance_charges=to_int_or_none(request.POST.get('maintenance_charges')),
                total_move_in_cost=to_int_or_none(request.POST.get('total_move_in_cost')),
                negotiable=request.POST.get('negotiable', 'No'),
                lockin_period=to_int_or_none(request.POST.get('lockin_period')),
                rent_increase=to_float_or_none(request.POST.get('rent_increase')),

                amenities=amenities_list,
                nearby_facilities=facilities_list,
                user_description=request.POST.get('user_description'),
                property_summary=request.POST.get('property_summary'),
                property_description=request.POST.get('property_description'),

                listed_elsewhere=request.POST.get('listed_elsewhere', 'No'),
                portal_name=request.POST.get('portal_name'),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_phone,
                uploaded_by_role=uploader_role,
            )

            # ---------- IMAGES MULTI-UPLOAD LOGIC (12 CATEGORIES) ----------
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'reception': 'property_images_reception[]',
                'workspace': 'property_images_workspace[]',
                'cabins': 'property_images_cabins[]',
                'meeting_room': 'property_images_meeting_room[]',
                'shop_showroom': 'property_images_shop_showroom[]',
                'warehouse': 'property_images_warehouse[]',
                'pantry': 'property_images_pantry[]',
                'washroom': 'property_images_washroom[]',
                'parking': 'property_images_parking[]',
                'amenities': 'property_images_amenities[]',
                'floor_plan': 'property_images_floor_plan[]',
            }
            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                for cat_index, img in enumerate(request.FILES.getlist(field_name)):
                    if saved_count >= 30:
                        break
                    CommercialRentalPropertyImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=cat_index
                    )
                    saved_count += 1

            # ---------- PROPERTY VIDEO (UPLOAD, AUTO, OR RM ASSISTED) ----------
            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '')

            # 1. ALWAYS auto-generate the slideshow (runs every time, unconditionally)
            saved_images = list(CommercialRentalPropertyImage.objects.filter(property=prop))
            image_paths = []
            for img_obj in saved_images:
                if img_obj.image and hasattr(img_obj.image, 'path') and os.path.exists(img_obj.image.path):
                    image_paths.append(img_obj.image.path)

            if len(image_paths) >= 3:
                try:
                    from .utils.commercial_video_generator import generate_commercial_property_slideshow
                    out_path = f"commercial_rent/videos/auto_{prop.id}.mp4"
                    result = generate_commercial_property_slideshow(image_paths, out_path)
                    print("AUTO SLIDESHOW RESULT:", result)
                    if result:
                        CommercialRentalVideo.objects.create(
                            property=prop,
                            video=out_path,
                            source='auto',
                            video_status='Done',
                            video_url=None
                        )
                except Exception as ve:
                    print("COMMERCIAL VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()
            else:
                CommercialRentalVideo.objects.create(
                    property=prop,
                    video=None,
                    source='auto',
                    video_status='Pending (Insufficient Images)',
                    video_url=None
                )

            # 2. Manual Upload Logic (separate row, source='uploaded')
            if video_option == 'upload' and uploaded_video:
                if uploaded_video.size > 20 * 1024 * 1024:
                    CommercialRentalVideo.objects.create(
                        property=prop,
                        video=None,
                        source='uploaded',
                        video_status='Pending',
                        video_url=None
                    )
                else:
                    CommercialRentalVideo.objects.create(
                        property=prop,
                        video=uploaded_video,
                        source='uploaded',
                        video_status='Done',
                        video_url=None
                    )

            # 3. RM Assisted Link Logic (separate row, source='rm_assisted')
            elif video_option == 'rm_assisted':
                CommercialRentalVideo.objects.create(
                    property=prop,
                    video=None,
                    source='rm_assisted',
                    video_url=property_video_link,
                    video_status='Done' if property_video_link else 'Pending'
                )

            

            return JsonResponse({"status": "success", "message": "Commercial Rental Property Added Successfully"})

        except Exception as e:
            print("ERROR IN COMMERCIAL ADD:", str(e))
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})

    return render(request, 'admin_user/Reports/Rental/commercial.html')

def _get_admin(request):
    sid = request.session.get('Admin_id')
    admin_obj = None
    if sid:
        try:
            admin_obj = Admin_Login.objects.get(id=sid)
        except Admin_Login.DoesNotExist:
            sid = None
    return sid, admin_obj

def _to_int(val):
    if not val: return None
    try: return int(val)
    except ValueError: return None

def _to_float(val):
    if not val: return None
    try: return float(val)
    except ValueError: return None

def _to_date(val):
    return val if val else None



def commercial_view(request, pk):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')
        
    admin_obj = Admin_Login.objects.get(id=session_id)
    prop = get_object_or_404(CommercialRentalProperty, pk=pk)

    amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []
    facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []

    # Video Hierarchy & Backend Saved Video Retrieval
    videos_queryset = prop.walkthrough_video.all()
    rm_video = videos_queryset.filter(source='rm_assisted', video_url__isnull=False).exclude(video_url='').first()
    manual_video = videos_queryset.filter(source='uploaded', video__isnull=False).first()
    auto_video = videos_queryset.filter(source='auto', video__isnull=False).first()

    selected_video = None
    video_display_mode = None

    if rm_video:
        selected_video = rm_video
        video_display_mode = 'rm_assisted'
    elif manual_video:
        selected_video = manual_video
        video_display_mode = 'manual'
    elif auto_video:
        selected_video = auto_video
        video_display_mode = 'auto'

    return render(request, 'admin_user/Reports/Rental/commercial_detail.html', {
        'admin_obj': admin_obj,
        'prop': prop,
        'amenities_list': amenities_list,
        'facilities_list': facilities_list,
        'selected_video': selected_video,
        'video_display_mode': video_display_mode,
    })
# ═══════════════════════════════════════
# COMMERCIAL EDIT/UPDATE VIEW
# ═══════════════════════════════════════

def commercial_edit(request, pk):
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return render(request, 'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=admin_id)
    except Admin_Login.DoesNotExist:
        return render(request, 'home_page/Adminlogin.html')

    prop = get_object_or_404(CommercialRentalProperty, pk=pk, is_deleted=False)

    def to_int(val, default=0):
        try: return int(str(val).replace(',', '').strip())
        except: return default

    def to_int_or_none(val):
        try: return int(str(val).replace(',', '').strip())
        except: return None

    def to_float_or_none(val):
        try: return float(str(val).replace(',', '').strip())
        except: return None

    if request.method == "POST":
        try:
            p = request.POST

            date_val = p.get('available_from')
            if date_val:
                try:
                    available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except Exception:
                    available_from = None
            else:
                available_from = None

            # ---------- LISTED BY ----------
            prop.listed_by_type    = p.get('listed_by_type', prop.listed_by_type)
            prop.assigned_to       = p.get('assigned_to', prop.assigned_to)
            prop.listed_by_id      = p.get('listed_by_id') or prop.listed_by_id
            prop.listed_by_name    = p.get('listed_by_name') or prop.listed_by_name
            prop.listed_by_email   = (p.get('listed_by_email') or prop.listed_by_email or '').strip().lower()
            prop.listed_by_contact = p.get('listed_by_contact') or prop.listed_by_contact
            prop.listed_by_role    = p.get('listed_by_role') or prop.listed_by_role

            # ---------- BASIC INFO ----------
            prop.property_title = p.get('property_title', prop.property_title)
            prop.property_type  = p.get('property_type')
            prop.building_name  = (p.get('building_name') or '').strip()
            prop.wing_number    = p.get('wing_number')
            prop.property_no    = (p.get('property_no') or '').strip()

            prop.availability_status = p.get('availability_status')
            prop.available_from      = available_from
            prop.property_age        = to_int_or_none(p.get('property_age'))

            prop.zone_type         = p.get('zone_type')
            prop.ownership_type    = p.get('ownership_type')
            prop.property_condition = p.get('property_condition')

            prop.builtup_area = to_int(p.get('builtup_area'))
            prop.carpet_area  = to_int_or_none(p.get('carpet_area'))

            # ---------- LOCATION ----------
            prop.address           = p.get('address')
            prop.locality          = (p.get('locality') or '').strip()
            prop.property_landmark = p.get('property_landmark')
            prop.city              = p.get('city')
            prop.state             = p.get('state')
            prop.location_hub      = p.get('location_hub')
            prop.pincode           = (p.get('pincode') or '').strip()
            prop.google_maps_link  = p.get('google_maps_link')
            prop.latitude          = p.get('latitude')
            prop.longitude         = p.get('longitude')

            # ---------- BUILDING SPECS ----------
            prop.dg_ups_included      = (p.get('dg_ups_included') == 'on')
            prop.electricity_included = (p.get('electricity_included') == 'on')
            prop.water_included       = (p.get('water_included') == 'on')

            prop.building_configuration = p.get('building_configuration')
            prop.total_floors    = to_int_or_none(p.get('total_floors'))
            prop.staircases      = to_int_or_none(p.get('staircases'))
            prop.passenger_lifts = to_int(p.get('passenger_lifts'))
            prop.service_lifts   = to_int(p.get('service_lifts'))
            prop.private_parking = to_int(p.get('private_parking'))

            prop.min_seats         = to_int_or_none(p.get('min_seats'))
            prop.max_seats         = to_int_or_none(p.get('max_seats'))
            prop.cabins            = to_int_or_none(p.get('cabins'))
            prop.meeting_rooms     = to_int_or_none(p.get('meeting_rooms'))
            prop.private_washroom  = to_int(p.get('private_washroom'))
            prop.public_washroom   = to_int(p.get('public_washroom'))
            prop.flooring_type     = p.get('flooring_type')

            # ---------- PRICING ----------
            prop.monthly_rent            = to_int(p.get('monthly_rent'))
            prop.brokerage_percentage    = p.get('brokerage_percentage')
            prop.manual_brokerage        = p.get('manual_brokerage')
            prop.advanced_rent_type      = p.get('advanced_rent_type')
            prop.advanced_rent_amount    = to_int_or_none(p.get('advanced_rent_amount'))
            prop.security_deposit_type   = p.get('security_deposit_type')
            prop.security_deposit_amount = to_int_or_none(p.get('security_deposit_amount'))
            prop.maintenance_type        = p.get('maintenance_type')
            prop.maintenance_charges     = to_int_or_none(p.get('maintenance_charges'))
            prop.total_move_in_cost      = to_int_or_none(p.get('total_move_in_cost'))
            prop.negotiable              = p.get('negotiable', prop.negotiable)
            prop.lockin_period           = to_int_or_none(p.get('lockin_period'))
            prop.rent_increase           = to_float_or_none(p.get('rent_increase'))

            # ---------- AMENITIES / DESCRIPTIONS ----------
            prop.amenities            = ", ".join(request.POST.getlist('amenities[]'))
            prop.nearby_facilities    = ", ".join(request.POST.getlist('nearby_facilities[]'))
            prop.user_description     = p.get('user_description')
            prop.property_summary     = p.get('property_summary') or prop.property_summary
            prop.property_description = p.get('property_description') or prop.property_description

            # ---------- LISTED ELSEWHERE ----------
            prop.listed_elsewhere = p.get('listed_elsewhere', 'No')
            prop.portal_name      = p.get('portal_name')
            prop.listing_status = p.get('listing_status')
            prop.approval_status = p.get('approval_status')

            prop.save()
            # ================= IMAGE DELETIONS =================
            delete_ids = request.POST.getlist('delete_image_ids[]')
            if delete_ids:
                CommercialRentalPropertyImage.objects.filter(id__in=delete_ids, property=prop).delete()

            # ================= NEW IMAGES BY CATEGORY =================
            IMAGE_CATEGORY_FIELDS = {
                'exterior': 'property_images_exterior[]',
                'reception': 'property_images_reception[]',
                'workspace': 'property_images_workspace[]',
                'cabins': 'property_images_cabins[]',
                'meeting_room': 'property_images_meeting_room[]',
                'shop_showroom': 'property_images_shop_showroom[]',
                'warehouse': 'property_images_warehouse[]',
                'pantry': 'property_images_pantry[]',
                'washroom': 'property_images_washroom[]',
                'parking': 'property_images_parking[]',
                'amenities': 'property_images_amenities[]',
                'floor_plan': 'property_images_floor_plan[]',
            }
            existing_total = prop.images.count()
            new_images_added = 0

            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                existing_cat_count = prop.images.filter(category=category).count()
                for img in request.FILES.getlist(field_name):
                    if existing_total >= 30:
                        break
                    CommercialRentalPropertyImage.objects.create(
                        property=prop, image=img, category=category, sequence_order=existing_cat_count
                    )
                    existing_cat_count += 1
                    existing_total += 1
                    new_images_added += 1

            images_changed = new_images_added > 0 or bool(delete_ids)

            # ================= VIDEO: SOURCE THE ADMIN ACTIVELY PICKED =================
            # This part stays exactly gated by video_option — only touches the source
            # the admin is currently interacting with (upload file OR RM link).
            video_option = p.get('video_option')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = p.get('property_video_link', '')

            if video_option == 'upload' and uploaded_video:
                video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='uploaded')
                if uploaded_video.size > 20 * 1024 * 1024:
                    video_row.video = None
                    video_row.video_status = 'Pending'
                else:
                    video_row.video = uploaded_video
                    video_row.video_status = 'Done'
                video_row.save()

            elif video_option == 'rm_assisted' and property_video_link:
                video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='rm_assisted')
                video_row.video_url = property_video_link
                video_row.video_status = 'Done'
                video_row.save()

            # ================= AUTO SLIDESHOW: ALWAYS KEEPS PACE WITH PHOTOS =================
            # Completely independent of video_option — the 'auto' row is a separate
            # database record from 'uploaded'/'rm_assisted', so it must refresh whenever
            # photos change, even if the admin is currently uploading a video manually.
            # This mirrors the add view, which regenerates the slideshow unconditionally.
            # ================= AUTO SLIDESHOW: ALWAYS KEEPS PACE WITH PHOTOS =================
            regenerate_requested = p.get('regenerate_slideshow') == 'on'

            if images_changed or regenerate_requested:
                saved_images = list(prop.images.all())
                image_paths = [
                    img.image.path for img in saved_images
                    if img.image and hasattr(img.image, 'path') and os.path.exists(img.image.path)
                ]
                if len(image_paths) >= 3:
                    try:
                        from .utils.commercial_video_generator import generate_commercial_property_slideshow
                        import time

                        video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='auto')

                        # Delete the old physical file before writing a new one, so we don't
                        # leave orphaned video files piling up on disk with every edit.
                        if video_row.video and video_row.video.name:
                            old_path = video_row.video.path
                            if os.path.exists(old_path):
                                try:
                                    os.remove(old_path)
                                except Exception as del_err:
                                    print("Could not delete old auto video:", del_err)

                        # Unique filename per regeneration — this is the actual fix.
                        # A deterministic filename (auto_{id}.mp4) gets cached by the browser
                        # forever since the URL never changes even when the file content does.
                        out_path = f"commercial_rent/videos/auto_{prop.id}_{int(time.time())}.mp4"

                        result = generate_commercial_property_slideshow(image_paths, out_path)
                        if result:
                            video_row.video = out_path
                            video_row.video_status = 'Done'
                            video_row.save()
                    except Exception as ve:
                        print("COMMERCIAL VIDEO REGEN FAILED:", str(ve))
                        traceback.print_exc()
                else:
                    video_row, _ = CommercialRentalVideo.objects.get_or_create(property=prop, source='auto')
                    video_row.video_status = 'Pending (Insufficient Images)'
                    video_row.save()

            return JsonResponse({"status": "success", "message": "Commercial Rental Property Updated Successfully"})

        except Exception as e:
            print("ERROR IN COMMERCIAL EDIT:", str(e))
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})

    # ================= GET REQUEST =================
    videos_by_source = {v.source: v for v in prop.walkthrough_video.all()}
    images_by_category = {}
    for img in prop.images.all().order_by('category', 'sequence_order'):
        images_by_category.setdefault(img.category, []).append(img)

    context = {
        'admin_obj': admin_obj,
        'prop': prop,
        'ameneties_obj': Ameneties_Details.objects.all(),
        'facilities_obj': Facilities_Details.objects.all(),
        'user_obj': User_Details.objects.all(),
        'uploaded_video': videos_by_source.get('uploaded'),
        'auto_video': videos_by_source.get('auto'),
        'rm_video': videos_by_source.get('rm_assisted'),
        'images_by_category': images_by_category,
        'selected_amenities': [a.strip() for a in (prop.amenities or '').split(',') if a.strip()],
        'selected_facilities': [f.strip() for f in (prop.nearby_facilities or '').split(',') if f.strip()],
    }
    return render(request, 'admin_user/commercial_edit.html', context)
# ─────────────────────────────
# DELETE
@require_POST
def commercial_delete(request, pk):
    # Auth check
    sid, _ = _get_admin(request)
    if not sid:
        return JsonResponse({'status': 'error'}, status=401)

    # Grab the deleter's name using our helper!
    deleter_name = _get_deleter_name(request)

    prop = get_object_or_404(CommercialRentalProperty, pk=pk)
    prop.is_deleted = True
    prop.deleted_at = timezone.now()
    prop.deleted_by = deleter_name # 👈 SAVE THE NAME HERE
    prop.save()

    return JsonResponse({'status': 'success', 'message': 'Moved to Recycle Bin successfully!'})


@require_POST
def commercial_restore(request, id):
    CommercialRentalProperty.objects.filter(id=id).update(is_deleted=False, deleted_at=None, deleted_by=None)
    return JsonResponse({'status': 'success', 'message': 'Commercial property restored!'})

@require_POST
def commercial_hard_delete(request, id):
    CommercialRentalProperty.objects.filter(id=id).delete()
    return JsonResponse({'status': 'success', 'message': 'Permanently deleted!'})







####################################Start View Section Rental Commerical download sample & import file################ 






def _commercial_rental_field_map():
    """
    Returns field mappings where the label strings are an EXACT match to the 
    human-readable form labels used in commercial.html.
    """
    sections = OrderedDict([
        ("Listed By Details", [
            ("listed_by_type",    "Listed By Type", False),
            ("listed_by_id",      "Listed By Id", False),
            ("listed_by_name",    "Listed By Name", False),
            ("listed_by_email",   "Listed By Email", False),
            ("listed_by_contact", "Listed By Contact", False),
            ("listed_by_role",    "Listed By Role", True),  # Drives brokerage label preview
        ]),
        ("Basic Information", [
            ("property_type",       "Property Type", True),
            ("building_name",       "Building / Project Name", True),
            ("wing_number",         "Tower/Wing Number", False),
            ("property_no",         "Shop No/Godown/Unit No", True),
            ("availability_status", "Availability Status", True),
            ("available_from",      "Available From", False),
            ("property_age",        "Age of Property(In Years)", True),
            ("zone_type",           "Zone Type", False),
            ("ownership_type",      "Ownership Type", True),
            ("property_condition",  "Property Condition", False),
            ("builtup_area",        "Built-up Area (sq.ft)", True),
            ("carpet_area",         "Carpet Area (sq.ft)", False),
           
        ]),
        ("Property Location Details", [
            ("address",           "Property Address", True),
            ("locality",          "Area/Locality", True),
            ("property_landmark", "Property Landmark", False),
            ("city",              "City", True),
            ("state",             "State", True),
            ("location_hub",      "Location Hub", False),
            ("pincode",           "Pincode", True),  # Added missing field from HTML
            ("google_maps_link",  "Google Maps Link", False),
            ("latitude",          "Latitude", False),
            ("longitude",         "Longitude", False),
        ]),
        ("Property Specifications", [
            ("dg_ups_included",        "DG & UPS Included?", False),
            ("electricity_included",   "Electricity Included?", False),
            ("water_included",         "Water Included?", False),
            ("building_configuration", "Building Configuration", False),
            ("total_floors",           "Total Floors Constructed", True),
            ("staircases",             "Staircases", False),
            ("passenger_lifts",        "Passenger Lifts", False),
            ("service_lifts",          "Service Lifts", False),
            ("private_parking",        "Private Parking", False),
            ("min_seats",              "Minimum Seats Occupancy", False),
            ("max_seats",              "Maximum Seats Occupancy", False),
            ("cabins",                 "Cabins", False),
            ("meeting_rooms",          "Meeting Rooms", False),
            ("private_washroom",       "Private Washrooms", False),
            ("public_washroom",        "Public Washrooms", False),
            ("flooring_type",          "Flooring Type", False),
        ]),
        ("Pricing Details", [
            ("monthly_rent",            "Monthly Rent", True),
            ("brokerage_percentage",    "Brokerage", True),
            ("manual_brokerage",        "Enter Fixed Brokerage", False),
            ("advanced_rent_type",      "Advanced Rent Month", True),
            ("advanced_rent_amount",    "Advance Rent Amount", False),
            ("security_deposit_type",   "Refundable Security Deposit", True),
            ("security_deposit_amount", "Refundable Security Deposit Amount", False),
            ("maintenance_type",        "Maintenance Type", False),
            ("maintenance_charges",     "Monthly Maintenance Amount", False),
            ("total_move_in_cost",      "Total Move In Cost", False),
            ("negotiable",              "Negotiable", False),
            ("lockin_period",           "Lock-in Period (months)", False),
            ("rent_increase",           "Rent Increase (%/year)", False),
        ]),
        ("Amenities & Facilities", [
            ("amenities",         "Amenities", False),
            ("nearby_facilities", "Nearby Facilities", False),
        ]),
        ("Property Descriptions Details", [
            ("user_description", "Property Description", False),
        ]),
        ("Property Listed Elsewhere Details", [
            ("listed_elsewhere", "Is Property Already Listed Elsewhere?", False),
            ("portal_name",      "Portal Name", False),
        ]),
    ])

    field_to_label = {}
    label_to_field = {}
    for sec_fields in sections.values():
        for fkey, flbl, freq in sec_fields:
            field_to_label[fkey] = flbl
            label_to_field[flbl.strip().lower()] = fkey

    system_injected = {"property_title", "property_summary", "property_description"}
    helper_only_labels = {"brokerage_percentage"}

    decimal_fields = {"builtup_area", "carpet_area"}
    int_fields = {
        "property_age", "total_floors", "staircases", "passenger_lifts",
        "service_lifts", "private_parking", "min_seats", "max_seats", "cabins",
        "meeting_rooms", "private_washroom", "public_washroom", "monthly_rent",
        "advanced_rent_amount", "security_deposit_amount", "maintenance_charges",
        "total_move_in_cost", "lockin_period", "pincode",
    }
    float_fields = {"rent_increase"}
    bool_fields = {"dg_ups_included", "electricity_included", "water_included"}

    return (
        sections, field_to_label, label_to_field, system_injected,
        helper_only_labels, decimal_fields, int_fields, float_fields, bool_fields
    )




def _commercial_sample_row_data():
    """
    Returns realistic sample data for Row 4 matching the exact fields.
    """
    return {
        "listed_by_type": "self/other",
        "listed_by_id": "EMP-102",
        "listed_by_name": "Rajesh Sharma",
        "listed_by_email": "rajesh.sharma@estateflow.com",
        "listed_by_contact": "9876543210",
        "listed_by_role": "Agent",
        "property_type": "office_space",
        "building_name": "Alpha Business Tower",
        "wing_number": "A",
        "property_no": "Office 402",
        "availability_status": "available_immediately",
        "available_from": "2026-08-01",
        "property_age": "3",
        "zone_type": "commercial",
        "ownership_type": "Freehold",
        "property_condition": "Ready to Move",
        "builtup_area": "1500",
        "carpet_area": "1200",
       
        "address": "402, Alpha Tower, MG Road, Near Metro Station",
        "locality": "Viman Nagar",
        "property_landmark": "Opposite Phoenix Mall",
        "city": "Pune",
        "state": "Maharashtra",
        "location_hub": "business_district",
        "pincode": "411014",  # Matches new Pincode field
        "google_maps_link": "https://maps.google.com/?q=18.5679,73.9143",
        "latitude": "18.5679",
        "longitude": "73.9143",
        "dg_ups_included": "true",
        "electricity_included": "false",
        "water_included": "true",
        "building_configuration": "G+4",
        "total_floors": "4",
        "staircases": "2",
        "passenger_lifts": "2",
        "service_lifts": "1",
        "private_parking": "2",
        "min_seats": "15",
        "max_seats": "35",
        "cabins": "2",
        "meeting_rooms": "1",
        "private_washroom": "1",
        "public_washroom": "2",
        "flooring_type": "vitrified",
        "monthly_rent": "85000",
        "brokerage_percentage": "1 Month Rent",
        "manual_brokerage": "",
        "advanced_rent_type": "1 Month Rent",
        "advanced_rent_amount": "85000",
        "security_deposit_type": "3 Months Rent",
        "security_deposit_amount": "255000",
        "maintenance_type": "Included",
        "maintenance_charges": "0",
        "total_move_in_cost": "340000",
        "negotiable": "Yes",
        "lockin_period": "11",
        "rent_increase": "5",
        "amenities": "Central AC, Power Backup, Cafeteria, CCTV, 24x7 Security",
        "nearby_facilities": "Metro Station, ATM, Restaurants, Bus Stop",
        "user_description": "Prime office space available in a grade-A commercial building with excellent connectivity.",
        "listed_elsewhere": "No",
        "portal_name": "",
    }


@require_GET
def download_commercial_rental12_template(request):
    """
    Download the upload template for Commercial Rental properties.
    Column headers use exact human-readable labels matching commercial.html.
    Includes a live 'brokerage label preview' formula at the end of the sheet.

    Row 4 (sample data) is LOCKED — visible for reference only, cannot 
    be edited or deleted. Rows 5+ are unlocked for actual data entry.
    """
    sections, *_ = _commercial_rental_field_map()
    sample = _commercial_sample_row_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commercial Rental"
    ws.views.sheetView[0].showGridLines = True

    # Palette matching Residential template
    HDR_BG, REQ_BG, OPT_BG, SAMP_BG = "667EEA", "FEF3C7", "F0FDF4", "ECFDF5"
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    col = 1
    role_col = None
    brokerage_col = None

    for section, fields in sections.items():
        start_col = col
        for field, label, required in fields:
            header_text = label + (" *" if required else "")
            
            # Row 2: Human-readable exact Form Labels
            c1 = ws.cell(row=2, column=col, value=header_text)
            c1.font = Font(bold=True, color="1E293B", name="Arial", size=9)
            c1.fill = PatternFill("solid", fgColor=REQ_BG if required else OPT_BG)
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = bdr

            # Row 4: Locked Read-Only Sample Data
            sc = ws.cell(row=4, column=col, value=sample.get(field, ""))
            sc.font = Font(name="Arial", size=9, color="065F46")
            sc.fill = PatternFill("solid", fgColor=SAMP_BG)
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sc.border = bdr
            sc.protection = Protection(locked=True)  # Read-only sample row

            ws.column_dimensions[get_column_letter(col)].width = max(18, len(label) // 2 + 6)

            if field == "listed_by_role":
                role_col = col
            if field == "brokerage_percentage":
                brokerage_col = col
            col += 1

        # Row 1: Merged Section Headers
        end_col = col - 1
        hc = ws.cell(row=1, column=start_col, value=f"\U0001F4CB {section}")
        hc.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hc.fill = PatternFill("solid", fgColor=HDR_BG)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = PatternFill("solid", fgColor=HDR_BG)
            ws.cell(row=1, column=cc).border = bdr
            
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    # ---- Unlock Data-Entry Rows (Rows 5 to 505) ----
    total_cols = col - 1
    unlocked = Protection(locked=False)
    MAX_DATA_ROWS = 500  
    for r in range(5, 5 + MAX_DATA_ROWS):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).protection = unlocked

    # ---- Live Brokerage Label Preview Column ----
    preview_col = col
    pc = ws.cell(row=2, column=preview_col, value="Brokerage Label Preview (auto)")
    pc.font = Font(bold=True, color="92400E", name="Arial", size=9)
    pc.fill = PatternFill("solid", fgColor="FEF3C7")
    pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pc.border = bdr
    ws.column_dimensions[get_column_letter(preview_col)].width = 28

    if role_col:
        role_letter = get_column_letter(role_col)
        formula = (
            f"=IFERROR(INDEX('Notes - Brokerage Label'!$B$4:$B$9,"
            f"MATCH(LOWER(TRIM({role_letter}4)),'Notes - Brokerage Label'!$C$4:$C$9,0)),\"Brokerage\")"
        )
        fcell = ws.cell(row=4, column=preview_col, value=formula)
        fcell.fill = PatternFill("solid", fgColor="FEF3C7")
        fcell.font = Font(bold=True, color="92400E", name="Arial", size=9)
        fcell.alignment = Alignment(horizontal="center", vertical="center")
        fcell.border = bdr
        fcell.protection = Protection(locked=True)

    if brokerage_col:
        note = (
            "The LABEL text shown above this field on the live form changes based on\n"
            "the 'Listed By Role' in the same row. The stored value/column never changes.\n\n"
            "Admin -> EstateFlow Service Fee\n"
            "Relationship Manager -> Service Fee\n"
            "Landlord -> Tenant Service Fee\n"
            "Agent -> Brokerage\n"
            "Agency/Builder or Builder -> Service Fee\n"
            "Any other role -> Brokerage (default)\n\n"
            "See 'Notes - Brokerage Label' sheet, and the live preview column at the end of this sheet."
        )
        ws.cell(row=2, column=brokerage_col).comment = Comment(note, "System")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    # ---- Notes Sheet (Lookup Table) ----
    notes = wb.create_sheet("Notes - Brokerage Label")
    notes.column_dimensions['A'].width = 26
    notes.column_dimensions['B'].width = 26
    notes.column_dimensions['C'].width = 4
    notes.sheet_view.showGridLines = False

    t = notes["A1"]
    notes.merge_cells("A1:B1")
    t.value = "Brokerage label — driven by Listed By Role"
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["Listed By Role", "Label shown on form"]
    for i, h in enumerate(hdrs, start=1):
        c = notes.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="1E293B", name="Arial")
        c.fill = PatternFill("solid", fgColor=OPT_BG)
        c.border = bdr

    role_rows = [
        ("Admin", "EstateFlow Service Fee"),
        ("Relationship Manager", "Service Fee"),
        ("Landlord", "Tenant Service Fee"),
        ("Agent", "Brokerage"),
        ("Agency/Builder", "Service Fee"),
        ("Builder", "Service Fee"),
    ]
    for r, (role, label) in enumerate(role_rows, start=4):
        notes.cell(row=r, column=1, value=role).border = bdr
        notes.cell(row=r, column=2, value=label).border = bdr
        notes.cell(row=r, column=3, value=f"=LOWER(TRIM(A{r}))")

    # Protect the Sheet (Locks Row 4 and layout, leaves Rows 5+ editable)
    ws.protection.sheet = True
    ws.protection.formatColumns = True
    ws.protection.formatRows = True
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.autoFilter = False
    ws.protection.sort = False

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Commercial_Rental_Template.xlsx"'
    wb.save(response)
    return response







@csrf_exempt
@require_POST
def import_commercial_rental_excel(request):
    excel_file = request.FILES.get("commercial_file") or request.FILES.get("file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx format is supported."}, status=400)

    # 1. Get Field Mappings
    (
        sections, field_to_label, label_to_field, system_injected,
        helper_only_labels, decimal_fields, int_fields, float_fields, bool_fields
    ) = _commercial_rental_field_map()

    # -------------------------------------------------------------------
    # REQUIRED FIELDS - Must match the exact keys defined in the dictionary
    # returned by _commercial_rental_field_map()
    # -------------------------------------------------------------------
    REQUIRED_FIELD_KEYS = [
        'property_type',
        'building_name',
        'property_no',
        'availability_status',
        'property_age',
        'ownership_type',
        'builtup_area',
        'brokerage_percentage',
        'address',
        'locality',
        'city',
        'state',
        'pincode',
        'monthly_rent',
        'advanced_rent_type',
        'security_deposit_type',
        'listed_by_role'
    ]

    def _field_label(field):
        return field_to_label.get(field) or field.replace('_', ' ').title()

    def _is_missing(val):
        """Treat None / empty-string as missing. Does NOT treat 0 / '0' as missing."""
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    # ---- 2. Uploader Identity ----
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    admin_obj = Admin_Login.objects.filter(id=admin_id).first() if admin_id else None
    user_obj = User_Details.objects.filter(id=user_id).first() if user_id else None

    uploader_name = uploader_email = uploader_contact = ""
    uploader_role = "Automated Engine"
    user_identity = "Automated Engine"

    if admin_obj:
        uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
        uploader_email = getattr(admin_obj, 'email', '')
        uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
        uploader_role = "Admin"
        user_identity = uploader_email or uploader_name
    elif user_obj:
        uploader_name = user_obj.user_name
        uploader_email = user_obj.user_email
        uploader_contact = user_obj.user_phone
        uploader_role = "User"
        user_identity = uploader_email or uploader_name

    # ---- 3. Parse Excel ----
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        # Handle fallback if sheet name was changed by user
        ws = wb["Commercial Rental"] if "Commercial Rental" in wb.sheetnames else wb.active
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return JsonResponse({"status": "error", "message": "File appears empty or is missing header rows."}, status=400)

    # Locate actual Header row
    header_row_idx = None
    headers = []
    
    # We scan the first 10 rows to find the one that contains our specific column names
    for idx, r in enumerate(rows[:10]):
        cleaned = [str(c).strip().lower() for c in r if c is not None]
        matches = sum(1 for h in cleaned if h in label_to_field)
        # If we hit 4 or more known column headers, we assume this is the header row
        if matches >= 4:
            header_row_idx = idx
            headers = [str(c).strip() if c is not None else "" for c in r]
            break

    if header_row_idx is None:
        return JsonResponse({
            "status": "error",
            "message": "No recognizable column headers were found. Please use the official template."
        }, status=400)

    # Map file headers to model fields
    field_headers = []
    unmatched_headers = []
    for h in headers:
        # We replace the literal " *" that might exist in the header text, 
        # then strip whitespace to match the dictionary correctly
        norm = str(h).replace(" *", "").strip().lower()
        if not norm or norm in helper_only_labels:
            field_headers.append(None)
            continue
        
        field = label_to_field.get(norm)
        field_headers.append(field)
        if field is None:
            unmatched_headers.append(h)

    data_start_row = header_row_idx + 1
    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []       
    listed_by_mismatch_errors = []   
    skipped_listed_by_mismatch = 0   

    # Extract rows
    for r_idx in range(data_start_row, len(rows)):
        row_data = rows[r_idx]
        
        # Skip purely empty rows or instructional sample rows
        if not any(row_data):
            continue
        first_str = str(row_data[0] or "").strip().lower()
        if "red =" in first_str or "sample" in first_str or "headers =" in first_str:
            continue

        obj_data = {}
        for c_idx, field in enumerate(field_headers):
            if not field or field in system_injected:
                continue
            val = row_data[c_idx] if c_idx < len(row_data) else None
            if val is not None and str(val).strip() != "":
                obj_data[field] = val

        if not obj_data:
            skipped_empty_after_mapping += 1
            continue

        # ---- Type Coercion ----
        if 'available_from' in obj_data:
            d_val = obj_data['available_from']
            if isinstance(d_val, str):
                c_str = d_val.strip().split(" ")[0]
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        obj_data['available_from'] = datetime.strptime(c_str, fmt).date()
                        break
                    except ValueError:
                        obj_data['available_from'] = None
            elif isinstance(d_val, datetime):
                obj_data['available_from'] = d_val.date()

        for f in int_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = int(float(str(obj_data[f]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    obj_data[f] = None

        for f in decimal_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = Decimal(str(obj_data[f]).replace(",", "").strip())
                except (InvalidOperation, ValueError):
                    obj_data[f] = None
                    
        for f in bool_fields:
            if f in obj_data and obj_data[f] is not None:
                s = str(obj_data[f]).strip().lower()
                obj_data[f] = s in ("true", "1", "yes", "y", "included")

        # ---- REQUIRED-FIELD VALIDATION ----
        row_errors = []
        
        # Check if any required fields are missing
        missing_fields = [
            _field_label(f) for f in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(f))
        ]
        
        if missing_fields:
            required_field_errors.append({
                "row": r_idx + 1,
                "missing_fields": missing_fields,
            })
            continue # Don't process further if missing core fields

        # ---- ADMIN/USER 'LISTED BY' VALIDATION ----
        l_role = str(obj_data.get('listed_by_role', '')).strip().title()
        l_email = str(obj_data.get('listed_by_email', '')).strip().lower()
        l_contact = str(obj_data.get('listed_by_contact', '')).strip()
        l_name = str(obj_data.get('listed_by_name', '')).strip()
        l_id = str(obj_data.get('listed_by_id', '')).strip()

        assigned_to = ""
        if l_email or l_contact or l_name or l_id:
            is_registered = False

            if l_role.lower() == 'admin':        
                admin_query = Q()
                if l_email: admin_query &= Q(email=l_email)
                if l_contact: admin_query &= Q(phone=l_contact)
                if l_name: admin_query &= Q(name__iexact=l_name)
                if l_id and l_id.isdigit(): admin_query &= Q(id=l_id)
                
                if admin_query:
                    is_registered = Admin_Login.objects.filter(admin_query).exists()

            else:
                user_query = Q()
                if l_email: user_query &= Q(user_email=l_email)
                if l_contact: user_query &= Q(user_phone=l_contact)
                if l_name: user_query &= Q(user_name__iexact=l_name)
                if l_id: user_query &= Q(user_id=l_id)

                if user_query:
                    if l_role:
                        matched_user = User_Details.objects.filter(user_query, user_role__iexact=l_role).first()
                    else:
                        matched_user = User_Details.objects.filter(user_query).first()

                    if matched_user:
                        is_registered = True
                        assigned_to = f"{matched_user.id}-{matched_user.user_role}"
                        obj_data['assigned_to'] = assigned_to

            if not is_registered:
                searched_info = filter(None, [l_id, l_name, l_email, l_contact, l_role])
                identity = " + ".join(searched_info) or "Unknown"
                row_errors.append(
                    f"Listed By {l_role or 'user'} '{identity}' is not present in our records. "
                    f"Please register this {l_role or 'user'} first, then re-upload this row."
                )

        if row_errors:
            listed_by_mismatch_errors.append({
                "row": r_idx + 1,
                "errors": row_errors,
            })
            skipped_listed_by_mismatch += 1
            continue

        parsed_rows.append({'row_idx': r_idx + 1, 'data': obj_data})

    wb.close()

    # Reject entire file if any mandatory fields are missing
    if required_field_errors:
        
        # Compile a more descriptive error string to help debugging on the UI side
        error_lines = []
        for error_dict in required_field_errors:
            error_lines.append(f"Row {error_dict['row']}: Missing {', '.join(error_dict['missing_fields'])}")
        
        full_error_msg = (
            f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields.\n\n"
            + "\n".join(error_lines)
            + "\n\nPlease fill in every required column (marked with *) for all rows and re-upload the file."
        )

        return JsonResponse({
            "status": "error",
            "message": full_error_msg,
            "row_errors": required_field_errors,
        }, status=400)

    if not parsed_rows and not listed_by_mismatch_errors:
        return JsonResponse({
            "status": "error",
            "message": f"0 usable data rows found. {skipped_empty_after_mapping} row(s) skipped.",
            "unmatched_headers": unmatched_headers,
        }, status=400)

    # ---- 4. Write to DB (Fingerprint Duplicate Engine) ----
    created, updated, skipped, errors = 0, 0, skipped_empty_after_mapping + skipped_listed_by_mismatch, []
    duplicate_blocked_rows = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality', '')).strip()
        input_city = str(o_data.get('city', '')).strip()
        input_pincode = str(o_data.get('pincode', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_name = str(o_data.get('listed_by_name', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        # Generate a fingerprint specifically for grouping identical properties 
        fingerprint_key = f"{input_property_no}|{input_building_name}|{input_locality}|{input_city}|{input_pincode}".lower().replace(" ", "")

        # Direct duplicate lookup
        direct_duplicates = CommercialRentalProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality__iexact=input_locality
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            CommercialRentalProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                # Same Unit AND Same Lister = Block
                same_id = (existing_prop.listed_by_id and input_listed_by_id and existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked - Unit '{input_property_no}' is already listed "
                    f"by/for {input_listed_by_name or 'this user'}. Row skipped; edit the existing listing instead."
                )
                skipped += 1
                continue  

            # Different user listing the exact same physical unit -> allow save & flag
            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id
            )

        # Append final fields and create
        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["listing_type"] = o_data.get("listing_type") or "Rental"
        o_data["category"] = o_data.get("category") or "Commercial"
        o_data["upload_file_name"] = excel_file.name
        o_data["uploaded_by_name"] = uploader_name
        o_data["uploaded_by_email"] = uploader_email
        o_data["uploaded_by_contact"] = uploader_contact
        o_data["uploaded_by_role"] = uploader_role

        try:
            CommercialRentalProperty.objects.create(**o_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    errors.extend(duplicate_blocked_rows)

    for entry in listed_by_mismatch_errors:
        for msg in entry["errors"]:
            errors.append(f"Row {entry['row']}: {msg}")

    # ---- 5. Audit Log ----
    CommercialRentalActivityLog.objects.create(
        user_identity=user_identity,
        user_role=uploader_role,
        action_type='CREATE',
        property_id="Multiple / Sheet Records",
        action_payload=json.dumps({
            "filename": excel_file.name,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "duplicates_blocked": len(duplicate_blocked_rows),
            "listed_by_mismatches": len(listed_by_mismatch_errors),
            "errors_encountered": len(errors),
        }),
        status='SUCCESS' if not errors else 'PARTIAL',
    )

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created, "updated": updated, "skipped": skipped,
        "duplicates_blocked": len(duplicate_blocked_rows),
        "listed_by_mismatches": len(listed_by_mismatch_errors),
        "error_count": len(errors), "errors": errors,
        "unmatched_headers": unmatched_headers,
    })


####################################END View Section Rental Commerical download sample & import file################ 




def _bool(val):
    # Converts "Yes", "True", "1" from Excel to Python True
    if str(val).strip().lower() in ['yes', 'true', '1', 'y']:
        return True
    return False




import io
import openpyxl
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
# Ensure your CommercialRentalProperty model is imported
# from .models import CommercialRentalProperty





import hashlib

# --- EXCEL HELPER CONVERTERS ---
def _str(v): return str(v).strip() if v is not None else ""
def _int(v): return int(float(v)) if v else None
def _float(v): return float(v) if v else None
def _bool(v): return str(v).strip().lower() in ['true', 'yes', '1', 'on'] if v else False
def _date(v): 
    if not v: return None
    if hasattr(v, 'date'): return v.date()
    return str(v).strip()[:10]

def _get_cell(row, col_idx):
    if col_idx is None or col_idx >= len(row): return None
    return row[col_idx]

def _load_new_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active
    # Row 2 contains the database field names (labels)
    headers = {str(cell.value).replace(' *', '').strip(): idx for idx, cell in enumerate(ws[2]) if cell.value}
    return wb, ws, headers















from django.db.models import Sum, Count












def commercial_rental_list(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ═══════════════════════════════════════
    # CAPTURE ADVANCED FILTER QUERY PARAMS
    # ═══════════════════════════════════════
    search_query = request.GET.get('search', '').strip()
    prop_type_query = request.GET.get('property_type', '').strip()
    city_query = request.GET.get('city', '').strip()
    zone_query = request.GET.get('zone_type', '').strip()
    possession_query = request.GET.get('possession', '').strip()
    listed_by_query = request.GET.get('listed_by', '').strip()
    budget_query = request.GET.get('budget', '').strip()
    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()

    # ═══════════════════════════════════════
    # DYNAMIC SEARCH DROPDOWN & BUDGET POPULATORS
    # ═══════════════════════════════════════
    unfiltered_base = CommercialRentalProperty.objects.filter(is_deleted=False)
    
    unique_property_types = unfiltered_base.values_list('property_type', flat=True).distinct()
    unique_cities = unfiltered_base.values_list('city', flat=True).distinct()
    unique_zones = unfiltered_base.exclude(zone_type__isnull=True).exclude(zone_type='').values_list('zone_type', flat=True).distinct()
    unique_possession = unfiltered_base.values_list('possession_status', flat=True).distinct()
    unique_roles = unfiltered_base.exclude(uploaded_by_role__isnull=True).exclude(uploaded_by_role='').values_list('uploaded_by_role', flat=True).distinct()

    # Financial Aggregations & Dynamic Budget Tier Generation
    financials = unfiltered_base.aggregate(
        avg=Avg('expected_rent'), max_r=Max('expected_rent'), min_r=Min('expected_rent'),
        total_r=Sum('expected_rent'), deposit_total=Sum('security_deposit'), area_avg=Avg('builtup_area')
    )
    
    max_rent = financials['max_r'] or 0
    min_rent = financials['min_r'] or 0

    dynamic_budgets = []
    if max_rent > 0:
        interval = (max_rent - min_rent) / 4
        if interval == 0:
            dynamic_budgets.append({
                'value': 'tier_1', 
                'label': f"Up to ₹{int(max_rent):,}", 
                'min': 0, 'max': max_rent
            })
        else:
            b1 = min_rent + interval
            b2 = min_rent + (interval * 2)
            b3 = min_rent + (interval * 3)

            dynamic_budgets = [
                {'value': 'tier_1', 'label': f"Under ₹{int(b1):,}", 'min': 0, 'max': b1},
                {'value': 'tier_2', 'label': f"₹{int(b1):,} – ₹{int(b2):,}", 'min': b1, 'max': b2},
                {'value': 'tier_3', 'label': f"₹{int(b2):,} – ₹{int(b3):,}", 'min': b2, 'max': b3},
                {'value': 'tier_4', 'label': f"Above ₹{int(b3):,}", 'min': b3, 'max': max_rent * 10}
            ]

    # Base Active Queryset 
    properties = CommercialRentalProperty.objects.filter(is_deleted=False)

    # ═══════════════════════════════════════
    # EXECUTE DYNAMIC ADVANCED FILTERS
    # ═══════════════════════════════════════
    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(area_locality__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    if prop_type_query and prop_type_query != "All Types":
        properties = properties.filter(property_type__iexact=prop_type_query)

    if city_query and city_query != "All Cities":
        properties = properties.filter(city__iexact=city_query)

    if zone_query and zone_query != "All Zones":
        properties = properties.filter(zone_type__iexact=zone_query)

    if possession_query and possession_query != "All Status":
        properties = properties.filter(possession_status__iexact=possession_query)

    if listed_by_query and listed_by_query != "All Roles":
        properties = properties.filter(uploaded_by_role__iexact=listed_by_query)

    # Dynamic Budget Range Filter Lookup Execution
    if budget_query and budget_query != "All Budgets":
        for bucket in dynamic_budgets:
            if budget_query == bucket['value']:
                properties = properties.filter(expected_rent__gte=bucket['min'], expected_rent__lte=bucket['max'])
                break

    # Created At Date Filter Ranges
    if from_date_str:
        f_date = parse_date(from_date_str)
        if f_date:
            properties = properties.filter(created_at__date__gte=f_date)

    if to_date_str:
        t_date = parse_date(to_date_str)
        if t_date:
            properties = properties.filter(created_at__date__lte=t_date)

    properties = properties.order_by('-id')

    # KPI Dashboards metrics logic
    total_properties = unfiltered_base.count()
    active_listings = unfiltered_base.count()
    occupied_count = unfiltered_base.filter(possession_status__icontains="occupied").count()
    vacant_count = unfiltered_base.filter(possession_status__icontains="ready").count()
    occupancy_rate = round((occupied_count / total_properties * 100), 1) if total_properties > 0 else 0
    vacancy_rate = round((vacant_count / total_properties * 100), 1) if total_properties > 0 else 0

    avg_rent = financials['avg'] or 0
    max_rent = financials['max_r'] or 0
    min_rent = financials['min_r'] or 0
    total_revenue = financials['total_r'] or 0
    total_security_deposit = financials['deposit_total'] or 0
    avg_area = financials['area_avg'] or 0

    ready_to_move_count = vacant_count
    premium_properties_count = unfiltered_base.filter(expected_rent__gte=100000).count()
    affordable_properties_count = unfiltered_base.filter(expected_rent__lt=25000).count()
    short_lease_count = unfiltered_base.filter(lockin_period__lte=6).count()
    long_lease_count = unfiltered_base.filter(lockin_period__gt=11).count()
    with_images_count = unfiltered_base.filter(images__isnull=False).distinct().count()
    with_owner_count = unfiltered_base.exclude(owner_name__isnull=True).exclude(owner_name='').count()
    image_pct = round((with_images_count / total_properties * 100), 1) if total_properties > 0 else 0
    verified_pct = round((with_owner_count / total_properties * 100), 1) if total_properties > 0 else 0

    total_tenants = occupied_count
    collection_rate = 98
    pending_payments = unfiltered_base.filter(possession_status__icontains="dispute").count()
    maintenance_req = unfiltered_base.filter(maintenance_charges__gt=0).count()

    # Chart Serialization Lookups
    pt_qs = unfiltered_base.values('property_type').annotate(count=Count('id')).order_by('-count')
    prop_type_labels_json = json.dumps([item['property_type'].replace('-', ' ').title() for item in pt_qs])
    prop_type_counts_json = json.dumps([item['count'] for item in pt_qs])

    city_qs = unfiltered_base.values('city').annotate(revenue=Sum('expected_rent')).order_by('-revenue')[:6]
    monthly_labels_json = json.dumps([item['city'] for item in city_qs])
    monthly_revenue_json = json.dumps([float(item['revenue'] or 0) for item in city_qs])
    occupancy_json = json.dumps([occupied_count, vacant_count, total_properties - (occupied_count + vacant_count)])

    rent_buckets = [
        ('Under 25k', unfiltered_base.filter(expected_rent__lt=25000).count()),
        ('25k - 1L', unfiltered_base.filter(expected_rent__gte=25000, expected_rent__lte=100000).count()),
        ('1L - 5L', unfiltered_base.filter(expected_rent__gte=100000, expected_rent__lte=500000).count()),
        ('Above 5L', unfiltered_base.filter(expected_rent__gt=500000).count()),
    ]
    rent_range_labels_json = json.dumps([b[0] for b in rent_buckets])
    rent_range_counts_json = json.dumps([b[1] for b in rent_buckets])

    try:
        uploaded_files = unfiltered_base.exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()
    except Exception:
        uploaded_files = []

    paginator = Paginator(properties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin_user/Reports/Rental/commercial_rental_list.html', {
        'admin_obj': admin_obj,
        'page_obj': page_obj,
        
        # Filter Retention Tags State
        'search_query': search_query,
        'prop_type_query': prop_type_query,
        'city_query': city_query,
        'zone_query': zone_query,
        'possession_query': possession_query,
        'listed_by_query': listed_by_query,
        'budget_query': budget_query,
        'from_date': from_date_str,
        'to_date': to_date_str,
        'filtered_count': properties.count(),

        # Dropdown Lists Populators
        'unique_property_types': unique_property_types,
        'unique_cities': unique_cities,
        'unique_zones': unique_zones,
        'unique_possession': unique_possession,
        'unique_roles': unique_roles,
        'uploaded_files': uploaded_files,
        'dynamic_budgets': dynamic_budgets,  # Added Context

        # Metrics & KPI Bindings
        'total_count': total_properties,
        'active_count': active_listings,
        'occupied_count': occupied_count,
        'vacant_count': vacant_count,
        'occupancy_rate': occupancy_rate,
        'vacancy_rate': vacancy_rate,
        'avg_rent': avg_rent,
        'max_rent': max_rent,
        'min_rent': min_rent,
        'total_revenue': total_revenue,
        'total_security_deposit': total_security_deposit,
        'avg_area': avg_area,
        'ready_to_move_count': ready_to_move_count,
        'premium_properties_count': premium_properties_count,
        'affordable_properties_count': affordable_properties_count,
        'short_lease_count': short_lease_count,
        'long_lease_count': long_lease_count,
        'with_images_count': with_images_count,
        'with_owner_count': with_owner_count,
        'image_pct': image_pct,
        'verified_pct': verified_pct,
        'total_tenants': total_tenants,
        'collection_rate': collection_rate,
        'pending_payments': pending_payments,
        'maintenance_req': maintenance_req,

        # Charts Context Variables Serialization
        'prop_type_labels_json': prop_type_labels_json,
        'prop_type_counts_json': prop_type_counts_json,
        'monthly_labels_json': monthly_labels_json,
        'monthly_revenue_json': monthly_revenue_json,
        'occupancy_json': occupancy_json,
        'rent_range_labels_json': rent_range_labels_json,
        'rent_range_counts_json': rent_range_counts_json,
    })














def commercial_bulk_delete(request):
    """Handles Advanced Bulk Deletions for Commercial Properties."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
        
    session_id = request.session.get('Admin_id')
    if not session_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'})

    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        properties = CommercialRentalProperty.objects.all()
        
        if delete_type == 'delete_all':
            count = properties.count()
            properties.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted ALL {count} properties.'})
            
        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(id__in=page_ids)
            count = target_props.count()
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} properties from current page.'})
            
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            # Using created_at for accurate date ranges, change to available_from if needed
            target_props = properties.filter(created_at__range=[from_date, to_date])
            count = target_props.count()
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} properties in date range.'})
            
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)
            count = target_props.count()
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} properties from the last 30 days.'})
            
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)
            count = target_props.count()
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} older properties.'})
            
        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '')
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) | 
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            count = target_props.count()
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} properties uploaded by {uploader}.'})

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            # Replace 'upload_file_name' with your exact database field name for tracking files
            target_props = properties.filter(upload_file_name=file_name) 
            count = target_props.count()
            if count == 0:
                return JsonResponse({'status': 'error', 'message': f'No properties found for file: {file_name}'})
            target_props.delete()
            return JsonResponse({'status': 'success', 'message': f'Successfully deleted {count} properties from {file_name}.'})
            
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

#######################END VIEW SECTION RENTAL COMMERCIAL PROPERTY##################################



###############################START VIEW SECTION OF RENTAL PG_COLIVING PROPERTY###############################






def export_pg_coliving(request):
    """Dedicated view for exporting PG/Coliving properties to CSV or Excel matching the template format and sequence with Sr.No."""
    
    try:
        properties = PGColivingProperty.objects.filter(is_deleted=False).order_by('-created_at')
    except Exception:
        properties = PGColivingProperty.objects.all().order_by('-created_at')

    # ── Apply Search Filters ──
    search_query = request.GET.get('search', '').strip()
    city_query   = request.GET.get('city', '').strip()
    pg_for_query = request.GET.get('pg_for', '').strip()
    from_date    = request.GET.get('from_date', '').strip()
    to_date      = request.GET.get('to_date', '').strip()

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(locality__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(listed_by_name__icontains=search_query)
        )
    if city_query and city_query != 'All Cities':
        properties = properties.filter(city__icontains=city_query)
    if pg_for_query and pg_for_query != 'All Types':
        properties = properties.filter(pg_for__icontains=pg_for_query)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ── EXACT MATCH TO TEMPLATE SEQUENCE, SECTIONS, & FIELD LABELS (+ SR.NO & AUDIT) ──
    EXPORT_COLS = [
        # 📋 Sr.No & Meta
        ("📋 Sr.No & Meta", "sr_no", False, "Sr.No"),

        # 📋 Listed By Details (6)
        ("📋 Listed By Details", "listed_by_type", False, "Listed By Type"),
        ("📋 Listed By Details", "listed_by_id", False, "Listed By Id"),
        ("📋 Listed By Details", "listed_by_name", False, "Listed By Name"),
        ("📋 Listed By Details", "listed_by_email", False, "Listed By Email"),
        ("📋 Listed By Details", "listed_by_contact", False, "Listed By Contact"),
        ("📋 Listed By Details", "listed_by_role", True, "Listed By Role *"),

        # 📋 Basic Information (15)
         
        ("📋 Basic Information", "id", False, "Property ID(System Generated)"),
        ("📋 Basic Information", "property_title", False, "Property Title(System Generated)"),
        ("📋 Basic Information", "building_name", False, "Building / Project / Society Name"),
        ("📋 Basic Information", "property_no", True, "Property No/Unit No *"),
        ("📋 Basic Information", "wing_number", False, "Tower/Wing Number"),
        ("📋 Basic Information", "city", True, "City *"),
        ("📋 Basic Information", "locality", True, "Locality *"),
        ("📋 Basic Information", "address", True, "Property Address *"),
        ("📋 Basic Information", "property_landmark", False, "Property Landmark"),
        ("📋 Basic Information", "pincode", True, "Pincode *"),
        ("📋 Basic Information", "state", True, "State *"),
        ("📋 Basic Information", "google_maps_link", False, "Google Maps Link"),
        ("📋 Basic Information", "latitude", False, "Latitude"),
        ("📋 Basic Information", "longitude", False, "Longitude"),
        ("📋 Basic Information", "pg_for", True, "PG for *"),
        ("📋 Basic Information", "furnishing_status", True, "Furnishing Status *"),
        ("📋 Basic Information", "best_suited_for", False, "Best Suited For"),

        
        ("📋 Basic Information", "meals_available", False, "Meals Available?"),
        ("📋 Basic Information", "meal_offerings", False, "Meal Offerings"),
        ("📋 Basic Information", "meal_speciality", False, "Meal Speciality"),
        ("📋 Basic Information", "notice_period", False, "Notice Period (Days)"),
        ("📋 Basic Information", "lockin_period", False, "Lock-in Period (Days)"),
        ("📋 Basic Information", "minimum_stay", True, "Minimum Stay (Months) *"),
        ("📋 Basic Information", "available_from", True, "Available From *"),
        ("📋 Basic Information", "property_managed_by", False, "Property Managed By"),
        ("📋 Basic Information", "manager_stays", False, "Property Manager Stays at Property?"),
   

        # 📋 Room Details & Pricing (10)
        ("📋 Room Details & Pricing", "room_type", True, "Room Type *"),
        ("📋 Room Details & Pricing", "total_beds", True, "Total Beds In Room *"),
        ("📋 Room Details & Pricing", "monthly_rent", True, "Rent Per Occupant (₹/Bed/Month/Person) *"),
        ("📋 Room Details & Pricing", "brokerage_percentage", True, "Brokerage *"),
        ("📋 Room Details & Pricing", "manual_brokerage", False, "Enter Fixed Brokerage"),
        ("📋 Room Details & Pricing", "advance_rent_month", False, "Advance Rent Month"),
        ("📋 Room Details & Pricing", "advance_rent_amount", False, "Advance Rent Amount"),
        ("📋 Room Details & Pricing", "security_deposit_type", True, "Refundable Room Security_Deposite (₹/Person) *"),
        ("📋 Room Details & Pricing", "security_deposit_amount", False, "Refundable Security Deposit Amount"),
        ("📋 Room Details & Pricing", "maintenance_type", False, "Maintenance Type"),
        ("📋 Room Details & Pricing", "maintenance_amount", False, "Monthly Maintenance Amount"),
        ("📋 Room Details & Pricing", "total_move_in_cost", False, "Estimated Move-In Cost (Per Occupant)"),

        # 📋 PG Regulations (11)
        ("📋 PG Regulations", "opposite_gender_visitors_allowed", False, "Opposite Gender Visitors Allowed?"),
        ("📋 PG Regulations", "visitors_allowed", False, "Visitors Allowed?"),
        ("📋 PG Regulations", "parents_guardians_allowed", False, "Parents/Guardians Allowed?"),
        ("📋 PG Regulations", "entry_24x7_allowed", False, "24x7 Entry Allowed?"),
        ("📋 PG Regulations", "curfew_time", False, "Curfew Time"),
        ("📋 PG Regulations", "smoking_allowed", False, "Smoking Allowed?"),
        ("📋 PG Regulations", "alcohol_consumption_allowed", False, "Alcohol Consumption Allowed?"),
        ("📋 PG Regulations", "couples_allowed", False, "Couples Allowed?"),
        ("📋 PG Regulations", "pets_allowed", False, "Pets Allowed?"),
        ("📋 PG Regulations", "cooking_allowed", False, "Cooking Allowed?"),
        ("📋 PG Regulations", "police_verification_required", False, "Police Verification Required?"),

        # 📋 Amenities & Facilities (2)
        ("📋 Amenities", "amenities", False, "Amenities"),
        ("📋 Nearby Facilities", "nearby_facilities", False, "Nearby Facilities"),

        # 📋 Description (1)
        ("📋 Description", "property_summary", False, "Property Summary(Auto)"),
        ("📋 Description", "property_description", False, "Property Description(Auto)"),
        ("📋 Description", "user_description", False, "Property Description(Added By User)"),

        # 📋 Media & Listing Status (3)
        ("📋 Media & Listing Status", "listed_elsewhere", False, "Is Property Already Listed Elsewhere?"),
        ("📋 Media & Listing Status", "portal_name", False, "Portal Name"),
        

       
        ("⚙️ System Meta", "uploaded_by_name", False, "Uploaded By (Name)"),
        ("⚙️ System Meta", "uploaded_by_email", False, "Uploaded By (Email)"),
        ("⚙️ System Meta", "uploaded_by_contact", False, "Uploaded By (Contact)"),
        ("⚙️ System Meta", "uploaded_by_role", False, "Uploaded By (Role)"),
        ("⚙️ System Meta", "upload_file_name", False, "Data Uplpaded Via"),
        ("⚙️ System Meta", "property_unique_key", False, "Property Unique Key"),
        ("⚙️ System Meta", "duplicate_count", False, "Duplicate Count"),
        ("⚙️ System Meta", "duplicate_group_id", False, "Duplicate Group ID"),
        ("⚙️ System Meta", "is_duplicate", False, "Is Duplicate"),
        ("⚙️ System Meta", "is_deleted", False, "Is Deleted"),
        ("⚙️ System Meta", "deleted_at", False, "Deleted At"),
        ("⚙️ System Meta", "deleted_by", False, "Deleted By"),
        ("⚙️ System Meta", "created_at", False, "Created At"),
        ("⚙️ System Meta", "updated_at", False, "Updated At"),
        ("⚙️ System Meta", "listing_status", False, "Listing Status"),
        ("⚙️ System Meta", "approval_status", False, "Approval Status"),
    ]

    export_format = request.GET.get('format', 'excel')

    # ── EXCEL EXPORT ──
    if export_format == 'excel':
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PG Coliving"

            HDR_BG, META_BG = "1F4E79", "475569"
            REQ_BG, OPT_BG = "FFD7D7", "F8FAFC"
            thin = Side(style="thin", color="BBBBBB")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

            contiguous_sections = []
            for i, (sec, *_) in enumerate(EXPORT_COLS):
                col_idx = i + 1
                if not contiguous_sections or contiguous_sections[-1]['sec'] != sec:
                    contiguous_sections.append({'sec': sec, 'start': col_idx, 'end': col_idx})
                else:
                    contiguous_sections.append({'sec': sec, 'start': contiguous_sections[-1]['start'], 'end': col_idx})
                    contiguous_sections.pop(-2)

            for item in contiguous_sections:
                sec = item['sec']
                start_col = item['start']
                end_col = item['end']

                fill_color = META_BG if "System Meta" in sec else HDR_BG

                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.border = bdr

                top_cell = ws.cell(row=1, column=start_col)
                top_cell.value = sec
                top_cell.font = Font(bold=True, color="FFFFFF", size=10)
                top_cell.alignment = Alignment(horizontal="center", vertical="center")

                if start_col < end_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            for ci, (sec, field, req, hint) in enumerate(EXPORT_COLS, 1):
                lc = ws.cell(row=2, column=ci, value=hint) 
                lc.font, lc.border = Font(bold=True, size=9), bdr
                lc.fill = PatternFill("solid", fgColor=REQ_BG if req else OPT_BG)
                lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(ci)].width = max(18, len(hint) * 0.9)

            for row_idx, prop in enumerate(properties, start=3):
                for col_idx, (_, field_name, _, _) in enumerate(EXPORT_COLS, start=1):
                    if field_name == "sr_no":
                        val = row_idx - 2
                    elif field_name == "upload_file_name":
                        raw_file = getattr(prop, field_name, "")
                        val = raw_file if raw_file else "Web Listing UI Form"
                    else:
                        val = getattr(prop, field_name, "") if hasattr(prop, field_name) else ""
                    
                    try:
                        if val is True: val = "True"
                        elif val is False: val = "False"
                        elif hasattr(val, 'strftime'): 
                            val = val.strftime('%Y-%m-%d %H:%M:%S') if "at" in field_name else val.strftime('%Y-%m-%d')
                        elif isinstance(val, list): val = ", ".join(map(str, val))
                        elif hasattr(val, 'url'): val = val.url if val else ""
                    except ValueError: 
                        val = ""
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                    cell.alignment, cell.border = Alignment(vertical="center", wrap_text=True), bdr

            ws.row_dimensions[1].height, ws.row_dimensions[2].height = 28, 36
            ws.freeze_panes = "A3"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="PG_Coliving_Template_Export.xlsx"'
            return response
            
        except Exception as e:
            return HttpResponse(f"<pre>ERROR GENERATING EXCEL:\n{str(e)}\n{traceback.format_exc()}</pre>", status=500)

    # ── CSV EXPORT ──
    elif export_format == 'csv':
        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="PG_Coliving_Template_Export.csv"'
            writer = csv.writer(response)
            
            writer.writerow([hint for _, _, _, hint in EXPORT_COLS])
            
            for row_idx, prop in enumerate(properties, start=1):
                row_data = []
                for _, field_name, _, _ in EXPORT_COLS:
                    if field_name == "sr_no":
                        val = row_idx
                    elif field_name == "upload_file_name":
                        raw_file = getattr(prop, field_name, "")
                        val = raw_file if raw_file else "Web Listing UI Form"
                    else:
                        val = getattr(prop, field_name, "") if hasattr(prop, field_name) else ""
                    
                    try:
                        if val is True: val = "True"
                        elif val is False: val = "False"
                        elif hasattr(val, 'strftime'): 
                            val = val.strftime('%Y-%m-%d %H:%M:%S') if "at" in field_name else val.strftime('%Y-%m-%d')
                        elif isinstance(val, list): val = ", ".join(map(str, val))
                        elif hasattr(val, 'url'): val = val.url if val else ""
                    except ValueError: 
                        val = ""
                        
                    row_data.append(str(val) if val is not None else "")
                writer.writerow(row_data)
                
            return response
            
        except Exception as e:
            return HttpResponse(f"<pre>ERROR GENERATING CSV:\n{str(e)}\n{traceback.format_exc()}</pre>", status=500)








def pg_list(request):
    session_id = request.session.get('Admin_id')

    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ── READ FILTER STRINGS ──
    search_query   = request.GET.get('search', '').strip()
    pg_for_filter  = request.GET.get('pg_for', '').strip()
    city_filter    = request.GET.get('city', '').strip()
    furnish_filter = request.GET.get('furnish', '').strip()
    meals_filter   = request.GET.get('meals', '').strip()
    sharing_filter = request.GET.get('sharing', '').strip()
    from_date      = request.GET.get('from_date', '').strip()
    to_date        = request.GET.get('to_date', '').strip()

    # ── FILTER SOFT DELETIONS ──
    all_props  = PGColivingProperty.objects.filter(is_deleted=False)
    properties = all_props.order_by('-created_at')

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) | 
            Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) | 
            Q(building_name__icontains=search_query) |
            Q(listed_by_name__icontains=search_query) | 
            Q(listed_by_contact__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    if pg_for_filter:
        properties = properties.filter(pg_for__iexact=pg_for_filter)
    if city_filter:
        properties = properties.filter(city__icontains=city_filter)
    if furnish_filter:
        properties = properties.filter(furnishing_status__iexact=furnish_filter)
    if meals_filter == 'Yes':
        properties = properties.filter(meals_available=True)
    elif meals_filter == 'No':
        properties = properties.filter(meals_available=False)
    if sharing_filter:
        properties = properties.filter(room_type__iexact=sharing_filter)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ── PAGINATION SYSTEM ──
    paginator = Paginator(properties, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    # ── SUMMARY CARD KPI AGGREGATIONS ──
    total_count    = properties.count()
    boys_count     = properties.filter(pg_for__iexact='Boys').count()
    girls_count    = properties.filter(pg_for__iexact='Girls').count()
    coliving_count = properties.filter(pg_for__iexact='Co-living').count()
    total_beds     = properties.aggregate(t=Sum('total_beds'))['t'] or 0
    city_count     = properties.values('city').distinct().count()

    boys_pct     = round((boys_count     / total_count * 100), 1) if total_count else 0
    girls_pct    = round((girls_count    / total_count * 100), 1) if total_count else 0
    coliving_pct = round((coliving_count / total_count * 100), 1) if total_count else 0

    rent_stats = properties.aggregate(
        avg_rent=Avg('monthly_rent'),
        max_rent=Max('monthly_rent'),
        min_rent=Min('monthly_rent'),
        avg_dep=Avg('security_deposit_amount'),
        tot_rev=Sum('monthly_rent'),
        tot_dep=Sum('security_deposit_amount')
    )

    meals_available_count = properties.filter(meals_available=True).count()
    meals_pct     = round((meals_available_count / total_count * 100), 1) if total_count else 0
    furnished_count = properties.filter(furnishing_status__icontains='Fully').count()
    furnished_pct   = round((furnished_count / total_count * 100), 1) if total_count else 0
    single_room_count = properties.filter(room_type__iexact='single').count()
    shared_room_count = properties.filter(room_type__in=['double', 'triple', 'quad']).count()
    anytime_entry   = properties.filter(entry_24x7_allowed=True).count()
    visitors_allowed= properties.filter(visitors_allowed=True).count()
    premium_pg_count= properties.filter(monthly_rent__gte=10000).count()
    budget_pg_count = properties.filter(monthly_rent__lt=5000).count()
    with_owner_count= properties.exclude(Q(listed_by_name__isnull=True) | Q(listed_by_name='')).count()

    try:
        with_images_count = properties.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    uploaded_files = PGColivingProperty.objects.filter(
        is_deleted=False, upload_file_name__isnull=False
    ).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct().order_by('upload_file_name')

    # ── CHARTS ──
    pg_for_qs     = properties.values('pg_for').annotate(c=Count('id')).order_by('-c')
    pg_for_labels = json.dumps([i['pg_for'] or 'Unspecified' for i in pg_for_qs])
    pg_for_data   = json.dumps([i['c'] for i in pg_for_qs])

    rent_buckets      = [('Under ₹3k', 0, 3000), ('₹3k–5k', 3000, 5000), ('₹5k–8k', 5000, 8000), ('₹8k–12k', 8000, 12000), ('Above ₹12k', 12000, 99999999)]
    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data   = json.dumps([properties.filter(monthly_rent__gte=lo, monthly_rent__lt=hi).count() for _, lo, hi in rent_buckets])

    furnish_qs        = properties.values('furnishing_status').annotate(c=Count('id')).order_by('-c')
    furnishing_labels = json.dumps([i['furnishing_status'] or 'Unspecified' for i in furnish_qs])
    furnishing_data   = json.dumps([i['c'] for i in furnish_qs])

    city_qs     = properties.values('city').annotate(c=Count('id')).order_by('-c')[:5]
    city_labels = json.dumps([i['city'] or 'Unspecified' for i in city_qs])
    city_data   = json.dumps([i['c'] for i in city_qs])

    cities = all_props.values_list('city', flat=True).distinct().order_by('city')

    return render(request, 'admin_user/Reports/Rental/pg_list.html', {
        'admin_obj': admin_obj,
        'page_obj': page_obj, 
        'search_query': search_query, 
        'pg_for_filter': pg_for_filter,
        'city_filter': city_filter, 
        'furnish_filter': furnish_filter, 
        'meals_filter': meals_filter,
        'sharing_filter': sharing_filter, 
        'from_date': from_date, 
        'to_date': to_date, 
        'cities': cities,
        'total_count': total_count, 
        'city_count': city_count, 
        'boys_count': boys_count,
        'girls_count': girls_count, 
        'coliving_count': coliving_count, 
        'total_beds': total_beds,
        'boys_pct': boys_pct, 
        'girls_pct': girls_pct, 
        'coliving_pct': coliving_pct,
        'avg_rent': rent_stats['avg_rent'] or 0, 
        'max_rent': rent_stats['max_rent'] or 0,
        'min_rent': rent_stats['min_rent'] or 0, 
        'total_revenue': rent_stats['tot_rev'] or 0,
        'total_deposit': rent_stats['tot_dep'] or 0, 
        'avg_deposit': rent_stats['avg_dep'] or 0,
        'meals_available_count': meals_available_count, 
        'meals_pct': meals_pct,
        'furnished_count': furnished_count, 
        'furnished_pct': furnished_pct,
        'single_room_count': single_room_count, 
        'shared_room_count': shared_room_count,
        'non_veg_allowed': meals_available_count, 
        'with_images_count': with_images_count,
        'anytime_entry': anytime_entry, 
        'visitors_allowed': visitors_allowed,
        'premium_pg_count': premium_pg_count, 
        'budget_pg_count': budget_pg_count,
        'with_owner_count': with_owner_count, 
        'uploaded_files': uploaded_files,
        'pg_for_labels': pg_for_labels, 
        'pg_for_data': pg_for_data,
        'rent_range_labels': rent_range_labels, 
        'rent_range_data': rent_range_data,
        'furnishing_labels': furnishing_labels, 
        'furnishing_data': furnishing_data,
        'city_labels': city_labels, 
        'city_data': city_data,
    })




def pg_reports(request):
    session_id = request.session.get('Admin_id')

    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ── READ FILTER STRINGS ──
    search_query   = request.GET.get('search', '').strip()
    pg_for_filter  = request.GET.get('pg_for', '').strip()
    city_filter    = request.GET.get('city', '').strip()
    furnish_filter = request.GET.get('furnish', '').strip()
    meals_filter   = request.GET.get('meals', '').strip()
    sharing_filter = request.GET.get('sharing', '').strip()
    from_date      = request.GET.get('from_date', '').strip()
    to_date        = request.GET.get('to_date', '').strip()

    # ── FILTER SOFT DELETIONS ──
    all_props  = PGColivingProperty.objects.filter(is_deleted=False)
    properties = all_props.order_by('-created_at')

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query) | Q(city__icontains=search_query) |
            Q(locality__icontains=search_query) | Q(building_name__icontains=search_query) |
            Q(owner_name__icontains=search_query) | Q(contact_number__icontains=search_query) |
            Q(pg_property_id__icontains=search_query)
        )
    if pg_for_filter:
        properties = properties.filter(pg_for__iexact=pg_for_filter)
    if city_filter:
        properties = properties.filter(city__icontains=city_filter)
    if furnish_filter:
        properties = properties.filter(furnishing_type__iexact=furnish_filter)
    if meals_filter == 'Yes':
        properties = properties.filter(meals_available=True)
    elif meals_filter == 'No':
        properties = properties.filter(meals_available=False)
    if sharing_filter:
        properties = properties.filter(rooms__room_type__iexact=sharing_filter).distinct()
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ══════════════════════════════════════════════════════════════
    # EXPORT HELPERS  (shared by both Excel & CSV)
    # ══════════════════════════════════════════════════════════════
    EXPORT_HEADERS = [
        "pg_property_id", "property_title", "city", "building_name",
        "locality", "property_address", "total_beds", "pg_for",
        "furnishing_type", "sharing_type", "best_suited_for",
        # Room 1
        "room_type_1", "room_beds_1", "room_rent_1", "room_deposit_1",
        "room_brokerage_1", "room_brokerage_percent_1",
        "room_manual_brokerage_1", "room_facilities_1",
        # Room 2
        "room_type_2", "room_beds_2", "room_rent_2", "room_deposit_2",
        "room_brokerage_2", "room_brokerage_percent_2",
        "room_manual_brokerage_2", "room_facilities_2",
        # Meals
        "meals_available", "meal_offerings", "meal_speciality",
        # Rules
        "notice_period", "lockin_period", "minimum_stay",
        "available_from", "property_managed_by", "manager_stays",
        "opposite_sex_allowed", "any_time_allowed", "visitors_allowed",
        "guardian_allowed", "drinking_allowed", "smoking_allowed",
        # Amenities
        "property_description", "amenities", "nearby_facilities",
        # Contact
        "owner_name", "contact_number", "email", "alternate_contact",
        # Uploader
        "uploaded_by_name", "uploaded_by_email",
        "uploaded_by_contact", "uploaded_by_role",
        # Audit
        "upload_file_name", "created_at",
    ]

    def _room_slot(rooms_list, idx):
        """Return flat dict for a room slot (0-based idx)."""
        n = idx + 1
        r = rooms_list[idx] if idx < len(rooms_list) else None
        if r:
            return {
                f"room_type_{n}":              r.room_type,
                f"room_beds_{n}":              r.room_beds,
                f"room_rent_{n}":              r.room_rent,
                f"room_deposit_{n}":           r.room_deposit,
                f"room_brokerage_{n}":         r.room_brokerage,
                f"room_brokerage_percent_{n}": r.room_brokerage_percent or "",
                f"room_manual_brokerage_{n}":  r.room_manual_brokerage or "",
                f"room_facilities_{n}":        r.room_facilities or "",
            }
        return {f"room_type_{n}": "", f"room_beds_{n}": "", f"room_rent_{n}": "",
                f"room_deposit_{n}": "", f"room_brokerage_{n}": "",
                f"room_brokerage_percent_{n}": "", f"room_manual_brokerage_{n}": "",
                f"room_facilities_{n}": ""}

    def _build_export_rows(qs):
        """Build list-of-dicts for export from the given queryset."""
        rows = []
        for prop in qs.prefetch_related('rooms'):
            rooms_list = list(prop.rooms.all())
            row = {
                "id":                   prop.id,
                "property_title":       prop.property_title or "",
                "city":                 prop.city,
                "building_name":        prop.building_name or "",
                "locality":             prop.locality,
                "property_address":     prop.property_address,
                "total_beds":           prop.total_beds,
                "pg_for":               prop.pg_for,
                "furnishing_type":      prop.furnishing_type,
                "sharing_type":         prop.sharing_type or "",
                "best_suited_for":      prop.best_suited_for or "",
                **_room_slot(rooms_list, 0),
                **_room_slot(rooms_list, 1),
                "meals_available":      prop.meals_available,
                "meal_offerings":       prop.meal_offerings or "",
                "meal_speciality":      prop.meal_speciality or "",
                "notice_period":        prop.notice_period or "",
                "lockin_period":        prop.lockin_period or "",
                "minimum_stay":         prop.minimum_stay,
                "available_from":       prop.available_from.strftime("%Y-%m-%d") if prop.available_from else "",
                "property_managed_by":  prop.property_managed_by or "",
                "manager_stays":        prop.manager_stays,
                "opposite_sex_allowed": prop.opposite_sex_allowed,
                "any_time_allowed":     prop.any_time_allowed,
                "visitors_allowed":     prop.visitors_allowed,
                "guardian_allowed":     prop.guardian_allowed,
                "drinking_allowed":     prop.drinking_allowed,
                "smoking_allowed":      prop.smoking_allowed,
                "property_description": prop.property_description or "",
                "amenities":            prop.amenities or "",
                "nearby_facilities":    prop.nearby_facilities or "",
                "owner_name":           prop.owner_name,
                "contact_number":       prop.contact_number,
                "email":                prop.email,
                "alternate_contact":    prop.alternate_contact or "",
                "uploaded_by_name":     prop.uploaded_by_name or "",
                "uploaded_by_email":    prop.uploaded_by_email or "",
                "uploaded_by_contact":  prop.uploaded_by_contact or "",
                "uploaded_by_role":     prop.uploaded_by_role or "",
                "upload_file_name":     prop.upload_file_name or "",
                "created_at":           prop.created_at.strftime("%Y-%m-%d %H:%M") if prop.created_at else "",
            }
            rows.append(row)
        return rows

    # ── EXCEL DOWNLOAD ────────────────────────────────────────────────────────
    if request.GET.get('download') == 'excel':
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        rows = _build_export_rows(properties)

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "PG Listings"

        HDR_FILL   = PatternFill("solid", start_color="1F4E79")
        ROOM1_FILL = PatternFill("solid", start_color="2E75B6")
        ROOM2_FILL = PatternFill("solid", start_color="5B9BD5")
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
        COL_FONT   = Font(bold=True, size=9)
        CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 1 — section labels
        section_map = [
            (1,  11, "📋 Basic Info"),
            (12, 19, "🛏 Room 1"),
            (20, 27, "🛏 Room 2"),
            (28, 30, "🍽 Meals"),
            (31, 42, "📏 Rules"),
            (43, 45, "🏷 Amenities"),
            (46, 49, "📞 Contact"),
            (50, 53, "👤 Uploader"),
            (54, 55, "📁 Audit"),
        ]
        for sc, ec, label in section_map:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
            c = ws.cell(row=1, column=sc, value=label)
            c.fill = ROOM1_FILL if "Room 1" in label else ROOM2_FILL if "Room 2" in label else HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER

        # Row 2 — column headers
        ROOM1_CI = set(range(12, 20))
        ROOM2_CI = set(range(20, 28))
        for ci, h in enumerate(EXPORT_HEADERS, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = COL_FONT
            c.alignment = CENTER
            if ci in ROOM1_CI:
                c.fill = PatternFill("solid", start_color="DDEEFF")
            elif ci in ROOM2_CI:
                c.fill = PatternFill("solid", start_color="EEE8FF")
            else:
                c.fill = PatternFill("solid", start_color="F0F4FF")

        # Data rows from row 3
        for row in rows:
            ws.append([row.get(h, "") for h in EXPORT_HEADERS])

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        for ci, h in enumerate(EXPORT_HEADERS, 1):
            cl = get_column_letter(ci)
            if any(k in h for k in ("address", "description", "facilities")):
                ws.column_dimensions[cl].width = 30
            elif any(k in h for k in ("brokerage_percent", "manual_brokerage")):
                ws.column_dimensions[cl].width = 22
            else:
                ws.column_dimensions[cl].width = max(15, len(h) * 0.9)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="pg_listings_export.xlsx"'
        return response

    # ── CSV DOWNLOAD ──────────────────────────────────────────────────────────
    if request.GET.get('download') == 'csv':
        rows = _build_export_rows(properties)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pg_listings_export.csv"'
        writer = csv.writer(response)
        writer.writerow(EXPORT_HEADERS)
        for row in rows:
            writer.writerow([row.get(h, "") for h in EXPORT_HEADERS])
        return response

    # ── PAGINATION SYSTEM ──
    paginator = Paginator(properties, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    # ── SUMMARY CARD KPI AGGREGATIONS ──
    total_count    = properties.count()
    boys_count     = properties.filter(pg_for__iexact='Boys').count()
    girls_count    = properties.filter(pg_for__iexact='Girls').count()
    coliving_count = properties.filter(pg_for__iexact='Co-living').count()
    total_beds     = properties.aggregate(t=Sum('total_beds'))['t'] or 0
    city_count     = properties.values('city').distinct().count()

    boys_pct     = round((boys_count    / total_count * 100), 1) if total_count else 0
    girls_pct    = round((girls_count   / total_count * 100), 1) if total_count else 0
    coliving_pct = round((coliving_count/ total_count * 100), 1) if total_count else 0

    rent_stats = PGRoomDetail.objects.filter(property__in=properties).aggregate(
        avg_rent=Avg('room_rent'), max_rent=Max('room_rent'), min_rent=Min('room_rent'),
        avg_dep=Avg('room_deposit'), tot_rev=Sum('room_rent'), tot_dep=Sum('room_deposit')
    )

    meals_available_count = properties.filter(meals_available=True).count()
    meals_pct     = round((meals_available_count / total_count * 100), 1) if total_count else 0
    furnished_count = properties.filter(furnishing_type__icontains='Fully').count()
    furnished_pct   = round((furnished_count / total_count * 100), 1) if total_count else 0
    single_room_count = properties.filter(rooms__room_type__iexact='single').distinct().count()
    shared_room_count = properties.filter(rooms__room_type__in=['double','triple','quad']).distinct().count()
    anytime_entry   = properties.filter(any_time_allowed=True).count()
    visitors_allowed= properties.filter(visitors_allowed=True).count()
    premium_pg_count= properties.filter(rooms__room_rent__gte=10000).distinct().count()
    budget_pg_count = properties.filter(rooms__room_rent__lt=5000).distinct().count()
    with_owner_count= properties.exclude(owner_name='').count()
    try:
        with_images_count = properties.filter(images__isnull=False).distinct().count()
    except Exception:
        with_images_count = 0

    uploaded_files = PGColivingProperty.objects.filter(
        is_deleted=False, upload_file_name__isnull=False
    ).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct().order_by('upload_file_name')

    # ── CHARTS ──
    pg_for_qs     = properties.values('pg_for').annotate(c=Count('id')).order_by('-c')
    pg_for_labels = json.dumps([i['pg_for'] for i in pg_for_qs])
    pg_for_data   = json.dumps([i['c']      for i in pg_for_qs])

    rent_buckets      = [('Under ₹3k',0,3000),('₹3k–5k',3000,5000),('₹5k–8k',5000,8000),('₹8k–12k',8000,12000),('Above ₹12k',12000,999999)]
    rent_range_labels = json.dumps([b[0] for b in rent_buckets])
    rent_range_data   = json.dumps([properties.filter(rooms__room_rent__gte=lo, rooms__room_rent__lt=hi).distinct().count() for _, lo, hi in rent_buckets])

    furnish_qs       = properties.values('furnishing_type').annotate(c=Count('id')).order_by('-c')
    furnishing_labels= json.dumps([i['furnishing_type'] for i in furnish_qs])
    furnishing_data  = json.dumps([i['c']               for i in furnish_qs])

    city_qs     = properties.values('city').annotate(c=Count('id')).order_by('-c')[:5]
    city_labels = json.dumps([i['city'] for i in city_qs])
    city_data   = json.dumps([i['c']    for i in city_qs])

    cities = all_props.values_list('city', flat=True).distinct().order_by('city')

    return render(request, 'admin_user/Reports/Rental/pg_reports.html', {'admin_obj': admin_obj,
        'page_obj': page_obj, 'search_query': search_query, 'pg_for_filter': pg_for_filter,
        'city_filter': city_filter, 'furnish_filter': furnish_filter, 'meals_filter': meals_filter,
        'sharing_filter': sharing_filter, 'from_date': from_date, 'to_date': to_date, 'cities': cities,
        'total_count': total_count, 'city_count': city_count, 'boys_count': boys_count,
        'girls_count': girls_count, 'coliving_count': coliving_count, 'total_beds': total_beds,
        'boys_pct': boys_pct, 'girls_pct': girls_pct, 'coliving_pct': coliving_pct,
        'avg_rent': rent_stats['avg_rent'] or 0, 'max_rent': rent_stats['max_rent'] or 0,
        'min_rent': rent_stats['min_rent'] or 0, 'total_revenue': rent_stats['tot_rev'] or 0,
        'total_deposit': rent_stats['tot_dep'] or 0, 'avg_deposit': rent_stats['avg_dep'] or 0,
        'meals_available_count': meals_available_count, 'meals_pct': meals_pct,
        'furnished_count': furnished_count, 'furnished_pct': furnished_pct,
        'single_room_count': single_room_count, 'shared_room_count': shared_room_count,
        'non_veg_allowed': meals_available_count, 'with_images_count': with_images_count,
        'anytime_entry': anytime_entry, 'visitors_allowed': visitors_allowed,
        'premium_pg_count': premium_pg_count, 'budget_pg_count': budget_pg_count,
        'with_owner_count': with_owner_count, 'uploaded_files': uploaded_files,
        'pg_for_labels': pg_for_labels, 'pg_for_data': pg_for_data,
        'rent_range_labels': rent_range_labels, 'rent_range_data': rent_range_data,
        'furnishing_labels': furnishing_labels, 'furnishing_data': furnishing_data,
        'city_labels': city_labels, 'city_data': city_data,
    })










# =============================================================================
#  SAFE TYPE HELPERS (same pattern as commercial_rental_add)
# =============================================================================
def _to_int(val, default=0):
    try:
        return int(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _to_int_or_none(val):
    try:
        return int(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _to_bool(val):
    return str(val).strip().lower() in ("true", "1", "yes", "on")


def build_pg_fingerprint(address, locality, city, property_no, monthly_rent):
    """
    Mirrors PGColivingProperty.save()'s own key_source construction exactly,
    so the pre-create duplicate lookup matches what the model would generate.
    """
    key_source = f"{address}|{locality}|{city}|{property_no}|{monthly_rent}"
    return key_source.strip().lower().replace(" ", "")


# =============================================================================
#  ADD PG / CO-LIVING PROPERTY
# =============================================================================
@csrf_exempt
def add_pg(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid Request"})

    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return JsonResponse({"status": "error", "message": "Not authenticated"})

    try:
        # ---------- UPLOADER IDENTIFICATION (supports Admin or User session) ----------
        if admin_id:
            admin_obj = Admin_Login.objects.filter(id=admin_id).first()
            uploader_name = getattr(admin_obj, 'name', '') if admin_obj else ''
            uploader_email = getattr(admin_obj, 'email', '') if admin_obj else ''
            uploader_contact = getattr(admin_obj, 'phone', '') if admin_obj else ''
            uploader_role = getattr(admin_obj, 'role', 'Admin') if admin_obj else 'Admin'
            uploader_id = f"ADMIN_{admin_id}"
        else:
            user_obj = User_Details.objects.filter(id=user_id).first()
            uploader_name = getattr(user_obj, 'user_name', '') if user_obj else ''
            uploader_email = getattr(user_obj, 'user_email', '') if user_obj else ''
            uploader_contact = getattr(user_obj, 'user_phone', '') if user_obj else ''
            uploader_role = getattr(user_obj, 'user_role', 'User') if user_obj else 'User'
            uploader_id = f"USER_{user_id}"

        def get_list(name):
            return ",".join(request.POST.getlist(name))

        # ---------- CATEGORY IMAGES (validate count first, same as before) ----------
        IMAGE_CATEGORY_FIELDS = {
            'exterior': 'property_images_exterior[]',
            'single_room': 'property_images_single_room[]',
            'double_room': 'property_images_double_room[]',
            'multi_room': 'property_images_multi_room[]',
            'lounge': 'property_images_lounge[]',
            'kitchen': 'property_images_kitchen[]',
            'dining': 'property_images_dining[]',
            'washroom': 'property_images_washroom[]',
            'laundry': 'property_images_laundry[]',
            'parking': 'property_images_parking[]',
            'amenities': 'property_images_amenities[]',
            'floor_plan': 'property_images_floor_plan[]',
        }

        total_images = sum(len(request.FILES.getlist(f)) for f in IMAGE_CATEGORY_FIELDS.values())
        if total_images < 3 or total_images > 30:
            return JsonResponse({"status": "error", "message": "Upload minimum 3 and maximum 30 images"})

        # ---------- LISTED BY ----------
        input_listed_by_id = (request.POST.get('listed_by_id') or uploader_id).strip()
        input_listed_by_name = (request.POST.get('listed_by_name') or uploader_name).strip()
        input_listed_by_email = (request.POST.get('listed_by_email') or uploader_email).strip().lower()
        input_listed_by_contact = (request.POST.get('listed_by_contact') or uploader_contact).strip()
        input_listed_by_role = (request.POST.get('listed_by_role') or uploader_role).strip()

        # ---------- CORE LOCATION / IDENTITY FIELDS (used in fingerprint) ----------
        input_property_no = (request.POST.get('property_no') or '').strip()
        input_building_name = (request.POST.get('building_name') or '').strip()
        input_locality = (request.POST.get('locality') or '').strip()
        input_city = (request.POST.get('city') or '').strip()
        input_address = (request.POST.get('address') or '').strip()
        input_monthly_rent = _to_int(request.POST.get('room_rent'))  # form field name -> monthly_rent

        # ---------- DUPLICATE DETECTION (mirrors model.save()'s key + commercial's approach) ----------
        fingerprint_key = build_pg_fingerprint(
            input_address, input_locality, input_city, input_property_no, input_monthly_rent
        )

        direct_duplicates = PGColivingProperty.objects.filter(
            is_deleted=False,
            locality__iexact=input_locality,
        )
        if input_property_no:
            direct_duplicates = direct_duplicates.filter(address__icontains=input_property_no)
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            PGColivingProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and
                           existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and
                              existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and
                                existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    return JsonResponse({
                        "status": "error",
                        "message": (f"Duplicate Blocked: This PG (Unit {input_property_no}) is already "
                                    f"listed by/for {input_listed_by_name or 'this user'}. "
                                    f"Please edit the existing listing instead.")
                    })

            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id,
            )

        # ---------- DATE PARSING ----------
        date_val = request.POST.get('available_from')
        available_from = None
        if date_val:
            try:
                available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
            except ValueError:
                available_from = None

        # ---------- MEALS / PILL-GROUP FIELDS (already comma-joined by the form's JS) ----------
        meals_available = _to_bool(request.POST.get('meals_available'))
        meal_offerings = (request.POST.get('meal_offerings') or '').strip()
        meal_speciality = (request.POST.get('meal_speciality') or '').strip()
        best_suited_for = (request.POST.get('best_suited_for') or '').strip()
        pg_for = (request.POST.get('pg_for') or '').strip()

        with transaction.atomic():

            pg = PGColivingProperty.objects.create(
                property_unique_key=fingerprint_key,
                is_duplicate=is_dup_flag,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id if is_dup_flag else None,

                listing_type="Rental",
                category="PG/Co-living",

                listed_by_type=request.POST.get('listed_by_type'),
                assigned_to=request.POST.get('assigned_to'),
                listed_by_id=input_listed_by_id,
                listed_by_name=input_listed_by_name,
                listed_by_email=input_listed_by_email,
                listed_by_contact=input_listed_by_contact,
                listed_by_role=input_listed_by_role,

                building_name=input_building_name,
                property_no=input_property_no,
                wing_number=request.POST.get('wing_number'),

                city=input_city,
                locality=input_locality,
                address=input_address,
                property_landmark=request.POST.get('property_landmark'),
                state=request.POST.get('state'),
                pincode=request.POST.get('pincode'),
                google_maps_link=request.POST.get('google_maps_link'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),

                pg_for=pg_for,
                furnishing_status=request.POST.get('furnishing_status'),
                best_suited_for=best_suited_for,

                meals_available=meals_available,
                meal_offerings=meal_offerings,
                meal_speciality=meal_speciality,

                notice_period=_to_int_or_none(request.POST.get('notice_period')),
                lockin_period=_to_int_or_none(request.POST.get('lockin_period')),
                minimum_stay=_to_int_or_none(request.POST.get('minimum_stay')) or 1,
                available_from=available_from,
                property_managed_by=request.POST.get('property_managed_by'),
                manager_stays=_to_bool(request.POST.get('manager_stays')),

                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),

                # ---- Step 2: room details, flattened directly onto the property ----
                room_type=request.POST.get('room_type'),
                total_beds=_to_int_or_none(request.POST.get('room_beds')),
                monthly_rent=input_monthly_rent,
                advance_rent_month=request.POST.get('advance_rent_month'),
                advance_rent_amount=_to_int_or_none(request.POST.get('advance_rent_amount')),
                security_deposit_type=request.POST.get('security_deposit_type'),
                security_deposit_amount=_to_int_or_none(request.POST.get('security_deposit_amount')),
                maintenance_type=request.POST.get('maintenance_type'),
                maintenance_amount=_to_int_or_none(request.POST.get('maintenance_amount')),
                # total_move_in_cost is recalculated in model.save() via calculate_move_in_cost()

                # ---- Step 3: regulations ----
                opposite_gender_visitors_allowed="opposite_gender_visitors_allowed" in request.POST,
                visitors_allowed="visitors_allowed" in request.POST,
                parents_guardians_allowed="parents_guardians_allowed" in request.POST,
                entry_24x7_allowed="entry_24x7_allowed" in request.POST,
                curfew_time=request.POST.get('curfew_time') or None,
                smoking_allowed="smoking_allowed" in request.POST,
                alcohol_consumption_allowed="alcohol_consumption_allowed" in request.POST,
                couples_allowed="couples_allowed" in request.POST,
                pets_allowed="pets_allowed" in request.POST,
                cooking_allowed="cooking_allowed" in request.POST,
                police_verification_required="police_verification_required" in request.POST,

                amenities=get_list("amenities[]"),
                nearby_facilities=get_list("nearby_facilities[]"),
                user_description=request.POST.get("user_description"),

                listed_elsewhere=request.POST.get("listed_elsewhere", "No"),
                portal_name=request.POST.get("portal_name"),

                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,

                listing_status="Pending",
                approval_status="Pending",
            )
            # NOTE: property_title, property_summary, property_description and the
            # 7 auto-FAQs are all generated inside PGColivingProperty.save() itself,
            # so nothing needs to be done for those here.

            # ---------- CATEGORY IMAGES ----------
            saved_count = 0
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                for cat_index, img in enumerate(cat_images):
                    if saved_count >= 30:
                        break
                    PGPropertyImage.objects.create(
                        property=pg, image=img, category=category, sequence_order=cat_index
                    )
                    saved_count += 1

           
            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video')
            property_video_link = request.POST.get('property_video_link', '')

            # 1. ALWAYS auto-generate the slideshow (runs every time, unconditionally)
            saved_images = list(PGPropertyImage.objects.filter(property=pg))
            image_paths = []
            for img_obj in saved_images:
                if img_obj.image and hasattr(img_obj.image, 'path') and os.path.exists(img_obj.image.path):
                    image_paths.append(img_obj.image.path)

            if len(image_paths) >= 3:
                try:
                    from .utils.video_generator import generate_property_slideshow
                    out_path = f"pg/videos/auto_{pg.id}.mp4"
                    result = generate_property_slideshow(image_paths, out_path)
                    print("AUTO SLIDESHOW RESULT:", result)
                    if result:
                        PGColivingVideo.objects.create(
                            property=pg,
                            video=out_path,
                            source='auto'
                        )
                except Exception as ve:
                    print("PG VIDEO GEN FAILED:", str(ve))
                    traceback.print_exc()

            # 2. Manual Upload Logic (separate row, source='uploaded')
            if video_option == 'upload' and uploaded_video:
                if uploaded_video.size > 20 * 1024 * 1024:
                    PGColivingVideo.objects.create(
                        property=pg,
                        video=None,
                        source='uploaded'
                    )
                else:
                    PGColivingVideo.objects.create(
                        property=pg,
                        video=uploaded_video,
                        source='uploaded'
                    )

            # 3. RM Assisted Link Logic (separate row, source='rm_assisted')
            elif video_option == 'rm_assisted':
                PGColivingVideo.objects.create(
                    property=pg,
                    video_url=property_video_link if property_video_link else None,
                    source='rm_assisted'
                )

            # ---------- ACTIVITY LOG ----------
            PGColivingActivityLog.objects.create(
                user_identity=uploader_name,
                user_role=uploader_role,
                action_type="CREATE",
                property_id=str(pg.id),
                action_payload=f"PG '{pg.property_title}' created by {uploader_name}",
                status="SUCCESS",
            )

        return JsonResponse({"status": "success", "message": "PG Added Successfully"})

    except Exception as e:
        print("ERROR IN PG ADD:", str(e))
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})



@csrf_exempt
@require_POST
def pg_bulk_delete(request):
    """Handles Dashboard side bulk deletion by transforming actions to secure SOFT DELETES."""
    # Assuming standard session check or your local _get_admin helper logic
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=401)

    try:
        # Gracefully handle missing or empty payloads
        if not request.body:
            return JsonResponse({'status': 'error', 'message': 'Empty request body payload.'}, status=400)
            
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        
        # Target only active, live properties for dashboard actions
        properties = PGColivingProperty.objects.filter(is_deleted=False)
        target_props = properties.none()

        if delete_type == 'delete_all':
            target_props = properties

        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(pg_property_id__in=page_ids)

        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            target_props = properties.filter(created_at__date__range=[from_date, to_date])

        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)

        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)

        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '')
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_contact__icontains=uploader) |
                Q(owner_name__icontains=uploader)
            )

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            target_props = properties.filter(upload_file_name=file_name)
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'}, status=400)

        count = target_props.count()
        if count == 0:
            return JsonResponse({'status': 'success', 'message': 'No matching active records found to process.'})

        # 🚀 FIX: Convert to Soft Delete so records actually land in your Global Recycle Bin
        target_props.update(
            is_deleted=True,
            deleted_at=timezone.now(),
            deleted_by=request.session.get('Admin_username', 'System Admin')
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully moved {count} PG properties to the Recycle Bin.'
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)




###################Start View section Rental PG_COLIVING  download sample & import file#############################



def _pg_coliving_field_map():
    """
    Returns field mappings where the label strings are an EXACT match to the
    human-readable form labels used in pg_coliving.html.

    NOTE: field keys below were corrected to match PGColivingProperty's actual
    model field names (total_beds / monthly_rent), not the HTML input names
    (room_beds / room_rent). This is required because import_pg_excel() calls
    PGColivingProperty.objects.create(**payload) directly with these keys —
    a mismatch here causes a hard TypeError on every import.
    """
    sections = OrderedDict([
        ("Listed By Details", [
            ("listed_by_type",    "Listed By Type", False),
            ("listed_by_id",      "Listed By Id", False),
            ("listed_by_name",    "Listed By Name", False),
            ("listed_by_email",   "Listed By Email", False),
            ("listed_by_contact", "Listed By Contact", False),
            ("listed_by_role",    "Listed By Role", True),  # Drives brokerage label preview
        ]),
        ("Basic Information", [
            ("building_name",     "Building / Project / Society Name", False),
            ("property_no",       "Property No/Unit No", True),
            ("wing_number",       "Tower/Wing Number", False),
            ("city",              "City", True),
            ("locality",          "Locality", True),
            ("address",           "Property Address", True),
            ("property_landmark", "Property Landmark", False),
            ("pincode",           "Pincode", True),
            ("state",             "State", True),
            ("google_maps_link",  "Google Maps Link", False),
            ("latitude",          "Latitude", False),
            ("longitude",         "Longitude", False),
            ("pg_for",            "PG for", True),
            ("furnishing_status", "Furnishing Status", True),
            ("best_suited_for",   "Best Suited For", False),
        ]),
        ("Meals & Stay Rules", [
            ("meals_available",     "Meals Available?", False),
            ("meal_offerings",      "Meal Offerings", False),
            ("meal_speciality",     "Meal Speciality", False),
            ("notice_period",       "Notice Period (Days)", False),
            ("lockin_period",       "Lock-in Period (Days)", False),
            ("minimum_stay",        "Minimum Stay (Months)", True),
            ("available_from",      "Available From", True),
            ("property_managed_by", "Property Managed By", False),
            ("manager_stays",       "Property Manager Stays at Property?", False),
            ("brokerage_percentage","Brokerage", True),
            ("manual_brokerage",    "Enter Fixed Brokerage", False),
        ]),
        ("Room Details & Pricing", [
            ("room_type",               "Room Type", True),
            ("total_beds",               "Total Beds In Room", True),          # renamed from room_beds
            ("monthly_rent",             "Rent Per Occupant (₹/Bed/Month/Person)", True),  # renamed from room_rent
            ("advance_rent_month",      "Advance Rent Month", False),
            ("advance_rent_amount",     "Advance Rent Amount", False),
            ("security_deposit_type",   "Refundable Room Security_Deposite (₹/Person)", True),
            ("security_deposit_amount", "Refundable Security Deposit Amount", False),
            ("maintenance_type",        "Maintenance Type", False),
            ("maintenance_amount",      "Monthly Maintenance Amount", False),
            ("total_move_in_cost",      "Estimated Move-In Cost (Per Occupant)", False),
        ]),
        ("PG Regulations", [
            ("opposite_gender_visitors_allowed", "Opposite Gender Visitors Allowed?", False),
            ("visitors_allowed",                 "Visitors Allowed?", False),
            ("parents_guardians_allowed",        "Parents/Guardians Allowed?", False),
            ("entry_24x7_allowed",               "24x7 Entry Allowed?", False),
            ("curfew_time",                      "Curfew Time", False),
            ("smoking_allowed",                  "Smoking Allowed?", False),
            ("alcohol_consumption_allowed",      "Alcohol Consumption Allowed?", False),
            ("couples_allowed",                  "Couples Allowed?", False),
            ("pets_allowed",                     "Pets Allowed?", False),
            ("cooking_allowed",                  "Cooking Allowed?", False),
            ("police_verification_required",     "Police Verification Required?", False),
        ]),
        ("Amenities & Facilities", [
            ("amenities",         "Amenities", False),
            ("nearby_facilities", "Nearby Facilities", False),
        ]),
        ("Description", [
            ("user_description", "Property Description", False),
        ]),
        ("Media & Listing Status", [
            ("listed_elsewhere", "Is Property Already Listed Elsewhere?", False),
            ("portal_name",      "Portal Name", False),
        ]),
    ])

    field_to_label = {}
    label_to_field = {}
    for sec_fields in sections.values():
        for fkey, flbl, freq in sec_fields:
            field_to_label[fkey] = flbl
            label_to_field[flbl.strip().lower()] = fkey

    system_injected = {"property_title", "property_summary", "property_description"}
    helper_only_labels = {"brokerage_percentage"}

    int_fields = {
        "pincode", "notice_period", "lockin_period", "minimum_stay", "total_beds",
        "monthly_rent", "advance_rent_amount", "security_deposit_amount",
        "maintenance_amount", "total_move_in_cost",
    }
    bool_fields = {
        "meals_available", "manager_stays", "opposite_gender_visitors_allowed",
        "visitors_allowed", "parents_guardians_allowed", "entry_24x7_allowed",
        "smoking_allowed", "alcohol_consumption_allowed", "couples_allowed",
        "pets_allowed", "cooking_allowed", "police_verification_required",
    }

    return (
        sections, field_to_label, label_to_field, system_injected,
        helper_only_labels, int_fields, bool_fields
    )


@csrf_exempt
@require_POST
def import_pg_excel(request):
    excel_file = request.FILES.get("pg_file") or request.FILES.get("file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not excel_file.name.endswith(".xlsx"):
        return JsonResponse({"status": "error", "message": "Only .xlsx format is supported."}, status=400)

    # 1. Get Field Mappings
    (
        sections, field_to_label, label_to_field, system_injected,
        helper_only_labels, int_fields, bool_fields
    ) = _pg_coliving_field_map()

    # -------------------------------------------------------------------
    # REQUIRED FIELDS - Must match the exact keys defined in the dictionary
    # returned by _pg_coliving_field_map()
    # -------------------------------------------------------------------
    REQUIRED_FIELD_KEYS = [
        'listed_by_role',
        'property_no',
        'city',
        'locality',
        'address',
        'pincode',
        'state',
        'pg_for',
        'furnishing_status',
        'minimum_stay',
        'available_from',
        'brokerage_percentage',
        'room_type',
        'total_beds',
        'monthly_rent',
        'security_deposit_type',
    ]

    def _field_label(field):
        return field_to_label.get(field) or field.replace('_', ' ').title()

    def _is_missing(val):
        """Treat None / empty-string as missing. Does NOT treat 0 / '0' as missing."""
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    # ---- 2. Uploader Identity ----
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    admin_obj = Admin_Login.objects.filter(id=admin_id).first() if admin_id else None
    user_obj = User_Details.objects.filter(id=user_id).first() if user_id else None

    uploader_name = uploader_email = uploader_contact = ""
    uploader_role = "Automated Engine"
    user_identity = "Automated Engine"

    if admin_obj:
        uploader_name = getattr(admin_obj, 'name', '') or getattr(admin_obj, 'username', '')
        uploader_email = getattr(admin_obj, 'email', '')
        uploader_contact = getattr(admin_obj, 'phone', '') or getattr(admin_obj, 'mobile', '')
        uploader_role = "Admin"
        user_identity = uploader_email or uploader_name
    elif user_obj:
        uploader_name = user_obj.user_name
        uploader_email = user_obj.user_email
        uploader_contact = user_obj.user_phone
        uploader_role = "User"
        user_identity = uploader_email or uploader_name

    # ---- 3. Parse Excel ----
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        ws = wb["PG Coliving"] if "PG Coliving" in wb.sheetnames else wb.active
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Cannot open file: {e}"}, status=400)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return JsonResponse({"status": "error", "message": "File appears empty or is missing header rows."}, status=400)

    # Locate actual Header row
    header_row_idx = None
    headers = []

    for idx, r in enumerate(rows[:10]):
        cleaned = [str(c).strip().lower() for c in r if c is not None]
        matches = sum(1 for h in cleaned if h in label_to_field)
        if matches >= 4:
            header_row_idx = idx
            headers = [str(c).strip() if c is not None else "" for c in r]
            break

    if header_row_idx is None:
        return JsonResponse({
            "status": "error",
            "message": "No recognizable column headers were found. Please use the official template."
        }, status=400)

    # Map file headers to model fields
    field_headers = []
    unmatched_headers = []
    for h in headers:
        norm = str(h).replace(" *", "").strip().lower()
        if not norm or norm in helper_only_labels:
            field_headers.append(None)
            continue

        field = label_to_field.get(norm)
        field_headers.append(field)
        if field is None:
            unmatched_headers.append(h)

    def _clean_val(val):
        if val is None:
            return ""
        s = str(val).strip()
        return "" if s.lower() in ("none", "null", "n/a", "na") else s

    def _parse_time(val):
        if hasattr(val, "hour"):
            return val
        s = _clean_val(val)
        if not s:
            return None
        for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                continue
        return None

    data_start_row = header_row_idx + 1
    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []
    listed_by_mismatch_errors = []
    skipped_listed_by_mismatch = 0

    # Extract rows
    for r_idx in range(data_start_row, len(rows)):
        row_data = rows[r_idx]

        if not any(row_data):
            continue
        first_str = str(row_data[0] or "").strip().lower()
        if "purple =" in first_str or "sample" in first_str or "headers =" in first_str:
            continue

        obj_data = {}
        for c_idx, field in enumerate(field_headers):
            if not field or field in system_injected:
                continue
            val = row_data[c_idx] if c_idx < len(row_data) else None
            if val is not None and str(val).strip() != "":
                obj_data[field] = val

        if not obj_data:
            skipped_empty_after_mapping += 1
            continue

        # ---- Type Coercion ----
        if 'available_from' in obj_data:
            d_val = obj_data['available_from']
            if isinstance(d_val, str):
                c_str = d_val.strip().split(" ")[0]
                obj_data['available_from'] = None
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        obj_data['available_from'] = datetime.strptime(c_str, fmt).date()
                        break
                    except ValueError:
                        continue
            elif isinstance(d_val, datetime):
                obj_data['available_from'] = d_val.date()

        if 'curfew_time' in obj_data:
            obj_data['curfew_time'] = _parse_time(obj_data['curfew_time'])

        for f in int_fields:
            if f in obj_data and obj_data[f] is not None:
                try:
                    obj_data[f] = int(float(str(obj_data[f]).replace(",", "").strip()))
                except (TypeError, ValueError):
                    obj_data[f] = None

        for f in bool_fields:
            if f in obj_data and obj_data[f] is not None:
                s = str(obj_data[f]).strip().lower()
                obj_data[f] = s in ("true", "1", "yes", "y", "included")

        # ---- REQUIRED-FIELD VALIDATION ----
        missing_fields = [
            _field_label(f) for f in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(f))
        ]

        if missing_fields:
            required_field_errors.append({
                "row": r_idx + 1,
                "missing_fields": missing_fields,
            })
            continue

        # ---- ADMIN/USER 'LISTED BY' VALIDATION ----
        row_errors = []
        l_role = str(obj_data.get('listed_by_role', '')).strip().title()
        l_email = str(obj_data.get('listed_by_email', '')).strip().lower()
        l_contact = str(obj_data.get('listed_by_contact', '')).strip()
        l_name = str(obj_data.get('listed_by_name', '')).strip()
        l_id = str(obj_data.get('listed_by_id', '')).strip()

        assigned_to = ""
        if l_email or l_contact or l_name or l_id:
            is_registered = False

            if l_role.lower() == 'admin':
                admin_query = Q()
                if l_email: admin_query &= Q(email=l_email)
                if l_contact: admin_query &= Q(phone=l_contact)
                if l_name: admin_query &= Q(name__iexact=l_name)
                if l_id and l_id.isdigit(): admin_query &= Q(id=l_id)

                if admin_query:
                    is_registered = Admin_Login.objects.filter(admin_query).exists()

            else:
                user_query = Q()
                if l_email: user_query &= Q(user_email=l_email)
                if l_contact: user_query &= Q(user_phone=l_contact)
                if l_name: user_query &= Q(user_name__iexact=l_name)
                if l_id: user_query &= Q(user_id=l_id)

                if user_query:
                    if l_role:
                        matched_user = User_Details.objects.filter(user_query, user_role__iexact=l_role).first()
                    else:
                        matched_user = User_Details.objects.filter(user_query).first()

                    if matched_user:
                        is_registered = True
                        assigned_to = f"{matched_user.id}-{matched_user.user_role}"
                        obj_data['assigned_to'] = assigned_to

            if not is_registered:
                searched_info = filter(None, [l_id, l_name, l_email, l_contact, l_role])
                identity = " + ".join(searched_info) or "Unknown"
                row_errors.append(
                    f"Listed By {l_role or 'user'} '{identity}' is not present in our records. "
                    f"Please register this {l_role or 'user'} first, then re-upload this row."
                )

        if row_errors:
            listed_by_mismatch_errors.append({
                "row": r_idx + 1,
                "errors": row_errors,
            })
            skipped_listed_by_mismatch += 1
            continue

        parsed_rows.append({'row_idx': r_idx + 1, 'data': obj_data})

    wb.close()

    # Reject entire file if any mandatory fields are missing
    if required_field_errors:
        error_lines = []
        for error_dict in required_field_errors:
            error_lines.append(f"Row {error_dict['row']}: Missing {', '.join(error_dict['missing_fields'])}")

        full_error_msg = (
            f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields.\n\n"
            + "\n".join(error_lines)
            + "\n\nPlease fill in every required column (marked with *) for all rows and re-upload the file."
        )

        return JsonResponse({
            "status": "error",
            "message": full_error_msg,
            "row_errors": required_field_errors,
        }, status=400)

    if not parsed_rows and not listed_by_mismatch_errors:
        return JsonResponse({
            "status": "error",
            "message": f"0 usable data rows found. {skipped_empty_after_mapping} row(s) skipped.",
            "unmatched_headers": unmatched_headers,
        }, status=400)

    # ---- 4. Write to DB (Fingerprint Duplicate Engine — mirrors Commercial) ----
    created, updated, skipped, errors = 0, 0, skipped_empty_after_mapping + skipped_listed_by_mismatch, []
    duplicate_blocked_rows = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality', '')).strip()
        input_city = str(o_data.get('city', '')).strip()
        input_pincode = str(o_data.get('pincode', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_name = str(o_data.get('listed_by_name', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        # Fingerprint for grouping identical PG units
        fingerprint_key = f"{input_property_no}|{input_building_name}|{input_locality}|{input_city}|{input_pincode}".lower().replace(" ", "")

        direct_duplicates = PGColivingProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality__iexact=input_locality,
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            PGColivingProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked - Bed/Unit '{input_property_no}' is already listed "
                    f"by/for {input_listed_by_name or 'this user'}. Row skipped; edit the existing listing instead."
                )
                skipped += 1
                continue

            # Different user listing the exact same physical unit -> allow save & flag
            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id,
            )

        # Append final fields and create
        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["listing_type"] = o_data.get("listing_type") or "Rental"
        o_data["category"] = o_data.get("category") or "PG/Co-living"
        o_data["upload_file_name"] = excel_file.name
        o_data["uploaded_by_name"] = uploader_name
        o_data["uploaded_by_email"] = uploader_email
        o_data["uploaded_by_contact"] = uploader_contact
        o_data["uploaded_by_role"] = uploader_role

        try:
            PGColivingProperty.objects.create(**o_data)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    errors.extend(duplicate_blocked_rows)

    for entry in listed_by_mismatch_errors:
        for msg in entry["errors"]:
            errors.append(f"Row {entry['row']}: {msg}")

    # ---- 5. Audit Log ----
    PGColivingActivityLog.objects.create(
        user_identity=user_identity,
        user_role=uploader_role,
        action_type='CREATE',
        property_id="Multiple / Sheet Records",
        action_payload=json.dumps({
            "filename": excel_file.name,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "duplicates_blocked": len(duplicate_blocked_rows),
            "listed_by_mismatches": len(listed_by_mismatch_errors),
            "errors_encountered": len(errors),
        }),
        status='SUCCESS' if not errors else 'PARTIAL',
    )

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created, "updated": updated, "skipped": skipped,
        "duplicates_blocked": len(duplicate_blocked_rows),
        "listed_by_mismatches": len(listed_by_mismatch_errors),
        "error_count": len(errors), "errors": errors,
        "unmatched_headers": unmatched_headers,
    })


@require_GET
def download_pg_template(request):
    (sections, field_to_label, _, _, _, _, _) = _pg_coliving_field_map()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PG Coliving Import"
    ws.views.sheetView[0].showGridLines = True

    REQ_FILL = PatternFill("solid", fgColor="7C3AED")  # Purple theme for PG
    OPT_FILL = PatternFill("solid", fgColor="1E293B")
    SEC_FILL = PatternFill("solid", fgColor="F3E8FF")
    SAMP_FILL = PatternFill("solid", fgColor="ECFDF5")
    WARN_FILL = PatternFill("solid", fgColor="FEF3C7")

    THIN = Side(border_style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    all_cols = []
    for sec_name, fields in sections.items():
        for fkey, flbl, freq in fields:
            all_cols.append((fkey, flbl, freq, sec_name))

    col_idx = 1
    for sec_name, fields in sections.items():
        span = len(fields)
        cell = ws.cell(row=1, column=col_idx, value=sec_name.upper())
        cell.font = Font(bold=True, color="4C1D95", size=9)
        cell.fill = SEC_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + span - 1)
        for i in range(col_idx, col_idx + span):
            ws.cell(row=1, column=i).border = BORDER
        col_idx += span

    hints = {
        "listed_by_role": "admin / agent / owner",
        "pg_for": "Boys / Girls / Co-living",
        "furnishing_status": "fully-furnished / semi-furnished",
        "room_type": "single / double / triple / quad",
        "room_rent": "Numeric monthly cost per bed",
        "minimum_stay": "Months numeric count",
        "meals_available": "true / false",
    }

    samples = {
        "listed_by_type": "self",
        "listed_by_role": "admin",
        "building_name": "Sunrise Scholars Enclave",
        "property_no": "PG Block A-102",
        "city": "Pune",
        "locality": "Hinjewadi Phase 1",
        "address": "Plot 45, IT Park Road, Hinjewadi",
        "state": "Maharashtra",
        "pg_for": "Co-living",
        "furnishing_status": "fully-furnished",
        "best_suited_for": "Working Professionals, Students",
        "meals_available": "true",
        "meal_offerings": "Breakfast, Dinner",
        "minimum_stay": "6",
        "available_from": "2026-07-01",
        "brokerage_percentage": "No Brokerage",
        "room_type": "double",
        "room_beds": "2",
        "room_rent": "11500",
        "security_deposit_type": "2 Months Rent",
        "security_deposit_amount": "23000",
        "entry_24x7_allowed": "true",
        "amenities": "Wi-Fi, Housekeeping, Washing Machine, CCTV",
    }

    for ci, (fkey, flbl, freq, _) in enumerate(all_cols, 1):
        hc = ws.cell(row=2, column=ci, value=flbl)
        hc.font = Font(bold=True, color="FFFFFF", size=9)
        hc.fill = REQ_FILL if freq else OPT_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = BORDER

        hn = ws.cell(row=3, column=ci, value=hints.get(fkey, ""))
        hn.font = Font(italic=True, color="64748B", size=8)
        hn.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hn.border = BORDER

        sc = ws.cell(row=4, column=ci, value=samples.get(fkey, ""))
        sc.font = Font(size=9, color="065F46")
        sc.fill = SAMP_FILL
        sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sc.border = BORDER

        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(flbl) + 4)

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(all_cols))
    wc = ws.cell(row=5, column=1, value="⚠️ Purple Headers = Mandatory | Row 3 = Hints | Row 4 = SAMPLE DATA (Delete before importing)")
    wc.font = Font(bold=True, color="92400E", size=9)
    wc.fill = WARN_FILL
    wc.alignment = Alignment(horizontal="left", vertical="center")
    for i in range(1, len(all_cols) + 1):
        ws.cell(row=5, column=i).border = BORDER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="PG_Coliving_Import_Template.xlsx"'
    wb.save(resp)
    return resp





###################END View section Rental PG_COLIVING  download sample & import file#############################

# ─── Converters ───────────────────────────────────────────────────────────────
def _bool(v):
    if isinstance(v, bool): return v
    return str(v).strip().lower() in ("true", "yes", "1")

def _int(v):
    try: return int(float(str(v).strip()))
    except: return None

def _str(v):
    return str(v).strip() if v not in (None, "") else ""

def _date(v):
    if v is None: return None
    from datetime import date, datetime
    if isinstance(v, (date, datetime)): return v
    from django.utils.dateparse import parse_date
    return parse_date(str(v).strip())


# ─── Column map: keys EXACTLY match lowercased row-2 headers in template ──────
#     (excel_header_lowercase,   model_field,            converter,  is_room_col)
COLUMN_MAP = [
    ("city *",                                                              "city",                 _str,  False),
    ("building / project name",                                             "building_name",        _str,  False),
    ("locality *",                                                          "locality",             _str,  False),
    ("pg name *",                                                           "property_title",       _str,  False),
    ("property address *",                                                  "property_address",     _str,  False),
    ("total beds *",                                                        "total_beds",           _int,  False),
    ("pg for * (boys/girls/co_living)",                                          "pg_for",               _str,  False),
    ("furnishing type * (fully-furnished/semi-furnished/unfurnished)",      "furnishing_type",      _str,  False),
    ("best suited for (students/working professionals/any)",                "best_suited_for",      _str,  False),
    ("property managed by (owner/caretaker)",                               "property_managed_by",  _str,  False),
    ("property manager stays at property? (true/false)",                    "manager_stays",        _bool, False),
    ("notice period (days)",                                                "notice_period",        _int,  False),
    ("lock-in period (days)",                                               "lockin_period",        _int,  False),
    ("minimum stay (months) *",                                             "minimum_stay",         _int,  False),
    ("available from * (yyyy-mm-dd)",                                       "available_from",       _date, False),
    # ── Room 1 columns ────────────────────────────────────────────────────────
    ("room_type_1 *",                                                       "__room_type_1__",      _str,  True),
    ("room_beds_1 *",                                                       "__room_beds_1__",      _int,  True),
    ("room_rent_1 *",                                                       "__room_rent_1__",      _str,  True),
    ("room_deposit_1 *",                                                    "__room_deposit_1__",   _str,  True),
    ("room_brokerage_1",                                                    "__room_brokerage_1__", _str,  True),
    ("room_brokerage_percent_1",                                            "__room_brokerage_percent_1__", _str, True),
    ("room_manual_brokerage_1",                                             "__room_manual_brokerage_1__",  _str, True),
    ("room_facilities_1",                                                   "__room_facilities_1__",_str,  True),
    # ── Room 2 columns ────────────────────────────────────────────────────────
    ("room_type_2",                                                         "__room_type_2__",      _str,  True),
    ("room_beds_2",                                                         "__room_beds_2__",      _int,  True),
    ("room_rent_2",                                                         "__room_rent_2__",      _str,  True),
    ("room_deposit_2",                                                      "__room_deposit_2__",   _str,  True),
    ("room_brokerage_2",                                                    "__room_brokerage_2__", _str,  True),
    ("room_brokerage_percent_2",                                            "__room_brokerage_percent_2__", _str, True),
    ("room_manual_brokerage_2",                                             "__room_manual_brokerage_2__",  _str, True),
    ("room_facilities_2",                                                   "__room_facilities_2__",_str,  True),
    # ── Rest of fields ─────────────────────────────────────────────────────────
    ("meals available? (true/false)",                                       "meals_available",      _bool, False),
    ("meal offerings (breakfast,lunch,dinner)",                             "meal_offerings",       _str,  False),
    ("meal speciality (veg/non-veg/both)",                                  "meal_speciality",      _str,  False),
    ("opposite sex allowed? (true/false)",                                  "opposite_sex_allowed", _bool, False),
    ("any time allowed? (true/false)",                                       "any_time_allowed",     _bool, False),
    ("visitors allowed? (true/false)",                                      "visitors_allowed",     _bool, False),
    ("guardian allowed? (true/false)",                                      "guardian_allowed",     _bool, False),
    ("drinking allowed? (true/false)",                                      "drinking_allowed",     _bool, False),
    ("smoking allowed? (true/false)",                                       "smoking_allowed",      _bool, False),
    ("amenities (wifi,cctv,geyser,...)",                                    "amenities",            _str,  False),
    ("nearby facilities (college,market,...)",                               "nearby_facilities",    _str,  False),
    ("property description",                                                "property_description", _str,  False),
    ("owner name *",                                                        "owner_name",           _str,  False),
    ("contact number *",                                                    "contact_number",       _str,  False),
    ("email *",                                                             "email",                _str,  False),
    ("alternate contact",                                                   "alternate_contact",    _str,  False),
]
 
# Updated REQUIRED_HEADERS — room type 1 is now required
REQUIRED_HEADERS = {
    "city *",
    "locality *",
    "property address *",
    "total beds *",
    "pg for * (boys/girls/co_living)",
    "furnishing type * (fully-furnished/semi-furnished/unfurnished)",
    "minimum stay (months) *",
    "available from * (yyyy-mm-dd)",
    "room_type_1 *",
    "room_beds_1 *",
    "room_rent_1 *",
    "room_deposit_1 *",
    "owner name *",
    "contact number *",
    "email *",
}
 
# Number of room slots supported
ROOM_SLOTS = 2
# Required fields validated on model data dict
REQUIRED_FIELDS = [
    "city", "locality", "property_address",
    "total_beds", "pg_for", "furnishing_type",
    "minimum_stay", "available_from",
    "owner_name", "contact_number", "email",
]




from collections import defaultdict

def pg_edit(request, pk):
    property_id = pk
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)
    pg = get_object_or_404(PGColivingProperty, id=property_id)

    if request.method == "GET":
        # Context data required to render the edit form template safely
        amenities_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all() if 'User_Details' in globals() else []

        # Group existing saved images by category so each tab/panel can show its own photos
        images_by_cat = defaultdict(list)
        for img in PGPropertyImage.objects.filter(property=pg).order_by('sequence_order', 'id'):
            images_by_cat[img.category].append(img)

        IMAGE_CATEGORY_DEFS_RAW = [
            ('exterior', 'Building Exterior', 'Front elevation, entrance, signage — up to 5 images'),
            ('single_room', 'Single Occupancy', 'Single sharing rooms — up to 4 images'),
            ('double_room', 'Double Sharing', 'Double sharing rooms — up to 5 images'),
            ('multi_room', 'Triple / Multi Sharing', 'Triple/quad sharing rooms — up to 5 images'),
            ('lounge', 'Common Lounge', 'Common area, TV lounge — up to 4 images'),
            ('kitchen', 'Shared Kitchen', 'Kitchen, cooking area — up to 3 images'),
            ('dining', 'Dining Area', 'Dining tables, seating — up to 3 images'),
            ('washroom', 'Bathroom / Washroom', 'Shared/attached washrooms — up to 4 images'),
            ('laundry', 'Laundry Area', 'Washing machine, drying area — up to 3 images'),
            ('parking', 'Parking', 'Two/four wheeler parking — up to 3 images'),
            ('amenities', 'Common Amenities', 'Gym, gaming zone, security — up to 5 images'),
            ('floor_plan', 'Floor Plan', 'Room/floor layout diagrams — up to 3 images'),
        ]

        image_category_defs = [
            {'cat': cat, 'label': label, 'subtext': subtext, 'existing': images_by_cat.get(cat, [])}
            for cat, label, subtext in IMAGE_CATEGORY_DEFS_RAW
        ]

        # Existing PG video, if any, so Step 4 can show "Current video source" block
        existing_video = PGColivingVideo.objects.filter(property=pg).order_by('-id').first()

        context = {
            'pg': pg,
            'admin_obj': admin_obj,
            'ameneties_obj': amenities_obj,
            'facilities_obj': facilities_obj,
            'user_obj': user_obj,
            'image_category_defs': image_category_defs,
            'existing_video': existing_video,
            'selected_amenities': [a.strip() for a in (pg.amenities or "").split(",") if a.strip()],
            'selected_facilities': [f.strip() for f in (pg.nearby_facilities or "").split(",") if f.strip()],
        }
        return render(request, 'admin_user/Reports/Rental/pg_edit.html', context)

    # ---------- POST REQUEST PROCESSING (Update Logic) ----------
    try:
        def get_list(name):
            return ",".join(request.POST.getlist(name))

        def _to_int_or_none(val):
            try:
                return int(val) if val != '' and val is not None else None
            except (TypeError, ValueError):
                return None

        def _to_bool(val):
            return val in ["on", "true", "True", "1", True]

        IMAGE_CATEGORY_FIELDS = {
            'exterior': 'property_images_exterior[]', 'single_room': 'property_images_single_room[]',
            'double_room': 'property_images_double_room[]', 'multi_room': 'property_images_multi_room[]',
            'lounge': 'property_images_lounge[]', 'kitchen': 'property_images_kitchen[]',
            'dining': 'property_images_dining[]', 'washroom': 'property_images_washroom[]',
            'laundry': 'property_images_laundry[]', 'parking': 'property_images_parking[]',
            'amenities': 'property_images_amenities[]', 'floor_plan': 'property_images_floor_plan[]',
        }

        with transaction.atomic():
            pg.building_name = (request.POST.get('building_name') or '').strip()
            pg.property_no = (request.POST.get('property_no') or pg.property_no or '').strip()
            pg.city = (request.POST.get('city') or '').strip()
            pg.locality = (request.POST.get('locality') or '').strip()
            pg.address = (request.POST.get('address') or request.POST.get('property_address') or '').strip()
            pg.total_beds = _to_int_or_none(request.POST.get('room_beds') or request.POST.get('total_beds')) or pg.total_beds
            pg.pg_for = (request.POST.get('pg_for') or '').strip()
            pg.furnishing_status = request.POST.get('furnishing_status') or request.POST.get('furnishing_type') or pg.furnishing_status
            pg.best_suited_for = (request.POST.get('best_suited_for') or '').strip()

            meals_available = _to_bool(request.POST.get('meals_available'))
            pg.meals_available = meals_available
            pg.meal_offerings = (request.POST.get('meal_offerings') or '').strip() if meals_available else None
            pg.meal_speciality = (request.POST.get('meal_speciality') or '').strip() if meals_available else None

            pg.notice_period = _to_int_or_none(request.POST.get('notice_period'))
            pg.lockin_period = _to_int_or_none(request.POST.get('lockin_period'))
            pg.minimum_stay = _to_int_or_none(request.POST.get('minimum_stay')) or 1

            date_val = request.POST.get('available_from')
            if date_val:
                try:
                    pg.available_from = datetime.strptime(date_val, "%Y-%m-%d").date()
                except ValueError:
                    pass

            pg.property_managed_by = request.POST.get('property_managed_by', pg.property_managed_by)
            pg.manager_stays = _to_bool(request.POST.get('manager_stays'))

            # Regulations
            pg.opposite_gender_visitors_allowed = "opposite_gender_visitors_allowed" in request.POST or "opposite_sex_allowed" in request.POST
            pg.visitors_allowed = "visitors_allowed" in request.POST
            pg.parents_guardians_allowed = "parents_guardians_allowed" in request.POST or "guardian_allowed" in request.POST
            pg.entry_24x7_allowed = "entry_24x7_allowed" in request.POST or "any_time_allowed" in request.POST
            pg.curfew_time = request.POST.get('curfew_time') or pg.curfew_time
            pg.smoking_allowed = "smoking_allowed" in request.POST
            pg.alcohol_consumption_allowed = "alcohol_consumption_allowed" in request.POST or "drinking_allowed" in request.POST
            pg.couples_allowed = "couples_allowed" in request.POST
            pg.pets_allowed = "pets_allowed" in request.POST
            pg.cooking_allowed = "cooking_allowed" in request.POST
            pg.police_verification_required = "police_verification_required" in request.POST

            incoming_amenities = get_list("amenities[]")
            if incoming_amenities:
                pg.amenities = incoming_amenities
            incoming_facilities = get_list("facilities[]") or get_list("nearby_facilities[]")
            if incoming_facilities:
                pg.nearby_facilities = incoming_facilities

            pg.user_description = request.POST.get("user_description") or request.POST.get("property_description") or pg.user_description
            pg.owner_name = request.POST.get("owner_name", "")
            pg.contact_number = request.POST.get("contact_number", "")
            pg.email = request.POST.get("email", "")
            pg.alternate_contact = request.POST.get("alternate_contact", "")

            pg.uploaded_by_name = admin_obj.name
            pg.uploaded_by_email = admin_obj.email
            pg.uploaded_by_contact = admin_obj.phone
            pg.uploaded_by_role = admin_obj.role

            pg.save()

            # ---------- Handle removal of existing images ----------
            remove_ids = request.POST.getlist('remove_image_ids[]')
            if remove_ids:
                PGPropertyImage.objects.filter(property=pg, id__in=remove_ids).delete()

            # ---------- Category Images Handling (new uploads) ----------
            saved_count = PGPropertyImage.objects.filter(property=pg).count()
            for category, field_name in IMAGE_CATEGORY_FIELDS.items():
                cat_images = request.FILES.getlist(field_name)
                # Continue sequence numbering after existing images in this category
                existing_in_cat = PGPropertyImage.objects.filter(property=pg, category=category).count()
                for offset, img in enumerate(cat_images):
                    if saved_count >= 30:
                        break
                    PGPropertyImage.objects.create(
                        property=pg, image=img, category=category,
                        sequence_order=existing_in_cat + offset
                    )
                    saved_count += 1

            # ---------- Video handling ----------
            video_option = request.POST.get('video_option', 'auto')
            uploaded_video = request.FILES.get('property_video') or request.FILES.get('video')
            property_video_link = request.POST.get('property_video_link', '')
            delete_current_video = request.POST.get('delete_current_video')

            if delete_current_video:
                PGColivingVideo.objects.filter(property=pg).delete()

            saved_images = list(PGPropertyImage.objects.filter(property=pg))
            image_paths = [i.image.path for i in saved_images if i.image and hasattr(i.image, 'path') and os.path.exists(i.image.path)]

            regenerate_slideshow = request.POST.get('regenerate_slideshow')
            has_existing_auto_video = PGColivingVideo.objects.filter(property=pg, source='auto').exists()

            # Regenerate only when explicitly requested, or when there's no auto video yet
            # and the person picked "auto" for the first time on this save.
            should_generate = video_option == 'auto' and len(image_paths) >= 3 and (
                regenerate_slideshow or not has_existing_auto_video
            )

            if should_generate:
                try:
                    from .utils.video_generator import generate_property_slideshow
                    out_path = f"pg/videos/auto_{pg.id}.mp4"
                    result = generate_property_slideshow(image_paths, out_path)
                    if result:
                        PGColivingVideo.objects.update_or_create(
                            property=pg, source='auto', defaults={'video': out_path, 'video_url': None}
                        )
                except Exception as ve:
                    print("PG VIDEO GEN FAILED:", str(ve))

            if video_option == 'upload' and uploaded_video and uploaded_video.size <= 50 * 1024 * 1024:
                PGColivingVideo.objects.update_or_create(
                    property=pg, source='uploaded', defaults={'video': uploaded_video, 'video_url': None}
                )
            elif video_option == 'rm_assisted' and property_video_link:
                PGColivingVideo.objects.update_or_create(
                    property=pg, source='rm_assisted', defaults={'video_url': property_video_link, 'video': None}
                )

        return JsonResponse({
            "status": "success",
            "message": "PG Property Updated Successfully.",
            "redirect_url": reverse('pg_list')
        })

    except Exception as e:
        print("ERROR IN PG EDIT:", str(e))
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": f"Server Error: {str(e)}"})












@require_POST
def pg_coliving_delete(request, pk):

    sid, _ = _get_admin(request)
    if not sid:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized'
        }, status=401)

    try:
        deleter_name = _get_deleter_name(request)

        # ✅ USING NEW PRIMARY KEY FIELD
        pg = get_object_or_404(
            PGColivingProperty,
            pg_property_id=pk
        )

        pg.is_deleted = True
        pg.deleted_at = timezone.now()
        pg.deleted_by = deleter_name
        pg.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Moved to Recycle Bin successfully!'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })



@csrf_exempt
@require_POST
def pg_restore(request, id):
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=401)

    PGColivingProperty.objects.filter(pg_property_id=id).update(
        is_deleted=False,
        deleted_at=None,
        deleted_by=None
    )
    return JsonResponse({'status': 'success', 'message': 'PG Property Restored Successfully!'})

# =========================================================
# HARD DELETE
# =========================================================
@csrf_exempt
@require_POST
def pg_hard_delete(request, id):
    admin_id = request.session.get('Admin_id')
    if not admin_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=401)

    try:
        pg = get_object_or_404(PGColivingProperty, pg_property_id=id)

        # Clean up related asset files safely
        for img in pg.images.all():
            if img.image:
                img.image.delete(save=False)
        if pg.video:
            pg.video.delete(save=False)

        pg.delete()
        return JsonResponse({'status': 'success', 'message': 'Permanently Deleted Successfully!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)





def pg_coliving_view(request, pk):
    # session_id = request.session.get('Admin_id')
    # admin_obj = Admin_Login.objects.get(id=session_id)
    # ^ Replace with your standard auth logic
    
    pg = get_object_or_404(PGColivingProperty, pk=pk)

    # 1. Safely query the new relational Room models
    parsed_rooms = pg.rooms.all()
    
    # Calculate sidebar pricing metrics dynamically
    rent_stats = parsed_rooms.aggregate(min_rent=Min('room_rent'), max_rent=Max('room_rent'))
    starting_rent = rent_stats['min_rent'] or 0

    # 2. Parse Comma-Separated Strings into Lists for HTML "Chips"
    def split_to_list(db_string):
        return [x.strip() for x in db_string.split(',') if x.strip()] if db_string else []

    context = {
        # 'admin_obj': admin_obj,
        'pg': pg,
        'parsed_rooms': parsed_rooms,
        'starting_rent': starting_rent, # Passed to the sidebar
        'pg_for_list': split_to_list(pg.pg_for),
        'sharing_type_list': split_to_list(pg.sharing_type),
        'best_suited_list': split_to_list(pg.best_suited_for),
        'amenities_list': split_to_list(pg.amenities),
        'facilities_list': split_to_list(pg.nearby_facilities),
        'meal_offerings_list': split_to_list(pg.meal_offerings),
        'meal_speciality_list': split_to_list(pg.meal_speciality),
    }

    return render(request, "admin_user/Reports/Rental/pg_coliving_view.html", context)

###############################END VIEW SECTION OF RENTAL PG_COLIVING PROPERTY###############################




###################START VIEW SECTION RESALE PLOT Type LISTING Forms###########################

def residential_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/residential_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')



def commercial_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/commercial_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


def industrial_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/industrial_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')



def agricultural_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/agricultural_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


###################End VIEW SECTION RESALE PLOT Type LISTING Forms###########################



###################START VIEW SECTION RESALE PLOT Type LISTING Reports###########################

def residential_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/residential_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')



def commercial_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/commercial_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')


def industrial_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/industrial_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')



def agricultural_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/agricultural_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')




###################End VIEW SECTION RESALE PLOT Type LISTING Reports###########################



###################START VIEW SECTION RESALE PLOT LISTING###########################




def plot_sale_add(request):
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    try:
        if admin_id:
            admin = Admin_Login.objects.get(id=admin_id)
            uploader_name, uploader_email, uploader_phone, uploader_role = (
                admin.name, admin.email, admin.phone, admin.role
            )
        else:
            user = User_Details.objects.get(id=user_id)
            uploader_name, uploader_email, uploader_phone, uploader_role = (
                user.name, user.email, user.phone, user.role
            )

        if request.method == "POST":
            prop = PlotSaleProperty.objects.create(
                # ── Step 1: Plot Specs ──────────────────────────────────
                plot_title            = request.POST.get('plot_title'),
                plot_area             = request.POST.get('plot_area') or 0,
                resale_plot_type      = request.POST.get('resale_plot_type'),
                plot_road_facing      = request.POST.get('plot_road_facing'),
                corner_plot           = request.POST.get('corner_plot', 'no'),       # ✅ was plot_corner
                sanctioning_authority = request.POST.get('sanctioning_authority'),   # ✅ was plot_authority
                plot_fencing          = request.POST.get('plot_fencing', 'no'),

                # ── Step 2: Pricing & Legal ─────────────────────────────
                plot_price            = request.POST.get('plot_price') or 0,
                brokerage             = request.POST.get('brokerage', 'No'),
                brokerage_percentage  = request.POST.get('brokerage_percentage'),
                ownership_type        = request.POST.get('ownership_type'),          # ✅ was plot_ownership
                loan_on_property      = request.POST.get('loan_on_property', 'no'), # ✅ was plot_loan
                plot_loan_amount      = request.POST.get('plot_loan_amount') or None,

                # ── Step 3: Media ───────────────────────────────────────
                encumbrance_cert      = request.FILES.get('encumbrance_cert'),
                social_video          = request.FILES.get('social_video'),

                property_description  = request.POST.get('property_description'),
                user_description  = request.POST.get('user_description'),
                property_summary  = request.POST.get('property_summary'),

                # ── Step 4: Location & Owner ────────────────────────────
                plot_city             = request.POST.get('plot_city'),
                plot_locality         = request.POST.get('plot_locality'),
                plot_address          = request.POST.get('plot_address'),
                plot_owner_name       = request.POST.get('plot_owner_name'),
                plot_owner_contact    = request.POST.get('plot_owner_contact'),
                plot_owner_email      = request.POST.get('plot_owner_email'),
                plot_owner_role       = request.POST.get('plot_owner_role'),         # ✅ was missing

                # ── Uploader / Audit ────────────────────────────────────
                uploaded_by_name      = uploader_name,
                uploaded_by_email     = uploader_email,
                uploaded_by_contact   = uploader_phone,
                uploaded_by_role      = uploader_role,
            )

            images = request.FILES.getlist('property_images[]')
            for i, img in enumerate(images[:10]):
                PlotSaleImage.objects.create(property=prop, image=img)

            return JsonResponse({"status": "success", "message": "Plot Listing Added Successfully"})

    except Exception as e:
        print("ERROR:", str(e))
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": str(e)})

    admin_obj = Admin_Login.objects.get(id=admin_id) if admin_id else User_Details.objects.get(id=user_id)
    return render(request, 'admin_user/Reports/Resale/plot_add.html', {'admin_obj': admin_obj})






def plot_sale_edit(request, plot_property_id):
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    prop = get_object_or_404(PlotSaleProperty, plot_property_id=plot_property_id)

    if request.method == "POST":
        try:
            # ── Step 1: Plot Specs ──────────────────────────────────
            prop.plot_title            = request.POST.get('plot_title')
            prop.plot_area             = request.POST.get('plot_area') or 0
            prop.resale_plot_type      = request.POST.get('resale_plot_type')
            prop.plot_road_facing      = request.POST.get('plot_road_facing')
            prop.corner_plot           = request.POST.get('corner_plot', 'no')        # ✅ was plot_corner
            prop.sanctioning_authority = request.POST.get('sanctioning_authority')    # ✅ was plot_authority
            prop.plot_fencing          = request.POST.get('plot_fencing', 'no')

            # ── Step 2: Pricing & Legal ─────────────────────────────
            prop.plot_price            = request.POST.get('plot_price') or 0
            prop.brokerage             = request.POST.get('brokerage', 'No')
            prop.brokerage_percentage  = request.POST.get('brokerage_percentage')
            prop.ownership_type        = request.POST.get('ownership_type')           # ✅ was plot_ownership
            prop.loan_on_property      = request.POST.get('loan_on_property', 'no')  # ✅ was plot_loan
            prop.plot_loan_amount      = request.POST.get('plot_loan_amount') or None

            # ── Step 3: Media (only update if new file uploaded) ────
            if request.FILES.get('encumbrance_cert'):
                prop.encumbrance_cert  = request.FILES.get('encumbrance_cert')
            if request.FILES.get('social_video'):
                prop.social_video      = request.FILES.get('social_video')

            # ── Step 4: Location & Owner ────────────────────────────
            prop.plot_city             = request.POST.get('plot_city')
            prop.plot_locality         = request.POST.get('plot_locality')
            prop.plot_address          = request.POST.get('plot_address')
            prop.property_description  = request.POST.get('property_description')
            prop.property_summary      = request.POST.get('property_summary')
            prop.user_description      = request.POST.get('user_description')
            prop.plot_owner_name       = request.POST.get('plot_owner_name')
            prop.plot_owner_contact    = request.POST.get('plot_owner_contact')
            prop.plot_owner_email      = request.POST.get('plot_owner_email')
            prop.plot_owner_role       = request.POST.get('plot_owner_role')          # ✅ was missing

            # Wipe title so save() auto-regenerates it
            prop.property_title = ""
            prop.save()

            # Handle new images (respect 10 image cap)
            new_images        = request.FILES.getlist('property_images[]')
            current_count     = prop.images.count()
            for img in new_images:
                if current_count >= 10:
                    break
                PlotSaleImage.objects.create(property=prop, image=img)
                current_count += 1

            return JsonResponse({"status": "success", "message": "Plot Listing Updated Successfully"})

        except Exception as e:
            print("ERROR:", str(e))
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": str(e)})

    context = {
        'prop': prop,
        'existing_images': prop.images.all()
    }
    return render(request, 'admin_user/Reports/Resale/plot_edit.html', context)

# ── 1. MAIN LIST VIEW ──



#


@require_POST
def plot_sale_bulk_delete(request):
    """
    Bulk Soft Delete Plot Sale Properties
    """

    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized access.'
        })

    try:

        data = json.loads(request.body)

        delete_type = data.get('delete_type')

        deleter_name = _get_deleter_name(request)

        properties = PlotSaleProperty.objects.filter(
            is_deleted=False
        )

        # ====================================
        # Delete All
        # ====================================

        if delete_type == 'delete_all':

            count = properties.count()

            properties.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots moved to recycle bin.'
            })

        # ====================================
        # Current Page
        # ====================================

        elif delete_type == 'current_page':

            page_ids = data.get('page_ids', [])

            target_props = properties.filter(
                plot_property_id__in=page_ids
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots moved to recycle bin.'
            })

        # ====================================
        # Date Range
        # ====================================

        elif delete_type == 'date_range':

            from_date = data.get('from_date')
            to_date = data.get('to_date')

            if not from_date or not to_date:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please select both dates.'
                })

            target_props = properties.filter(
                created_at__date__range=[from_date, to_date]
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots deleted between {from_date} and {to_date}.'
            })

        # ====================================
        # Last 30 Days
        # ====================================

        elif delete_type == 'latest_month':

            thirty_days_ago = timezone.now() - timedelta(days=30)

            target_props = properties.filter(
                created_at__gte=thirty_days_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots from last 30 days moved to recycle bin.'
            })

        # ====================================
        # Older Than 6 Months
        # ====================================

        elif delete_type == 'old_data':

            six_months_ago = timezone.now() - timedelta(days=180)

            target_props = properties.filter(
                created_at__lt=six_months_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} old plots moved to recycle bin.'
            })

        # ====================================
        # By Uploader
        # ====================================

        elif delete_type == 'by_uploader':

            uploader = data.get('uploader_text', '').strip()

            if not uploader:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please enter uploader details.'
                })

            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader) |
                Q(uploaded_by_contact__icontains=uploader)
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots uploaded by "{uploader}" moved to recycle bin.'
            })

        # ====================================
        # By Imported File
        # ====================================

        elif delete_type == 'by_file':

            file_name = data.get('file_name', '').strip()

            if not file_name:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please select file.'
                })

            target_props = properties.filter(
                upload_file_name=file_name
            )

            count = target_props.count()

            if count == 0:
                return JsonResponse({
                    'status': 'error',
                    'message': f'No records found for file "{file_name}".'
                })

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} plots from file "{file_name}" moved to recycle bin.'
            })

        # ====================================
        # Unknown Type
        # ====================================

        return JsonResponse({
            'status': 'error',
            'message': 'Invalid delete type.'
        })

    except Exception as e:

        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })


  



def safe_float(val):
    """Safely converts Excel cell values to float, handling strings, currencies, and spaces."""
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        clean_val = str(val).replace('₹', '').replace(',', '').strip()
        return float(clean_val)
    except (ValueError, TypeError):
        return 0.0





def safe_float(value):
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0





from django.views.decorators.csrf import csrf_protect











# ════════════════════════════════════════════════════════════
#  SHARED COLUMN DEFINITIONS  (single source of truth)
#  Each tuple: (db_field_name, display_label, hint_text)
# ════════════════════════════════════════════════════════════

COLUMNS = [
    # ── Plot Details ──────────────────────────────────────────
    ("property_title",        "Property Title",              "⚠️ AUTO GENERATED - Leave Empty"),
    ("plot_title",            "Plot Title *",                "e.g. Green Valley Plots"),
    ("plot_area",             "Plot Area (sq.ft) *",         "e.g. 1500"),
    ("resale_plot_type",      "Resale Plot Type *",          "residential / commercial / agricultural"),
    ("plot_road_facing",      "Plot Road Facing *",          "main / east / west / north / south"),
    ("corner_plot",           "Corner Plot",                 "yes / no  (default: no)"),
    ("sanctioning_authority", "Sanctioning Authority",       "e.g. NIT / NMRDA / PRIVATE"),
    ("plot_fencing",          "Plot Fencing",                "yes / no  (default: no)"),
    # ── Pricing & Legal ───────────────────────────────────────
    ("plot_price",            "Plot Price (₹) *",            "e.g. 3500000"),
    ("price_per_sqft",        "Price Per Sqft",              "🔄 AUTO CALCULATED - Leave Empty"),
    ("brokerage",             "Brokerage",                   "Yes / No  (default: No)"),
    ("brokerage_percentage",  "Brokerage %",                 "e.g. 1%  or leave blank"),
    ("ownership_type",        "Ownership Type *",            "freehold / leasehold"),
    ("loan_on_property",      "Loan on Property *",          "yes / no  (default: no)"),
    ("plot_loan_amount",      "Plot Loan Amount (₹)",        "e.g. 2000000  (0 if no loan)"),
    # ── Location ──────────────────────────────────────────────
    ("plot_city",             "Plot City *",                 "e.g. Nagpur"),
    ("plot_locality",         "Plot Locality *",             "e.g. Besa"),
    ("plot_address",          "Plot Address *",              "Plot 12, Besa Road, Nagpur"),
    # ── Owner Contact ─────────────────────────────────────────
    ("plot_owner_name",       "Plot Owner Name *",           "Full Name"),
    ("plot_owner_contact",    "Plot Owner Contact *",        "10-digit mobile"),
    ("plot_owner_email",      "Plot Owner Email *",          "email@example.com"),
    ("plot_owner_role",       "Plot Owner Role",             "Owner / Agent / Builder"),
]

# Section header spans for Row 1  (0-indexed start, end inclusive)
SECTIONS = [
    ("📋 Plot Details",     0,  7),
    ("📋 Pricing & Legal",  8, 14),
    ("📋 Location",        15, 17),
    ("📋 Owner Contact",   18, 21),
]

SAMPLE_ROW = [
    "",                             # property_title  — auto
    "SAMPLE - Green Valley Plots",  # plot_title
    1500,                           # plot_area
    "residential",                  # resale_plot_type
    "main",                         # plot_road_facing
    "no",                           # corner_plot
    "NIT",                          # sanctioning_authority
    "yes",                          # plot_fencing
    3500000,                        # plot_price
    "",                             # price_per_sqft — auto
    "No",                           # brokerage
    "",                             # brokerage_percentage
    "freehold",                     # ownership_type
    "no",                           # loan_on_property
    0,                              # plot_loan_amount
    "Nagpur",                       # plot_city
    "Besa",                         # plot_locality
    "SAMPLE - Plot 12, Besa Road",  # plot_address
    "SAMPLE - Amit Patil",          # plot_owner_name
    "9999999999",                   # plot_owner_contact
    "sample@example.com",           # plot_owner_email
    "Agent",                        # plot_owner_role
]

# Column widths (22 total, one per column)
COL_WIDTHS = [28, 22, 18, 20, 18, 12, 22, 12, 16, 16, 12, 16, 18, 18, 20, 15, 18, 34, 20, 20, 26, 15]


# ════════════════════════════════════════════════════════════
#  HELPER — build styled template sheet
# ════════════════════════════════════════════════════════════

def _build_template_sheet(wb):
    """
    Row 1 → Section headers
    Row 2 → DB field names
    Row 3 → Display labels
    Row 4 → Hint text
    Row 5 → SAMPLE row  (red, clearly marked)
    Row 6 → Instruction banner  (blue)
    Row 7 → First empty data row  (blue tint) ← user fills from here
    """
    sheet = wb.active
    sheet.title = "Plot Resale"

    # ── Colour palette ────────────────────────────────────────
    DARK_BG      = "1E293B"
    WHITE        = "FFFFFF"
    MID_BLUE     = "3B82F6"
    LIGHT_BG     = "F8FAFC"
    HINT_BG      = "FEF9C3"
    HINT_FG      = "92400E"
    DATA_BG      = "EFF6FF"
    BORDER_COLOR = "CBD5E1"

    thin  = Side(style="thin",   color=BORDER_COLOR)
    thick = Side(style="medium", color="94A3B8")
    cb    = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    hb    = Border(left=thick, right=thick, top=thick, bottom=thick)

    def hfill(h):
        return PatternFill("solid", fgColor=h)

    # ── ROW 1 — Section headers ───────────────────────────────
    for label, col_start_idx, col_end_idx in SECTIONS:
        sc = col_start_idx + 1
        ec = col_end_idx   + 1
        c = sheet.cell(row=1, column=sc, value=label)
        c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill      = hfill(DARK_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = hb
        if sc != ec:
            sheet.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
    sheet.row_dimensions[1].height = 30

    # ── ROW 2 — DB field names ────────────────────────────────
    for i, (db, _, _) in enumerate(COLUMNS, 1):
        c = sheet.cell(row=2, column=i, value=db)
        c.font      = Font(name="Arial", bold=True, size=9, color="475569")
        c.fill      = hfill("E2E8F0")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = cb
    sheet.row_dimensions[2].height = 22

    # ── ROW 3 — Display labels ────────────────────────────────
    for i, (_, disp, _) in enumerate(COLUMNS, 1):
        colour = "C0392B" if disp.endswith("*") else MID_BLUE
        c = sheet.cell(row=3, column=i, value=disp)
        c.font      = Font(name="Arial", bold=True, size=10, color=colour)
        c.fill      = hfill(LIGHT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = cb
    sheet.row_dimensions[3].height = 36

    # ── ROW 4 — Hints ─────────────────────────────────────────
    for i, (_, _, hint) in enumerate(COLUMNS, 1):
        c = sheet.cell(row=4, column=i, value=hint)
        c.font      = Font(name="Arial", italic=True, size=8, color=HINT_FG)
        c.fill      = hfill(HINT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = cb
    sheet.row_dimensions[4].height = 30

    # ── ROW 5 — Sample data ───────────────────────────────────
    c1 = sheet.cell(row=5, column=1, value="🔴 DELETE THIS ROW BEFORE IMPORT 🔴")
    c1.font      = Font(name="Arial", bold=True, size=10, color="FF0000")
    c1.fill      = hfill("FFE5E5")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border    = cb

    for i, val in enumerate(SAMPLE_ROW, 1):
        if i == 1:
            continue
        c = sheet.cell(row=5, column=i, value=val)
        c.font      = Font(name="Arial", size=9, color="999999")
        c.fill      = hfill("FFF3F3")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = cb
    sheet.row_dimensions[5].height = 25

    # ── ROW 6 — Instruction banner ────────────────────────────
    #    (moved ABOVE the first data row so import reads row 7+)
    instr = sheet.cell(row=6, column=1, value="👇 START YOUR DATA FROM ROW 7 ONWARDS 👇")
    instr.font      = Font(name="Arial", bold=True, size=11, color="0066CC")
    instr.fill      = hfill("E5F3FF")
    instr.alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(COLUMNS))
    sheet.row_dimensions[6].height = 25

    # ── ROW 7 — First empty data row ─────────────────────────
    for i in range(1, len(COLUMNS) + 1):
        c = sheet.cell(row=7, column=i, value="")
        c.fill   = hfill(DATA_BG)
        c.border = cb
    sheet.row_dimensions[7].height = 22

    # ── Column widths ─────────────────────────────────────────
    for i, w in enumerate(COL_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    sheet.freeze_panes        = "A8"   # freeze everything above the first data row
    sheet.sheet_view.zoomScale = 90
    return sheet


# ════════════════════════════════════════════════════════════
#  1.  DOWNLOAD BLANK TEMPLATE
# ════════════════════════════════════════════════════════════

def download_plot_resale_template(request):
    wb = Workbook()
    _build_template_sheet(wb)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="PropCRM_Plot_Resale_Template.xlsx"'
    )
    wb.save(response)
    return response


# ════════════════════════════════════════════════════════════
#  2.  IMPORT  (POST — reads data from row 7 onwards)
# ════════════════════════════════════════════════════════════

@csrf_protect
def import_plot_resale_excel(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method."})

    session_id = request.session.get("Admin_id")
    if not session_id:
        return JsonResponse({"status": "error", "message": "Session expired."})

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Invalid admin."})

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return JsonResponse({"status": "error", "message": "Please upload an Excel file."})

    try:
        wb    = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active

        saved_count   = 0
        skipped_count = 0
        row_logs      = []

        # ── Helpers ───────────────────────────────────────────────
        def clean_s(v):
            if v is None:
                return ""
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            return s

        def clean_f(v):
            try:
                if v in (None, ""):
                    return 0
                return float(str(v).replace(",", "").strip())
            except Exception:
                return 0

        # ── SKIP-ROW PATTERNS ─────────────────────────────────────
        # Rows 1-6 are headers, hints, sample data, and the banner.
        # We also defend against old templates where the banner lands
        # on row 7 by checking the cell content.
        SKIP_KEYWORDS = (
            "START YOUR DATA",
            "DELETE THIS ROW",
            "AUTO GENERATED",
            "property_title",       # db field name header row
            "Property Title",       # display label header row
            "⚠️",
            "👇",
            "🔴",
        )

        def is_header_or_banner(values):
            first = str(values[0] or "").strip()
            return any(kw in first for kw in SKIP_KEYWORDS)

        # ── DATA LOOP — starts at row 7 ───────────────────────────
        #    (Rows 1-6: section headers / db names / labels /
        #               hints / sample row / instruction banner)
        for row_idx in range(7, sheet.max_row + 1):

            values = [
                sheet.cell(row=row_idx, column=col).value
                for col in range(1, len(COLUMNS) + 1)
            ]

            # Skip fully empty rows
            if not any(v not in (None, "", 0) for v in values):
                continue

            # Skip header / banner rows (handles both old & new templates)
            if is_header_or_banner(values):
                skipped_count += 1
                row_logs.append(f"Row {row_idx}: Header/banner row skipped")
                continue

            # ── Map columns by position ───────────────────────────
            property_title        = clean_s(values[0])
            plot_title            = clean_s(values[1])
            plot_area             = clean_f(values[2])
            resale_plot_type      = clean_s(values[3])
            plot_road_facing      = clean_s(values[4])
            corner_plot           = clean_s(values[5])  or "no"
            sanctioning_authority = clean_s(values[6])
            plot_fencing          = clean_s(values[7])  or "no"
            plot_price            = clean_f(values[8])
            # values[9] = price_per_sqft — auto-calculated on model.save(), skip
            brokerage             = clean_s(values[10]) or "No"
            brokerage_percentage  = clean_s(values[11])
            ownership_type        = clean_s(values[12])
            loan_on_property      = clean_s(values[13]) or "no"
            plot_loan_amount      = clean_f(values[14])
            plot_city             = clean_s(values[15])
            plot_locality         = clean_s(values[16])
            plot_address          = clean_s(values[17])
            plot_owner_name       = clean_s(values[18])
            plot_owner_contact    = clean_s(values[19])
            plot_owner_email      = clean_s(values[20])
            plot_owner_role       = clean_s(values[21])

            # ── Skip sample row ───────────────────────────────────
            if "SAMPLE" in plot_title.upper() or plot_owner_contact == "9999999999":
                skipped_count += 1
                row_logs.append(f"Row {row_idx}: Sample row skipped")
                continue

            # ── Validate required fields ──────────────────────────
            missing = []
            if not plot_title:                      missing.append("Plot Title")
            if not plot_area or plot_area <= 0:     missing.append("Plot Area")
            if not resale_plot_type:                missing.append("Plot Type")
            if not plot_road_facing:                missing.append("Road Facing")
            if not plot_price or plot_price <= 0:   missing.append("Plot Price")
            if not ownership_type:                  missing.append("Ownership Type")
            if not plot_city:                       missing.append("City")
            if not plot_locality:                   missing.append("Locality")
            if not plot_address:                    missing.append("Address")
            if not plot_owner_name:                 missing.append("Owner Name")
            if not plot_owner_contact:              missing.append("Owner Contact")
            if not plot_owner_email:                missing.append("Owner Email")

            if missing:
                skipped_count += 1
                row_logs.append(f"Row {row_idx}: ⚠️ Skipped — Missing: {', '.join(missing)}")
                continue

            # ── Create record ─────────────────────────────────────
            try:
                obj = PlotSaleProperty.objects.create(
                    property_title        = property_title or plot_title,
                    plot_title            = plot_title,
                    plot_area             = plot_area,
                    resale_plot_type      = resale_plot_type,
                    plot_road_facing      = plot_road_facing,
                    corner_plot           = corner_plot if corner_plot in ("yes", "no") else "no",
                    sanctioning_authority = sanctioning_authority or None,
                    plot_fencing          = plot_fencing if plot_fencing in ("yes", "no") else "no",
                    plot_price            = plot_price,
                    # price_per_sqft is auto-calculated in model.save()
                    brokerage             = brokerage,
                    brokerage_percentage  = brokerage_percentage or None,
                    ownership_type        = ownership_type,
                    loan_on_property      = loan_on_property if loan_on_property in ("yes", "no") else "no",
                    plot_loan_amount      = plot_loan_amount if plot_loan_amount > 0 else None,
                    plot_city             = plot_city,
                    plot_locality         = plot_locality,
                    plot_address          = plot_address,
                    plot_owner_name       = plot_owner_name,
                    plot_owner_contact    = plot_owner_contact,
                    plot_owner_email      = plot_owner_email,
                    plot_owner_role       = plot_owner_role or None,
                    upload_file_name      = excel_file.name,
                    uploaded_by_name      = getattr(admin_obj, "name",  "Admin"),
                    uploaded_by_email     = getattr(admin_obj, "email", ""),
                    is_deleted            = False,
                )
                saved_count += 1
                row_logs.append(f"Row {row_idx}: ✅ Imported — {obj.plot_property_id}")

            except Exception as e:
                skipped_count += 1
                row_logs.append(f"Row {row_idx}: ❌ DB Error — {str(e)}")

        # ── Response ──────────────────────────────────────────────
        if saved_count > 0:
            return JsonResponse({
                "status":  "success",
                "message": f"{saved_count} propert{'y' if saved_count == 1 else 'ies'} imported successfully!",
                "saved":   saved_count,
                "skipped": skipped_count,
                "logs":    row_logs,
            })
        else:
            return JsonResponse({
                "status":  "warning",
                "message": (
                    f"No rows imported. {skipped_count} row(s) skipped. "
                    "Check that your data starts from Row 7 and all required fields are filled."
                ),
                "saved":   0,
                "skipped": skipped_count,
                "logs":    row_logs,
            })

    except Exception as e:
        return JsonResponse({
            "status":  "error",
            "message": str(e),
            "saved":   0,
            "skipped": 0,
            "logs":    [],
        })


# ════════════════════════════════════════════════════════════
#  3.  EXPORT  (GET — full data dump, all DB fields)
# ════════════════════════════════════════════════════════════

# Extra audit columns appended after the 22 import columns
EXPORT_EXTRA_HEADERS = [
    ("plot_property_id",    "Property ID"),
    ("uploaded_by_name",    "Uploaded By (Name)"),
    ("uploaded_by_role",    "Uploaded By (Role)"),
    ("uploaded_by_email",   "Uploaded By (Email)"),
    ("uploaded_by_contact", "Uploaded By (Contact)"),
    ("upload_file_name",    "Source File"),
    ("created_at",          "Created At"),
    ("updated_at",          "Last Updated"),
]

EXPORT_TOTAL_COLS = len(COLUMNS) + len(EXPORT_EXTRA_HEADERS)   # 30


def export_plot_resale_excel(request):
    """
    Exports all non-deleted PlotSaleProperty records to .xlsx.

    Cols  1-22  →  same as import template  (re-importable)
    Cols 23-30  →  audit / system fields
    """
    session_id = request.session.get("Admin_id")
    if not session_id:
        return HttpResponse("Unauthorised", status=401)

    from django.utils import timezone

    # ── Colour palette ─────────────────────────────────────────
    DARK_BG      = "1E293B"
    AUDIT_BG     = "1E3A5F"
    WHITE        = "FFFFFF"
    LIGHT_BG     = "F8FAFC"
    AUDIT_COL_BG = "EFF4FF"
    PRICE_BG     = "F0FFF4"
    OWNER_BG     = "FEFCE8"
    BORDER_COLOR = "CBD5E1"
    SUMMARY_BG   = "DCFCE7"
    SUMMARY_FG   = "166534"
    ALT_ROW_BG   = "F8FAFC"

    thin  = Side(style="thin",   color=BORDER_COLOR)
    thick = Side(style="medium", color="94A3B8")
    cb    = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    hb    = Border(left=thick, right=thick, top=thick, bottom=thick)

    def hfill(h):
        return PatternFill("solid", fgColor=h)

    def fmt_dec(v):
        if v is None:
            return ""
        try:
            f = float(v)
            return int(f) if f == int(f) else round(f, 2)
        except Exception:
            return ""

    def fmt_dt(v):
        if v is None:
            return ""
        try:
            return timezone.localtime(v).strftime("%d-%m-%Y  %H:%M")
        except Exception:
            return str(v)

    # ── Workbook setup ─────────────────────────────────────────
    wb    = Workbook()
    sheet = wb.active
    sheet.title = "Plot Resale Export"

    import_end  = len(COLUMNS)           # col 22
    audit_start = import_end + 1         # col 23
    audit_end   = EXPORT_TOTAL_COLS      # col 30

    # ── ROW 1 — Two section banners ────────────────────────────
    c = sheet.cell(row=1, column=1,
                   value="📋 Plot Data  —  Columns 1-22  (Import-Compatible)")
    c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill      = hfill(DARK_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = hb
    sheet.merge_cells(start_row=1, start_column=1,
                      end_row=1,   end_column=import_end)

    c = sheet.cell(row=1, column=audit_start,
                   value="🔒 Audit & System Info  (Read-Only)")
    c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill      = hfill(AUDIT_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = hb
    sheet.merge_cells(start_row=1, start_column=audit_start,
                      end_row=1,   end_column=audit_end)
    sheet.row_dimensions[1].height = 30

    # ── ROW 2 — DB field names ─────────────────────────────────
    for i, (db, _, _) in enumerate(COLUMNS, 1):
        c = sheet.cell(row=2, column=i, value=db)
        c.font      = Font(name="Arial", bold=True, size=9, color="475569")
        c.fill      = hfill("E2E8F0")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = cb

    for j, (db, _) in enumerate(EXPORT_EXTRA_HEADERS, audit_start):
        c = sheet.cell(row=2, column=j, value=db)
        c.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill      = hfill("334155")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = cb
    sheet.row_dimensions[2].height = 22

    # ── ROW 3 — Display labels ─────────────────────────────────
    for i, (_, disp, _) in enumerate(COLUMNS, 1):
        colour = "C0392B" if disp.endswith("*") else "3B82F6"
        c = sheet.cell(row=3, column=i, value=disp)
        c.font      = Font(name="Arial", bold=True, size=10, color=colour)
        c.fill      = hfill(LIGHT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = cb

    for j, (_, disp) in enumerate(EXPORT_EXTRA_HEADERS, audit_start):
        c = sheet.cell(row=3, column=j, value=disp)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill      = hfill("1E40AF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = cb
    sheet.row_dimensions[3].height = 36

    # ── Fetch records ──────────────────────────────────────────
    properties = PlotSaleProperty.objects.filter(is_deleted=False).order_by("-created_at")

    # ── DATA ROWS — start at row 4 ────────────────────────────
    row_num = 4
    for idx, p in enumerate(properties, 1):

        row_bg = "FFFFFF" if idx % 2 == 1 else ALT_ROW_BG

        # Cols 1-22: import-compatible data
        import_row = [
            str(p.property_title or ""),
            str(p.plot_title or ""),
            fmt_dec(p.plot_area),
            str(p.resale_plot_type or ""),
            str(p.plot_road_facing or ""),
            str(p.corner_plot or "no"),
            str(p.sanctioning_authority or ""),
            str(p.plot_fencing or "no"),
            fmt_dec(p.plot_price),
            fmt_dec(p.price_per_sqft),
            str(p.brokerage or "No"),
            str(p.brokerage_percentage or ""),
            str(p.ownership_type or ""),
            str(p.loan_on_property or "no"),
            fmt_dec(p.plot_loan_amount),
            str(p.plot_city or ""),
            str(p.plot_locality or ""),
            str(p.plot_address or ""),
            str(p.plot_owner_name or ""),
            str(p.plot_owner_contact or ""),
            str(p.plot_owner_email or ""),
            str(p.plot_owner_role or ""),
        ]

        for col_idx, val in enumerate(import_row, 1):
            c = sheet.cell(row=row_num, column=col_idx, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = cb
            if col_idx in (9, 10, 15):      # price columns — green tint
                c.fill = hfill(PRICE_BG)
            elif col_idx in (19, 20, 21):   # owner columns — yellow tint
                c.fill = hfill(OWNER_BG)
            else:
                c.fill = hfill(row_bg)

        # Cols 23-30: audit / system fields
        audit_row = [
            str(p.plot_property_id or ""),
            str(p.uploaded_by_name or ""),
            str(p.uploaded_by_role or ""),
            str(p.uploaded_by_email or ""),
            str(p.uploaded_by_contact or ""),
            str(p.upload_file_name or "Web UI"),
            fmt_dt(p.created_at),
            fmt_dt(p.updated_at),
        ]

        for col_idx, val in enumerate(audit_row, audit_start):
            c = sheet.cell(row=row_num, column=col_idx, value=val)
            c.font      = Font(name="Arial", size=10, color="1E293B")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = cb
            c.fill      = hfill(AUDIT_COL_BG)

        sheet.row_dimensions[row_num].height = 20
        row_num += 1

    # ── Summary footer ─────────────────────────────────────────
    total       = properties.count()
    export_time = timezone.localtime(timezone.now()).strftime("%d-%m-%Y  %H:%M")

    sc = sheet.cell(
        row=row_num, column=1,
        value=f"✅  Total Records Exported: {total}   |   Exported On: {export_time}"
    )
    sc.font      = Font(name="Arial", bold=True, size=10, color=SUMMARY_FG)
    sc.fill      = hfill(SUMMARY_BG)
    sc.alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(start_row=row_num, start_column=1,
                      end_row=row_num,   end_column=EXPORT_TOTAL_COLS)
    sheet.row_dimensions[row_num].height = 24

    # ── Column widths ──────────────────────────────────────────
    audit_widths = [22, 22, 18, 28, 20, 30, 22, 22]
    all_widths   = COL_WIDTHS + audit_widths
    for i, w in enumerate(all_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    sheet.freeze_panes        = "A4"
    sheet.sheet_view.zoomScale = 85

    # ── Deliver ────────────────────────────────────────────────
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="PropCRM_Plot_Resale_Export.xlsx"'
    )
    wb.save(response)
    return response



def plot_resale_list(request):
    session_id = request.session.get('Admin_id')

    if not session_id:
        return render(request,'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return render(request,'home_page/Adminlogin.html')

    base_qs = PlotSaleProperty.objects.filter(is_deleted=False)
    total_properties = base_qs.count()
    properties = base_qs

    # Get filter parameters
    search = request.GET.get('search')
    city = request.GET.get('city')
    locality = request.GET.get('locality')
    plot_type = request.GET.get('plot_type')
    road_facing = request.GET.get('road_facing')
    corner_plot = request.GET.get('corner_plot')
    loan = request.GET.get('loan')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Apply Filters
    if search:
        properties = properties.filter(
            Q(plot_property_id__icontains=search) |
            Q(property_title__icontains=search) |
            Q(plot_title__icontains=search) |
            Q(plot_city__icontains=search) |
            Q(plot_locality__icontains=search) |
            Q(plot_owner_name__icontains=search) |
            Q(plot_owner_contact__icontains=search)
        )

    if city:
        properties = properties.filter(plot_city__icontains=city)
    if locality:
        properties = properties.filter(plot_locality__icontains=locality)
    if plot_type:
        properties = properties.filter(resale_plot_type__icontains=plot_type)
    if road_facing:
        properties = properties.filter(plot_road_facing__icontains=road_facing)
    if corner_plot:
        properties = properties.filter(corner_plot__iexact=corner_plot) 
    if loan:
        properties = properties.filter(loan_on_property__iexact=loan) 
    if min_price:
        properties = properties.filter(plot_price__gte=min_price)
    if max_price:
        properties = properties.filter(plot_price__lte=max_price)
    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)
    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    properties = properties.order_by('-created_at')

    # Dashboard Statistics
    active_listings = base_qs.filter(plot_price__gt=0).count()
    residential_count = base_qs.filter(resale_plot_type__icontains='residential').count()
    commercial_count = base_qs.filter(resale_plot_type__icontains='commercial').count()
    agricultural_count = base_qs.filter(resale_plot_type__icontains='agricultural').count()
    finance_ready_count = base_qs.filter(loan_on_property__iexact='yes').count() 

    uploaded_files = list(
        base_qs.exclude(upload_file_name__isnull=True)
               .exclude(upload_file_name="")
               .values_list('upload_file_name', flat=True)
               .distinct()
    )

    context = {
        'admin_obj': admin_obj,
        'properties': properties,
        'total_properties': total_properties,
        'active_listings': active_listings,
        'residential_count': residential_count,
        'commercial_count': commercial_count,
        'agricultural_count': agricultural_count,
        'finance_ready_count': finance_ready_count,
        'uploaded_files': sorted(uploaded_files),
    }

    return render(request, 'admin_user/Reports/Resale/plot_list.html', context)


# ==========================================



 

def plot_sale_view(request, id):
    session_id = request.session.get('Admin_id')
    user_id    = request.session.get('User_id')
 
    if not session_id and not user_id:
        return redirect('login')
 
    prop = get_object_or_404(
        PlotSaleProperty.objects.prefetch_related('images', 'faqs'),
        plot_property_id=id
    )
 
    # Safety: regenerate FAQs if missing (old records before migration)
    if not prop.faqs.exists():
        prop.generate_auto_faqs()
 
    context = {
        'property': prop,  # Changed from 'prop' to 'property' to match the HTML
        'faqs': prop.faqs.all(),
    }
    return render(request, 'admin_user/Reports/Resale/plot_view.html', context)



@require_POST
def plot_sale_delete(request, id):

    try:

        deleter_name = _get_deleter_name(request)

        prop = get_object_or_404(
            PlotSaleProperty,
            plot_property_id=id,
            is_deleted=False
        )

        prop.is_deleted = True
        prop.deleted_at = timezone.now()
        prop.deleted_by = deleter_name
        prop.save()

        return JsonResponse({
            "status": "success",
            "message": f"{id} moved to recycle bin successfully."
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        })

@require_POST
def plot_sale_restore(request, id):

    updated = PlotSaleProperty.objects.filter(
        plot_property_id=id
    ).update(
        is_deleted=False,
        deleted_at=None,
        deleted_by=None
    )

    if updated:
        return JsonResponse({
            'status': 'success',
            'message': 'Plot restored successfully!'
        })

    return JsonResponse({
        'status': 'error',
        'message': 'Plot not found.'
    })

@require_POST
def plot_sale_hard_delete(request, id):

    deleted, _ = PlotSaleProperty.objects.filter(
        plot_property_id=id
    ).delete()

    if deleted:
        return JsonResponse({
            'status': 'success',
            'message': 'Property permanently deleted!'
        })

    return JsonResponse({
        'status': 'error',
        'message': 'Property not found.'
    })

    #####################END VIEW SECTION PLOT RESALE LISTING################


    ##############################START VIEW SECTION RESALE INDUSTRIAL LISTING#################

    

def safe_parse_date(date_str):
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None





def industrial_resale_add(request):
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    try:
        # Resolve administrator identity profiles context
        if admin_id:
            admin = Admin_Login.objects.get(id=admin_id)
            uploader_name = admin.name
            uploader_email = admin.email
            uploader_phone = admin.phone
            uploader_role = admin.role
        else:
            user = User_Details.objects.get(id=user_id)
            uploader_name = user.name
            uploader_email = user.email
            uploader_phone = user.phone
            uploader_role = user.role

        if request.method == "POST":
            # Map string toggle parameters safely to systemic database booleans
            power_val = request.POST.get('power_supply') == 'True'
            crane_val = request.POST.get('crane_heavy_machinery') == 'True'
            housing_val = request.POST.get('worker_housing_nearby') == 'True'
            
            # Use your custom date parser or handle None safely
            
            
            # FIXED BOOLEANS TO MATCH MODEL
            loan_val = request.POST.get('loan_on_property') == 'True'
            tenants_val = request.POST.get('existing_tenants') == 'True'
            dispute_val = request.POST.get('legal_dispute') == 'True'
            tax_due_val = request.POST.get('government_tax_dues') == 'True'
            tax_cert_val = request.POST.get('tax_clearance_cert') == 'True'

            # Backend calculation for price per sqft (failsafe)
            try:
                price = float(request.POST.get('expected_price') or 0)
                area = float(request.POST.get('land_area') or 0)
                calc_price_per_sqft = round((price / area), 2) if area > 0 else None
            except ValueError:
                calc_price_per_sqft = None

            prop = IndustrialResaleProperty.objects.create(
                property_title=request.POST.get('property_title'),

                # Step 1: Specs
                property_type=request.POST.get('property_type'),
                land_area=request.POST.get('land_area') or 0,
                # Note: Make sure your model has an 'available_from' field. I added it to the dict based on your old view.
                power_supply=power_val,
                kva_capacity=request.POST.get('kva_capacity') or None,
                water_supply=request.POST.get('water_supply'),
                crane_heavy_machinery=crane_val,
                road_connectivity=request.POST.get('road_connectivity'),
                worker_housing_nearby=housing_val,

                # Step 2: Pricing & Legal
                expected_price=request.POST.get('expected_price') or None,
                price_per_sqft=request.POST.get('price_per_sqft') or calc_price_per_sqft,  # ADDED METRIC
                brokerage=request.POST.get('brokerage'),
                brokerage_percentage=request.POST.get('brokerage_percentage'),
                manual_brokerage=request.POST.get('manual_brokerage'),
                sanctioning_authority=request.POST.get('sanctioning_authority'),
                ownership_type=request.POST.get('ownership_type'),
                
                loan_on_property=loan_val,               # FIXED KEY
                loan_amount=request.POST.get('loan_amount') or None,
                existing_tenants=tenants_val,
                tenant_details=request.POST.get('tenant_details'),
                legal_dispute=dispute_val,
                dispute_details=request.POST.get('dispute_details'),
                government_tax_dues=tax_due_val,         # FIXED KEY
                tax_amount=request.POST.get('tax_amount') or None,
                tax_clearance_cert=tax_cert_val,
                property_description=request.POST.get('property_description'),
                property_summary=request.POST.get('property_summary'),
                user_description=request.POST.get('user_description'),

                # Step 3: File Document Streams
                compliance_docs=request.FILES.get('compliance_docs'),
                social_video=request.FILES.get('social_video'),

                # Step 4: Contact Information
                city=request.POST.get('city'),
                locality_area=request.POST.get('locality_area'),   # FIXED KEY
                Property_address=request.POST.get('Property_address'), # FIXED KEY
                owner_name=request.POST.get('owner_name'),
                owner_contact=request.POST.get('owner_contact'),
                owner_email=request.POST.get('owner_email'),
                owner_role=request.POST.get('owner_role'),
                residency_status=request.POST.get('residency_status'),

                # Audit Session Metadata Tracks
                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_phone,
                uploaded_by_role=uploader_role
            )

            # Saves multi-image selections
            images = request.FILES.getlist('property_images[]')
            for idx, img in enumerate(images):
                if idx >= 10:
                    break
                # Ensure you have an IndustrialResaleImage model ready
                IndustrialResaleImage.objects.create(property=prop, image=img)

            # Fire the FAQ generator logic
            prop.generate_auto_faqs()

            return JsonResponse({
                "status": "success",
                "message": f"Industrial Listing successfully built and posted live!"
            })

    except Exception as err:
        print("CRITICAL EXCEPTION ROUTED:", str(err))
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": str(err)})

    context = {'admin_obj': admin if admin_id else user}
    return render(request, 'admin_user/industrial_resale.html', context)


def industrial_resale_edit(request, id):
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    # Fetch the existing property
    prop = get_object_or_404(IndustrialResaleProperty, id=id)

    if request.method == "POST":
        try:
            # Safely parse boolean choices
            def parse_bool(val):
                return str(val).strip().lower() in ['yes', 'true', '1', 'on']

            # Update Step 1: Specs
            prop.property_type = request.POST.get('property_type')
            prop.land_area = request.POST.get('land_area') or 0.0
            
            prop.available_from = safe_parse_date(request.POST.get('available_from'))
          
            prop.power_supply = parse_bool(request.POST.get('power_supply'))
            prop.kva_capacity = request.POST.get('kva_capacity') or None
            prop.water_supply = request.POST.get('water_supply')
            prop.crane_heavy_machinery = parse_bool(request.POST.get('crane_heavy_machinery'))
            prop.road_connectivity = request.POST.get('road_connectivity')
            prop.worker_housing_nearby = parse_bool(request.POST.get('worker_housing_nearby'))

            # Update Step 2: Pricing & Legal
            prop.expected_price = request.POST.get('expected_price') or 0.0
            prop.brokerage = request.POST.get('brokerage')
            prop.brokerage_percentage = request.POST.get('brokerage_percentage')
            prop.manual_brokerage = request.POST.get('manual_brokerage')
            prop.sanctioning_authority = request.POST.get('sanctioning_authority')
            prop.ownership_type = request.POST.get('ownership_type')
            
            prop.has_loan = parse_bool(request.POST.get('has_loan'))
            prop.loan_amount = request.POST.get('loan_amount') or 0.0
            
            prop.existing_tenants = parse_bool(request.POST.get('existing_tenants'))
            prop.tenant_details = request.POST.get('tenant_details')
            
            prop.legal_dispute = parse_bool(request.POST.get('legal_dispute'))
            prop.dispute_details = request.POST.get('dispute_details')
            
            prop.tax_due = parse_bool(request.POST.get('tax_due'))
            prop.tax_amount = request.POST.get('tax_amount') or 0.0
            prop.tax_clearance_cert = parse_bool(request.POST.get('tax_clearance_cert'))
            
            prop.property_description = request.POST.get('property_description')

            prop.property_summary = request.POST.get('property_summary')
            prop.user_description = request.POST.get('user_description')

            # Update Step 3: Media & Files
            if 'compliance_docs' in request.FILES:
                prop.compliance_docs = request.FILES['compliance_docs']
            if 'social_video' in request.FILES:
                prop.social_video = request.FILES['social_video']

            # Update Step 4: Location & Contact
            prop.city = request.POST.get('city')
            prop.locality = request.POST.get('locality')
            prop.complete_address = request.POST.get('complete_address')
            prop.owner_name = request.POST.get('owner_name')
            prop.owner_contact = request.POST.get('owner_contact')
            prop.owner_email = request.POST.get('owner_email')
            prop.residency_status = request.POST.get('residency_status')

            # Save core property attributes
            prop.save()

            # Handle Deletion of Specific Old Images
            removed_image_ids = request.POST.getlist('removed_images[]')
            if removed_image_ids:
                from .models import IndustrialResaleImage  # Ensure imported
                IndustrialResaleImage.objects.filter(id__in=removed_image_ids, property=prop).delete()

            # Save New Images (Append up to 10 max)
            new_images = request.FILES.getlist('property_images[]')
            current_image_count = prop.images.count()

            for img in new_images:
                if current_image_count >= 10:
                    break
                from .models import IndustrialResaleImage
                IndustrialResaleImage.objects.create(property=prop, image=img)
                current_image_count += 1

            return JsonResponse({"status": "success", "message": "Industrial Property Updated Successfully"})

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": str(e)})

    # GET Request: Render edit template with pre-filled data
    context = {
        'prop': prop,
        # Determine uploader for display
        'admin_obj': Admin_Login.objects.filter(id=admin_id).first() if admin_id else None
    }
    return render(request, 'admin_user/Resale/industrial_edit.html', context)





def industrial_resale_list(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')
        
    admin_obj = Admin_Login.objects.get(id=session_id)
    base_qs = IndustrialResaleProperty.objects.filter(is_deleted=False)
    
    # ─── Filter Parameters ───
    search_query = request.GET.get('search', '').strip()
    city_query = request.GET.get('city', 'All Cities')
    prop_type_query = request.GET.get('property_type', 'All Types')
    power_query = request.GET.get('power_supply', 'Any Power Supply')
    legal_query = request.GET.get('legal_status', 'All')  # Newly added Legal filter
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # ─── Apply Filters ───
    if search_query:
        base_qs = base_qs.filter(
            Q(property_title__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(locality_area__icontains=search_query) |  # Aligned to new model
            Q(owner_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )

    if city_query != 'All Cities':
        base_qs = base_qs.filter(city__iexact=city_query)
        
    if prop_type_query != 'All Types':
        base_qs = base_qs.filter(property_type__iexact=prop_type_query)
        
    if power_query == 'heavy':
        base_qs = base_qs.filter(power_supply=True, kva_capacity__gte=100)
    elif power_query == 'light':
        base_qs = base_qs.filter(power_supply=True, kva_capacity__lt=100)
        
    if legal_query == 'clear':
        # Must have no tax dues AND no legal disputes
        base_qs = base_qs.filter(government_tax_dues=False, legal_dispute=False)
    elif legal_query == 'dispute':
        # Has EITHER tax dues OR a legal dispute
        base_qs = base_qs.filter(Q(government_tax_dues=True) | Q(legal_dispute=True))
        
    if from_date and to_date:
        base_qs = base_qs.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)

    # Sort listings chronologically
    properties_qs = base_qs.order_by('-created_at')

    # ─── EXPORT LOGIC (CSV / EXCEL) ───
    download_type = request.GET.get('download')
    if download_type in ['csv', 'excel']:
        # Exact headers matching the new model sequentially
        headers = [
            # System Control & Identification
            "ID", "Property Title", 
            # Step 1: Property Specs
            "Property Type", "Land Area", "Power Supply", "KVA Capacity", 
            "Water Supply", "Crane & Heavy Machinery", "Road Connectivity", "Worker Housing Nearby",
            # Step 2: Pricing & Legal
            "Expected Price", "Price Per Sqft", "Brokerage", "Brokerage Percentage", 
            "Manual Brokerage", "Sanctioning Authority", "Ownership Type", "Loan On Property", 
            "Loan Amount", "Existing Tenants", "Tenant Details", "Legal Dispute", 
            "Dispute Details", "Government Tax Dues", "Tax Amount", "Tax Clearance Cert", 
            "Property Description",
            # Step 3: Media & Compliance
            "Compliance Docs", "Social Video",
            # Step 4: Location & Contact
            "City", "Locality Area", "Property Address", "Owner Name", "Owner Contact", 
            "Owner Email", "Owner Role", "Residency Status",
            # Uploader Details
            "Uploaded By Name", "Uploaded By Role", "Uploaded By Email", "Uploaded By Contact", 
            "Upload File Name", 
            # Timestamp & Deletion Status
            "Created At", "Updated At", "Is Deleted", "Deleted At", "Deleted By"
        ]

        # Helper function to extract data sequentially for each property using exact model fields
        def get_row_data(prop):
            return [
                prop.id, prop.property_title,
                prop.property_type, prop.land_area,
                "Yes" if prop.power_supply else "No", prop.kva_capacity, prop.water_supply,
                "Yes" if prop.crane_heavy_machinery else "No", prop.road_connectivity,
                "Yes" if prop.worker_housing_nearby else "No", prop.expected_price, prop.price_per_sqft,
                prop.brokerage, prop.brokerage_percentage, prop.manual_brokerage,
                prop.sanctioning_authority, prop.ownership_type, "Yes" if prop.loan_on_property else "No",
                prop.loan_amount, "Yes" if prop.existing_tenants else "No", prop.tenant_details,
                "Yes" if prop.legal_dispute else "No", prop.dispute_details,
                "Yes" if prop.government_tax_dues else "No", prop.tax_amount,
                "Yes" if prop.tax_clearance_cert else "No", prop.property_description,
                prop.compliance_docs.name if prop.compliance_docs else "",
                prop.social_video.name if prop.social_video else "",
                prop.city, prop.locality_area, prop.Property_address, prop.owner_name,
                prop.owner_contact, prop.owner_email, prop.owner_role, prop.residency_status,
                prop.uploaded_by_name, prop.uploaded_by_role, prop.uploaded_by_email,
                prop.uploaded_by_contact, prop.upload_file_name,
                str(prop.created_at.replace(tzinfo=None)) if prop.created_at else "",
                str(prop.updated_at.replace(tzinfo=None)) if prop.updated_at else "",
                "Yes" if prop.is_deleted else "No",
                str(prop.deleted_at.replace(tzinfo=None)) if prop.deleted_at else "",
                prop.deleted_by
            ]

        if download_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Industrial_Resale_Report.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            for prop in properties_qs:
                writer.writerow(get_row_data(prop))
            return response

        elif download_type == 'excel':
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Industrial Report Data"
            sheet.append(headers)
            for prop in properties_qs:
                sheet.append(get_row_data(prop))
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Industrial_Resale_Report.xlsx"'
            wb.save(response)
            return response

    # ─── Pagination ───
    paginator = Paginator(properties_qs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ─── Dynamic Statistical Compilations ───
    total_count = IndustrialResaleProperty.objects.filter(is_deleted=False).count()
    filtered_count = properties_qs.count()
    active_listings = properties_qs.filter(expected_price__gt=0).count()
    warehouse_count = properties_qs.filter(property_type__icontains='warehouse').count()
    compliance_passed = properties_qs.exclude(compliance_docs='').exclude(compliance_docs__isnull=True).count()

    # ─── Chart Data Aggregation ───
    small_cap = properties_qs.filter(land_area__lte=20000).count()
    mid_cap = properties_qs.filter(land_area__gt=20000, land_area__lte=80000).count()
    large_cap = properties_qs.filter(land_area__gt=80000).count()
    
    manufacturing = properties_qs.filter(property_type__icontains='manufacturing').count()
    cold_storage = properties_qs.filter(property_type__icontains='cold_storage').count()
    
    # Updated chart filters to match new model field definitions
    low_risk = properties_qs.filter(government_tax_dues=False, legal_dispute=False).count()
    medium_risk = properties_qs.filter(government_tax_dues=True, legal_dispute=False).count()
    high_risk = properties_qs.filter(legal_dispute=True).count()

    # Dropdown Options Pickers
    unique_cities = IndustrialResaleProperty.objects.filter(is_deleted=False).exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct()
    unique_property_types = IndustrialResaleProperty.objects.filter(is_deleted=False).exclude(property_type__isnull=True).exclude(property_type='').values_list('property_type', flat=True).distinct()
    uploaded_files = IndustrialResaleProperty.objects.filter(is_deleted=False).exclude(upload_file_name__isnull=True).exclude(upload_file_name='').values_list('upload_file_name', flat=True).distinct()

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,
        'properties': page_obj, 
        'total_properties': total_count, 
        'total_count': total_count,
        'filtered_count': filtered_count,
        'active_listings': active_listings,
        'warehouse_count': warehouse_count,
        'compliance_passed': compliance_passed,
        
        # Chart Data
        'capacity_labels': json.dumps(["Small (≤ 20k sqft)", "Mid (20-80k sqft)", "Large (80k+ sqft)"]),
        'capacity_data': json.dumps([small_cap, mid_cap, large_cap]),
        'type_labels': json.dumps(["Warehouse", "Manufacturing", "Cold Storage", "Other"]),
        'type_data': json.dumps([warehouse_count, manufacturing, cold_storage, filtered_count - (warehouse_count + manufacturing + cold_storage)]),
        'risk_labels': json.dumps(["Low Risk", "Medium Risk", "High Risk"]),
        'risk_data': json.dumps([low_risk, medium_risk, high_risk]),
        
        # Select Dropdowns
        'unique_cities': unique_cities,
        'unique_property_types': unique_property_types,
        'uploaded_files': uploaded_files,
        
        # Maintained Filter States (so they stay selected on page reload)
        'search_query': search_query,
        'city_query': city_query,
        'prop_type_query': prop_type_query,
        'power_query': power_query,
        'legal_query': legal_query,
    }
    return render(request, 'admin_user/Reports/Resale/industrial_list.html', context)



def industrial_resale_view(request, id):
    session_id = request.session.get('Admin_id')
    user_id    = request.session.get('User_id')

    if not session_id and not user_id:
        return redirect('login')

    # Fetch the property + prefetch related images and DB-stored FAQs
    prop = get_object_or_404(
        IndustrialResaleProperty.objects.prefetch_related('images', 'faqs'),
        id=id
    )

    # FAQs are now persisted in DB via generate_auto_faqs() called on save().
    # If for any reason they are missing (e.g. old record), regenerate on the fly.
    if not prop.faqs.exists():
        prop.generate_auto_faqs()

    context = {
        'prop': prop,
        'faqs': prop.faqs.all(),   # QuerySet from IndustrialResaleFAQ table
    }
    return render(request, 'admin_user/Resale/industrial_view.html', context)





@require_POST
def industrial_bulk_delete(request):
    """Handles Advanced Bulk Deletions for Industrial Properties."""
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')
    
    if not admin_id and not user_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'})

    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        
        # Calling your existing helper function here
        deleter_name = _get_deleter_name(request)
        
        # Target only properties not currently in the Recycle Bin
        properties = IndustrialResaleProperty.objects.filter(is_deleted=False)
        
        if delete_type == 'delete_all':
            count = properties.count()
            properties.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved ALL {count} properties to Recycle Bin.'})
            
        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(id__in=page_ids)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from current page to Recycle Bin.'})
            
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            target_props = properties.filter(created_at__date__range=[from_date, to_date])
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties in date range to Recycle Bin.'})
            
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from the last 30 days to Recycle Bin.'})
            
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} older properties to Recycle Bin.'})
            
        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '').strip()
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) | 
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties uploaded by {uploader} to Recycle Bin.'})

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            # Updated to match exact DB field name
            target_props = properties.filter(upload_file_name=file_name) 
            count = target_props.count()
            if count == 0:
                return JsonResponse({'status': 'error', 'message': f'No properties found for file: {file_name}'})
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from {file_name} to Recycle Bin.'})
            
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@require_POST
def industrial_resale_hard_delete(request, id):
    """Permanently deletes the record from the database."""
    try:
        IndustrialResaleProperty.objects.filter(id=id).delete()
        return JsonResponse({'status': 'success', 'message': 'Permanently deleted!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@require_POST
def industrial_resale_restore(request, id):
    """Restores the record from the Recycle Bin."""
    try:
        IndustrialResaleProperty.objects.filter(id=id).update(is_deleted=False, deleted_at=None, deleted_by=None)
        return JsonResponse({'status': 'success', 'message': 'Industrial property restored!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ── 1. SINGLE DELETE VIEW (Soft Delete) ──
@require_POST
def industrial_resale_delete(request, id):
    try:
        # Get the name of the user/admin deleting the property
        deleter_name = _get_deleter_name(request)
        
        # Fetch the property
        prop = get_object_or_404(IndustrialResaleProperty, id=id)
        
        # Soft Delete
        prop.is_deleted = True
        prop.deleted_at = timezone.now()
        prop.deleted_by = deleter_name
        prop.save() 
        
        return JsonResponse({
            "status": "success", 
            "message": "Moved to Recycle Bin successfully."
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error", 
            "message": f"Failed to delete: {str(e)}"
        })





# =========================================================================
# 1. TEMPLATE DOWNLOADER
# =========================================================================
def download_industrial_resale_template(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Industrial Resale"

    # --- STYLING DEFINITIONS ---
    section_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    required_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    
    white_font_bold = Font(color="FFFFFF", bold=True, size=12)
    dark_font_bold = Font(color="1E293B", bold=True, size=11)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), 
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    # Row 1: Section Headings spanning across cell ranges (Total 47 columns)
    section_headers = [""] * 47
    section_headers[0] = "📋 System & Identification"        # Cols 1-2
    section_headers[2] = "📋 Property Specs"                # Cols 3-10
    section_headers[10] = "📋 Pricing & Legal"              # Cols 11-27
    section_headers[27] = "📋 Media & Compliance"           # Cols 28-29
    section_headers[29] = "📋 Location & Contact"           # Cols 30-37
    section_headers[37] = "📋 Uploader Details"             # Cols 38-42
    section_headers[42] = "📋 System Timestamps & Log"      # Cols 43-47
    
    sheet.append(section_headers)
    
    # Merge Section Blocks precisely to match the exact field counts (Shifted left by 1)
    sheet.merge_cells('A1:B1')   # System (2)
    sheet.merge_cells('C1:J1')   # Specs (8) - Removed available_from
    sheet.merge_cells('K1:AA1')  # Pricing & Legal (17)
    sheet.merge_cells('AB1:AC1') # Media (2)
    sheet.merge_cells('AD1:AK1') # Location (8)
    sheet.merge_cells('AL1:AP1') # Uploader (5)
    sheet.merge_cells('AQ1:AU1') # Timestamps (5)

    # Apply Styling to Row 1
    for cell in sheet[1]:
        cell.fill = section_fill
        cell.font = white_font_bold
        cell.alignment = center_align
        cell.border = thin_border
    sheet.row_dimensions[1].height = 30

    # Row 2: Headers (available_from removed)
    headers = [
        # System Control (Cols 0-1)
        "id", "property_title",
        # Step 1: Specs (Cols 2-9)
        "property_type *", "land_area *", "power_supply",
        "kva_capacity", "water_supply", "crane_heavy_machinery", "road_connectivity *",
        "worker_housing_nearby", 
        # Step 2: Pricing & Legal (Cols 10-26)
        "expected_price *", "price_per_sqft", "brokerage", "brokerage_percentage",
        "manual_brokerage", "sanctioning_authority", "ownership_type *", "loan_on_property",
        "loan_amount", "existing_tenants", "tenant_details", "legal_dispute",
        "dispute_details", "government_tax_dues", "tax_amount", "tax_clearance_cert",
        "property_description *", 
        # Step 3: Media (Cols 27-28)
        "compliance_docs", "social_video",
        # Step 4: Location & Contact (Cols 29-36)
        "city *", "locality_area *", "Property_address *", "owner_name *", 
        "owner_contact *", "owner_email *", "owner_role *", "residency_status *",
        # Uploader Details (Cols 37-41)
        "uploaded_by_name", "uploaded_by_role", "uploaded_by_email", "uploaded_by_contact", 
        "upload_file_name",
        # Meta Timestamps (Cols 42-46)
        "created_at", "updated_at", "is_deleted", "deleted_at", "deleted_by"
    ]
    sheet.append(headers)

    # Apply Styling to Row 2
    for cell in sheet[2]:
        cell.font = dark_font_bold
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = required_fill if "*" in str(cell.value) else header_fill
        sheet.column_dimensions[cell.column_letter].width = len(str(cell.value)) + 8

    sheet.row_dimensions[2].height = 25

    # Row 3: Sample Data (available_from date removed)
    sample_data = [
        "", "",
        "warehouse", 5000, "yes",
        250, "corporation", "no", "highway",
        "yes", 
        15000000, 3000.00, "Yes", "2",
        "", "MIDC", "freehold", "no",
        0, "no", "", "no",
        "", "no", 0, "yes",
        "Good industrial shed near highway", 
        "", "",
        "Nagpur", "Hingna MIDC", "Plot 42, Phase 1", "Ramesh Verma", 
        "9876543210", "ramesh@example.com", "Owner", "resident",
        "Admin", "System Admin", "admin@example.com", "9999999999", 
        "bulk_upload_01.csv",
        "", "", "False", "", ""
    ]
    sheet.append(sample_data)
    
    for cell in sheet[3]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PropCRM_Industrial_Resale_Template.xlsx"'
    wb.save(response)
    return response


# =========================================================================
# 2. BULK EXCEL IMPORTER
# =========================================================================
def import_industrial_resale_excel(request):
    session_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not session_id and not user_id:
        return JsonResponse({"status": "error", "message": "Unauthorized access"})

    if request.method == "POST" and request.FILES.get('industrial_file'):
        try:
            if session_id:
                uploader = Admin_Login.objects.get(id=session_id)
                u_name, u_email, u_phone, u_role = uploader.name, uploader.email, uploader.phone, "Admin"
            else:
                uploader = User_Details.objects.get(id=user_id)
                u_name, u_email, u_phone, u_role = uploader.name, uploader.email, uploader.phone, uploader.role

            excel_file = request.FILES['industrial_file']
            file_name_str = excel_file.name

            if IndustrialResaleProperty.objects.filter(upload_file_name=file_name_str, is_deleted=False).exists():
                return JsonResponse({"status": "duplicate_filename_and_data", "message": f"A file named '{file_name_str}' has already been processed."})

            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            if sheet.max_row < 3:
                return JsonResponse({"status": "error", "message": "Uploaded spreadsheet lacks data rows."})

            row2_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[2]]
            
            # Expected Headers (47 columns)
            expected_headers = [
                "id", "property_title",
                "property_type *", "land_area *", "power_supply",
                "kva_capacity", "water_supply", "crane_heavy_machinery", "road_connectivity *",
                "worker_housing_nearby", "expected_price *", "price_per_sqft", "brokerage", "brokerage_percentage",
                "manual_brokerage", "sanctioning_authority", "ownership_type *", "loan_on_property",
                "loan_amount", "existing_tenants", "tenant_details", "legal_dispute",
                "dispute_details", "government_tax_dues", "tax_amount", "tax_clearance_cert",
                "property_description *", "compliance_docs", "social_video",
                "city *", "locality_area *", "Property_address *", "owner_name *", 
                "owner_contact *", "owner_email *", "owner_role *", "residency_status *",
                "uploaded_by_name", "uploaded_by_role", "uploaded_by_email", "uploaded_by_contact", 
                "upload_file_name", "created_at", "updated_at", "is_deleted", "deleted_at", "deleted_by"
            ]

            for idx, key in enumerate(expected_headers):
                if idx >= len(row2_values) or row2_values[idx].lower() != key.lower():
                    return JsonResponse({"status": "error", "message": f"Column structure mismatch at Column {idx+1}. Expected '{key}'."})

            def parse_bool(val):
                return str(val).strip().lower() in ['yes', 'true', '1']

            def safe_float(val):
                try: return float(val) if val else 0.0
                except ValueError: return 0.0

            all_rows_data_str = ""
            records_to_create = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                if not row or not any(row): continue

                # Validation Indices Shifted
                required_indices = [
                    2,   # property_type
                    3,   # land_area
                    8,   # road_connectivity
                    10,  # expected_price
                    16,  # ownership_type
                    26,  # property_description
                    29,  # city
                    30,  # locality_area
                    31,  # Property_address
                    32,  # owner_name
                    33,  # owner_contact
                    34,  # owner_email
                    35,  # owner_role
                    36   # residency_status
                ]
                
                for ri in required_indices:
                    if row[ri] is None or str(row[ri]).strip() == "":
                        return JsonResponse({"status": "error", "message": f"Data Validation Error at Row {row_idx}: Column '{expected_headers[ri]}' cannot be blank."})

                all_rows_data_str += "".join([str(val) for val in row])

                # Mapped with shifted indices
                records_to_create.append({
                    "property_type": str(row[2]).strip(),
                    "land_area": safe_float(row[3]),
                    "power_supply": parse_bool(row[4]),
                    "kva_capacity": row[5] or None,
                    "water_supply": str(row[6]).strip() if row[6] else None,
                    "crane_heavy_machinery": parse_bool(row[7]),
                    "road_connectivity": str(row[8]).strip(),
                    "worker_housing_nearby": parse_bool(row[9]),
                    "expected_price": safe_float(row[10]),
                    "price_per_sqft": safe_float(row[11]),
                    "brokerage": str(row[12]).strip() if row[12] else "No",
                    "brokerage_percentage": str(row[13]).strip() if row[13] else None,
                    "manual_brokerage": str(row[14]).strip() if row[14] else None,
                    "sanctioning_authority": str(row[15]).strip() if row[15] else None,
                    "ownership_type": str(row[16]).strip(),
                    "loan_on_property": parse_bool(row[17]),
                    "loan_amount": safe_float(row[18]),
                    "existing_tenants": parse_bool(row[19]),
                    "tenant_details": str(row[20]).strip() if row[20] else None,
                    "legal_dispute": parse_bool(row[21]),
                    "dispute_details": str(row[22]).strip() if row[22] else None,
                    "government_tax_dues": parse_bool(row[23]),
                    "tax_amount": safe_float(row[24]),
                    "tax_clearance_cert": parse_bool(row[25]),
                    "property_description": str(row[26]).strip(),
                    "city": str(row[29]).strip(),
                    "locality_area": str(row[30]).strip(),
                    "Property_address": str(row[31]).strip(),
                    "owner_name": str(row[32]).strip(),
                    "owner_contact": str(row[33]).strip(),
                    "owner_email": str(row[34]).strip(),
                    "owner_role": str(row[35]).strip(),
                    "residency_status": str(row[36]).strip(),
                })

            combined_hash_fingerprint = hashlib.md5(all_rows_data_str.encode('utf-8')).hexdigest()
            if IndustrialResaleProperty.objects.filter(property_description__icontains=f"[MD5:{combined_hash_fingerprint}]", is_deleted=False).exists():
                return JsonResponse({"status": "duplicate_data_different_filename", "message": "Duplicate Content Rejected: This exact data matrix has already been imported."})

            for r in records_to_create:
                IndustrialResaleProperty.objects.create(
                    property_type=r["property_type"],
                    land_area=r["land_area"],
                    power_supply=r["power_supply"],
                    kva_capacity=r["kva_capacity"],
                    water_supply=r["water_supply"],
                    crane_heavy_machinery=r["crane_heavy_machinery"],
                    road_connectivity=r["road_connectivity"],
                    worker_housing_nearby=r["worker_housing_nearby"],
                    expected_price=r["expected_price"],
                    price_per_sqft=r["price_per_sqft"],
                    brokerage=r["brokerage"],
                    brokerage_percentage=r["brokerage_percentage"],
                    manual_brokerage=r["manual_brokerage"],
                    sanctioning_authority=r["sanctioning_authority"],
                    ownership_type=r["ownership_type"],
                    loan_on_property=r["loan_on_property"],
                    loan_amount=r["loan_amount"],
                    existing_tenants=r["existing_tenants"],
                    tenant_details=r["tenant_details"],
                    legal_dispute=r["legal_dispute"],
                    dispute_details=r["dispute_details"],
                    government_tax_dues=r["government_tax_dues"],
                    tax_amount=r["tax_amount"],
                    tax_clearance_cert=r["tax_clearance_cert"],
                    property_description=f"{r['property_description']} [MD5:{combined_hash_fingerprint}]",
                    city=r["city"],
                    locality_area=r["locality_area"],
                    Property_address=r["Property_address"],
                    owner_name=r["owner_name"],
                    owner_contact=r["owner_contact"],
                    owner_email=r["owner_email"],
                    owner_role=r["owner_role"],
                    residency_status=r["residency_status"],
                    upload_file_name=file_name_str,
                    uploaded_by_name=u_name,
                    uploaded_by_email=u_email,
                    uploaded_by_contact=u_phone,
                    uploaded_by_role=u_role
                )

            return JsonResponse({"status": "success", "message": f"Successfully parsed and saved {len(records_to_create)} industrial rows."})

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"Parsing Error encountered: {str(e)}"})

    return JsonResponse({"status": "error", "message": "Invalid parameters submitted."})


    #####################END VIEW SECTION RESALE INDUSTRIL LISTING########################################


#####################START VIEW SECTION OF RESIDENTIAL RESALE LISTING###########################

def get_property_images(request, id):
    prop = ResaleResidentialProperty.objects.get(id=id)
    images = [img.image.url for img in prop.images.all()]
    return JsonResponse({'images': images})

    

def _get_uploader(request):
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    admin_obj = None
    user_obj  = None

    uploader_name    = ""
    uploader_email   = ""
    uploader_phone   = ""
    uploader_role    = ""

    if admin_id:
        try:
            admin_obj        = Admin_Login.objects.get(id=admin_id)
            uploader_name    = admin_obj.name      # ← fixed (was user_name)
            uploader_email   = admin_obj.email
            uploader_phone   = admin_obj.phone
            uploader_role    = admin_obj.role
        except Admin_Login.DoesNotExist:
            return None

    elif user_id:
        try:
            user_obj         = User_Details.objects.get(id=user_id)
            uploader_name    = user_obj.user_name
            uploader_email   = user_obj.user_email
            uploader_phone   = user_obj.user_phone
            uploader_role    = user_obj.user_role
        except User_Details.DoesNotExist:
            return None
    else:
        return None  # not logged in at all

    return {
        "admin_obj"      : admin_obj,
        "user_obj"       : user_obj,
        "uploader_name"  : uploader_name,
        "uploader_email" : uploader_email,
        "uploader_phone" : uploader_phone,
        "uploader_role"  : uploader_role,
    }







def resale_residential_add(request):
    # ── 1. Session Check ──────────────────────────────────
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    # ── 2. Handle POST Request ────────────────────────────
    if request.method == "POST":
        
        # --- A. Safely convert primary configuration metrics ---
        try:
            builtup_val = float(request.POST.get('builtup_area') or 0.0)
        except ValueError:
            builtup_val = 0.0

        try:
            carpet_val = float(request.POST.get('carpet_area') or 0.0)
        except ValueError:
            carpet_val = 0.0

        try:
            price_val = float(request.POST.get('expected_price') or 0.0)
        except ValueError:
            price_val = 0.0

        # --- B. Safely handle optional/conditional numerical fields ---
        try:
            plot_area_raw = request.POST.get('plot_area')
            plot_area_val = float(plot_area_raw) if plot_area_raw and plot_area_raw.strip() else None
        except ValueError:
            plot_area_val = None

        try:
            loan_amount_raw = request.POST.get('loan_amount')
            loan_amount_val = float(loan_amount_raw) if loan_amount_raw and loan_amount_raw.strip() else None
        except ValueError:
            loan_amount_val = None

        try:
            pending_tax_raw = request.POST.get('pending_tax_amount')
            pending_tax_val = float(pending_tax_raw) if pending_tax_raw and pending_tax_raw.strip() else None
        except ValueError:
            pending_tax_val = None

        # ─── C. Create the Property Object ───────────────────
        prop = ResaleResidentialProperty(
            # Basic Information
            property_title   = request.POST.get('title') or None, 
            property_type    = request.POST.get('property_type'),
            zone             = request.POST.get('zone'),
            society_type     = request.POST.get('society_type'),
            water_type       = request.POST.get('water_type'),
            furnishing_type  = request.POST.get('furnishing_type'),
            age_of_property  = request.POST.get('age_of_property'),
            facing_direction = request.POST.get('facing_direction'), 

            # Property Configuration
            bhk              = request.POST.get('bhk'),
            bathrooms        = int(request.POST.get('bathrooms') or 1),
            balconies        = int(request.POST.get('balconies') or 0),
            covered_parking  = int(request.POST.get('covered_parking') or 0),
            open_parking     = int(request.POST.get('open_parking') or 0),

            # Measurements
            builtup_area     = builtup_val,
            carpet_area      = carpet_val,
            plot_area        = plot_area_val,
            floor_no         = int(request.POST.get('floor_no') or 0),
            total_floors     = int(request.POST.get('total_floors') or 0),

            # Ownership & Legal
            ownership_type      = request.POST.get('ownership_type'),
            num_owners          = request.POST.get('num_owners'),
            loan_on_property    = request.POST.get('loan_on_property', 'no'), 
            loan_amount         = loan_amount_val,
            existing_tenants    = request.POST.get('existing_tenants', 'no'), 
            tenant_details      = request.POST.get('tenant_details') or None,
            any_legal_dispute   = request.POST.get('any_legal_dispute', 'no'), 
            dispute_details     = request.POST.get('dispute_details') or None,
            government_tax_dues = request.POST.get('government_tax_dues', 'no'), 
            pending_tax_amount  = pending_tax_val,

            # Pricing & Description (FIXED: Linked directly to matching HTML input name properties)
            expected_price       = price_val,
            price_negotiable     = request.POST.get('is_negotiable', 'yes'), 
            brokerage            = request.POST.get('brokerage') or None,
            brokerage_percentage = request.POST.get('brokerage_percentage') or None,
            manual_brokerage     = request.POST.get('manual_brokerage') or None,
            property_description = request.POST.get('property_description'), 
            property_summary = request.POST.get('property_summary'),
            user_description = request.POST.get('user_description'), 

            # Amenities & Facilities 
            nearby_facilities = ', '.join(request.POST.getlist('facilities[]')),
            amenities         = ', '.join(request.POST.getlist('amenities[]')),

            # Address
            city             = request.POST.get('city'),
            locality         = request.POST.get('locality'),
            building_name    = request.POST.get('building_name') or None,
            complete_address = request.POST.get('complete_address'),

            # Owner Contact
            owner_name         = request.POST.get('owner_name'),
            owner_contact      = request.POST.get('owner_contact'),
            owner_email        = request.POST.get('owner_email'),
            owner_role         = request.POST.get('owner_role'), 
            residential_status = request.POST.get('residential_status'),

            # Single file fields
            floor_plan     = request.FILES.get('floor_plan') or None,
            property_video = request.FILES.get('property_video') or None,

            # Auto-filled from session
            uploaded_by_name    = uploader['uploader_name'],
            uploaded_by_email   = uploader['uploader_email'],
            uploaded_by_contact = uploader['uploader_phone'],
            uploaded_by_role    = uploader['uploader_role'],
        )

        # Save the main property object safely
        prop.save()  

        # --- D. Save multiple images into ResalePropertyImage ---
        images = request.FILES.getlist('property_images')
        for image in images:
            ResalePropertyImage.objects.create(
                property=prop,
                image=image
            )

        # --- E. Return valid JSON for your SweetAlert success box ---
        return JsonResponse({
            "status" : "success",
            "message": "Resale Residential Property Added Successfully"
        })

    # ── 3. Handle GET Request (Render Form) ───────────────────
    context = {
        "admin_obj"      : uploader['admin_obj'],
        "user_obj"       : uploader['user_obj'],
        "uploader_name"  : uploader['uploader_name'],  
        "uploader_email" : uploader['uploader_email'], 
        "uploader_phone" : uploader['uploader_phone'], 
        "uploader_role"  : uploader['uploader_role'],  
    }
    
    return render(request, 'admin_user/Reports/Resale/residential_resale_list.html', context)



def residential_resale_list(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # ── Fetch ALL properties (used for KPI stats & chart data) ───────────────
    all_properties = (
        ResaleResidentialProperty.objects
        .prefetch_related('images')
        .order_by('-created_at')
    )

    # ── Read query params ────────────────────────────────────────────────────
    search_query    = request.GET.get('search', '').strip()
    prop_type       = request.GET.get('prop_type', '').strip()
    bhk_filter      = request.GET.get('bhk', '').strip()
    furnish         = request.GET.get('furnish', '').strip()
    # zone_filter     = request.GET.get('zone', '').strip()
    ownership       = request.GET.get('ownership', '').strip()
    negotiable      = request.GET.get('negotiable', '').strip()
    from_date       = request.GET.get('from_date', '').strip()
    to_date         = request.GET.get('to_date', '').strip()

    # ── Apply filters ────────────────────────────────────────────────────────
    properties = all_properties

    if search_query:
        properties = properties.filter(
            Q(property_title__icontains=search_query)  |
            Q(city__icontains=search_query)           |
            Q(locality__icontains=search_query)       |
            Q(listed_by_name__icontains=search_query)     |
            Q(bhk__icontains=search_query)            |
            Q(building_name__icontains=search_query)
        )

    if prop_type:
        properties = properties.filter(property_type=prop_type)

    if bhk_filter:
        properties = properties.filter(bhk=bhk_filter)

    if furnish:
        properties = properties.filter(furnishing_status=furnish)

    # if zone_filter:
    #     properties = properties.filter(zone=zone_filter)

    if ownership:
        properties = properties.filter(ownership_status=ownership)

    if negotiable:
        properties = properties.filter(price_negotiable=negotiable)

    if from_date:
        properties = properties.filter(created_at__date__gte=from_date)

    if to_date:
        properties = properties.filter(created_at__date__lte=to_date)

    # ── Thumbnail + helper attributes for each filtered property ─────────────
    for prop in properties:
        prop.thumbnail = prop.images.first()

        prop.nearby_facilities_list = (
            [f.strip() for f in prop.nearby_facilities.split(',')]
            if prop.nearby_facilities else []
        )
        prop.amenities_list = (
            [a.strip() for a in prop.amenities.split(',')]
            if prop.amenities else []
        )
        prop.image_count = prop.images.count()
        prop.image_urls  = [img.image.url for img in prop.images.all()]

    # ════════════════════════════════════════════════════════════════════════
    # KPI STATS (Calculated using exact form model keys: selling_price, furnishing_status, etc.)
    # ════════════════════════════════════════════════════════════════════════
    total_count = all_properties.count()

    # ── Row 1 — Inventory ────────────────────────────────────────────────────
    total_negotiable  = all_properties.filter(price_negotiable='yes').count()
    total_furnished   = all_properties.filter(furnishing_status='fully').count()
    total_freehold    = all_properties.filter(ownership_status='Self Owned').count()
    total_with_images = all_properties.filter(images__isnull=False).distinct().count()

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    negotiable_pct = pct(total_negotiable,  total_count)
    furnished_pct  = pct(total_furnished,    total_count)
    freehold_pct   = pct(total_freehold,     total_count)
    images_pct     = pct(total_with_images, total_count)

    # ── Row 2 — Pricing (Using selling_price instead of expected_price) ───────
    price_agg = all_properties.aggregate(
        avg      = Avg('selling_price'),
        max_val  = Max('selling_price'),
        min_val  = Min('selling_price'),
        avg_sqft = Avg('price_per_sqft'),
        avg_area = Avg('builtup_area'),
    )
    avg_price      = price_agg['avg']
    max_price      = price_agg['max_val']
    min_price      = price_agg['min_val']
    avg_price_sqft = price_agg['avg_sqft']
    avg_builtup    = price_agg['avg_area']
    total_with_loan = all_properties.filter(property_loan='yes').count()

    # ── Row 3 — Legal & Status (Using government_tax instead of government_tax_dues) ───
    no_dispute_count  = all_properties.filter(any_legal_dispute='no').count()
    dispute_count     = all_properties.filter(any_legal_dispute='yes').count()
    tax_pending_count = all_properties.filter(government_tax='yes').count()
    tenant_occupied   = all_properties.filter(existing_tenants='yes').count()
    premium_count     = all_properties.filter(selling_price__gte=10000000).count()    # >= 1 Cr

    # ── Row 4 — Listing Quality ──────────────────────────────────────────────
    with_video_count = (
        all_properties
        .exclude(property_video__isnull=True)
        .exclude(property_video='')
        .count()
    )
    with_floor_plan = (
        all_properties
        .exclude(floor_plan__isnull=True)
        .exclude(floor_plan='')
        .count()
    )
    with_owner_count = (
        all_properties
        .exclude(listed_by_name__isnull=True)
        .exclude(listed_by_name='')
        .count()
    )
    budget_count = all_properties.filter(selling_price__lt=3000000).count()            # < 30 L

    # ── Charts ───────────────────────────────────────────────────────────────
    property_type_counts = dict(
        all_properties.values('property_type')
        .annotate(count=Count('id'))
        .values_list('property_type', 'count')
    )
    bhk_counts = dict(
        all_properties.values('bhk')
        .annotate(count=Count('id'))
        .values_list('bhk', 'count')
    )
    fully_furnished = all_properties.filter(furnishing_status='fully').count()
    semi_furnished  = all_properties.filter(furnishing_status='semi').count()
    unfurnished     = all_properties.filter(furnishing_status='unfurnished').count()

    # zone_counts = dict(
    #     all_properties.values('zone')
    #     .annotate(count=Count('id'))
    #     .values_list('zone', 'count')
    # )

    # ── Unique values for Select2 searchable dropdowns ───────────────────────
    unique_prop_types  = list(
        all_properties.values_list('property_type', flat=True)
        .distinct().order_by('property_type')
    )
    unique_bhk_values  = list(
        all_properties.values_list('bhk', flat=True)
        .distinct().order_by('bhk')
    )
    # unique_zones       = list(
    #     all_properties.values_list('zone', flat=True)
    #     .distinct().order_by('zone')
    # )
    unique_cities      = list(
        all_properties.values_list('city', flat=True)
        .distinct().order_by('city')
    )

    try:
        uploaded_files = (
            all_properties
            .exclude(upload_file_name__isnull=True)
            .exclude(upload_file_name='')
            .values_list('upload_file_name', flat=True)
            .distinct()
        )
    except Exception:
        uploaded_files = []

    # ── Context ──────────────────────────────────────────────────────────────
    context = {
        'admin_obj'  : admin_obj,
        'properties' : properties,

        # Counts
        'filtered_count' : properties.count(),
        'total_count'    : total_count,

        # Active search params
        'search_query'   : search_query,
        'prop_type_query': prop_type,
        'bhk_query'      : bhk_filter,
        'furnish_query'  : furnish,
        
        'ownership_query': ownership,
        'negotiable_query': negotiable,
        'from_date'      : from_date,
        'to_date'        : to_date,

        # Row 1 — Inventory
        'total_negotiable' : total_negotiable,
        'total_furnished'  : total_furnished,
        'total_freehold'   : total_freehold,
        'total_with_images': total_with_images,
        'negotiable_pct'   : negotiable_pct,
        'furnished_pct'    : furnished_pct,
        'freehold_pct'     : freehold_pct,
        'images_pct'       : images_pct,

        # Row 2 — Pricing
        'avg_price'       : avg_price,
        'max_price'       : max_price,
        'min_price'       : min_price,
        'avg_price_sqft'  : avg_price_sqft,
        'total_with_loan' : total_with_loan,

        # Row 3 — Legal
        'no_dispute_count' : no_dispute_count,
        'dispute_count'    : dispute_count,
        'tax_pending_count': tax_pending_count,
        'tenant_occupied'  : tenant_occupied,
        'avg_builtup'      : avg_builtup,
        'premium_count'    : premium_count,

        # Row 4 — Quality
        'with_video_count': with_video_count,
        'with_floor_plan' : with_floor_plan,
        'with_owner_count': with_owner_count,
        'budget_count'    : budget_count,

        # Charts
        'property_type_counts': property_type_counts,
        'bhk_counts'          : bhk_counts,
        'fully_furnished'     : fully_furnished,
        'semi_furnished'      : semi_furnished,
        'unfurnished'         : unfurnished,
        

        # Select2 unique options
        'unique_prop_types' : unique_prop_types,
        'unique_bhk_values' : unique_bhk_values,

        'unique_cities'     : unique_cities,

        'uploaded_files': uploaded_files,
    }

    return render(request, 'admin_user/Reports/Resale/residential_resale_list.html', context)






def resale_residential_edit(request, id):
    # ── 1. Session Check ──────────────────────────────────
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    # ── 2. Get Property Instance ──────────────────────────
    prop = get_object_or_404(ResaleResidentialProperty, id=id)

    # ── 3. Handle POST (UPDATE DATA) ──────────────────────
    if request.method == "POST":
        # --- A. Safe Numeric Conversion & Validation ---
        try: builtup_val = float(request.POST.get('builtup_area') or 0.0)
        except ValueError: builtup_val = 0.0

        try: carpet_val = float(request.POST.get('carpet_area') or 0.0)
        except ValueError: carpet_val = 0.0

        try: price_val = float(request.POST.get('expected_price') or 0.0)
        except ValueError: price_val = 0.0

        try:
            plot_raw = request.POST.get('plot_area')
            plot_val = float(plot_raw) if plot_raw and plot_raw.strip() else None
        except ValueError: plot_val = None

        try:
            loan_raw = request.POST.get('loan_amount')
            loan_val = float(loan_raw) if loan_raw and loan_raw.strip() else None
        except ValueError: loan_val = None

        try:
            tax_raw = request.POST.get('pending_tax_amount')
            tax_val = float(tax_raw) if tax_raw and tax_raw.strip() else None
        except ValueError: tax_val = None

        # --- B. Map Fields Sequentially as per Database Model Layout ---
        # Basic Information
        prop.property_title   = request.POST.get('property_title') or None  # Empty allows fallback title builder
        prop.property_type    = request.POST.get('property_type')
        prop.zone             = request.POST.get('zone')
        prop.society_type     = request.POST.get('society_type')
        prop.water_type       = request.POST.get('water_type')
        prop.furnishing_type  = request.POST.get('furnishing_type')
        prop.age_of_property  = request.POST.get('age_of_property')
        prop.facing_direction = request.POST.get('facing_direction')

        # Property Configuration
        prop.bhk             = request.POST.get('bhk')
        prop.bathrooms       = int(request.POST.get('bathrooms') or 1)
        prop.balconies       = int(request.POST.get('balconies') or 0)
        prop.covered_parking = int(request.POST.get('covered_parking') or 0)
        prop.open_parking    = int(request.POST.get('open_parking') or 0)

        # Measurements
        prop.builtup_area = builtup_val
        prop.carpet_area  = carpet_val
        prop.plot_area    = plot_val
        prop.floor_no     = int(request.POST.get('floor_no') or 0)
        prop.total_floors = int(request.POST.get('total_floors') or 0)

        # Ownership & Legal
        prop.ownership_type      = request.POST.get('ownership_type')
        prop.num_owners          = request.POST.get('num_owners')
        prop.loan_on_property    = request.POST.get('loan_on_property', 'no')
        prop.loan_amount         = loan_val if prop.loan_on_property == 'yes' else None
        prop.existing_tenants    = request.POST.get('existing_tenants', 'no')
        prop.tenant_details      = request.POST.get('tenant_details') if prop.existing_tenants == 'yes' else None
        prop.any_legal_dispute   = request.POST.get('any_legal_dispute', 'no')
        prop.dispute_details     = request.POST.get('dispute_details') if prop.any_legal_dispute == 'yes' else None
        prop.government_tax_dues = request.POST.get('government_tax_dues', 'no')
        prop.pending_tax_amount  = tax_val if prop.government_tax_dues == 'yes' else None

        # Pricing & Description
        prop.expected_price       = price_val
        prop.price_negotiable     = request.POST.get('price_negotiable', 'yes')
        prop.brokerage            = request.POST.get('brokerage') or None
        prop.brokerage_percentage = request.POST.get('brokerage_percentage') or None
        prop.manual_brokerage     = request.POST.get('manual_brokerage') or None
        prop.property_description = request.POST.get('property_description')
        prop.property_summary = request.POST.get('property_summary')
        prop.user_description = request.POST.get('user_description')

        # Amenities & Facilities
        prop.nearby_facilities = ', '.join(request.POST.getlist('facilities[]'))
        prop.amenities         = ', '.join(request.POST.getlist('amenities[]'))

        # Address
        prop.city             = request.POST.get('city')
        prop.locality         = request.POST.get('locality')
        prop.building_name    = request.POST.get('building_name') or None
        prop.complete_address = request.POST.get('complete_address')

        # Owner Contact Info
        prop.owner_name         = request.POST.get('owner_name')
        prop.owner_contact      = request.POST.get('owner_contact')
        prop.owner_email        = request.POST.get('owner_email')
        prop.owner_role         = request.POST.get('owner_role')
        prop.residential_status = request.POST.get('residential_status')

        # File Upload Fields
        if request.FILES.get('floor_plan'):
            prop.floor_plan = request.FILES.get('floor_plan')
        if request.FILES.get('property_video'):
            prop.property_video = request.FILES.get('property_video')

        # Save updates safely (executes system automated formulas for title, price/sqft, and programmatic FAQs)
        prop.save()

        # --- C. Handle Dynamic Image Deletion Array ---
        deleted_images = request.POST.getlist('deleted_images[]')
        if deleted_images:
            ResalePropertyImage.objects.filter(id__in=deleted_images, property=prop).delete()

        # --- D. Handle New Image Queue Processing Array ---
        new_images = request.FILES.getlist('property_images')
        for img in new_images:
            ResalePropertyImage.objects.create(property=prop, image=img)

        return JsonResponse({
            "status": "success",
            "message": "Resale Residential Property Updated Successfully"
        })

    # ── 4. Handle GET (LOAD EDIT FORM) ───────────────────
    # Stored string splitters mapping tags back to array lists for template rendering checks
    prop_facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []
    prop_amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []
    existing_images = prop.images.all()

    context = {
        "prop": prop,
        "prop_facilities_list": prop_facilities_list,
        "prop_amenities_list": prop_amenities_list,
        "existing_images": existing_images,
        "admin_obj": uploader['admin_obj'],
        "user_obj": uploader['user_obj'],
    }
    return render(request, 'admin_user/Reports/Resale/residential_resale_edit.html', context)





def resale_residential_view(request, pk):
    # ── Session Check ──
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    # Fetch Resale Residential Object via PK
    prop = get_object_or_404(ResaleResidentialProperty, pk=pk)

    # Convert comma-separated string datasets into lists for badge generation loop arrays
    facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []
    amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []

    context = {
        "prop": prop,
        "images": prop.images.all(),
        "facilities_list": facilities_list,
        "amenities_list": amenities_list,
        "admin_obj": uploader['admin_obj'],
        "user_obj": uploader['user_obj'],
    }
    
    return render(request, 'admin_user/Reports/Resale/residential_resale_view.html', context)

# ─────────────────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────────────────


# ═════════════════════════════════════════════════════════════
# SINGLE SOFT DELETE
# ═════════════════════════════════════════════════════════════
@require_POST
def resale_residential_delete(request, pk):
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    deleter_name = _get_deleter_name(request)

    prop = get_object_or_404(
        ResaleResidentialProperty,
        pk=pk,
        is_deleted=False
    )

    prop.is_deleted = True
    prop.deleted_at = timezone.now()
    prop.deleted_by = deleter_name
    prop.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by'])

    return JsonResponse({
        "status": "success",
        "message": f"{prop.property_id} moved to Recycle Bin successfully!"
    })


# ═════════════════════════════════════════════════════════════
# RESTORE SINGLE PROPERTY
# ═════════════════════════════════════════════════════════════
@require_POST
def resale_restore(request, id):

    uploader = _get_uploader(request)
    if uploader is None:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized access.'
        })

    prop = get_object_or_404(
        ResaleResidentialProperty,
        id=id
    )

    prop.is_deleted = False
    prop.deleted_at = None
    prop.deleted_by = None
    prop.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by'])

    return JsonResponse({
        'status': 'success',
        'message': f'{prop.property_id} restored successfully!'
    })


# ═════════════════════════════════════════════════════════════
# HARD DELETE SINGLE PROPERTY
# ═════════════════════════════════════════════════════════════
@require_POST
def resale_hard_delete(request, id):

    uploader = _get_uploader(request)
    if uploader is None:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized access.'
        })

    prop = get_object_or_404(
        ResaleResidentialProperty,
        id=id
    )

    property_id = prop.property_id

    # Delete related images automatically
    prop.images.all().delete()

    # Delete property
    prop.delete()

    return JsonResponse({
        'status': 'success',
        'message': f'{property_id} permanently deleted!'
    })


# ═════════════════════════════════════════════════════════════
# BULK DELETE / RECYCLE BIN SYSTEM
# ═════════════════════════════════════════════════════════════
@require_POST
def resale_residential_bulk_delete(request):

    uploader = _get_uploader(request)
    if uploader is None:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized access.'
        })

    deleter_name = _get_deleter_name(request)

    try:
        data = json.loads(request.body)

        delete_type = data.get('delete_type')

        # Only active records
        properties = ResaleResidentialProperty.objects.filter(
            is_deleted=False
        )

        # ═══════════════════════════════════════
        # DELETE ALL
        # ═══════════════════════════════════════
        if delete_type == 'delete_all':

            count = properties.count()

            properties.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # CURRENT PAGE DELETE
        # ═══════════════════════════════════════
        elif delete_type == 'current_page':

            page_ids = data.get('page_ids', [])

            target_props = properties.filter(id__in=page_ids)

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # DATE RANGE DELETE
        # ═══════════════════════════════════════
        elif delete_type == 'date_range':

            from_date = data.get('from_date')
            to_date = data.get('to_date')

            if not from_date or not to_date:
                return JsonResponse({
                    'status': 'error',
                    'message': 'From date and To date are required.'
                })

            target_props = properties.filter(
                created_at__date__range=[from_date, to_date]
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} resale properties deleted between selected dates.'
            })

        # ═══════════════════════════════════════
        # LAST 30 DAYS
        # ═══════════════════════════════════════
        elif delete_type == 'latest_month':

            thirty_days_ago = timezone.now() - timedelta(days=30)

            target_props = properties.filter(
                created_at__gte=thirty_days_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} recent resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # OLD DATA
        # ═══════════════════════════════════════
        elif delete_type == 'old_data':

            six_months_ago = timezone.now() - timedelta(days=180)

            target_props = properties.filter(
                created_at__lt=six_months_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} old resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # DELETE BY UPLOADER
        # ═══════════════════════════════════════
        elif delete_type == 'by_uploader':

            uploader_text = data.get('uploader_text', '').strip()

            if not uploader_text:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Uploader name is required.'
                })

            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader_text) |
                Q(uploaded_by_email__icontains=uploader_text) |
                Q(uploaded_by_role__icontains=uploader_text) |
                Q(uploaded_by_contact__icontains=uploader_text)
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # DELETE BY FILE
        # ═══════════════════════════════════════
        elif delete_type == 'by_file':

            file_name = data.get('file_name', '').strip()

            if not file_name:
                return JsonResponse({
                    'status': 'error',
                    'message': 'File name is required.'
                })

            target_props = properties.filter(
                upload_file_name__iexact=file_name
            )

            count = target_props.count()

            if count == 0:
                return JsonResponse({
                    'status': 'error',
                    'message': f'No properties found for file: {file_name}'
                })

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'{count} resale properties moved to recycle bin.'
            })

        # ═══════════════════════════════════════
        # PERMANENT DELETE RECYCLE BIN
        # ═══════════════════════════════════════
        elif delete_type == 'permanent_deleted':

            deleted_props = ResaleResidentialProperty.objects.filter(
                is_deleted=True
            )

            count = deleted_props.count()

            # Delete related images first
            for prop in deleted_props:
                prop.images.all().delete()

            deleted_props.delete()

            return JsonResponse({
                'status': 'success',
                'message': f'{count} properties permanently deleted.'
            })

        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Unknown delete criteria.'
            })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })




def generate_row_hash(row_values):
    """Generates a secure MD5 signature unique to the actual structural data values inside a row."""
    data_string = "|".join([str(v).strip().lower() for v in row_values if v is not None])
    return hashlib.md5(data_string.encode('utf-8')).hexdigest()





@csrf_exempt
@require_POST
def resale_residential_import_excel(request):
    # ---- 1. Session & Request Validation ----
    session_id = request.session.get('Admin_id')
    if not session_id:
        return JsonResponse({'status': 'error', 'message': 'Session expired. Please log in again.'})

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded. Please select a valid sheet.'})

    file_name_string = str(excel_file.name).strip()
    if not file_name_string.endswith('.xlsx'):
        return JsonResponse({'status': 'error', 'message': 'Invalid format. Only .xlsx extensions allowed.'})

    # ---- 2. Uploader Identity Injection ----
    try:
        uploader_obj = Admin_Login.objects.get(id=session_id)
        current_uploader_name = getattr(uploader_obj, 'name', '') or getattr(uploader_obj, 'username', 'System Admin')
        current_uploader_email = getattr(uploader_obj, 'email', 'admin@crm.com')
        current_uploader_contact = getattr(uploader_obj, 'phone', '') or getattr(uploader_obj, 'mobile', '0000000000')
        current_uploader_role = getattr(uploader_obj, 'role', 'admin')
        user_identity = current_uploader_email or current_uploader_name
    except Exception:
        current_uploader_name = "System Admin"
        current_uploader_email = "admin@crm.com"
        current_uploader_contact = "0000000000"
        current_uploader_role = "admin"
        user_identity = "System Admin"

    # ---- 3. Parse Excel Workbook ----
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Cannot open file: {e}'})

    # Helper to normalize headers
    def normalize_header(h):
        if not h: return ""
        s = str(h).lower().strip().replace('\n', ' ')
        s = re.sub(r'\(.*?\)', '', s).replace('*', '').strip()
        return s

    HEADER_MAP_ALIASES = {
        'property id': 'id',
        'listing status': 'listing_status',
        'approval status': 'approval_status',
        'listed by type': 'listed_by_type',
        'assigned to': 'assigned_to',
        'listed by id': 'listed_by_id',
        'listed by name': 'listed_by_name',
        'listed by email': 'listed_by_email',
        'listed by contact': 'listed_by_contact',
        'listed by role': 'listed_by_role',
        'property title': 'property_title',
        'property type': 'property_type',
        'flat/house/unit no': 'property_no',
        'flat/house no/unit no': 'property_no',
        'society type': 'society_type',
        'tower/wing number': 'wing_no',
        'floor no': 'floor_no',
        'water type': 'water_type',
        'furnishing status': 'furnishing_status',
        'age of property': 'property_age',
        'facing direction': 'facing_direction',
        'occupancy status': 'occupancy_status',
        'bhk': 'bhk',
        'super built-up area': 'super_builtup_area',
        'built-up area': 'builtup_area',
        'carpet area': 'carpet_area',
        'plot area': 'plot_area',
        'building configuration': 'building_configuration',
        'total floors constructed': 'total_floors',
        'bathrooms': 'bathrooms',
        'balconies': 'balconies',
        'covered parking': 'covered_parking',
        'open parking available?': 'open_parking',
        'no. of owners': 'no_of_owners',
        'ownership status': 'ownership_status',
        'ownership document type': 'ownership_document_type',
        'title clarity status': 'title_clarity_status',
        'encumbrance status': 'encumbrance_status',
        'loan on property?': 'property_loan',
        'loan amount': 'loan_amount',
        'existing tenants?': 'existing_tenants',
        'tenant details': 'tenant_details',
        'any legal dispute?': 'any_legal_dispute',
        'dispute details': 'dispute_details',
        'government tax dues?': 'government_tax',
        'pending tax amount': 'pending_tax_amount',
        'sanctioning / approving authority': 'sanctioning_authority',
        'selling price': 'selling_price',
        'price per sq.ft': 'price_per_sqft',
        'selling price negotiable?': 'price_negotiable',
        'brokerage': 'brokerage_percentage',
        'enter fixed brokerage': 'manual_brokerage',
        'fixed brokerage amount': 'manual_brokerage',
        'city': 'city',
        'locality': 'locality',
        'building/society name': 'building_name',
        'complete address': 'address',
        'property landmark': 'property_landmark',
        'state': 'state',
        'google maps link': 'google_maps_link',
        'latitude': 'latitude',
        'longitude': 'longitude',
        'amenities': 'amenities',
        'nearby facilities': 'nearby_facilities',
        'property description': 'user_description',
        'is property already listed elsewhere?': 'listed_elsewhere',
        'listed elsewhere?': 'listed_elsewhere',
        'portal name': 'portal_name'
    }

    # Dynamic Header Detection
    row1_vals = [normalize_header(cell.value) for cell in ws[1]]
    row2_vals = [normalize_header(cell.value) for cell in ws[2]] if ws.max_row >= 2 else []

    r1_matches = sum(1 for h in row1_vals if h in HEADER_MAP_ALIASES)
    r2_matches = sum(1 for h in row2_vals if h in HEADER_MAP_ALIASES)

    if r1_matches > r2_matches and r1_matches > 5:
        target_headers = row1_vals
        data_min_row = 2
    else:
        target_headers = row2_vals
        data_min_row = 3

    col_map = {}
    for idx, norm_header in enumerate(target_headers):
        if norm_header in HEADER_MAP_ALIASES:
            col_map[HEADER_MAP_ALIASES[norm_header]] = idx

    REQUIRED_FIELD_KEYS = ['property_type', 'property_no', 'bhk', 'selling_price', 'address', 'city', 'locality', 'state']
    missing_keys = [k for k in REQUIRED_FIELD_KEYS if k not in col_map]
    if missing_keys:
        return JsonResponse({
            'status': 'error',
            'message': f'Required core columns missing or misnamed in headers: {", ".join(missing_keys)}'
        }, status=400)

    def _is_missing(val):
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    parsed_rows = []
    skipped_empty_after_mapping = 0
    required_field_errors = []
    listed_by_mismatch_errors = []
    skipped_listed_by_mismatch = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_min_row, values_only=True), start=data_min_row):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def get_val(key, default=''):
            if key in col_map:
                v = row[col_map[key]]
                return str(v).strip() if v is not None and str(v).strip().lower() != 'none' else default
            return default

        obj_data = {
            'id': get_val('id', None),
            'listing_type': "Resale",
            'category': "Residential",
            'listing_status': get_val('listing_status', 'Draft'),
            'approval_status': get_val('approval_status', 'Pending'),
            'listed_by_type': get_val('listed_by_type'),
            'assigned_to': get_val('assigned_to'),
            'listed_by_id': get_val('listed_by_id'),
            'listed_by_name': get_val('listed_by_name'),
            'listed_by_email': get_val('listed_by_email'),
            'listed_by_contact': get_val('listed_by_contact'),
            'listed_by_role': get_val('listed_by_role'),
            'property_title': get_val('property_title'),
            'property_type': get_val('property_type'),
            'property_no': get_val('property_no'),
            'society_type': get_val('society_type'),
            'wing_no': get_val('wing_no'),
            'floor_no': get_val('floor_no'),
            'water_type': get_val('water_type'),
            'furnishing_status': get_val('furnishing_status'),
            'property_age': get_val('property_age'),
            'facing_direction': get_val('facing_direction'),
            'occupancy_status': get_val('occupancy_status'),
            'bhk': get_val('bhk'),
            'super_builtup_area': get_val('super_builtup_area'),
            'builtup_area': get_val('builtup_area'),
            'carpet_area': get_val('carpet_area'),
            'plot_area': get_val('plot_area'),
            'building_configuration': get_val('building_configuration'),
            'total_floors': get_val('total_floors'),
            'bathrooms': get_val('bathrooms'),
            'balconies': get_val('balconies'),
            'covered_parking': get_val('covered_parking'),
            'open_parking': get_val('open_parking'),
            'no_of_owners': get_val('no_of_owners'),
            'ownership_status': get_val('ownership_status'),
            'ownership_document_type': get_val('ownership_document_type'),
            'title_clarity_status': get_val('title_clarity_status'),
            'encumbrance_status': get_val('encumbrance_status'),
            'property_loan': get_val('property_loan', 'no').lower(),
            'loan_amount': get_val('loan_amount'),
            'existing_tenants': get_val('existing_tenants', 'no').lower(),
            'tenant_details': get_val('tenant_details'),
            'any_legal_dispute': get_val('any_legal_dispute', 'no').lower(),
            'dispute_details': get_val('dispute_details'),
            'government_tax': get_val('government_tax', 'no').lower(),
            'pending_tax_amount': get_val('pending_tax_amount'),
            'sanctioning_authority': get_val('sanctioning_authority'),
            'selling_price': get_val('selling_price'),
            'price_per_sqft': get_val('price_per_sqft'),
            'price_negotiable': get_val('price_negotiable', 'yes').lower(),
            'brokerage_percentage': get_val('brokerage_percentage'),
            'manual_brokerage': get_val('manual_brokerage'),
            'city': get_val('city'),
            'locality': get_val('locality'),
            'building_name': get_val('building_name'),
            'address': get_val('address'),
            'property_landmark': get_val('property_landmark'),
            'state': get_val('state'),
            'google_maps_link': get_val('google_maps_link'),
            'latitude': get_val('latitude'),
            'longitude': get_val('longitude'),
            'nearby_facilities': get_val('nearby_facilities'),
            'amenities': get_val('amenities'),
            'user_description': get_val('user_description'),
            'listed_elsewhere': get_val('listed_elsewhere'),
            'portal_name': get_val('portal_name'),
        }

        # 1. Required Field Validation Guard
        missing_fields = [k for k in REQUIRED_FIELD_KEYS if _is_missing(obj_data.get(k))]
        if missing_fields:
            required_field_errors.append({"row": row_idx, "missing_fields": missing_fields})
            continue

        # 2. Uploader/Agent Verification Guard (Matches Rental Logic)
        l_role = str(obj_data.get('listed_by_role', '')).strip().title()
        l_email = str(obj_data.get('listed_by_email', '')).strip().lower()
        l_contact = str(obj_data.get('listed_by_contact', '')).strip()
        l_name = str(obj_data.get('listed_by_name', '')).strip()
        l_id = str(obj_data.get('listed_by_id', '')).strip()

        assigned_to = ""
        if l_email or l_contact or l_name or l_id:
            is_registered = False
            matched_user = None

            if l_role.lower() == 'admin':
                admin_query = Q()
                if l_email: admin_query &= Q(email=l_email)
                if l_contact: admin_query &= Q(phone=l_contact)
                if l_name: admin_query &= Q(name__iexact=l_name)
                if l_id and l_id.isdigit(): admin_query &= Q(id=l_id)
                
                if admin_query and Admin_Login.objects.filter(admin_query).exists():
                    is_registered = True
            else:
                user_query = Q()
                if l_email: user_query &= Q(user_email=l_email)
                if l_contact: user_query &= Q(user_phone=l_contact)
                if l_name: user_query &= Q(user_name__iexact=l_name)
                if l_id: user_query &= Q(user_id=l_id)

                if user_query:
                    if l_role:
                        matched_user = User_Details.objects.filter(user_query, user_role__iexact=l_role).first()
                    else:
                        matched_user = User_Details.objects.filter(user_query).first()
                    
                    if matched_user:
                        is_registered = True
                        assigned_to = f"{matched_user.id}-{matched_user.user_role}"
                        obj_data['assigned_to'] = assigned_to

            # Fallback / Error Capture for Unregistered Users
            if not is_registered:
                # If testing with template sample accounts like Rahul Sharma, assign to system uploader instead of blocking hard
                obj_data['listed_by_name'] = current_uploader_name
                obj_data['listed_by_email'] = current_uploader_email
                obj_data['listed_by_contact'] = current_uploader_contact
                obj_data['listed_by_role'] = current_uploader_role

        # Safe Numerical Type Casts
        try: obj_data['super_builtup_area'] = Decimal(str(obj_data['super_builtup_area'])) if obj_data['super_builtup_area'] else None
        except Exception: obj_data['super_builtup_area'] = None
        try: obj_data['builtup_area'] = Decimal(str(obj_data['builtup_area'])) if obj_data['builtup_area'] else Decimal('0.0')
        except Exception: obj_data['builtup_area'] = Decimal('0.0')
        try: obj_data['carpet_area'] = Decimal(str(obj_data['carpet_area'])) if obj_data['carpet_area'] else Decimal('0.0')
        except Exception: obj_data['carpet_area'] = Decimal('0.0')
        try: obj_data['plot_area'] = Decimal(str(obj_data['plot_area'])) if obj_data['plot_area'] else None
        except Exception: obj_data['plot_area'] = None
        try: obj_data['selling_price'] = Decimal(str(obj_data['selling_price']))
        except Exception: obj_data['selling_price'] = Decimal('0.0')
        try: obj_data['price_per_sqft'] = Decimal(str(obj_data['price_per_sqft'])) if obj_data['price_per_sqft'] else None
        except Exception: obj_data['price_per_sqft'] = None
        try: obj_data['loan_amount'] = Decimal(str(obj_data['loan_amount'])) if obj_data['loan_amount'] else None
        except Exception: obj_data['loan_amount'] = None
        try: obj_data['pending_tax_amount'] = Decimal(str(obj_data['pending_tax_amount'])) if obj_data['pending_tax_amount'] else None
        except Exception: obj_data['pending_tax_amount'] = None
        try: obj_data['manual_brokerage'] = int(float(obj_data['manual_brokerage'])) if obj_data['manual_brokerage'] else 0
        except Exception: obj_data['manual_brokerage'] = 0
        try: obj_data['floor_no'] = int(float(obj_data['floor_no'])) if obj_data['floor_no'] else 1
        except Exception: obj_data['floor_no'] = 1
        try: obj_data['total_floors'] = int(float(obj_data['total_floors'])) if obj_data['total_floors'] else 1
        except Exception: obj_data['total_floors'] = 1
        try: obj_data['bathrooms'] = int(float(obj_data['bathrooms'])) if obj_data['bathrooms'] else 0
        except Exception: obj_data['bathrooms'] = 0
        try: obj_data['balconies'] = int(float(obj_data['balconies'])) if obj_data['balconies'] else 0
        except Exception: obj_data['balconies'] = 0
        try: obj_data['covered_parking'] = int(float(obj_data['covered_parking'])) if obj_data['covered_parking'] else 0
        except Exception: obj_data['covered_parking'] = 0

        parsed_rows.append({'row_idx': row_idx, 'data': obj_data})

    wb.close()

    if required_field_errors:
        return JsonResponse({
            "status": "error",
            "message": f"Upload Denied: {len(required_field_errors)} row(s) are missing mandatory fields.",
            "row_errors": required_field_errors
        }, status=400)

    if not parsed_rows:
        return JsonResponse({
            "status": "error",
            "message": "0 usable data rows found in uploaded Excel file.",
        }, status=400)

    # ---- 4. Database Writing & Fingerprint Duplicate Guard Engine ----
    created, updated, skipped = 0, 0, skipped_empty_after_mapping + skipped_listed_by_mismatch
    duplicate_blocked_rows = []
    errors = []

    for item in parsed_rows:
        o_data = item['data']
        row_idx = item['row_idx']

        input_property_no = str(o_data.get('property_no', '')).strip()
        input_building_name = str(o_data.get('building_name', '')).strip()
        input_locality = str(o_data.get('locality', '')).strip()
        input_city = str(o_data.get('city', '')).strip()

        input_listed_by_id = str(o_data.get('listed_by_id', '')).strip()
        input_listed_by_email = str(o_data.get('listed_by_email', '')).strip().lower()
        input_listed_by_contact = str(o_data.get('listed_by_contact', '')).strip()

        # Generate unique property fingerprint key
        fingerprint_key = generate_property_fingerprint(
            input_property_no,
            input_building_name,
            input_locality,
            o_data.get('pincode', '') if 'pincode' in o_data else ''
        )

        # Direct case-insensitive match on the same unit in locality/building
        direct_duplicates = ResaleResidentialProperty.objects.filter(
            is_deleted=False,
            property_no__iexact=input_property_no,
            locality__iexact=input_locality,
            city__iexact=input_city
        )
        if input_building_name:
            direct_duplicates = direct_duplicates.filter(building_name__iexact=input_building_name)

        existing_duplicates = (
            ResaleResidentialProperty.objects.filter(property_unique_key=fingerprint_key, is_deleted=False)
            | direct_duplicates
        ).distinct()

        is_dup_flag = False
        dup_group_id = fingerprint_key
        total_dup_count = 1
        hard_blocked = False

        if existing_duplicates.exists():
            for existing_prop in existing_duplicates:
                same_id = (existing_prop.listed_by_id and input_listed_by_id and existing_prop.listed_by_id.strip() == input_listed_by_id)
                same_email = (existing_prop.listed_by_email and input_listed_by_email and existing_prop.listed_by_email.strip().lower() == input_listed_by_email)
                same_contact = (existing_prop.listed_by_contact and input_listed_by_contact and existing_prop.listed_by_contact.strip() == input_listed_by_contact)

                if same_id or same_email or same_contact:
                    hard_blocked = True
                    break

            if hard_blocked:
                duplicate_blocked_rows.append(
                    f"Row {row_idx}: Duplicate Blocked — Unit '{input_property_no}' is already listed by this user."
                )
                skipped += 1
                continue

            # Level 2: Different user listing the exact same unit -> allow save & flag as duplicate
            is_dup_flag = True
            total_dup_count = existing_duplicates.count() + 1
            existing_duplicates.update(
                is_duplicate=True,
                duplicate_count=total_dup_count,
                duplicate_group_id=dup_group_id
            )

        o_data["property_unique_key"] = fingerprint_key
        o_data["is_duplicate"] = is_dup_flag
        o_data["duplicate_count"] = total_dup_count
        o_data["duplicate_group_id"] = dup_group_id if is_dup_flag else None

        o_data["upload_file_name"] = file_name_string
        o_data["uploaded_by_name"] = current_uploader_name
        o_data["uploaded_by_email"] = current_uploader_email
        o_data["uploaded_by_contact"] = current_uploader_contact
        o_data["uploaded_by_role"] = current_uploader_role

        prop_id_val = o_data.pop('id', None)
        try:
            if prop_id_val and ResaleResidentialProperty.objects.filter(id=prop_id_val).exists():
                ResaleResidentialProperty.objects.filter(id=prop_id_val).update(**o_data)
                updated += 1
            else:
                ResaleResidentialProperty.objects.create(**o_data)
                created += 1
        except Exception as e:
            errors.append(f"Row {row_idx} processing failure: {str(e)}")

    errors.extend(duplicate_blocked_rows)

    return JsonResponse({
        "status": "success" if not errors else "partial_error",
        "message": f"{created} Created | {updated} Updated | {skipped} Skipped due to system rules.",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    })


def resale_residential_sample_excel(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return redirect('login')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resale Template"
    
    # Ensure grid lines show up clearly
    ws.views.sheetView[0].showGridLines = True

    # ── STYLING ENGINE DEFINITIONS ────────────────────────────────────────
    font_family = "Segoe UI"
    
    # Fonts
    section_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    data_font = Font(name=font_family, size=11, bold=False, color="000000")
    
    # Fills
    section_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Corporate Indigo Block
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Light Slate Header Row
    
    # Borders
    thin_border = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── HARDCODED HORIZONTAL BLOCKS (Synchronized with DB layout) ──────────
    # Arranged precisely sequentially side-by-side matching your database model structure
    sections_config = [
        {
            "title": "STEP 1: BASIC INFO & CONFIGURATION",
            "fields": [
                ('property_title', 'Auto Generated By System'), ('property_type', 'apartment'), ('zone', 'north'), ('society_type', 'gated community'),
                ('water_type', 'municipal'), ('furnishing_type', 'semi-furnished'), ('age_of_property', '1-3 years'),
                ('facing_direction', 'North-East'), ('bhk', '3 BHK'), ('bathrooms', 2), ('balconies', 1),
                ('covered_parking', 1), ('open_parking', 0), ('builtup_area', 1450.00), ('carpet_area', 1120.00),
                ('plot_area', ''), ('floor_no', 4), ('total_floors', 12)
            ]
        },
        {
            "title": "STEP 2: LEGAL & PRICING DETAILS",
            "fields": [
                ('ownership_type', 'freehold'), ('num_owners', '1'), ('loan_on_property', 'no'),
                ('loan_amount', ''), ('existing_tenants', 'no'), ('tenant_details', ''),
                ('any_legal_dispute', 'no'), ('dispute_details', ''), ('government_tax_dues', 'no'),
                ('pending_tax_amount', ''), ('expected_price', 7500000.00), ('price_per_sqft', 5172.41),
                ('price_negotiable', 'yes'), ('brokerage', 'no'), ('brokerage_percentage', ''),
                ('manual_brokerage', ''), ('property_description', 'Spacious apartment layout.')
            ]
        },
        {
            "title": "STEP 3: AMENITIES & LOCATION",
            "fields": [
                ('nearby_facilities', 'School, Metro'), ('amenities', 'Lift, Security, Gym'), ('city', 'Nagpur'),
                ('locality', 'Dharampeth'), ('building_name', 'Sunshine Towers'),
                ('complete_address', 'Flat 402, Sunshine Towers, Dharampeth, Nagpur'), ('owner_name', 'Rahul Sharma'),
                ('owner_contact', '9876543210'), ('owner_email', 'rahul.sharma@example.com'),
                ('owner_role', 'Owner'), ('residential_status', 'resident')
            ]
        }
    ]

    # Flatten out arrays to compute complete ranges
    total_columns = sum(len(sec["fields"]) for sec in sections_config)
    
    # ── GENERATE ROW 1: MERGED SECTION BANNERS ────────────────────────────
    current_col = 1
    for sec in sections_config:
        sec_length = len(sec["fields"])
        start_cell = ws.cell(row=1, column=current_col)
        end_col_idx = current_col + sec_length - 1
        
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end_col_idx)
        
        start_cell.value = sec["title"]
        start_cell.font = section_font
        start_cell.fill = section_fill
        start_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply borders across merged blocks cleanly
        for col in range(current_col, end_col_idx + 1):
            ws.cell(row=1, column=col).border = cell_border
            
        current_col += sec_length
    ws.row_dimensions[1].height = 30

    # ── GENERATE ROW 2 & 3: DB FIELD HEADERS & SAMPLE DATA ────────────────
    current_col = 1
    for sec in sections_config:
        for field_name, sample_val in sec["fields"]:
            # Row 2: Database Headers
            header_cell = ws.cell(row=2, column=current_col, value=field_name)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = center_align
            header_cell.border = cell_border
            
            # Row 3: Sample Values
            data_cell = ws.cell(row=3, column=current_col, value=sample_val)
            data_cell.font = data_font
            data_cell.alignment = center_align if isinstance(sample_val, (int, float)) else left_align
            data_cell.border = cell_border
            
            current_col += 1
            
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 22

    # Auto-fit column widths dynamically to prevent text truncation
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    # ── TRANSMIT ATTACHMENT RESPONSE ──────────────────────────────────────
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resale_residential_listing_template.xlsx"'
    wb.save(response)
    return response








# ── MASTER DATA CONFIGURATION MATRIX (DEFINED ONCE) ──────────────────
EXPORT_COLUMNS_BLUEPRINT = [
    # System Control Section
    {'field': 'property_id', 'section_name': 'SYSTEM CONTROL & IDENTIFICATION'},
    {'field': 'property_title', 'section_name': 'SYSTEM CONTROL & IDENTIFICATION'},
    
    # Step 1 Section
    {'field': 'property_type', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'zone', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'society_type', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'water_type', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'furnishing_type', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'age_of_property', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'facing_direction', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'bhk', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'bathrooms', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'balconies', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'covered_parking', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'open_parking', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'builtup_area', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'carpet_area', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'plot_area', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'floor_no', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    {'field': 'total_floors', 'section_name': 'STEP 1: BASIC INFO & CONFIGURATION'},
    
    # Step 2 Section
    {'field': 'ownership_type', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'num_owners', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'loan_on_property', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'loan_amount', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'existing_tenants', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'tenant_details', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'any_legal_dispute', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'dispute_details', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'government_tax_dues', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'pending_tax_amount', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'expected_price', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'price_per_sqft', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'price_negotiable', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'brokerage', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'brokerage_percentage', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'manual_brokerage', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    {'field': 'property_description', 'section_name': 'STEP 2: LEGAL & PRICING DETAILS'},
    
    # Step 3 Section
    {'field': 'nearby_facilities', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'amenities', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'city', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'locality', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'building_name', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'complete_address', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'owner_name', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'owner_contact', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'owner_email', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'owner_role', 'section_name': 'STEP 3: AMENITIES & LOCATION'},
    {'field': 'residential_status', 'section_name': 'STEP 3: AMENITIES & LOCATION'},

    # Step 4 Auditing Meta
    {'field': 'uploaded_by_name', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'uploaded_by_email', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'uploaded_by_contact', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'uploaded_by_role', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'upload_file_name', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'created_at', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'updated_at', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'is_deleted', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'deleted_at', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'},
    {'field': 'deleted_by', 'section_name': 'STEP 4: PHOTOS & PUBLISH SYSTEM'}
]


def export_resale_csv(request):
    """
    Exports all Resale Residential properties to CSV pulling directly from the master matrix sequence.
    """
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="resale_residential_complete_database.csv"'

    writer = csv.writer(response)
    
    # Extract headers straight out of the master dictionary array
    headers = [item['field'] for item in EXPORT_COLUMNS_BLUEPRINT]
    writer.writerow(headers)

    queryset = ResaleResidentialProperty.objects.all()
    for prop in queryset:
        row = []
        for item in EXPORT_COLUMNS_BLUEPRINT:
            val = getattr(prop, item['field'], '')
            if val is None:
                val = ''
            elif isinstance(val, Decimal):
                val = float(val)
            elif hasattr(val, 'strftime'):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            row.append(val)
        writer.writerow(row)

    return response


def export_resale_excel(request):
    """
    Generates a stylized Excel sheet pulling from the master configuration blueprint array.
    """
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    wb = Workbook()
    ws = wb.active
    ws.title = "Resale Template"
    ws.views.sheetView[0].showGridLines = True

    # ── SYSTEM PALETTE CONFIGURATIONS ──
    sys_control_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Slate
    step1_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")       # Dark Gray
    step2_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")       # Deep Purple
    step3_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")       # Slate Midnight
    meta_audit_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")    # Mid Gray

    zebra_even_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_step_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_db_field = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data_row = Font(name="Segoe UI", size=10, bold=False, color="111827")

    thin_border_side = Side(border_style="thin", color="E5E7EB")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center")

    # ── RENDER HEADERS FROM THE SINGLE MASTER BLUEPRINT MATRIX ──
    for col_idx, column_meta in enumerate(EXPORT_COLUMNS_BLUEPRINT, start=1):
        # Render Row 1 Section Group Block
        cell_step = ws.cell(row=1, column=col_idx, value=column_meta['section_name'])
        cell_step.font = font_step_title
        cell_step.alignment = center_alignment
        cell_step.border = grid_border
        
        # Style sections dynamically based on identity mapping
        if "SYSTEM CONTROL" in column_meta['section_name']:
            cell_step.fill = sys_control_fill
        elif "STEP 1" in column_meta['section_name']:
            cell_step.fill = step1_fill
        elif "STEP 2" in column_meta['section_name']:
            cell_step.fill = step2_fill
        elif "STEP 3" in column_meta['section_name']:
            cell_step.fill = step3_fill
        else:
            cell_step.fill = meta_audit_fill

        # Render Row 2 Property Column Names
        cell_field = ws.cell(row=2, column=col_idx, value=column_meta['field'])
        cell_field.font = font_db_field
        cell_field.alignment = left_alignment
        cell_field.fill = meta_audit_fill
        cell_field.border = grid_border

    # ── INJECT RETRIEVED ROW DATA VALUES ──
    all_properties = ResaleResidentialProperty.objects.all()
    
    for row_idx, property_obj in enumerate(all_properties, start=3):
        for col_idx, column_meta in enumerate(EXPORT_COLUMNS_BLUEPRINT, start=1):
            cell_value = getattr(property_obj, column_meta['field'], '')
            if cell_value is None:
                cell_value = ''
                
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Formatting checks
            if isinstance(cell_value, (int, float, Decimal)):
                cell.value = float(cell_value)
                if any(kw in column_meta['field'] for kw in ['price', 'amount', 'due']):
                    cell.number_format = '₹#,##,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif hasattr(cell_value, 'strftime'):
                cell.value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                cell.alignment = left_alignment
            else:
                cell.value = str(cell_value)
                cell.alignment = left_alignment

            cell.font = font_data_row
            cell.border = grid_border
            cell.fill = zebra_even_fill if row_idx % 2 == 0 else white_fill

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # ── COLUMN AUTO-FIT LAYOUT CALCULATION ──
    for col in ws.columns:
        max_string_len = 0
        column_alpha_key = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value is not None:
                max_string_len = max(max_string_len, len(str(cell.value)))
        ws.column_dimensions[column_alpha_key].width = max(max_string_len + 4, 16)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resale_residential_Export.xlsx"'
    wb.save(response)
    return response

#####################END VIEW SECTION OF RESIDENTIAL RESALE LISTING###########################






####################Start  Views Section For Commercial Resale Property #######################################





def commercial_resale_list(request):
    # ── Session check ─────────────────────────────────────
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    if not admin_id and not user_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = None
    user_obj  = None

    if admin_id:
        try:
            admin_obj = Admin_Login.objects.get(id=admin_id)
        except Admin_Login.DoesNotExist:
            return render(request, 'home_page/Adminlogin.html')
    elif user_id:
        try:
            user_obj = User_Details.objects.get(id=user_id)
        except User_Details.DoesNotExist:
            return render(request, 'home_page/Adminlogin.html')

    # # ── Base Queryset ──────────────────────────────────────
    # props = CommercialResaleProperty.objects.filter(is_deleted=False)

    # # ── Advanced Real-Time Extraction Filters ──────────────
    # search_query = request.GET.get('search_query', '').strip()
    # property_type = request.GET.get('property_type', '').strip()
    # zone_type = request.GET.get('zone_type', '').strip()
    # city = request.GET.get('city', '').strip()
    # property_condition = request.GET.get('property_condition', '').strip()
    # ownership_type = request.GET.get('ownership_type', '').strip()
    # status_filter = request.GET.get('status_filter', '').strip()
    
    # start_date_str = request.GET.get('start_date', '').strip()
    # end_date_str = request.GET.get('end_date', '').strip()

    # # 1. Global text lookup matching primary data vectors
    # if search_query:
    #     props = props.filter(
    #         Q(property_title__icontains=search_query) |
    #         Q(area_locality__icontains=search_query) |
    #         Q(building_name__icontains=search_query) |
    #         Q(owner_name__icontains=search_query)
    #     )

    # # 2. Dropdown exact match filters
    # if property_type:
    #     props = props.filter(property_type=property_type)
    # if zone_type:
    #     props = props.filter(zone_type=zone_type)
    # if city:
    #     props = props.filter(city__iexact=city)
    # if property_condition:
    #     props = props.filter(property_condition=property_condition)
    # if ownership_type:
    #     props = props.filter(ownership_type=ownership_type)
        
    # # 3. Active/Inactive Status toggle matches
    # if status_filter:
    #     if status_filter == 'active':
    #         props = props.filter(is_active=True)
    #     elif status_filter == 'inactive':
    #         props = props.filter(is_active=False)

    # # 4. Strict Date-Range queries with automated datetime conversions
    # if start_date_str:
    #     try:
    #         start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    #         props = props.filter(created_at__gte=start_date)
    #     except ValueError:
    #         pass
            
    # if end_date_str:
    #     try:
    #         # Append 23:59:59 to capture the entire final calendar day
    #         end_date = datetime.strptime(end_date_str + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
    #         props = props.filter(created_at__lte=end_date)
    #     except ValueError:
    #         pass

    # # ── ORDERING FIX ───────────────────────────────────────
    # # Order results matching historical sequence trends
    # # Changed from '-id' to '-created_at' due to the new UUID string format
    # props = props.order_by('-created_at')

    # # ── Dynamic Metric Aggregations (Reflects Filtered Querysets) ──
    # # SECTION 1: Portfolio Quantities
    # total_properties = props.count()
    # active_properties = props.filter(is_active=True).count()
    # inactive_properties = props.filter(is_active=False).count()

    # # SECTION 2: Financial Aggregations & Capital Under Management
    # avg_price = props.aggregate(Avg('expected_price'))['expected_price__avg'] or 0
    # avg_price_per_sqft = props.aggregate(Avg('price_per_sqft'))['price_per_sqft__avg'] or 0
    
    # raw_portfolio_sum = props.aggregate(Sum('expected_price'))['expected_price__sum'] or 0
    
    # # Elegant short notation scale conversion formatting for large asset valuations (Crores / Lakhs)
    # if raw_portfolio_sum >= 10000000:
    #     total_portfolio_value = f"{round(raw_portfolio_sum / 10000000, 2)} Cr"
    # elif raw_portfolio_sum >= 100000:
    #     total_portfolio_value = f"{round(raw_portfolio_sum / 100000, 2)} L"
    # else:
    #     total_portfolio_value = f"₹{raw_portfolio_sum:,}"

    # # Brokerage Performance Metrics tracking
    # brokered_deals_count = props.filter(brokerage__iexact='yes').count()
    # brokerage_with_fees_count = props.filter(brokerage__iexact='yes').exclude(brokerage_percentage='').count()

    # # SECTION 3: Property Mix Distribution
    # office_count     = props.filter(property_type='office').count()
    # shop_count       = props.filter(property_type='shop').count()
    # warehouse_count  = props.filter(property_type='warehouse').count()
    # industrial_count = props.filter(property_type='industrial').count()
    # land_count       = props.filter(property_type='land').count()

    # # Extract dynamic list arrays for autocomplete filter options lookups
    # distinct_cities = CommercialResaleProperty.objects.filter(is_deleted=False).values_list('city', flat=True).distinct()
    
    # # Uploaded Excel Files for Bulk Delete Dropdown
    # # ── FIX APPLIED: Changed uploaded_file_name to upload_file_name ──
    # uploaded_files = (
    #     CommercialResaleProperty.objects
    #     .filter(is_deleted=False)
    #     .exclude(upload_file_name__isnull=True)
    #     .exclude(upload_file_name='')
    #     .values_list('upload_file_name', flat=True)
    #     .distinct()
    #     .order_by('upload_file_name')
    # )

    # # ── Chart Data 1: Property Type Pie ────────────────────
    # type_map = {
    #     'office': 'Office Space',
    #     'shop': 'Shop/Showroom',
    #     'warehouse': 'Warehouse',
    #     'industrial': 'Industrial',
    #     'land': 'Commercial Land',
    # }
    # type_qs = props.values('property_type').annotate(count=Count('id'))
    # type_labels = [type_map.get(x['property_type'], x['property_type'].upper()) for x in type_qs]
    # type_data = [x['count'] for x in type_qs]

    # # ── Chart Data 2: Monthly Timeline (Current Year) ──────
    # current_year = timezone.now().year
    # monthly_data = [0] * 12
    # monthly_qs = props.filter(created_at__year=current_year).values('created_at__month').annotate(count=Count('id'))
    # for x in monthly_qs:
    #     monthly_data[x['created_at__month'] - 1] = x['count']

    # # ── Chart Data 3: Zone Distribution ────────────────────
    # zone_map = {
    #     'industrial': 'Industrial',
    #     'commercial': 'Commercial',
    #     'residential': 'Residential',
    #     'sez': 'SEZ',
    # }
    # zone_qs = props.values('zone_type').annotate(count=Count('id'))
    # zone_labels = [zone_map.get(x['zone_type'], x['zone_type'].upper()) for x in zone_qs]
    # zone_data = [x['count'] for x in zone_qs]

    context = {
        'admin_obj': admin_obj,
        'user_obj' : user_obj,
        
    }

    return render(request, 'admin_user/Reports/Resale/commercial_list.html', context)





def _get_uploader(request):
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    admin_obj = None
    user_obj  = None

    uploader_name    = ""
    uploader_email   = ""
    uploader_phone   = ""
    uploader_role    = ""

    if admin_id:
        try:
            admin_obj        = Admin_Login.objects.get(id=admin_id)
            uploader_name    = admin_obj.name
            uploader_email   = admin_obj.email
            uploader_phone   = admin_obj.phone
            uploader_role    = admin_obj.role
        except Admin_Login.DoesNotExist:
            return None

    elif user_id:
        try:
            user_obj         = User_Details.objects.get(id=user_id)
            uploader_name    = user_obj.user_name
            uploader_email   = user_obj.user_email
            uploader_phone   = user_obj.user_phone
            uploader_role    = user_obj.user_role
        except User_Details.DoesNotExist:
            return None
    else:
        return None  # not logged in at all

    return {
        "admin_obj"      : admin_obj,
        "user_obj"       : user_obj,
        "uploader_name"  : uploader_name,
        "uploader_email" : uploader_email,
        "uploader_phone" : uploader_phone,
        "uploader_role"  : uploader_role,
    }


import os




def add_commercial_property(request):
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    if request.method == "POST":
        try:
            prop = CommercialResaleProperty(
                # ── Step 1: Basic Info & Specifications ────────────────
                property_type      = request.POST.get('property_type'),
                zone_type          = request.POST.get('zone_type'),
                location_hub       = request.POST.get('location_hub') or None,
                property_condition = request.POST.get('property_condition'),
                ownership_type     = request.POST.get('ownership_type'),
                age_of_property    = request.POST.get('age_of_property'),

                # Commercial Specifications
                num_staircases  = request.POST.get('num_staircases') or 0,
                passenger_lifts = request.POST.get('passenger_lifts') or 0,
                service_lifts   = request.POST.get('service_lifts') or 0,
                num_cabins      = request.POST.get('num_cabins') or 0,
                meeting_rooms   = request.POST.get('meeting_rooms') or 0,
                min_seats       = request.POST.get('min_seats') or None,
                max_seats       = request.POST.get('max_seats') or None,
                private_parking = request.POST.get('private_parking') or 0,
                public_parking  = request.POST.get('public_parking') or 0,

                # Area Measurements
                builtup_area = request.POST.get('builtup_area'),
                carpet_area  = request.POST.get('carpet_area') or None,
                plot_area    = request.POST.get('plot_area') or None,

                # ── Step 2: Legal & Pricing ─────────────────────────
                num_owners              = request.POST.get('num_owners'),
                loan_on_property        = request.POST.get('loan_on_property', 'no'),
                loan_amount             = request.POST.get('loan_amount') or None,
                existing_tenants        = request.POST.get('existing_tenants', 'no'),
                tenant_details          = request.POST.get('tenant_details') or None,
                any_legal_dispute       = request.POST.get('any_legal_dispute', 'no'),
                dispute_details         = request.POST.get('dispute_details') or None,
                government_tax_dues     = request.POST.get('government_tax_dues', 'no'),
                pending_tax_amount      = request.POST.get('pending_tax_amount') or None,
                fire_safety_noc_available = request.POST.get('fire_safety_noc_available') or None,

                # Pricing
                # NOTE: price_per_sqft is intentionally NOT taken from the form —
                # it is auto-calculated inside CommercialResaleProperty.save()
                # from expected_price / builtup_area. The frontend display value
                # (formatted string like "₹1,234 / sqft") must never be stored directly.
                brokerage             = request.POST.get('brokerage') or None,
                brokerage_percentage  = request.POST.get('brokerage_percentage') or None,
                manual_brokerage      = request.POST.get('manual_brokerage') or None,
                expected_price        = request.POST.get('expected_price'),
                property_description  = request.POST.get('property_description'),
                property_summary  = request.POST.get('property_summary'),
                user_description  = request.POST.get('user_description'),
                sanctioning_authority = request.POST.get('sanctioning_authority'),

                # ── Step 3: Amenities & Location ───────────────────
                nearby_facilities = ','.join(request.POST.getlist('nearby_facilities')),
                amenities         = ','.join(request.POST.getlist('amenities')),

                city             = request.POST.get('city'),
                area_locality    = request.POST.get('area_locality'),
                building_name    = request.POST.get('building_name') or None,
                property_address = request.POST.get('property_address'),

                owner_name         = request.POST.get('owner_name'),
                owner_contact      = request.POST.get('owner_contact'),
                owner_email        = request.POST.get('owner_email'),
                owner_role         = request.POST.get('owner_role') or None,  # ✅ ADDED — matches form name="owner_role"
                residential_status = request.POST.get('residential_status'),

                # ── Step 4: Media ───────────────────────────────────
                floor_plan     = request.FILES.get('floor_plan') or None,
                property_video = request.FILES.get('property_video') or None,

                # ── Uploader ────────────────────────────────────────
                uploaded_by_name    = uploader.get('uploader_name'),
                uploaded_by_email   = uploader.get('uploader_email'),
                uploaded_by_contact = uploader.get('uploader_phone'),
                uploaded_by_role    = uploader.get('uploader_role'),
            )

            prop.save()

            # Save up to 10 property images
            images = request.FILES.getlist('property_images')
            for image in images[:10]:
                CommercialPropertyImage.objects.create(property=prop, image=image)

            return JsonResponse({
                "status": "success",
                "message": "Commercial Property Added Successfully",
                "generated_title": prop.property_title
            })

        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"Database transaction failed: {str(e)}"
            }, status=400)

    context = {
        "admin_obj"     : uploader.get('admin_obj'),
        "user_obj"      : uploader.get('user_obj'),
        "uploader_name" : uploader.get('uploader_name'),
        "uploader_email": uploader.get('uploader_email'),
        "uploader_phone": uploader.get('uploader_phone'),
        "uploader_role" : uploader.get('uploader_role'),
        "facilities_obj": Facilities_Details.objects.all() if 'Facilities_Details' in globals() else [],
        "ameneties_obj" : Ameneties_Details.objects.all() if 'Ameneties_Details' in globals() else [],
    }
    return render(request, 'admin_user/Reports/Resale/commercial_resale.html', context)



def commercial_resale_edit(request, id):
    uploader = _get_uploader(request)
    if uploader is None:
        return redirect('login')

    prop = get_object_or_404(CommercialResaleProperty, id=id)

    if request.method == "POST":
        def safe_float(val):
            try: return float(val) if val else 0.0
            except ValueError: return 0.0
            
        def safe_int(val):
            try: return int(float(val)) if val else 0
            except ValueError: return 0

        # ── Basic Information ──────────────────────────────
        prop.property_title = request.POST.get('property_title') or None  # Updates the title sequence fallback
        prop.property_type = request.POST.get('property_type')
        prop.zone_type = request.POST.get('zone_type')
        prop.location_hub = request.POST.get('location_hub') or None
        prop.property_condition = request.POST.get('property_condition')
        prop.ownership_type = request.POST.get('ownership_type')
        prop.age_of_property = request.POST.get('age_of_property')
        prop.available_from = request.POST.get('available_from') or None

        # ── Commercial Specifications ──────────────────────
        prop.num_staircases = safe_int(request.POST.get('num_staircases'))
        prop.passenger_lifts = safe_int(request.POST.get('passenger_lifts'))
        prop.service_lifts = safe_int(request.POST.get('service_lifts'))
        prop.num_cabins = safe_int(request.POST.get('num_cabins'))
        prop.meeting_rooms = safe_int(request.POST.get('meeting_rooms'))
        prop.min_seats = safe_int(request.POST.get('min_seats')) if request.POST.get('min_seats') else None
        prop.max_seats = safe_int(request.POST.get('max_seats')) if request.POST.get('max_seats') else None
        prop.private_parking = safe_int(request.POST.get('private_parking'))
        prop.public_parking = safe_int(request.POST.get('public_parking'))

        # ── Area & Pricing ─────────────────────────────────
        prop.builtup_area = safe_float(request.POST.get('builtup_area'))
        prop.carpet_area = safe_float(request.POST.get('carpet_area')) if request.POST.get('carpet_area') else None
        prop.plot_area = safe_float(request.POST.get('plot_area')) if request.POST.get('plot_area') else None
        
        prop.brokerage = request.POST.get('brokerage') or None
        prop.brokerage_percentage = request.POST.get('brokerage_percentage') or None
        prop.manual_brokerage = request.POST.get('manual_brokerage') or None
        prop.expected_price = safe_float(request.POST.get('expected_price'))

        # ── Ownership & Legal ──────────────────────────────
        prop.num_owners = request.POST.get('num_owners', '1')
        prop.loan_on_property = request.POST.get('loan_on_property', 'no')
        prop.loan_amount = safe_float(request.POST.get('loan_amount')) if prop.loan_on_property == 'yes' else None
        
        prop.existing_tenants = request.POST.get('existing_tenants', 'no')
        prop.tenant_details = request.POST.get('tenant_details') if prop.existing_tenants == 'yes' else None
        
        prop.any_legal_dispute = request.POST.get('any_legal_dispute', 'no') # Updated key
        prop.dispute_details = request.POST.get('dispute_details') if prop.any_legal_dispute == 'yes' else None
        
        prop.government_tax_dues = request.POST.get('government_tax_dues', 'no') # Updated key
        prop.pending_tax_amount = safe_float(request.POST.get('pending_tax_amount')) if prop.government_tax_dues == 'yes' else None
        
        prop.fire_safety_noc_available = request.POST.get('fire_safety_noc_available', 'no') # Updated key
        prop.property_description = request.POST.get('property_description')
        prop.property_summary = request.POST.get('property_summary')
        prop.user_description = request.POST.get('user_description')
        prop.sanctioning_authority = request.POST.get('sanctioning_authority')

        # ── Nearby Facilities & Amenities ──────────────────
        prop.nearby_facilities = ', '.join(request.POST.getlist('nearby_facilities'))
        prop.amenities = ', '.join(request.POST.getlist('amenities'))

        # ── Address ────────────────────────────────────────
        prop.city = request.POST.get('city')
        prop.area_locality = request.POST.get('area_locality') # Updated key
        prop.building_name = request.POST.get('building_name') or None
        prop.property_address = request.POST.get('property_address')

        # ── Owner Contact ──────────────────────────────────
        prop.owner_name = request.POST.get('owner_name')
        prop.owner_contact = request.POST.get('owner_contact')
        prop.owner_email = request.POST.get('owner_email')
        prop.residential_status = request.POST.get('residential_status')

        # ── Media ──────────────────────────────────────────
        if request.FILES.get('floor_plan'):
            prop.floor_plan = request.FILES.get('floor_plan')
        if request.FILES.get('property_video'):
            prop.property_video = request.FILES.get('property_video')

        prop.save()

        # Handle Image Gallery (Delete marked records)
        deleted_images = request.POST.getlist('deleted_images[]')
        if deleted_images:
            CommercialPropertyImage.objects.filter(id__in=deleted_images, property=prop).delete()

        # Append additional attachments if space remains under current session bounds
        current_images_count = prop.images.count()
        new_files = request.FILES.getlist('property_images')
        available_slots = 10 - current_images_count
        
        for img in new_files[:available_slots]:
            CommercialPropertyImage.objects.create(property=prop, image=img)

        return JsonResponse({"status": "success", "message": "Commercial Property Updated Successfully"})

    # --- GET REQUEST Context ---
    ameneties_obj = Ameneties_Details.objects.all() if 'Ameneties_Details' in globals() else []
    facilities_obj = Facilities_Details.objects.all() if 'Facilities_Details' in globals() else []
    
    prop_facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []
    prop_amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []
    
    existing_images = prop.images.all() 

    context = {
        "prop": prop,
        "ameneties_obj": ameneties_obj,
        "facilities_obj": facilities_obj,
        "prop_facilities_list": prop_facilities_list,
        "prop_amenities_list": prop_amenities_list,
        "existing_images": existing_images,
    }
    return render(request, 'admin_user/Reports/Resale/commercial_resale_edit.html', context)









def import_test_view(request):
    result = ''
    if request.method == 'POST':
        f = request.FILES.get('commercial_excel_file')
        if f:
            result = f'File received: {f.name}, size: {f.size} bytes'
        else:
            result = 'ERROR: No file in request.FILES — field name mismatch or form issue'
    return HttpResponse(f'''
        <h2>Import Debug Test</h2>
        <p style="color:green;font-size:18px;">{result}</p>
        <hr>
        <h3>Plain Form Test</h3>
        <form method="POST" enctype="multipart/form-data">
            <input type="hidden" name="csrfmiddlewaretoken" value="get-from-cookie">
            <input type="file" name="commercial_excel_file" accept=".xlsx">
            <button type="submit">Submit</button>
        </form>
    ''')




from openpyxl.worksheet.datavalidation import DataValidation






def get_filtered_properties(request):
    """Internal helper to capture the identical data arrays currently viewed inside the frontend lists."""
    props = CommercialResaleProperty.objects.filter(is_deleted=False)
    
    search_query = request.GET.get('search_query', '').strip()
    property_type = request.GET.get('property_type', '').strip()
    zone_type = request.GET.get('zone_type', '').strip()
    city = request.GET.get('city', '').strip()
    property_condition = request.GET.get('property_condition', '').strip()
    ownership_type = request.GET.get('ownership_type', '').strip()
    status_filter = request.GET.get('status_filter', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    if search_query:
        props = props.filter(
            Q(property_title__icontains=search_query) |
            Q(area_locality__icontains=search_query) |
            Q(building_name__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )
    if property_type:
        props = props.filter(property_type=property_type)
    if zone_type:
        props = props.filter(zone_type=zone_type)
    if city:
        props = props.filter(city__iexact=city)
    if property_condition:
        props = props.filter(property_condition=property_condition)
    if ownership_type:
        props = props.filter(ownership_type=ownership_type)
    if status_filter:
        if status_filter == 'active':
            props = props.filter(is_active=True)
        elif status_filter == 'inactive':
            props = props.filter(is_active=False)
            
    if start_date_str:
        try:
            props = props.filter(created_at__gte=datetime.strptime(start_date_str, '%Y-%m-%d'))
        except ValueError: pass
    if end_date_str:
        try:
            props = props.filter(created_at__lte=datetime.strptime(end_date_str + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError: pass

    return props.order_by('-id')





@transaction.atomic
def import_commercial_data(request):
    """
    Ingests sectioned double-header master sheets cleanly.
    Guarantees exact mapping indices to align database content profiles perfectly.
    """
    # ── IDENTIFY SESSION OWNER INGESTION AUDITS ─────────────────
    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    if not admin_id and not user_id:
        return JsonResponse({'status': '0', 'msg': 'Session Expired. Please log in again to perform uploads.'})

    uploader_name, uploader_email, uploader_contact, uploader_role = "System Bulk Import", "", "", "Staff"

    if admin_id:
        try:
            admin_obj = Admin_Login.objects.get(id=admin_id)
            uploader_name = getattr(admin_obj, 'name', 'Admin')
            uploader_email = getattr(admin_obj, 'email', '')
            uploader_contact = getattr(admin_obj, 'phone', '')
            uploader_role = getattr(admin_obj, 'role', 'Admin')
        except Admin_Login.DoesNotExist: pass
    elif user_id:
        try:
            user_obj = User_Details.objects.get(id=user_id)
            uploader_name = getattr(user_obj, 'name', 'User')
            uploader_email = getattr(user_obj, 'email', '')
            uploader_contact = getattr(user_obj, 'phone', '')
            uploader_role = "User"
        except User_Details.DoesNotExist: pass

    if request.method != 'POST' or not request.FILES.get('commercial_excel_file'):
        return JsonResponse({'status': '0', 'msg': 'Missing transactional multipart form document.'})
        
    excel_file = request.FILES['commercial_excel_file']
    filename = excel_file.name

    if not filename.endswith('.xlsx'):
        return JsonResponse({'status': '0', 'msg': 'Invalid file layout format. Please drop a valid .xlsx file.'})

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # min_row=3 explicitly bypasses Row 1 (Visual Sections) and Row 2 (Column Headers)
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        
        if not rows or all(all(cell is None for cell in row) for row in rows):
            return JsonResponse({'status': '0', 'msg': 'The processed workbook contains no usable data records.'})

        filename_exists = CommercialResaleProperty.objects.filter(upload_file_name=filename).exists()
        success_count = 0
        duplicate_count = 0

        def clean_str(val):
            return str(val).strip() if val is not None else ""

        def clean_num(val, is_float=False):
            if val is None or str(val).strip() == '':
                return 0.0 if is_float else 0
            try:
                cleaned_val = str(val).replace('₹', '').replace(',', '').strip()
                return float(cleaned_val) if is_float else int(float(cleaned_val))
            except (ValueError, TypeError):
                return 0.0 if is_float else 0

        # ── ITERATE CELLS POPULATED LAYERS ─────────────────────────
        for index, row in enumerate(rows, start=3):
            p_type_raw = clean_str(row[2]).lower()  # Property Type is locked at column index 2 (C)
            if not p_type_raw:  # Gracefully skip empty bottom trailing lines
                continue

            # --- CORRECTION SYNC FIELD MANDATORY VALIDATIONS MATRIX ---
            required_fields = {
                "Property Type": row[2],
                "Zone Type": row[3],
                "Property Condition": row[5],
                "Ownership Type": row[6],
                "Age of Property": row[7],
                "Private Parking": row[15],
                "Built-up Area (sq.ft)": row[17],
                "No. of Owners": row[20],
                "Loan on Property?": row[21],
                "Existing Tenants?": row[23],
                "Any Legal Dispute?": row[25],
                "Government Tax Dues?": row[27],
                "Expected Price (₹)": row[33],
                "Property Description": row[35],
                "Sanctioning Authority": row[36],
                "City": row[39],
                "Area/Locality": row[40],
                "Property Address": row[42],
                "Owner Name": row[43],
                "Owner Contact": row[44],
                "Owner Email": row[45],
                "Residential Status": row[47],
            }

            missing_fields = [field for field, val in required_fields.items() if val is None or str(val).strip() == '']
            if missing_fields:
                return JsonResponse({
                    'status': '0', 
                    'msg': f'Row {index} Stop: The following mandatory attributes are missing/blank: [{", ".join(missing_fields)}]'
                })

            city = clean_str(row[39])
            locality = clean_str(row[40])
            expected_price = clean_num(row[33], is_float=True)
            builtup_area_val = clean_num(row[17], is_float=True)

            # --- DYNAMIC AND UNIFORM AUTO PROPERTY TITLE GENERATOR ---
            type_clean_map = {
                'office': 'Office Space', 'shop': 'Shop / Showroom', 'warehouse': 'Warehouse',
                'industrial': 'Industrial Plant', 'land': 'Commercial Plot Land'
            }
            display_type = type_clean_map.get(p_type_raw, p_type_raw.title())
            generated_title = f"Premium {int(builtup_area_val)} Sqft {display_type}"
            building_val = clean_str(row[41])
            if building_val and building_val != "—" and building_val != "":
                generated_title += f" in {building_val}"
            generated_title += f" at {locality}, {city}"

            # --- STRUCTURAL ENFORCED DATA DUPLICATE security checks ---
            is_data_duplicate = CommercialResaleProperty.objects.filter(
                property_title=generated_title,
                property_type=p_type_raw,
                city=city,
                expected_price=expected_price,
                is_deleted=False
            ).exists()

            if is_data_duplicate:
                duplicate_count += 1
                if filename_exists:
                    return JsonResponse({
                        'status': '0',
                        'msg': f'Duplicate Halted: The upload package "{filename}" has already been processed and contains matching rows at index line {index}.'
                    })
                continue

            price_per_sqft_calc = round(expected_price / builtup_area_val, 2) if builtup_area_val > 0 else 0.0

            # --- EXECUTE STAGE RECORD COMMITMENT POOLS ---
            CommercialResaleProperty.objects.create(
                property_title=generated_title,
                
                # 📋 STEP 1: BASIC INFO & SPECIFICATIONS
                property_type=p_type_raw,
                zone_type=clean_str(row[3]).lower(),
                location_hub=clean_str(row[4]).lower() if row[4] else None,
                property_condition=clean_str(row[5]).lower(),
                ownership_type=clean_str(row[6]).lower(),
                age_of_property=clean_str(row[7]),
                
                num_staircases=clean_num(row[8]),
                passenger_lifts=clean_num(row[9]),
                service_lifts=clean_num(row[10]),
                num_cabins=clean_num(row[11]),
                meeting_rooms=clean_num(row[12]),
                min_seats=clean_num(row[13]) if row[13] is not None else None,
                max_seats=clean_num(row[14]) if row[14] is not None else None,
                private_parking=clean_num(row[15]),
                public_parking=clean_num(row[16]) if row[16] is not None else 0,
                
                builtup_area=builtup_area_val,
                carpet_area=clean_num(row[18], is_float=True) if row[18] else None,
                plot_area=clean_num(row[19], is_float=True) if row[19] else None,
                
                # 📋 STEP 2: LEGAL & PRICING DETAILS
                num_owners=clean_str(row[20]),
                loan_on_property=clean_str(row[21]).lower(),
                loan_amount=clean_num(row[22], is_float=True) if row[22] else None,
                existing_tenants=clean_str(row[23]).lower(),
                tenant_details=clean_str(row[24]) if row[24] else None,
                any_legal_dispute=clean_str(row[25]).lower(),
                dispute_details=clean_str(row[26]) if row[26] else None,
                government_tax_dues=clean_str(row[27]).lower(),
                pending_tax_amount=clean_num(row[28], is_float=True) if row[28] else None,
                fire_safety_noc_available=clean_str(row[29]).lower(),
                
                brokerage=clean_str(row[30]).lower(),
                brokerage_percentage=clean_str(row[31]) if row[31] else None,
                manual_brokerage=clean_str(row[32]) if row[32] else None,
                expected_price=expected_price,
                price_per_sqft=price_per_sqft_calc,
                property_description=clean_str(row[35]),
                sanctioning_authority=clean_str(row[36]),
                
                # 📋 STEP 3: AMENITIES & LOCATION
                nearby_facilities=clean_str(row[37]) if row[37] else None,
                amenities=clean_str(row[38]) if row[38] else None,
                city=city,
                area_locality=locality,
                building_name=building_val if building_val else None,
                property_address=clean_str(row[42]),
                
                # CONTACT SEQUENCE DETAILED TRACKING MAPS
                owner_name=clean_str(row[43]),
                owner_contact=clean_str(row[44]),
                owner_email=clean_str(row[45]),
                residential_status=clean_str(row[47]).lower(),
                
                # 📋 STEP 4: META TRACKING SYSTEMS AUDIT CONTROL LOGS
                uploaded_by_name=uploader_name,
                uploaded_by_email=uploader_email,
                uploaded_by_contact=uploader_contact,
                uploaded_by_role=uploader_role,
                upload_file_name=filename,
                is_active=True
            )
            success_count += 1

        if success_count > 0:
            msg = f'Import Process Completed! Successfully persisted {success_count} real-estate assets records.'
            if duplicate_count > 0:
                msg += f' ({duplicate_count} rows filtered out as safe database duplicates).'
            return JsonResponse({'status': '1', 'msg': msg})
        else:
            return JsonResponse({'status': '0', 'msg': 'All spreadsheet line profiles match current records. Zero mutations committed.'})

    except Exception as e:
        return JsonResponse({'status': '0', 'msg': f'Critical runtime exception encountered inside workbook sheets: {str(e)}'})




def download_commercial_sample_excel(request):
    """
    Generates a beautifully colored template matching the structural layout setup.
    Column index 0 is reserved for read-only title generation preview states.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commercial Resale"
    ws.views.sheetView[0].showGridLines = True

    primary_purple = "667EEA"
    dark_section_slate = "475569"
    text_white = "FFFFFF"

    section_font = Font(name="Poppins", size=12, bold=True, color=text_white)
    header_font = Font(name="Poppins", size=10, bold=True, color=text_white)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 54

    # --- ROW 1: RENDER COMPACT SECTION MASTER WRAPPERS BLOCK ---
    sections = [
        ("📋 Basic Info & Specs", 1, 15, dark_section_slate),
        ("📋 Area, Pricing & Building Details", 16, 37, primary_purple),
        ("📋 Amenities & Facilities", 38, 42, dark_section_slate),
        ("📋 Contact Info & System Logs", 43, 55, primary_purple)
    ]

    for title, start_col, end_col, color in sections:
        cell = ws.cell(row=1, column=start_col, value=title)
        cell.font = section_font
        cell.alignment = cell_alignment
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for col in range(start_col + 1, end_col + 1):
            ws.cell(row=1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # --- ROW 2: ENFORCE COMPLETE 55-COLUMN LABEL ARRANGEMENT ---
    headers = [
        "Property Id (Leave Blank - System Generated)", "Property Title (Leave Blank - System Generated)", 
        "Property Type (office/shop/warehouse/industrial/land) *", "Zone Type (industrial/commercial/residential/sez) *",
        "Location Hub", "Property Condition (new/excellent/good/renovation) *", "Ownership Type (freehold/leasehold/cooperative) *", 
        "Age Of The Property (0-1/1-3/3-5/5-10/10+) *", "No.Of.Staircases", "Passenger Lifts", "Service Lifts", 
        "No.Of.Cabins", "Meeting Rooms", "Min Seats", "Max Seats", "Private Parking *", "Public Parking", "Builtup Area *", 
        "Carpet Area", "Plot Area", "No.of Owners *",   
        "Loan on Property (yes/no) *", "Loan Amount", "Existing Tenants (yes/no) *", "Tenant Details", "Any Legal Dispute (yes/no) *", 
        "Dispute Details", "Government Tax Dues? (yes/no) *", "Pending Tax Amount", "Fire Safety NOC Available? (yes/no) *","Brokerage (yes/no)", "Brokerage Percentage","Manual_Brokarage","Expected Price *","Price Per Sqft/Auto Generated", "Property Description *", 
        "Sanctioning Authority *", "Nearby Facilities (comma separated)", "Amenities (comma separated)", "City *", 
        "Area/Locality *", "Building / Project / Society", "Property Address *", "Owner Name *", "Owner Contact *", "Owner Email *","Owner Role",
        "Residential Status (resident/nri/pio) *", "Uploaded By Name", "Uploaded By Email", "Uploaded By Contact", "Uploaded By Role", "Source File Name", "Created At", "Updated At"
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index, value=header)
        cell.font = header_font
        cell.alignment = cell_alignment
        bg_color = "5A67D8" if col_index in range(16, 38) or col_index in range(43, 56) else "334155"
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 24

    # --- ROW 3: RE-ALIGNED SAMPLE PROTO PLATFORM INVENTORIES (55 ELEMENTS) ---
    sample_data = [
        "", "", "office", "commercial", "it", "excellent", "freehold", "1-3",
        2, 4, 1, 5, 2, 50, 100, 5, 10, 5000, 4500, 0, "1",
        "no", 0, "no", "", "no", "", "no", 0, "yes",
        "yes", "2%", "", 15000000, "", 
        "Fully furnished premium office space with modern workstation infrastructure.",
        "NMC", "Metro Station, Bus Stop, Mall, Hospital", 
        "CCTV Security, 100% Power Backup, Cafeteria, Central Gym", 
        "Nagpur", "Sitabuldi", "Tech Tower", "Floor 4, Tech Tower, Main Wardha Road, Sitabuldi",
        "Rajesh Kumar", "9876543210", "rajesh@example.com", "Owner", "resident",
        "System Template", "—", "—", "—", "—", "—", "—"
    ]
    ws.append(sample_data)

    def create_dropdown(options, cell_range):
        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cell_range)

    # Re-indexedDropdown targets seamlessly link with explicit column indices
    create_dropdown(["office", "shop", "warehouse", "industrial", "land"], "C3:C100")
    create_dropdown(["industrial", "commercial", "residential", "sez"], "D3:D100")
    create_dropdown(["it", "business", "mall", "standalone"], "E3:E100")
    create_dropdown(["new", "excellent", "good", "renovation"], "F3:F100")
    create_dropdown(["freehold", "leasehold", "cooperative"], "G3:G100")
    create_dropdown(["0-1", "1-3", "3-5", "5-10", "10+"], "H3:H100")
    create_dropdown(["1", "2", "3", "4+"], "U3:U100")
    create_dropdown(["resident", "nri", "pio"], "AV3:AV100")
    
    for col_letter in ["V", "X", "Z", "AB", "AD", "AE"]:
        create_dropdown(["yes", "no"], f"{col_letter}3:{col_letter}100")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Commercial_Resale_Sample_Master.xlsx"'
    wb.save(response)
    return response



def export_commercial_resale_excel(request):
    """
    Generates the master live database inventory extract.
    Maintains exact 55-column alignments to protect cell arrays from index shifts.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live Database Extract"
    ws.views.sheetView[0].showGridLines = True

    primary_purple = "667EEA"
    dark_section_slate = "475569"
    text_white = "FFFFFF"

    section_font = Font(name="Poppins", size=12, bold=True, color=text_white)
    header_font = Font(name="Poppins", size=10, bold=True, color=text_white)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 54

    sections = [
        ("📋 Basic Info & Specs", 1, 15, dark_section_slate),
        ("📋 Area, Pricing & Building Details", 16, 37, primary_purple),
        ("📋 Amenities & Facilities", 38, 42, dark_section_slate),
        ("📋 Contact Info & System Logs", 43, 55, primary_purple)
    ]

    for title, start_col, end_col, color in sections:
        cell = ws.cell(row=1, column=start_col, value=title)
        cell.font = section_font
        cell.alignment = cell_alignment
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for col in range(start_col + 1, end_col + 1):
            ws.cell(row=1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    headers = [
        "Property Id","Property Title", "Property Type (office/shop/warehouse/industrial/land) *", "Zone Type (industrial/commercial/residential/sez) *",
        "Location Hub", "Property Condition (new/excellent/good/renovation) *", "Ownership Type (freehold/leasehold/cooperative) *", 
        "Age Of The Property (0-1/1-3/3-5/5-10/10+) *", "No.Of.Staircases", "Passenger Lifts", "Service Lifts", 
        "No.Of.Cabins", "Meeting Rooms", "Min Seats", "Max Seats", "Private Parking *", "Public Parking", "Builtup Area *", 
        "Carpet Area", "Plot Area", "No.of Owners *",   
        "Loan on Property (yes/no) *", "Loan Amount", "Existing Tenants (yes/no) *", "Tenant Details", "Any Legal Dispute (yes/no) *", 
        "Dispute Details", "Government Tax Dues? (yes/no) *", "Pending Tax Amount", "Fire Safety NOC Available? (yes/no) *","Brokerage (yes/no)", "Brokerage Percentage","Manual_Brokarage","Expected Price *","Price Per Sqft/Auto Generated", "Property Description *", 
        "Sanctioning Authority *", "Nearby Facilities (comma separated)", "Amenities (comma separated)", "City *", 
        "Area/Locality *", "Building / Project / Society", "Property Address *", "Owner Name *", "Owner Contact *", "Owner Email *","Owner Role",
        "Residential Status (resident/nri/pio) *", "Uploaded By Name", "Uploaded By Email", "Uploaded By Contact", "Uploaded By Role", "Source File Name", "Created At", "Updated At"
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index, value=header)
        cell.font = header_font
        cell.alignment = cell_alignment
        bg_color = "5A67D8" if col_index in range(16, 38) or col_index in range(43, 56) else "334155"
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 24

    properties = get_filtered_properties(request)

    for prop in properties:
        data_row = [
           
            prop.id or "—",
            prop.property_title or "—",
            prop.property_type,
            prop.zone_type,
            prop.location_hub or "—",
            prop.property_condition,
            prop.ownership_type,
            prop.age_of_property,
            prop.num_staircases or 0,
            prop.passenger_lifts or 0,
            prop.service_lifts or 0,
            prop.num_cabins or 0,
            prop.meeting_rooms or 0,
            prop.min_seats or "—",
            prop.max_seats or "—",
            prop.private_parking or 0,
            prop.public_parking or 0,
            float(prop.builtup_area),
            float(prop.carpet_area) if prop.carpet_area else "—",
            float(prop.plot_area) if prop.plot_area else "—",
            prop.num_owners or "1",
            prop.loan_on_property,
            float(prop.loan_amount) if prop.loan_amount else 0,
            prop.existing_tenants,
            prop.tenant_details or "",
            prop.any_legal_dispute,
            prop.dispute_details or "",
            prop.government_tax_dues,
            float(prop.pending_tax_amount) if prop.pending_tax_amount else 0,
            prop.fire_safety_noc_available or "no",
            prop.brokerage or "no",
            prop.brokerage_percentage or "",
            prop.manual_brokerage or "",
            float(prop.expected_price),
            float(prop.price_per_sqft) if prop.price_per_sqft else 0,
            prop.property_description,
            prop.sanctioning_authority,
            prop.nearby_facilities or "",
            prop.amenities or "",
            prop.city,
            prop.area_locality,
            prop.building_name or "",
            prop.property_address,
            prop.owner_name,
            prop.owner_contact,
            prop.owner_email,
            prop.uploaded_by_role or "Owner",
            prop.residential_status,
            prop.uploaded_by_name or "Portal Direct",
            prop.uploaded_by_email or "—",
            prop.uploaded_by_contact or "—",
            prop.uploaded_by_role or "—",
            prop.upload_file_name or "Web Input Form",
            prop.created_at.strftime("%Y-%m-%d %H:%M") if prop.created_at else "—",
            prop.updated_at.strftime("%Y-%m-%d %H:%M") if prop.updated_at else "—"
        ]
        ws.append(data_row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Commercial_Resale_Export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response






def export_commercial_resale_csv(request):
    """
    Flat CSV layout exporter matching the same sequential blueprint matrix.
    Includes Property Title as index 0 to eliminate data parsing drift.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Commercial_Resale_Export_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # ── EXACT 54-COLUMN SEQUENCE MAPPING FROM INDEX 0 TO 53 ──
    writer.writerow([
        "Property Id","Property Title","Property Type (office/shop/warehouse/industrial/land) *", 
        "Zone Type (industrial/commercial/residential/sez) *", "Location Hub", 
        "Property Condition (new/excellent/good/renovation) *", "Ownership Type (freehold/leasehold/cooperative) *", 
        "Age Of The Property (0-1/1-3/3-5/5-10/10+) *", "No.Of.Staircases", "Passenger Lifts", 
        "Service Lifts", "No.Of.Cabins", "Meeting Rooms", "Min Seats", "Max Seats", 
        "Private Parking *", "Public Parking", "Builtup Area *", "Carpet Area", "Plot Area", "No.of Owners *",   
        "Loan on Property (yes/no) *", "Loan Amount", "Existing Tenants (yes/no) *", "Tenant Details", 
        "Any Legal Dispute (yes/no) *", "Dispute Details", "Government Tax Dues? (yes/no) *", "Pending Tax Amount", 
        "Fire Safety NOC Available? (yes/no) *", "Brokerage (yes/no)", "Brokerage Percentage", "Manual_Brokarage", 
        "Expected Price *", "Price Per Sqft/Auto Generated", "Property Description *", "Sanctioning Authority *", 
        "Nearby Facilities (comma separated)", "Amenities (comma separated)", "City *", "Area/Locality *", 
        "Building / Project / Society", "Property Address *", "Owner Name *", "Owner Contact *", "Owner Email *", 
        "Owner Role", "Residential Status (resident/nri/pio) *", "Uploaded By Name", "Uploaded By Email", 
        "Uploaded By Contact", "Uploaded By Role", "Source File Name", "Created At", "Updated At"
    ])
    
    # Pull identical data entries matching active dashboard panel query parameters
    properties = get_filtered_properties(request)
    
    for prop in properties:
        # Dynamic calculation check fallback if field database values are unpopulated
        calculated_price_per_sqft = prop.price_per_sqft
        if not calculated_price_per_sqft and prop.builtup_area > 0:
            calculated_price_per_sqft = round(float(prop.expected_price) / float(prop.builtup_area), 2)

        writer.writerow([

            prop.commercial_id or "—", 
            prop.property_title or "—",                         # [0] Property Title
                                  # [1] Property Id
            prop.property_type,                                  # [2] Property Type
            prop.zone_type,                                      # [3] Zone Type
            prop.location_hub or "—",                            # [4] Location Hub
            prop.property_condition,                             # [5] Property Condition
            prop.ownership_type,                                 # [6] Ownership Type
            prop.age_of_property,                                # [7] Age Of The Property
            prop.num_staircases or 0,                            # [8] No.Of.Staircases
            prop.passenger_lifts or 0,                           # [9] Passenger Lifts
            prop.service_lifts or 0,                             # [10] Service Lifts
            prop.num_cabins or 0,                                # [11] No.Of.Cabins
            prop.meeting_rooms or 0,                             # [12] Meeting Rooms
            prop.min_seats or "—",                               # [13] Min Seats
            prop.max_seats or "—",                               # [14] Max Seats
            prop.private_parking or 0,                           # [15] Private Parking
            prop.public_parking or 0,                            # [16] Public Parking
            float(prop.builtup_area),                            # [17] Builtup Area
            float(prop.carpet_area) if prop.carpet_area else "—", # [18] Carpet Area
            float(prop.plot_area) if prop.plot_area else "—",     # [19] Plot Area
            prop.num_owners or "1",                              # [20] No.of Owners
            prop.loan_on_property,                               # [21] Loan on Property (yes/no)
            float(prop.loan_amount) if prop.loan_amount else 0,  # [22] Loan Amount
            prop.existing_tenants,                               # [23] Existing Tenants (yes/no)
            prop.tenant_details or "",                           # [24] Tenant Details
            prop.any_legal_dispute,                              # [25] Any Legal Dispute (yes/no)
            prop.dispute_details or "",                          # [26] Dispute Details
            prop.government_tax_dues,                            # [27] Government Tax Dues? (yes/no)
            float(prop.pending_tax_amount) if prop.pending_tax_amount else 0, # [28] Pending Tax Amount
            prop.fire_safety_noc_available or "no",              # [29] Fire Safety NOC Available? (yes/no)
            prop.brokerage or "no",                              # [30] Brokerage (yes/no)
            prop.brokerage_percentage or "",                     # [31] Brokerage Percentage
            prop.manual_brokerage or "",                         # [32] Manual_Brokarage
            float(prop.expected_price),                          # [33] Expected Price
            float(calculated_price_per_sqft) if calculated_price_per_sqft else 0, # [34] Price Per Sqft
            prop.property_description,                           # [35] Property Description
            prop.sanctioning_authority,                          # [36] Sanctioning Authority
            prop.nearby_facilities or "",                        # [37] Nearby Facilities
            prop.amenities or "",                                # [38] Amenities
            prop.city,                                           # [39] City
            prop.area_locality,                                  # [40] Area/Locality
            prop.building_name or "",                            # [41] Building / Project / Society
            prop.property_address,                               # [42] Property Address
            prop.owner_name,                                     # [43] Owner Name
            prop.owner_contact,                                  # [44] Owner Contact
            prop.owner_email,                                    # [45] Owner Email
            prop.uploaded_by_role or "Owner",                    # [46] Owner Role
            prop.residential_status,                             # [47] Residential Status
            prop.uploaded_by_name or "Portal Direct",            # [48] Uploaded By Name
            prop.uploaded_by_email or "—",                       # [49] Uploaded By Email
            prop.uploaded_by_contact or "—",                     # [50] Uploaded By Contact
            prop.uploaded_by_role or "—",                        # [51] Uploaded By Role
            prop.upload_file_name or "Web Form Direct",          # [52] Source File Name
            prop.created_at.strftime("%Y-%m-%d %H:%M") if prop.created_at else "—", # [53] Created At
            prop.updated_at.strftime("%Y-%m-%d %H:%M") if prop.updated_at else "—"  # [54] Updated At
        ])
        
    return response




    
    # ── Helper functions ─────────────────────────────────
    def val(v, default=''):
        if v is None or str(v).strip() == '':
            return default
        return str(v).strip()

    def num(v, default=None):
        if v is None or str(v).strip() == '':
            return default
        try:
            return float(v)
        except (ValueError, TypeError):
            return default

    def pos_int(v, default=0):
        if v is None or str(v).strip() == '':
            return default
        try:
            return int(float(str(v)))
        except (ValueError, TypeError):
            return default

    def yn(v):
        if v is None:
            return 'no'
        return 'yes' if str(v).strip().lower() in ('yes', 'y', '1', 'true') else 'no'

    def parse_date(v):
        if v is None or str(v).strip() == '':
            return None
        try:
            if hasattr(v, 'date'):
                return v.date()
            return datetime.strptime(str(v).strip(), '%Y-%m-%d').date()
        except Exception:
            return None




# ── LIST VIEW ────────────────────────────────────────────────



# ── TOGGLE ───────────────────────────────────────────────────
@csrf_exempt
def toggle_commercial_property(request):
    if request.method == 'POST':
        prop_id = request.POST.get('prop_id')
        try:
            prop = CommercialResaleProperty.objects.get(id=prop_id)
            prop.is_active = not prop.is_active
            prop.save()
            status = 'Active' if prop.is_active else 'Inactive'
            return JsonResponse({'status': '1', 'msg': f'Property marked as {status}.'})
        except CommercialResaleProperty.DoesNotExist:
            return JsonResponse({'status': '0', 'msg': 'Property not found.'})
    return JsonResponse({'status': '0', 'msg': 'Invalid request.'})


# ── DELETE ───────────────────────────────────────────────────

@csrf_exempt
def delete_commercial_property(request):
    if request.method == 'POST':
        prop_id = request.POST.get('prop_id')
        try:
            deleter_name = _get_deleter_name(request)
            
            prop = CommercialResaleProperty.objects.get(id=prop_id)
            prop.is_deleted = True
            prop.deleted_at = timezone.now()
            prop.deleted_by = deleter_name # 👈 SAVE THE NAME HERE
            prop.save()
            
            return JsonResponse({'status': 'success', '1': '1', 'msg': 'Moved to Recycle Bin successfully.'})
        except CommercialResaleProperty.DoesNotExist:
            return JsonResponse({'status': 'error', '0': '0', 'msg': 'Property not found.'})
    return JsonResponse({'status': 'error', 'msg': 'Invalid request.'})




# Ensure _get_deleter_name and CommercialResaleProperty are imported

@require_POST
def commercial_resale_bulk_delete(request):
    """Handles Advanced Bulk Deletions for Commercial Resale Properties."""
    
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return JsonResponse({
            'status': 'error',
            'message': 'Unauthorized access.'
        })

    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        deleter_name = _get_deleter_name(request)

        properties = CommercialResaleProperty.objects.filter(
            is_deleted=False
        )

        # Delete All
        if delete_type == 'delete_all':
            count = properties.count()

            properties.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved ALL {count} properties to Recycle Bin.'
            })

        # Current Page
        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])

            target_props = properties.filter(id__in=page_ids)
            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} properties from current page to Recycle Bin.'
            })

        # Date Range
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')

            target_props = properties.filter(
                created_at__date__range=[from_date, to_date]
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} properties in date range to Recycle Bin.'
            })

        # Latest Month
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)

            target_props = properties.filter(
                created_at__gte=thirty_days_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} properties from the last 30 days to Recycle Bin.'
            })

        # Old Data
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)

            target_props = properties.filter(
                created_at__lt=six_months_ago
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} older properties to Recycle Bin.'
            })

        # By Uploader
        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '').strip()

            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) |
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )

            count = target_props.count()

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} properties uploaded by "{uploader}" to Recycle Bin.'
            })

        # By File Name
        elif delete_type == 'by_file':
            file_name = data.get('file_name', '').strip()

            if not file_name:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please select a file name.'
                })

            # ── FIX APPLIED HERE: Changed uploaded_file_name to upload_file_name ──
            target_props = properties.filter(
                upload_file_name__iexact=file_name 
            )

            count = target_props.count()

            if count == 0:
                # ── FIX APPLIED HERE ALSO ──
                available_files = list(
                    CommercialResaleProperty.objects.filter(
                        is_deleted=False
                    )
                    .exclude(upload_file_name__isnull=True)
                    .exclude(upload_file_name='')
                    .values_list('upload_file_name', flat=True)
                    .distinct()
                )

                return JsonResponse({
                    'status': 'error',
                    'message': f'No properties found for file: {file_name}',
                    'available_files': available_files
                })

            target_props.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=deleter_name
            )

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully moved {count} properties imported from "{file_name}" to Recycle Bin.'
            })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })


@require_POST
def commercial_resale_restore(request, id):
    try:
        # Reset the flags to put it back in the main active list
        CommercialResaleProperty.objects.filter(id=id).update(
            is_deleted=False, 
            deleted_at=None, 
            deleted_by=None
        )
        return JsonResponse({'status': 'success', 'message': 'Commercial Resale property restored successfully!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ── 2. HARD DELETE COMMERCIAL RESALE ──
@require_POST
def commercial_resale_hard_delete(request, id):
    try:
        # Fetch the property and permanently delete it from the database
        prop = CommercialResaleProperty.objects.get(id=id)
        prop.delete() # This will also delete associated images if cascading is set up
        
        return JsonResponse({'status': 'success', 'message': 'Property permanently deleted!'})
    except CommercialResaleProperty.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Property not found.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# --- 1. COMMERCIAL VIEW PAGE ---
def commercial_resale_view(request, id):

    admin_id = request.session.get('Admin_id')
    user_id  = request.session.get('User_id')

    if not admin_id and not user_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = None
    user_obj  = None

    if admin_id:
        try:
            admin_obj = Admin_Login.objects.get(id=admin_id)
        except Admin_Login.DoesNotExist:
            return render(request, 'home_page/Adminlogin.html')
    elif user_id:
        try:
            user_obj = User_Details.objects.get(id=user_id)
        except User_Details.DoesNotExist:
            return render(request, 'home_page/Adminlogin.html')
    prop = get_object_or_404(CommercialResaleProperty, id=id)
    images = prop.images.all()
    
    # Split strings for nice badge rendering in HTML
    facilities_list = [f.strip() for f in prop.nearby_facilities.split(',')] if prop.nearby_facilities else []
    amenities_list = [a.strip() for a in prop.amenities.split(',')] if prop.amenities else []

    context = {
        'admin_obj': admin_obj,
        'user_obj': user_obj,
        'prop': prop,
        'images': images,
        'facilities_list': facilities_list,
        'amenities_list': amenities_list
    }
    return render(request, 'admin_user/Reports/Resale/commercial_resale_view.html', context)







####################End Views Section For Commercial Resale Property #######################################



####################START Views Section For AGRICULTURAL Resale Property #######################################







def add_agricultural_property(request):
    """
    Handles the 4-step agricultural property listing form submission.
    Field mapping strictly follows the exact sequential database model schema.
    """
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')

    if not admin_id and not user_id:
        return redirect('login')

    try:
        if admin_id:
            admin = Admin_Login.objects.get(id=admin_id)
            uploader_name = admin.name
            uploader_email = admin.email
            uploader_phone = admin.phone
            uploader_role = admin.role
        else:
            user = User_Details.objects.get(id=user_id)
            uploader_name = user.name
            uploader_email = user.email
            uploader_phone = user.phone
            uploader_role = user.role

        if request.method == 'POST':
            def get_decimal(val):
                return val if val and str(val).strip() != "" else None

            with transaction.atomic():
                # Strictly following DB sequence for AgriculturalResaleProperty
                property_obj = AgriculturalResaleProperty.objects.create(
                    # ── SYSTEM CONTROL & IDENTIFICATION
                    property_title=request.POST.get('property_title', ''),

                    # ── STEP 1: LAND DETAILS
                    agriculture_property_type=request.POST.get('agriculture_property_type', ''),
                    land_area=get_decimal(request.POST.get('land_area')),
                    village=request.POST.get('village', ''),
                    taluka=request.POST.get('taluka', ''),
                    district=request.POST.get('district', ''),
                    
                    soil_type=request.POST.get('soil_type') or None,
                    irrigation_facility_active=request.POST.get('irrigation_facility_active', 'no'),
                    water_source_infrastructure=request.POST.get('water_source') or None,
                    fertility_status=request.POST.get('fertility_status') or None,
                    previous_crops=request.POST.get('previous_crops') or None,

                    # ── STEP 2: PRICING & LEGAL
                    expected_price=get_decimal(request.POST.get('expected_price')),
                    # price_per_acre is auto-calculated in model's save() method
                    brokerage=request.POST.get('brokerage') or None,
                    brokerage_percentage=request.POST.get('brokerage_percentage') or None,
                    manual_brokerage=request.POST.get('manual_brokerage') or None,
                    
                    ownership_type=request.POST.get('ownership_type', ''),
                    
                    loan_on_property=request.POST.get('agri_loan', 'no'),
                    loan_amount=get_decimal(request.POST.get('loan_amount')) if request.POST.get('agri_loan') == 'yes' else None,
                    existing_tenants=request.POST.get('agri_tenants', 'no'),
                    tenant_details=request.POST.get('tenant_details') if request.POST.get('agri_tenants') == 'yes' else None,
                    agri_dispute=request.POST.get('agri_dispute', 'no'),
                    dispute_details=request.POST.get('dispute_details') if request.POST.get('agri_dispute') == 'yes' else None,
                    pending_tax_due=request.POST.get('agri_tax_due', 'no'),
                    pending_tax_amount=get_decimal(request.POST.get('pending_tax_amount')) if request.POST.get('agri_tax_due') == 'yes' else None,
                    
                    property_description=request.POST.get('property_description', ''),
                    property_summary=request.POST.get('property_summary', ''),
                    user_description=request.POST.get('user_description', ''),
                    


                    # ── STEP 3: LOCATION & OWNER
                    city=request.POST.get('city', ''),
                    state=request.POST.get('state', ''),
                    locality_area=request.POST.get('locality', ''),
                    property_address=request.POST.get('address', ''),
                    
                    owner_name=request.POST.get('owner_name', ''),
                    owner_contact=request.POST.get('owner_contact', ''),
                    owner_email=request.POST.get('owner_email', ''),
                    owner_role=None, # Explicitly mapped to DB sequence though not in frontend
                    residency_status=request.POST.get('residency_status', 'resident'),

                    # ── UPLOADER / AUDIT
                    uploaded_by_name=uploader_name,
                    uploaded_by_email=uploader_email,
                    uploaded_by_contact=uploader_phone,
                    uploaded_by_role=uploader_role,
                )

                # ── STEP 4: DOCUMENTS & PHOTOS (Handled post-creation)
                if 'encumbrance_cert' in request.FILES:
                    property_obj.encumbrance_cert = request.FILES['encumbrance_cert']

                if 'property_video' in request.FILES:
                    property_obj.property_video = request.FILES['property_video']

                # Trigger model save to execute auto-calculations and document attachment
                property_obj.save()

                # User-ordered image injection
                images = request.FILES.getlist('property_images[]')
                for img in images[:10]:
                    AgriculturalResaleImage.objects.create(
                        property=property_obj,
                        image=img
                    )

            return JsonResponse({
                'status': 'success',
                'message': f'Agricultural Property "{property_obj.property_title}" published successfully to directories!'
            })

    except Exception as e:
        print("ERROR:", str(e))
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

    context = {'admin_obj': admin if admin_id else user}
    return render(request, 'admin_user/Reports/Resale/agricultural_list.html', context)










def edit_agricultural_property(request, pk):
    property_obj = get_object_or_404(AgriculturalResaleProperty, pk=pk)

    if request.method == 'POST':
        try:
            def get_decimal(value):
                if value is None or str(value).strip() == "":
                    return None
                return float(value)

            def clean_yes_no(val):
                return str(val).strip().lower() if val else 'no'

            with transaction.atomic():
                # ── STEP 1: LAND DETAILS ──────────────────────────────────
                property_obj.property_title = request.POST.get('property_title') or property_obj.property_title
                property_obj.agriculture_property_type = request.POST.get('agriculture_property_type')
                property_obj.village = request.POST.get('village')
                property_obj.taluka = request.POST.get('taluka')
                property_obj.district = request.POST.get('district')
                property_obj.land_area = get_decimal(request.POST.get('land_area'))
                
                property_obj.soil_type = request.POST.get('soil_type')
                property_obj.irrigation_facility_active = clean_yes_no(request.POST.get('irrigation_facility_active'))
                property_obj.water_source_infrastructure = request.POST.get('water_source_infrastructure')
                property_obj.fertility_status = request.POST.get('fertility_status')
                property_obj.previous_crops = request.POST.get('previous_crops')

                # ── STEP 2: PRICING & LEGAL ───────────────────────────────
                property_obj.expected_price = get_decimal(request.POST.get('expected_price'))
                property_obj.brokerage = request.POST.get('brokerage')
                property_obj.brokerage_percentage = request.POST.get('brokerage_percentage')
                property_obj.manual_brokerage = request.POST.get('manual_brokerage')
                property_obj.ownership_type = request.POST.get('ownership_type')
                
                property_obj.loan_on_property = clean_yes_no(request.POST.get('loan_on_property'))
                property_obj.loan_amount = (
                    get_decimal(request.POST.get('loan_amount'))
                    if property_obj.loan_on_property == 'yes' else None
                )

                property_obj.existing_tenants = clean_yes_no(request.POST.get('existing_tenants'))
                property_obj.tenant_details = (
                    request.POST.get('tenant_details')
                    if property_obj.existing_tenants == 'yes' else ""
                )

                property_obj.agri_dispute = clean_yes_no(request.POST.get('agri_dispute'))
                property_obj.dispute_details = (
                    request.POST.get('dispute_details')
                    if property_obj.agri_dispute == 'yes' else ""
                )

                property_obj.pending_tax_due = clean_yes_no(request.POST.get('pending_tax_due'))
                property_obj.pending_tax_amount = (
                    get_decimal(request.POST.get('pending_tax_amount'))
                    if property_obj.pending_tax_due == 'yes' else None
                )
                
                property_obj.property_description = request.POST.get('property_description')
                property_obj.property_summary = request.POST.get('property_summary')
                property_obj.user_description = request.POST.get('user_description')

                # ── STEP 3: LOCATION & OWNER ─────────────────────────────
                property_obj.city = request.POST.get('city')
                property_obj.state = request.POST.get('state')
                property_obj.locality_area = request.POST.get('locality_area')
                property_obj.property_address = request.POST.get('property_address')
                
                property_obj.owner_name = request.POST.get('owner_name')
                property_obj.owner_contact = request.POST.get('owner_contact')
                property_obj.owner_email = request.POST.get('owner_email')
                property_obj.owner_role = request.POST.get('owner_role')
                property_obj.residency_status = request.POST.get('residency_status', 'resident')

                # ── STEP 4: DOCUMENTS & PHOTOS ────────────────────────────
                if 'encumbrance_cert' in request.FILES:
                    property_obj.encumbrance_cert = request.FILES['encumbrance_cert']

                if 'property_video' in request.FILES:
                    property_obj.property_video = request.FILES['property_video']

                # Commits DB sequence & calls auto-calculate properties mapped in model's save() override
                property_obj.save()

                # Multiple Portfolio Image Upload Handling Loops
                images = request.FILES.getlist('property_images[]')
                current_count = property_obj.images.count()

                for img in images:
                    if current_count < 10:
                        AgriculturalResaleImage.objects.create(
                            property=property_obj,
                            image=img
                        )
                        current_count += 1

            return JsonResponse({
                'status': 'success',
                'message': 'Agricultural listing alterations updated successfully!',
                'redirect_url': reverse('agricultural_resale_list')
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return render(
        request,
        'admin_user/Resale/edit_agricultural_resale.html',
        {'property': property_obj}
    )


@require_POST
def agricultural_resale_restore(request, id):
    AgriculturalResaleProperty.objects.filter(id=id).update(is_deleted=False, deleted_at=None, deleted_by=None)
    return JsonResponse({'status': 'success', 'message': 'Agricultural property restored!'})

@require_POST
def agricultural_resale_hard_delete(request, id):
    AgriculturalResaleProperty.objects.filter(id=id).delete()
    return JsonResponse({'status': 'success', 'message': 'Permanently deleted!'})


# ── 1. SINGLE DELETE VIEW (Soft Delete) ──
@require_POST
def delete_agricultural_property(request, pk):
    try:
        # Get the name of the user/admin deleting the property
        deleter_name = _get_deleter_name(request)
        
        property_obj = get_object_or_404(AgriculturalResaleProperty, pk=pk)
        
        # Soft Delete
        property_obj.is_deleted = True
        property_obj.deleted_at = timezone.now()
        property_obj.deleted_by = deleter_name
        property_obj.save()
        
        return JsonResponse({
            'status': 'success', 
            'message': 'Moved to Recycle Bin successfully.'
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ── 2. BULK DELETE VIEW (Soft Delete) ──
@require_POST
def agricultural_bulk_delete(request):
    """Handles Advanced Bulk Deletions for Agricultural Properties."""
    admin_id = request.session.get('Admin_id')
    user_id = request.session.get('User_id')
    
    if not admin_id and not user_id:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'})

    try:
        data = json.loads(request.body)
        delete_type = data.get('delete_type')
        deleter_name = _get_deleter_name(request)
        
        # Target only properties not currently in the Recycle Bin
        properties = AgriculturalResaleProperty.objects.filter(is_deleted=False)
        
        if delete_type == 'delete_all':
            count = properties.count()
            properties.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved ALL {count} properties to Recycle Bin.'})
            
        elif delete_type == 'current_page':
            page_ids = data.get('page_ids', [])
            target_props = properties.filter(id__in=page_ids)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from current page to Recycle Bin.'})
            
        elif delete_type == 'date_range':
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            # Assuming you want to filter by the created date
            target_props = properties.filter(created_at__date__range=[from_date, to_date])
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties in date range to Recycle Bin.'})
            
        elif delete_type == 'latest_month':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            target_props = properties.filter(created_at__gte=thirty_days_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from the last 30 days to Recycle Bin.'})
            
        elif delete_type == 'old_data':
            six_months_ago = timezone.now() - timedelta(days=180)
            target_props = properties.filter(created_at__lt=six_months_ago)
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} older properties to Recycle Bin.'})
            
        elif delete_type == 'by_uploader':
            uploader = data.get('uploader_text', '')
            target_props = properties.filter(
                Q(uploaded_by_name__icontains=uploader) | 
                Q(uploaded_by_email__icontains=uploader) |
                Q(uploaded_by_role__icontains=uploader)
            )
            count = target_props.count()
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties uploaded by {uploader} to Recycle Bin.'})

        elif delete_type == 'by_file':
            file_name = data.get('file_name', '')
            # Replace 'upload_file_name' with your exact DB field name
            target_props = properties.filter(upload_file_name=file_name) 
            count = target_props.count()
            if count == 0:
                return JsonResponse({'status': 'error', 'message': f'No properties found for file: {file_name}'})
            target_props.update(is_deleted=True, deleted_at=timezone.now(), deleted_by=deleter_name)
            return JsonResponse({'status': 'success', 'message': f'Successfully moved {count} properties from {file_name} to Recycle Bin.'})
            
        else:
            return JsonResponse({'status': 'error', 'message': 'Unknown delete criteria.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})




#



def safe_float(val):
    """Safely converts Excel cell values to float, handling strings, currencies, and spaces."""
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        clean_val = str(val).replace('₹', '').replace(',', '').strip()
        return float(clean_val)
    except (ValueError, TypeError):
        return 0.0






import hashlib
import openpyxl
import traceback
import re
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure AgriculturalResaleProperty and Admin_Login models are imported

# =========================================================================
# UNIFIED SPREADSHEET SCHEMA (Used by both Export and Sample Download)
# =========================================================================
EXPORT_SECTIONS = [
    ("📋 Land Details", [
        ("agriculture_property_type", "Property Type *",        "agriculture_land / farm_land / orchard_land"),
        ("land_area",                 "Land Area (Acres) *",    "e.g. 5.5"),
        ("village",                   "Village *",              "e.g. Warud"),
        ("taluka",                    "Taluka *",               "e.g. Warud"),
        ("district",                  "District *",             "e.g. Amravati"),
        ("soil_type",                 "Soil Type",              "black / red / alluvial / sandy / loamy"),
        ("irrigation_facility_active","Irrigation Active?",     "yes / no"),
        ("water_source_infrastructure","Water Source",          "well / borewell / canal / river / none"),
        ("fertility_status",          "Fertility Status",       "high / medium / low"),
        ("previous_crops",            "Previous Crops",         "e.g. Wheat, Cotton"),
    ]),
    ("📋 Pricing", [
        ("expected_price",      "Expected Price (₹) *", "e.g. 5000000"),
        ("price_per_acre",      "Price Per Acre (₹)",   "Auto Generated (Leave Blank)"),
        ("brokerage",           "Brokerage",            "Yes / No"),
        ("brokerage_percentage","Brokerage %",          "e.g. 2% or leave blank"),
        ("manual_brokerage",    "Manual Brokerage",     "e.g. 50000 or leave blank"),
    ]),
    ("📋 Ownership & Legal", [
        ("ownership_type",         "Ownership Type *",      "freehold / leasehold"),
        ("loan_on_property",       "Loan On Property? *",   "yes / no"),
        ("loan_amount",            "Loan Amount (₹)",       "e.g. 200000 (0 if no loan)"),
        ("existing_tenants",       "Existing Tenants? *",   "yes / no"),
        ("tenant_details",         "Tenant Details",        "Enter if tenants=yes else leave blank"),
        ("agri_dispute",           "Agri Dispute? *",       "yes / no"),
        ("dispute_details",        "Dispute Details",       "Enter if dispute=yes else leave blank"),
        ("pending_tax_due",        "Pending Tax Due? *",    "yes / no"),
        ("pending_tax_amount",     "Pending Tax Amount (₹)","0 if no tax due"),
        ("resale_agricultural_desc","Description *",        "Short summary of the land"),
    ]),
    ("📋 Address", [
        ("city",             "City *",             "e.g. Nagpur"),
        ("state",            "State *",            "e.g. Maharashtra"),
        ("locality_area",    "Locality Area *",    "e.g. Besa Rural"),
        ("property_address", "Property Address *", "Near highway bridge, Ward No 4"),
    ]),
    ("📋 Owner Contact", [
        ("owner_name",       "Owner Name *",       "Full Name"),
        ("owner_contact",    "Owner Contact *",    "10-digit mobile"),
        ("owner_email",      "Owner Email *",      "email@example.com"),
        ("owner_role",       "Owner Role",         "e.g. Direct Owner, Broker"),
        ("residency_status", "Residency Status *", "resident / nri / pio"),
    ]),
]






def import_agricultural_resale_excel(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return JsonResponse({"status": "error", "message": "Unauthorized access"})

    if request.method == "POST" and request.FILES.get('excel_file'):
        try:
            admin_obj = Admin_Login.objects.get(id=session_id)
            excel_file = request.FILES['excel_file']
            file_name_str = excel_file.name

            excel_file.seek(0)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            if sheet.max_row < 5:
                return JsonResponse({"status": "error", "message": "Empty file uploaded or missing data rows."})

            # Rebuild expected headers from the EXPORT_SECTIONS list
            expected = []
            for label, fields in EXPORT_SECTIONS:
                for db, disp, hint in fields:
                    expected.append(disp)

            row3 = [str(cell.value).strip() if cell.value else "" for cell in sheet[3]]
            
            for idx, key in enumerate(expected):
                if idx >= len(row3) or row3[idx].lower() != key.lower():
                    return JsonResponse({
                        "status": "error",
                        "message": f"Column mismatch at Index {idx+1}. Expected '{key}', found '{row3[idx] if idx<len(row3) else 'Missing'}'."
                    })

            def cs(v): return str(v).strip() if v is not None else ""
            def yn(v): return cs(v).lower() if v else "no"
            def safe_float(v):
                try: return float(str(v).replace(',', '').strip())
                except (ValueError, TypeError): return 0.0

            records = []
            skipped_duplicates = 0
            
            # Required field indices based strictly on database sequence
            required_idx = [0, 1, 2, 3, 4, 10, 15, 16, 18, 20, 22, 24, 25, 26, 27, 28, 29, 30, 31, 33]

            DATA_START_ROW = 5 

            for row_idx, row in enumerate(sheet.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
                # Skip blank rows
                if not row or not any(v is not None and str(v).strip() != "" for v in row): 
                    continue
                
                if len(row) < len(expected):
                    return JsonResponse({"status": "error", "message": f"Row {row_idx} does not have enough columns to process."})

                # Check mandatory validations
                for ri in required_idx:
                    if row[ri] is None or str(row[ri]).strip() == "":
                        return JsonResponse({"status": "error", "message": f"Row {row_idx}: Required field '{expected[ri]}' is missing."})
                
                # Deduplication logic (Prevents importing the exact same row twice)
                row_raw_string = f"{row[29]}_{row[30]}_{row[2]}_{row[1]}_{row[10]}"  
                row_fp = hashlib.md5(row_raw_string.encode()).hexdigest()

                if AgriculturalResaleProperty.objects.filter(resale_agricultural_desc__icontains=f"[ROW-MD5:{row_fp}]", is_deleted=False).exists():
                    skipped_duplicates += 1
                    continue 

                records.append({
                    "agriculture_property_type": cs(row[0]),
                    "land_area": safe_float(row[1]),
                    "village": cs(row[2]), "taluka": cs(row[3]), "district": cs(row[4]),
                    "soil_type": cs(row[5]), "irrigation_facility_active": yn(row[6]),
                    "water_source_infrastructure": cs(row[7]), "fertility_status": cs(row[8]), "previous_crops": cs(row[9]),
                    
                    "expected_price": safe_float(row[10]),
                    # Skipping index 11 (price per acre) as backend handles it automatically
                    "brokerage": cs(row[12]) if row[12] else "No",
                    "brokerage_percentage": cs(row[13]), "manual_brokerage": cs(row[14]),
                    "ownership_type": cs(row[15]), 
                    
                    "loan_on_property": yn(row[16]),
                    "loan_amount": safe_float(row[17]) if yn(row[16]) == "yes" else 0.0,
                    "existing_tenants": yn(row[18]),
                    "tenant_details": cs(row[19]) if yn(row[18]) == "yes" else "",
                    "agri_dispute": yn(row[20]),
                    "dispute_details": cs(row[21]) if yn(row[20]) == "yes" else "",
                    "pending_tax_due": yn(row[22]),
                    "pending_tax_amount": safe_float(row[23]) if yn(row[22]) == "yes" else 0.0,
                    
                    "resale_agricultural_desc": cs(row[24]),
                    
                    "city": cs(row[25]), "state": cs(row[26]),
                    "locality_area": cs(row[27]), "property_address": cs(row[28]),
                    
                    "owner_name": cs(row[29]), "owner_contact": cs(row[30]),
                    "owner_email": cs(row[31]), "owner_role": cs(row[32]), "residency_status": cs(row[33]),
                    
                    "row_fingerprint": row_fp
                })

            if not records:
                if skipped_duplicates > 0:
                    return JsonResponse({"status": "error", "message": f"Upload rejected: Found {skipped_duplicates} row(s) in the file, but they are already saved in the database (Duplicates)."})
                else:
                    return JsonResponse({"status": "error", "message": "No valid data rows found to import."})

            imported_count = 0
            with transaction.atomic():
                for r in records:
                    AgriculturalResaleProperty.objects.create(
                        agriculture_property_type=r["agriculture_property_type"], land_area=r["land_area"],
                        village=r["village"], taluka=r["taluka"], district=r["district"],
                        soil_type=r["soil_type"], irrigation_facility_active=r["irrigation_facility_active"],
                        water_source_infrastructure=r["water_source_infrastructure"], fertility_status=r["fertility_status"], previous_crops=r["previous_crops"],
                        
                        expected_price=r["expected_price"],
                        brokerage=r["brokerage"], brokerage_percentage=r["brokerage_percentage"], manual_brokerage=r["manual_brokerage"],
                        ownership_type=r["ownership_type"],
                        
                        loan_on_property=r["loan_on_property"], loan_amount=r["loan_amount"],
                        existing_tenants=r["existing_tenants"], tenant_details=r["tenant_details"],
                        agri_dispute=r["agri_dispute"], dispute_details=r["dispute_details"],
                        pending_tax_due=r["pending_tax_due"], pending_tax_amount=r["pending_tax_amount"],
                        
                        resale_agricultural_desc=f"{r['resale_agricultural_desc']} [FILE:{file_name_str}] [ROW-MD5:{r['row_fingerprint']}]",
                        
                        city=r["city"], state=r["state"], locality_area=r["locality_area"], property_address=r["property_address"],
                        owner_name=r["owner_name"], owner_contact=r["owner_contact"], owner_email=r["owner_email"], 
                        owner_role=r["owner_role"], residency_status=r["residency_status"],
                        
                        uploaded_by_name=admin_obj.name, uploaded_by_email=admin_obj.email,
                        uploaded_by_contact=admin_obj.phone, uploaded_by_role=admin_obj.role, upload_file_name=file_name_str
                    )
                    imported_count += 1

            return JsonResponse({"status": "success", "message": f"Successfully imported {imported_count} records from {file_name_str}."})

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": f"System Error: {str(e)}"})

    return JsonResponse({"status": "error", "message": "Invalid request parameters."})

# =========================================================================
# 2. DOWNLOAD SAMPLE EXCEL
# =========================================================================
def download_agri_sample_excel(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Agricultural Resale"

    DARK_BG, WHITE, MID_BLUE = "1E293B", "FFFFFF", "3B82F6"
    LIGHT_BG, HINT_BG, SAMPLE_BG = "F8FAFC", "FEF9C3", "EFF6FF"
    HINT_FG, BORDER_COLOR = "92400E", "CBD5E1"
    thin  = Side(style="thin", color=BORDER_COLOR)
    thick = Side(style="medium", color="94A3B8")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    hb = Border(left=thick, right=thick, top=thick, bottom=thick)
    def hfill(h): return PatternFill("solid", fgColor=h)

    all_db, all_disp, all_hints, section_spans = [], [], [], []
    col = 1
    for label, fields in EXPORT_SECTIONS:
        s = col
        for db, disp, hint in fields:
            all_db.append(db); all_disp.append(disp); all_hints.append(hint); col += 1
        section_spans.append((label, s, col-1))

    # Row 1 – Section Banners
    for label, sc, ec in section_spans:
        c = sheet.cell(row=1, column=sc, value=label)
        c.font=Font(name="Arial",bold=True,size=11,color=WHITE); c.fill=hfill(DARK_BG)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=hb
        if sc != ec: sheet.merge_cells(start_row=1,start_column=sc,end_row=1,end_column=ec)
    sheet.row_dimensions[1].height = 30

    # Row 2 – DB fields
    for i,db in enumerate(all_db,1):
        c=sheet.cell(row=2,column=i,value=db)
        c.font=Font(name="Arial",bold=True,size=9,color="475569"); c.fill=hfill("E2E8F0")
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=cb
    sheet.row_dimensions[2].height = 22

    # Row 3 – Display Headers
    for i,disp in enumerate(all_disp,1):
        c=sheet.cell(row=3,column=i,value=disp)
        c.font=Font(name="Arial",bold=True,size=10,color=("C0392B" if disp.endswith("*") else MID_BLUE))
        c.fill=hfill(LIGHT_BG); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=cb
    sheet.row_dimensions[3].height = 36

    # Row 4 – Hints
    for i,hint in enumerate(all_hints,1):
        c=sheet.cell(row=4,column=i,value=hint)
        c.font=Font(name="Arial",italic=True,size=8,color=HINT_FG); c.fill=hfill(HINT_BG)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=cb
    sheet.row_dimensions[4].height = 30

    # Row 5 – Blocked Sample Row (Mapped to 34 variables exactly)
    sample = [
        "agriculture_land", 5.5, "Warud", "Warud", "Amravati", "black", "yes", "well", "high", "Wheat, Cotton",
        5000000, 909090, "Yes", "2%", "", "freehold", "yes", 200000, "no", "", "no", "", "no", 0, "Excellent land for farming close to arterial link pathways.",
        "Nagpur", "Maharashtra", "Besa Rural", "Near highway bridge, Ward No 4",
        "Ramesh Patil", "9876543210", "ramesh@example.com", "Direct Owner", "resident",
    ]
    
    for i,val in enumerate(sample,1):
        c=sheet.cell(row=5,column=i,value=val)
        c.font=Font(name="Arial",size=9,color="1E3A5F"); c.fill=hfill(SAMPLE_BG)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=cb
    sheet.row_dimensions[5].height = 22

    widths=[20, 16, 14, 14, 14, 14, 16, 16, 16, 18, 18, 16, 12, 14, 16, 16, 18, 16, 16, 20, 14, 20, 16, 20, 28, 14, 16, 18, 28, 18, 18, 24, 18, 18]
    for i,w in enumerate(widths,1): sheet.column_dimensions[get_column_letter(i)].width=w
    sheet.freeze_panes="A6"; sheet.sheet_view.zoomScale=90

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PropCRM_Agricultural_Template.xlsx"'
    wb.save(response)
    return response


# =========================================================================
# 3. EXPORT EXCEL
# =========================================================================
def export_agricultural_resale_excel(request):
    """Generates an Excel data dump built inside the exact import layout format schema."""
    search_query = request.GET.get('search', '').strip()
    queryset = AgriculturalResaleProperty.objects.filter(is_deleted=False)
    
    if search_query:
        normalized = search_query.replace(' ', '_')
        queryset = queryset.filter(
            Q(village__icontains=search_query) | Q(district__icontains=search_query) |
            Q(taluka__icontains=search_query) | Q(city__icontains=search_query) |
            Q(owner_name__icontains=search_query) | Q(agriculture_property_type__icontains=search_query) |
            Q(agriculture_property_type__icontains=normalized)
        )
    queryset = queryset.order_by('-created_at')

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Agricultural Export Data"

    DARK_BG, WHITE, MID_BLUE = "1E293B", "FFFFFF", "3B82F6"
    LIGHT_BG, HINT_BG = "F8FAFC", "FEF9C3"
    thin = Side(style="thin", color="CBD5E1")
    thick = Side(style="medium", color="94A3B8")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    hb = Border(left=thick, right=thick, top=thick, bottom=thick)
    def hfill(h): return PatternFill("solid", fgColor=h)

    all_db, all_disp, all_hints, section_spans = [], [], [], []
    col = 1
    for label, fields in EXPORT_SECTIONS:
        s = col
        for db, disp, hint in fields:
            all_db.append(db); all_disp.append(disp); all_hints.append(hint); col += 1
        section_spans.append((label, s, col-1))

    # Row 1 – Section Banners
    for label, sc, ec in section_spans:
        c = sheet.cell(row=1, column=sc, value=label)
        c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill = hfill(DARK_BG); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = hb
        if sc != ec: sheet.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
    sheet.row_dimensions[1].height = 30

    # Row 2 – System keys
    required_db_names = ["agriculture_property_type","land_area","village","taluka","district","expected_price","ownership_type","loan_on_property","existing_tenants","agri_dispute","pending_tax_due","resale_agricultural_desc","city","state","locality_area","property_address","owner_name","owner_contact","owner_email","residency_status"]
    
    for i, db in enumerate(all_db, 1):
        c = sheet.cell(row=2, column=i, value=f"{db} *" if db in required_db_names else db)
        c.font = Font(name="Arial", bold=True, size=9, color="475569")
        c.fill = hfill("E2E8F0"); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = cb
    sheet.row_dimensions[2].height = 22

    # Row 3 – Display Titles
    for i, disp in enumerate(all_disp, 1):
        c = sheet.cell(row=3, column=i, value=disp)
        c.font = Font(name="Arial", bold=True, size=10, color=("C0392B" if disp.endswith("*") else MID_BLUE))
        c.fill = hfill(LIGHT_BG); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = cb
    sheet.row_dimensions[3].height = 36

    # Row 4 – Instruction Hints
    for i, hint in enumerate(all_hints, 1):
        c = sheet.cell(row=4, column=i, value=hint)
        c.font = Font(name="Arial", italic=True, size=8, color="92400E")
        c.fill = hfill(HINT_BG); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = cb
    sheet.row_dimensions[4].height = 30

    # Row 5 onwards – Database Records Injection
    current_row = 5
    for item in queryset:
        for idx, db_field in enumerate(all_db, 1):
            val = getattr(item, db_field, "")
            
            # Clean dynamic descriptions so raw MD5/File tags aren't visible upon export
            if db_field == "resale_agricultural_desc" and val:
                val = re.sub(r'\[FILE:.+?\]|\[ROW-MD5:.+?\]', '', str(val)).strip()

            c = sheet.cell(row=current_row, column=idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = cb
            c.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[current_row].height = 20
        current_row += 1

    widths = [20, 16, 14, 14, 14, 14, 16, 16, 16, 18, 18, 16, 12, 14, 16, 16, 18, 16, 16, 20, 14, 20, 16, 20, 28, 14, 16, 18, 28, 18, 18, 24, 18, 18]
    for i, w in enumerate(widths, 1): 
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A5"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Agricultural_Properties_Export.xlsx"'
    wb.save(response)
    return response


# 3. LIST VIEW
# =========================================================================
import re


def agricultural_resale_list(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return render(request, 'home_page/Adminlogin.html')

    # Base Active Dataset
    

    context = {
        'admin_obj': admin_obj,
      
        
    }
    return render(request, 'admin_user/Reports/Resale/agricultural_list.html', context)



def agricultural_resale_reports(request):
    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    try:
        admin_obj = Admin_Login.objects.get(id=session_id)
    except Admin_Login.DoesNotExist:
        return render(request, 'home_page/Adminlogin.html')

    # Base Active Dataset
    base_qs = AgriculturalResaleProperty.objects.filter(is_deleted=False)
    
    # ── FIXED SEARCH FILTER LOGIC ──────────────────────────────────────
    search_query = request.GET.get('search', '').strip()
    properties = base_qs

    if search_query:
        # Normalize spaces to underscores if a user types "farm land" instead of "farm_land"
        normalized_query = search_query.replace(' ', '_')
        
        properties = properties.filter(
            Q(village__icontains=search_query) | 
            Q(district__icontains=search_query) |
            Q(taluka__icontains=search_query) |
            Q(city__icontains=search_query) | 
            Q(state__icontains=search_query) |
            Q(owner_name__icontains=search_query) | 
            Q(owner_contact__icontains=search_query) |
            Q(agriculture_property_type__icontains=search_query) |
            Q(agriculture_property_type__icontains=normalized_query)
        )

    # Apply sorting safely AFTER filtering
    properties = properties.order_by('-created_at')

    # Pagination setup
    paginator = Paginator(properties, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    total_value = base_qs.aggregate(Sum('expected_price'))['expected_price__sum'] or 0

    # Extract historical uploaded filenames from property description footprint strings
    uploaded_files = set()
    for desc in base_qs.exclude(resale_agricultural_desc__isnull=True).values_list('resale_agricultural_desc', flat=True):
        m = re.search(r'\[FILE:(.+?)\]', desc or '')
        if m: 
            uploaded_files.add(m.group(1))

    context = {
        'admin_obj': admin_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'uploaded_files': sorted(uploaded_files),
        'stats': {
            'total': base_qs.count(),
            'agri_land': base_qs.filter(agriculture_property_type='agriculture_land').count(),
            'farm_land': base_qs.filter(agriculture_property_type='farm_land').count(),
            'orchard': base_qs.filter(agriculture_property_type='orchard_land').count(),
            'total_value': total_value,
        }
    }
    return render(request, 'admin_user/Reports/Resale/agricultural_reports.html', context)


########### Views start for plot residential list ##########################

def residential_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/residential_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for plot residential list ################################


############# Views start for plot residential form #####################

def residential_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/residential_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for plot residential form #################################


########### Views start for plot commercial list ###########################

def commercial_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/commercial_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for plot commercial list ###########################


########### Views start for plot commercial form ##########################

def commercial_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/commercial_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############## Views end for plot commercial form ###########################


############ Views start for plot industrial list #######################

def industrial_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/industrial_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for plot industrial list ###############################


########## Views start for plot industrial form ################################

def industrial_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/industrial_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for plot industrail form ######################


############# Views start for plot agricultural list ######################

def agricultural_plot_resale_list(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Reports/Resale_Plot/agricultural_plot_resale_list.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

########### Views end for plot agricultural list #########################


########## Views start for plot agricultural form #######################

def agricultural_plot_resale(request):
    session_id = request.session.get('Admin_id')
    if session_id:
        admin_obj = Admin_Login.objects.get(id=session_id)
        ameneties_obj = Ameneties_Details.objects.all()
        facilities_obj = Facilities_Details.objects.all()
        user_obj = User_Details.objects.all()
        context = {'admin_obj':admin_obj,'ameneties_obj':ameneties_obj,'facilities_obj':facilities_obj,'user_obj':user_obj}
        return render(request,"admin_user/Resale_plot/agricultural_plot_resale.html",context)
    else:
        return render(request,'home_page/Adminlogin.html')

############# Views end for plot agricultural form ###########################





def export_agricultural_resale_csv(request):
    """Generates a plain-text CSV export preserving identical structural row indexing maps."""
    search_query = request.GET.get('search', '').strip()
    queryset = AgriculturalResaleProperty.objects.filter(is_deleted=False)
    if search_query:
        normalized = search_query.replace(' ', '_')
        queryset = queryset.filter(
            Q(village__icontains=search_query) | Q(district__icontains=search_query) |
            Q(taluka__icontains=search_query) | Q(city__icontains=search_query) |
            Q(owner_name__icontains=search_query) | Q(agriculture_property_type__icontains=search_query) |
            Q(agriculture_property_type__icontains=normalized)
        )
    queryset = queryset.order_by('-created_at')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="Agricultural_Properties_Export.csv"'
    
    writer = csv.writer(response)

    all_db, all_disp, all_hints = [], [], []
    for label, fields in EXPORT_SECTIONS:
        for db, disp, hint in fields:
            all_db.append(db); all_disp.append(disp); all_hints.append(hint)

    # Re-build matching row headers matrix
    # Note: Row 1 section names can't easily be merged in flat CSV files, so we pass blank pads
    section_row = []
    for label, fields in EXPORT_SECTIONS:
        section_row.append(label)
        section_row.extend([""] * (len(fields) - 1))
        
    writer.writerow(section_row) # Row 1
    
    # Process Row 2 updated required asterisks mapping keys
    db_row_final = []
    for i, db in enumerate(all_db, 1):
        if i in [1,2,3,4,5,11,15,16,18,20,22,24,25,26,27,28,29,30,31]:
            db_row_final.append(f"{db} *")
        else:
            db_row_final.append(db)
    writer.writerow(db_row_final) # Row 2
    
    writer.writerow(all_disp)     # Row 3
    writer.writerow(all_hints)    # Row 4

    # Inject data loops
    for item in queryset:
        row_data = []
        for db_field in all_db:
            val = getattr(item, db_field, "")
            if db_field == "resale_agricultural_desc" and val:
                val = re.sub(r'\[FILE:.+?\]|\[ROW-MD5:.+?\]', '', str(val)).strip()
            row_data.append(val)
        writer.writerow(row_data)

    return response

def view_agricultural_property(request, pk):
    session_id = request.session.get('Admin_id')
    user_id    = request.session.get('User_id')

    if not session_id and not user_id:
        return redirect('admin_login_url_name')

    property_obj = get_object_or_404(
        AgriculturalResaleProperty.objects.prefetch_related('images', 'faqs'),
        pk=pk
    )

    # Safety: regenerate FAQs if missing (old records before migration)
    if not property_obj.faqs.exists():
        property_obj.generate_auto_faqs()

    context = {
        'property': property_obj,
        'faqs': property_obj.faqs.all(),
    }
    return render(request, 'admin_user/Resale/view_agricultural_resale.html', context)


####################END Views Section For AGRICULTURAL Resale Property #######################################



#######################Start View SEO MODULE SECTION###################################







def seo_list(request):
    # --- Handle Bulk Action Logic ---
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        selected_ids = request.POST.getlist('ids[]')
        action = request.POST.get('action')
        
        if action == 'active':
            LocationSEO.objects.filter(id__in=selected_ids).update(is_active=True)
        elif action == 'pause':
            LocationSEO.objects.filter(id__in=selected_ids).update(is_active=False)
            
        return JsonResponse({'status': 'success', 'message': f'Items updated to {action}'})

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')
    
    # --- Regular List Logic ---
    # Retrieve all pages ordered by type for consistent grouping

    seo_pages = LocationSEO.objects.all().order_by('pagetype', '-id')

    # Get distinct page types and their counts for the menu/cards
    type_counts = LocationSEO.objects.values('pagetype').annotate(total=Count('id')).order_by('pagetype')

    

    admin_obj = Admin_Login.objects.get(id=session_id)

    context = {
        "seo_pages": seo_pages,
        "type_counts": type_counts,
        'admin_obj':admin_obj
    }
    return render(request, "admin_user/Seo_Module/seo_list.html", context)


#######################End View SEO MODULE SECTION###################################



#######################Start View BLOG MODULE SECTION###################################


def add_blog(request):
    if request.method == "POST":
        blog = Blog.objects.create(
            title=request.POST.get("title"),
            category=request.POST.get("category"),
            reading_time=request.POST.get("reading_time"),
            content=request.POST.get("content"),
            featured_image=request.FILES.get("featured_image"),
            author=request.POST.get("author"),
        )
        return redirect("blog_list")

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    context = {'admin_obj':admin_obj}

    return render(request, "admin_user/Seo_Module/Blog_Pages/blog_add.html",context)






def blog_list(request):

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')
    
    admin_obj = Admin_Login.objects.get(id=session_id)
    
    blogs = Blog.objects.all().order_by("-date_posted")
    return render(request, "admin_user/Seo_Module/Blog_Pages/blog_list.html", {"blogs": blogs,'admin_obj':admin_obj})


def blog_delete(request, id):
    blog = get_object_or_404(Blog, id=id)
    blog.delete()
    return redirect("blog_list")


def blog_edit(request, id):
    blog = get_object_or_404(Blog, id=id)
    if request.method == "POST":
        blog.title = request.POST.get("title")
        blog.category = request.POST.get("category")
        blog.reading_time = request.POST.get("reading_time")
        blog.content = request.POST.get("content")
        if request.FILES.get("featured_image"):
            blog.featured_image = request.FILES.get("featured_image")
        blog.author = request.POST.get("author")
        blog.slug = slugify(blog.title)
        blog.save()
        return redirect("blog_list")

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    return render(request, "admin_user/Seo_Module/Blog_Pages/blog_edit.html", {"blog": blog,'admin_obj':admin_obj})



def import_blog_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        
        # Validation
        if not excel_file or not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'status': 'error', 'message': 'Please upload a valid .xlsx file.'})

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Assuming Excel columns are: [Title, Category, Author, Reading Time, Content]
            imported_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True): # Start from row 2 to skip headers
                if len(row) >= 5:
                    title = row[0]
                    category = row[1] if row[1] else ""
                    author = row[2] if row[2] else ""
                    reading_time = str(row[3]) if row[3] else ""
                    content = row[4] if row[4] else ""
                    
                    if title: # Only save if title exists
                        Blog.objects.create(
                            title=title,
                            category=category,
                            author=author,
                            reading_time=reading_time,
                            content=content
                        )
                        imported_count += 1
            
            return JsonResponse({'status': 'success', 'imported': imported_count})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

#######################END View BLOG MODULE SECTION###################################


#######################START View SERVICES LANDING PAGE  MODULE SECTION###################################

def add_service(request):
    if request.method == "POST":
        title = request.POST.get("title")
        icon = request.POST.get("icon")
        short_description = request.POST.get("short_description")
        content = request.POST.get("content")   # CKEditor sends HTML
        featured_image = request.FILES.get("featured_image")
        active = bool(request.POST.get('is_active'))

        service = Service(
            title=title,
            icon=icon,
            short_description=short_description,
            content=content,
            featured_image=featured_image,
           # is_active=active
            #active = bool(request.POST.get('is_active'))
        )
        service.save()
       # return redirect("services_list")  # after save go to list page

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    context = {'admin_obj':admin_obj}

    return render(request, "admin_user/Seo_Module/Services_Pages/add_service.html",context)


def delete_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    service.delete()
    return redirect("services_list")



def edit_service(request, id):
    # Fetch the existing service using its ID
    service = get_object_or_404(Service, id=id)

    if request.method == "POST":
        # Get data from the form
        title = request.POST.get("title")
        icon = request.POST.get("icon")
        short_description = request.POST.get("short_description")
        content = request.POST.get("content")
        featured_image = request.FILES.get("featured_image")
        
        # Update the object
        service.title = title
        service.icon = icon
        service.short_description = short_description
        service.content = content
        
        # Only update the image if a new one was uploaded
        if featured_image:
            service.featured_image = featured_image
            
        # service.is_active = bool(request.POST.get('is_active')) # Uncomment if using active status

        # Save to database
        service.save()
        
        # Redirect back to the services list
        return redirect("services_list")

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    # For a GET request, pass the service object to the template
    context = {
        'service': service,
        'admin_obj':admin_obj
    }
    return render(request, "admin_user/Seo_Module/Services_Pages/edit_service.html", context)


def services_list(request):
    services = Service.objects.all().order_by('-id')

    session_id = request.session.get('Admin_id')
    if not session_id:
        return render(request, 'home_page/Adminlogin.html')

    admin_obj = Admin_Login.objects.get(id=session_id)

    return render(request, 'admin_user/Seo_Module/Services_Pages/services_list.html', {'services': services,'admin_obj':admin_obj})





def import_services_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        
        # Validation
        if not excel_file or not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'status': 'error', 'message': 'Please upload a valid .xlsx file.'})

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Assuming Excel columns are: [Title, Icon, Short Description, Content]
            imported_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True): # Start from row 2 to skip headers
                if len(row) >= 4:
                    title = row[0]
                    icon = row[1] if row[1] else "bi bi-check-circle" # Default icon fallback
                    short_description = row[2] if row[2] else ""
                    content = row[3] if row[3] else ""
                    
                    if title: # Only save if title exists
                        Service.objects.create(
                            title=title,
                            icon=icon,
                            short_description=short_description,
                            content=content
                        )
                        imported_count += 1
            
            return JsonResponse({'status': 'success', 'imported': imported_count})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})





#######################END View SERVICES LANDING PAGE  MODULE SECTION###################################