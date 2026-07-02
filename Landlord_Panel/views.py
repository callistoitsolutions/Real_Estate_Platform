from django.shortcuts import render,HttpResponse

# Create your views here.
from django.shortcuts import render
from django.contrib.auth.decorators import login_required



# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect
from Landlord_Panel .models import *
from Main_App .models import *
from Admin_App.models import *

# Create your views here.

from django.shortcuts import render,HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Create your views here.
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect,get_object_or_404
from Admin_App .models import *
from Main_App .models import *
from seo .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from openpyxl import load_workbook
from django.template.loader import render_to_string
import traceback
import json
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator  # ← ADD THIS
import csv
import csv
import json
from django.db.models import Count, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction  
from django.utils import timezone
import csv
from datetime import datetime, date
import io
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
import io
from django.urls import reverse,NoReverseMatch

from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
# I added 'Sum' to the end of this line:
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect

############## Views start for calculate profile strength ##########################

def calculate_profile_strength(user_obj):
    """Calculates the profile completion percentage dynamically based on the user's role."""
    strength = 0
    
    # ---------------------------------------------------------
    # 1. BASE DETAILS (Applies to everyone - Total 50%)
    # ---------------------------------------------------------
    if user_obj.user_name: strength += 10
    if user_obj.user_email: strength += 10
    if user_obj.user_phone: strength += 10
    if user_obj.user_password: strength += 10
    if user_obj.user_role: strength += 10
    
    # ---------------------------------------------------------
    # 2. ROLE-SPECIFIC DETAILS (Remaining 50%)
    # ---------------------------------------------------------
    role = user_obj.user_role

    if role == 'Vendor':
        # Vendor requires 5 specific things (10% each)
        if getattr(user_obj, 'user_state', None): strength += 10
        if getattr(user_obj, 'user_city', None): strength += 10
        if getattr(user_obj, 'user_address', None): strength += 10
        if getattr(user_obj, 'user_profile', None) and user_obj.user_profile.name: strength += 10
        if getattr(user_obj, 'user_service_type', None): strength += 10 # <-- Change to your actual DB field
        if getattr(user_obj, 'user_company_name', None): strength += 10 # <-- Change to your actual DB field
        if getattr(user_obj, 'user_profile', None) and user_obj.user_profile.name: strength += 10
        
    elif role in ['Agent', 'Agency/Builder']:
        # Agents/Agencies require different fields (10% each)
        if getattr(user_obj, 'user_state', None): strength += 10
        if getattr(user_obj, 'user_city', None): strength += 10
        if getattr(user_obj, 'user_address', None): strength += 10
        if getattr(user_obj, 'user_license_number', None): strength += 10 # <-- Change to your actual DB field
        if getattr(user_obj, 'user_profile', None) and user_obj.user_profile.name: strength += 10
        
    else:
        # Default for Landlord, Tenant, and Buyer
        if getattr(user_obj, 'user_state', None): strength += 15
        if getattr(user_obj, 'user_city', None): strength += 15
        if getattr(user_obj, 'user_address', None): strength += 10
        if getattr(user_obj, 'user_profile', None) and user_obj.user_profile.name: strength += 10

    # Ensure it never accidentally goes over 100
    return min(strength, 100)

############# Views end for calculate profile strength ##########################






def landlord_dashboard(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    # (Your original logic stays untouched here!)
    completion_score = calculate_profile_strength(user_obj)
    
    context = {
        'user_obj': user_obj,
        # Pass the object's role so the template behaves normally for the Landlord UI
        'user_role': user_obj.user_role, 
        'profile_completion_percentage': completion_score,
    }
    
    return render(request, "landlord/landlord_dashboard.html", context)













############## Views start for update landlord profile page ######################

def Update_Profile_Landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role
    }
    
    return render(request, "landlord/Profile/landlord_profile.html", context) 

########### Views end for update landlord profile page ##########################


############# Views start for rental forms list for landlord ###################

def residential_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role
    }
    
    return render(request, "landlord/Reports/Rental/residential_list.html", context)

############ Views end for rental forms list for landlord #######################


############# Views start for rental forms for landlord #########################

def residential_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Rental/residential.html", context)

############# Views end for rental forms for landlord #############################


########### Views start for commercial rent forms for landlord #######################

def commercial_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role
    }
    
    return render(request, "landlord/Reports/Rental/commercial_list.html", context)

############## Views end for commercial rent form for landlord #########################


############ Views start for commercial forms for landlord ############################

def commercial_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Rental/commercial.html", context)

############### Views end for commercial forms for landlord #############################


############ Views start for pg list forms for landlord #############################

def pg_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Rental/pg_list.html", context)

############ Views end for pg list forms for landlord ##########################


############# Views start for pg forms for landlord ########################

def pg_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Rental/pg_coliving.html", context)

########### Views end for pg forms for landlord #############################


############# Views start for resale property list #########################

def residential_resale_landlord_list(request):
   # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/residential_resale_list.html", context)

############# Views end for resale property list #######################


############ Views start for resale residential property ########################

def residential_resale_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()
    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/residential_resale.html", context)

########## Views end for resale residential property ############################


######### Views start for resale commercial property list #######################

def commercial_resale_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/commercial_list.html", context)

############# Views end for resale commercial property list ######################


########### Views start for resale property form commercial #######################

def commercial_resale_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/commercial_resale.html", context)

########## Views end for resale property form commercial ############################


############ Views start for resale plot property list ##########################

def plot_resale_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/plot_list.html", context)

############# Views end for resale plot property list ############################


############# Views start for residential plot list from landlord ###############

def plot_resale_res_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/Plots/res_plots_list.html", context)

########### Views end for resale residnetial plot list from landlord ################


############ Views start for residential plot from landlord ###################

def plot_resale_res_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/Plots/res_plots.html", context)

########### Views end for residential plot from landlord ########################


############# Views start for commercial plot list from landlord ##################

def plot_resale_comm_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/Plots/comm_plots_list.html", context)

############ Views end for commercial plot list from landlord #########################


############### Views start for commercial plot from landlord ##################

def plot_resale_comm_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/Plots/comm_plots.html", context)

############ Views end for commercial plot from landlord ##########################


########### Views start for industrial plot list from landlord ####################

def plot_resale_ind_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/Plots/ind_plots_list.html", context)

############# Views end for industrial plot list from landlord #####################


############ Views start for industrial plot from landlord #####################

def plot_resale_ind_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/Plots/ind_plots.html", context)

############# Views end for industrial plot from landlord ########################


############ Views start for agricultural property list from landlord ##############

def plot_resale_agri_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/Plots/agri_plots_list.html", context)

########## Views end for agricultural property list from landlord ######################


############ Views start for agricultural plots from landlord #####################

def plot_resale_agri_landlord(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/Plots/agri_plots.html", context)

############# Views end for agricultural plots from landlord ########################


############ Views start for resale industrial property list #####################

def industry_resale_landlord_list(request):
    # 1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
    }
    
    return render(request, "landlord/Reports/Resale/industrial_list.html", context)


############### Views end for resale industrial property list #########################


############ Views start for industrial resale property form #######################

def industry_resale_landlord(request):
     #1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/industrial_resale.html", context)

############### Views end for industry resale property form ############################


############# Views start for agricultural resale property form ####################

def agriculture_resale_landlord(request):
    #1. Retrieve BOTH possible session IDs from the browser
    user_id = request.session.get('User_id')
    admin_id = request.session.get('Admin_id') 
    logged_in_role = request.session.get('user_type')

    # 2. VIP Access Control
    is_valid_landlord = (user_id and logged_in_role == "Landlord")
    is_valid_admin = (admin_id and logged_in_role == "Admin" and 'impersonate_id' in request.session)

    # If they aren't a valid Landlord, AND they aren't an Admin trying to impersonate... kick them out.
    if not is_valid_landlord and not is_valid_admin:
        return redirect('login') 

    # 3. The ID Swap
    if is_valid_admin:
        # Admin is visiting: pull the target Landlord's ID
        dashboard_user_id = request.session.get('impersonate_id')
    else:
        # Normal Landlord is visiting: use their normal ID
        dashboard_user_id = user_id

    # 4. Data Fetching: Get the full user object using the final decided ID
    user_obj = User_Details.objects.get(id=dashboard_user_id)

    ameneties_obj = Ameneties_Details.objects.all()
    facilities_obj = Facilities_Details.objects.all()

    
    context = {
        'user_obj': user_obj,
        'user_role': user_obj.user_role,
        'ameneties_obj':ameneties_obj,
        'facilities_obj':facilities_obj
    }
    
    return render(request, "landlord/Forms/Resale/agricultural_resale.html", context)

############# Views end for agricultural resale property form #######################








